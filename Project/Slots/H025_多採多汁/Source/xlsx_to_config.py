from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SOURCE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SOURCE_DIR.parent

VARIANTS = {
    "92A": (SOURCE_DIR / "H025192A.xlsx", PROJECT_DIR / "config_92A.js"),
    "92B": (SOURCE_DIR / "H025192B.xlsx", PROJECT_DIR / "config_92B.js"),
    "94A": (SOURCE_DIR / "H025194A.xlsx", PROJECT_DIR / "config_94A.js"),
    "94B": (SOURCE_DIR / "H025194B.xlsx", PROJECT_DIR / "config_94B.js"),
}

INITIAL_BLOCKS = (
    ("U4:AA123", "AC4:AI123", "B19:D27"),
    ("BE4:BK123", "BM4:BS123", "AL19:AN27"),
    ("CO4:CU123", "CW4:DC123", "BV19:BX27"),
    ("DY4:EE123", "EG4:EM123", "DF19:DH27"),
    ("FI4:FO123", "FQ4:FW123", "EP19:ER27"),
)

DROP_BLOCKS = (
    ("U4:AA28", "AC4:AI28", "AK4:AK28", "C19:C20", "G19:I27"),
    ("BG4:BM28", "BO4:BU28", "BW4:BW28", "AO19:AO20", "AS19:AU27"),
    ("CS4:CY28", "DA4:DG28", "DI4:DI28", "CA19:CA20", "CE19:CG27"),
    ("EE4:EK28", "EM4:ES28", "EU4:EU28", "DM19:DM20", "DQ19:DS27"),
    ("FQ4:FW28", "FY4:GE28", "GG4:GG28", "EY19:EY20", "FC19:FE27"),
    ("HC4:HI28", "HK4:HQ28", "HS4:HS28", "GK19:GK20", "GO19:GQ27"),
)


def _floatify(value: Any) -> Any:
    if value is None or isinstance(value, str):
        return value
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return float(value)
    raise TypeError(f"Unsupported cell value: {value!r}")


def _matrix(ws: Any, cell_range: str, *, transpose: bool = False) -> list[Any]:
    values = [[_floatify(cell.value) for cell in row] for row in ws[cell_range]]
    if transpose:
        values = [list(column) for column in zip(*values)]
    if len(values) == 1:
        return values[0]
    return values


def _require_shape(key: str, value: list[Any], expected: tuple[int, ...]) -> None:
    actual: tuple[int, ...]
    if value and isinstance(value[0], list):
        actual = (len(value), len(value[0]))
    else:
        actual = (len(value),)
    if actual != expected:
        raise ValueError(f"{key}: expected shape {expected}, got {actual}")


def _parse_card_range(label: Any) -> tuple[float, float] | None:
    match = re.match(r"^\(\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*\]$", str(label or "").strip())
    if not match:
        return None
    return float(match.group(1)), float(match.group(2))


def _card_entries(rows: list[tuple[Any, ...]], header_index: int, weight_column: int) -> list[dict[str, Any]]:
    cards: list[dict[str, Any]] = []
    for row in rows[header_index + 2:]:
        label = row[15] if len(row) > 15 else None
        if label is None:
            break
        weight_value = row[weight_column] if len(row) > weight_column else 0
        if not isinstance(weight_value, (int, float)) or weight_value <= 0:
            continue
        label_text = str(label).strip()
        if label_text.lower() in {"fg trigger", "free game"}:
            cards.append({"type": "free_game", "weight": int(weight_value)})
            continue
        card_range = _parse_card_range(label_text)
        if card_range is not None:
            cards.append({
                "type": "range",
                "min": card_range[0],
                "max": card_range[1],
                "weight": int(weight_value),
            })
    return cards


def _parse_card_system(ws: Any) -> dict[str, Any]:
    rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
    sections: dict[str, int] = {}
    for index, row in enumerate(rows):
        if len(row) > 16 and str(row[15] or "").strip() == "Range":
            sections[str(row[16] or "").strip().lower()] = index

    if "base game" not in sections or "free game" not in sections:
        raise ValueError("Card: Base Game or Free Game section is missing")

    base_index = sections["base game"]
    free_index = sections["free game"]
    buy_index = sections.get("buy feature")
    newbie_bg = _card_entries(rows, base_index, 42)
    oldhand_bg = _card_entries(rows, base_index, 43)
    newbie_fg = _card_entries(rows, free_index, 42)
    oldhand_fg = _card_entries(rows, free_index, 43)
    buy_feature = _card_entries(rows, buy_index, 42) if buy_index is not None else []
    if not all((newbie_bg, oldhand_bg, newbie_fg, oldhand_fg)):
        raise ValueError("Card: active newbie/oldhand BG/FG weights are incomplete")

    return {
        "enabled": True,
        "retry_limit": 5000,
        "newbie": {
            "normal_bet": {"weight_bg": newbie_bg, "weight_fg": newbie_fg},
        },
        "oldhand": {
            "normal_bet": {"weight_bg": oldhand_bg, "weight_fg": oldhand_fg},
            "buy_feature": {"weight_fg": buy_feature},
        },
    }


def generate_config(xlsx_path: Path) -> dict[str, Any]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        required = {
            "Overview",
            "Description",
            "Card",
            "BaseGameSymbol",
            "BaseGameSymbolDrop",
            "FreeGameSymbol",
            "FreeGameSymbolDrop",
            "工作表1",
        }
        missing = required.difference(workbook.sheetnames)
        if missing:
            raise ValueError(f"Missing worksheets: {', '.join(sorted(missing))}")

        config: dict[str, Any] = {
            "linkpoint": _matrix(workbook["Overview"], "B36:K42"),
            "card_system": _parse_card_system(workbook["Card"]),
        }

        for prefix, sheet_name in (
            ("baseGame", "BaseGameSymbol"),
            ("FreeGame", "FreeGameSymbol"),
        ):
            ws = workbook[sheet_name]
            for index, (symbol_range, weight_range, my_range) in enumerate(INITIAL_BLOCKS, 1):
                config[f"{prefix}Symbol{index}"] = _matrix(ws, symbol_range, transpose=True)
                config[f"{prefix}SymbolWeight{index}"] = _matrix(ws, weight_range, transpose=True)
                config[f"{prefix}MY{index}"] = _matrix(ws, my_range, transpose=True)

        for prefix, sheet_name in (
            ("BaseGame", "BaseGameSymbolDrop"),
            ("FreeGame", "FreeGameSymbolDrop"),
        ):
            ws = workbook[sheet_name]
            for index, ranges in enumerate(DROP_BLOCKS, 1):
                symbol_range, random_range, parallel_range, method_range, my_range = ranges
                config[f"{prefix}Drop{index}"] = _matrix(ws, symbol_range, transpose=True)
                config[f"{prefix}DropRWeight{index}"] = _matrix(ws, random_range, transpose=True)
                config[f"{prefix}DropPWeight{index}"] = _matrix(ws, parallel_range, transpose=True)
                config[f"{prefix}Dropmethod{index}"] = _matrix(ws, method_range, transpose=True)
                config[f"{prefix}DropMy{index}"] = _matrix(ws, my_range, transpose=True)

        description = workbook["Description"]
        config.update(
            {
                "ReelWeight": _matrix(description, "D5:D9", transpose=True),
                "DropWeight": _matrix(description, "D18:M23", transpose=True),
                "Eliminate": _matrix(description, "D37:D45", transpose=True),
                "FreeReelWeight": _matrix(description, "G5:G9", transpose=True),
                "FreeDropWeight": _matrix(description, "P18:Y23", transpose=True),
                "FreeEliminate": _matrix(description, "H37:H45", transpose=True),
            }
        )

        scratch = workbook["工作表1"]
        for key, cell_range in (
            ("baseredraw", "B1:B59"),
            ("freeredraw", "C1:C58"),
            ("multipleRange", "A1:A57"),
            ("baseredrawB", "E1:E59"),
            ("freeredrawB", "F1:F58"),
            ("basemultiple", "P1:P58"),
            ("freemultiple", "Q1:Q58"),
        ):
            config[key] = _matrix(scratch, cell_range, transpose=True)

        validate_config(config)
        return config
    finally:
        workbook.close()


def validate_config(config: dict[str, Any]) -> None:
    _require_shape("linkpoint", config["linkpoint"], (7, 10))
    for prefix in ("baseGame", "FreeGame"):
        for index in range(1, 6):
            _require_shape(f"{prefix}Symbol{index}", config[f"{prefix}Symbol{index}"], (7, 120))
            _require_shape(
                f"{prefix}SymbolWeight{index}", config[f"{prefix}SymbolWeight{index}"], (7, 120)
            )
            _require_shape(f"{prefix}MY{index}", config[f"{prefix}MY{index}"], (3, 9))
    for prefix in ("BaseGame", "FreeGame"):
        for index in range(1, 7):
            _require_shape(f"{prefix}Drop{index}", config[f"{prefix}Drop{index}"], (7, 25))
            _require_shape(
                f"{prefix}DropRWeight{index}", config[f"{prefix}DropRWeight{index}"], (7, 25)
            )
            _require_shape(f"{prefix}DropPWeight{index}", config[f"{prefix}DropPWeight{index}"], (25,))
            _require_shape(f"{prefix}Dropmethod{index}", config[f"{prefix}Dropmethod{index}"], (2,))
            _require_shape(f"{prefix}DropMy{index}", config[f"{prefix}DropMy{index}"], (3, 9))
    for key, expected in (
        ("ReelWeight", (5,)),
        ("DropWeight", (10, 6)),
        ("Eliminate", (9,)),
        ("FreeReelWeight", (5,)),
        ("FreeDropWeight", (10, 6)),
        ("FreeEliminate", (9,)),
        ("baseredraw", (59,)),
        ("freeredraw", (58,)),
        ("multipleRange", (57,)),
        ("baseredrawB", (59,)),
        ("freeredrawB", (58,)),
        ("basemultiple", (58,)),
        ("freemultiple", (58,)),
    ):
        _require_shape(key, config[key], expected)

    for key, value in config.items():
        if _contains_none(value):
            raise ValueError(f"{key}: contains blank cells")


def _contains_none(value: Any) -> bool:
    if isinstance(value, list):
        return any(_contains_none(item) for item in value)
    return value is None


def load_config(path: Path) -> dict[str, Any]:
    raw = path.read_text(encoding="utf-8-sig").strip()
    raw = re.sub(r"^const\s+data\s*=\s*", "", raw)
    return json.loads(raw.rstrip(";"))


def write_config(path: Path, config: dict[str, Any]) -> None:
    payload = json.dumps(config, ensure_ascii=False, indent=2)
    path.write_text(f"const data = {payload};\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate H025 config files from the source workbooks")
    parser.add_argument("--all", action="store_true", help="Generate all 92A/92B/94A/94B variants")
    parser.add_argument("--check", action="store_true", help="Validate generated output after writing")
    parser.add_argument("--xlsx", type=Path, help="Generate one workbook")
    parser.add_argument("--output", type=Path, help="Output path used with --xlsx")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if args.all:
        jobs = list(VARIANTS.values())
    elif args.xlsx:
        if not args.output:
            raise SystemExit("--output is required with --xlsx")
        jobs = [(args.xlsx, args.output)]
    else:
        raise SystemExit("Use --all, or provide --xlsx and --output")

    for xlsx_path, output_path in jobs:
        xlsx_path = xlsx_path.resolve()
        output_path = output_path.resolve()
        config = generate_config(xlsx_path)
        write_config(output_path, config)
        if args.check:
            written = load_config(output_path)
            validate_config(written)
            if written != config:
                raise ValueError(f"Read-back mismatch: {output_path}")
        print(f"Generated {output_path.name} from {xlsx_path.name}")


if __name__ == "__main__":
    main()
