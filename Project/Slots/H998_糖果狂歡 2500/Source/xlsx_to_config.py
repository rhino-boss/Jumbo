from __future__ import annotations

import argparse
import json
import math
import re
import warnings
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


warnings.filterwarnings("ignore", message="Unknown extension is not supported and will be removed")
warnings.filterwarnings("ignore", message="Conditional Formatting extension is not supported and will be removed")


SOURCE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SOURCE_DIR.parent
DEFAULT_SOURCE = SOURCE_DIR / "H998192.xlsx"

GAME_ID = "101001"
PARSHEET_ID = "H9981"
GAME_NAME = "糖果狂歡 2500"
ENGLISH_NAME = "Sugar Bonanza 2500"

SOURCE_SHEETS = [
    "Base Game Symbol (1)",
    "Base Game Symbol (2)",
    "Base Game Symbol (3)",
    "Base Game Symbol (4)",
    "Extra Bet Symbol (1)",
    "Extra Bet Symbol (2)",
    "Extra Bet Symbol (3)",
    "Free Game Symbol (1)",
    "Free Game Symbol (2)",
    "Feature Buy Symbol (1)",
    "Feature Buy Symbol (2)",
    "Super Feature Buy Symbol (1)",
    "Super Feature Buy Symbol (2)",
]

# Keep these names for the existing H998 Simulator and DemoGame adapters.
LEGACY_STRIP_NAMES = [
    "BG_strip",
    "BG_strip (2)",
    "BG_strip (3)",
    "BG_strip (4)",
    "EB_strip",
    "EB_strip (2)",
    "EB_strip (3)",
    "FG_strip",
    "FG_strip (2)",
    "FB_strip",
    "FB_strip (2)",
    "SB_strip",
    "SB_strip (2)",
]

PARAMETER_SOURCE_TO_SHEET = {
    "Buy Feature FG Symbol (1)": "Feature Buy Symbol (1)",
    "Buy Feature FG Symbol (2)": "Feature Buy Symbol (2)",
    "Super Feature FG Symbol (1)": "Super Feature Buy Symbol (1)",
    "Super Feature FG Symbol (2)": "Super Feature Buy Symbol (2)",
}


def to_int(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    return int(round(float(value)))


def cumulative(values: list[int]) -> list[int]:
    running = 0
    result: list[int] = []
    for value in values:
        running += max(0, int(value))
        result.append(running)
    return result


def cumulative_by_row(rows: list[list[int]]) -> list[list[int]]:
    if not rows:
        return []
    running = [0] * len(rows[0])
    result: list[list[int]] = []
    for row in rows:
        running = [running[index] + max(0, int(value)) for index, value in enumerate(row)]
        result.append(list(running))
    return result


def normalize_weights(values: list[int]) -> list[int]:
    """Reduce weights by their common divisor without changing probabilities."""
    divisor = 0
    for value in values:
        if value > 0:
            divisor = math.gcd(divisor, int(value))
    if divisor <= 1:
        return [int(value) for value in values]
    return [int(value) // divisor for value in values]


def find_row(ws: Any, column: int, expected: str, start: int = 1) -> int:
    for row in range(start, ws.max_row + 1):
        if str(ws.cell(row, column).value or "").strip() == expected:
            return row
    raise ValueError(f"{ws.title}: could not find {expected!r} in column {column}")


def parse_overview(ws: Any) -> dict[str, Any]:
    model = str(ws["B2"].value or "").strip()
    excel_version = str(ws["B3"].value or "").strip()
    visible_row = find_row(ws, 1, "Visible Window Size")
    layout_visible = [to_int(ws.cell(visible_row, column).value) for column in range(2, 8)]

    pay_header = find_row(ws, 1, "Symbol")
    if str(ws.cell(pay_header, 9).value or "").strip() != "Id":
        raise ValueError(f"{ws.title}!I{pay_header}: expected Id header")

    symbol_str: dict[str, str] = {}
    symbol_id: list[int] = []
    pay_table: list[list[int]] = []
    row = pay_header + 1
    while ws.cell(row, 1).value is not None:
        symbol_code = str(ws.cell(row, 1).value).strip()
        symbol = to_int(ws.cell(row, 9).value)
        if symbol != len(symbol_id):
            raise ValueError(f"{ws.title}!I{row}: symbol IDs must be contiguous from zero")
        symbol_str[str(symbol)] = symbol_code
        symbol_id.append(symbol)
        pay_table.append([to_int(ws.cell(row, column).value) for column in range(3, 9)])
        row += 1

    free_header = find_row(ws, 1, "Free Spins Setting")
    retrigger_header = find_row(ws, 1, "Retrigger Setting")

    def read_awards(header_row: int) -> dict[str, int]:
        awards: dict[str, int] = {}
        row_index = header_row + 2
        while isinstance(ws.cell(row_index, 1).value, (int, float)):
            awards[str(to_int(ws.cell(row_index, 1).value))] = to_int(ws.cell(row_index, 2).value)
            row_index += 1
        return awards

    mode_names = ("normal", "extrabet", "featurebuy", "superfeaturebuy")
    rtp_targets = {
        name: float(ws.cell(row, 3).value)
        for name, row in zip(mode_names, range(6, 10))
    }
    bet_factors = {
        name: float(ws.cell(row, 2).value)
        for name, row in zip(mode_names, range(6, 10))
    }
    return {
        "model": model,
        "excel_version": excel_version,
        "default_coin_in": to_int(ws["A6"].value),
        "bet_factors": bet_factors,
        "rtp_targets": rtp_targets,
        "layout_visible": layout_visible,
        "window_size": max(layout_visible),
        "reel_num": len(layout_visible),
        "free_spin_awards": read_awards(free_header),
        "retrigger_awards": read_awards(retrigger_header),
        "pay_awards": [4, 5, 6, 8, 10, 12],
        "pay_table": pay_table,
        "symbol_str": symbol_str,
        "symbol_id": symbol_id,
    }


def parse_parameter_block(ws: Any, start_row: int, name: str) -> dict[str, Any]:
    table_weights = {
        "normal": [to_int(ws.cell(row, 3).value) for row in range(start_row + 2, start_row + 5)],
        "extrabet": [to_int(ws.cell(row, 4).value) for row in range(start_row + 2, start_row + 5)],
    }

    schedule_header = find_row(ws, 2, "高表場次", start_row)
    schedules: list[dict[str, Any]] = []
    row = schedule_header + 1
    while isinstance(ws.cell(row, 2).value, (int, float)):
        schedules.append(
            {
                "high": to_int(ws.cell(row, 2).value),
                "low": to_int(ws.cell(row, 3).value),
                "weights": {
                    "normal": to_int(ws.cell(row, 4).value),
                    "featurebuy": to_int(ws.cell(row, 5).value),
                    "superfeaturebuy": to_int(ws.cell(row, 6).value),
                },
            }
        )
        row += 1

    multiplier_header = find_row(ws, 2, "C2 Multiplier", start_row)
    source_row = multiplier_header + 1
    source_names = [str(ws.cell(source_row, column).value or "").strip() for column in range(3, 9)]
    multipliers: list[int] = []
    columns: list[list[int]] = [[] for _ in range(6)]
    row = source_row + 1
    while isinstance(ws.cell(row, 1).value, (int, float)):
        multipliers.append(to_int(ws.cell(row, 1).value))
        for index, column in enumerate(range(3, 9)):
            columns[index].append(to_int(ws.cell(row, column).value))
        row += 1

    return {
        "name": name,
        "table_weights": table_weights,
        "free_spin_mix": schedules,
        "multipliers": multipliers,
        "multiplier_sources": source_names,
        "multiplier_weights": {
            source: normalize_weights(weights)
            for source, weights in zip(source_names, columns)
        },
    }


def parse_parameter(ws: Any) -> dict[str, dict[str, Any]]:
    # A4:H43 is only a CHOOSE(A3, A44:H83, A84:H123) preview area.
    # The canonical source blocks are the last [A] and [B] occurrences.
    source_rows: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        label = str(ws.cell(row, 1).value or "").strip()
        if not re.fullmatch(r"\[[A-Za-z0-9_-]+\]", label):
            continue
        source_rows[label[1:-1]] = row
    if "A" not in source_rows or "B" not in source_rows:
        raise ValueError(f"{ws.title}: expected parameter blocks A and B")
    return {
        name: parse_parameter_block(ws, source_rows[name], name)
        for name in ("A", "B")
    }


def parse_strip_sheet(ws: Any, code_to_id: dict[str, int]) -> dict[str, Any]:
    symbols: list[list[int]] = []
    weights: list[list[int]] = []
    for row in range(4, ws.max_row + 1):
        if not isinstance(ws.cell(row, 11).value, (int, float)):
            continue
        raw_symbols = [ws.cell(row, column).value for column in range(12, 18)]
        if all(value is None for value in raw_symbols):
            continue
        reel_symbols: list[int] = []
        for column, value in zip(range(12, 18), raw_symbols):
            if value is None or str(value).strip() == "":
                reel_symbols.append(-1)
                continue
            code = str(value or "").strip()
            if code not in code_to_id:
                raise ValueError(f"{ws.title}!{ws.cell(row, column).coordinate}: unknown symbol {code!r}")
            reel_symbols.append(code_to_id[code])
        reel_weights = [to_int(ws.cell(row, column).value) for column in range(33, 39)]
        symbols.append(reel_symbols)
        weights.append(reel_weights)

    if not symbols:
        raise ValueError(f"{ws.title}: no reel rows found")
    reel_lengths = [
        sum(1 for row in symbols if row[reel] >= 0)
        for reel in range(6)
    ]
    for reel in range(6):
        if sum(row[reel] for row in weights) <= 0:
            raise ValueError(f"{ws.title}: reel {reel + 1} has no positive stop weight")

    multiplier_values: list[int] = []
    multiplier_weights: list[int] = []
    for row in range(4, ws.max_row + 1):
        label = ws.cell(row, 40).value
        weight = ws.cell(row, 41).value
        if label is None:
            continue
        match = re.fullmatch(r"\s*(\d+)x\s*", str(label), re.IGNORECASE)
        if match:
            multiplier_values.append(int(match.group(1)))
            multiplier_weights.append(to_int(weight))

    return {
        "symbols": symbols,
        "weights": weights,
        "weights_cum": cumulative_by_row(weights),
        "reel_lengths": reel_lengths,
        "multiplier_values": multiplier_values,
        "multiplier_weights": normalize_weights(multiplier_weights),
    }


def parse_card_section(ws: Any, title: str, include_trigger: bool = False) -> list[dict[str, Any]]:
    title_rows = [
        row
        for row in range(1, ws.max_row + 1)
        if str(ws.cell(row, 1).value or "").strip() == title
    ]
    if not title_rows:
        raise ValueError(f"{ws.title}: card section {title!r} not found")
    # Summary rows at the top repeat mode names. The actual card table is the
    # last exact section title and is followed by its own Lower/Upper header.
    title_row = title_rows[-1]
    header_row = find_row(ws, 1, "Lower", title_row)
    cards: list[dict[str, Any]] = []
    row = header_row + 2
    while isinstance(ws.cell(row, 1).value, (int, float)) and isinstance(ws.cell(row, 2).value, (int, float)):
        cards.append(
            {
                "type": "range",
                "min": float(ws.cell(row, 1).value),
                "max": float(ws.cell(row, 2).value),
                "table": str(ws.cell(row, 13).value or "").strip(),
                "weight": to_int(ws.cell(row, 18).value),
            }
        )
        row += 1
    if include_trigger:
        while row <= ws.max_row and str(ws.cell(row, 1).value or "").strip() != "FG Trigger":
            row += 1
        if row <= ws.max_row:
            cards.append(
                {
                    "type": "free_game",
                    "table": str(ws.cell(row, 13).value or "").strip(),
                    "weight": to_int(ws.cell(row, 18).value),
                }
            )
    return cards


def parse_card_profile(ws: Any, include_feature_modes: bool) -> dict[str, Any]:
    free_cards = parse_card_section(ws, "Free Game")
    profile: dict[str, Any] = {
        "normal_bet": {
            "weight_bg": parse_card_section(ws, "Base Game", include_trigger=True),
            "weight_fg": free_cards,
        },
        "extra_bet": {
            "weight_bg": parse_card_section(ws, "Extra Bet", include_trigger=True),
            "weight_fg": free_cards,
        },
    }
    if include_feature_modes:
        profile["buy_feature"] = {"weight_fg": parse_card_section(ws, "Buy Feature")}
        profile["super_feature"] = {"weight_fg": parse_card_section(ws, "Super Feature")}
    return profile


def parse_card_system(workbook: Any) -> dict[str, Any]:
    return {
        "enabled": True,
        "newbie": parse_card_profile(workbook["Multiplier_Weight_Newbie"], False),
        "oldhand": parse_card_profile(workbook["Multiplier_Weight_Oldhand"], True),
    }


def selected_mix(block: dict[str, Any], mode: str) -> tuple[int, int]:
    choices = [item for item in block["free_spin_mix"] if item["weights"][mode] > 0]
    if len(choices) != 1:
        raise ValueError(f"Parameter block {block['name']}: expected one active {mode} free-spin mix")
    return int(choices[0]["low"]), int(choices[0]["high"])


def pad_strip_tables(strip_data: list[dict[str, Any]]) -> tuple[list[Any], list[Any]]:
    max_rows = max(len(item["symbols"]) for item in strip_data)
    padded_symbols: list[list[list[int]]] = []
    padded_weights_cum: list[list[list[int]]] = []
    for item in strip_data:
        symbols = [list(row) for row in item["symbols"]]
        weights_cum = [list(row) for row in item["weights_cum"]]
        last_cum = list(weights_cum[-1])
        while len(symbols) < max_rows:
            symbols.append([-1] * 6)
            weights_cum.append(list(last_cum))
        padded_symbols.append(symbols)
        padded_weights_cum.append(weights_cum)
    return padded_symbols, padded_weights_cum


def build_config(source_path: Path) -> dict[str, Any]:
    workbook = load_workbook(source_path, read_only=False, data_only=True)
    required = {
        "Overview",
        "Parameter",
        "Multiplier_Weight_Newbie",
        "Multiplier_Weight_Oldhand",
        *SOURCE_SHEETS,
    }
    missing = sorted(required.difference(workbook.sheetnames))
    if missing:
        raise ValueError(f"{source_path.name}: missing worksheets: {', '.join(missing)}")

    overview = parse_overview(workbook["Overview"])
    parameters = parse_parameter(workbook["Parameter"])
    code_to_id = {code: int(symbol) for symbol, code in overview["symbol_str"].items()}
    strips = [parse_strip_sheet(workbook[name], code_to_id) for name in SOURCE_SHEETS]
    card_system = parse_card_system(workbook)
    workbook.close()

    arr_reels, arr_weights_cum = pad_strip_tables(strips)
    default_block = parameters["A"]
    low_spins, high_spins = selected_mix(default_block, "normal")
    retrigger_totals = set(overview["retrigger_awards"].values())
    if len(retrigger_totals) != 1:
        raise ValueError(f"Overview: retrigger awards must use one fixed spin count, got {retrigger_totals}")
    retrigger_total = retrigger_totals.pop()
    retrigger_high = round(retrigger_total * high_spins / (low_spins + high_spins))
    retrigger_low = retrigger_total - retrigger_high

    multiplier_sources = default_block["multiplier_sources"]
    multiplier_weights = default_block["multiplier_weights"]
    source_weights = [multiplier_weights[name] for name in multiplier_sources]
    strip_multiplier_weights: dict[str, dict[str, list[int]]] = {}
    for source_name, expected_values, expected_weights in zip(
        multiplier_sources,
        [default_block["multipliers"]] * len(multiplier_sources),
        source_weights,
    ):
        sheet_name = PARAMETER_SOURCE_TO_SHEET.get(source_name, source_name)
        strip_index = SOURCE_SHEETS.index(sheet_name)
        strip_values = strips[strip_index]["multiplier_values"]
        strip_weights = strips[strip_index]["multiplier_weights"]
        if strip_values != expected_values or strip_weights != expected_weights:
            raise ValueError(f"Parameter and {source_name} C2 multiplier data do not match")
        strip_multiplier_weights[source_name] = {
            "values": strip_values,
            "weights": strip_weights,
        }

    model_match = re.fullmatch(r"H9981(?P<rtp>\d{2})(?P<variant>[A-Za-z0-9_-]*)", overview["model"])
    if not model_match:
        raise ValueError(f"Overview model {overview['model']!r} does not match H9981xx naming")

    config: dict[str, Any] = {
        "source_xlsx": f"Source/{source_path.name}",
        "source_game_id": overview["model"],
        "model": overview["model"],
        "game_id": GAME_ID,
        "parsheet_id": PARSHEET_ID,
        "game_name": GAME_NAME,
        "display_name": GAME_NAME,
        "english_name": ENGLISH_NAME,
        "game_version": overview["excel_version"],
        "excel_version": overview["excel_version"],
        "target_rtp": int(model_match.group("rtp")) / 100,
        "rtp_targets": overview["rtp_targets"],
        "window_size": overview["window_size"],
        "reel_num": overview["reel_num"],
        "layout_visible": overview["layout_visible"],
        "default_coin_in": overview["default_coin_in"],
        "normalbet": overview["bet_factors"]["normal"],
        "extrabet": overview["bet_factors"]["extrabet"],
        "featurebuy": overview["bet_factors"]["featurebuy"],
        "superfeaturebuy": overview["bet_factors"]["superfeaturebuy"],
        "mode_normalbet": 0,
        "mode_extrabet": 1,
        "mode_featurebuy": 2,
        "mode_superfeaturebuy": 3,
        "max_spin_free_game": 50,
        "initial_free_spins_low": low_spins,
        "initial_free_spins_high": high_spins,
        "retrigger_free_spins_low": retrigger_low,
        "retrigger_free_spins_high": retrigger_high,
        "super_feature_guaranteed_multiplier": 2500,
        "free_spin_awards": overview["free_spin_awards"],
        "retrigger_awards": overview["retrigger_awards"],
        "strip_name_map": LEGACY_STRIP_NAMES,
        "source_strip_names": SOURCE_SHEETS,
        "arr_reels": arr_reels,
        "arr_reels_weight_cum": arr_weights_cum,
        "reels_len": [item["reel_lengths"] for item in strips],
        "strip_multiplier_weights": strip_multiplier_weights,
        "value_multiplier": default_block["multipliers"],
        "weight_table_normal_bet": default_block["table_weights"]["normal"],
        "weight_table_extra_bet": default_block["table_weights"]["extrabet"],
        "weight_multiplier_fg_low": source_weights[0],
        "weight_multiplier_fg_high": source_weights[1],
        "weight_multiplier_fb_low": source_weights[2],
        "weight_multiplier_fb_high": source_weights[3],
        "weight_multiplier_sb_low": source_weights[4],
        "weight_multiplier_sb_high": source_weights[5],
        "parameter_blocks": parameters,
        "pay_awards": overview["pay_awards"],
        "pay_table": overview["pay_table"],
        "symbol_str": overview["symbol_str"],
        "symbol_codes": [overview["symbol_str"][str(symbol)] for symbol in overview["symbol_id"]],
        "symbol_id": overview["symbol_id"],
        "symbols_special": [0, 1, 2],
        "symbols_score": [symbol for symbol in overview["symbol_id"] if symbol >= 3],
        "card_system": card_system,
    }
    validate_config(config, source_path)
    return config


def validate_config(config: dict[str, Any], source_path: Path) -> None:
    if tuple(config["parameter_blocks"]) != ("A", "B"):
        raise ValueError(f"expected canonical Parameter blocks A/B, got {tuple(config['parameter_blocks'])}")
    if len(config["arr_reels"]) != len(SOURCE_SHEETS):
        raise ValueError("expected 13 reel-strip tables")
    if config["pay_awards"] != [4, 5, 6, 8, 10, 12]:
        raise ValueError(f"unexpected pay awards: {config['pay_awards']}")
    if config["symbol_id"] != list(range(len(config["symbol_id"]))):
        raise ValueError("symbol IDs are not contiguous")
    if len(config["value_multiplier"]) != 15:
        raise ValueError("expected 15 C2 multiplier values")
    if config["super_feature_guaranteed_multiplier"] not in config["value_multiplier"]:
        raise ValueError("Super Feature guaranteed multiplier is missing from C2 multiplier values")
    for key in (
        "weight_multiplier_fg_low",
        "weight_multiplier_fg_high",
        "weight_multiplier_fb_low",
        "weight_multiplier_fb_high",
        "weight_multiplier_sb_low",
        "weight_multiplier_sb_high",
    ):
        if len(config[key]) != len(config["value_multiplier"]) or sum(config[key]) <= 0:
            raise ValueError(f"invalid multiplier weights: {key}")
    for table_index, table in enumerate(config["arr_reels_weight_cum"]):
        for reel in range(config["reel_num"]):
            if table[config["reels_len"][table_index][reel] - 1][reel] <= 0:
                raise ValueError(f"table {table_index}, reel {reel + 1}: no selectable stops")
    filename_match = re.fullmatch(r"H9981(?P<rtp>\d{2})[A-Za-z0-9_-]*\.xlsx", source_path.name, re.IGNORECASE)
    if not filename_match:
        raise ValueError(f"unsupported source filename: {source_path.name}")
    expected_rtp = int(filename_match.group("rtp")) / 100
    if abs(config["rtp_targets"]["normal"] - expected_rtp) > 0.01:
        raise ValueError(
            f"{source_path.name}: Normal Bet RTP {config['rtp_targets']['normal']:.6f} "
            f"does not match the filename's {expected_rtp:.2%} profile"
        )


def derive_output_path(source_path: Path) -> Path:
    match = re.fullmatch(r"H9981(?P<rtp>\d{2})(?P<variant>[A-Za-z0-9_-]*)", source_path.stem, re.IGNORECASE)
    if not match:
        raise ValueError(f"Unsupported xlsx name: {source_path.name}; expected H998192.xlsx or H998194.xlsx")
    return PROJECT_DIR / f"config_{match.group('rtp')}{match.group('variant')}.js"


def load_js_config(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8").strip()
    start = text.find("{")
    end = text.rfind("}")
    if start < 0 or end <= start:
        raise ValueError(f"Unsupported config format: {path}")
    return json.loads(text[start : end + 1])


def write_js_config(path: Path, data: dict[str, Any]) -> None:
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    path.write_text(f"window.H998_BOX_DATA = {payload};\n", encoding="utf-8")


def process_source(source_path: Path, output_path: Path, check: bool) -> None:
    generated = build_config(source_path)
    if check:
        current = load_js_config(output_path)
        if generated != current:
            changed = sorted(key for key in set(generated) | set(current) if generated.get(key) != current.get(key))
            raise ValueError(f"Config mismatch for {output_path.name}: {changed}")
        print(f"Config is current: {output_path}")
        return
    write_js_config(output_path, generated)
    if load_js_config(output_path) != generated:
        raise ValueError(f"Write verification failed: {output_path}")
    print(f"Config written: {output_path} <- {source_path.name}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build H998 config JSON/JS from H998192.xlsx and H998194.xlsx")
    parser.add_argument("--source", "--xlsx", dest="source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--all", action="store_true", help="process every Source/H9981*.xlsx workbook")
    parser.add_argument("--check", action="store_true", help="compare generated data with existing output files")
    args = parser.parse_args()

    sources = (
        sorted(path for path in SOURCE_DIR.glob("H9981*.xlsx") if not path.name.startswith("~$"))
        if args.all
        else [args.source.resolve()]
    )
    if not sources:
        raise FileNotFoundError(f"No H9981*.xlsx files found in {SOURCE_DIR}")
    if args.output and len(sources) != 1:
        raise ValueError("--output can only be used with one --source")

    for source_path in sources:
        if not source_path.is_file():
            raise FileNotFoundError(source_path)
        output_path = args.output.resolve() if args.output else derive_output_path(source_path)
        process_source(source_path, output_path, args.check)
    print(f"Processed {len(sources)} xlsx file(s).")


if __name__ == "__main__":
    main()
