from __future__ import annotations

import argparse
import json
import math
import re
import shutil
from copy import copy
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
PROJECT = HERE.parent
SLOTS = PROJECT.parent
BASE_XLSX = HERE / "H0271.xlsx"
BASE_CONFIG = PROJECT / "config.js"
TEMPLATE = SLOTS / "H016_幸運王牌" / "Source" / "H016192A.xlsx"
THRESHOLD = 1_000_000_000
RANGES = [
    (-1, 0), (0, 1), (1, 2), (2, 3), (3, 4), (4, 5), (5, 6),
    (6, 7), (7, 8), (8, 9), (9, 10), (10, 15), (15, 20),
    (20, 25), (25, 30), (30, 35), (35, 40), (40, 45), (45, 50),
    (50, 60), (60, 70), (70, 80), (80, 90), (90, 100), (100, 120),
    (120, 140), (140, 160), (160, 180), (180, 200), (200, 250),
    (250, 300), (300, 350), (350, 400), (400, 450), (450, 500),
    (500, 550), (550, 600), (600, 650), (650, 700), (700, 750),
    (750, 800), (800, 850), (850, 900), (900, 950), (950, 1000),
    (1000, 2000), (2000, 3000), (3000, 4000), (4000, 5000),
    (5000, 6000), (6000, 7000), (7000, 8000), (8000, 9000),
    (9000, 10000), (10000, 20000), (20000, 30000), (30000, 40000),
    (40000, 50000), (50000, 60000), (60000, 70000), (70000, 80000),
    (80000, 90000), (90000, 100000), (100000, 9999999),
]


def load_js(path: Path) -> dict[str, Any]:
    match = re.search(r"\{.*\}", path.read_text(encoding="utf-8-sig"), re.S)
    if not match:
        raise ValueError(f"Cannot find JSON payload in {path}")
    return json.loads(match.group())


def latest_report(mode: int) -> Path:
    reports = sorted((PROJECT / "Record").glob(f"H0271_*_betmode{mode}_107.xlsx"))
    if not reports:
        raise FileNotFoundError(f"Missing 10^7 Card-Off report for bet mode {mode}")
    return reports[-1]


def report_rows(
    path: Path, count_col: str, pay_col: str, *, exclude_bg_triggers: bool = False
) -> list[dict[str, float]]:
    frame = pd.read_excel(path, sheet_name="Multiplier Line")
    by_upper = {float(row.Interval_Upper): row for row in frame.itertuples()}
    result = []
    previous_trigger_count = 0.0
    previous_trigger_pay = 0.0
    for lower, upper in RANGES:
        row = by_upper.get(float(upper))
        count = float(getattr(row, count_col)) if row is not None else 0.0
        pay = float(getattr(row, pay_col)) if row is not None else 0.0
        if exclude_bg_triggers and row is not None:
            cumulative_count = float(getattr(row, "bg_trigger_fg_cnt_lte_upper"))
            cumulative_pay = float(getattr(row, "bg_trigger_fg_pay_lte_upper"))
            count -= cumulative_count - previous_trigger_count
            pay -= cumulative_pay - previous_trigger_pay
            previous_trigger_count = cumulative_count
            previous_trigger_pay = cumulative_pay
        result.append({"lower": lower, "upper": upper, "count": count, "pay": pay})
    return result


def means_and_mass(rows: list[dict[str, float]], coin_in: float, cap: float) -> tuple[list[float], list[float]]:
    allowed = [row for row in rows if row["upper"] <= cap and row["count"] > 0]
    if not allowed:
        raise ValueError(f"No reachable ranges under cap {cap}")
    values = [row["pay"] / row["count"] / coin_in for row in allowed]
    total = sum(row["count"] for row in allowed)
    mass = [row["count"] / total for row in allowed]
    return values, mass


def integerize(probabilities: list[float], total: int = THRESHOLD) -> list[int]:
    exact = [max(0.0, value) * total for value in probabilities]
    weights = [int(value) for value in exact]
    remainder = total - sum(weights)
    order = sorted(range(len(weights)), key=lambda i: exact[i] - weights[i], reverse=True)
    for index in order[:remainder]:
        weights[index] += 1
    return weights


def tilted_weights(rows: list[dict[str, float]], target: float, coin_in: float, cap: float) -> list[int]:
    allowed_indices = [i for i, row in enumerate(rows) if row["upper"] <= cap and row["count"] > 0]
    values, mass = means_and_mass(rows, coin_in, cap)
    minimum, maximum = min(values), max(values)
    if not minimum <= target <= maximum:
        raise ValueError(f"Target {target:.8f} is outside reachable mean [{minimum:.8f}, {maximum:.8f}] under cap {cap}")

    def distribution(lam: float) -> list[float]:
        logs = [math.log(p) + lam * v for p, v in zip(mass, values)]
        top = max(logs)
        raw = [math.exp(value - top) for value in logs]
        scale = sum(raw)
        return [value / scale for value in raw]

    low, high = -100.0, 100.0
    for _ in range(160):
        middle = (low + high) / 2
        probs = distribution(middle)
        mean = sum(prob * value for prob, value in zip(probs, values))
        if mean < target:
            low = middle
        else:
            high = middle
    allowed_weights = integerize(distribution((low + high) / 2))
    weights = [0] * len(rows)
    for index, weight in zip(allowed_indices, allowed_weights):
        weights[index] = weight
    return weights


def cards(rows: list[dict[str, float]], weights: list[int]) -> list[dict[str, Any]]:
    return [
        {"type": "range", "min": row["lower"], "max": row["upper"], "weight": int(weight)}
        for row, weight in zip(rows, weights)
    ]


def weighted_mean(rows: list[dict[str, float]], weights: list[int], coin_in: float) -> float:
    return sum(
        weight / THRESHOLD * row["pay"] / row["count"] / coin_in
        for row, weight in zip(rows, weights) if weight and row["count"]
    )


def style_like(source, target) -> None:
    target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)


def write_section(ws, start: int, title: str, rows: list[dict[str, float]], weights: list[int], coin_in: float) -> None:
    ws.cell(start - 4, 1).value = title
    ws.cell(start - 2, 1).value = "Simulate"
    ws.cell(start - 2, 2).value = sum(row["count"] for row in rows)
    for offset, (row, weight) in enumerate(zip(rows, weights)):
        r = start + offset
        ws.cell(r, 1).value = f"({row['lower']}, {row['upper']}]"
        ws.cell(r, 2).value = int(row["count"])
        ws.cell(r, 3).value = float(row["pay"])
        ws.cell(r, 4).value = f"=IFERROR(B{r}/$B${start - 2},0)"
        ws.cell(r, 5).value = f"=IFERROR(C{r}/B{r}/{coin_in},0)"
        ws.cell(r, 7).value = f"({row['lower']}, {row['upper']}]"
        ws.cell(r, 8).value = 1
        ws.cell(r, 9).value = f"=IFERROR(B{r}/$B${start - 2},0)"
        ws.cell(r, 10).value = f"=IFERROR(K{r}/$B$2,0)"
        ws.cell(r, 11).value = int(weight)
        ws.cell(r, 12).value = f"=IFERROR(K{r}/$B$2,0)"
        ws.cell(r, 13).value = f"=IFERROR(L{r}*E{r},0)"
        ws.cell(r, 15).value = row["lower"]
        ws.cell(r, 16).value = row["upper"]
    ws.cell(start + 64, 11).value = f"=SUM(K{start}:K{start + 63})"
    ws.cell(start + 64, 13).value = f"=SUM(M{start}:M{start + 63})"


def build_variant(rtp: int, normal_report: Path, buy_report: Path) -> tuple[Path, dict[str, Any]]:
    output = HERE / f"H0271{rtp}A.xlsx"
    shutil.copy2(TEMPLATE, output)
    workbook = load_workbook(output, read_only=False, data_only=False, keep_links=True)
    overview = workbook["Overview"]
    overview["B2"] = f"H0271{rtp}"
    overview["B3"] = "0.0.0.0"
    overview["A6"] = "Base Bet"
    overview["B6"] = "Board Cells"
    overview["A7"] = 100
    overview["B7"] = 30
    overview["A10"] = "Coin in"
    overview["B10"] = "Price(x)"
    overview["C10"] = "Bet Type"
    overview["A11"], overview["B11"], overview["C11"] = 100, 1, "Normal Bet"
    overview["A12"], overview["B12"], overview["C12"] = 200, 2, "Extra Bet"
    overview["A13"], overview["B13"], overview["C13"] = 10000, 100, "Buy Feature"
    for col in range(2, 8):
        overview.cell(15, col).value = col - 1
        overview.cell(16, col).value = 5
    overview["A15"] = "Reel #"
    overview["A16"] = "Visible Window Size"

    normal_bg = report_rows(
        normal_report, "base_game_cnt", "base_game_pay", exclude_bg_triggers=True
    )
    normal_fg = report_rows(normal_report, "free_game_cnt", "free_game_pay")
    buy_package = report_rows(buy_report, "free_game_cnt_BF", "free_game_pay_BF")
    coin_in = 500.0
    # Bonus Game target is 2%. A 100x FG package keeps ranges reachable and
    # retry counts healthy, therefore the corresponding entry cycle is 5,000.
    fg_cycle = 5000.0
    fg_entry_probability = 1.0 / fg_cycle
    fg_entry_weight = round(fg_entry_probability / (1.0 - fg_entry_probability) * THRESHOLD)
    fg_target_mean = 100.0
    newbie_bg_target = 0.93
    oldhand_bg_target = rtp / 100.0
    buy_target_mean = 96.5

    newbie_bg = tilted_weights(normal_bg, newbie_bg_target, coin_in, 30)
    newbie_fg = tilted_weights(normal_fg, fg_target_mean, coin_in, 120)
    # A compact BG support avoids an RTP that only converges through extremely
    # rare 1,000x+ cards. The target is reachable below 30x for both families.
    old_bg = tilted_weights(normal_bg, oldhand_bg_target, coin_in, 30)
    old_fg_small = tilted_weights(normal_fg, fg_target_mean, coin_in, 20000)
    old_fg_medium = list(old_fg_small)
    old_fg_big = list(old_fg_small)
    # H016 layout shares the Feature card table between player profiles. Use
    # the strictest 120x cap so the same table is valid for Newbie as well.
    buy_small = tilted_weights(buy_package, buy_target_mean, coin_in, 120)
    buy_medium = list(buy_small)
    buy_big = list(buy_small)

    detail = workbook["Detail"]
    detail["B2"] = THRESHOLD
    detail["B3"] = 100
    write_section(detail, 15, f"Oldhand Normal Bet BG ({rtp} Game RTP family)", normal_bg, old_bg, coin_in)
    write_section(detail, 86, "Oldhand Normal Bet FG", normal_fg, old_fg_small, coin_in)
    write_section(detail, 163, "Oldhand Buy Feature", buy_package, buy_small, coin_in)
    write_section(detail, 234, "Oldhand Buy Feature Big Bet (independent weights)", buy_package, buy_big, coin_in)
    detail["A7"], detail["B7"] = "Normal Bet", 1
    detail["C7"] = f"=SUM(M15:M78)"
    detail["D7"] = f"=SUM(M86:M149)*{fg_entry_probability:.15f}"
    detail["E7"] = "=SUM(C7:D7)"
    detail["F7"] = fg_entry_probability
    detail["G7"] = "=1/F7"
    detail["A8"], detail["B8"] = "Buy Feature", 100
    detail["C8"] = 0
    detail["D8"] = "=SUM(M163:M226)/B8"
    detail["E8"] = "=SUM(C8:D8)"

    newbie = workbook["Detail_Newbie"]
    newbie["B2"] = THRESHOLD
    newbie["B3"] = 100
    write_section(newbie, 15, "Newbie Normal Bet BG", normal_bg, newbie_bg, coin_in)
    write_section(newbie, 86, "Newbie Normal Bet FG", normal_fg, newbie_fg, coin_in)
    newbie["A7"], newbie["B7"] = "Normal Bet", 1
    newbie["C7"] = "=SUM(M15:M78)"
    newbie["D7"] = f"=SUM(M86:M149)*{fg_entry_probability:.15f}"
    newbie["E7"] = "=SUM(C7:D7)"
    newbie["F7"] = fg_entry_probability
    newbie["G7"] = "=1/F7"

    mw = workbook["Multiplier_Weight"]
    mw["A1"] = "H027 Card Range Weight Summary"
    for row, (lower, upper) in enumerate(RANGES, 4):
        mw.cell(row, 1).value = f"({lower}, {upper}]"
        mw.cell(row, 2).value = newbie_bg[row - 4]
        mw.cell(row, 3).value = newbie_fg[row - 4]
        mw.cell(row, 4).value = old_bg[row - 4]
        mw.cell(row, 5).value = old_fg_small[row - 4]
        mw.cell(row, 6).value = buy_small[row - 4]
        mw.cell(row, 7).value = buy_big[row - 4]

    if workbook.calculation is not None:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    workbook.save(output)
    workbook.close()

    base = load_js(BASE_CONFIG)
    base.update({
        "model": f"H0271{rtp}",
        "parsheet_id": f"H0271{rtp}",
        "excel_version": "0.0.0.0",
        "runtime_version": "0.0.0.0",
        "rtp_label": rtp,
        "config_type": "rtp_variant",
        "config_code": f"{rtp}A",
        "source_xlsx": "H0271.xlsx",
        "source_multiplier_xlsx": output.name,
    })
    free_card = {"type": "free_game", "weight": fg_entry_weight}
    newbie_normal = {"weight_bg": cards(normal_bg, newbie_bg) + [free_card], "weight_fg": cards(normal_fg, newbie_fg)}

    def old_normal(fg_weights: list[int]) -> dict[str, Any]:
        return {"weight_bg": cards(normal_bg, old_bg) + [dict(free_card)], "weight_fg": cards(normal_fg, fg_weights)}

    def feature(weights: list[int]) -> dict[str, Any]:
        return {"weight_fg": cards(buy_package, weights)}

    base["card_system"] = {
        "enabled": True,
        "retry_limit": 10000,
        "weight_threshold": THRESHOLD,
        "card_multiplier_denominator": "normal_bet_base_cost",
        "fg_entry_cycle_target": fg_cycle,
        "newbie": {
            "normal_bet": newbie_normal,
            "buy_feature": feature(buy_small),
        },
        "oldhand": {
            "normal_bet": {
                "small_bet": old_normal(old_fg_small),
                "medium_bet": old_normal(old_fg_medium),
                "big_bet": old_normal(old_fg_big),
            },
            "buy_feature": {
                "small_bet": feature(buy_small),
                "medium_bet": feature(buy_medium),
                "big_bet": feature(buy_big),
            },
        },
    }
    diagnostics = {
        "rtp_family": rtp,
        "newbie_bg_mean": weighted_mean(normal_bg, newbie_bg, coin_in),
        "oldhand_bg_mean": weighted_mean(normal_bg, old_bg, coin_in),
        "fg_package_mean": weighted_mean(normal_fg, old_fg_small, coin_in),
        "fg_entry_probability": fg_entry_weight / (THRESHOLD + fg_entry_weight),
        "buy_package_mean": weighted_mean(buy_package, buy_small, coin_in),
        "normal_report": normal_report.name,
        "buy_report": buy_report.name,
    }
    base["card_system"]["calibration"] = diagnostics
    return output, base


def write_config(path: Path, config: dict[str, Any]) -> None:
    payload = json.dumps(config, ensure_ascii=False, indent=2)
    path.write_text(f"// Generated from Source/{config['source_multiplier_xlsx']} by Source/rtp_variant_sync.py.\nconst data = {payload};\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build/check H027 92A and 94A multiplier workbooks and configs")
    parser.add_argument("--check", action="store_true", help="Compare generated configs with current files")
    args = parser.parse_args()
    if not TEMPLATE.is_file():
        raise FileNotFoundError(TEMPLATE)
    normal = latest_report(0)
    buy = latest_report(2)
    results = []
    for rtp in (92, 94):
        workbook, config = build_variant(rtp, normal, buy)
        config_path = PROJECT / f"config_{rtp}A.js"
        if args.check and config_path.is_file():
            current = load_js(config_path)
            if current != config:
                raise ValueError(f"{config_path.name} differs from generated data")
        else:
            write_config(config_path, config)
        results.append({"xlsx": workbook.name, "config": config_path.name, **config["card_system"]["calibration"]})
    print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
