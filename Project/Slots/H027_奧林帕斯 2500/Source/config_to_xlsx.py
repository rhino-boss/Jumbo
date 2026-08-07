import argparse
import copy
import json
from pathlib import Path

from openpyxl import load_workbook

from xlsx_to_config import (
    DEFAULT_CARD_SOURCE,
    DEFAULT_OUTPUT,
    DEFAULT_SOURCE,
    STRIP_SHEETS,
    build_config,
    find_bet_row,
    find_first_bet_row,
    find_row,
)


BASE_DIR = Path(__file__).resolve().parent


def load_js_config(path):
    text = path.read_text(encoding="utf-8-sig").strip()
    start = text.find("{")
    end = text.rfind("}")
    if start < 0 or end <= start:
        raise ValueError(f"Unsupported config format: {path}")
    return json.loads(text[start : end + 1])


def save_atomic(workbook, path):
    temporary = path.with_name(path.stem + ".tmp.xlsx")
    workbook.save(temporary)
    workbook.close()
    check = load_workbook(temporary, read_only=True, data_only=False)
    check.close()
    temporary.replace(path)


def write_overview(workbook, config):
    ws = workbook["Overview"]
    ws["B2"] = config["source_model"]
    ws["B3"] = config["excel_version"]
    ws["A7"] = config["default_coin_in"]
    visible_row = find_row(ws, 1, "Visible Window Size")
    ws.cell(visible_row, 2).value = config["window_size"]

    normal_row = find_first_bet_row(ws, "Normal Bet Oldhand", "Normal Bet")
    newbie_row = find_first_bet_row(ws, "Normal Bet Newbie", "Normal Bet")
    extra_row = find_bet_row(ws, "Extra Bet")
    feature_row = find_bet_row(ws, "Buy Feature")
    for row, bet in (
        (normal_row, config["normalbet"]),
        (newbie_row, config["normalbet"]),
        (extra_row, config["extrabet"]),
        (feature_row, config["featurebuy"]),
    ):
        ws.cell(row, 2).value = bet
    components = config.get("rtp_components", {})
    for mode, bg_cell, fg_cell in (
        ("normal", "B18", "B19"),
        ("extra", "B22", "B23"),
        ("featurebuy", "B26", "B27"),
    ):
        ws[bg_cell] = components.get(mode, {}).get("bg")
        ws[fg_cell] = components.get(mode, {}).get("fg")

    pay_header = find_row(ws, 1, "Pay Table：")
    for index, (code, symbol_id, values) in enumerate(
        zip(config["symbol_codes"], config["symbol_ids"], config["pay_table"]),
        start=pay_header + 2,
    ):
        ws.cell(index, 1).value = code
        ws.cell(index, 9).value = symbol_id
        for column, value in enumerate(values, start=3):
            ws.cell(index, column).value = value


def write_weight_block(ws, label, block):
    label_col = None
    anchor = None
    for candidate in (10, 7):
        try:
            anchor = find_row(ws, candidate, label)
            label_col = candidate
            break
        except ValueError:
            continue
    if anchor is None:
        raise ValueError(f"Could not find {label!r} in {ws.title}")
    for index, value in enumerate(block["multipliers"]):
        ws.cell(anchor + 1, label_col + 1 + index).value = value
    for row_offset, table_name in enumerate(block["table_names"], start=2):
        ws.cell(anchor + row_offset, label_col).value = table_name
        values = block["weights"][table_name]
        for index, value in enumerate(values):
            ws.cell(anchor + row_offset, label_col + 1 + index).value = value


def write_parameter(workbook, config):
    ws = workbook["Parameter"]
    profile = config["parameter"]["normal"]

    anchor = find_row(ws, 2, "weight_base_game_table")
    for offset, (name, weight) in enumerate(
        zip(profile["base_reel_names"], profile["base_reel_weights"]), start=2
    ):
        ws.cell(anchor + offset, 2).value = name
        ws.cell(anchor + offset, 3).value = weight

    anchor = find_row(ws, 2, "free_game_table_setting")
    free = profile["free_table"]
    for offset, (name, initial, retrigger) in enumerate(
        zip(free["names"], free["initial"], free["retrigger"]), start=2
    ):
        ws.cell(anchor + offset, 2).value = name
        ws.cell(anchor + offset, 3).value = initial
        ws.cell(anchor + offset, 4).value = retrigger

    anchor = find_row(ws, 2, "weight_use_super_multiplier")
    use_c3 = profile["use_c3"]
    for offset, (name, weight) in enumerate(zip(use_c3["table_names"], use_c3["weights"]), start=2):
        ws.cell(anchor + offset, 2).value = name
        reel_weights = use_c3.get("weights_by_reel", {}).get(name, [weight] * 6)
        for reel, column in enumerate(range(3, 9)):
            ws.cell(anchor + offset, column).value = reel_weights[reel]

    anchor = find_row(ws, 2, "multiplier_level")
    for offset, value in enumerate(config["multiplier_levels"], start=2):
        ws.cell(anchor + offset, 2).value = offset - 1
        ws.cell(anchor + offset, 3).value = value

    write_weight_block(ws, "weight_super_multiplier", config["parameter"]["super_multiplier"])
    write_weight_block(ws, "weight_C2_multiplier", profile["c2"])
    write_weight_block(ws, "weight_C3_multiplier", profile["c3"])


def write_strips(workbook, config):
    id_to_code = dict(zip(config["symbol_ids"], config["symbol_codes"]))
    for sheet_name, strip in zip(config["strip_names"], config["strips"]):
        if sheet_name not in STRIP_SHEETS:
            continue
        ws = workbook[sheet_name]
        for column in range(19, 25):
            ws.column_dimensions[ws.cell(3, column).column_letter].width = 7
        for column in range(26, 32):
            ws.column_dimensions[ws.cell(3, column).column_letter].width = 11
        max_length = max(strip["reel_lengths"])
        for offset in range(max_length + 1):
            row = 4 + offset
            ws.cell(row, 11).value = offset if offset < max_length else None
            for reel in range(6):
                symbol_cell = ws.cell(row, 12 + reel)
                weight_cell = ws.cell(row, 26 + reel)
                if offset < strip["reel_lengths"][reel]:
                    symbol_cell.value = id_to_code[strip["symbols"][offset][reel]]
                    weight_cell.value = strip["weights"][offset][reel]
                else:
                    symbol_cell.value = None
                    weight_cell.value = None
                id_cell = ws.cell(row, 19 + reel)
                id_cell.value = f'=IF({symbol_cell.coordinate}="","",VLOOKUP({symbol_cell.coordinate},$A$4:$I$15,9,FALSE))'
                id_cell.number_format = "0"


def write_cards(workbook, card_system):
    ws = workbook["Multiplier_Weight"]
    paths = {
        3: card_system["newbie"]["normal_bet"]["weight_bg"],
        4: card_system["newbie"]["normal_bet"]["weight_fg"],
        5: card_system["oldhand"]["normal_bet"]["weight_bg"],
        6: card_system["oldhand"]["normal_bet"]["weight_fg"],
        7: card_system["oldhand"]["buy_feature"]["weight_fg"],
    }
    for column, cards in paths.items():
        for offset, card in enumerate(cards, start=4):
            ws.cell(offset, column).value = card["weight"]


def compare_config(expected, actual):
    changed = [key for key in expected if expected[key] != actual.get(key)]
    extra = [key for key in actual if key not in expected]
    return changed, extra


def main():
    parser = argparse.ArgumentParser(description="Write H027 config data back to H027 xlsx files")
    parser.add_argument("--config", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--card-source", type=Path, default=DEFAULT_CARD_SOURCE)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--card-output", type=Path)
    parser.add_argument("--in-place", action="store_true")
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()

    config_path = args.config.resolve()
    source_path = args.source.resolve()
    card_source_path = args.card_source.resolve()
    config = load_js_config(config_path)

    if args.check:
        current = build_config(source_path, card_source_path)
        changed, extra = compare_config(config, current)
        if changed or extra:
            raise ValueError(f"Config/xlsx mismatch. Changed keys: {changed}; extra keys: {extra}")
        print("Config and xlsx are round-trip consistent.")
        return

    output_path = source_path if args.in_place else (args.output.resolve() if args.output else source_path.with_name(source_path.stem + "_from_config.xlsx"))
    card_output_path = card_source_path if args.in_place else (args.card_output.resolve() if args.card_output else card_source_path.with_name(card_source_path.stem + "_from_config.xlsx"))

    workbook = load_workbook(source_path, read_only=False, data_only=False)
    write_overview(workbook, config)
    write_parameter(workbook, config)
    write_strips(workbook, config)
    save_atomic(workbook, output_path)

    if config.get("card_system", {}).get("enabled"):
        card_workbook = load_workbook(card_source_path, read_only=False, data_only=False)
        write_cards(card_workbook, config["card_system"])
        save_atomic(card_workbook, card_output_path)

    expected = copy.deepcopy(config)
    expected["model"] = output_path.stem
    generated = build_config(output_path, card_output_path)
    changed, extra = compare_config(expected, generated)
    if changed or extra:
        raise ValueError(f"Round-trip verification failed. Changed keys: {changed}; extra keys: {extra}")
    print(f"Xlsx round-trip verified: {output_path}")
    if config.get("card_system", {}).get("enabled"):
        print(f"Card xlsx round-trip verified: {card_output_path}")


if __name__ == "__main__":
    main()
