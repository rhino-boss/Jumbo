from __future__ import annotations

import argparse
import json
import re
from copy import deepcopy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import rtp_variant_sync as model


HERE = Path(__file__).resolve().parent
PROJECT = HERE.parent


def load_js(path: Path) -> dict[str, Any]:
    match = re.search(r"\{.*\}", path.read_text(encoding="utf-8-sig"), re.S)
    if not match:
        raise ValueError(f"Cannot find JSON payload in {path}")
    return json.loads(match.group())


def workbook_weights(sheet, start: int) -> list[int]:
    values = [sheet.cell(row, 11).value for row in range(start, start + 64)]
    if any(not isinstance(value, (int, float)) or int(value) != value or value < 0 for value in values):
        raise ValueError(f"{sheet.title}!K{start}:K{start + 63} must be non-negative integers")
    result = [int(value) for value in values]
    if sum(result) != model.THRESHOLD:
        raise ValueError(f"{sheet.title}!K{start}:K{start + 63} sums to {sum(result):,}, expected {model.THRESHOLD:,}")
    return result


def make_cards(weights: list[int]) -> list[dict[str, Any]]:
    return [
        {"type": "range", "min": lower, "max": upper, "weight": weight}
        for (lower, upper), weight in zip(model.RANGES, weights)
    ]


def config_from_xlsx(path: Path) -> dict[str, Any]:
    match = re.fullmatch(r"H0271(92|94)A", path.stem)
    if not match:
        raise ValueError(f"Unsupported RTP workbook name: {path.name}")
    rtp = int(match.group(1))
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        overview = workbook["Overview"]
        expected_model = f"H0271{rtp}"
        if str(overview["B2"].value).strip() != expected_model:
            raise ValueError(f"Overview!B2 must be {expected_model}")
        version = str(overview["B3"].value).strip()
        if not re.fullmatch(r"\d+\.\d+\.\d+\.\d+", version):
            raise ValueError("Overview!B3 must be a four-part numeric version")
        if version.split(".")[0] != "0":
            raise ValueError("RTP workbook major version must match H0271.xlsx version 0")
        detail = workbook["Detail"]
        newbie = workbook["Detail_Newbie"]
        newbie_bg = workbook_weights(newbie, 15)
        newbie_fg = workbook_weights(newbie, 86)
        old_bg = workbook_weights(detail, 15)
        old_fg = workbook_weights(detail, 86)
        buy = workbook_weights(detail, 163)
        probability = float(newbie["F7"].value)
        if not 0 < probability < 1:
            raise ValueError("Detail_Newbie!F7 must be the FG entry probability")
        entry_weight = round(probability / (1 - probability) * model.THRESHOLD)
    finally:
        workbook.close()

    base = load_js(model.BASE_CONFIG)
    base.update({
        "model": f"H0271{rtp}",
        "parsheet_id": f"H0271{rtp}",
        "excel_version": version,
        "runtime_version": version,
        "rtp_label": rtp,
        "config_type": "rtp_variant",
        "config_code": f"{rtp}A",
        "source_xlsx": "H0271.xlsx",
        "source_multiplier_xlsx": path.name,
    })
    free_card = {"type": "free_game", "weight": entry_weight}

    def normal(bg: list[int], fg: list[int]) -> dict[str, Any]:
        return {"weight_bg": make_cards(bg) + [deepcopy(free_card)], "weight_fg": make_cards(fg)}

    def feature() -> dict[str, Any]:
        return {"weight_fg": make_cards(buy)}

    base["card_system"] = {
        "enabled": True,
        "retry_limit": 10000,
        "weight_threshold": model.THRESHOLD,
        "card_multiplier_denominator": "normal_bet_base_cost",
        "fg_entry_cycle_target": 1 / probability,
        "newbie": {"normal_bet": normal(newbie_bg, newbie_fg), "buy_feature": feature()},
        "oldhand": {
            "normal_bet": {
                tier: normal(old_bg, old_fg)
                for tier in ("small_bet", "medium_bet", "big_bet")
            },
            "buy_feature": {
                tier: feature()
                for tier in ("small_bet", "medium_bet", "big_bet")
            },
        },
    }
    normal_report = model.latest_report(0)
    buy_report = model.latest_report(2)
    normal_bg_rows = model.report_rows(normal_report, "base_game_cnt", "base_game_pay", exclude_bg_triggers=True)
    normal_fg_rows = model.report_rows(normal_report, "free_game_cnt", "free_game_pay")
    buy_rows = model.report_rows(buy_report, "free_game_cnt_BF", "free_game_pay_BF")
    base["card_system"]["calibration"] = {
        "rtp_family": rtp,
        "newbie_bg_mean": model.weighted_mean(normal_bg_rows, newbie_bg, 500),
        "oldhand_bg_mean": model.weighted_mean(normal_bg_rows, old_bg, 500),
        "fg_package_mean": model.weighted_mean(normal_fg_rows, old_fg, 500),
        "fg_entry_probability": entry_weight / (model.THRESHOLD + entry_weight),
        "buy_package_mean": model.weighted_mean(buy_rows, buy, 500),
        "normal_report": normal_report.name,
        "buy_report": buy_report.name,
    }
    return base


def write_config(path: Path, config: dict[str, Any]) -> None:
    payload = json.dumps(config, ensure_ascii=False, indent=2)
    path.write_text(
        f"// Generated from Source/{config['source_multiplier_xlsx']} by Source/rtp_xlsx_config.py.\n"
        f"const data = {payload};\n",
        encoding="utf-8",
    )


def write_weights(sheet, start: int, cards: list[dict[str, Any]]) -> None:
    range_cards = [card for card in cards if card.get("type", "range") == "range"]
    if len(range_cards) != 64:
        raise ValueError("Config range card list must contain exactly 64 entries")
    for offset, card in enumerate(range_cards):
        lower, upper = model.RANGES[offset]
        if float(card["min"]) != lower or float(card["max"]) != upper:
            raise ValueError(f"Range order mismatch at index {offset}")
        sheet.cell(start + offset, 11).value = int(card["weight"])


def config_to_xlsx(config_path: Path, xlsx_path: Path) -> None:
    config = load_js(config_path)
    workbook = load_workbook(xlsx_path, read_only=False, data_only=False, keep_links=True)
    try:
        detail = workbook["Detail"]
        newbie = workbook["Detail_Newbie"]
        cs = config["card_system"]
        write_weights(newbie, 15, cs["newbie"]["normal_bet"]["weight_bg"])
        write_weights(newbie, 86, cs["newbie"]["normal_bet"]["weight_fg"])
        small = cs["oldhand"]["normal_bet"]["small_bet"]
        write_weights(detail, 15, small["weight_bg"])
        write_weights(detail, 86, small["weight_fg"])
        write_weights(detail, 163, cs["oldhand"]["buy_feature"]["small_bet"]["weight_fg"])
        write_weights(detail, 234, cs["oldhand"]["buy_feature"]["big_bet"]["weight_fg"])
        workbook["Overview"]["B3"] = str(config["excel_version"])
        if workbook.calculation is not None:
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
        workbook.save(xlsx_path)
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="H027 RTP XLSX/config bidirectional converter")
    parser.add_argument("command", choices=("export", "import", "check"))
    parser.add_argument("--rtp", type=int, choices=(92, 94), action="append")
    args = parser.parse_args()
    rtps = args.rtp or [92, 94]
    results = []
    for rtp in rtps:
        xlsx = HERE / f"H0271{rtp}A.xlsx"
        config_path = PROJECT / f"config_{rtp}A.js"
        if args.command == "import":
            config_to_xlsx(config_path, xlsx)
        generated = config_from_xlsx(xlsx)
        if args.command == "export":
            write_config(config_path, generated)
        else:
            current = load_js(config_path)
            if current != generated:
                changed = [key for key in generated if current.get(key) != generated[key]]
                raise ValueError(f"{config_path.name} differs from {xlsx.name}; changed keys: {changed}")
        results.append({"xlsx": xlsx.name, "config": config_path.name, "status": "PASS"})
    print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
