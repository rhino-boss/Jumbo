"""Refresh dynamic H027-vs-competitor sections from a Simulator report."""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_RECORD = ROOT / "Record" / "H0271_0001_2608111732_betmode0_106.xlsx"
DEFAULT_COMPETITOR = ROOT / "其他" / "參考資料" / "analysis_gates_of_olympus_1000_metrics.json"
DEFAULT_REPORT = ROOT / "其他" / "競品參考數值比較.md"


def pct(value: float) -> str:
    return f"{value * 100:.4f}%"


def pp(value: float) -> str:
    return f"{value * 100:+.4f} pp"


def table(headers: list[str], rows: list[list[object]]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "|" + "|".join("---:" if index >= 2 else "---" for index in range(len(headers))) + "|",
    ]
    lines.extend("| " + " | ".join(str(item) for item in row) + " |" for row in rows)
    return "\n".join(lines)


def replace_between(text: str, start: str, end: str | None, replacement: str) -> str:
    left = text.index(start)
    right = text.index(end, left) if end else len(text)
    return text[:left] + replacement.rstrip() + "\n\n" + text[right:]


def read_record(path: Path) -> dict:
    workbook = load_workbook(path, data_only=True, read_only=True)
    base = dict(workbook["Base Info"].iter_rows(min_row=2, values_only=True))
    cascades = list(workbook["Cascade"].iter_rows(min_row=2, values_only=True))
    multipliers = list(workbook["C2-C3 Multiplier"].iter_rows(min_row=2, values_only=True))
    workbook.close()
    return {"base": base, "cascades": cascades, "multipliers": multipliers}


def competitor_combo_counts(rows: list[dict]) -> dict[str, int]:
    result = {str(row["combo"]): int(row["count"]) for row in rows}
    result["9+"] = result.get("9", 0) + result.get("10+", 0)
    return result


def simulator_combo_counts(rows: list[tuple]) -> tuple[dict[str, int], dict[str, int]]:
    bg, fg = {}, {}
    for combo, bg_count, fg_count in rows:
        key = str(combo) if int(combo) < 9 else "9+"
        bg[key] = bg.get(key, 0) + int(bg_count)
        fg[key] = fg.get(key, 0) + int(fg_count)
    return bg, fg


def value_counts(rows: list[tuple]) -> tuple[dict[int, int], dict[int, int]]:
    bg, fg = defaultdict(int), defaultdict(int)
    for value, bg_count, fg_count in rows:
        if int(value) <= 1000:
            bg[int(value)] += int(bg_count)
            fg[int(value)] += int(fg_count)
    return dict(bg), dict(fg)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--record", type=Path, default=DEFAULT_RECORD)
    parser.add_argument("--competitor", type=Path, default=DEFAULT_COMPETITOR)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    args = parser.parse_args()

    record = read_record(args.record.resolve())
    h = record["base"]
    competitor = json.loads(args.competitor.read_text(encoding="utf-8"))["analysis"]
    c = competitor["basic"]
    fg_spins = round(float(h["avg_fg_spins"]) * int(h["fg_trigger_count"]))
    h_cycle = 1 / float(h["fg_trigger_rate"])

    basis = f"""### 比較基準

| 項目 | Gates of Olympus 1000 | H027 現在版本 |
|---|---|---|
| 來源 | `game_responses-gates of olympus 1000.xlsx` | `{args.record.name}`（`config_92A.js` + `Simulator.py`） |
| 版本 | 競品實際 Response 樣本 | `H0271` / `{h['version']}` |
| 樣本 | {c['paid_spins']:,} 個付費 Spin、{c['fg_sessions']:,} 場 FG、{c['fg_spins']:,} 個 FG Spin | {int(h['total_rounds']):,} 個 Normal Bet Round、{int(h['fg_trigger_count']):,} 場 FG、{fg_spins:,} 個 FG Spin |
| Card System | 競品實際遊玩資料 | 關閉 |
| Bet Mode | 一般投注 | Normal Bet，Bet Multi 1 |
| 輪帶 | 實際盤面觀測分布 | BG 為 `BG_Symbol` + `BG_Symbol (2)` 1:1；FG 為 `FG_Symbol` + `FG_Symbol (2)`，初始 Spins 排程 8:7；各 Table 的 R1～R6 各 300 格 |

H027 數值為本次 {int(h['total_rounds']):,} 局模擬樣本，不是理論精確值。競品 FG 只有 {c['fg_sessions']} 場，FG RTP、FG 平均倍數與最大得分的波動會明顯高於 H027 樣本。"""

    core_rows = [
        ["RTP", "Total RTP", pct(c["rtp_total"]), pct(h["rtp_total"]), pp(h["rtp_total"] - c["rtp_total"])],
        ["RTP", "BG RTP", pct(c["rtp_bg"]), pct(h["rtp_bg"]), pp(h["rtp_bg"] - c["rtp_bg"])],
        ["RTP", "FG RTP", pct(c["rtp_fg"]), pct(h["rtp_fg"]), pp(h["rtp_fg"] - c["rtp_fg"])],
        ["Hit Rate", "BG Hit Rate", pct(c["hit_rate_bg"]), pct(h["hit_rate_bg"]), pp(h["hit_rate_bg"] - c["hit_rate_bg"])],
        ["Hit Rate", "FG Hit Rate", pct(c["hit_rate_fg"]), pct(h["hit_rate_fg"]), pp(h["hit_rate_fg"] - c["hit_rate_fg"])],
        ["FG", "FG Trigger Rate", pct(c["fg_trigger_rate"]), pct(h["fg_trigger_rate"]), pp(h["fg_trigger_rate"] - c["fg_trigger_rate"])],
        ["FG", "平均觸發週期", f"{c['fg_cycle']:.2f} 局／次", f"{h_cycle:.2f} 局／次", f"H027 慢 {h_cycle - c['fg_cycle']:.2f} 局"],
        ["FG", "平均 Free Spins", f"{c['avg_fg_spins']:.2f}", f"{h['avg_fg_spins']:.4f}", f"{h['avg_fg_spins'] - c['avg_fg_spins']:+.4f}"],
        ["FG", "FG 平均得分倍數", f"{c['avg_fg_multiplier']:.2f}×", f"{h['avg_fg_multiplier']:.2f}×", f"{h['avg_fg_multiplier'] - c['avg_fg_multiplier']:+.2f}×"],
        ["Win", "單次付費事件最大得分", f"{c['max_score_multiplier']:.2f}×", f"{h['max_win_x']:,.2f}×", f"{h['max_win_x'] - c['max_score_multiplier']:+,.2f}×"],
    ]
    core = "### 核心指標比較\n\n" + table(
        ["類別", "指標", "競品樣本", "H027 現在版本", "差異"], core_rows
    ) + (
        "\n\nH027 的 BG RTP、BG Hit Rate 與 FG 週期已接近競品方向；"
        "主要差距集中在 FG Hit Rate、FG RTP 與 FG 平均得分。"
        "目前仍未校正正式 RTP，以上為輪帶排列調整後的副作用紀錄。"
    )

    cbg = competitor_combo_counts(competitor["combo"]["bg"])
    cfg = competitor_combo_counts(competitor["combo"]["fg"])
    hbg, hfg = simulator_combo_counts(record["cascades"])
    c_bg_den = sum(cbg.get(str(value), 0) for value in range(1, 9)) + cbg["9+"]
    c_fg_den = sum(cfg.get(str(value), 0) for value in range(1, 9)) + cfg["9+"]
    h_bg_den = sum(hbg.get(str(value), 0) for value in range(1, 9)) + hbg.get("9+", 0)
    h_fg_den = sum(hfg.get(str(value), 0) for value in range(1, 9)) + hfg.get("9+", 0)
    combo_rows = []
    for label in [str(value) for value in range(1, 9)] + ["9+"]:
        crb, chb = cbg.get(label, 0) / c_bg_den, hbg.get(label, 0) / h_bg_den
        crf, chf = cfg.get(label, 0) / c_fg_den, hfg.get(label, 0) / h_fg_den
        combo_rows.append([label, pct(crb), pct(chb), pp(chb - crb), pct(crf), pct(chf), pp(chf - crf)])
    combo = f"""## 消除率

比較只納入「至少發生 1 次得獎消除」的 BG／FG Spin，Combo 0 不納入分母。各 Scene 的 Combo 1～9+ 合計為 100%；`9+` 合併所有消除 9 次以上的 Spin。競品分母為 BG {c_bg_den:,} Spin、FG {c_fg_den:,} Spin；H027 分母為 BG {h_bg_den:,} Spin、FG {h_fg_den:,} Spin。

{table(['Combo', '競品 BG', 'H027 BG', 'BG 差異', '競品 FG', 'H027 FG', 'FG 差異'], combo_rows)}

H027 BG Combo 1 與競品的差異為 {pp(hbg.get('1', 0) / h_bg_den - cbg.get('1', 0) / c_bg_den)}；FG Combo 1 的差異為 {pp(hfg.get('1', 0) / h_fg_den - cfg.get('1', 0) / c_fg_den)}。高 Combo 尾端仍受輪帶群組排列與掉落次數影響。"""

    cballs = competitor["multiplier_ball"]
    h_bg_ball_spins = round(float(h["multiplier_ball_rate_bg"]) * int(h["total_rounds"]))
    h_fg_ball_spins = round(float(h["multiplier_ball_rate_fg"]) * fg_spins)
    h_bg_balls = round(float(h["avg_multiplier_balls_bg"]) * int(h["total_rounds"]))
    h_fg_balls = round(float(h["avg_multiplier_balls_fg"]) * fg_spins)
    balls = f"""## 倍數球出現率

採 Spin-level 口徑：該次 BG／FG Spin 的初始盤面或任一次掉落中，至少出現 1 顆倍數球，該 Spin 計數 1 次。同一顆球跨 Cascade 留盤不重複計數。H027 目前 C3 自然出現率為 0，因此下表 H027 來源全為 C2。

{table(
    ['Scene', '競品 Spin 數', '競品有球 Spin', '競品出現率', 'H027 Spin 數', 'H027 有球 Spin', 'H027 出現率', '差異'],
    [
        ['BG', f"{cballs['bg']['spin_count']:,}", f"{cballs['bg']['spins_with_ball']:,}", pct(cballs['bg']['spin_appearance_rate']), f"{int(h['total_rounds']):,}", f"{h_bg_ball_spins:,}", pct(h['multiplier_ball_rate_bg']), pp(h['multiplier_ball_rate_bg'] - cballs['bg']['spin_appearance_rate'])],
        ['FG', f"{cballs['fg']['spin_count']:,}", f"{cballs['fg']['spins_with_ball']:,}", pct(cballs['fg']['spin_appearance_rate']), f"{fg_spins:,}", f"{h_fg_ball_spins:,}", pct(h['multiplier_ball_rate_fg']), pp(h['multiplier_ball_rate_fg'] - cballs['fg']['spin_appearance_rate'])],
    ],
)}

{table(
    ['Scene', '競品去重後球數', '競品平均球數／Spin', 'H027 球數', 'H027 平均球數／Spin', '平均球數差異'],
    [
        ['BG', f"{cballs['bg']['deduplicated_ball_count']:,}", f"{cballs['bg']['balls_per_spin']:.4f}", f"{h_bg_balls:,}", f"{h['avg_multiplier_balls_bg']:.4f}", f"{h['avg_multiplier_balls_bg'] - cballs['bg']['balls_per_spin']:+.4f}"],
        ['FG', f"{cballs['fg']['deduplicated_ball_count']:,}", f"{cballs['fg']['balls_per_spin']:.4f}", f"{h_fg_balls:,}", f"{h['avg_multiplier_balls_fg']:.4f}", f"{h['avg_multiplier_balls_fg'] - cballs['fg']['balls_per_spin']:+.4f}"],
    ],
)}

H027 BG 倍數球出現率差異為 {pp(h['multiplier_ball_rate_bg'] - cballs['bg']['spin_appearance_rate'])}；FG 差異為 {pp(h['multiplier_ball_rate_fg'] - cballs['fg']['spin_appearance_rate'])}。"""

    h_bg_values, h_fg_values = value_counts(record["multipliers"])
    c_bg_values = {int(row["value"]): int(row["count"]) for row in cballs["bg"]["value_distribution"]}
    c_fg_values = {int(row["value"]): int(row["count"]) for row in cballs["fg"]["value_distribution"]}
    values = sorted(set(c_bg_values) | set(c_fg_values) | set(h_bg_values) | set(h_fg_values))
    c_bg_total, c_fg_total = sum(c_bg_values.values()), sum(c_fg_values.values())
    h_bg_total, h_fg_total = sum(h_bg_values.values()), sum(h_fg_values.values())
    value_rows = []
    for value in values:
        cb, hb = c_bg_values.get(value, 0) / c_bg_total, h_bg_values.get(value, 0) / h_bg_total
        cf, hf = c_fg_values.get(value, 0) / c_fg_total, h_fg_values.get(value, 0) / h_fg_total
        value_rows.append([f"{value}×", pct(cb), pct(hb), pp(hb - cb), pct(cf), pct(hf), pp(hf - cf)])
    multiplier = f"""## 倍數球上的倍率分布

口徑為各倍率球數 ÷ 該 Scene 全部倍數球數。H027 倍率權重依競品觀測球數寫入；表內差異為模擬抽樣波動。競品倍率池包含 1000×，但本批競品與 H027 樣本都沒有抽到；2500× 不是競品倍率球池成員，因此不列入。

{table(['倍率', '競品 BG', 'H027 BG', 'BG 差異', '競品 FG', 'H027 FG', 'FG 差異'], value_rows)}

倍率分布仍與競品權重一致；後續若要控制得分，應調整賠率或 FG 累積倍數機制，而不是改動已對齊的倍率權重。"""

    text = args.report.read_text(encoding="utf-8")
    text = replace_between(text, "### 比較基準", "### 核心指標比較", basis)
    text = replace_between(text, "### 核心指標比較", "## 符號分布", core)
    text = replace_between(text, "## 消除率", "## 倍數球出現率", combo)
    text = replace_between(text, "## 倍數球出現率", "## 倍數球上的倍率分布", balls)
    text = replace_between(text, "## 倍數球上的倍率分布", None, multiplier)
    args.report.write_text(text.rstrip() + "\n", encoding="utf-8")
    print(f"Updated: {args.report.resolve()}")


if __name__ == "__main__":
    main()
