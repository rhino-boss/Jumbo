import argparse
import copy
import json
import re
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
DEFAULT_SOURCE = BASE_DIR / "H0271.xlsx"
DEFAULT_CARD_SOURCE = BASE_DIR / "H027194A.xlsx"
DEFAULT_OUTPUT = PROJECT_DIR / "config_92A.js"

METADATA = {
    "game_id": "101027",
    "parsheet_id": "H0271",
    "display_name": "Olympus 2500",
    "game_name_zh": "奧林帕斯 2500",
    "mode_normalbet": 0,
    "mode_extrabet": 1,
    "mode_featurebuy": 2,
    "supported_bet_modes": [0, 1, 2],
    "extra_bet_multiplier": 2,
    "extra_fg_probability_multiplier": 5,
    "has_super_feature": False,
    "has_wild": False,
    "has_jackpot": True,
    "multiplier_max_value": 2500,
    "max_free_spins": 50,
    "reference_presentation": "參考資料/260630_Olympus 2500.pptx",
    "rule_document": "game_rule.md",
    "model_status": "rules_confirmed_math_draft",
    "pending_math_items": [
        "Extra Bet dedicated reel and card weights",
        "Formal RTP target confirmation and final calibration",
        "C3 multiplier pool and appearance weights",
    ],
}

STRIP_SHEETS = [
    "BG_Symbol",
    "BG_Symbol (2)",
    "BG_Symbol (3)",
    "FG_Symbol",
    "FG_Symbol (2)",
    "FG_Symbol (3)",
    "BF_Symbol",
]


def to_int(value, default=0):
    if value is None or value == "":
        return default
    return int(value)


def cumulative(values):
    total = 0
    result = []
    for value in values:
        total += int(value)
        result.append(total)
    return result


def find_row(ws, column, value):
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, column).value).strip() == value:
            return row
    raise ValueError(f"Could not find {value!r} in {ws.title}")


def find_bet_row(ws, label):
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 4).value).strip() == label:
            return row
    raise ValueError(f"Could not find bet type {label!r} in {ws.title}")


def find_first_bet_row(ws, *labels):
    for label in labels:
        try:
            return find_bet_row(ws, label)
        except ValueError:
            continue
    raise ValueError(f"Could not find any bet type {labels!r} in {ws.title}")


def to_float(value, default=None):
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def parse_overview(ws):
    pay_header = find_row(ws, 1, "Pay Table：")
    symbol_codes = []
    symbol_ids = []
    pay_table = []
    row = pay_header + 2
    while ws.cell(row, 1).value is not None:
        symbol_codes.append(str(ws.cell(row, 1).value).strip())
        symbol_ids.append(to_int(ws.cell(row, 9).value))
        pay_table.append([to_int(ws.cell(row, col).value) for col in range(3, 9)])
        row += 1

    visible_row = find_row(ws, 1, "Visible Window Size")
    # Older H019-derived sheets split Normal Bet into Newbie/Oldhand rows.
    # Support both split Newbie/Oldhand rows and one-row-per-mode layouts so
    # the converter follows the workbook instead of depending on stale labels.
    normal_row = find_first_bet_row(ws, "Normal Bet Oldhand", "Normal Bet")
    newbie_row = find_first_bet_row(ws, "Normal Bet Newbie", "Normal Bet")
    extra_row = find_first_bet_row(ws, "Extra Bet")
    featurebuy_row = find_first_bet_row(ws, "Buy Feature")
    return {
        "model": str(ws["B2"].value).strip(),
        "excel_version": str(ws["B3"].value).strip(),
        "default_coin_in": to_int(ws["A7"].value),
        "reel_num": 6,
        "window_size": to_int(ws.cell(visible_row, 2).value),
        "symbol_codes": symbol_codes,
        "symbol_ids": symbol_ids,
        "pay_table": pay_table,
        "pay_count_bounds": [8, 10, 12],
        "scatter_pay_counts": [4, 5, 6],
        "normalbet": to_int(ws.cell(normal_row, 2).value),
        "extrabet": to_int(ws.cell(extra_row, 2).value),
        "featurebuy": to_int(ws.cell(featurebuy_row, 2).value),
        "rtp_targets": {
            "normal_newbie": to_float(ws.cell(newbie_row, 3).value),
            "normal_oldhand": to_float(ws.cell(normal_row, 3).value),
            "normal": to_float(ws.cell(normal_row, 3).value),
            "extrabet": to_float(ws.cell(extra_row, 3).value),
            "featurebuy": to_float(ws.cell(featurebuy_row, 3).value),
        },
        "rtp_components": {
            "normal": {"bg": to_float(ws["B18"].value), "fg": to_float(ws["B19"].value)},
            "extra": {"bg": to_float(ws["B22"].value), "fg": to_float(ws["B23"].value)},
            "featurebuy": {"bg": to_float(ws["B26"].value), "fg": to_float(ws["B27"].value)},
        },
    }


def parse_named_weights(ws, label_col, weight_col, start_row):
    names = []
    weights = []
    row = start_row
    while ws.cell(row, label_col).value is not None:
        names.append(str(ws.cell(row, label_col).value).strip())
        weights.append(to_int(ws.cell(row, weight_col).value))
        row += 1
    return names, weights


def parse_free_table(ws, name_col, free_col, retrigger_col, start_row):
    names = []
    free = []
    retrigger = []
    row = start_row
    while ws.cell(row, name_col).value is not None:
        names.append(str(ws.cell(row, name_col).value).strip())
        free.append(to_int(ws.cell(row, free_col).value))
        retrigger.append(to_int(ws.cell(row, retrigger_col).value))
        row += 1
    return {"names": names, "initial": free, "retrigger": retrigger}


def parse_horizontal_weight_block(ws, label, label_cols=(10, 7)):
    anchor_row = None
    label_col = None
    for candidate in label_cols:
        try:
            anchor_row = find_row(ws, candidate, label)
            label_col = candidate
            break
        except ValueError:
            continue
    if anchor_row is None:
        raise ValueError(f"Could not find {label!r} in {ws.title}")
    header_row = anchor_row + 1
    multipliers = []
    col = label_col + 1
    while ws.cell(header_row, col).value is not None:
        multipliers.append(to_int(ws.cell(header_row, col).value))
        col += 1

    names = []
    weights = {}
    row = header_row + 1
    while ws.cell(row, label_col).value is not None:
        name = str(ws.cell(row, label_col).value).strip()
        if name.startswith("#"):
            break
        names.append(name)
        weights[name] = [to_int(ws.cell(row, label_col + 1 + index).value) for index in range(len(multipliers))]
        row += 1
    return {
        "multipliers": multipliers,
        "table_names": names,
        "weights": weights,
        "weights_cum": {name: cumulative(values) for name, values in weights.items()},
    }


def parse_named_single_weights(ws, label, label_col=2, weight_col=3):
    anchor_row = find_row(ws, label_col, label)
    names = []
    weights = []
    row = anchor_row + 2
    while ws.cell(row, label_col).value is not None:
        name = str(ws.cell(row, label_col).value).strip()
        if name.startswith("#"):
            break
        names.append(name)
        weights.append(to_int(ws.cell(row, weight_col).value))
        row += 1
    return {"table_names": names, "weights": weights, "weights_cum": cumulative(weights)}


def parse_use_c3_weights(ws):
    anchor_row = find_row(ws, 2, "weight_use_super_multiplier")
    names = []
    weights = []
    weights_by_reel = {}
    row = anchor_row + 2
    while ws.cell(row, 2).value is not None:
        name = str(ws.cell(row, 2).value).strip()
        if name.startswith("#"):
            break
        reel_weights = [to_int(ws.cell(row, column).value) for column in range(3, 9)]
        names.append(name)
        weights.append(reel_weights[0])
        weights_by_reel[name] = reel_weights
        row += 1
    return {
        "table_names": names,
        "weights": weights,
        "weights_by_reel": weights_by_reel,
        "denominator": 10000,
    }


def parse_multiplier_levels(ws):
    anchor_row = find_row(ws, 2, "multiplier_level")
    levels = []
    row = anchor_row + 2
    while ws.cell(row, 2).value is not None:
        levels.append(to_int(ws.cell(row, 3).value))
        row += 1
    return levels


def parse_parameter(ws):
    base = parse_named_single_weights(ws, "weight_base_game_table")
    free_anchor = find_row(ws, 2, "free_game_table_setting")
    free_table = parse_free_table(ws, 2, 3, 4, free_anchor + 2)
    use_c3 = parse_use_c3_weights(ws)
    c2 = parse_horizontal_weight_block(ws, "weight_C2_multiplier")
    c3 = parse_horizontal_weight_block(ws, "weight_C3_multiplier")
    super_multiplier = parse_horizontal_weight_block(ws, "weight_super_multiplier")
    multiplier_levels = parse_multiplier_levels(ws)

    normal = {
        "base_reel_names": base["table_names"],
        "base_reel_weights": base["weights"],
        "base_reel_weights_cum": base["weights_cum"],
        "free_table": free_table,
        "use_c3": use_c3,
        "c2": c2,
        "c3": c3,
    }
    featurebuy = copy.deepcopy(normal)
    featurebuy["base_reel_names"] = ["BF_Symbol"]
    featurebuy["base_reel_weights"] = [1]
    featurebuy["base_reel_weights_cum"] = [1]
    return {
        "multiplier_levels": multiplier_levels,
        "super_multiplier": super_multiplier,
        "normal": normal,
        "featurebuy": featurebuy,
    }


def parse_card_range(label):
    if label is None:
        return None
    match = re.fullmatch(r"\(\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*\]", str(label).strip())
    if not match:
        return None
    return float(match.group(1)), float(match.group(2))


def parse_card_system(ws):
    profile_map = {
        3: ("newbie", "normal_bet", "weight_bg"),
        4: ("newbie", "normal_bet", "weight_fg"),
        5: ("oldhand", "normal_bet", "weight_bg"),
        6: ("oldhand", "normal_bet", "weight_fg"),
        7: ("oldhand", "buy_feature", "weight_fg"),
    }
    profiles = {
        "newbie": {"normal_bet": {"weight_bg": [], "weight_fg": []}},
        "oldhand": {
            "normal_bet": {"weight_bg": [], "weight_fg": []},
            "buy_feature": {"weight_fg": []},
        },
    }
    row = 4
    while ws.cell(row, 2).value is not None:
        label = str(ws.cell(row, 2).value).strip()
        range_pair = parse_card_range(label)
        for col, (player, mode, segment) in profile_map.items():
            weight = to_int(ws.cell(row, col).value)
            card = None
            if label.lower() == "free game":
                card = {"type": "free_game", "weight": weight}
            elif range_pair is not None:
                card = {"type": "range", "min": range_pair[0], "max": range_pair[1], "weight": weight}
            if card is not None:
                profiles[player][mode][segment].append(card)
        row += 1
    return {"enabled": True, "retry_limit": 5000, **profiles}


def disabled_card_system():
    return {
        "enabled": False,
        "retry_limit": 0,
        "newbie": {"normal_bet": {"weight_bg": [], "weight_fg": []}},
        "oldhand": {
            "normal_bet": {"weight_bg": [], "weight_fg": []},
            "buy_feature": {"weight_fg": []},
        },
    }


def parse_strip_sheet(ws, code_to_id):
    reel_symbols = [[] for _ in range(6)]
    reel_weights = [[] for _ in range(6)]
    row = 4
    while ws.cell(row, 11).value is not None:
        for reel, col in enumerate(range(12, 18)):
            code = ws.cell(row, col).value
            if code is None or str(code).strip() == "":
                continue
            code = str(code).strip()
            if code == "WW":
                continue
            if code not in code_to_id:
                raise ValueError(f"Unknown symbol {code!r} in {ws.title}!{ws.cell(row, col).coordinate}")
            reel_symbols[reel].append(code_to_id[code])
            reel_weights[reel].append(to_int(ws.cell(row, 26 + reel).value))
        row += 1
    reel_lengths = [len(values) for values in reel_symbols]
    max_length = max(reel_lengths, default=0)
    rows = [[reel_symbols[reel][index] if index < reel_lengths[reel] else -1 for reel in range(6)] for index in range(max_length)]
    weights = [[reel_weights[reel][index] if index < reel_lengths[reel] else 0 for reel in range(6)] for index in range(max_length)]
    return {"symbols": rows, "weights": weights, "reel_lengths": reel_lengths}


def build_config(source_path, card_source_path=None):
    workbook = load_workbook(source_path, read_only=False, data_only=True)
    missing_sheets = [name for name in STRIP_SHEETS if name not in workbook.sheetnames]
    if missing_sheets:
        workbook.close()
        joined = ", ".join(missing_sheets)
        raise ValueError(
            f"{source_path.name} is missing reel worksheets required by Parameter: {joined}. "
            "Provide the complete H027 parsheet before regenerating config."
        )
    overview = parse_overview(workbook["Overview"])
    parameter = parse_parameter(workbook["Parameter"])
    if "Multiplier_Weight" in workbook.sheetnames:
        card_system = parse_card_system(workbook["Multiplier_Weight"])
    elif card_source_path is not None and card_source_path.exists():
        card_workbook = load_workbook(card_source_path, read_only=False, data_only=True)
        if "Multiplier_Weight" not in card_workbook.sheetnames:
            card_workbook.close()
            workbook.close()
            raise ValueError(f"{card_source_path.name} has no Multiplier_Weight worksheet")
        card_system = parse_card_system(card_workbook["Multiplier_Weight"])
        card_workbook.close()
    else:
        card_system = disabled_card_system()
    code_to_id = dict(zip(overview["symbol_codes"], overview["symbol_ids"]))
    strip_data = [parse_strip_sheet(workbook[name], code_to_id) for name in STRIP_SHEETS]
    workbook.close()

    config = dict(METADATA)
    config.update(overview)
    # Use the actual source filename as the model name so generated reports and
    # configs retain the selected PAR sheet variant.
    config["source_model"] = overview["model"]
    config["model"] = source_path.stem
    config["multiplier_levels"] = parameter["multiplier_levels"]
    config["parameter"] = parameter
    config["card_system"] = card_system
    config["strip_names"] = STRIP_SHEETS
    config["strips"] = strip_data
    return config


def derive_output_path(source_path):
    if source_path.stem.upper() == "H0271":
        return DEFAULT_OUTPUT
    match = re.fullmatch(r"H0271(?P<rtp>\d{2})(?P<variant>[A-Za-z0-9_-]*)", source_path.stem, re.IGNORECASE)
    if not match:
        raise ValueError(f"Unsupported xlsx name: {source_path.name}; expected H0271.xlsx or a name such as H027192A.xlsx")
    return PROJECT_DIR / f"config_{match.group('rtp')}{match.group('variant')}.js"


def load_js_config(path):
    text = path.read_text(encoding="utf-8").strip()
    start = text.find("{")
    end = text.rfind("}")
    if start < 0 or end <= start:
        raise ValueError(f"Unsupported config format: {path}")
    return json.loads(text[start : end + 1])


def write_js_config(path, data):
    path.write_text("const data = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")


def process_source(source_path, output_path, check=False, card_source_path=None):
    generated = build_config(source_path, card_source_path)
    if check:
        current = load_js_config(output_path)
        if generated != current:
            changed = [key for key in generated if generated[key] != current.get(key)]
            raise ValueError(f"Config mismatch for {output_path.name}: {changed}")
        print(f"Config is current: {output_path}")
    else:
        write_js_config(output_path, generated)
        if load_js_config(output_path) != generated:
            raise ValueError(f"Write verification failed: {output_path}")
        print(f"Config written: {output_path} <- {source_path.name}")


def main():
    parser = argparse.ArgumentParser(description="Build H027 config files from xlsx")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--card-source", type=Path, default=DEFAULT_CARD_SOURCE)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--all", action="store_true")
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()

    candidates = sorted(path for path in BASE_DIR.glob("H0271*.xlsx") if not path.name.startswith("~$"))
    sources = []
    if args.all:
        for path in candidates:
            workbook = load_workbook(path, read_only=True, data_only=False)
            is_complete = "Parameter" in workbook.sheetnames and all(name in workbook.sheetnames for name in STRIP_SHEETS)
            workbook.close()
            if is_complete:
                sources.append(path)
    else:
        sources = [args.source.resolve()]
    if not sources:
        raise FileNotFoundError(f"No H0271*.xlsx files found in {BASE_DIR}")

    for source_path in sources:
        output_path = args.output.resolve() if args.output and len(sources) == 1 else derive_output_path(source_path)
        process_source(source_path, output_path, args.check, args.card_source.resolve() if args.card_source else None)
    print(f"Processed {len(sources)} xlsx file(s).")


if __name__ == "__main__":
    main()
