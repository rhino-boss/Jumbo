"""H027 model <-> config sync, following the H028 single-tool workflow.

    model_sync.py export [--check]       H0271.xlsx -> config_92A.js
    model_sync.py import [--check]       config_92A.js -> H0271.xlsx

Only BG_Symbol and FG_Symbol are part of the H027 config model.  Historical
"(2)" / "(3)" worksheets may remain in the workbook, but are deliberately
ignored by both directions.
"""
import argparse
import json
import math
import os
import posixpath
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = BASE_DIR / "H0271.xlsx"
DEFAULT_CONFIG = BASE_DIR.parent / "config_92A.js"
SYMBOL_SHEETS = ("BG_Symbol", "FG_Symbol")
REEL_COUNT = 6
STRIP_LENGTH = 300
LEVEL_COUNT = 25

FIXED_METADATA = {
    "game_id": "101027",
    "parsheet_id": "H0271",
    "display_name": "Olympus 2500",
    "game_name_zh": "奧林帕斯 2500",
    "mode_normalbet": 0,
    "mode_extrabet": 1,
    "mode_featurebuy": 2,
    "supported_bet_modes": [0, 1, 2],
    "extra_bet_multiplier": 2,
    "extra_fg_probability_multiplier": 5,
    "has_super_feature": False,
    "has_wild": False,
    "has_jackpot": True,
    "max_free_spins": 50,
    "reference_presentation": "參考資料/260630_Olympus 2500.pptx",
    "rule_document": "game_rule.md",
    "model_status": "rules_confirmed_math_draft",
    "pending_math_items": [
        "Extra Bet dedicated reel and card weights",
        "Formal RTP target confirmation and final calibration",
        "C3 multiplier pool and appearance weights",
    ],
}

CELL_PATTERN = re.compile(
    r'<c(?P<selfattrs>\s+[^>]*?\br="(?P<selfref>[A-Z]{1,3}[1-9][0-9]*)"[^>]*)/>|'
    r'<c(?P<attrs>\s+[^>]*?\br="(?P<ref>[A-Z]{1,3}[1-9][0-9]*)"[^>/]*)>'
    r'(?P<body>.*?)</c>',
    re.DOTALL,
)


def cumulative(values):
    result, total = [], 0
    for value in values:
        total += int(value)
        result.append(total)
    return result


def require_int(value, label):
    if value is None or isinstance(value, bool):
        raise ValueError(f"{label} must be numeric, got {value!r}")
    number = int(value)
    if float(value) != number:
        raise ValueError(f"{label} must be an integer, got {value!r}")
    return number


def load_js_config(path):
    text = path.read_text(encoding="utf-8-sig")
    match = re.fullmatch(r"\s*const\s+data\s*=\s*(\{.*\})\s*;?\s*", text, re.DOTALL)
    if not match:
        raise ValueError(f"Unsupported config format: {path}")
    return json.loads(match.group(1))


def write_js_config(path, config):
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + ".tmp")
    temporary.write_text(
        "const data = " + json.dumps(config, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    os.replace(temporary, path)


def symbol_map(sheet):
    result = {}
    for row in range(4, 16):
        code = sheet.cell(row, 1).value
        symbol_id = sheet.cell(row, 9).value
        if code is not None and symbol_id is not None:
            result[str(code)] = require_int(symbol_id, f"{sheet.title}!I{row}")
    if not result:
        raise ValueError(f"No symbol mapping found in {sheet.title}!A4:I15")
    return result


def read_weights(sheet, row):
    return [require_int(sheet.cell(row, col).value, f"{sheet.title}!{sheet.cell(row, col).coordinate}")
            for col in range(11, 36)]


def make_multiplier_table(levels, names, weights):
    return {
        "multipliers": levels,
        "table_names": names,
        "weights": weights,
        "weights_cum": {name: cumulative(weights[name]) for name in names},
    }


def read_parameter(sheet):
    levels = [require_int(sheet.cell(row, 3).value, f"Parameter!C{row}")
              for row in range(28, 53)]
    header_sets = ([sheet.cell(3, col).value for col in range(11, 36)],
                   [sheet.cell(8, col).value for col in range(11, 36)],
                   [sheet.cell(18, col).value for col in range(11, 36)])
    for row_number, header in zip((3, 8, 18), header_sets):
        parsed = [require_int(value, f"Parameter row {row_number}") for value in header]
        if parsed != levels:
            raise ValueError(f"Parameter multiplier headers on row {row_number} do not match C28:C52")

    super_weights = {"Super Ball": read_weights(sheet, 4)}
    c2_weights = {"BG_Symbol": read_weights(sheet, 9), "FG_Symbol": read_weights(sheet, 12)}
    c3_weights = {"BG_Symbol": read_weights(sheet, 19), "FG_Symbol": read_weights(sheet, 22)}
    use_c3_by_reel = {
        "BG_Symbol": [require_int(sheet.cell(18, col).value, f"Parameter!{sheet.cell(18, col).coordinate}") for col in range(3, 9)],
        "FG_Symbol": [require_int(sheet.cell(21, col).value, f"Parameter!{sheet.cell(21, col).coordinate}") for col in range(3, 9)],
    }
    initial = require_int(sheet["C11"].value, "Parameter!C11")
    retrigger = require_int(sheet["D11"].value, "Parameter!D11")
    profile = {
        "base_reel_names": ["BG_Symbol"],
        "base_reel_weights": [require_int(sheet["C4"].value, "Parameter!C4")],
        "base_reel_weights_cum": cumulative([require_int(sheet["C4"].value, "Parameter!C4")]),
        "free_table": {"names": ["FG_Symbol"], "initial": [initial], "retrigger": [retrigger]},
        "use_c3": {
            "table_names": list(SYMBOL_SHEETS),
            "weights": [use_c3_by_reel[name][0] for name in SYMBOL_SHEETS],
            "weights_by_reel": use_c3_by_reel,
            "denominator": 10000,
        },
        "c2": make_multiplier_table(levels, list(SYMBOL_SHEETS), c2_weights),
        "c3": make_multiplier_table(levels, list(SYMBOL_SHEETS), c3_weights),
    }
    return levels, {
        "multiplier_levels": levels,
        "super_multiplier": make_multiplier_table(levels, ["Super Ball"], super_weights),
        "normal": profile,
        "featurebuy": json.loads(json.dumps(profile)),
    }


def read_strip(sheet):
    mapping = symbol_map(sheet)
    symbols, weights = [], []
    for row in range(4, 4 + STRIP_LENGTH):
        symbol_row, weight_row = [], []
        for reel in range(REEL_COUNT):
            code = sheet.cell(row, 12 + reel).value
            if str(code) not in mapping:
                raise ValueError(f"Unknown symbol {code!r} in {sheet.title}!{sheet.cell(row, 12 + reel).coordinate}")
            symbol_row.append(mapping[str(code)])
            weight_row.append(require_int(sheet.cell(row, 26 + reel).value,
                                          f"{sheet.title}!{sheet.cell(row, 26 + reel).coordinate}"))
        symbols.append(symbol_row)
        weights.append(weight_row)
    return {"symbols": symbols, "weights": weights, "reel_lengths": [STRIP_LENGTH] * REEL_COUNT}


def build_config(source_path):
    workbook = load_workbook(source_path, read_only=False, data_only=False)
    try:
        missing = [name for name in ("Overview", "Parameter", *SYMBOL_SHEETS) if name not in workbook.sheetnames]
        if missing:
            raise ValueError(f"Missing required worksheet(s): {missing}")
        overview = workbook["Overview"]
        model = str(overview["B2"].value)
        version = str(overview["B3"].value)
        symbol_codes = [str(overview.cell(row, 1).value) for row in range(30, 42)]
        symbol_ids = [require_int(overview.cell(row, 9).value, f"Overview!I{row}") for row in range(30, 42)]
        pay_table = [[require_int(overview.cell(row, col).value, f"Overview!{overview.cell(row, col).coordinate}")
                      for col in range(3, 9)] for row in range(30, 42)]
        windows = [require_int(overview.cell(17, col).value, f"Overview!{overview.cell(17, col).coordinate}")
                   for col in range(2, 8)]
        if len(set(windows)) != 1:
            raise ValueError(f"H027 config supports one window_size, got {windows}")
        levels, parameter = read_parameter(workbook["Parameter"])
        config = dict(FIXED_METADATA)
        config.update({
            "multiplier_max_value": max(levels),
            "model": model,
            "excel_version": version,
            "default_coin_in": require_int(overview["A7"].value, "Overview!A7"),
            "reel_num": REEL_COUNT,
            "window_size": windows[0],
            "symbol_codes": symbol_codes,
            "symbol_ids": symbol_ids,
            "pay_table": pay_table,
            "pay_count_bounds": [8, 10, 12],
            "scatter_pay_counts": [4, 5, 6],
            "normalbet": require_int(overview["B11"].value, "Overview!B11"),
            "extrabet": require_int(overview["B12"].value, "Overview!B12"),
            "featurebuy": require_int(overview["B13"].value, "Overview!B13"),
            "source_model": model,
            "multiplier_levels": levels,
            "parameter": parameter,
            "card_system": {
                "enabled": False,
                "retry_limit": 0,
                "newbie": {"normal_bet": {"weight_bg": [], "weight_fg": []}},
                "oldhand": {
                    "normal_bet": {"weight_bg": [], "weight_fg": []},
                    "buy_feature": {"weight_fg": []},
                },
            },
            "strip_names": list(SYMBOL_SHEETS),
            "strips": [read_strip(workbook[name]) for name in SYMBOL_SHEETS],
        })
        return config
    finally:
        workbook.close()


def compare_config(generated, output_path):
    if not output_path.exists():
        raise FileNotFoundError(output_path)
    existing = load_js_config(output_path)
    if generated != existing:
        changed = [key for key in generated if existing.get(key) != generated[key]]
        extra = [key for key in existing if key not in generated]
        raise ValueError(f"Config differs. Changed keys: {changed}; extra keys: {extra}")


def add_update(updates, sheet, address, value, key):
    updates.setdefault(sheet, {})[address] = (value, key)


def validate_config(config):
    if config.get("strip_names") != list(SYMBOL_SHEETS):
        raise ValueError(f"strip_names must be exactly {list(SYMBOL_SHEETS)}")
    if len(config.get("strips", [])) != 2:
        raise ValueError("config must contain exactly two strips")
    levels = config.get("multiplier_levels", [])
    if len(levels) != LEVEL_COUNT:
        raise ValueError(f"multiplier_levels must contain {LEVEL_COUNT} entries")
    for profile_name in ("normal", "featurebuy"):
        profile = config["parameter"][profile_name]
        if profile["base_reel_names"] != ["BG_Symbol"] or profile["free_table"]["names"] != ["FG_Symbol"]:
            raise ValueError(f"parameter.{profile_name} may only use BG_Symbol and FG_Symbol")


def build_updates(config):
    validate_config(config)
    updates = {}
    levels = config["multiplier_levels"]
    normal = config["parameter"]["normal"]
    add_update(updates, "Overview", "B2", config["model"], "model")
    add_update(updates, "Overview", "B3", config["excel_version"], "excel_version")
    add_update(updates, "Overview", "A7", config["default_coin_in"], "default_coin_in")
    for address, key in (("B11", "normalbet"), ("B12", "extrabet"), ("B13", "featurebuy")):
        add_update(updates, "Overview", address, config[key], key)
    for col in range(2, 8):
        add_update(updates, "Overview", f"{chr(64 + col)}17", config["window_size"], "window_size")
    for index, row in enumerate(range(30, 42)):
        add_update(updates, "Overview", f"A{row}", config["symbol_codes"][index], "symbol_codes")
        add_update(updates, "Overview", f"I{row}", config["symbol_ids"][index], "symbol_ids")
        for offset, col in enumerate("CDEFGH"):
            add_update(updates, "Overview", f"{col}{row}", config["pay_table"][index][offset], "pay_table")

    add_update(updates, "Parameter", "B4", "BG_Symbol", "normal.base_reel_names")
    add_update(updates, "Parameter", "C4", normal["base_reel_weights"][0], "normal.base_reel_weights")
    add_update(updates, "Parameter", "B11", "FG_Symbol", "normal.free_table.names")
    add_update(updates, "Parameter", "C11", normal["free_table"]["initial"][0], "normal.free_table.initial")
    add_update(updates, "Parameter", "D11", normal["free_table"]["retrigger"][0], "normal.free_table.retrigger")
    for row in (5, 6):
        add_update(updates, "Parameter", f"B{row}", None, "unused table")
        add_update(updates, "Parameter", f"C{row}", 0, "unused table")
    for row in (12, 13):
        add_update(updates, "Parameter", f"B{row}", None, "unused table")
        add_update(updates, "Parameter", f"C{row}", 0, "unused table")
        add_update(updates, "Parameter", f"D{row}", 0, "unused table")

    for index, value in enumerate(levels):
        col = column_name(11 + index)
        for row in (3, 8, 18):
            add_update(updates, "Parameter", f"{col}{row}", value, "multiplier_levels")
        add_update(updates, "Parameter", f"B{28 + index}", index + 1, "multiplier level index")
        add_update(updates, "Parameter", f"C{28 + index}", value, "multiplier_levels")

    table_rows = {"BG_Symbol": 18, "FG_Symbol": 21}
    for name, row in table_rows.items():
        add_update(updates, "Parameter", f"B{row}", name, "normal.use_c3.table_names")
        for reel, value in enumerate(normal["use_c3"]["weights_by_reel"][name], start=3):
            add_update(updates, "Parameter", f"{column_name(reel)}{row}", value, "normal.use_c3.weights_by_reel")
    for row in (19, 20, 22, 23):
        add_update(updates, "Parameter", f"B{row}", None, "unused table")
        for col in range(3, 9):
            add_update(updates, "Parameter", f"{column_name(col)}{row}", 0, "unused table")

    multiplier_rows = (("super_multiplier", "Super Ball", 4), ("c2", "BG_Symbol", 9),
                       ("c2", "FG_Symbol", 12), ("c3", "BG_Symbol", 19), ("c3", "FG_Symbol", 22))
    for table_key, name, row in multiplier_rows:
        table = config["parameter"][table_key] if table_key == "super_multiplier" else normal[table_key]
        add_update(updates, "Parameter", f"J{row}", name, f"{table_key}.table_names")
        for index, value in enumerate(table["weights"][name]):
            add_update(updates, "Parameter", f"{column_name(11 + index)}{row}", value, f"{table_key}.weights")
    for row in (10, 11, 13, 14, 20, 21, 23, 24):
        add_update(updates, "Parameter", f"J{row}", None, "unused table")
        for col in range(11, 36):
            add_update(updates, "Parameter", f"{column_name(col)}{row}", 0, "unused table")

    id_to_code = dict(zip(config["symbol_ids"], config["symbol_codes"]))
    for sheet_name, strip in zip(SYMBOL_SHEETS, config["strips"]):
        if len(strip["symbols"]) != STRIP_LENGTH or len(strip["weights"]) != STRIP_LENGTH:
            raise ValueError(f"{sheet_name} must have exactly {STRIP_LENGTH} rows")
        for row_index in range(STRIP_LENGTH):
            if len(strip["symbols"][row_index]) != REEL_COUNT or len(strip["weights"][row_index]) != REEL_COUNT:
                raise ValueError(f"{sheet_name} row {row_index} must have {REEL_COUNT} reels")
            row = 4 + row_index
            for reel in range(REEL_COUNT):
                symbol_id = require_int(strip["symbols"][row_index][reel], f"{sheet_name}.symbols")
                if symbol_id not in id_to_code:
                    raise ValueError(f"Unknown symbol ID {symbol_id} in {sheet_name} row {row_index}")
                add_update(updates, sheet_name, f"{column_name(12 + reel)}{row}", id_to_code[symbol_id], "strips.symbols")
                add_update(updates, sheet_name, f"{column_name(19 + reel)}{row}", symbol_id, "Symbol ID cache")
                add_update(updates, sheet_name, f"{column_name(26 + reel)}{row}",
                           require_int(strip["weights"][row_index][reel], f"{sheet_name}.weights"), "strips.weights")
    return updates


def column_name(index):
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def values_equal(left, right):
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return float(left) == float(right)
    return left == right


def changed_cells(source_path, updates):
    workbook = load_workbook(source_path, read_only=False, data_only=True)
    try:
        changed = []
        for sheet_name, cells in updates.items():
            sheet = workbook[sheet_name]
            for address, (new_value, key) in cells.items():
                old_value = sheet[address].value
                if not values_equal(old_value, new_value):
                    changed.append((sheet_name, address, old_value, new_value, key))
        return changed
    finally:
        workbook.close()


def workbook_sheet_parts(archive):
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels_root}
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    result = {}
    for sheet in workbook_root.findall(f".//{{{main_ns}}}sheet"):
        target = rel_targets[sheet.attrib[f"{{{rel_ns}}}id"]]
        result[sheet.attrib["name"]] = target.lstrip("/") if target.startswith("/") else posixpath.normpath(posixpath.join("xl", target))
    return result


def number_text(value):
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and math.isfinite(value):
        return format(value, ".15g")
    raise TypeError(f"Unsupported numeric value: {value!r}")


def render_cell(attrs, value):
    attrs = re.sub(r'\s+t="[^"]*"', "", attrs.rstrip())
    if value is None:
        return f"<c{attrs}></c>"
    if isinstance(value, str):
        space = ' xml:space="preserve"' if value != value.strip() else ""
        return f'<c{attrs} t="inlineStr"><is><t{space}>{escape(value)}</t></is></c>'
    return f"<c{attrs}><v>{number_text(value)}</v></c>"


def patch_sheet(xml_bytes, sheet_name, cell_updates):
    text = xml_bytes.decode("utf-8")
    remaining = set(cell_updates)

    def replace(match):
        address = match.group("ref") or match.group("selfref")
        if address not in cell_updates:
            return match.group(0)
        attrs = match.group("attrs") or match.group("selfattrs") or ""
        body = match.group("body") or ""
        value, _key = cell_updates[address]
        if re.search(r"<f(?:\s|>)", body):
            if not isinstance(value, (int, float)) or isinstance(value, bool):
                raise TypeError(f"Formula cache must be numeric: {sheet_name}!{address}={value!r}")
            attrs = re.sub(r'\s+t="[^"]*"', "", attrs.rstrip())
            value_xml = f"<v>{number_text(value)}</v>"
            if re.search(r"<v(?:\s[^>]*)?>.*?</v>", body, re.DOTALL):
                body = re.sub(r"<v(?:\s[^>]*)?>.*?</v>", value_xml, body, count=1, flags=re.DOTALL)
            else:
                end = body.find("</f>") + 4
                body = body[:end] + value_xml + body[end:]
            replacement = f"<c{attrs}>{body}</c>"
        else:
            replacement = render_cell(attrs, value)
        remaining.discard(address)
        return replacement

    patched = CELL_PATTERN.sub(replace, text)
    if remaining:
        sample = ", ".join(sorted(remaining)[:10])
        raise ValueError(f"Mapped cell(s) missing from {sheet_name}: {sample}")
    payload = patched.encode("utf-8")
    ET.fromstring(payload)
    return payload


def calculation_metadata(filename, payload):
    if filename == "xl/_rels/workbook.xml.rels":
        return re.sub(r'<Relationship\b[^>]*\bType="[^"]*/calcChain"[^>]*/>', "", payload.decode("utf-8")).encode("utf-8")
    if filename == "[Content_Types].xml":
        return re.sub(r'<Override\b[^>]*\bPartName="/xl/calcChain\.xml"[^>]*/>', "", payload.decode("utf-8")).encode("utf-8")
    if filename == "xl/workbook.xml":
        text = payload.decode("utf-8")
        def update(match):
            attrs = match.group(1)
            for name, value in (("calcMode", "auto"), ("fullCalcOnLoad", "1"), ("forceFullCalc", "1")):
                attrs = re.sub(rf'\b{name}="[^"]*"', f'{name}="{value}"', attrs) if re.search(rf'\b{name}="', attrs) else attrs + f' {name}="{value}"'
            return f"<calcPr{attrs}/>"
        text, count = re.subn(r"<calcPr\b([^>]*)/>", update, text, count=1)
        if not count:
            text = text.replace("</workbook>", '<calcPr calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"/></workbook>')
        return text.encode("utf-8")
    return payload


def write_patched_workbook(source_path, output_path, updates, force=False):
    source_path, output_path = source_path.resolve(), output_path.resolve()
    if output_path.exists() and output_path != source_path and not force:
        raise FileExistsError(f"Output exists; pass --force: {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = output_path.with_name(output_path.name + ".tmp")
    if temporary.exists():
        temporary.unlink()
    try:
        with zipfile.ZipFile(source_path, "r") as source_zip:
            parts = workbook_sheet_parts(source_zip)
            part_updates = {parts[name]: (name, cells) for name, cells in updates.items()}
            with zipfile.ZipFile(temporary, "w") as output_zip:
                for info in source_zip.infolist():
                    if info.filename == "xl/calcChain.xml":
                        continue
                    payload = source_zip.read(info.filename)
                    if info.filename in part_updates:
                        sheet_name, cells = part_updates[info.filename]
                        payload = patch_sheet(payload, sheet_name, cells)
                    payload = calculation_metadata(info.filename, payload)
                    output_zip.writestr(info, payload)
        os.replace(temporary, output_path)
    finally:
        if temporary.exists():
            temporary.unlink()


def verify_output(output_path, config):
    generated = build_config(output_path)
    if generated != config:
        changed = [key for key in generated if config.get(key) != generated[key]]
        extra = [key for key in config if key not in generated]
        raise ValueError(f"Round-trip verification failed. Changed keys: {changed}; extra keys: {extra}")


def run_export(argv):
    parser = argparse.ArgumentParser(description="Build H027 config from H0271.xlsx")
    parser.add_argument("--source", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--output", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args(argv)
    generated = build_config(args.source.resolve())
    if args.check:
        compare_config(generated, args.output.resolve())
        print(f"Config is current: {args.output.resolve()}")
    else:
        write_js_config(args.output.resolve(), generated)
        compare_config(generated, args.output.resolve())
        print(f"Config written and verified: {args.output.resolve()}")


def run_import(argv):
    parser = argparse.ArgumentParser(description="Write H027 config data back to H0271.xlsx")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--source", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--in-place", action="store_true")
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args(argv)
    if args.in_place and args.output:
        parser.error("--in-place and --output cannot be combined")
    config = load_js_config(args.config.resolve())
    updates = build_updates(config)
    changed = changed_cells(args.source.resolve(), updates)
    print(f"Mapped cells: {sum(len(cells) for cells in updates.values()):,}")
    print(f"Changed cells: {len(changed):,}")
    for sheet_name, address, old, new, key in changed[:20]:
        print(f"  {sheet_name}!{address}: {old!r} -> {new!r} ({key})")
    if len(changed) > 20:
        print(f"  ... and {len(changed) - 20:,} more")
    if args.check:
        print("Check only; no xlsx was written.")
        return
    output = args.source.resolve() if args.in_place else (args.output.resolve() if args.output else args.source.resolve().with_name("H0271_from_config.xlsx"))
    write_patched_workbook(args.source.resolve(), output, updates, force=args.force or args.in_place)
    verify_output(output, config)
    print(f"Xlsx written and round-trip verified: {output}")


def main():
    parser = argparse.ArgumentParser(prog="model_sync.py", usage="model_sync.py {export,import} [options]")
    parser.add_argument("command", choices=("export", "import"))
    args, rest = parser.parse_known_args()
    (run_export if args.command == "export" else run_import)(rest)


if __name__ == "__main__":
    main()
