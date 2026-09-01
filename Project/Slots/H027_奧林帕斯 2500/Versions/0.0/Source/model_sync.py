"""H027 model <-> config sync, following the H028 single-tool workflow.

    model_sync.py export [--check]       H0271.xlsx -> config.js
    model_sync.py import [--check]       config.js -> H0271.xlsx

BG uses reconstructed competitor Reel Sets 0-2; FG uses Reel Sets 3-4.
BF_Symbol is the Buy Feature entry alias because the supplied reconstruction
does not contain the competitor's entry-only Reel Set 6. Cascade replacement
continues backward on the selected circular strip; symbol drop-weight tables
are deliberately not part of the runtime model.
"""
import argparse
import json
import math
import os
import posixpath
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = BASE_DIR / "H0271.xlsx"
DEFAULT_CONFIG = BASE_DIR.parent / "config.js"
SYMBOL_SHEETS = (
    "BG_Symbol", "BG_Symbol (2)", "BG_Symbol (3)",
    "FG_Symbol", "FG_Symbol (2)", "BF_Symbol",
)
NORMAL_BG_SHEETS = ("BG_Symbol", "BG_Symbol (2)", "BG_Symbol (3)")
FEATUREBUY_BG_SHEETS = ("BF_Symbol",)
FG_SHEETS = ("FG_Symbol", "FG_Symbol (2)")
REEL_COUNT = 6
MAX_STRIP_LENGTH = 300
LEVEL_COUNT = 25
SOURCE_REEL_SETS = {
    "BG_Symbol": 0, "BG_Symbol (2)": 1, "BG_Symbol (3)": 2,
    "FG_Symbol": 3, "FG_Symbol (2)": 4, "BF_Symbol": 0,
}
REEL_SOURCE_WORKBOOK = r"C:\Users\rhinshen\Mine\個人工作區\市場資訊\H5\遊戲資源\PP - Gates of Olympus 1000\還原輪帶_Gates_of_Olympus_1000.xlsx"

FIXED_METADATA = {
    "game_id": "101027",
    "parsheet_id": "H0271",
    "config_type": "base",
    "config_code": "base",
    "is_competitor_model": True,
    "initial_version_rule": "competitor_model_starts_at_0",
    "display_name": "Olympus 2500",
    "game_name_zh": "奧林帕斯 2500",
    "mode_normalbet": 0,
    "mode_extrabet": 1,
    "mode_featurebuy": 2,
    "supported_bet_modes": [0, 1, 2],
    "extra_bet_multiplier": 2,
    "extra_fg_probability_multiplier": 5,
    "has_super_feature": False,
    "has_wild": False,
    "has_jackpot": True,
    "max_free_spins": 50,
    "fg_trigger_count": 4,
    "fg_retrigger_count": 3,
    "cascade_limit": 100,
    "denom": 0.002,
    "bet_options": [1, 2, 3, 4, 5, 6, 8, 10, 12, 16, 20, 30, 40, 60, 100, 200, 300, 600, 1000, 1500],
    "initial_balance": 10000,
    "drop_mode": "cascade_drop",
    "cascade_symbol_source": "reel_strip",
    "reel_source_workbook": REEL_SOURCE_WORKBOOK,
    "reel_set_usage": {
        "BG": {"sets": [0, 1, 2], "weights": [9698, 19298, 2344]},
        "FG": {"sets": [3, 4], "spin_counts": [6, 9]},
        "BF_FG": {"sets": [3, 4], "spin_counts": [5, 10]},
        "BF_ENTRY": {
            "source": "direct_generator",
            "rule": "exactly four C1 on R2-R5; remaining 26 cells uniformly use nine regular symbols",
        },
    },
    "bet_tier_thresholds": {"small_bet_lt": 2, "medium_bet_lte": 100},
    "link": {"enabled": False},
    "rtp_accounting": {
        "link": "none",
        "bonus_game": "free_game",
        "game": "base_game",
        "target_status": "pending",
    },
    "reference_presentation": "參考資料/260630_Olympus 2500.pptx",
    "rule_document": "game_rule.md",
    "model_status": "rules_confirmed_math_draft",
    "pending_math_items": [
        "Extra Bet dedicated reel and card weights",
        "Formal RTP target confirmation and final calibration",
        "C3 multiplier pool and appearance weights",
    ],
}

CELL_PATTERN = re.compile(
    r'<c(?P<selfattrs>\s+[^>]*?\br="(?P<selfref>[A-Z]{1,3}[1-9][0-9]*)"[^>]*)/>|'
    r'<c(?P<attrs>\s+[^>]*?\br="(?P<ref>[A-Z]{1,3}[1-9][0-9]*)"[^>/]*)>'
    r'(?P<body>.*?)</c>',
    re.DOTALL,
)


def cumulative(values):
    result, total = [], 0
    for value in values:
        total += int(value)
        result.append(total)
    return result


def require_int(value, label):
    if value is None or isinstance(value, bool):
        raise ValueError(f"{label} must be numeric, got {value!r}")
    number = int(value)
    if float(value) != number:
        raise ValueError(f"{label} must be an integer, got {value!r}")
    return number


def validate_super_probability_curve(weights, label, denominator=10000):
    if len(weights) != 6:
        raise ValueError(f"{label} must contain exactly six weights for initial ball counts 1-6")
    if any(value < 0 or value > denominator for value in weights):
        raise ValueError(f"{label} weights must be within 0-{denominator}: {weights}")
    if any(left < right for left, right in zip(weights, weights[1:])):
        raise ValueError(f"{label} per-C2 weights must not increase as initial ball count increases: {weights}")
    visibility = [1.0 - (1.0 - value / denominator) ** count for count, value in enumerate(weights, start=1)]
    if any(left > right + 1e-12 for left, right in zip(visibility, visibility[1:])):
        raise ValueError(
            f"{label} probability of at least one C3 must not decrease as initial ball count increases: "
            f"weights={weights}, visibility={visibility}"
        )
    if visibility[0] >= visibility[-1]:
        raise ValueError(
            f"{label} probability of at least one C3 must be higher at six C2 than at one C2: "
            f"weights={weights}, visibility={visibility}"
        )


def validate_drop_combo_curve(weights, label, denominator=10000):
    if len(weights) != 5:
        raise ValueError(f"{label} must contain exactly five weights for Combo 1, 2, 3, 4 and 5+")
    if any(value < 0 or value > denominator for value in weights):
        raise ValueError(f"{label} weights must be within 0-{denominator}: {weights}")
    if any(left < right for left, right in zip(weights, weights[1:])):
        raise ValueError(f"{label} per-C2 weights must not increase as Combo increases: {weights}")


def load_js_config(path):
    text = path.read_text(encoding="utf-8-sig")
    match = re.fullmatch(r"\s*const\s+data\s*=\s*(\{.*\})\s*;?\s*", text, re.DOTALL)
    if not match:
        raise ValueError(f"Unsupported config format: {path}")
    return json.loads(match.group(1))


def write_js_config(path, config):
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + ".tmp")
    temporary.write_text(
        "const data = " + json.dumps(config, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )
    os.replace(temporary, path)


def symbol_map(sheet):
    result = {}
    for row in range(4, 16):
        code = sheet.cell(row, 1).value
        symbol_id = sheet.cell(row, 9).value
        if code is not None and symbol_id is not None:
            result[str(code)] = require_int(symbol_id, f"{sheet.title}!I{row}")
    if not result:
        raise ValueError(f"No symbol mapping found in {sheet.title}!A4:I15")
    return result


def read_weights(sheet, row):
    return [require_int(sheet.cell(row, col).value, f"{sheet.title}!{sheet.cell(row, col).coordinate}")
            for col in range(11, 36)]


def make_multiplier_table(levels, names, c2_weights, c3_weights):
    return {
        "multipliers": levels,
        "table_names": names,
        "weights_c2": c2_weights,
        "weights_c2_cum": {name: cumulative(c2_weights[name]) for name in names},
        "weights_c3": c3_weights,
        "weights_c3_cum": {name: cumulative(c3_weights[name]) for name in names},
    }


def read_parameter(sheet):
    levels = [require_int(sheet.cell(row, 3).value, f"Parameter!C{row}")
              for row in range(18, 43)]
    header_sets = (
        [sheet.cell(3, col).value for col in range(11, 36)],
        [sheet.cell(13, col).value for col in range(11, 36)],
    )
    for row_number, header in zip((3, 13), header_sets):
        parsed = [require_int(value, f"Parameter row {row_number}") for value in header]
        if parsed != levels:
            raise ValueError(f"Parameter multiplier headers on row {row_number} do not match C18:C42")

    c2_rows = {
        "BG_Symbol": 4, "BG_Symbol (2)": 5, "BG_Symbol (3)": 6,
        "FG_Symbol": 7, "FG_Symbol (2)": 8, "BF_Symbol": 9,
    }
    c3_rows = {
        "BG_Symbol": 14, "BG_Symbol (2)": 15, "BG_Symbol (3)": 16,
        "FG_Symbol": 17, "FG_Symbol (2)": 18, "BF_Symbol": 19,
    }
    use_super_rows = {
        "BG_Symbol": 24, "BG_Symbol (2)": 25, "BG_Symbol (3)": 26,
        "FG_Symbol": 27, "FG_Symbol (2)": 28, "BF_Symbol": 29,
    }
    c2_multiplier_weights = {name: read_weights(sheet, row) for name, row in c2_rows.items()}
    c3_multiplier_weights = {name: read_weights(sheet, row) for name, row in c3_rows.items()}
    use_super_by_initial_count = {
        name: [require_int(sheet.cell(row, col).value, f"Parameter!{sheet.cell(row, col).coordinate}")
               for col in range(11, 17)]
        for name, row in use_super_rows.items()
    }
    for name, weights in use_super_by_initial_count.items():
        validate_super_probability_curve(weights, f"Parameter use Super row {name}")
    drop_combo_labels = [str(sheet.cell(33, col).value) for col in range(11, 16)]
    if drop_combo_labels != ["1", "2", "3", "4", "5+"]:
        raise ValueError(
            "Parameter!K33:O33 must be Combo labels 1, 2, 3, 4, 5+; "
            f"got {drop_combo_labels}"
        )
    drop_combo_rows = {
        "BG_Symbol": 34, "BG_Symbol (2)": 35, "BG_Symbol (3)": 36,
        "FG_Symbol": 37, "FG_Symbol (2)": 38, "BF_Symbol": 39,
    }
    use_super_by_drop_combo = {}
    for name, row in drop_combo_rows.items():
        if str(sheet.cell(row, 10).value) != name:
            raise ValueError(f"Parameter!J{row} must be {name!r}")
        weights = [
            require_int(sheet.cell(row, col).value, f"Parameter!{sheet.cell(row, col).coordinate}")
            for col in range(11, 16)
        ]
        validate_drop_combo_curve(weights, f"Parameter drop Combo row {name}")
        use_super_by_drop_combo[name] = weights
    base_names, base_weights = [], []
    for row in (4, 5, 6):
        name = sheet.cell(row, 2).value
        weight = require_int(sheet.cell(row, 3).value, f"Parameter!C{row}")
        if name is not None:
            if str(name) not in NORMAL_BG_SHEETS:
                raise ValueError(f"Parameter!B{row} must be one of {NORMAL_BG_SHEETS}, got {name!r}")
            base_names.append(str(name))
            base_weights.append(weight)
    if not base_names:
        raise ValueError("Parameter!B4:C6 must contain at least one active BG table")
    free_names, initial_counts, retrigger_counts = [], [], []
    for row in (11, 12):
        name = sheet.cell(row, 2).value
        initial = require_int(sheet.cell(row, 3).value, f"Parameter!C{row}")
        retrigger = require_int(sheet.cell(row, 4).value, f"Parameter!D{row}")
        if name is None:
            if initial or retrigger:
                raise ValueError(f"Parameter!B{row} must name an FG table when counts are non-zero")
            continue
        name = str(name)
        if name not in FG_SHEETS:
            raise ValueError(f"Parameter!B{row} must be one of {FG_SHEETS}, got {name!r}")
        if initial < 0 or retrigger < 0 or ((initial == 0) != (retrigger == 0)):
            raise ValueError(f"Parameter!C{row}:D{row} must both be zero or both be positive")
        free_names.append(name)
        initial_counts.append(initial)
        retrigger_counts.append(retrigger)
    if free_names != list(FG_SHEETS):
        raise ValueError(f"Parameter!B11:B12 must be exactly {list(FG_SHEETS)}")
    if sum(initial_counts) != 15 or sum(retrigger_counts) != 5:
        raise ValueError(
            "Parameter FG table counts must preserve 15 initial spins and +5 retrigger spins"
        )
    profile = {
        "base_reel_names": base_names,
        "base_reel_weights": base_weights,
        "base_reel_weights_cum": cumulative(base_weights),
        "free_table": {"names": free_names, "initial": initial_counts, "retrigger": retrigger_counts},
        "c2_to_c3": {
            "table_names": list(SYMBOL_SHEETS),
            "initial_ball_counts": [1, 2, 3, 4, 5, 6],
            "weights_by_initial_ball_count": use_super_by_initial_count,
            "drop_combo_buckets": ["1", "2", "3", "4", "5+"],
            "weights_by_drop_combo": use_super_by_drop_combo,
            "denominator": 10000,
        },
        "multiplier": make_multiplier_table(
            levels, list(SYMBOL_SHEETS), c2_multiplier_weights, c3_multiplier_weights
        ),
    }
    featurebuy = json.loads(json.dumps(profile))
    featurebuy["base_reel_names"] = list(FEATUREBUY_BG_SHEETS)
    featurebuy["base_reel_weights"] = [1]
    featurebuy["base_reel_weights_cum"] = [1]
    featurebuy["free_table"]["initial"] = [5, 10]
    featurebuy["free_table"]["retrigger"] = [2, 3]
    bf_c2_weights = list(c2_multiplier_weights["BF_Symbol"])
    bf_c3_weights = list(c3_multiplier_weights["BF_Symbol"])
    featurebuy["multiplier"] = make_multiplier_table(
        levels,
        list(SYMBOL_SHEETS),
        {name: list(bf_c2_weights) for name in SYMBOL_SHEETS},
        {name: list(bf_c3_weights) for name in SYMBOL_SHEETS},
    )
    bf_initial = list(use_super_by_initial_count["BF_Symbol"])
    bf_drop = list(use_super_by_drop_combo["BF_Symbol"])
    featurebuy["c2_to_c3"]["weights_by_initial_ball_count"] = {
        name: list(bf_initial) for name in SYMBOL_SHEETS
    }
    featurebuy["c2_to_c3"]["weights_by_drop_combo"] = {
        name: list(bf_drop) for name in SYMBOL_SHEETS
    }
    return levels, {
        "multiplier_levels": levels,
        "normal": profile,
        "featurebuy": featurebuy,
    }


def read_strip(sheet):
    mapping = symbol_map(sheet)
    reel_lengths = []
    for reel in range(REEL_COUNT):
        positive_rows = [
            index for index in range(MAX_STRIP_LENGTH)
            if require_int(
                sheet.cell(4 + index, 26 + reel).value or 0,
                f"{sheet.title}!{sheet.cell(4 + index, 26 + reel).coordinate}",
            ) > 0
        ]
        if not positive_rows:
            raise ValueError(f"{sheet.title} R{reel + 1} has no active reel positions")
        reel_lengths.append(max(positive_rows) + 1)
    row_count = max(reel_lengths)
    symbols, weights = [], []
    for index in range(row_count):
        row = 4 + index
        symbol_row, weight_row = [], []
        for reel in range(REEL_COUNT):
            code = sheet.cell(row, 12 + reel).value
            weight = require_int(sheet.cell(row, 26 + reel).value or 0,
                                 f"{sheet.title}!{sheet.cell(row, 26 + reel).coordinate}")
            if index < reel_lengths[reel] and str(code) not in mapping:
                raise ValueError(f"Unknown symbol {code!r} in {sheet.title}!{sheet.cell(row, 12 + reel).coordinate}")
            if index < reel_lengths[reel] and weight <= 0:
                raise ValueError(f"{sheet.title} active weight must be positive at R{reel + 1} line {index}")
            if index >= reel_lengths[reel] and weight != 0:
                raise ValueError(f"{sheet.title} inactive weight must be zero at R{reel + 1} line {index}")
            symbol_row.append(mapping.get(str(code), next(iter(mapping.values()))))
            weight_row.append(weight)
        symbols.append(symbol_row)
        weights.append(weight_row)
    return {
        "symbols": symbols,
        "weights": weights,
        "reel_lengths": reel_lengths,
        "cascade_source": "reel_strip",
        "source_reel_set": SOURCE_REEL_SETS[sheet.title],
    }


def build_config(source_path):
    workbook = load_workbook(source_path, read_only=False, data_only=False)
    try:
        missing = [name for name in ("Overview", "Parameter", *SYMBOL_SHEETS) if name not in workbook.sheetnames]
        if missing:
            raise ValueError(f"Missing required worksheet(s): {missing}")
        overview = workbook["Overview"]
        model = str(overview["B2"].value)
        version = str(overview["B3"].value)
        if model != "H0271":
            raise ValueError(f"Overview!B2 must be H0271, got {model!r}")
        if not re.fullmatch(r"\d", version):
            raise ValueError(f"Base excel_version must be exactly one digit, got {version!r}")
        symbol_codes = [str(overview.cell(row, 1).value) for row in range(30, 42)]
        symbol_ids = [require_int(overview.cell(row, 9).value, f"Overview!I{row}") for row in range(30, 42)]
        pay_table = [[require_int(overview.cell(row, col).value, f"Overview!{overview.cell(row, col).coordinate}")
                      for col in range(3, 9)] for row in range(30, 42)]
        windows = [require_int(overview.cell(17, col).value, f"Overview!{overview.cell(17, col).coordinate}")
                   for col in range(2, 8)]
        if len(set(windows)) != 1:
            raise ValueError(f"H027 config supports one window_size, got {windows}")
        levels, parameter = read_parameter(workbook["Parameter"])
        strips = [read_strip(workbook[name]) for name in SYMBOL_SHEETS]
        # 2026-08-31 移除「共用加權百分位」（linked stop）：SPS 的 gameSetting 沒有
        # 對應欄位、screenGenerator 也是六輪各自抽停點，這個機制只存在於本地模型，
        # 是兩邊 Hit Rate／FG RTP 對不上的唯一機制性差異。
        config = dict(FIXED_METADATA)
        config.update({
            "multiplier_max_value": max(levels),
            "model": model,
            "excel_version": version,
            "default_coin_in": require_int(overview["A7"].value, "Overview!A7"),
            "reel_num": REEL_COUNT,
            "window_size": windows[0],
            "symbol_codes": symbol_codes,
            "symbol_ids": symbol_ids,
            "pay_table": pay_table,
            "pay_count_bounds": [8, 10, 12],
            "scatter_pay_counts": [4, 5, 6],
            "normalbet": require_int(overview["B11"].value, "Overview!B11"),
            "extrabet": require_int(overview["B12"].value, "Overview!B12"),
            "featurebuy": require_int(overview["B13"].value, "Overview!B13"),
            "source_model": model,
            "multiplier_levels": levels,
            "parameter": parameter,
            "card_system": {
                "enabled": False,
                "retry_limit": 10000,
                "reason": "RTP/Variant source workbook is not available",
            },
            "strip_names": list(SYMBOL_SHEETS),
            "strips": strips,
        })
        return config
    finally:
        workbook.close()


def compare_config(generated, output_path):
    if not output_path.exists():
        raise FileNotFoundError(output_path)
    existing = load_js_config(output_path)
    if generated != existing:
        changed = [key for key in generated if existing.get(key) != generated[key]]
        extra = [key for key in existing if key not in generated]
        raise ValueError(f"Config differs. Changed keys: {changed}; extra keys: {extra}")


def add_update(updates, sheet, address, value, key):
    updates.setdefault(sheet, {})[address] = (value, key)


def validate_config(config):
    if config.get("strip_names") != list(SYMBOL_SHEETS):
        raise ValueError(f"strip_names must be exactly {list(SYMBOL_SHEETS)}")
    if len(config.get("strips", [])) != len(SYMBOL_SHEETS):
        raise ValueError(f"config must contain exactly {len(SYMBOL_SHEETS)} strips")
    levels = config.get("multiplier_levels", [])
    if len(levels) != LEVEL_COUNT:
        raise ValueError(f"multiplier_levels must contain {LEVEL_COUNT} entries")
    for profile_name in ("normal", "featurebuy"):
        profile = config["parameter"][profile_name]
        expected_bg = list(NORMAL_BG_SHEETS if profile_name == "normal" else FEATUREBUY_BG_SHEETS)
        if profile["base_reel_names"] != expected_bg or profile["free_table"]["names"] != list(FG_SHEETS):
            raise ValueError(
                f"parameter.{profile_name} must use base tables {expected_bg} and FG tables {list(FG_SHEETS)}"
            )
        use_super = profile["c2_to_c3"]
        denominator = int(use_super.get("denominator", 10000))
        for name in SYMBOL_SHEETS:
            validate_super_probability_curve(
                use_super["weights_by_initial_ball_count"][name],
                f"parameter.{profile_name}.c2_to_c3.{name}",
                denominator,
            )
            validate_drop_combo_curve(
                use_super["weights_by_drop_combo"][name],
                f"parameter.{profile_name}.c2_to_c3.drop.{name}",
                denominator,
            )
        if use_super.get("drop_combo_buckets") != ["1", "2", "3", "4", "5+"]:
            raise ValueError(
                f"parameter.{profile_name}.c2_to_c3.drop_combo_buckets must be "
                "['1', '2', '3', '4', '5+']"
            )
        if profile_name == "normal":
            initial_bg = use_super["weights_by_initial_ball_count"]["BG_Symbol"]
            initial_bg_2 = use_super["weights_by_initial_ball_count"]["BG_Symbol (2)"]
            drop_bg = use_super["weights_by_drop_combo"]["BG_Symbol"]
            drop_bg_2 = use_super["weights_by_drop_combo"]["BG_Symbol (2)"]
            if any(bg_2 * 5 != bg for bg, bg_2 in zip(initial_bg, initial_bg_2)):
                raise ValueError("normal BG_Symbol (2) initial C3 weights must equal BG_Symbol / 5")
            if any(bg_2 * 5 != bg for bg, bg_2 in zip(drop_bg, drop_bg_2)):
                raise ValueError("normal BG_Symbol (2) drop C3 weights must equal BG_Symbol / 5")
            if any(drop * 2 != initial for initial, drop in zip(initial_bg[:5], drop_bg)):
                raise ValueError("normal BG_Symbol initial C3 weights 1-5 must equal drop Combo weights x2")
        multiplier = profile["multiplier"]
        for kind in ("c2", "c3"):
            weights = multiplier[f"weights_{kind}"]
            cumulative_weights = multiplier[f"weights_{kind}_cum"]
            for name in SYMBOL_SHEETS:
                if len(weights[name]) != LEVEL_COUNT or sum(weights[name]) <= 0:
                    raise ValueError(
                        f"parameter.{profile_name}.multiplier.weights_{kind}.{name} "
                        f"must contain {LEVEL_COUNT} values with a positive total"
                    )
                if cumulative(weights[name]) != cumulative_weights[name]:
                    raise ValueError(
                        f"parameter.{profile_name}.multiplier.weights_{kind}_cum.{name} is stale"
                    )
                if kind == "c3" and any(
                    weight > 0 for level, weight in zip(levels, weights[name]) if level >= 100
                ):
                    raise ValueError(
                        f"parameter.{profile_name}.multiplier.weights_c3.{name} "
                        "must be zero at 100x and above"
                    )


def build_updates(config):
    validate_config(config)
    updates = {}
    levels = config["multiplier_levels"]
    normal = config["parameter"]["normal"]
    add_update(updates, "Overview", "B2", config["model"], "model")
    version_value = config["excel_version"]
    if re.fullmatch(r"\d+", str(version_value)):
        version_value = int(version_value)
    add_update(updates, "Overview", "B3", version_value, "excel_version")
    add_update(updates, "Overview", "A7", config["default_coin_in"], "default_coin_in")
    for address, key in (("B11", "normalbet"), ("B12", "extrabet"), ("B13", "featurebuy")):
        add_update(updates, "Overview", address, config[key], key)
    for col in range(2, 8):
        add_update(updates, "Overview", f"{chr(64 + col)}17", config["window_size"], "window_size")
    for index, row in enumerate(range(30, 42)):
        add_update(updates, "Overview", f"A{row}", config["symbol_codes"][index], "symbol_codes")
        add_update(updates, "Overview", f"I{row}", config["symbol_ids"][index], "symbol_ids")
        for offset, col in enumerate("CDEFGH"):
            add_update(updates, "Overview", f"{col}{row}", config["pay_table"][index][offset], "pay_table")

    for index, row in enumerate((4, 5, 6)):
        add_update(updates, "Parameter", f"B{row}", normal["base_reel_names"][index], "normal.base_reel_names")
        add_update(updates, "Parameter", f"C{row}", normal["base_reel_weights"][index], "normal.base_reel_weights")
    for index, row in enumerate((11, 12)):
        add_update(updates, "Parameter", f"B{row}", normal["free_table"]["names"][index], "normal.free_table.names")
        add_update(updates, "Parameter", f"C{row}", normal["free_table"]["initial"][index], "normal.free_table.initial")
        add_update(updates, "Parameter", f"D{row}", normal["free_table"]["retrigger"][index], "normal.free_table.retrigger")
    for index, value in enumerate(levels):
        col = column_name(11 + index)
        add_update(updates, "Parameter", f"{col}3", value, "multiplier_levels")
        add_update(updates, "Parameter", f"{col}13", value, "multiplier_levels")
        add_update(updates, "Parameter", f"B{18 + index}", index + 1, "multiplier level index")
        add_update(updates, "Parameter", f"C{18 + index}", value, "multiplier_levels")

    table_rows = {
        "BG_Symbol": 24, "BG_Symbol (2)": 25, "BG_Symbol (3)": 26,
        "FG_Symbol": 27, "FG_Symbol (2)": 28, "BF_Symbol": 29,
    }
    for name, row in table_rows.items():
        add_update(updates, "Parameter", f"J{row}", name, "normal.c2_to_c3.table_names")
        for column, value in enumerate(normal["c2_to_c3"]["weights_by_initial_ball_count"][name], start=11):
            add_update(updates, "Parameter", f"{column_name(column)}{row}", value, "normal.c2_to_c3.weights_by_initial_ball_count")
    add_update(updates, "Parameter", "J22", "weight_C2_to_C3_by_initial_count", "initial conversion header")
    add_update(updates, "Parameter", "J23", "Initial C2 Count", "initial conversion label")
    add_update(updates, "Parameter", "J32", "weight_C2_to_C3_by_drop_combo", "drop parameter header")
    add_update(updates, "Parameter", "J33", "Combo", "drop parameter label")
    for column, label in enumerate(("1", "2", "3", "4", "5+"), start=11):
        add_update(updates, "Parameter", f"{column_name(column)}33", label, "drop_combo_buckets")
    drop_combo_rows = {
        "BG_Symbol": 34, "BG_Symbol (2)": 35, "BG_Symbol (3)": 36,
        "FG_Symbol": 37, "FG_Symbol (2)": 38, "BF_Symbol": 39,
    }
    for name, row in drop_combo_rows.items():
        add_update(updates, "Parameter", f"J{row}", name, "normal.c2_to_c3.table_names")
        for column, value in enumerate(normal["c2_to_c3"]["weights_by_drop_combo"][name], start=11):
            add_update(updates, "Parameter", f"{column_name(column)}{row}", value, "normal.c2_to_c3.weights_by_drop_combo")
    add_update(
        updates,
        "Parameter",
        "J40",
        "# How to use: Weight/10000; each dropped C2 rolls once by Combo; Combo 5+ uses the last column",
        "drop parameter note",
    )
    add_update(updates, "Parameter", "J2", "weight_multiplier", "multiplier header")
    add_update(updates, "Parameter", "J3", "C2 Multiplier", "multiplier label")
    multiplier_rows = (("BG_Symbol", 4), ("BG_Symbol (2)", 5),
                       ("BG_Symbol (3)", 6), ("FG_Symbol", 7),
                       ("FG_Symbol (2)", 8), ("BF_Symbol", 9))
    for name, row in multiplier_rows:
        table = normal["multiplier"]
        add_update(updates, "Parameter", f"J{row}", name, "multiplier.table_names")
        for index, value in enumerate(table["weights_c2"][name]):
            add_update(updates, "Parameter", f"{column_name(11 + index)}{row}", value, "multiplier.weights_c2")
    add_update(updates, "Parameter", "J12", "weight_multiplier_C3", "C3 multiplier header")
    add_update(updates, "Parameter", "J13", "C3 Multiplier", "C3 multiplier label")
    c3_multiplier_rows = (("BG_Symbol", 14), ("BG_Symbol (2)", 15),
                          ("BG_Symbol (3)", 16), ("FG_Symbol", 17),
                          ("FG_Symbol (2)", 18), ("BF_Symbol", 19))
    for name, row in c3_multiplier_rows:
        table = normal["multiplier"]
        add_update(updates, "Parameter", f"J{row}", name, "multiplier.table_names")
        for index, value in enumerate(table["weights_c3"][name]):
            add_update(updates, "Parameter", f"{column_name(11 + index)}{row}", value, "multiplier.weights_c3")
    id_to_code = dict(zip(config["symbol_ids"], config["symbol_codes"]))
    for sheet_name, strip in zip(SYMBOL_SHEETS, config["strips"]):
        row_count = len(strip["symbols"])
        if row_count < 1 or row_count > MAX_STRIP_LENGTH or len(strip["weights"]) != row_count:
            raise ValueError(f"{sheet_name} must have 1-{MAX_STRIP_LENGTH} rows")
        reel_lengths = strip.get("reel_lengths")
        if not isinstance(reel_lengths, list) or len(reel_lengths) != REEL_COUNT:
            raise ValueError(f"{sheet_name}.reel_lengths must contain {REEL_COUNT} values")
        for reel, length in enumerate(reel_lengths):
            if require_int(length, f"{sheet_name}.reel_lengths") < 1 or length > row_count:
                raise ValueError(f"{sheet_name} invalid R{reel + 1} length {length}")
        first_symbols = strip["symbols"][0]
        for row_index in range(MAX_STRIP_LENGTH):
            if row_index < row_count and (len(strip["symbols"][row_index]) != REEL_COUNT or len(strip["weights"][row_index]) != REEL_COUNT):
                raise ValueError(f"{sheet_name} row {row_index} must have {REEL_COUNT} reels")
            row = 4 + row_index
            for reel in range(REEL_COUNT):
                active = row_index < reel_lengths[reel]
                symbol_id = require_int(
                    strip["symbols"][row_index][reel] if row_index < row_count else first_symbols[reel],
                    f"{sheet_name}.symbols",
                )
                if symbol_id not in id_to_code:
                    raise ValueError(f"Unknown symbol ID {symbol_id} in {sheet_name} row {row_index}")
                if active:
                    add_update(updates, sheet_name, f"{column_name(12 + reel)}{row}", id_to_code[symbol_id], "strips.symbols")
                    add_update(updates, sheet_name, f"{column_name(19 + reel)}{row}", symbol_id, "Symbol ID cache")
                if active:
                    add_update(updates, sheet_name, f"{column_name(26 + reel)}{row}",
                               require_int(strip["weights"][row_index][reel], f"{sheet_name}.weights"),
                               "strips.weights")
    return updates


def column_name(index):
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def column_index(name):
    result = 0
    for character in name:
        result = result * 26 + ord(character) - 64
    return result


def values_equal(left, right):
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return float(left) == float(right)
    return left == right


def changed_cells(source_path, updates):
    workbook = load_workbook(source_path, read_only=False, data_only=True)
    try:
        changed = []
        for sheet_name, cells in updates.items():
            sheet = workbook[sheet_name]
            for address, (new_value, key) in cells.items():
                old_value = sheet[address].value
                if not values_equal(old_value, new_value):
                    changed.append((sheet_name, address, old_value, new_value, key))
        return changed
    finally:
        workbook.close()


def workbook_sheet_parts(archive):
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels_root}
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    result = {}
    for sheet in workbook_root.findall(f".//{{{main_ns}}}sheet"):
        target = rel_targets[sheet.attrib[f"{{{rel_ns}}}id"]]
        result[sheet.attrib["name"]] = target.lstrip("/") if target.startswith("/") else posixpath.normpath(posixpath.join("xl", target))
    return result


def number_text(value):
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and math.isfinite(value):
        return format(value, ".15g")
    raise TypeError(f"Unsupported numeric value: {value!r}")


def render_cell(attrs, value):
    attrs = re.sub(r'\s+t="[^"]*"', "", attrs.rstrip())
    if value is None:
        return f"<c{attrs}></c>"
    if isinstance(value, str):
        space = ' xml:space="preserve"' if value != value.strip() else ""
        return f'<c{attrs} t="inlineStr"><is><t{space}>{escape(value)}</t></is></c>'
    return f"<c{attrs}><v>{number_text(value)}</v></c>"


def patch_sheet(xml_bytes, sheet_name, cell_updates):
    text = xml_bytes.decode("utf-8")
    remaining = set(cell_updates)

    def replace(match):
        address = match.group("ref") or match.group("selfref")
        if address not in cell_updates:
            return match.group(0)
        attrs = match.group("attrs") or match.group("selfattrs") or ""
        body = match.group("body") or ""
        value, _key = cell_updates[address]
        if _key == "strips linked runtime":
            replacement = render_cell(attrs, value)
        elif re.search(r"<f(?:\s|>)", body):
            if not isinstance(value, (int, float)) or isinstance(value, bool):
                raise TypeError(f"Formula cache must be numeric: {sheet_name}!{address}={value!r}")
            attrs = re.sub(r'\s+t="[^"]*"', "", attrs.rstrip())
            value_xml = f"<v>{number_text(value)}</v>"
            if re.search(r"<v(?:\s[^>]*)?>.*?</v>", body, re.DOTALL):
                body = re.sub(r"<v(?:\s[^>]*)?>.*?</v>", value_xml, body, count=1, flags=re.DOTALL)
            else:
                end = body.find("</f>") + 4
                body = body[:end] + value_xml + body[end:]
            replacement = f"<c{attrs}>{body}</c>"
        else:
            replacement = render_cell(attrs, value)
        remaining.discard(address)
        return replacement

    patched = CELL_PATTERN.sub(replace, text)
    if remaining:
        by_row = {}
        for address in remaining:
            row_number = int(re.search(r"\d+$", address).group())
            by_row.setdefault(row_number, []).append(address)
        for row_number, addresses in by_row.items():
            row_pattern = re.compile(
                rf'(<row\b[^>]*\br="{row_number}"[^>]*>)(.*?)(</row>)', re.DOTALL
            )
            row_match = row_pattern.search(patched)
            if not row_match:
                sample = ", ".join(sorted(addresses)[:10])
                raise ValueError(f"Mapped row missing from {sheet_name}: {sample}")
            additions = []
            for address in sorted(addresses, key=lambda item: column_index(re.match(r"[A-Z]+", item).group())):
                column = re.match(r"[A-Z]+", address).group()
                donor = re.search(
                    rf'<c\b(?P<attrs>[^>]*\br="{column}8"[^>]*)>', patched
                )
                style = re.search(r'\bs="([^"]+)"', donor.group("attrs")) if donor else None
                attrs = f' r="{address}"' + (f' s="{style.group(1)}"' if style else "")
                additions.append(render_cell(attrs, cell_updates[address][0]))
            replacement = row_match.group(1) + row_match.group(2) + "".join(additions) + row_match.group(3)
            patched = patched[:row_match.start()] + replacement + patched[row_match.end():]
            remaining.difference_update(addresses)
    if remaining:
        sample = ", ".join(sorted(remaining)[:10])
        raise ValueError(f"Mapped cell(s) missing from {sheet_name}: {sample}")
    payload = patched.encode("utf-8")
    ET.fromstring(payload)
    return payload


def calculation_metadata(filename, payload):
    if filename == "xl/_rels/workbook.xml.rels":
        return re.sub(r'<Relationship\b[^>]*\bType="[^"]*/calcChain"[^>]*/>', "", payload.decode("utf-8")).encode("utf-8")
    if filename == "[Content_Types].xml":
        return re.sub(r'<Override\b[^>]*\bPartName="/xl/calcChain\.xml"[^>]*/>', "", payload.decode("utf-8")).encode("utf-8")
    if filename == "xl/workbook.xml":
        text = payload.decode("utf-8")
        def update(match):
            attrs = match.group(1)
            for name, value in (("calcMode", "auto"), ("fullCalcOnLoad", "1"), ("forceFullCalc", "1")):
                attrs = re.sub(rf'\b{name}="[^"]*"', f'{name}="{value}"', attrs) if re.search(rf'\b{name}="', attrs) else attrs + f' {name}="{value}"'
            return f"<calcPr{attrs}/>"
        text, count = re.subn(r"<calcPr\b([^>]*)/>", update, text, count=1)
        if not count:
            text = text.replace("</workbook>", '<calcPr calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"/></workbook>')
        return text.encode("utf-8")
    return payload


def write_patched_workbook(source_path, output_path, updates, force=False):
    source_path, output_path = source_path.resolve(), output_path.resolve()
    if output_path.exists() and output_path != source_path and not force:
        raise FileExistsError(f"Output exists; pass --force: {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = output_path.with_name(output_path.name + ".tmp")
    if temporary.exists():
        temporary.unlink()
    try:
        with zipfile.ZipFile(source_path, "r") as source_zip:
            parts = workbook_sheet_parts(source_zip)
            part_updates = {parts[name]: (name, cells) for name, cells in updates.items()}
            with zipfile.ZipFile(temporary, "w") as output_zip:
                for info in source_zip.infolist():
                    if info.filename == "xl/calcChain.xml":
                        continue
                    payload = source_zip.read(info.filename)
                    if info.filename in part_updates:
                        sheet_name, cells = part_updates[info.filename]
                        payload = patch_sheet(payload, sheet_name, cells)
                    payload = calculation_metadata(info.filename, payload)
                    output_zip.writestr(info, payload)
        os.replace(temporary, output_path)
    finally:
        if temporary.exists():
            temporary.unlink()


def verify_output(output_path, config):
    generated = build_config(output_path)
    if generated != config:
        changed = [key for key in generated if config.get(key) != generated[key]]
        extra = [key for key in config if key not in generated]
        raise ValueError(f"Round-trip verification failed. Changed keys: {changed}; extra keys: {extra}")


def run_export(argv):
    parser = argparse.ArgumentParser(description="Build H027 config from H0271.xlsx")
    parser.add_argument("--source", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--output", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args(argv)
    generated = build_config(args.source.resolve())
    if args.check:
        compare_config(generated, args.output.resolve())
        print(f"Config is current: {args.output.resolve()}")
    else:
        write_js_config(args.output.resolve(), generated)
        compare_config(generated, args.output.resolve())
        print(f"Config written and verified: {args.output.resolve()}")


def run_import(argv):
    parser = argparse.ArgumentParser(description="Write H027 config data back to H0271.xlsx")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--source", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--in-place", action="store_true")
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args(argv)
    if args.in_place and args.output:
        parser.error("--in-place and --output cannot be combined")
    config = load_js_config(args.config.resolve())
    updates = build_updates(config)
    changed = changed_cells(args.source.resolve(), updates)
    print(f"Mapped cells: {sum(len(cells) for cells in updates.values()):,}")
    print(f"Changed cells: {len(changed):,}")
    for sheet_name, address, old, new, key in changed[:20]:
        print(f"  {sheet_name}!{address}: {old!r} -> {new!r} ({key})")
    if len(changed) > 20:
        print(f"  ... and {len(changed) - 20:,} more")
    if args.check:
        print("Check only; no xlsx was written.")
        return
    output = args.source.resolve() if args.in_place else (args.output.resolve() if args.output else args.source.resolve().with_name("H0271_from_config.xlsx"))
    write_patched_workbook(args.source.resolve(), output, updates, force=args.force or args.in_place)
    verify_output(output, config)
    print(f"Xlsx written and round-trip verified: {output}")


def run_set_c3_rate(argv):
    parser = argparse.ArgumentParser(
        description="Set the independent per-C2 conversion rate used by initial and drop C3 rolls"
    )
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--weight", type=int, help="Fixed probability weight over the configured denominator")
    parser.add_argument("--initial-weights", help="Six comma-separated per-C2 weights for initial C2 counts 1-6")
    parser.add_argument("--drop-weights", help="Five comma-separated per-C2 weights for drop Combo 1,2,3,4,5+")
    parser.add_argument(
        "--profile",
        choices=("normal", "featurebuy", "all"),
        default="all",
        help="Limit the update to one parameter profile (default: all)",
    )
    parser.add_argument(
        "--table",
        action="append",
        choices=SYMBOL_SHEETS,
        help="Limit the update to one or more reel tables (default: every table)",
    )
    parser.add_argument(
        "--table-index",
        action="append",
        type=int,
        choices=range(len(SYMBOL_SHEETS)),
        help="Limit the update by zero-based SYMBOL_SHEETS index; useful for table names containing spaces",
    )
    args = parser.parse_args(argv)
    config_path = args.config.resolve()
    config = load_js_config(config_path)
    profile_names = ("normal", "featurebuy") if args.profile == "all" else (args.profile,)
    selected_tables = list(args.table or [])
    selected_tables.extend(SYMBOL_SHEETS[index] for index in (args.table_index or []))
    table_names = tuple(dict.fromkeys(selected_tables)) if selected_tables else SYMBOL_SHEETS
    if args.weight is None and args.initial_weights is None and args.drop_weights is None:
        parser.error("provide --weight, --initial-weights, or --drop-weights")
    if args.weight is not None and (args.initial_weights is not None or args.drop_weights is not None):
        parser.error("--weight cannot be combined with --initial-weights or --drop-weights")

    def parse_curve(raw, expected, option):
        if raw is None:
            return None
        try:
            values = [int(value.strip()) for value in raw.split(",")]
        except ValueError as error:
            parser.error(f"{option} must contain integers: {error}")
        if len(values) != expected:
            parser.error(f"{option} must contain exactly {expected} values")
        return values

    initial_weights = parse_curve(args.initial_weights, 6, "--initial-weights")
    drop_weights = parse_curve(args.drop_weights, 5, "--drop-weights")
    for profile_name in profile_names:
        conversion = config["parameter"][profile_name]["c2_to_c3"]
        denominator = int(conversion.get("denominator", 10000))
        if args.weight is not None and (args.weight < 0 or args.weight > denominator):
            parser.error(f"--weight must be within 0-{denominator}")
        for table_name in table_names:
            if args.weight is not None:
                conversion["weights_by_initial_ball_count"][table_name] = [args.weight] * 6
                conversion["weights_by_drop_combo"][table_name] = [args.weight] * 5
            else:
                if initial_weights is not None:
                    conversion["weights_by_initial_ball_count"][table_name] = list(initial_weights)
                if drop_weights is not None:
                    conversion["weights_by_drop_combo"][table_name] = list(drop_weights)
    validate_config(config)
    write_js_config(config_path, config)
    compare_config(config, config_path)
    print(
        f"C2-to-C3 rate updated over denominator {config['parameter']['normal']['c2_to_c3']['denominator']} "
        f"for {', '.join(profile_names)} / {', '.join(table_names)}: {config_path}"
    )


def rebalance_multiplier_weights(weights, levels, targets):
    total = sum(weights)
    if total <= 0:
        raise ValueError("multiplier weight total must be positive")
    target_indices = {}
    duplicate_target_indices = set()
    for multiplier, probability in targets.items():
        matches = [index for index, value in enumerate(levels) if value == multiplier]
        if not matches:
            raise ValueError(f"multiplier {multiplier} is not present in multiplier_levels")
        target_indices[matches[0]] = int(round(probability * total))
        duplicate_target_indices.update(matches[1:])
    reserved = sum(target_indices.values())
    if reserved > total:
        raise ValueError(f"target multiplier weights {reserved} exceed total {total}")
    excluded = set(target_indices) | duplicate_target_indices
    candidates = [index for index, value in enumerate(weights) if index not in excluded and value > 0]
    if not candidates and reserved != total:
        raise ValueError("no non-target multiplier weights are available for rebalancing")
    remaining = total - reserved
    source_total = sum(weights[index] for index in candidates)
    result = [0] * len(weights)
    if source_total > 0:
        exact = [(index, weights[index] * remaining / source_total) for index in candidates]
        for index, value in exact:
            result[index] = int(value)
        remainder = remaining - sum(result)
        order = sorted(exact, key=lambda item: (-(item[1] - int(item[1])), item[0]))
        for index, _ in order[:remainder]:
            result[index] += 1
    for index, value in target_indices.items():
        result[index] = value
    if sum(result) != total:
        raise ValueError(f"rebalanced multiplier weights must sum to {total}, got {sum(result)}")
    return result


def run_set_multiplier_rates(argv):
    parser = argparse.ArgumentParser(
        description="Set exact multiplier probabilities while preserving the other positive weights proportionally"
    )
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--kind", choices=("c2", "c3"), required=True)
    parser.add_argument("--profile", choices=("normal", "featurebuy", "all"), default="all")
    parser.add_argument("--table-index", action="append", type=int, choices=range(len(SYMBOL_SHEETS)))
    parser.add_argument(
        "--target",
        action="append",
        required=True,
        help="Multiplier probability as MULTIPLIER=PERCENT, for example 500=0.2",
    )
    args = parser.parse_args(argv)
    targets = {}
    for item in args.target:
        try:
            multiplier_text, percent_text = item.split("=", 1)
            multiplier = int(multiplier_text)
            probability = float(percent_text) / 100.0
        except (ValueError, TypeError) as error:
            parser.error(f"invalid --target {item!r}: {error}")
        if probability < 0 or probability > 1:
            parser.error(f"target probability must be within 0%-100%: {item}")
        targets[multiplier] = probability
    config_path = args.config.resolve()
    config = load_js_config(config_path)
    profile_names = ("normal", "featurebuy") if args.profile == "all" else (args.profile,)
    table_names = tuple(SYMBOL_SHEETS[index] for index in args.table_index) if args.table_index else SYMBOL_SHEETS
    for profile_name in profile_names:
        multiplier = config["parameter"][profile_name]["multiplier"]
        levels = multiplier["multipliers"]
        raw_key = f"weights_{args.kind}"
        cumulative_key = f"weights_{args.kind}_cum"
        for table_name in table_names:
            updated = rebalance_multiplier_weights(multiplier[raw_key][table_name], levels, targets)
            multiplier[raw_key][table_name] = updated
            multiplier[cumulative_key][table_name] = cumulative(updated)
    validate_config(config)
    write_js_config(config_path, config)
    compare_config(config, config_path)
    formatted_targets = ", ".join(f"{value}x={probability * 100:g}%" for value, probability in targets.items())
    print(
        f"{args.kind.upper()} multiplier rates updated for {', '.join(profile_names)} / "
        f"{', '.join(table_names)} ({formatted_targets}): {config_path}"
    )


def main():
    parser = argparse.ArgumentParser(
        prog="model_sync.py",
        usage="model_sync.py {export,import,set-c3-rate,set-multiplier-rates} [options]",
    )
    parser.add_argument("command", choices=("export", "import", "set-c3-rate", "set-multiplier-rates"))
    args, rest = parser.parse_known_args()
    commands = {
        "export": run_export,
        "import": run_import,
        "set-c3-rate": run_set_c3_rate,
        "set-multiplier-rates": run_set_multiplier_rates,
    }
    commands[args.command](rest)


if __name__ == "__main__":
    main()
