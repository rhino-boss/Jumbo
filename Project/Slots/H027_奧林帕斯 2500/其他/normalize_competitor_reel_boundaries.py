"""Keep H027 v1 reels competitor-exact while moving C1 away from sheet edges."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
COMPETITOR_BOOK = (
    ROOT.parents[3]
    / "市場資訊"
    / "H5"
    / "遊戲資源"
    / "PP - Gates of Olympus 1000"
    / "還原輪帶_Gates_of_Olympus_1000.xlsx"
)
CONFIG_PATHS = (ROOT / "config.js", ROOT / "config_92A.js", ROOT / "config_94A.js")
REPORT_PATH = ROOT / "其他" / "competitor_reel_validation_v1.json"
MODEL_XLSX = ROOT / "Source" / "H0271.xlsx"
EDGE_GUARD = 5
C1 = 1
BASE_LINKED_OFFSETS = {
    "BG_Symbol": [0, 250000, 656250, 765625, 234375, 984375],
    "BG_Symbol (2)": [0, 921875, 625000, 421875, 312500, 31250],
    "BG_Symbol (3)": [0, 484375, 250000, 328125, 15625, 671875],
    "FG_Symbol": [0, 703125, 734375, 781250, 250000, 656250],
    "FG_Symbol (2)": [0, 484375, 234375, 343750, 343750, 250000],
    "BF_Symbol": [0, 0, 0, 0, 0, 0],
}


def load_js(path: Path) -> tuple[str, dict]:
    text = path.read_text(encoding="utf-8-sig")
    start, end = text.index("{"), text.rindex("}") + 1
    return text[:start], json.loads(text[start:end])


def write_js(path: Path, prefix: str, data: dict) -> None:
    path.write_text(prefix + json.dumps(data, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")


def competitor_reels() -> dict[int, list[list[str]]]:
    workbook = load_workbook(COMPETITOR_BOOK, read_only=True, data_only=True)
    result = {}
    try:
        for set_id in range(5):
            rows = list(workbook[f"Reel Set {set_id}"].iter_rows(min_row=2, values_only=True))
            result[set_id] = [
                [str(row[1 + reel]) for row in rows if row[1 + reel] is not None]
                for reel in range(6)
            ]
    finally:
        workbook.close()
    return result


def cyclic_shift(source: list[str], target: list[str]) -> int | None:
    if len(source) != len(target):
        return None
    for shift in range(len(source)):
        if source[shift:] + source[:shift] == target:
            return shift
    return None


def edge_safe(sequence: list[int]) -> bool:
    return C1 not in sequence[:EDGE_GUARD] and C1 not in sequence[-EDGE_GUARD:]


def preferred_rotation(sequence: list[int]) -> int:
    length = len(sequence)
    candidates = [shift for shift in range(length) if edge_safe(sequence[shift:] + sequence[:shift])]
    if not candidates:
        raise ValueError(f"No cyclic rotation can keep C1 outside the first/last {EDGE_GUARD} stops")
    return min(candidates, key=lambda shift: (min(shift, length - shift), shift))


def rotate_strip(strip: dict, shifts: list[int]) -> list[dict]:
    lengths = [int(value) for value in strip["reel_lengths"]]
    rows = len(strip["symbols"])
    details = []
    for reel, (length, shift) in enumerate(zip(lengths, shifts)):
        old_symbols = [int(strip["symbols"][row][reel]) for row in range(length)]
        new_symbols = old_symbols[shift:] + old_symbols[:shift]
        for row in range(length):
            strip["symbols"][row][reel] = new_symbols[row]
            strip["weights"][row][reel] = 1
        for row in range(length, rows):
            # Inactive rectangular padding is ignored by reel_lengths and is
            # canonicalized to C1 internally; XLSX displays these cells blank.
            strip["symbols"][row][reel] = C1
            strip["weights"][row][reel] = 0

        details.append(
            {
                "reel": reel + 1,
                "length": length,
                "additional_rotation_applied": shift,
                "c1_positions": [index for index, symbol in enumerate(new_symbols) if symbol == C1],
                "edge_safe": edge_safe(new_symbols),
            }
        )
    return details


def validate_xlsx(report: dict, competitor: dict[int, list[list[str]]]) -> None:
    # Normal mode is intentional: repeated random cell access on an openpyxl
    # read-only worksheet rescans XML and is dramatically slower here.
    workbook = load_workbook(MODEL_XLSX, read_only=False, data_only=False)
    try:
        for name, table in report["tables"].items():
            sheet = workbook[name]
            set_id = int(table["source_reel_set"])
            lengths = [int(item["length"]) for item in table["reels"]]
            row_count = max(lengths)
            for reel, length in enumerate(lengths):
                symbols = [str(sheet.cell(4 + row, 12 + reel).value) for row in range(length)]
                weights = [sheet.cell(4 + row, 26 + reel).value for row in range(length)]
                if cyclic_shift(competitor[set_id][reel], symbols) is None:
                    raise ValueError(f"XLSX mismatch: {name} R{reel + 1}")
                if any(weight != 1 for weight in weights):
                    raise ValueError(f"XLSX active Symbol Weight must all be 1: {name} R{reel + 1}")
                c1_positions = [index for index, symbol in enumerate(symbols) if symbol == "C1"]
                if any(index < EDGE_GUARD or index >= length - EDGE_GUARD for index in c1_positions):
                    raise ValueError(f"XLSX C1 edge violation: {name} R{reel + 1} {c1_positions}")
                for row in range(length, 300):
                    if sheet.cell(4 + row, 12 + reel).value not in (None, ""):
                        raise ValueError(f"XLSX inactive symbol is not blank: {name} R{reel + 1} line {row}")
                    if sheet.cell(4 + row, 26 + reel).value not in (None, ""):
                        raise ValueError(f"XLSX inactive weight is not blank: {name} R{reel + 1} line {row}")
            for row in range(row_count, 300):
                if sheet.cell(4 + row, 11).value not in (None, ""):
                    raise ValueError(f"XLSX inactive line number is not blank: {name} line {row}")
    finally:
        workbook.close()
    report["xlsx"] = {
        "path": str(MODEL_XLSX),
        "competitor_cyclic_match": True,
        "c1_outside_first_last_5": True,
        "active_symbol_weights_all_1": True,
        "inactive_rows_blank": True,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    competitor = competitor_reels()
    base_prefix, base = load_js(CONFIG_PATHS[0])
    codes = dict(zip(base["symbol_ids"], base["symbol_codes"]))
    plan = {}
    report = {"competitor_workbook": str(COMPETITOR_BOOK), "edge_guard": EDGE_GUARD, "tables": {}}

    for name, strip in zip(base["strip_names"], base["strips"]):
        set_id = int(strip["source_reel_set"])
        shifts = []
        before_match = []
        for reel, length in enumerate(strip["reel_lengths"]):
            sequence = [int(strip["symbols"][row][reel]) for row in range(int(length))]
            sequence_codes = [codes[symbol] for symbol in sequence]
            match = cyclic_shift(competitor[set_id][reel], sequence_codes)
            if match is None:
                raise ValueError(f"{name} R{reel + 1} is not a cyclic copy of competitor Reel Set {set_id}")
            before_match.append(match)
            shifts.append(preferred_rotation(sequence))
        plan[name] = shifts
        report["tables"][name] = {
            "source_reel_set": set_id,
            "cyclic_match_before": before_match,
            "reels": rotate_strip(strip, shifts),
        }

    # Validate the normalized base against the competitor one more time.
    for name, strip in zip(base["strip_names"], base["strips"]):
        set_id = int(strip["source_reel_set"])
        for reel, length in enumerate(strip["reel_lengths"]):
            sequence = [codes[int(strip["symbols"][row][reel])] for row in range(int(length))]
            match = cyclic_shift(competitor[set_id][reel], sequence)
            if match is None:
                raise ValueError(f"Post-normalization mismatch: {name} R{reel + 1}")
            report["tables"][name]["reels"][reel]["competitor_cyclic_shift"] = match
            strip["linked_stop_offsets"][reel] = (
                (BASE_LINKED_OFFSETS[name][reel] - round(match * 1_000_000 / int(length))) % 1_000_000
                if int(strip.get("linked_stop_weight", 0)) > 0
                else 0
            )
            report["tables"][name]["reels"][reel]["symbol_weight"] = 1

    if MODEL_XLSX.exists():
        validate_xlsx(report, competitor)

    if args.apply:
        write_js(CONFIG_PATHS[0], base_prefix, base)
        base_by_name = dict(zip(base["strip_names"], base["strips"]))
        for path in CONFIG_PATHS[1:]:
            prefix, config = load_js(path)
            by_name = dict(zip(config["strip_names"], config["strips"]))
            for name, shifts in plan.items():
                rotate_strip(by_name[name], shifts)
                by_name[name]["linked_stop_offsets"] = list(base_by_name[name]["linked_stop_offsets"])
            write_js(path, prefix, config)
        REPORT_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(f"Updated configs and wrote {REPORT_PATH}")
    else:
        print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
