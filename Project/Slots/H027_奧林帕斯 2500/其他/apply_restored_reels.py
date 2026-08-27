from __future__ import annotations

import hashlib
import json
import shutil
from copy import deepcopy
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\rhinshen\Mine\個人工作區\市場資訊\H5\遊戲資源\PP - Gates of Olympus 1000\還原輪帶_Gates_of_Olympus_1000.xlsx")
CONFIG_PATHS = [ROOT / "config.js", ROOT / "config_92A.js", ROOT / "config_94A.js"]
SNAPSHOT_FILES = CONFIG_PATHS + [
    ROOT / "Simulator.py",
    ROOT / "index.html",
    ROOT / "Source" / "H0271.xlsx",
    ROOT / "Source" / "H027192A.xlsx",
    ROOT / "Source" / "H027194A.xlsx",
]

TABLES = [
    ("BG_Symbol", "Reel Set 0"),
    ("BG_Symbol (2)", "Reel Set 1"),
    ("BG_Symbol (3)", "Reel Set 2"),
    ("FG_Symbol", "Reel Set 3"),
    ("FG_Symbol (2)", "Reel Set 4"),
    # Set 6 entry strips are not present in the supplied workbook.  BF uses
    # Set 0 as its source strip and applies the existing four-reel C1 guarantee.
    ("BF_Symbol", "Reel Set 0"),
]


def load_js(path: Path) -> dict:
    text = path.read_text(encoding="utf-8-sig")
    return json.loads(text[text.index("{"): text.rindex("}") + 1])


def write_js(path: Path, data: dict):
    path.write_text("const data = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def snapshot():
    stamp = datetime.now().strftime("%y%m%d_%H%M%S")
    target = ROOT / "Versions" / "0.0" / f"pre_restored_reels_{stamp}"
    target.mkdir(parents=True, exist_ok=False)
    rows = []
    for source in SNAPSHOT_FILES:
        relative = source.relative_to(ROOT)
        destination = target / relative
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, destination)
        rows.append(f"{sha256(destination)}  {relative.as_posix()}")
    (target / "SHA256SUMS.txt").write_text("\n".join(rows) + "\n", encoding="utf-8")
    return target


def load_reels(symbol_to_id: dict[str, int]):
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    loaded = {}
    for table_name, sheet_name in TABLES:
        sheet = workbook[sheet_name]
        header = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        if header[:7] != ["Line #", "R1", "R2", "R3", "R4", "R5", "R6"]:
            raise ValueError(f"Unexpected reel header in {sheet_name}: {header[:7]}")
        reels = [[] for _ in range(6)]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            for reel in range(6):
                code = row[reel + 1]
                if code is not None:
                    if str(code) not in symbol_to_id:
                        raise ValueError(f"Unknown symbol {code!r} in {sheet_name} R{reel + 1}")
                    reels[reel].append(symbol_to_id[str(code)])
        lengths = [len(reel) for reel in reels]
        if min(lengths) < 5:
            raise ValueError(f"Invalid reel lengths in {sheet_name}: {lengths}")
        row_count = max(lengths)
        symbols = []
        weights = []
        for index in range(row_count):
            symbols.append([reels[reel][index] if index < lengths[reel] else reels[reel][0] for reel in range(6)])
            weights.append([1 if index < lengths[reel] else 0 for reel in range(6)])
        loaded[table_name] = {
            "symbols": symbols,
            "weights": weights,
            "reel_lengths": lengths,
            "cascade_source": "reel_strip",
            "source_reel_set": int(sheet_name.rsplit(" ", 1)[-1]),
        }
    return loaded


def copy_table_values(source: dict, target_names: list[str], reference_by_target: dict[str, str]):
    output = {}
    for target in target_names:
        reference = reference_by_target[target]
        if reference not in source:
            reference = next(iter(source))
        output[target] = deepcopy(source[reference])
    return output


def update_profile(profile: dict, is_feature: bool):
    if is_feature:
        profile["base_reel_names"] = ["BF_Symbol"]
        profile["base_reel_weights"] = [1]
        profile["base_reel_weights_cum"] = [1]
        profile["free_table"] = {
            "names": ["FG_Symbol", "FG_Symbol (2)"],
            "initial": [5, 10],
            "retrigger": [2, 3],
        }
    else:
        weights = [9699, 19298, 2344]
        profile["base_reel_names"] = ["BG_Symbol", "BG_Symbol (2)", "BG_Symbol (3)"]
        profile["base_reel_weights"] = weights
        profile["base_reel_weights_cum"] = [weights[0], weights[0] + weights[1], sum(weights)]
        profile["free_table"] = {
            "names": ["FG_Symbol", "FG_Symbol (2)"],
            "initial": [6, 9],
            "retrigger": [2, 3],
        }

    names = [name for name, _ in TABLES]
    reference = {
        "BG_Symbol": "BG_Symbol",
        "BG_Symbol (2)": "BG_Symbol",
        "BG_Symbol (3)": "BG_Symbol (2)",
        "FG_Symbol": "FG_Symbol",
        "FG_Symbol (2)": "FG_Symbol (2)",
        "BF_Symbol": "BG_Symbol",
    }
    use_super = profile["use_super_multiplier"]
    use_super["table_names"] = names
    use_super["weights_by_initial_ball_count"] = copy_table_values(
        use_super["weights_by_initial_ball_count"], names, reference
    )
    c2 = profile["c2"]
    c2["table_names"] = names
    c2["weights"] = copy_table_values(c2["weights"], names, reference)
    c2["weights_cum"] = copy_table_values(c2["weights_cum"], names, reference)
    c3 = profile["c3"]
    c3["table_names"] = names
    c3["weights"] = copy_table_values(c3["weights"], names, reference)
    c3["weights_cum"] = copy_table_values(c3["weights_cum"], names, reference)


def update_config(path: Path, strips: dict):
    config = load_js(path)
    config["strip_names"] = [name for name, _ in TABLES]
    config["strips"] = [deepcopy(strips[name]) for name in config["strip_names"]]
    config["cascade_symbol_source"] = "reel_strip"
    config["reel_source_workbook"] = str(SOURCE)
    config["reel_set_usage"] = {
        "BG": {"sets": [0, 1, 2], "weights": [9699, 19298, 2344]},
        "FG": {"sets": [3, 4], "spin_counts": [6, 9]},
        "BF_FG": {"sets": [3, 4], "spin_counts": [5, 10]},
        "BF_ENTRY": {"source_set": 0, "rule": "force C1 on reels 1-4 because Set 6 is absent from source workbook"},
    }
    update_profile(config["parameter"]["normal"], False)
    update_profile(config["parameter"]["featurebuy"], True)
    write_js(path, config)


def main():
    if not SOURCE.exists():
        raise FileNotFoundError(SOURCE)
    base = load_js(CONFIG_PATHS[0])
    symbol_to_id = dict(zip(base["symbol_codes"], base["symbol_ids"]))
    strips = load_reels(symbol_to_id)
    target = snapshot()
    for path in CONFIG_PATHS:
        update_config(path, strips)
    print(f"snapshot={target}")
    print(f"source={SOURCE}")
    for name, _ in TABLES:
        print(f"{name}: {strips[name]['reel_lengths']} (Set {strips[name]['source_reel_set']})")


if __name__ == "__main__":
    main()
