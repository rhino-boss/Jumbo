from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(r"C:\Users\rhinshen\Mine\個人工作區\市場資訊\H5\遊戲資源\PP - Gates of Olympus 1000\還原輪帶_Gates_of_Olympus_1000.xlsx")
RESPONSE_DIR = SOURCE.parent / "遊玩資料"


def main():
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    print(SOURCE)
    for sheet in workbook.worksheets:
        print(f"\n[{sheet.title}] rows={sheet.max_row} cols={sheet.max_column}")
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 12), values_only=True):
            print(row[: min(sheet.max_column, 12)])

    for response in sorted(RESPONSE_DIR.glob("*.xlsx")):
        source = load_workbook(response, read_only=True, data_only=True)
        print(f"\n[Response: {response.name}] sheets={source.sheetnames}")
        sheet = source["game_data"]
        header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        print(f"rows={sheet.max_row} columns={header}")
        for row in sheet.iter_rows(min_row=2, max_row=min(sheet.max_row, 5), values_only=True):
            print(row)


if __name__ == "__main__":
    main()
