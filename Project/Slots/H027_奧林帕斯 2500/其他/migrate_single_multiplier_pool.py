from __future__ import annotations

import json
import math
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
VERSION = ROOT / "Versions" / "1.0"
LEVELS = [2, 3, 4, 5, 6, 8, 10, 12, 15, 20, 25, 50, 100, 250, 500, 1000, 2500]
TARGETS = {
    "BG": [11.50, 9.67, 10.37, 14.39, 18.33, 15.53, 9.41, 4.81, 1.53, 2.06, .92, .92, .52, .04, 0, 0, 0],
    "FG": [43.61, 23.87, 14.00, 11.57, 2.91, 1.86, 1.29, .32, .24, .16, .08, .08, 0, 0, 0, 0, 0],
    "BF": [46.29, 24.84, 11.80, 10.09, 3.20, 1.41, .63, .63, .40, .23, .15, .25, .08, 0, 0, 0, 0],
}
SCENE_BY_TABLE = {
    "BG_Symbol": "BG", "BG_Symbol (2)": "BG", "BG_Symbol (3)": "BG",
    "FG_Symbol": "FG", "FG_Symbol (2)": "FG", "BF_Symbol": "BF",
}


def largest_remainder(values: list[float], total: int = 10000) -> list[int]:
    scale = sum(values)
    exact = [value / scale * total for value in values]
    result = [math.floor(value) for value in exact]
    for index in sorted(range(len(values)), key=lambda i: exact[i] - result[i], reverse=True)[:total - sum(result)]:
        result[index] += 1
    return result


WEIGHTS = {scene: largest_remainder(values) for scene, values in TARGETS.items()}


def cumulative(values: list[int]) -> list[int]:
    total = 0
    result = []
    for value in values:
        total += value
        result.append(total)
    return result


def load_js(path: Path) -> tuple[dict, str]:
    text = path.read_text(encoding="utf-8-sig")
    start, end = text.index("{"), text.rindex("}")
    return json.loads(text[start:end + 1]), text[:start]


def write_js(path: Path, data: dict, prefix: str) -> None:
    path.write_text(prefix + json.dumps(data, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")


def pool_for_tables(table_names: list[str], featurebuy: bool) -> dict:
    weights = {}
    for name in table_names:
        scene = "BF" if featurebuy else SCENE_BY_TABLE[name]
        row = WEIGHTS[scene] + [0] * (25 - len(WEIGHTS[scene]))
        weights[name] = row
    return {
        "multipliers": [2, 3, 4, 5, 6, 8, 10, 12, 15, 20, 25, 50, 100, 250, 500, 1000, 2500] + [2500] * 8,
        "table_names": table_names,
        "weights": weights,
        "weights_cum": {name: cumulative(row) for name, row in weights.items()},
    }


def migrate_config(path: Path) -> None:
    data, prefix = load_js(path)
    parameter = data["parameter"]
    parameter.pop("super_multiplier", None)
    for profile_name in ("normal", "featurebuy"):
        profile = parameter[profile_name]
        conversion = profile.pop("use_super_multiplier")
        profile["c2_to_c3"] = conversion
        profile["multiplier"] = pool_for_tables(list(conversion["table_names"]), profile_name == "featurebuy")
        profile.pop("c2", None)
        profile.pop("c3", None)
    write_js(path, data, prefix)


def copy_style(source, target) -> None:
    target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def migrate_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook["Parameter"]
        for row in range(2, 6):
            for column in range(10, 36):
                sheet.cell(row, column).value = None
        for row in range(17, 26):
            for column in range(10, 36):
                sheet.cell(row, column).value = None
        sheet["J7"] = "weight_multiplier"
        sheet["J8"] = "Multiplier"
        table_rows = {"BG_Symbol": 9, "BG_Symbol (2)": 10, "BG_Symbol (3)": 11,
                      "FG_Symbol": 12, "FG_Symbol (2)": 13, "BF_Symbol": 14}
        levels = [sheet.cell(28 + index, 3).value for index in range(25)]
        for name, row in table_rows.items():
            sheet.cell(row, 10, name)
            by_value = dict(zip(LEVELS, WEIGHTS[SCENE_BY_TABLE[name]]))
            used = set()
            for index, value in enumerate(levels, start=11):
                weight = by_value.get(int(value), 0) if int(value) not in used else 0
                used.add(int(value))
                sheet.cell(row, index, weight)
        sheet["B16"] = "weight_C2_to_C3_by_initial_count"
        sheet["B17"] = "Initial C2 Count"
        sheet["B24"] = "# Each initial C2 rolls independently: Weight/10000; 6+ uses column 6"
        sheet["B26"] = "multiplier_level_C2_C3"
        sheet["B27"] = "Level"
        sheet["C27"] = "Multiplier"
        sheet["B55"] = "weight_C2_to_C3_by_drop_combo"
        sheet["B56"] = "Combo"
        sheet["B63"] = "# Each dropped C2 rolls independently: Weight/10000; Combo 5+ uses the last column"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(path)
    finally:
        workbook.close()


def main() -> None:
    for base in (ROOT, VERSION):
        for name in ("config.js", "config_92A.js", "config_94A.js"):
            migrate_config(base / name)
        migrate_workbook(base / "Source" / "H0271.xlsx")
    print("Competitor multiplier weights:", WEIGHTS)
    print("Migrated Base/v1 configs and H0271.xlsx to one C2/C3 multiplier pool.")


if __name__ == "__main__":
    main()
