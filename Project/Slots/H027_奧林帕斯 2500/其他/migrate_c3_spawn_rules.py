from __future__ import annotations

import json
import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
VERSION = ROOT / "Versions" / "1.0"
CONFIG_PATHS = [
    ROOT / "config.js",
    ROOT / "config_92A.js",
    ROOT / "config_94A.js",
    VERSION / "config.js",
    VERSION / "config_92A.js",
    VERSION / "config_94A.js",
]


def load_js(path: Path) -> tuple[dict, str]:
    text = path.read_text(encoding="utf-8-sig")
    start, end = text.find("{"), text.rfind("}")
    if start < 0 or end <= start:
        raise ValueError(f"Invalid config: {path}")
    return json.loads(text[start : end + 1]), text[:start]


def write_js(path: Path, data: dict, prefix: str) -> None:
    path.write_text(prefix + json.dumps(data, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")


def update_config(path: Path) -> None:
    data, prefix = load_js(path)
    for profile_name in ("normal", "featurebuy"):
        block = data["parameter"][profile_name]["use_super_multiplier"]
        block["drop_combo_buckets"] = ["1", "2", "3", "4", "5+"]
        block["weights_by_drop_combo"] = {
            name: [int(value) for value in block["weights_by_initial_ball_count"][name][:5]]
            for name in block["table_names"]
        }
    write_js(path, data, prefix)


def copy_style(source, target) -> None:
    target._style = copy(source._style)
    target.number_format = source.number_format
    target.protection = copy(source.protection)
    target.alignment = copy(source.alignment)


def update_workbook(source: Path, output: Path) -> None:
    if source.resolve() != output.resolve():
        shutil.copy2(source, output)
    workbook = load_workbook(output)
    try:
        sheet = workbook["Parameter"]
        for target_row, source_row in ((55, 16), (56, 17), (63, 24)):
            for column in range(2, 8):
                copy_style(sheet.cell(source_row, column), sheet.cell(target_row, column))
        for target_row, source_row in zip(range(57, 63), range(18, 24)):
            for column in range(2, 8):
                copy_style(sheet.cell(source_row, column), sheet.cell(target_row, column))
        for merged in ("B55:G55", "B63:G63"):
            if merged not in {str(item) for item in sheet.merged_cells.ranges}:
                sheet.merge_cells(merged)

        sheet["B55"] = "weight_use_super_multiplier_drop"
        for column, label in enumerate(("1", "2", "3", "4", "5+"), start=3):
            sheet.cell(56, column, label)

        config, _ = load_js(ROOT / "config.js")
        block = config["parameter"]["normal"]["use_super_multiplier"]
        for row, name in zip(range(57, 63), block["table_names"]):
            sheet.cell(row, 2, name)
            weights = block["weights_by_drop_combo"][name]
            for column, value in enumerate(weights, start=3):
                sheet.cell(row, column, int(value))
        sheet["B63"] = "# How to use: Weight/10000; each dropped C2 rolls once by Combo; Combo 5+ uses the last column"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(output)
    finally:
        workbook.close()


def main() -> None:
    for path in CONFIG_PATHS:
        update_config(path)
        print(f"Updated config: {path}")

    version_workbook = VERSION / "Source" / "H0271.xlsx"
    update_workbook(version_workbook, version_workbook)
    print(f"Updated workbook: {version_workbook}")

    active_workbook = ROOT / "Source" / "H0271.xlsx"
    lock_file = active_workbook.with_name("~$" + active_workbook.name)
    if lock_file.exists():
        staged = active_workbook.with_name("H0271_C3機制更新版.xlsx")
        update_workbook(active_workbook, staged)
        print(f"Active workbook appears open; staged workbook: {staged}")
    else:
        update_workbook(active_workbook, active_workbook)
        print(f"Updated workbook: {active_workbook}")


if __name__ == "__main__":
    main()
