from pathlib import Path

from openpyxl import load_workbook


root = Path(__file__).resolve().parent.parent
for path in sorted((root / "Source").glob("*.xlsx")):
    print(f"\n[{path.name}]")
    workbook = load_workbook(path, read_only=False, data_only=False)
    print(workbook.sheetnames)
    for sheet in workbook.worksheets:
        labels = []
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    text = cell.value.strip()
                    if any(token in text.lower() for token in ("weight", "multiplier", "version", "symbol id", "symbol weight")):
                        labels.append(f"{cell.coordinate}={text}")
        if labels:
            print(sheet.title, " | ".join(labels[:40]))
    if "Multiplier_Weight" in workbook.sheetnames:
        sheet = workbook["Multiplier_Weight"]
        for row in range(1, min(sheet.max_row, 70) + 1):
            values = [sheet.cell(row, column).value for column in range(1, min(sheet.max_column, 8) + 1)]
            if any(value not in (None, "") for value in values):
                print(row, values)
        for detail_name in ("Detail_Newbie", "Detail"):
            detail = workbook[detail_name]
            print(f"[{detail_name} samples]")
            for row in list(range(10, 20)) + list(range(80, 90)) + list(range(155, 166)) + list(range(228, 237)):
                values = [detail.cell(row, column).value for column in range(1, 12)]
                if any(value not in (None, "") for value in values):
                    print(row, values)
