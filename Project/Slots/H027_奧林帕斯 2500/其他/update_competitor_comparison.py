"""Refresh dynamic H027-vs-competitor sections from a Simulator report."""

from __future__ import annotations

import argparse
import importlib.util
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_RECORD = ROOT / "Record" / "H0271_0001_2608121346_betmode0_2000000.xlsx"
DEFAULT_COMPETITOR = ROOT / "其他" / "參考資料" / "analysis_gates_of_olympus_1000_metrics.json"
DEFAULT_RESPONSES = ROOT / "其他" / "參考資料" / "game_responses-gates of olympus 1000.xlsx"
DEFAULT_STACK_METRICS = ROOT / "其他" / "參考資料" / "stack_distribution_metrics.json"
DEFAULT_CONFIG = ROOT / "config_92A.js"
DEFAULT_REPORT = ROOT / "其他" / "競品參考數值比較.md"
ANALYZER_PATH = ROOT / "其他" / "analyze_gates_competitor.py"
STACK_ANALYZER_PATH = ROOT / "其他" / "analyze_stack_distribution.py"
SCENE_TABLES = {
    "BG": (("BG_Symbol", 1), ("BG_Symbol (2)", 1)),
    "FG": (("FG_Symbol", 8), ("FG_Symbol (2)", 7)),
}
SYMBOL_ORDER = ("C1", "C2", "M1", "M2", "M3", "M4", "A", "K", "Q", "J", "TE")
SYMBOL_LABELS = {"C1": "Scatter / C1", "C2": "Multiplier / C2"}


def pct(value: float) -> str:
    return f"{value * 100:.4f}%"


def pp(value: float) -> str:
    return f"{value * 100:+.4f} pp"


def table(headers: list[str], rows: list[list[object]]) -> str:
    lines = [
        "| " + " | ".join(headers) + " |",
        "|" + "|".join("---:" if index >= 2 else "---" for index in range(len(headers))) + "|",
    ]
    lines.extend("| " + " | ".join(str(item) for item in row) + " |" for row in rows)
    return "\n".join(lines)


def replace_between(text: str, start: str, end: str | None, replacement: str) -> str:
    left = text.index(start)
    right = text.index(end, left) if end else len(text)
    return text[:left] + replacement.rstrip() + "\n\n" + text[right:]


def load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load module: {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


def load_config(path: Path) -> dict:
    match = re.fullmatch(
        r"\s*const\s+data\s*=\s*(\{.*\})\s*;?\s*",
        path.read_text(encoding="utf-8-sig"),
        re.DOTALL,
    )
    if not match:
        raise ValueError(f"Unsupported config format: {path}")
    return json.loads(match.group(1))


def competitor_reel_distributions(response_path: Path) -> dict:
    analyzer = load_module("h027_competitor_report", ANALYZER_PATH)
    analysis = analyzer.analyze(response_path)
    sessions = analysis["_sessions"]
    scene_spins = {
        "BG": [session.bg for session in sessions],
        "FG": [spin for session in sessions for spin in session.fg_spins],
    }
    result = {}
    for scene, spins in scene_spins.items():
        result[scene] = {}
        for mode in ("initial", "drop"):
            counts = [[0] * 13 for _ in range(6)]
            screen_count = 0
            for spin in spins:
                screens = spin.screens[:1] if mode == "initial" else spin.screens[1:]
                for screen in screens:
                    if len(screen) != 30:
                        continue
                    screen_count += 1
                    for reel in range(6):
                        for row in range(5):
                            counts[reel][int(screen[row * 6 + reel])] += 1
            denominator = screen_count * 5
            by_code = {}
            for symbol_id, code in analyzer.SYMBOL_ID_TO_CODE.items():
                if code in SYMBOL_ORDER:
                    by_code[code] = [counts[reel][symbol_id] / denominator for reel in range(6)]
            result[scene][mode] = {"screen_count": screen_count, "by_code": by_code}
    return result


def h027_reel_distributions(config: dict) -> dict:
    by_name = dict(zip(config["strip_names"], config["strips"]))
    code_by_id = dict(zip(config["symbol_ids"], config["symbol_codes"]))
    result = {}
    for scene, tables in SCENE_TABLES.items():
        result[scene] = {}
        initial = {code: [0.0] * 6 for code in SYMBOL_ORDER}
        initial_den = [0.0] * 6
        drops = {code: [0.0] * 6 for code in SYMBOL_ORDER}
        drop_den = [0.0] * 6
        for table_name, table_weight in tables:
            strip = by_name[table_name]
            for reel in range(6):
                length = int(strip["reel_lengths"][reel])
                initial_den[reel] += table_weight * length
                for row in range(length):
                    code = code_by_id[int(strip["symbols"][row][reel])]
                    if code in initial:
                        initial[code][reel] += table_weight
                for symbol_index, symbol_id in enumerate(config["symbol_ids"]):
                    weight = table_weight * int(strip["drop_weights"][symbol_index][reel])
                    drop_den[reel] += weight
                    code = code_by_id[int(symbol_id)]
                    if code in drops:
                        drops[code][reel] += weight
        for reel in range(6):
            for code in SYMBOL_ORDER:
                initial[code][reel] /= initial_den[reel]
                drops[code][reel] /= drop_den[reel]
        result[scene]["initial"] = {"by_code": initial}
        result[scene]["drop"] = {"by_code": drops}
    return result


def render_symbol_distribution(competitor: dict, h027: dict) -> str:
    parts = [
        "## 符號分布",
        "",
        "競品依 Response 盤面拆成 R1～R6；每輪每個盤面有 5 格。H027 BG 初始分布為兩張 300 格 Table 以 1:1 合併；FG 依初始排程 8:7 加權。掉落使用各 Table 的 Symbol Weight，並沿用相同場景權重。輪帶中的 `Multiplier / C2` 是倍數球候選；進盤後才依 `weight_use_super_multiplier` 決定保留為 C2 或轉成 C3，因此符號分布表合併列為同一種倍數球。",
    ]
    labels = {
        ("BG", "initial"): "BG 初始 R1-R6",
        ("BG", "drop"): "BG 掉落 R1-R6",
        ("FG", "initial"): "FG 初始 R1-R6",
        ("FG", "drop"): "FG 掉落 R1-R6",
    }
    for scene, mode in labels:
        rows = []
        for code in SYMBOL_ORDER:
            label = SYMBOL_LABELS.get(code, code)
            rows.append([label, "競品", *[pct(value) for value in competitor[scene][mode]["by_code"][code]]])
            rows.append([label, "H027", *[pct(value) for value in h027[scene][mode]["by_code"][code]]])
        parts.extend([
            "",
            f"### {labels[(scene, mode)]}",
            "",
            table(["Symbol", "模型", "R1", "R2", "R3", "R4", "R5", "R6"], rows),
        ])
    return "\n".join(parts)


def render_stack_distribution(metrics: dict) -> str:
    parts = [
        "## 初始盤面堆疊分布",
        "",
        "堆疊定義為同一初始盤面、同一 Reel 內由上到下連續相同符號的最大 run，採 cell-weighted 口徑。例如 `A A A K J` 記為 3 個 Stack 3 cells 與 2 個 Stack 1 cells。競品使用實際初始盤面；H027 對每張 300 格輪帶枚舉所有 stop，BG 兩表按 1:1、FG 兩表按 8:7 加權。掉落不是由連續輪帶 stop 產生，因此本指標只比較初始盤面。",
    ]
    for scene in ("BG", "FG"):
        rows = []
        for length in range(1, 6):
            for model, key in (("競品", "competitor"), ("H027", "h027")):
                values = metrics[key][scene]["by_reel"]
                rows.append([str(length), model, *[pct(values[f"R{reel}"][f"stack_{length}"]) for reel in range(1, 7)]])
        maximum = max(
            abs(metrics["h027"][scene]["by_reel"][f"R{reel}"][f"stack_{length}"]
                - metrics["competitor"][scene]["by_reel"][f"R{reel}"][f"stack_{length}"])
            for reel in range(1, 7) for length in range(1, 6)
        )
        parts.extend([
            "",
            f"### {scene} Stack 1-5／R1-R6",
            "",
            table(["Stack", "模型", "R1", "R2", "R3", "R4", "R5", "R6"], rows),
            "",
            f"{scene} 各 Reel 的最大絕對差異為 {maximum * 100:.4f} pp；Stack 4／5 均與競品同為 0。",
        ])
    symbol_rows = []
    max_symbol_2 = {"BG": 0.0, "FG": 0.0}
    for code in ("M1", "M2", "M3", "M4", "A", "K", "Q", "J", "TE"):
        row = [code]
        for scene in ("BG", "FG"):
            comp = metrics["competitor"][scene]["by_symbol"][code]
            game = metrics["h027"][scene]["by_symbol"][code]
            comp_2, game_2 = 1 - comp["stack_1"], 1 - game["stack_1"]
            comp_3 = sum(comp[f"stack_{length}"] for length in range(3, 6))
            game_3 = sum(game[f"stack_{length}"] for length in range(3, 6))
            max_symbol_2[scene] = max(max_symbol_2[scene], abs(game_2 - comp_2))
            row.extend([pct(comp_2), pct(game_2), pct(comp_3), pct(game_3)])
        symbol_rows.append(row)
    parts.extend([
        "",
        "### 各符號 Stack 2+／3+",
        "",
        table([
            "Symbol", "競品 BG 2+", "H027 BG 2+", "競品 BG 3+", "H027 BG 3+",
            "競品 FG 2+", "H027 FG 2+", "競品 FG 3+", "H027 FG 3+",
        ], symbol_rows),
        "",
        f"BG 各一般符號的 Stack 2+ 最大差異為 {max_symbol_2['BG'] * 100:.2f} pp；FG 最大差異為 {max_symbol_2['FG'] * 100:.2f} pp。C1／C2 不列入各符號表。",
    ])
    return "\n".join(parts)


def validate_distribution_inputs(competitor_reels: dict, h027_reels: dict, stack_metrics: dict) -> None:
    for source_name, source in (("competitor", competitor_reels), ("H027", h027_reels)):
        for scene in ("BG", "FG"):
            for mode in ("initial", "drop"):
                for reel in range(6):
                    total = sum(source[scene][mode]["by_code"][code][reel] for code in SYMBOL_ORDER)
                    if abs(total - 1.0) > 1e-9:
                        raise ValueError(
                            f"{source_name} {scene} {mode} R{reel + 1} symbol rates sum to {total}"
                        )
    for source_name in ("competitor", "h027"):
        for scene in ("BG", "FG"):
            for reel in range(1, 7):
                total = sum(
                    stack_metrics[source_name][scene]["by_reel"][f"R{reel}"][f"stack_{length}"]
                    for length in range(1, 6)
                )
                if abs(total - 1.0) > 1e-9:
                    raise ValueError(f"{source_name} {scene} R{reel} stack rates sum to {total}")


def read_record(path: Path) -> dict:
    workbook = load_workbook(path, data_only=True, read_only=True)
    base = dict(workbook["Base Info"].iter_rows(min_row=2, values_only=True))
    cascades = list(workbook["Cascade"].iter_rows(min_row=2, values_only=True))
    multipliers = list(workbook["C2-C3 Multiplier"].iter_rows(min_row=2, values_only=True))
    symbol_hit_sheet = workbook["Symbol Hit Rate"]
    symbol_hit_headers = [cell.value for cell in next(symbol_hit_sheet.iter_rows(min_row=1, max_row=1))]
    symbol_hits = [dict(zip(symbol_hit_headers, row)) for row in symbol_hit_sheet.iter_rows(min_row=2, values_only=True)]
    workbook.close()
    return {"base": base, "cascades": cascades, "multipliers": multipliers, "symbol_hits": symbol_hits}


def competitor_combo_counts(rows: list[dict]) -> dict[str, int]:
    result = {str(row["combo"]): int(row["count"]) for row in rows}
    result["9+"] = result.get("9", 0) + result.get("10+", 0)
    return result


def simulator_combo_counts(rows: list[tuple]) -> tuple[dict[str, int], dict[str, int]]:
    bg, fg = {}, {}
    for combo, bg_count, fg_count in rows:
        key = str(combo) if int(combo) < 9 else "9+"
        bg[key] = bg.get(key, 0) + int(bg_count)
        fg[key] = fg.get(key, 0) + int(fg_count)
    return bg, fg


def value_counts(rows: list[tuple]) -> tuple[dict[int, int], dict[int, int]]:
    bg, fg = defaultdict(int), defaultdict(int)
    for value, bg_count, fg_count in rows:
        if int(value) <= 2500:
            bg[int(value)] += int(bg_count)
            fg[int(value)] += int(fg_count)
    return dict(bg), dict(fg)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--record", type=Path, default=DEFAULT_RECORD)
    parser.add_argument("--competitor", type=Path, default=DEFAULT_COMPETITOR)
    parser.add_argument("--responses", type=Path, default=DEFAULT_RESPONSES)
    parser.add_argument("--stack-metrics", type=Path, default=DEFAULT_STACK_METRICS)
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    args = parser.parse_args()

    record = read_record(args.record.resolve())
    h = record["base"]
    competitor = json.loads(args.competitor.read_text(encoding="utf-8"))["analysis"]
    config = load_config(args.config.resolve())
    competitor_reels = competitor_reel_distributions(args.responses.resolve())
    h027_reels = h027_reel_distributions(config)
    stack_analyzer = load_module("h027_stack_report", STACK_ANALYZER_PATH)
    stack_metrics = stack_analyzer.analyze(args.config.resolve(), args.responses.resolve())
    validate_distribution_inputs(competitor_reels, h027_reels, stack_metrics)
    args.stack_metrics.write_text(
        json.dumps(stack_metrics, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    symbol_distribution = render_symbol_distribution(competitor_reels, h027_reels)
    stack_distribution = render_stack_distribution(stack_metrics)
    c = competitor["basic"]
    fg_spins = round(float(h["avg_fg_spins"]) * int(h["fg_trigger_count"]))
    h_cycle = 1 / float(h["fg_trigger_rate"])

    basis = f"""### 比較基準

| 項目 | Gates of Olympus 1000 | H027 現在版本 |
|---|---|---|
| 來源 | `game_responses-gates of olympus 1000.xlsx` | `{args.record.name}`（`config_92A.js` + `Simulator.py`） |
| 版本 | 競品實際 Response 樣本 | `H0271` / `{h['version']}` |
| 樣本 | {c['paid_spins']:,} 個付費 Spin、{c['fg_sessions']:,} 場 FG、{c['fg_spins']:,} 個 FG Spin | {int(h['total_rounds']):,} 個 Normal Bet Round、{int(h['fg_trigger_count']):,} 場 FG、{fg_spins:,} 個 FG Spin |
| Card System | 競品實際遊玩資料 | 關閉 |
| Bet Mode | 一般投注 | Normal Bet，Bet Multi 1 |
| 輪帶 | 實際盤面觀測分布 | BG 為 `BG_Symbol` + `BG_Symbol (2)` 1:1；FG 為 `FG_Symbol` + `FG_Symbol (2)`，初始 Spins 排程 8:7；各 Table 的 R1～R6 各 300 格 |

H027 數值為本次 {int(h['total_rounds']):,} 局模擬樣本，不是理論精確值。競品 FG 只有 {c['fg_sessions']} 場，FG RTP、FG 平均倍數與最大得分的波動會明顯高於 H027 樣本。"""

    core_rows = [
        ["RTP", "Total RTP", pct(c["rtp_total"]), pct(h["rtp_total"]), pp(h["rtp_total"] - c["rtp_total"])],
        ["RTP", "BG RTP", pct(c["rtp_bg"]), pct(h["rtp_bg"]), pp(h["rtp_bg"] - c["rtp_bg"])],
        ["RTP", "FG RTP", pct(c["rtp_fg"]), pct(h["rtp_fg"]), pp(h["rtp_fg"] - c["rtp_fg"])],
        ["Hit Rate", "BG Hit Rate", pct(c["hit_rate_bg"]), pct(h["hit_rate_bg"]), pp(h["hit_rate_bg"] - c["hit_rate_bg"])],
        ["Hit Rate", "FG Hit Rate", pct(c["hit_rate_fg"]), pct(h["hit_rate_fg"]), pp(h["hit_rate_fg"] - c["hit_rate_fg"])],
        ["FG", "FG Trigger Rate", pct(c["fg_trigger_rate"]), pct(h["fg_trigger_rate"]), pp(h["fg_trigger_rate"] - c["fg_trigger_rate"])],
        ["FG", "平均觸發週期", f"{c['fg_cycle']:.2f} 局／次", f"{h_cycle:.2f} 局／次", f"H027 慢 {h_cycle - c['fg_cycle']:.2f} 局"],
        ["FG", "平均 Free Spins", f"{c['avg_fg_spins']:.2f}", f"{h['avg_fg_spins']:.4f}", f"{h['avg_fg_spins'] - c['avg_fg_spins']:+.4f}"],
        ["FG", "FG 平均得分倍數", f"{c['avg_fg_multiplier']:.2f}×", f"{h['avg_fg_multiplier']:.2f}×", f"{h['avg_fg_multiplier'] - c['avg_fg_multiplier']:+.2f}×"],
        ["Win", "單次付費事件最大得分", f"{c['max_score_multiplier']:.2f}×", f"{h['max_win_x']:,.2f}×", f"{h['max_win_x'] - c['max_score_multiplier']:+,.2f}×"],
    ]
    core = "### 核心指標比較\n\n" + table(
        ["類別", "指標", "競品樣本", "H027 現在版本", "差異"], core_rows
    ) + (
        f"\n\nH027 的 BG Hit Rate 與競品相差 {pp(h['hit_rate_bg'] - c['hit_rate_bg'])}，"
        f"FG Hit Rate 相差 {pp(h['hit_rate_fg'] - c['hit_rate_fg'])}；"
        f"FG 平均觸發週期則比競品慢 {h_cycle - c['fg_cycle']:.2f} 局。"
        f"目前 Total RTP 高出 {pp(h['rtp_total'] - c['rtp_total'])}，"
        "主要數學差距仍是 BG／FG RTP 與 FG 平均得分；正式 RTP 尚未校正。"
    )

    symbol_order = ["M1", "M2", "M3", "M4", "A", "K", "Q", "J", "TE"]
    competitor_symbol_hits = competitor["symbol_hit_rate_by_bucket"]
    h027_symbol_hits = {row["Symbol"]: row for row in record["symbol_hits"]}

    def symbol_hit_table(scene: str) -> str:
        competitor_rows = {
            row["symbol"]: row["buckets"]
            for row in competitor_symbol_hits[scene.lower()]["symbols"]
        }
        scene_prefix = scene.upper()
        rows = []
        for symbol in symbol_order:
            c_buckets = competitor_rows[symbol]
            h_row = h027_symbol_hits[symbol]
            row = [symbol]
            for label, column_label in [("8-9", "8_9"), ("10-11", "10_11"), ("12+", "12_Plus")]:
                competitor_rate = float(c_buckets[label]["hit_rate"])
                h027_rate = float(h_row[f"{scene_prefix}_{column_label}_Hit_Rate"])
                row.extend([pct(competitor_rate), pct(h027_rate), pp(h027_rate - competitor_rate)])
            rows.append(row)
        return table(
            ["Symbol", "競品 8～9", "H027 8～9", "差異", "競品 10～11", "H027 10～11", "差異", "競品 12+", "H027 12+", "差異"],
            rows,
        )

    symbol_hit = f"""## 得分符號 Hit Rate

Hit Rate 口徑為「得獎 Cascade 次數 ÷ 該 Scene 全部 Spin」。同一 Spin 的同一符號若在後續 Cascade 再次以 8 顆以上得獎，會再計 1 Hit；因此這是得獎事件頻率，不是「至少中過一次」的 Spin 占比。只列 Any-8 的一般得分符號；C1 依 Scatter 顆數賠付，C2／C3 不直接形成得分，因此不列入。

### BG

競品分母為 {int(competitor_symbol_hits['bg']['spin_count']):,} 個 BG Spin；H027 分母為 {int(h['total_rounds']):,} 個 BG Spin。

{symbol_hit_table('BG')}

### FG

競品分母為 {int(competitor_symbol_hits['fg']['spin_count']):,} 個 FG Spin；H027 分母為 {fg_spins:,} 個 FG Spin。競品 FG 僅 {int(competitor_symbol_hits['fg']['spin_count']):,} Spin，單一符號與高顆數區間的抽樣波動較大。

{symbol_hit_table('FG')}"""

    cbg = competitor_combo_counts(competitor["combo"]["bg"])
    cfg = competitor_combo_counts(competitor["combo"]["fg"])
    hbg, hfg = simulator_combo_counts(record["cascades"])
    c_bg_den = sum(cbg.get(str(value), 0) for value in range(1, 9)) + cbg["9+"]
    c_fg_den = sum(cfg.get(str(value), 0) for value in range(1, 9)) + cfg["9+"]
    h_bg_den = sum(hbg.get(str(value), 0) for value in range(1, 9)) + hbg.get("9+", 0)
    h_fg_den = sum(hfg.get(str(value), 0) for value in range(1, 9)) + hfg.get("9+", 0)
    combo_rows = []
    for label in [str(value) for value in range(1, 9)] + ["9+"]:
        crb, chb = cbg.get(label, 0) / c_bg_den, hbg.get(label, 0) / h_bg_den
        crf, chf = cfg.get(label, 0) / c_fg_den, hfg.get(label, 0) / h_fg_den
        combo_rows.append([label, pct(crb), pct(chb), pp(chb - crb), pct(crf), pct(chf), pp(chf - crf)])
    combo = f"""## 消除率

比較只納入「至少發生 1 次得獎消除」的 BG／FG Spin，Combo 0 不納入分母。各 Scene 的 Combo 1～9+ 合計為 100%；`9+` 合併所有消除 9 次以上的 Spin。競品分母為 BG {c_bg_den:,} Spin、FG {c_fg_den:,} Spin；H027 分母為 BG {h_bg_den:,} Spin、FG {h_fg_den:,} Spin。

{table(['Combo', '競品 BG', 'H027 BG', 'BG 差異', '競品 FG', 'H027 FG', 'FG 差異'], combo_rows)}

H027 BG Combo 1 與競品的差異為 {pp(hbg.get('1', 0) / h_bg_den - cbg.get('1', 0) / c_bg_den)}；FG Combo 1 的差異為 {pp(hfg.get('1', 0) / h_fg_den - cfg.get('1', 0) / c_fg_den)}。高 Combo 尾端仍受輪帶群組排列與掉落次數影響。"""

    cballs = competitor["multiplier_ball"]
    h_bg_ball_spins = round(float(h["multiplier_ball_rate_bg"]) * int(h["total_rounds"]))
    h_fg_ball_spins = round(float(h["multiplier_ball_rate_fg"]) * fg_spins)
    h_bg_balls = round(float(h["avg_multiplier_balls_bg"]) * int(h["total_rounds"]))
    h_fg_balls = round(float(h["avg_multiplier_balls_fg"]) * fg_spins)
    balls = f"""## 倍數球出現率

採 Spin-level 口徑：該次 BG／FG Spin 的初始盤面或任一次掉落中，至少出現 1 顆倍數球，該 Spin 計數 1 次。同一顆球跨 Cascade 留盤不重複計數。H027 輪帶先產生倍數球候選，再依初始盤面球數對應的 `weight_use_super_multiplier` 決定為一般 C2 或 C3 / Super；同一 Spin 的掉落球沿用該次初始球數欄位。

{table(
    ['Scene', '競品 Spin 數', '競品有球 Spin', '競品出現率', 'H027 Spin 數', 'H027 有球 Spin', 'H027 出現率', '差異'],
    [
        ['BG', f"{cballs['bg']['spin_count']:,}", f"{cballs['bg']['spins_with_ball']:,}", pct(cballs['bg']['spin_appearance_rate']), f"{int(h['total_rounds']):,}", f"{h_bg_ball_spins:,}", pct(h['multiplier_ball_rate_bg']), pp(h['multiplier_ball_rate_bg'] - cballs['bg']['spin_appearance_rate'])],
        ['FG', f"{cballs['fg']['spin_count']:,}", f"{cballs['fg']['spins_with_ball']:,}", pct(cballs['fg']['spin_appearance_rate']), f"{fg_spins:,}", f"{h_fg_ball_spins:,}", pct(h['multiplier_ball_rate_fg']), pp(h['multiplier_ball_rate_fg'] - cballs['fg']['spin_appearance_rate'])],
    ],
)}

{table(
    ['Scene', '競品去重後球數', '競品平均球數／Spin', 'H027 球數', 'H027 平均球數／Spin', '平均球數差異'],
    [
        ['BG', f"{cballs['bg']['deduplicated_ball_count']:,}", f"{cballs['bg']['balls_per_spin']:.4f}", f"{h_bg_balls:,}", f"{h['avg_multiplier_balls_bg']:.4f}", f"{h['avg_multiplier_balls_bg'] - cballs['bg']['balls_per_spin']:+.4f}"],
        ['FG', f"{cballs['fg']['deduplicated_ball_count']:,}", f"{cballs['fg']['balls_per_spin']:.4f}", f"{h_fg_balls:,}", f"{h['avg_multiplier_balls_fg']:.4f}", f"{h['avg_multiplier_balls_fg'] - cballs['fg']['balls_per_spin']:+.4f}"],
    ],
)}

H027 BG 倍數球出現率差異為 {pp(h['multiplier_ball_rate_bg'] - cballs['bg']['spin_appearance_rate'])}；FG 差異為 {pp(h['multiplier_ball_rate_fg'] - cballs['fg']['spin_appearance_rate'])}。"""

    h_bg_values, h_fg_values = value_counts(record["multipliers"])
    c_bg_values = {int(row["value"]): int(row["count"]) for row in cballs["bg"]["value_distribution"]}
    c_fg_values = {int(row["value"]): int(row["count"]) for row in cballs["fg"]["value_distribution"]}
    values = sorted(set(c_bg_values) | set(c_fg_values) | set(h_bg_values) | set(h_fg_values))
    c_bg_total, c_fg_total = sum(c_bg_values.values()), sum(c_fg_values.values())
    h_bg_total, h_fg_total = sum(h_bg_values.values()), sum(h_fg_values.values())
    c_bg_high = sum(count for value, count in c_bg_values.items() if value >= 10) / c_bg_total
    c_fg_high = sum(count for value, count in c_fg_values.items() if value >= 10) / c_fg_total
    h_bg_high = sum(count for value, count in h_bg_values.items() if value >= 10) / h_bg_total
    h_fg_high = sum(count for value, count in h_fg_values.items() if value >= 10) / h_fg_total
    value_rows = []
    for value in values:
        cb, hb = c_bg_values.get(value, 0) / c_bg_total, h_bg_values.get(value, 0) / h_bg_total
        cf, hf = c_fg_values.get(value, 0) / c_fg_total, h_fg_values.get(value, 0) / h_fg_total
        value_rows.append([f"{value}×", pct(cb), pct(hb), pp(hb - cb), pct(cf), pct(hf), pp(hf - cf)])
    multiplier = f"""## 倍數球上的倍率分布

口徑為各倍率球數 ÷ 該 Scene 全部倍數球數，採 Spin 結束時的最終顯示倍率。H027 的 C2 一般倍率池只含 2～8×；所有 10× 以上初始倍率均由 `weight_super_multiplier` 產生，且 C3 每次得獎消除後會升一級。Super 初始權重已反向補償升級效果，因此比較的是升級後分布，而不是直接把競品最終分布原樣抄入初始權重。

{table(['倍率', '競品 BG', 'H027 BG', 'BG 差異', '競品 FG', 'H027 FG', 'FG 差異'], value_rows)}

10× 以上大倍數球占全部倍數球：競品 BG {pct(c_bg_high)}、H027 BG {pct(h_bg_high)}（{pp(h_bg_high - c_bg_high)}）；競品 FG {pct(c_fg_high)}、H027 FG {pct(h_fg_high)}（{pp(h_fg_high - c_fg_high)}）。

BG 樣本用來校正 Super pool 的主要形狀；競品 FG 的 10× 以上球數很少，FG 高倍細項僅作觀察，不宜逐格追樣本雜訊。2500× 不是競品直接抽取倍率，但 H027 的 C3 可由較低倍率逐級升至 2500×，因此仍列入最終顯示分布。"""

    text = args.report.read_text(encoding="utf-8")
    toc_entry = "- [得分符號 Hit Rate](#得分符號-hit-rate)\n"
    if toc_entry not in text:
        text = text.replace("- [消除率](#消除率)\n", toc_entry + "- [消除率](#消除率)\n")
    text = replace_between(text, "### 比較基準", "### 核心指標比較", basis)
    text = replace_between(text, "### 核心指標比較", "## 符號分布", core)
    text = replace_between(text, "## 符號分布", "## 初始盤面堆疊分布", symbol_distribution)
    text = replace_between(text, "## 初始盤面堆疊分布", "## 得分符號 Hit Rate", stack_distribution)
    if "## 得分符號 Hit Rate" in text:
        text = replace_between(text, "## 得分符號 Hit Rate", "## 消除率", symbol_hit)
    else:
        text = text.replace("## 消除率", symbol_hit.rstrip() + "\n\n## 消除率")
    text = replace_between(text, "## 消除率", "## 倍數球出現率", combo)
    text = replace_between(text, "## 倍數球出現率", "## 倍數球上的倍率分布", balls)
    text = replace_between(text, "## 倍數球上的倍率分布", None, multiplier)
    args.report.write_text(text.rstrip() + "\n", encoding="utf-8")
    written = args.report.read_text(encoding="utf-8")
    required_headings = (
        "## Overview", "## 符號分布", "## 初始盤面堆疊分布",
        "## 得分符號 Hit Rate", "## 消除率", "## 倍數球出現率",
        "## 倍數球上的倍率分布",
    )
    for heading in required_headings:
        if written.count(heading) != 1:
            raise ValueError(f"Report heading must appear exactly once: {heading}")
    print(f"Updated: {args.report.resolve()}")


if __name__ == "__main__":
    main()
