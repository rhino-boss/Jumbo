from collections import defaultdict
from copy import copy
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment


ROOT = Path(__file__).resolve().parent
CURRENT_WORKBOOK = Path(r"C:\Users\rhinshen\Mine\個人工作區\市場資訊\H5\遊戲資源\PP - Gates of Olympus 1000\還原輪帶_Gates_of_Olympus_1000.xlsx")
TEMPLATE_FILE = Path.cwd() / "Project" / "Slots" / "C027_奧林帕斯 2500" / "Source" / "C0271.xlsx"
OUTPUT_FILE = ROOT / "還原輪帶_Gates_of_Olympus_1000.xlsx"

SYMBOL_NAMES = {
    1: "C1",
    3: "M1",
    4: "M2",
    5: "M3",
    6: "M4",
    7: "A",
    8: "K",
    9: "Q",
    10: "J",
    11: "TE",
    12: "C2",
}


def copy_cell_style(source, target):
    target._style = copy(source._style)
    if source.has_style:
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def prepare_sheet(workbook, title, template_sheet, max_data_rows):
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = template_sheet.sheet_view.showGridLines
    sheet.sheet_view.zoomScale = template_sheet.sheet_view.zoomScale
    sheet.freeze_panes = None

    for column in range(11, 18):
        letter = template_sheet.cell(1, column).column_letter
        sheet.column_dimensions[letter].width = template_sheet.column_dimensions[letter].width
    for row in range(1, max_data_rows + 4):
        if template_sheet.row_dimensions[row].height is not None:
            sheet.row_dimensions[row].height = template_sheet.row_dimensions[row].height
        for column in range(11, 18):
            copy_cell_style(template_sheet.cell(row, column), sheet.cell(row, column))

    sheet["K2"] = "Symbol"
    headers = ["Line #", "R1", "R2", "R3", "R4", "R5", "R6"]
    for offset, value in enumerate(headers, start=11):
        sheet.cell(3, offset, value)
    return sheet


def load_reconstructed_data():
    workbook = load_workbook(CURRENT_WORKBOOK, read_only=True, data_only=True)
    long_sheet = workbook["Reel_Strips_Long"]
    reels = defaultdict(lambda: defaultdict(dict))
    for reel_set, _role, reel_name, index, symbol_id, _symbol_code in long_sheet.iter_rows(min_row=2, values_only=True):
        reels[int(reel_set)][int(str(reel_name).removeprefix("R"))][int(index)] = int(symbol_id)
    reconstructed = {
        reel_set: [
            [values[index] for index in sorted(values)]
            for _reel, values in sorted(reels[reel_set].items())
        ]
        for reel_set in sorted(reels)
    }

    entry_sheet = workbook["Set6_BF_Entry"]
    boards = []
    for row in entry_sheet.iter_rows(min_row=5, values_only=True):
        value = row[1]
        if not value:
            continue
        board = tuple(int(symbol) for symbol in str(value).split(","))
        if len(board) == 30:
            boards.append(board)
    return reconstructed, boards


def build_workbook():
    reconstructed, set6_boards = load_reconstructed_data()
    template_workbook = load_workbook(TEMPLATE_FILE, data_only=False)
    template_sheet = template_workbook["BG_Symbol"]

    workbook = Workbook()
    workbook.remove(workbook.active)

    for reel_set in range(5):
        reels = reconstructed[reel_set]
        max_length = max(len(reel) for reel in reels)
        sheet = prepare_sheet(workbook, f"Reel Set {reel_set}", template_sheet, max_length)
        sheet["K2"].comment = Comment(
            "循環輪帶起點可任意旋轉。陣列方向為上方至下方；補牌時由目前可見頂端的前一格反向循環取值。",
            "Codex",
        )
        for line_index in range(max_length):
            row = line_index + 4
            sheet.cell(row, 11, line_index)
            for reel_index, reel in enumerate(reels, start=1):
                symbol = reel[line_index] if line_index < len(reel) else None
                sheet.cell(row, 11 + reel_index, SYMBOL_NAMES.get(symbol, str(symbol)) if symbol is not None else None)

    entry_sheet = prepare_sheet(workbook, "BF Entry Set 6", template_sheet, len(set6_boards))
    entry_sheet["K2"] = "BF Entry Boards"
    entry_sheet["K2"].comment = Comment(
        "Set 6 是 Buy Feature 強制進場板，不是一般循環輪帶。每個觀測盤面固定四個 C1 Scatter。",
        "Codex",
    )
    entry_sheet["K3"] = "Board #"
    for board_index, board in enumerate(set6_boards):
        row = board_index + 4
        entry_sheet.cell(row, 11, board_index)
        for reel in range(6):
            column_symbols = [SYMBOL_NAMES[board[board_row * 6 + reel]] for board_row in range(5)]
            entry_sheet.cell(row, 12 + reel, " / ".join(column_symbols))

    workbook.active = 0
    workbook.save(OUTPUT_FILE)

    check = load_workbook(OUTPUT_FILE, read_only=True, data_only=True)
    expected_sheets = [f"Reel Set {index}" for index in range(5)] + ["BF Entry Set 6"]
    if check.sheetnames != expected_sheets:
        raise RuntimeError(f"Unexpected sheet order: {check.sheetnames}")
    for sheet_name in expected_sheets:
        sheet = check[sheet_name]
        if sheet["K3"].value not in ("Line #", "Board #"):
            raise RuntimeError(f"Invalid table header in {sheet_name}")
        if [sheet.cell(3, column).value for column in range(12, 18)] != ["R1", "R2", "R3", "R4", "R5", "R6"]:
            raise RuntimeError(f"Invalid reel headers in {sheet_name}")

    print(OUTPUT_FILE)
    print(check.sheetnames)
    print(f"set6_boards={len(set6_boards)}")


if __name__ == "__main__":
    build_workbook()
