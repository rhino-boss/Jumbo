"""Derive competitor reel-set and per-stop usage from response initial screens."""

from __future__ import annotations

import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


COMP_ROOT = Path(r"C:\Users\rhinshen\Mine\個人工作區\市場資訊\H5\遊戲資源\PP - Gates of Olympus 1000")
REEL_BOOK = COMP_ROOT / "還原輪帶_Gates_of_Olympus_1000.xlsx"
DATA_DIR = COMP_ROOT / "遊玩資料"
OUTPUT = Path(__file__).resolve().parent / "competitor_stop_weights.json"
ID_TO_CODE = {1: "C1", 3: "M1", 4: "M2", 5: "M3", 6: "M4", 7: "A", 8: "K", 9: "Q", 10: "J", 11: "TE", 12: "C2"}


def number(value):
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def empty(value):
    return value is None or str(value).strip() == ""


def parse_screen(value):
    if empty(value):
        return None
    try:
        values = [int(item) for item in str(value).split(",")]
    except ValueError:
        return None
    return values if len(values) == 30 and all(item in ID_TO_CODE for item in values) else None


def load_reels():
    workbook = load_workbook(REEL_BOOK, read_only=True, data_only=True)
    result = {}
    try:
        for set_id in range(5):
            sheet = workbook[f"Reel Set {set_id}"]
            rows = list(sheet.iter_rows(min_row=2, values_only=True))
            result[set_id] = []
            for reel in range(6):
                result[set_id].append([str(row[1 + reel]) for row in rows if row[1 + reel] is not None])
    finally:
        workbook.close()
    return result


def matching_starts(sequence, window):
    length = len(sequence)
    return [
        start for start in range(length)
        if all(sequence[(start + offset) % length] == symbol for offset, symbol in enumerate(window))
    ]


def main():
    reels = load_reels()
    set_counts = Counter()
    fs_set_counts = Counter()
    stop_counts = defaultdict(Counter)
    ambiguity = Counter()
    seen = set()

    for path in sorted(DATA_DIR.glob("*.xlsx")):
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook["game_data"]
            header = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            columns = {str(name): index for index, name in enumerate(header) if name is not None}
            required = {"rid", "index", "stime", "reel_set", "s", "rs_p", "rs_t", "fs"}
            if not required.issubset(columns):
                continue
            is_buy = "buyfeature" in path.name.lower()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if empty(row[columns["rid"]]) or not empty(row[columns["rs_t"]]):
                    continue
                rs_p = number(row[columns["rs_p"]])
                if rs_p not in (None, 0):
                    continue
                set_id = number(row[columns["reel_set"]])
                screen = parse_screen(row[columns["s"]])
                if set_id not in reels or screen is None:
                    continue
                key = (row[columns["stime"]], row[columns["index"]], row[columns["s"]])
                if key in seen:
                    continue
                seen.add(key)
                fs_index = number(row[columns["fs"]]) or 0
                scene = "BF" if is_buy else ("FG" if fs_index > 0 else "BG")
                set_counts[(scene, set_id)] += 1
                if scene in {"FG", "BF"}:
                    fs_set_counts[(scene, fs_index, set_id)] += 1
                for reel in range(6):
                    window = [ID_TO_CODE[screen[row_index * 6 + reel]] for row_index in range(5)]
                    candidates = matching_starts(reels[set_id][reel], window)
                    ambiguity[(scene, "no_match" if not candidates else "unique" if len(candidates) == 1 else "multi")] += 1
                    if candidates:
                        share = 1.0 / len(candidates)
                        for start in candidates:
                            stop_counts[(scene, set_id, reel)][start] += share
        finally:
            workbook.close()

    payload = {
        "source": str(REEL_BOOK),
        "set_counts": {
            scene: {str(set_id): set_counts[(scene, set_id)] for set_id in range(5)}
            for scene in ("BG", "FG", "BF")
        },
        "fs_set_counts": {
            scene: {
                str(fs_index): {str(set_id): fs_set_counts[(scene, fs_index, set_id)] for set_id in range(5)}
                for fs_index in sorted({key[1] for key in fs_set_counts if key[0] == scene})
            }
            for scene in ("FG", "BF")
        },
        "ambiguity": {
            scene: {kind: ambiguity[(scene, kind)] for kind in ("unique", "multi", "no_match")}
            for scene in ("BG", "FG", "BF")
        },
        "stop_counts": {},
        "stop_stats": {},
    }
    for (scene, set_id, reel), counts in sorted(stop_counts.items()):
        values = [
            counts.get(stop, 0.0) for stop in range(len(reels[set_id][reel]))
        ]
        payload["stop_counts"].setdefault(scene, {}).setdefault(str(set_id), {})[f"R{reel + 1}"] = values
        payload["stop_stats"].setdefault(scene, {}).setdefault(str(set_id), {})[f"R{reel + 1}"] = {
            "total": sum(values), "zero": sum(value == 0 for value in values),
            "min_positive": min((value for value in values if value > 0), default=0), "max": max(values),
        }
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"set_counts": payload["set_counts"], "ambiguity": payload["ambiguity"], "stop_stats": payload["stop_stats"]}, ensure_ascii=False, indent=2))
    print(f"written={OUTPUT}")


if __name__ == "__main__":
    main()
