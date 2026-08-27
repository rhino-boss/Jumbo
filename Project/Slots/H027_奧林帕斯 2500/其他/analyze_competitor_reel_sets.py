from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\rhinshen\Mine\個人工作區\市場資訊\H5\遊戲資源\PP - Gates of Olympus 1000")
DATA_DIR = ROOT / "遊玩資料"


def empty(value):
    return value is None or str(value).strip() == ""


def number(value):
    if empty(value):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def main():
    counts = Counter()
    seen = set()
    duplicate_rows = 0
    for path in sorted(DATA_DIR.glob("*.xlsx")):
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook["game_data"]
        header = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        columns = {str(name): index for index, name in enumerate(header) if name is not None}
        required = {"rid", "index", "reel_set", "s", "rs_p", "rs_t", "fs"}
        if not required.issubset(columns):
            print(f"skip {path.name}: missing {sorted(required - columns.keys())}")
            continue
        is_buy = "buyfeature" in path.name.lower()
        local = Counter()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            reel_set = number(row[columns["reel_set"]])
            if reel_set is None or empty(row[columns["s"]]) or not empty(row[columns["rs_t"]]):
                continue
            rs_p = number(row[columns["rs_p"]])
            if rs_p not in (None, 0):
                continue
            key = (row[columns["rid"]], row[columns["index"]], reel_set, row[columns["s"]])
            if key in seen:
                duplicate_rows += 1
                continue
            seen.add(key)
            fs = number(row[columns["fs"]])
            if is_buy:
                scene = "BF_ENTRY" if reel_set == 6 else "BF_FG"
            else:
                scene = "FG" if fs is not None and fs > 0 else "BG"
            local[(scene, reel_set)] += 1
            counts[(scene, reel_set)] += 1
        print(f"\n{path.name}")
        for item, count in sorted(local.items()):
            print(item, count)
    print(f"\nunique_initial_screens={len(seen)} duplicate_rows={duplicate_rows}")
    for scene in ("BG", "FG", "BF_FG", "BF_ENTRY"):
        rows = [(reel_set, count) for (row_scene, reel_set), count in counts.items() if row_scene == scene]
        total = sum(count for _, count in rows)
        print(f"\n{scene} total={total}")
        for reel_set, count in sorted(rows):
            print(f"Set {reel_set}: {count} ({count / total:.6%})")


if __name__ == "__main__":
    main()
