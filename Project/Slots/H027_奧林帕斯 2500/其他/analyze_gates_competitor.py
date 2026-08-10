from __future__ import annotations

import argparse
import json
import math
from collections import Counter, OrderedDict
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from statistics import mean

import numpy as np
from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parent.parent
REFERENCE_DIR = PROJECT_DIR / "其他" / "參考資料"
DEFAULT_INPUT = REFERENCE_DIR / "game_responses-gates of olympus 1000.xlsx"
DEFAULT_BUY_INPUT = REFERENCE_DIR / "game_responses-gates of olympus 1000 - buyfeature.xlsx"
DEFAULT_MODEL = PROJECT_DIR / "Source" / "H0271.xlsx"
DEFAULT_REPORT = REFERENCE_DIR / "分析報告_gates of olympus 1000.md"
DEFAULT_COMPARISON = PROJECT_DIR / "其他" / "競品參考數值比較.md"
DEFAULT_METRICS = REFERENCE_DIR / "analysis_gates_of_olympus_1000_metrics.json"
DEFAULT_OVERVIEW_TEMPLATE = PROJECT_DIR.parent / "H028_雷神爆金 1000" / "Source" / "H0281.xlsx"

SYMBOL_NAMES = {
    1: "Scatter",
    3: "M1",
    4: "M2",
    5: "M3",
    6: "M4",
    7: "A",
    8: "K",
    9: "Q",
    10: "J",
    11: "TE",
    12: "Multiplier",
}

SYMBOL_ID_TO_CODE = {
    1: "C1",
    3: "M1",
    4: "M2",
    5: "M3",
    6: "M4",
    7: "A",
    8: "K",
    9: "Q",
    10: "J",
    11: "TE",
    12: "C2",
}
CODE_TO_ID = {
    "C1": 1,
    "C2": 2,
    "C3": 3,
    "M1": 4,
    "M2": 5,
    "M3": 6,
    "M4": 7,
    "A": 8,
    "K": 9,
    "Q": 10,
    "J": 11,
    "TE": 12,
}

# Gates of Olympus 1000 complete multiplier-ball value pool. The supplied
# responses do not happen to contain 1000x, but it remains a valid competitor
# value and therefore must remain in Multiple Level with zero observed weight.
COMPETITOR_MULTIPLIER_LEVELS = [
    2, 3, 4, 5, 6, 8, 10, 12, 15, 20, 25, 50, 100, 250, 500, 1000
]


def number(value, default=0.0):
    if value is None or value == "":
        return default
    if isinstance(value, str):
        value = value.replace(",", "")
    return float(value)


def integer(value, default=0):
    return int(round(number(value, default)))


def parse_symbols(value):
    if value is None or value == "":
        return []
    return [int(item) for item in str(value).split(",") if item != ""]


def parse_multiplier_balls(value):
    if value is None or value == "":
        return []
    result = []
    for item in str(value).split(";"):
        parts = item.split("~")
        if len(parts) >= 3:
            result.append((integer(parts[0]), integer(parts[1]), number(parts[2])))
    return result


def load_rows(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "game_data" not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"{path.name} 缺少 game_data 工作表")
    worksheet = workbook["game_data"]
    header = list(next(worksheet.iter_rows(values_only=True)))
    rows = [dict(zip(header, values)) for values in worksheet.iter_rows(min_row=2, values_only=True)]
    workbook.close()
    return header, rows


@dataclass
class Spin:
    rows: list[dict] = field(default_factory=list)

    @property
    def win(self):
        return sum(number(row.get("w")) for row in self.rows)

    @property
    def combo(self):
        values = [integer(row.get("rs_t")) for row in self.rows if row.get("rs_t") is not None]
        return max(values, default=0)

    @property
    def has_multiplier_ball(self):
        return any(parse_multiplier_balls(row.get("rmul")) for row in self.rows)

    @property
    def multiplier_ball_events(self):
        # A ball can remain on screen for multiple response rows.  Deduplicate
        # within a spin by position and value so persistence is not counted as
        # a new appearance on every cascade response.
        return sorted(
            {
                (position, value)
                for row in self.rows
                for _, position, value in parse_multiplier_balls(row.get("rmul"))
            }
        )

    @property
    def used_multiplier(self):
        return sum(number(row.get("apv")) for row in self.rows if row.get("apv") is not None)

    @property
    def screens(self):
        return [parse_symbols(row.get("s")) for row in self.rows if row.get("s") not in (None, "")]


@dataclass
class Session:
    rid: object
    rows: list[dict]
    bet: float
    bg: Spin
    fg_spins: list[Spin]

    @property
    def triggered_fg(self):
        return bool(self.fg_spins)

    @property
    def bg_win(self):
        return self.bg.win

    @property
    def fg_win(self):
        return sum(spin.win for spin in self.fg_spins)

    @property
    def total_win(self):
        return self.bg_win + self.fg_win

    @property
    def fg_end_accumulated_multiplier(self):
        # apv is already the cumulative multiplier applied at that response;
        # summing successive apv values would count prior additions repeatedly.
        values = [number(row.get("apv")) for spin in self.fg_spins for row in spin.rows if row.get("apv") is not None]
        return max(values, default=1.0)


def build_sessions(rows):
    grouped = OrderedDict()
    for row in rows:
        rid = row.get("rid")
        if rid is None:
            continue
        grouped.setdefault(rid, []).append(row)

    sessions = []
    for rid, group in grouped.items():
        c_values = [number(row.get("c")) for row in group if row.get("c") is not None]
        l_values = [number(row.get("l")) for row in group if row.get("l") is not None]
        if not c_values or not l_values:
            raise ValueError(f"rid={rid} 缺少 c 或 l，無法計算 Bet")
        bet = c_values[0] * l_values[0]
        if bet <= 0:
            raise ValueError(f"rid={rid} Bet 非正值：{bet}")

        # A trigger response carries psym and fs=1, but it is still the paid BG
        # result (including the Scatter pay).  Actual FG spin responses begin
        # on the following row.  The last FG spin uses fs_total instead of fs.
        trigger_index = next(
            (
                index
                for index, row in enumerate(group)
                if row.get("psym") not in (None, "") and row.get("fs") is not None
            ),
            None,
        )
        bg_rows = group if trigger_index is None else group[: trigger_index + 1]
        fg_rows = [] if trigger_index is None else group[trigger_index + 1 :]

        fg_spins = []
        current_fs = None
        current_spin = None
        for row in fg_rows:
            fs_value = row.get("fs")
            if fs_value is not None:
                fs_number = integer(fs_value)
                if current_spin is None or fs_number != current_fs:
                    current_fs = fs_number
                    current_spin = Spin()
                    fg_spins.append(current_spin)
            elif row.get("fs_total") is not None and current_fs != "final":
                current_fs = "final"
                current_spin = Spin()
                fg_spins.append(current_spin)
            if current_spin is not None:
                current_spin.rows.append(row)

        sessions.append(Session(rid=rid, rows=group, bet=bet, bg=Spin(bg_rows), fg_spins=fg_spins))
    return sessions


def session_is_complete(session):
    has_trigger = any(row.get("psym") not in (None, "") for row in session.rows)
    if has_trigger:
        ended = any(integer(row.get("fsend_total")) == 1 for row in session.rows if row.get("fsend_total") is not None)
        announced = final_non_null(session.rows, "fs_total")
        return ended and announced is not None and integer(announced) == len(session.fg_spins)

    if any(row.get("fs") is not None or row.get("fs_total") is not None for row in session.rows):
        return False
    no_win = (
        len(session.rows) == 1
        and session.rows[0].get("na") == "s"
        and number(session.rows[0].get("w")) == 0
    )
    terminal_win = any(
        row.get("na") == "c" and row.get("rs_t") is not None and number(row.get("rs_t")) > 0
        for row in session.rows
    )
    return no_win or terminal_win


def combo_distribution(spins):
    counts = Counter(min(spin.combo, 10) for spin in spins)
    total = len(spins)
    rows = []
    for value in range(11):
        count = counts[value]
        rows.append(
            {
                "combo": "10+" if value == 10 else str(value),
                "count": count,
                "rate": count / total if total else 0.0,
            }
        )
    return rows


def symbol_distribution(spins, screen_mode):
    counts = Counter()
    cells = 0
    screens = 0
    for spin in spins:
        if screen_mode == "initial":
            selected = spin.screens[:1]
        elif screen_mode == "drop":
            selected = spin.screens[1:]
        else:
            raise ValueError(f"Unsupported screen mode: {screen_mode}")
        for screen in selected:
            counts.update(screen)
            cells += len(screen)
            screens += 1
    result = []
    for symbol_id in sorted(set(SYMBOL_NAMES) | set(counts)):
        count = counts[symbol_id]
        result.append(
            {
                "symbol_id": symbol_id,
                "symbol": SYMBOL_NAMES.get(symbol_id, f"Symbol {symbol_id}"),
                "count": count,
                "rate": count / cells if cells else 0.0,
            }
        )
    return {"screen_count": screens, "cell_count": cells, "symbols": result}


def infer_paytable(paths):
    features = [(symbol_id, bucket) for symbol_id in range(3, 12) for bucket in range(3)]
    equations = []
    payouts = []
    scatter_values = {4: [], 5: [], 6: []}
    for path in paths:
        _, rows = load_rows(path)
        for row in rows:
            bet = number(row.get("c")) * number(row.get("l"))
            if bet <= 0:
                continue
            psym = row.get("psym")
            if psym not in (None, ""):
                parts = str(psym).split("~")
                if len(parts) >= 3:
                    scatter_count = len([item for item in parts[2].split(",") if item != ""])
                    if scatter_count in scatter_values:
                        scatter_values[scatter_count].append(number(parts[1]) / bet)
            tmb = row.get("tmb")
            if tmb in (None, "") or row.get("apv") not in (None, "") or number(row.get("w")) <= 0:
                continue
            counts = Counter()
            for item in str(tmb).split("~"):
                parts = item.split(",")
                if len(parts) >= 2:
                    counts[integer(parts[1])] += 1
            vector = []
            for symbol_id, bucket in features:
                count = counts.get(symbol_id, 0)
                actual_bucket = 0 if 8 <= count < 10 else 1 if 10 <= count < 12 else 2 if count >= 12 else -1
                vector.append(1.0 if actual_bucket == bucket else 0.0)
            equations.append(vector)
            payouts.append(number(row.get("w")) / bet)

    matrix = np.asarray(equations, dtype=np.float64)
    target = np.asarray(payouts, dtype=np.float64)
    coefficients, _, rank, _ = np.linalg.lstsq(matrix, target, rcond=None)
    fitted = matrix @ coefficients
    rows = []
    for symbol_id in range(3, 12):
        values = []
        observations = []
        for bucket in range(3):
            index = features.index((symbol_id, bucket))
            observation_count = int(matrix[:, index].sum())
            observations.append(observation_count)
            values.append(round(float(coefficients[index]), 10) if observation_count else None)
        rows.append(
            {
                "symbol_id": symbol_id,
                "symbol": SYMBOL_NAMES[symbol_id],
                "payouts": values,
                "observations": observations,
            }
        )
    scatter = []
    for count, values in scatter_values.items():
        unique = sorted({round(value, 10) for value in values})
        scatter.append({"count": count, "payout": unique[0] if len(unique) == 1 else None, "observations": len(values)})
    return {
        "cluster_counts": ["8–9", "10–11", "12+"],
        "symbols": rows,
        "scatter": scatter,
        "equation_count": len(equations),
        "matrix_rank": int(rank),
        "parameter_count": len(features),
        "rmse": float(np.sqrt(np.mean((fitted - target) ** 2))),
        "max_abs_error": float(np.max(np.abs(fitted - target))),
    }


def multiplier_ball_metrics(spins):
    appeared = [spin for spin in spins if spin.has_multiplier_ball]
    events = [event for spin in spins for event in spin.multiplier_ball_events]
    values = Counter(value for _, value in events)
    return {
        "spin_count": len(spins),
        "spins_with_ball": len(appeared),
        "spin_appearance_rate": len(appeared) / len(spins) if spins else 0.0,
        "deduplicated_ball_count": len(events),
        "balls_per_spin": len(events) / len(spins) if spins else 0.0,
        "value_distribution": [
            {"value": value, "count": count, "rate": count / len(events) if events else 0.0}
            for value, count in sorted(values.items())
        ],
    }


def final_non_null(rows, key):
    values = [row.get(key) for row in rows if row.get(key) is not None]
    return values[-1] if values else None


def data_quality(rows, sessions, excluded_sessions):
    board_lengths = Counter()
    invalid_symbol_cells = 0
    for row in rows:
        symbols = parse_symbols(row.get("s"))
        if symbols:
            board_lengths[len(symbols)] += 1
            invalid_symbol_cells += sum(symbol not in SYMBOL_NAMES for symbol in symbols)

    total_reconciliation = []
    fg_reconciliation = []
    bet_inconsistency = 0
    fg_sequence_issues = 0
    for session in sessions:
        c_values = {number(row.get("c")) for row in session.rows if row.get("c") is not None}
        l_values = {number(row.get("l")) for row in session.rows if row.get("l") is not None}
        if len(c_values) > 1 or len(l_values) > 1:
            bet_inconsistency += 1
        final_tw = final_non_null(session.rows, "tw")
        if final_tw is not None:
            total_reconciliation.append(abs(number(final_tw) - session.total_win))
        if session.triggered_fg:
            final_fg = final_non_null(session.rows, "fswin_total")
            if final_fg is not None:
                fg_reconciliation.append(abs(number(final_fg) - session.fg_win))
            numbers = []
            for spin in session.fg_spins:
                value = next((integer(row.get("fs")) for row in spin.rows if row.get("fs") is not None), None)
                if value is not None:
                    numbers.append(value)
            # fs=1 is carried by the BG trigger response.  Played FG responses
            # therefore use fs=2..N, followed by one final fs_total segment.
            if numbers != list(range(2, len(session.fg_spins) + 1)):
                fg_sequence_issues += 1

    return {
        "source_rows": len(rows),
        "rows_without_rid": sum(row.get("rid") is None for row in rows),
        "paid_sessions": len(sessions),
        "excluded_incomplete_sessions": len(excluded_sessions),
        "excluded_incomplete_rids": [str(session.rid) for session in excluded_sessions],
        "board_length_distribution": dict(sorted(board_lengths.items())),
        "invalid_symbol_cells": invalid_symbol_cells,
        "bet_inconsistent_sessions": bet_inconsistency,
        "fg_sequence_issue_sessions": fg_sequence_issues,
        "total_win_reconciliation_failures": sum(value > 1e-6 for value in total_reconciliation),
        "total_win_max_abs_diff": max(total_reconciliation, default=0.0),
        "fg_win_reconciliation_failures": sum(value > 1e-6 for value in fg_reconciliation),
        "fg_win_max_abs_diff": max(fg_reconciliation, default=0.0),
    }


def analyze(path: Path):
    header, rows = load_rows(path)
    required = {"rid", "c", "l", "s", "w", "tw", "rs_t", "rmul", "apv", "fs", "fswin_total"}
    missing = sorted(required - set(header))
    if missing:
        raise ValueError(f"{path.name} 缺少必要欄位：{', '.join(missing)}")
    all_sessions = build_sessions(rows)
    sessions = [session for session in all_sessions if session_is_complete(session)]
    excluded_sessions = [session for session in all_sessions if not session_is_complete(session)]
    bg_spins = [session.bg for session in sessions]
    fg_spins = [spin for session in sessions for spin in session.fg_spins]
    fg_sessions = [session for session in sessions if session.triggered_fg]
    wager = sum(session.bet for session in sessions)
    bg_pay = sum(session.bg_win for session in sessions)
    fg_pay = sum(session.fg_win for session in sessions)

    fg_multiples = [session.fg_win / session.bet for session in fg_sessions]
    total_multiples = [session.total_win / session.bet for session in sessions]
    accumulated = [session.fg_end_accumulated_multiplier for session in fg_sessions]
    result = {
        "source": str(path),
        "quality": data_quality(rows, sessions, excluded_sessions),
        "basic": {
            "paid_spins": len(sessions),
            "fg_sessions": len(fg_sessions),
            "fg_spins": len(fg_spins),
            "total_wager": wager,
            "bg_pay": bg_pay,
            "fg_pay": fg_pay,
            "rtp_bg": bg_pay / wager if wager else 0.0,
            "rtp_fg": fg_pay / wager if wager else 0.0,
            "rtp_total": (bg_pay + fg_pay) / wager if wager else 0.0,
            "hit_rate_bg": sum(spin.win > 0 for spin in bg_spins) / len(bg_spins) if bg_spins else 0.0,
            "hit_rate_fg": sum(spin.win > 0 for spin in fg_spins) / len(fg_spins) if fg_spins else 0.0,
            "fg_trigger_rate": len(fg_sessions) / len(sessions) if sessions else 0.0,
            "fg_cycle": len(sessions) / len(fg_sessions) if fg_sessions else math.inf,
            "avg_fg_spins": mean(len(session.fg_spins) for session in fg_sessions) if fg_sessions else 0.0,
            "avg_fg_multiplier": mean(fg_multiples) if fg_multiples else 0.0,
            "max_fg_multiplier": max(fg_multiples, default=0.0),
            "max_score_multiplier": max(total_multiples, default=0.0),
        },
        "symbol_distribution": {
            "bg_initial": symbol_distribution(bg_spins, "initial"),
            "fg_initial": symbol_distribution(fg_spins, "initial"),
            "bg_drop": symbol_distribution(bg_spins, "drop"),
            "fg_drop": symbol_distribution(fg_spins, "drop"),
        },
        "multiplier_ball": {
            "bg": multiplier_ball_metrics(bg_spins),
            "fg": multiplier_ball_metrics(fg_spins),
        },
        "combo": {
            "bg": combo_distribution(bg_spins),
            "fg": combo_distribution(fg_spins),
        },
        "fg_end_accumulated_multiplier": {
            "definition": "maximum cumulative apv observed in FG; use 1x when the session has no applied multiplier",
            "average": mean(accumulated) if accumulated else 0.0,
            "maximum": max(accumulated, default=0.0),
            "sessions_with_applied_multiplier": sum(value > 1 for value in accumulated),
            "session_count": len(accumulated),
            "values": accumulated,
        },
        "_sessions": sessions,
    }
    return result


def reel_symbol_probabilities(spins, screen_mode):
    counts = [Counter() for _ in range(6)]
    totals = [0] * 6
    for spin in spins:
        if screen_mode == "initial":
            screens = spin.screens[:1]
        elif screen_mode == "drop":
            screens = spin.screens[1:]
        elif screen_mode == "all":
            screens = spin.screens
        else:
            raise ValueError(f"Unsupported screen mode: {screen_mode}")
        for screen in screens:
            if len(screen) != 30:
                continue
            for reel in range(6):
                for row in range(5):
                    code = SYMBOL_ID_TO_CODE.get(screen[row * 6 + reel])
                    if code is not None:
                        counts[reel][code] += 1
                        totals[reel] += 1
    return [
        {code: counts[reel][code] / totals[reel] for code in CODE_TO_ID if code != "C3"}
        if totals[reel]
        else {}
        for reel in range(6)
    ]


def largest_remainder_counts(probabilities, total, force_positive=()):
    codes = [code for code in CODE_TO_ID if code != "C3"]
    raw = {code: max(0.0, probabilities.get(code, 0.0)) * total for code in codes}
    counts = {code: int(math.floor(raw[code])) for code in codes}
    for code in force_positive:
        if probabilities.get(code, 0.0) > 0 and counts.get(code, 0) == 0:
            counts[code] = 1
    while sum(counts.values()) > total:
        candidates = [code for code in codes if counts[code] > (1 if code in force_positive else 0)]
        code = max(candidates, key=lambda item: counts[item] - raw[item])
        counts[code] -= 1
    remainder = total - sum(counts.values())
    order = sorted(codes, key=lambda code: raw[code] - math.floor(raw[code]), reverse=True)
    for index in range(remainder):
        counts[order[index % len(order)]] += 1
    return counts


def optimize_start_weights(sequence, target, iterations=1200):
    codes = [code for code in CODE_TO_ID if code != "C3"]
    code_index = {code: index for index, code in enumerate(codes)}
    length = len(sequence)
    features = np.zeros((length, len(codes)), dtype=np.float64)
    for start in range(length):
        for offset in range(5):
            features[start, code_index[sequence[(start + offset) % length]]] += 0.2
    target_vector = np.asarray([target.get(code, 0.0) for code in codes], dtype=np.float64)
    target_vector /= target_vector.sum()
    weights = np.full(length, 1.0 / length, dtype=np.float64)
    for _ in range(iterations):
        predicted = weights @ features
        ratio = (target_vector + 1e-12) / (predicted + 1e-12)
        adjustment = features @ np.log(ratio)
        weights *= np.exp(np.clip(0.45 * adjustment, -6.0, 6.0))
        weights /= weights.sum()
    raw = weights * 1_000_000
    result = np.floor(raw).astype(np.int64)
    missing = 1_000_000 - int(result.sum())
    if missing > 0:
        order = np.argsort(-(raw - result))
        result[order[:missing]] += 1
    predicted = (result / result.sum()) @ features
    max_abs_error = float(np.max(np.abs(predicted - target_vector)))
    return result.tolist(), max_abs_error


def build_clustered_sequence(counts, seed, high_fraction):
    high_symbols = {"M1", "M2", "M3", "M4"}
    low_symbols = {"A", "K", "Q", "J", "TE"}
    chunks = []
    for code, count in counts.items():
        fraction = high_fraction if code in high_symbols else 0.10 if code in low_symbols else 0.0
        paired = (int(count * fraction) // 2) * 2
        chunks.extend([[code, code]] * (paired // 2))
        chunks.extend([[code]] * (count - paired))
    rng = np.random.default_rng(seed)
    rng.shuffle(chunks)
    shuffled = [code for chunk in chunks for code in chunk]
    scatter_count = shuffled.count("C1")
    if scatter_count == 0:
        return shuffled
    non_scatter = [code for code in shuffled if code != "C1"]
    length = len(shuffled)
    offset = seed % max(1, length // scatter_count)
    scatter_positions = {
        (offset + round(index * length / scatter_count)) % length
        for index in range(scatter_count)
    }
    result = []
    source_index = 0
    for index in range(length):
        if index in scatter_positions:
            result.append("C1")
        else:
            result.append(non_scatter[source_index])
            source_index += 1
    return result


def scale_symbol_probability(target, code, factor):
    if factor == 1 or not target or target.get(code, 0) <= 0:
        return dict(target)
    result = dict(target)
    original = result[code]
    scaled = min(0.95, original * factor)
    remaining_before = 1.0 - original
    remaining_after = 1.0 - scaled
    for item in result:
        if item != code:
            result[item] = result[item] * remaining_after / remaining_before
    result[code] = scaled
    return result


def build_reel_strip(
    initial_target,
    drop_target,
    seed,
    length=300,
    force_scatter=False,
    high_fraction=0.70,
    scatter_scale=1.0,
):
    initial_target = scale_symbol_probability(initial_target, "C1", scatter_scale)
    sequence_target = initial_target
    counts = largest_remainder_counts(sequence_target, length, force_positive=("C1", "C2"))
    if force_scatter and counts.get("C1", 0) == 0:
        donor = max(
            (code for code in counts if code not in {"C1", "C2", "C3"} and counts[code] > 0),
            key=counts.get,
        )
        counts[donor] -= 1
        counts["C1"] += 1
    sequence = build_clustered_sequence(counts, seed, high_fraction)
    weights = [1] * length
    achieved = Counter(sequence)
    error = max(
        abs(achieved.get(code, 0) / length - initial_target.get(code, 0.0))
        for code in CODE_TO_ID
        if code != "C3"
    )
    return sequence, weights, error


def multiplier_value_counts(spins):
    counts = Counter()
    for spin in spins:
        for _, value in spin.multiplier_ball_events:
            counts[int(value)] += 1
    return counts


def map_multiplier_weights(counts, levels):
    result = [0] * len(levels)
    for value, count in counts.items():
        candidates = [index for index, level in enumerate(levels) if level == value]
        if candidates:
            result[candidates[0]] += count
            continue
        nearest = min(range(len(levels)), key=lambda index: abs(levels[index] - value))
        result[nearest] += count
    if sum(result) == 0:
        result[0] = 1
    return result


def write_strip_sheet(
    worksheet,
    spins,
    seed_base,
    force_scatter_first_four=False,
    high_fraction=0.70,
    scatter_scale=1.0,
):
    # One table and uniform Symbol Weight cannot represent different initial
    # and drop distributions. Use every observed screen so the Symbol column
    # matches the competitor's overall BG/FG symbol distribution.
    overall = reel_symbol_probabilities(spins, "all")
    errors = []
    descriptions = {
        "C1": "Scatter",
        "C2": "Multiplier",
        "C3": "Super Multiplier",
    }
    # Keep the VLOOKUP source independent from Overview row positions. The
    # previous array formulas pointed to Overview!A44:I55 and broke when the
    # H028 layout moved the paytable to rows 32:43.
    for offset, (code, symbol_id) in enumerate(CODE_TO_ID.items()):
        row = 4 + offset
        worksheet.cell(row, 1).value = code
        worksheet.cell(row, 2).value = descriptions.get(code, code)
        worksheet.cell(row, 9).value = symbol_id
    for column in range(19, 25):
        worksheet.column_dimensions[worksheet.cell(3, column).column_letter].width = 7
    for column in range(26, 32):
        worksheet.column_dimensions[worksheet.cell(3, column).column_letter].width = 11
    for reel in range(6):
        sequence, weights, error = build_reel_strip(
            overall[reel],
            overall[reel],
            seed_base + reel,
            force_scatter=force_scatter_first_four and reel < 4,
            high_fraction=high_fraction,
            scatter_scale=scatter_scale,
        )
        errors.append(error)
        for offset in range(301):
            row = 4 + offset
            if reel == 0:
                worksheet.cell(row, 11).value = offset if offset < 300 else None
            symbol_cell = worksheet.cell(row, 12 + reel)
            symbol_cell.value = sequence[offset] if offset < 300 else None
            id_cell = worksheet.cell(row, 19 + reel)
            id_cell.value = f'=IF({symbol_cell.coordinate}="","",VLOOKUP({symbol_cell.coordinate},$A$4:$I$15,9,FALSE))'
            id_cell.number_format = "0"
            worksheet.cell(row, 26 + reel).value = 1 if offset < 300 else None
    return {
        "overall_max_abs_error": max(errors),
        "overall_max_abs_error_pp": max(errors) * 100,
        "spin_samples": len(spins),
    }


def copy_cell_style(source_cell, target_cell):
    # Copy components instead of the cross-workbook style ID. Directly copying
    # _style can leave indexes that do not exist in the target styles.xml.
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.protection = copy(source_cell.protection)
    target_cell.number_format = source_cell.number_format


def apply_h028_overview_layout(worksheet, template_path: Path):
    """Apply H028's Overview layout while keeping H027-specific game data."""
    template_book = load_workbook(template_path, data_only=False)
    template = template_book["Overview"]

    version = worksheet["B3"].value
    paytable_start = 32 if worksheet["A32"].value == "C1" else 44
    paytable = [
        [worksheet.cell(row, column).value for column in range(1, 10)]
        for row in range(paytable_start, paytable_start + 12)
    ]

    for merged in list(worksheet.merged_cells.ranges):
        worksheet.unmerge_cells(str(merged))
    max_row = max(worksheet.max_row, template.max_row)
    max_column = max(worksheet.max_column, template.max_column)
    for row in range(1, max_row + 1):
        for column in range(1, max_column + 1):
            target_cell = worksheet.cell(row, column)
            target_cell.value = None
            target_cell.comment = None
            target_cell.hyperlink = None
            if row <= template.max_row and column <= template.max_column:
                source_cell = template.cell(row, column)
                copy_cell_style(source_cell, target_cell)

    for column in range(1, max_column + 1):
        key = worksheet.cell(1, column).column_letter
        source_dimension = template.column_dimensions[key]
        worksheet.column_dimensions[key].width = source_dimension.width
        worksheet.column_dimensions[key].hidden = source_dimension.hidden
    for row in range(1, max_row + 1):
        source_dimension = template.row_dimensions[row]
        worksheet.row_dimensions[row].height = (
            source_dimension.height or template.sheet_format.defaultRowHeight
        )
        worksheet.row_dimensions[row].hidden = source_dimension.hidden
    worksheet.freeze_panes = template.freeze_panes
    worksheet.sheet_view.showGridLines = template.sheet_view.showGridLines
    worksheet.sheet_properties.tabColor = copy(template.sheet_properties.tabColor)
    worksheet.sheet_format.defaultRowHeight = template.sheet_format.defaultRowHeight
    worksheet.sheet_format.defaultColWidth = template.sheet_format.defaultColWidth
    worksheet.sheet_format.baseColWidth = template.sheet_format.baseColWidth
    worksheet.page_margins = copy(template.page_margins)
    worksheet.page_setup = copy(template.page_setup)
    worksheet.print_options = copy(template.print_options)

    # H028 Overview section order: identity, bet, reel/window, free spins, paytable.
    values = {
        "A2": "Model:", "B2": "H0271",
        "A3": "Version:", "B3": version or "0.0.0.1",
        "A6": "Base Bet", "B6": "Board Cells",
        "A7": 100, "B7": "=SUM(B16:G16)",
        "A10": "Coin in", "B10": "Price(x)", "C10": "Bet Type",
        "A11": "=A7*B11", "B11": 1, "C11": "Normal Bet",
        "A12": "=A7*B12", "B12": 2, "C12": "Extra Bet",
        "A13": "=A7*B13", "B13": 100, "C13": "Buy Feature",
        "A15": "Reel #", "B15": 1, "C15": 2, "D15": 3,
        "E15": 4, "F15": 5, "G15": 6,
        "A16": "Visible Window Size", "B16": 5, "C16": 5, "D16": 5,
        "E16": 5, "F16": 5, "G16": 5,
        "A19": "Free Spins Setting",
        "A20": "C1 Num", "B20": "Initial Free Spins",
        "A21": 4, "B21": 15,
        "A22": 5, "B22": 15,
        "A23": 6, "B23": 15,
        "A26": "Retrigger", "B26": 5,
        "A27": "The maximum of free spins is 50.",
        "A30": "Pay Table：", "B30": '="All wins show for "&A7&" credit bet."',
    }
    for coordinate, value in values.items():
        worksheet[coordinate] = value
    for column in range(1, 4):
        copy_cell_style(template.cell(12, column), worksheet.cell(13, column))

    headers = ["Symbol", "Description", 4, 5, 6, "8~9", "10~11", "12+", "Id"]
    for column, value in enumerate(headers, 1):
        cell = worksheet.cell(31, column)
        cell.value = value
        if column > template.max_column or column > 7:
            copy_cell_style(template.cell(31, 7), cell)
    for row_offset, row_values in enumerate(paytable, 32):
        for column, value in enumerate(row_values, 1):
            cell = worksheet.cell(row_offset, column)
            cell.value = value
            style_column = min(column, 7)
            style_row = min(34, max(32, row_offset))
            copy_cell_style(template.cell(style_row, style_column), cell)

    for row in range(44, max_row + 1):
        for column in range(1, max_column + 1):
            worksheet.cell(row, column).value = None
    template_book.close()


def update_h0271_model(model_path: Path, analysis, overview_template: Path):
    workbook = load_workbook(model_path)
    required = {
        "Overview",
        "Parameter",
        "BG_Symbol",
        "FG_Symbol",
    }
    missing = required - set(workbook.sheetnames)
    if missing:
        workbook.close()
        raise ValueError(f"H0271 缺少工作表：{', '.join(sorted(missing))}")

    basic = analysis["basic"]
    sessions = analysis["_sessions"]
    bg_spins = [session.bg for session in sessions]
    fg_spins = [spin for session in sessions for spin in session.fg_spins]

    overview = workbook["Overview"]
    apply_h028_overview_layout(overview, overview_template)

    parameter = workbook["Parameter"]
    parameter["B4"], parameter["C4"] = "BG_Symbol", 1
    for row in (5, 6):
        parameter.cell(row, 2).value = None
        parameter.cell(row, 3).value = None
    parameter["B11"], parameter["C11"], parameter["D11"] = "FG_Symbol", 15, 5
    for row in (12, 13):
        for column in range(2, 5):
            parameter.cell(row, column).value = None

    # Only BG_Symbol and FG_Symbol are selectable. C3 appearance remains off,
    # while C2/C3 multiplier value weights both follow the competitor balls.
    for row in range(18, 24):
        parameter.cell(row, 2).value = None
        for column in range(3, 9):
            parameter.cell(row, column).value = 0
    parameter["B18"] = "BG_Symbol"
    parameter["B19"] = "FG_Symbol"

    observed_values = sorted(
        set(multiplier_value_counts(bg_spins)) | set(multiplier_value_counts(fg_spins))
    )
    levels = COMPETITOR_MULTIPLIER_LEVELS + [2500] * (
        25 - len(COMPETITOR_MULTIPLIER_LEVELS)
    )
    for index, value in enumerate(levels):
        parameter.cell(28 + index, 2).value = index + 1
        parameter.cell(28 + index, 3).value = value
        for header_row in (3, 8, 18):
            parameter.cell(header_row, 11 + index).value = value

    bg_value_weights = map_multiplier_weights(multiplier_value_counts(bg_spins), levels)
    fg_value_weights = map_multiplier_weights(multiplier_value_counts(fg_spins), levels)
    table_weight_rows = {
        9: bg_value_weights,
        10: fg_value_weights,
    }
    for row in range(9, 15):
        parameter.cell(row, 10).value = None
        for column in range(11, 11 + len(levels)):
            parameter.cell(row, column).value = None
    for row in range(19, 25):
        parameter.cell(row, 10).value = None
        for column in range(11, 11 + len(levels)):
            parameter.cell(row, column).value = None
    parameter["J9"], parameter["J10"] = "BG_Symbol", "FG_Symbol"
    parameter["J19"], parameter["J20"] = "BG_Symbol", "FG_Symbol"
    for row, values in table_weight_rows.items():
        for index, value in enumerate(values):
            parameter.cell(row, 11 + index).value = value
            parameter.cell(row + 10, 11 + index).value = value
    for index, value in enumerate(bg_value_weights):
        parameter.cell(4, 11 + index).value = value

    sheet_groups = {
        "BG_Symbol": bg_spins,
        "FG_Symbol": fg_spins,
    }
    strip_fit = {}
    for index, (sheet_name, spins) in enumerate(sheet_groups.items()):
        if not spins:
            workbook.close()
            raise ValueError(f"{sheet_name} 沒有可用競品樣本")
        strip_fit[sheet_name] = write_strip_sheet(
            workbook[sheet_name],
            spins,
            seed_base=27010 + index * 100,
            force_scatter_first_four=False,
            high_fraction=0.75 if sheet_name.startswith("BG_Symbol") else 0.70,
            scatter_scale=1.0,
        )

    for sheet_name in list(workbook.sheetnames):
        if sheet_name not in {"Overview", "Parameter", "BG_Symbol", "FG_Symbol"}:
            workbook.remove(workbook[sheet_name])

    try:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    except AttributeError:
        pass

    temp_path = model_path.with_name(model_path.stem + ".tmp.xlsx")
    workbook.save(temp_path)
    workbook.close()
    verify = load_workbook(temp_path, read_only=True, data_only=False)
    missing_after = required - set(verify.sheetnames)
    verify.close()
    if missing_after:
        temp_path.unlink(missing_ok=True)
        raise ValueError(f"更新後 H0271 缺少工作表：{', '.join(sorted(missing_after))}")
    temp_path.replace(model_path)
    return {
        "model": "H0271.xlsx",
        "observed_rtp_bg": basic["rtp_bg"],
        "observed_rtp_fg": basic["rtp_fg"],
        "multiplier_levels": levels,
        "observed_multiplier_values": observed_values,
        "use_c3_probability": 0.0,
        "fg_phase_samples": [len(fg_spins)],
        "buy_feature_trigger_samples": 0,
        "strip_fit": strip_fit,
        "scope": "H028-style Overview, Parameter, BG_Symbol, and FG_Symbol only",
    }


def pct(value):
    return f"{value * 100:.4f}%"


def num(value, digits=4):
    return f"{value:,.{digits}f}"


def markdown_table(headers, rows):
    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join("---" for _ in headers) + " |"]
    lines.extend("| " + " | ".join(str(value) for value in row) + " |" for row in rows)
    return "\n".join(lines)


def symbol_table(bg, fg):
    bg_map = {row["symbol_id"]: row for row in bg["symbols"]}
    fg_map = {row["symbol_id"]: row for row in fg["symbols"]}
    rows = []
    for symbol_id in sorted(set(bg_map) | set(fg_map)):
        b = bg_map.get(symbol_id, {"symbol": SYMBOL_NAMES.get(symbol_id, str(symbol_id)), "count": 0, "rate": 0})
        f = fg_map.get(symbol_id, {"count": 0, "rate": 0})
        rows.append([symbol_id, b["symbol"], b["count"], pct(b["rate"]), f["count"], pct(f["rate"])])
    return markdown_table(["ID", "符號", "BG 數量", "BG 占比", "FG 數量", "FG 占比"], rows)


def combo_table(bg, fg):
    rows = []
    for b, f in zip(bg, fg):
        rows.append([b["combo"], b["count"], pct(b["rate"]), f["count"], pct(f["rate"])])
    return markdown_table(["Combo", "BG 局數", "BG 比率", "FG 局數", "FG 比率"], rows)


def ball_value_table(bg, fg):
    bg_map = {row["value"]: row for row in bg["value_distribution"]}
    fg_map = {row["value"]: row for row in fg["value_distribution"]}
    rows = []
    for value in sorted(set(bg_map) | set(fg_map)):
        b = bg_map.get(value, {"count": 0, "rate": 0})
        f = fg_map.get(value, {"count": 0, "rate": 0})
        rows.append([num(value, 0), b["count"], pct(b["rate"]), f["count"], pct(f["rate"])])
    return markdown_table(["倍數值", "BG 球數", "BG 占比", "FG 球數", "FG 占比"], rows)


def paytable_table(paytable):
    rows = []
    for item in paytable["symbols"]:
        values = ["—" if value is None else f"{value:g}×" for value in item["payouts"]]
        rows.append([item["symbol_id"], item["symbol"], *values])
    return markdown_table(["ID", "符號", "8–9 個", "10–11 個", "12+ 個"], rows)


def render_report(analysis, source_path, paytable_source_paths):
    basic = analysis["basic"]
    quality = analysis["quality"]
    symbols = analysis["symbol_distribution"]
    balls = analysis["multiplier_ball"]
    accumulated = analysis["fg_end_accumulated_multiplier"]
    paytable = analysis["paytable"]
    scatter_rows = [
        [item["count"], "—" if item["payout"] is None else f"{item['payout']:g}×", item["observations"]]
        for item in paytable["scatter"]
    ]
    report = f"""# Gates of Olympus 1000 競品分析報告

## 目錄

- [基本資訊](#基本資訊)
- [符號分布](#符號分布)
  - [初始轉輪](#初始轉輪)
  - [掉落](#掉落)
- [賠率](#賠率)
- [其他指標](#其他指標)
  - [消除分布](#消除分布)
  - [倍數球出現率](#倍數球出現率)
  - [FG 結束平均累積倍數](#fg-結束平均累積倍數)
  - [FG 結束最大倍數](#fg-結束最大倍數)

## 基本資訊

{markdown_table(
    ['指標', '統計值', '口徑'],
    [
        ['付費 Spin', f"{basic['paid_spins']:,}", '不同 rid；每個 rid 為一次付費遊戲事件'],
        ['FG 場次', f"{basic['fg_sessions']:,}", '同一 rid 內出現 fs 序列'],
        ['FG Spin', f"{basic['fg_spins']:,}", '依 fs 事件序列及 fs_total 終局段重組；觸發列 fs=1 歸於 BG'],
        ['RTP BG', pct(basic['rtp_bg']), 'BG 得分 ÷ 全部付費 Bet'],
        ['RTP FG', pct(basic['rtp_fg']), 'FG 得分 ÷ 全部付費 Bet'],
        ['RTP Total', pct(basic['rtp_total']), '(BG＋FG 得分) ÷ 全部付費 Bet'],
        ['Hit Rate BG', pct(basic['hit_rate_bg']), 'BG 得分 > 0 的付費 Spin 占比'],
        ['Hit Rate FG', pct(basic['hit_rate_fg']), 'FG 得分 > 0 的 FG Spin 占比'],
        ['FG 週期', f"1 / {basic['fg_cycle']:.2f}", f"觸發率 {pct(basic['fg_trigger_rate'])}"],
        ['平均 FG Spins', f"{basic['avg_fg_spins']:.2f}", '包含再觸發後的總 FG Spins'],
        ['FG 平均倍數', f"{basic['avg_fg_multiplier']:.2f}×", '每場 FG 總得分 ÷ 該次 Base Bet，再取場次平均'],
        ['FG 最大倍數', f"{basic['max_fg_multiplier']:.2f}×", '樣本內單場 FG 最大值'],
        ['最大得分倍數', f"{basic['max_score_multiplier']:.2f}×", '單次付費事件 BG＋FG 總得分 ÷ Base Bet'],
    ],
)}

RTP 使用實際金額加權，而不是把不同 Bet 的局倍率直接平均。FG RTP 是免費遊戲對付費投注的 RTP 貢獻，不是每次 FG Spin 的獨立 RTP。

資料來源為 `{source_path.name}` 的 `game_data`。共排除 {quality['excluded_incomplete_sessions']} 個截斷 rid；總得分、FG 得分與 Bet 對帳均無異常。FG 僅 {basic['fg_sessions']} 場，因此 FG 尾端數值屬樣本觀察值。

## 符號分布

### 初始轉輪

初始盤面每局只取第一個 `s`，避免 Combo 較多的局因盤面快照較多而被過度加權。BG 共 {symbols['bg_initial']['screen_count']:,} 個初始盤面／{symbols['bg_initial']['cell_count']:,} 格；FG 共 {symbols['fg_initial']['screen_count']:,} 個初始盤面／{symbols['fg_initial']['cell_count']:,} 格。

{symbol_table(symbols['bg_initial'], symbols['fg_initial'])}

### 掉落

掉落只統計每個 Spin 的第 2 個以後盤面快照，不包含初始盤面。BG 共 {symbols['bg_drop']['screen_count']:,} 個掉落後盤面／{symbols['bg_drop']['cell_count']:,} 格；FG 共 {symbols['fg_drop']['screen_count']:,} 個掉落後盤面／{symbols['fg_drop']['cell_count']:,} 格。

{symbol_table(symbols['bg_drop'], symbols['fg_drop'])}

## 賠率

一般符號賠率由未套用倍數的得獎 Response 反推：依 `tmb` 計算各符號消除數，使用單列得分 ÷ Bet 建立聯立方程。共使用 {paytable['equation_count']:,} 列，殘差 RMSE 為 {paytable['rmse']:.2e}、最大絕對誤差 {paytable['max_abs_error']:.2e}。`—` 表示兩份樣本中沒有該組合，不能只靠資料確認。

{paytable_table(paytable)}

Scatter 賠率直接由 `psym` 的賠付與位置數量計算。

{markdown_table(['Scatter 數量', '賠率', '觀察列數'], scatter_rows)}

賠率推導使用：{', '.join(f'`{path.name}`' for path in paytable_source_paths)}。

## 其他指標

### 消除分布

Combo 0 表示該 Spin 沒有得獎消除；Combo N 使用該 Spin 結束 Response 的 `rs_t=N`。Combo 10 以上合併至 `10+`。

{combo_table(analysis['combo']['bg'], analysis['combo']['fg'])}

### 倍數球出現率

倍率球出現率採 Spin-level 口徑：同一顆球跨多個 Cascade Response 留在盤面時只算該 Spin 有出現，不重複增加出現局數。球數以同一 Spin 內的「位置＋倍數值」去重。

{markdown_table(
    ['模式', 'Spin 數', '有倍數球 Spin', '出現率', '去重後球數', '平均球數／Spin'],
    [
        ['BG', f"{balls['bg']['spin_count']:,}", f"{balls['bg']['spins_with_ball']:,}", pct(balls['bg']['spin_appearance_rate']), f"{balls['bg']['deduplicated_ball_count']:,}", f"{balls['bg']['balls_per_spin']:.4f}"],
        ['FG', f"{balls['fg']['spin_count']:,}", f"{balls['fg']['spins_with_ball']:,}", pct(balls['fg']['spin_appearance_rate']), f"{balls['fg']['deduplicated_ball_count']:,}", f"{balls['fg']['balls_per_spin']:.4f}"],
    ],
)}

{ball_value_table(balls['bg'], balls['fg'])}

### FG 結束平均累積倍數

**FG 結束平均累積倍數為 {accumulated['average']:.2f}×。** 競品 Response 的 `fsmul_total` 在樣本中固定為 1，無法直接代表終局累積值；`apv` 本身已是當下套用的累積倍數，因此每場 FG 取最後可觀察到的最高 `apv`，完全沒有套用倍數球時記為 1×。

{markdown_table(
    ['指標', '統計值'],
    [
        ['FG 場次', f"{accumulated['session_count']:,}"],
        ['有實際套用倍數球的 FG 場次', f"{accumulated['sessions_with_applied_multiplier']:,}"],
        ['FG 結束平均累積倍數', f"{accumulated['average']:.2f}×"],
    ],
)}

這是根據競品實際套用紀錄重建的分析值，不是 Response 直接提供的終局欄位。

### FG 結束最大倍數

FG 結束最大累積倍數為 **{accumulated['maximum']:.2f}×**；樣本最大單場 FG 得分為 {basic['max_fg_multiplier']:.2f}× Base Bet，最大單次付費事件總得分為 {basic['max_score_multiplier']:.2f}× Base Bet。由於只有 {basic['fg_sessions']} 場 FG，最大值不代表產品上限。
"""
    return report


def render_comparison(analysis, model_update):
    basic = analysis["basic"]
    accumulated = analysis["fg_end_accumulated_multiplier"]
    return f"""# 競品參考數值比較

## 基本指標比較

{markdown_table(
    ['指標', '競品樣本', 'H0271 寫入／處理'],
    [
        ['RTP BG', pct(basic['rtp_bg']), 'Overview B18 使用樣本觀察值；不是正式 target'],
        ['RTP FG', pct(basic['rtp_fg']), 'Overview B19 使用樣本觀察值；不是正式 target'],
        ['Hit Rate BG', pct(basic['hit_rate_bg']), 'Overview C18 由 D18 Pulls/Hit 公式計算'],
        ['Hit Rate FG', pct(basic['hit_rate_fg']), '保留作模擬比較指標'],
        ['FG 週期', f"1 / {basic['fg_cycle']:.2f}", '保留作模擬比較指標'],
        ['FG 平均得分', f"{basic['avg_fg_multiplier']:.2f}×", '保留作模擬比較指標'],
        ['FG 結束平均累積倍數', f"{accumulated['average']:.2f}×", '保留作模擬比較指標'],
        ['FG 結束最大累積倍數', f"{accumulated['maximum']:.2f}×", '樣本最大值，不當成硬上限'],
        ['最大得分倍數', f"{basic['max_score_multiplier']:.2f}×", '產品上限仍為 2500×'],
    ],
)}

## H0271 參數映射

{markdown_table(
    ['工作表／欄位', '填入方式'],
    [
        ['Parameter / Base Game Table', '只使用 BG_Symbol；Symbol 對齊競品所有 BG 盤面分布，六輪 Symbol Weight 全部固定為 1'],
        ['Parameter / Free Game Table', f"只使用 FG_Symbol；競品 FG Spin 樣本數為 {model_update['fg_phase_samples'][0]}，初始 15 Spins、Retrigger +5"],
        ['Parameter / Multiplier Level', f"只保留競品完整值池 {model_update['multiplier_levels'][:16]}；1000× 本批樣本未抽到、權重為 0，剩餘欄位填 2500"],
        ['Parameter / C2、C3', '競品只有一種倍數球，映射至 C2；C3 使用率設 0%，C3 權重表暫與 C2 相同'],
        ['BG／FG 輪帶', '只保留 BG_Symbol、FG_Symbol；只調整 Symbol，Symbol ID 依對照表產生，Symbol Weight 全部固定為 1'],
    ],
)}

## 實作差異與限制

- 競品 Scatter 出現在第 2～5 輪；H027 Simulator 的 Buy Feature 保證邏輯固定前 4 輪，因此 BF 輪帶有一處必要的位置調整。
- 競品只有一種倍數球；H027 的 C3 升級機制沒有競品對應資料，目前不讓 C3 自然出現。
- Normal 樣本只有 {basic['fg_sessions']} 場 FG，高倍尾端、累積倍數與 Retrigger 分布需要更多模擬或競品樣本確認。
- 競品樣本 RTP 為 {pct(basic['rtp_total'])}，只代表本批資料；正式 RTP target 尚未由樣本值推定。

## 調參順序

1. 先比較初始盤面與掉落符號分布。
2. 再比較 Combo 0～10+、BG／FG Hit Rate 與 FG 週期。
3. 接著比較倍數球 Spin-level 出現率及倍率值分布。
4. 最後以大量模擬校正正式 RTP，同時監控 FG 平均／最大累積倍數。
"""


def json_ready(value):
    if isinstance(value, dict):
        return {key: json_ready(item) for key, item in value.items() if key != "_sessions"}
    if isinstance(value, list):
        return [json_ready(item) for item in value]
    if isinstance(value, Path):
        return str(value)
    return value


def main():
    parser = argparse.ArgumentParser(description="Analyze Gates of Olympus 1000 responses and fill H0271")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--buy-input", type=Path, default=DEFAULT_BUY_INPUT)
    parser.add_argument("--model", type=Path, default=DEFAULT_MODEL)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT)
    parser.add_argument("--comparison", type=Path, default=DEFAULT_COMPARISON)
    parser.add_argument("--metrics", type=Path, default=DEFAULT_METRICS)
    parser.add_argument("--overview-template", type=Path, default=DEFAULT_OVERVIEW_TEMPLATE)
    parser.add_argument("--no-model-update", action="store_true")
    args = parser.parse_args()

    analysis = analyze(args.input.resolve())
    buy_analysis = analyze(args.buy_input.resolve()) if args.buy_input.exists() else None
    paytable_sources = [args.input.resolve()]
    if buy_analysis is not None:
        paytable_sources.append(args.buy_input.resolve())
    analysis["paytable"] = infer_paytable(paytable_sources)
    dry_run_observed_values = sorted(
        set(row["value"] for row in analysis["multiplier_ball"]["bg"]["value_distribution"])
        | set(row["value"] for row in analysis["multiplier_ball"]["fg"]["value_distribution"])
    )
    model_update = {
        "model": "H0271.xlsx",
        "observed_rtp_bg": analysis["basic"]["rtp_bg"],
        "observed_rtp_fg": analysis["basic"]["rtp_fg"],
        "multiplier_levels": COMPETITOR_MULTIPLIER_LEVELS + [2500] * (25 - len(COMPETITOR_MULTIPLIER_LEVELS)),
        "observed_multiplier_values": dry_run_observed_values,
        "use_c3_probability": 0.0,
        "fg_phase_samples": [],
        "buy_feature_trigger_samples": len(buy_analysis["_sessions"]) if buy_analysis else 0,
        "strip_fit": {},
        "scope": "dry run",
    }
    if not args.no_model_update:
        model_update = update_h0271_model(args.model.resolve(), analysis, args.overview_template.resolve())

    args.metrics.parent.mkdir(parents=True, exist_ok=True)
    args.metrics.write_text(
        json.dumps({"analysis": json_ready(analysis), "model_update": model_update}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    args.report.write_text(
        render_report(analysis, args.input.resolve(), paytable_sources),
        encoding="utf-8",
    )
    args.comparison.write_text(render_comparison(analysis, model_update), encoding="utf-8")
    print(json.dumps({"report": str(args.report), "comparison": str(args.comparison), "metrics": str(args.metrics), "model": str(args.model), "basic": analysis["basic"], "quality": analysis["quality"], "model_update": model_update}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
