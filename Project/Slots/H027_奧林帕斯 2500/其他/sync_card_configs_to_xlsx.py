"""Synchronize H027 card range weights from config_92A/94A into Source XLSX."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent


def load_config(path: Path) -> dict:
    text = path.read_text(encoding="utf-8-sig")
    return json.loads(text[text.index("{"): text.rindex("}") + 1])


def range_cards(cards: list[dict]) -> list[dict]:
    return [card for card in cards if card.get("type", "range") == "range"]


def write_block(sheet, start_row: int, cards: list[dict]) -> None:
    ranges = range_cards(cards)
    if len(ranges) != 64:
        raise ValueError(f"Expected 64 range cards, found {len(ranges)}")
    for offset, card in enumerate(ranges):
        sheet.cell(start_row + offset, 11).value = int(card["weight"])
    free_cards = [card for card in cards if card.get("type") == "free_game"]
    if free_cards:
        sheet.cell(start_row + 64, 11).value = int(free_cards[0]["weight"])


def sync(config_path: Path, workbook_path: Path) -> None:
    config = load_config(config_path)
    card_system = config["card_system"]
    newbie_normal = card_system["newbie"]["normal_bet"]
    oldhand_normal = card_system["oldhand"]["normal_bet"]
    oldhand_buy = card_system["oldhand"]["buy_feature"]
    small_normal = oldhand_normal["small_bet"]
    small_buy = oldhand_buy["small_bet"]
    big_buy = oldhand_buy["big_bet"]

    workbook = load_workbook(workbook_path)
    workbook["Overview"]["B3"] = str(config["excel_version"])
    write_block(workbook["Detail_Newbie"], 15, newbie_normal["weight_bg"])
    write_block(workbook["Detail_Newbie"], 86, newbie_normal["weight_fg"])
    write_block(workbook["Detail"], 15, small_normal["weight_bg"])
    write_block(workbook["Detail"], 86, small_normal["weight_fg"])
    write_block(workbook["Detail"], 163, small_buy["weight_fg"])
    write_block(workbook["Detail"], 234, big_buy["weight_fg"])

    summary = workbook["Multiplier_Weight"]
    columns = [
        range_cards(newbie_normal["weight_bg"]),
        range_cards(newbie_normal["weight_fg"]),
        range_cards(small_normal["weight_bg"]),
        range_cards(small_normal["weight_fg"]),
        range_cards(small_buy["weight_fg"]),
        range_cards(big_buy["weight_fg"]),
    ]
    for column, cards in enumerate(columns, start=2):
        for offset, card in enumerate(cards, start=4):
            summary.cell(offset, column).value = int(card["weight"])

    workbook.save(workbook_path)

    # Read-back verification against all explicitly mapped range weights.
    check = load_workbook(workbook_path, data_only=False, read_only=True)
    for column, cards in enumerate(columns, start=2):
        actual = [check["Multiplier_Weight"].cell(row, column).value for row in range(4, 68)]
        expected = [int(card["weight"]) for card in cards]
        if actual != expected:
            raise ValueError(f"Round-trip mismatch in Multiplier_Weight column {column}")
    check.close()
    print(f"Synced and verified: {workbook_path.name}")


if __name__ == "__main__":
    sync(ROOT / "config_92A.js", ROOT / "Source" / "H027192A.xlsx")
    sync(ROOT / "config_94A.js", ROOT / "Source" / "H027194A.xlsx")
