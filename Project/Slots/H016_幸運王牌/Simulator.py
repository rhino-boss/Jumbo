"""H016 幸運王牌 simulator.

The runner, environment settings, batch workflow and Excel report layout follow
the shared slot development specification.  Runtime math is loaded from Config;
direct H0161.xlsx loading remains available only as a legacy compatibility path:
5x4 Ways, cascades, golden-symbol retention, WW/W2 conversion and fixed combo
multipliers.
"""

from __future__ import annotations

import argparse
import bisect
import copy
import importlib.util
import json
import math
import os
import random
import re
import sys
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

# Loaded only after BASE_DIR has been validated as H016. Notebook sessions may
# retain another game's ``fast_simulator`` and ``__file__`` in global state.
fast_simulator = None

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)

# ===== User Settings (slot_development_specification.md) =====

CONFIG_FILE = "config.js"
CONFIG_RTP_FILE = "config_92A.js"
TOTAL_ROUNDS = 100_000
BET_MULTI = 1
BET_MODE = 0  # 0=Normal Bet, 2=Buy Feature, 3=Buy Super Feature
CARD_SYSTEM_ENABLED = True
CARD_SYSTEM_IS_NEWBIE = False

RUN_ALL_COMBINATIONS = True
BATCH_RUNS = [
    # Test
    # {"config_file": "config.js", "config_rtp_file": "config_92A.js", "bet_mode": 0, "total_rounds": 10**7+2, "card_system_enabled": True, "card_system_is_newbie": False},   # Test
    # {"config_file": "config.js", "config_rtp_file": "config_92A.js", "bet_mode": 0, "total_rounds": 10**7, "card_system_enabled": True, "card_system_is_newbie": True},   # Test
    {"config_file": "config.js", "config_rtp_file": "config_92A.js", "bet_mode": 3, "total_rounds": 10**5, "card_system_enabled": False, "card_system_is_newbie": False},  # Test
    # {"config_file": "config.js", "config_rtp_file": "config_92A.js", "bet_mode": 0, "total_rounds": 10**6, "card_system_enabled": False, "card_system_is_newbie": False},  # Test
    # ## 自然機率
    # {"config_file": "config.js", "config_rtp_file": "config_92A.js", "bet_mode": 0, "total_rounds": 10**9, "card_system_enabled": False, "card_system_is_newbie": False},  # 自然機率
    # {"config_file": "config.js", "config_rtp_file": "config_92A.js", "bet_mode": 3, "total_rounds": 10**8, "card_system_enabled": False, "card_system_is_newbie": False},  # 自然機率
    # {"config_file": "config.js", "config_rtp_file": "config_92A.js", "bet_mode": 3, "total_rounds": 10**9, "card_system_enabled": False, "card_system_is_newbie": False},  # 自然機率
    ## SCR
    # {"config_file": "config.js", "config_rtp_file": "config_92A.js", "bet_mode": 0, "total_rounds": 10**8, "card_system_enabled": True, "card_system_is_newbie": True},  # SCR
    # {"config_file": "config.js", "config_rtp_file": "config_92A.js", "bet_mode": 0, "total_rounds": 10**8, "card_system_enabled": True, "card_system_is_newbie": False},  # SCR
    # {"config_file": "config.js", "config_rtp_file": "config_94A.js", "bet_mode": 0, "total_rounds": 10**8, "card_system_enabled": True, "card_system_is_newbie": True},  # SCR
    # {"config_file": "config.js", "config_rtp_file": "config_94A.js", "bet_mode": 0, "total_rounds": 10**8, "card_system_enabled": True, "card_system_is_newbie": False},  # SCR
    # {"config_file": "config.js", "config_rtp_file": "config_92A.js", "bet_mode": 2, "total_rounds": 10**7, "card_system_enabled": True, "card_system_is_newbie": False},  # SCR
    # {"config_file": "config.js", "config_rtp_file": "config_92A.js", "bet_mode": 3, "total_rounds": 10**7, "card_system_enabled": True, "card_system_is_newbie": False},  # SCR
]
THREADS = max(1, min(8, os.cpu_count() or 1))
OUTPUT_REPORT = True
FAST_SIMULATION = True
SHOW_CONSOLE_SUMMARY = True
SHOW_CONSOLE_DETAIL = False
RUN_SINGLE_SPIN_DEBUG = False
DEBUG_ROUNDS = 1
CARD_RETRY_LIMIT = 10_000
RNG_SEED = 46_046

BASE_BET = 100.0
MODE_NORMALBET = 0
MODE_FEATUREBUY = 2
MODE_SUPERBUY = 3
SUPPORTED_BET_MODES = (MODE_NORMALBET, MODE_FEATUREBUY, MODE_SUPERBUY)
WW = 0
W2 = 1
C1 = 2
SCORE_SYMBOLS = tuple(range(3, 11))
GOLD_MIN = 11
GOLD_MAX = 18
MULTIPLIER_THRESHOLDS = (
    0,
    1,
    2,
    3,
    4,
    5,
    6,
    7,
    8,
    9,
    10,
    15,
    20,
    25,
    30,
    35,
    40,
    45,
    50,
    60,
    70,
    80,
    90,
    100,
    120,
    140,
    160,
    180,
    200,
    250,
    300,
    350,
    400,
    450,
    500,
    550,
    600,
    650,
    700,
    750,
    800,
    850,
    900,
    950,
    1000,
    2000,
    3000,
    4000,
    5000,
    6000,
    7000,
    8000,
    9000,
    10000,
    20000,
    30000,
    40000,
    50000,
    60000,
    70000,
    80000,
    90000,
    100000,
    9999999,
)


def _env_bool(name: str, default: bool) -> bool:
    value = os.environ.get(name)
    if value is None:
        return default
    value = value.strip().lower()
    if value in {"1", "true", "yes", "on"}:
        return True
    if value in {"0", "false", "no", "off"}:
        return False
    raise ValueError(f"{name} must be true/false, got {value!r}")


CONFIG_FILE = os.environ.get("H016_CONFIG_FILE", CONFIG_FILE)
CONFIG_RTP_FILE = os.environ.get("H016_CONFIG_RTP_FILE", CONFIG_RTP_FILE)
# A no-argument batch run will immediately use its first combo.  Select that
# Config before base-dir discovery so startup immediately loads the same
# natural/RTP pair that the selected batch will use. Explicit positional runs
# use the CONFIG_FILE / CONFIG_RTP_FILE defaults above.
_explicit_cli_run = len(sys.argv) > 1 and str(sys.argv[1]).lstrip("+-").isdigit()
if "H016_CONFIG_FILE" not in os.environ and _env_bool("H016_RUN_ALL_COMBINATIONS", RUN_ALL_COMBINATIONS) and BATCH_RUNS and not _explicit_cli_run and os.environ.get("H016_BATCH_CHILD") != "1":
    CONFIG_FILE = str(BATCH_RUNS[0].get("config_file", CONFIG_FILE))
    CONFIG_RTP_FILE = str(BATCH_RUNS[0].get("config_rtp_file", CONFIG_RTP_FILE))


def _load_config(path: Path) -> dict[str, Any]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _load_xlsx_config(path)
    raw = path.read_text(encoding="utf-8-sig")
    start, end = raw.find("{"), raw.rfind("}")
    if start < 0 or end <= start:
        raise ValueError(f"Config does not contain a JSON object: {path}")
    return json.loads(raw[start : end + 1])


def _version_major(value: Any) -> str:
    return str(value or "").strip().split(".", 1)[0]


def validate_config_pair(
    natural: dict[str, Any],
    rtp: dict[str, Any],
    natural_name: str,
    rtp_name: str,
) -> None:
    """Validate the natural-probability and RTP/Card configs before a batch."""
    natural_game_id = str(natural.get("game_id", ""))
    rtp_game_id = str(rtp.get("game_id", ""))
    if natural_game_id != rtp_game_id:
        raise ValueError(f"Config game_id mismatch: {natural_name}={natural_game_id!r}, " f"{rtp_name}={rtp_game_id!r}")
    if natural_game_id != "H016":
        raise ValueError(f"Unsupported game_id {natural_game_id!r}; expected 'H016'")

    natural_version = str(natural.get("excel_version", ""))
    rtp_version = str(rtp.get("excel_version", ""))
    if _version_major(natural_version) != _version_major(rtp_version):
        raise ValueError(f"Config version mismatch: {natural_name}={natural_version!r}, " f"{rtp_name}={rtp_version!r}")

    # RTP/Card config is not allowed to silently replace the natural board.
    natural_symbol_ids = set(map(str, natural.get("symbol_names", {})))
    rtp_symbol_ids = set(map(str, rtp.get("symbol_names", {})))
    if natural_symbol_ids and rtp_symbol_ids and natural_symbol_ids != rtp_symbol_ids:
        raise ValueError(f"Incompatible symbol IDs between {natural_name} and {rtp_name}")

    immutable_keys = (
        "reel_num",
        "window_size",
        "pays",
        "tables",
        "table_selection",
        "free_game_mix",
        "free_spins",
        "retrigger_spins",
        "free_spin_cap",
        "buy_price",
        "super_buy_price",
    )
    for key in immutable_keys:
        if key in natural and key in rtp and natural[key] != rtp[key]:
            raise ValueError(f"Incompatible Config field {key!r}: {rtp_name} must not " f"change the natural-probability data from {natural_name}")


def compose_runtime_config(natural: dict[str, Any], rtp: dict[str, Any]) -> dict[str, Any]:
    """Use natural game data plus only RTP/Card metadata from the RTP config."""
    merged = copy.deepcopy(natural)
    merged["card_system"] = copy.deepcopy(rtp.get("card_system") or {"enabled": False, "profiles": {}})
    for key in (
        "parsheet_id",
        "excel_version",
        "rtp_label",
        "runtime_version",
        "source_multiplier_xlsx",
    ):
        if key in rtp:
            merged[key] = copy.deepcopy(rtp[key])
    return merged


def _xlsx_number(value: Any, label: str) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{label} must be numeric, got {value!r}")
    return float(value)


def _xlsx_table(ws, name_to_id: dict[str, int], multipliers: list[int], reel_source=None, drop_label_source=None, drop_weight_source=None) -> dict[str, Any]:
    reel_source = reel_source or ws
    drop_label_source = drop_label_source or ws
    drop_weight_source = drop_weight_source or ws
    reels: list[list[int]] = [[] for _ in range(5)]
    weights: list[list[float]] = [[] for _ in range(5)]
    for row_number in range(4, 404):
        for reel in range(5):
            symbol_name = reel_source.cell(row_number, 11 + reel).value
            if symbol_name in (None, ""):
                continue
            if symbol_name not in name_to_id:
                raise ValueError(f"{ws.title} R{reel + 1} row {row_number}: unknown symbol {symbol_name!r}")
            reels[reel].append(name_to_id[str(symbol_name)])
            weights[reel].append(_xlsx_number(ws.cell(row_number, 23 + reel).value, f"{ws.title} R{reel + 1} weight row {row_number}"))
    if any(not reel for reel in reels):
        raise ValueError(f"{ws.title}: five non-empty reels are required")
    if any(sum(reel_weights) <= 0 for reel_weights in weights):
        raise ValueError(f"{ws.title}: each reel must have positive total weight")

    drop_values: list[list[int]] = [[] for _ in range(5)]
    drop_weights: list[list[float]] = [[] for _ in range(5)]
    for row_number in range(4, 23):
        # Variant AF labels are linked Excel array formulas.  Use the matching
        # scene's canonical label cells while keeping every variant's own
        # AF:AK numeric Symbol Drop Weight values.
        symbol_name = drop_label_source.cell(row_number, 32).value
        if symbol_name in (None, ""):
            continue
        if symbol_name not in name_to_id:
            raise ValueError(f"{ws.title} AF{row_number}: unknown drop symbol {symbol_name!r}")
        symbol_id = name_to_id[str(symbol_name)]
        for reel in range(5):
            weight = _xlsx_number(drop_weight_source.cell(row_number, 33 + reel).value, f"{ws.title} drop R{reel + 1} row {row_number}")
            if weight < 0:
                raise ValueError(f"{ws.title}: drop weights cannot be negative")
            drop_values[reel].append(symbol_id)
            drop_weights[reel].append(weight)
    if any(not values or sum(reel_weights) <= 0 for values, reel_weights in zip(drop_values, drop_weights)):
        raise ValueError(f"{ws.title}: Symbol Drop Weight must define five positive-total reels")

    random_values = [int(_xlsx_number(ws.cell(row, 29).value, f"{ws.title} Random Wild value")) for row in range(4, 8)]
    random_weights = [_xlsx_number(ws.cell(row, 30).value, f"{ws.title} Random Wild weight") for row in range(4, 8)]
    if random_values != [0, 2, 3, 4] or len(random_values) != len(random_weights):
        raise ValueError(f"{ws.title}: Random Wild must define 0/2/3/4 and matching weights")

    return {
        "reels": reels,
        "weights": weights,
        "drop_values": drop_values,
        "drop_weights": drop_weights,
        "random_wild": {"values": random_values, "weights": random_weights},
        "multipliers": multipliers,
    }


def _load_xlsx_config(path: Path) -> dict[str, Any]:
    # Normal mode is materially faster for repeatedly reading the six 200-stop
    # physical reel tables by coordinate.
    workbook = load_workbook(path, read_only=False, data_only=False)
    required_sheets = {
        "Overview",
        "Parameter",
        "BG_Symbol",
        "FG_Symbol",
        "BF_Symbol",
        "SF_Symbol",
        "SF_Symbol (2)",
        "SF_Symbol (3)",
    }
    missing = required_sheets.difference(workbook.sheetnames)
    if missing:
        raise ValueError(f"{path.name}: missing sheets {sorted(missing)}")
    overview = workbook["Overview"]
    parameter = workbook["Parameter"]

    base_bet = _xlsx_number(overview["A7"].value, "Overview!A7 Base Bet")
    symbol_names: dict[str, str] = {}
    name_to_id: dict[str, int] = {}
    for row in range(30, 49):
        name = overview.cell(row, 1).value
        symbol_id = overview.cell(row, 8).value
        if name in (None, "") or symbol_id in (None, ""):
            continue
        numeric_id = int(_xlsx_number(symbol_id, f"Overview!H{row}"))
        symbol_names[str(numeric_id)] = str(name)
        name_to_id[str(name)] = numeric_id
    expected_ids = set(range(19))
    if set(map(int, symbol_names)) != expected_ids:
        raise ValueError(f"{path.name}: Overview symbol ids must be 0..18")

    pays: dict[str, list[float]] = {}
    for row in range(33, 41):
        symbol_id = int(_xlsx_number(overview.cell(row, 8).value, f"Overview!H{row}"))
        pays[str(symbol_id)] = [_xlsx_number(overview.cell(row, col).value, f"Overview pay row {row}") / base_bet for col in (5, 6, 7)]

    bg_multipliers = [1, 2, 3, 5]
    fg_multipliers = [2, 4, 6, 10]
    bg_names = ["BG_Symbol", "BG_Symbol (2)", "BG_Symbol (3)"]
    fg_names = ["FG_Symbol", "FG_Symbol (2)", "FG_Symbol (3)"]
    sf_names = ["SF_Symbol", "SF_Symbol (2)", "SF_Symbol (3)"]
    missing_variants = set(bg_names + fg_names + sf_names).difference(workbook.sheetnames)
    if missing_variants:
        raise ValueError(f"{path.name}: missing sheets {sorted(missing_variants)}")
    bg_tables = [_xlsx_table(workbook[name], name_to_id, bg_multipliers, drop_label_source=workbook["BG_Symbol"]) for name in bg_names]
    fg_tables = [_xlsx_table(workbook[name], name_to_id, fg_multipliers, drop_label_source=workbook["FG_Symbol"]) for name in fg_names]
    sf_tables = [_xlsx_table(workbook[name], name_to_id, fg_multipliers, drop_label_source=workbook["SF_Symbol"]) for name in sf_names]
    # BF is an entry-only board and never cascades, so its AF labels are unused.
    # They are linked array formulas; use the canonical BG labels only to keep a
    # structurally complete runtime table for the shared config schema.
    bf_table = _xlsx_table(
        workbook["BF_Symbol"],
        name_to_id,
        bg_multipliers,
        drop_label_source=workbook["BG_Symbol"],
        drop_weight_source=workbook["BG_Symbol"],
    )
    tables = {
        "bg_1": bg_tables[0],
        "bg_2": bg_tables[1],
        "bg_3": bg_tables[2],
        "fg_1": fg_tables[0],
        "fg_2": fg_tables[1],
        "fg_3": fg_tables[2],
        "sf_1": sf_tables[0],
        "sf_2": sf_tables[1],
        "sf_3": sf_tables[2],
        "bg_high": bg_tables[0],
        "bg_low": bg_tables[1],
        "buy": bf_table,
        "fg_high_a": fg_tables[0],
        "fg_high_k": fg_tables[1],
        "fg_high_q": fg_tables[2],
        "fg_high_j": fg_tables[0],
        "fg_low": fg_tables[0],
        "super": sf_tables[0],
    }

    def table_selection(rows: range, prefix: str) -> list[dict[str, Any]]:
        result = [{"table": f"{prefix}_{index}", "weight": _xlsx_number(parameter.cell(row, 3).value, f"Parameter!C{row}")} for index, row in enumerate(rows, start=1)]
        if sum(float(item["weight"]) for item in result) <= 0:
            raise ValueError(f"Parameter {prefix} table selection must have positive total weight")
        return result

    return {
        "game_id": "H016",
        "parsheet_id": str(overview["B2"].value or path.stem),
        "excel_version": str(overview["B3"].value or "1.0.0.0"),
        "name_zh": "幸運王牌",
        "rtp_label": None,
        "reel_num": 5,
        "window_size": 4,
        "max_ways": 1024,
        "symbol_names": symbol_names,
        "pays": pays,
        "tables": tables,
        "table_selection": {
            "base": table_selection(range(4, 7), "bg"),
            "free": table_selection(range(11, 14), "fg"),
            "retrigger": table_selection(range(18, 21), "fg"),
            "super_free": table_selection(range(25, 28), "sf"),
            "super_retrigger": table_selection(range(32, 35), "sf"),
        },
        "free_game_mix": {
            "choices": [{"high": int(_xlsx_number(overview["B22"].value, "Overview!B22")), "low": 0, "weight": 1}],
            "high_variant_weights": [1, 0, 0, 0],
        },
        "free_spins": int(_xlsx_number(overview["B22"].value, "Overview!B22")),
        "retrigger_spins": int(_xlsx_number(overview["C22"].value, "Overview!C22")),
        "free_spin_cap": 50,
        "buy_price": _xlsx_number(overview["B12"].value, "Overview!B12"),
        "super_buy_price": _xlsx_number(overview["B13"].value, "Overview!B13"),
        "card_system": {"enabled": False, "profiles": {}},
        "source_xlsx": path.name,
    }


_CONFIG_DISCOVERY_CACHE: dict[Path, dict[str, Any]] = {}


def _is_h016_config(path: Path) -> bool:
    if not path.is_file():
        return False
    try:
        data = _load_config(path)
    except (OSError, ValueError, json.JSONDecodeError):
        return False
    required = {"tables", "symbol_names", "pays", "free_game_mix"}
    valid = str(data.get("game_id")) == "H016" and required.issubset(data)
    if valid:
        _CONFIG_DISCOVERY_CACHE[path.resolve()] = data
    return valid


def _load_runtime_config(path: Path) -> dict[str, Any]:
    """Reuse the config parsed while locating the project directory."""
    resolved = path.resolve()
    cached = _CONFIG_DISCOVERY_CACHE.pop(resolved, None)
    return cached if cached is not None else _load_config(resolved)


def resolve_base_dir() -> Path:
    """Locate H016 safely when run as a file, `%run`, or a Notebook cell."""
    cwd = Path.cwd().resolve()
    candidates: list[Path] = []
    override = os.environ.get("H016_BASE_DIR")
    if override:
        candidates.append(Path(override).expanduser())

    file_value = globals().get("__file__")
    if file_value and not str(file_value).startswith("<"):
        candidates.append(Path(file_value).resolve().parent)
    candidates.append(cwd)

    for parent in (cwd, *cwd.parents):
        candidates.extend(
            [
                parent / "Project" / "Slots" / "H016_幸運王牌",
                parent / "Project_AI" / "Slots" / "H016_幸運王牌",
                parent / "Slots" / "H016_幸運王牌",
            ]
        )

    checked: list[Path] = []
    for candidate in candidates:
        try:
            candidate = candidate.resolve()
        except OSError:
            continue
        if candidate in checked:
            continue
        checked.append(candidate)
        if _is_h016_config(candidate / CONFIG_FILE):
            return candidate

    locations = "\n  - ".join(str(path / CONFIG_FILE) for path in checked)
    raise FileNotFoundError(f"Cannot locate a valid H016 {CONFIG_FILE}. Checked:\n  - {locations}\n" "Set H016_BASE_DIR to the H016_幸運王牌 folder when running from another workspace.")


BASE_DIR = resolve_base_dir()


def load_fast_simulator(base_dir: Path):
    """Load H016's exact fast core, ignoring stale Notebook modules."""
    module_name = "fast_simulator"
    module_path = (base_dir / "其他" / "fast_simulator.py").resolve()
    if not module_path.is_file():
        return None
    source_stat = module_path.stat()
    source_signature = (source_stat.st_mtime_ns, source_stat.st_size)

    existing = sys.modules.get(module_name)
    existing_path = getattr(existing, "__file__", None)
    try:
        same_module = bool(existing_path) and Path(existing_path).resolve() == module_path
    except OSError:
        same_module = False
    if same_module and getattr(existing, "FAST_SIMULATOR_API_VERSION", 1) >= 9 and getattr(existing, "__h016_source_signature__", None) == source_signature:
        return existing

    # Keep one canonical module name. Numba persists this name inside its disk
    # cache; a Notebook-only alias makes the next batch child unable to unpickle
    # the compiled function environment.
    spec = importlib.util.spec_from_file_location(module_name, module_path)
    if spec is None or spec.loader is None:
        return None
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    module.__h016_source_signature__ = source_signature
    return module


fast_simulator = load_fast_simulator(BASE_DIR)
OUTPUT_DIR = BASE_DIR / "Record"
SIMULATOR_PATH = BASE_DIR / "Simulator.py"
TOTAL_ROUNDS = int(os.environ.get("H016_TOTAL_ROUNDS", TOTAL_ROUNDS))
BET_MULTI = int(os.environ.get("H016_BET_MULTI", BET_MULTI))
BET_MODE = int(os.environ.get("H016_BET_MODE", BET_MODE))
CARD_SYSTEM_ENABLED = _env_bool("H016_CARD_SYSTEM_ENABLED", CARD_SYSTEM_ENABLED)
CARD_SYSTEM_IS_NEWBIE = _env_bool("H016_CARD_SYSTEM_IS_NEWBIE", CARD_SYSTEM_IS_NEWBIE)
RUN_ALL_COMBINATIONS = _env_bool("H016_RUN_ALL_COMBINATIONS", RUN_ALL_COMBINATIONS)
OUTPUT_REPORT = _env_bool("H016_OUTPUT_REPORT", OUTPUT_REPORT)
FAST_SIMULATION = _env_bool("H016_FAST_SIMULATION", FAST_SIMULATION)
SHOW_CONSOLE_SUMMARY = _env_bool("H016_SHOW_CONSOLE_SUMMARY", SHOW_CONSOLE_SUMMARY)
SHOW_CONSOLE_DETAIL = _env_bool("H016_SHOW_CONSOLE_DETAIL", SHOW_CONSOLE_DETAIL)
RUN_SINGLE_SPIN_DEBUG = _env_bool("H016_RUN_SINGLE_SPIN_DEBUG", RUN_SINGLE_SPIN_DEBUG)
THREADS = max(1, int(os.environ.get("H016_THREADS", THREADS)))
RNG_SEED = int(os.environ.get("H016_RNG_SEED", RNG_SEED))
CFG_NATURAL = _load_runtime_config(BASE_DIR / CONFIG_FILE)
CFG_RTP = CFG_NATURAL if Path(CONFIG_RTP_FILE) == Path(CONFIG_FILE) else _load_config(BASE_DIR / CONFIG_RTP_FILE)
validate_config_pair(CFG_NATURAL, CFG_RTP, CONFIG_FILE, CONFIG_RTP_FILE)
CFG = compose_runtime_config(CFG_NATURAL, CFG_RTP)
CARD_SYSTEM_ENABLED = CARD_SYSTEM_ENABLED and bool(CFG.get("card_system", {}).get("enabled"))
CARD_RETRY_LIMIT = max(1, int(CFG.get("card_system", {}).get("retry_limit", CARD_RETRY_LIMIT)))
GAME_ID = str(CFG.get("game_id", "H016"))
PARSHEET_ID = str(CFG.get("parsheet_id", "H016192"))
GAME_NAME = str(CFG.get("name_zh", "幸運王牌"))
MATH_VERSION = str(CFG_RTP.get("excel_version", CFG.get("excel_version", "")))
SYMBOL_STR = {int(k): v for k, v in CFG["symbol_names"].items()}


def canonical(symbol: int) -> int:
    return symbol - 8 if GOLD_MIN <= symbol <= GOLD_MAX else symbol


def weighted_pick(rng: random.Random, values: list[Any], weights: list[float]) -> Any:
    total = sum(max(0.0, float(w)) for w in weights)
    if not values:
        raise ValueError("No values available for weighted selection")
    if total <= 0:
        return values[0]
    target = rng.random() * total
    for value, weight in zip(values, weights):
        target -= max(0.0, float(weight))
        if target < 0:
            return value
    return values[-1]


def weighted_pick_index(rng: random.Random, weights: list[float]) -> int:
    if not weights:
        raise ValueError("No weights available for weighted selection")
    total = sum(max(0.0, float(weight)) for weight in weights)
    if total <= 0:
        return 0
    target = rng.random() * total
    for index, weight in enumerate(weights):
        target -= max(0.0, float(weight))
        if target < 0:
            return index
    return len(weights) - 1


@dataclass
class Reel:
    symbols: list[int]
    stop_cumulative: list[float]
    stop_total: float

    @staticmethod
    def _index(rng: random.Random, cumulative: list[float], total: float) -> int:
        position = rng.random() * total
        return min(bisect.bisect_right(cumulative, position), len(cumulative) - 1)

    def window(self, rng: random.Random, size: int) -> list[int]:
        stop = self._index(rng, self.stop_cumulative, self.stop_total)
        return [self.symbols[(stop + offset) % len(self.symbols)] for offset in range(size)]

    def pick(self, rng: random.Random) -> int:
        return self.symbols[self._index(rng, self.stop_cumulative, self.stop_total)]


@dataclass
class Table:
    reels: list[Reel]
    drop_values: list[list[int]]
    drop_weights: list[list[float]]
    random_wild_values: list[int]
    random_wild_weights: list[float]
    multipliers: list[int]


@dataclass
class SpinResult:
    pay: float = 0.0
    scatter_count: int = 0
    cascades: int = 0
    max_multiplier: int = 1
    golden_converted: int = 0
    w2_events: int = 0
    w2_counts: Counter = field(default_factory=Counter)
    m1_present: bool = False
    initial_gold_count: int = 0
    symbol_hits: Counter = field(default_factory=Counter)
    symbol_pay: Counter = field(default_factory=Counter)
    symbol_length_hits: Counter = field(default_factory=Counter)
    symbol_length_pay: Counter = field(default_factory=Counter)
    initial_symbols: Counter = field(default_factory=Counter)
    drop_symbols: Counter = field(default_factory=Counter)
    initial_board: list[list[int]] = field(default_factory=list)
    final_board: list[list[int]] = field(default_factory=list)


@dataclass
class RoundResult:
    pay_bg: float = 0.0
    pay_fg: float = 0.0
    fg_triggered: bool = False
    fg_spins: int = 0
    retriggers: int = 0
    cascades_bg: int = 0
    cascades_fg: int = 0
    max_multiplier: int = 1
    golden_converted: int = 0
    w2_events: int = 0
    bg_w2_events: int = 0
    fg_w2_events: int = 0
    bg_w2_counts: Counter = field(default_factory=Counter)
    fg_w2_counts: Counter = field(default_factory=Counter)
    bg_m1_spins: int = 0
    fg_m1_spins: int = 0
    bg_hit_spins: int = 0
    fg_hit_spins: int = 0
    bg_gold_spins: int = 0
    fg_gold_spins: int = 0
    bg_gold_symbols: int = 0
    fg_gold_symbols: int = 0
    special_symbol_cnt: int = 0
    bg_trigger_fg_cnt: int = 0
    bg_trigger_fg_pay: float = 0.0
    combo_fg: Counter = field(default_factory=Counter)
    symbol_hits: Counter = field(default_factory=Counter)
    symbol_pay: Counter = field(default_factory=Counter)
    bg_symbol_length_hits: Counter = field(default_factory=Counter)
    bg_symbol_length_pay: Counter = field(default_factory=Counter)
    fg_symbol_length_hits: Counter = field(default_factory=Counter)
    fg_symbol_length_pay: Counter = field(default_factory=Counter)
    bg_initial_symbols: Counter = field(default_factory=Counter)
    bg_drop_symbols: Counter = field(default_factory=Counter)
    fg_initial_symbols: Counter = field(default_factory=Counter)
    fg_drop_symbols: Counter = field(default_factory=Counter)

    @property
    def pay(self) -> float:
        return self.pay_bg + self.pay_fg


class LuckyAce:
    def __init__(self, config: dict[str, Any], seed: int, card_enabled: bool, newbie: bool):
        self.config = config
        self.rng = random.Random(seed)
        self.card_enabled = card_enabled and bool(config.get("card_system", {}).get("enabled"))
        self.profile = "weight_1" if newbie else "weight_2"
        self.bg_trigger_cap = self.profile_bg_trigger_cap(self.profile)
        self.tables = {name: self._prepare(raw) for name, raw in config["tables"].items()}
        self.retry_total = 0
        self.retry_limit_exceeded = 0
        self.retry_limit_bg_range = 0
        self.retry_limit_bg_freegame = 0
        self.retry_limit_fg = 0
        self.card_draws: Counter = Counter()

    def profile_bg_trigger_cap(self, profile: str | None = None) -> float:
        """Maximum enabled BG range-card interval for the selected profile."""
        profile = profile or self.profile
        card_system = self.config.get("card_system") or {}
        profiles = card_system.get("profiles") or {}
        profile_config = profiles.get(profile) or profiles.get(str(card_system.get("default_profile", "weight_2"))) or {}
        maxima = [float(card["max"]) for card in profile_config.get("base_game") or [] if str(card.get("type")) == "range" and float(card.get("weight", 0)) > 0 and "max" in card]
        return max(maxima) if maxima else math.inf

    def bg_trigger_within_cap(self, pay: float) -> bool:
        return pay / (BASE_BET * BET_MULTI) <= self.bg_trigger_cap

    @staticmethod
    def _prepare(raw: dict[str, Any]) -> Table:
        reels = []
        for symbols, weights in zip(raw["reels"], raw["weights"]):
            stop_cumulative, running = [], 0.0
            for weight in weights:
                running += max(0.0, float(weight))
                stop_cumulative.append(running)
            reels.append(Reel(list(map(int, symbols)), stop_cumulative, running))
        random_wild = raw["random_wild"]
        return Table(
            reels,
            [list(map(int, values)) for values in raw["drop_values"]],
            [list(map(float, weights)) for weights in raw["drop_weights"]],
            list(map(int, random_wild["values"])),
            list(map(float, random_wild["weights"])),
            list(map(int, raw.get("multipliers") or [])),
        )

    def board(self, table_name: str) -> list[list[int]]:
        table = self.tables[table_name]
        return [table.reels[reel].window(self.rng, 4) for reel in range(5)]

    def evaluate(self, board: list[list[int]]) -> tuple[float, set[tuple[int, int]], list[tuple[int, int, int, float]]]:
        total, hits, details = 0.0, set(), []
        for target in SCORE_SYMBOLS:
            counts, positions = [], []
            for reel in range(5):
                matched = [(reel, row) for row, symbol in enumerate(board[reel]) if symbol in (WW, W2) or canonical(symbol) == target]
                if not matched:
                    break
                counts.append(len(matched))
                positions.append(matched)
            length = len(counts)
            if length < 3:
                continue
            ways = math.prod(counts)
            raw_pay = float(self.config["pays"][str(target)][length - 3]) * ways
            if raw_pay <= 0:
                continue
            total += raw_pay
            for group in positions:
                hits.update(group)
            details.append((target, length, ways, raw_pay))
        return total, hits, details

    def add_w2(
        self,
        board: list[list[int]],
        table: Table,
        gold: list[tuple[int, int]],
        super_mode: bool = False,
    ) -> int:
        if not gold:
            return 0
        count = int(weighted_pick(self.rng, table.random_wild_values, table.random_wild_weights))
        if count <= 0:
            return 0
        source = self.rng.choice(gold)
        candidates = [(reel, row) for reel in range(1, 5) for row, symbol in enumerate(board[reel]) if symbol not in (WW, W2, C1)]
        # JHS101003 Super Buy first excludes every golden symbol from the
        # additional Random Wild candidates. It falls back to the full list
        # only when fewer than four non-golden positions remain.
        if super_mode:
            non_gold = [position for position in candidates if not GOLD_MIN <= board[position[0]][position[1]] <= GOLD_MAX]
            if len(non_gold) >= 4:
                candidates = non_gold
        if len(candidates) < count:
            return 0
        board[source[0]][source[1]] = W2
        self.rng.shuffle(candidates)
        for reel, row in candidates[:count]:
            board[reel][row] = W2
        return count

    def spin(
        self,
        table_name: str,
        free_game: bool = False,
        super_mode: bool = False,
    ) -> SpinResult:
        table = self.tables[table_name]
        multipliers = table.multipliers or ([2, 4, 6, 10] if free_game else [1, 2, 3, 5])
        board = self.board(table_name)
        result = SpinResult(initial_board=[reel[:] for reel in board])
        result.initial_symbols.update((reel, symbol) for reel, symbols in enumerate(board) for symbol in symbols)
        result.m1_present = any(canonical(symbol) == 3 for symbols in board for symbol in symbols)
        result.initial_gold_count = sum(GOLD_MIN <= symbol <= GOLD_MAX for symbols in board for symbol in symbols)
        pending_gold: list[tuple[int, int]] = []
        w2_used = False
        while True:
            # 101003 behavior: BG locks after the first successful Random Wild
            # event; FG evaluates every cascade that converted new gold symbols.
            if pending_gold and (free_game or not w2_used):
                made = self.add_w2(board, table, pending_gold, super_mode)
                if made:
                    result.w2_events += 1
                    result.w2_counts[made] += 1
                    w2_used = True
            pending_gold = []
            raw_pay, hit_positions, details = self.evaluate(board)
            if raw_pay <= 0:
                break
            multiplier = multipliers[min(result.cascades, len(multipliers) - 1)]
            result.pay += raw_pay * multiplier * BASE_BET * BET_MULTI
            result.max_multiplier = max(result.max_multiplier, multiplier)
            result.cascades += 1
            for symbol, length, ways, symbol_raw_pay in details:
                result.symbol_hits[symbol] += ways
                result.symbol_pay[symbol] += symbol_raw_pay * multiplier * BASE_BET * BET_MULTI
                result.symbol_length_hits[(symbol, length)] += 1
                result.symbol_length_pay[(symbol, length)] += symbol_raw_pay * multiplier * BASE_BET * BET_MULTI
            for reel, row in hit_positions:
                symbol = board[reel][row]
                if GOLD_MIN <= symbol <= GOLD_MAX:
                    board[reel][row] = WW
                    pending_gold.append((reel, row))
                    result.golden_converted += 1
                else:
                    board[reel][row] = -1
            for reel in range(5):
                for row in range(4):
                    if board[reel][row] != -1:
                        continue
                    symbol = int(weighted_pick(self.rng, table.drop_values[reel], table.drop_weights[reel]))
                    board[reel][row] = symbol
                    result.drop_symbols[(reel, symbol)] += 1
                    result.m1_present |= canonical(symbol) == 3
        result.scatter_count = sum(symbol == C1 for reel in board for symbol in reel)
        result.final_board = [reel[:] for reel in board]
        return result

    def pick_card(self, section: str, profile: str | None = None) -> dict[str, Any]:
        profile = profile or self.profile
        cards = self.config["card_system"]["profiles"][profile][section]
        card_index = weighted_pick_index(self.rng, [float(card["weight"]) for card in cards])
        self.card_draws[(profile, section, card_index)] += 1
        return dict(cards[card_index])

    @staticmethod
    def card_matches(card: dict[str, Any], pay: float) -> bool:
        ratio = pay / (BASE_BET * BET_MULTI)
        return float(card["min"]) < ratio <= float(card["max"])

    def natural_base_spin(self) -> SpinResult:
        """Draw one complete Normal Bet BG from the natural table selection."""
        selections = self.config.get("table_selection", {}).get("base")
        if selections:
            table_name = str(
                weighted_pick(
                    self.rng,
                    [str(item["table"]) for item in selections],
                    [float(item["weight"]) for item in selections],
                )
            )
        else:
            table_name = "bg_high"
        return self.spin(table_name)

    def card_spin(self, card: dict[str, Any]) -> SpinResult:
        if card.get("type") == "free_game":
            for _ in range(CARD_RETRY_LIMIT):
                # A Normal Bet FG card must keep rerolling the complete natural
                # BG selection.  BF_Symbol is reserved for paid Buy entry and
                # would otherwise discard the triggering round's BG award.
                spin = self.natural_base_spin()
                if spin.scatter_count >= 3 and self.bg_trigger_within_cap(spin.pay):
                    return spin
                self.retry_total += 1
            self.retry_limit_exceeded += 1
            self.retry_limit_bg_freegame += 1
            return spin
        for _ in range(CARD_RETRY_LIMIT):
            # Card System only filters a completed natural result.  Every retry
            # must redraw the complete BG table selection before checking the
            # fixed card; card.table is report metadata, not a table override.
            spin = self.natural_base_spin()
            if spin.scatter_count < 3 and self.card_matches(card, spin.pay):
                return spin
            self.retry_total += 1
        self.retry_limit_exceeded += 1
        self.retry_limit_bg_range += 1
        return spin

    def free_queue(self, super_mode: bool = False) -> list[str]:
        group = "super_free" if super_mode else "free"
        selections = self.config.get("table_selection", {}).get(group)
        if selections:
            values = [str(item["table"]) for item in selections]
            weights = [float(item["weight"]) for item in selections]
            return [str(weighted_pick(self.rng, values, weights)) for _ in range(int(self.config["free_spins"]))]
        mix = self.config["free_game_mix"]
        choices = mix.get("choices") or mix.get("groups", {}).get("E")
        if not choices:
            raise ValueError("free_game_mix must provide choices or group E")
        choice = weighted_pick(self.rng, choices, [float(item["weight"]) for item in choices])
        queue = ["high"] * int(choice["high"]) + ["low"] * int(choice["low"])
        self.rng.shuffle(queue)
        return queue

    def high_table(self) -> str:
        return str(weighted_pick(self.rng, ["fg_high_a", "fg_high_k", "fg_high_q", "fg_high_j"], self.config["free_game_mix"]["high_variant_weights"]))

    def free_session(self, super_mode: bool = False) -> RoundResult:
        result = RoundResult(fg_triggered=True)
        remaining, played = int(self.config["free_spins"]), 0
        queue = self.free_queue(super_mode)
        while remaining > 0 and played < int(self.config["free_spin_cap"]):
            remaining -= 1
            played += 1
            surface = queue.pop(0) if queue else "low"
            table_name = surface if surface in self.tables else self.high_table() if surface == "high" else "fg_low"
            spin = self.spin(table_name, free_game=True, super_mode=super_mode)
            result.pay_fg += spin.pay
            result.fg_spins += 1
            result.cascades_fg += spin.cascades
            result.max_multiplier = max(result.max_multiplier, spin.max_multiplier)
            result.golden_converted += spin.golden_converted
            result.w2_events += spin.w2_events
            result.fg_w2_events += spin.w2_events
            result.fg_w2_counts.update(spin.w2_counts)
            result.fg_m1_spins += int(spin.m1_present)
            result.fg_hit_spins += int(spin.pay > 0)
            result.fg_gold_spins += int(spin.initial_gold_count > 0)
            result.fg_gold_symbols += spin.initial_gold_count
            result.special_symbol_cnt += int(spin.scatter_count > 0)
            result.combo_fg[min(spin.cascades, 5)] += 1
            result.symbol_hits.update(spin.symbol_hits)
            result.symbol_pay.update(spin.symbol_pay)
            result.fg_symbol_length_hits.update(spin.symbol_length_hits)
            result.fg_symbol_length_pay.update(spin.symbol_length_pay)
            result.fg_initial_symbols.update(spin.initial_symbols)
            result.fg_drop_symbols.update(spin.drop_symbols)
            if spin.scatter_count >= 3 and played + remaining < int(self.config["free_spin_cap"]):
                add = min(int(self.config["retrigger_spins"]), int(self.config["free_spin_cap"]) - played - remaining)
                remaining += add
                result.retriggers += int(add > 0)
                group = "super_retrigger" if super_mode else "retrigger"
                selections = self.config.get("table_selection", {}).get(group)
                if selections:
                    values = [str(item["table"]) for item in selections]
                    weights = [float(item["weight"]) for item in selections]
                    queue.extend(str(weighted_pick(self.rng, values, weights)) for _ in range(add))
                else:
                    queue.extend(["high"] + ["low"] * max(0, add - 1))
        return result

    def card_feature(self, section: str, super_mode: bool = False) -> RoundResult:
        profile = "weight_1" if section == "super_feature" else self.profile
        card = self.pick_card(section, profile)
        for _ in range(CARD_RETRY_LIMIT):
            result = self.free_session(super_mode)
            if self.card_matches(card, result.pay_fg):
                return result
            self.retry_total += 1
        self.retry_limit_exceeded += 1
        self.retry_limit_fg += 1
        return result

    @staticmethod
    def merge(target: RoundResult, source: RoundResult) -> None:
        target.pay_bg += source.pay_bg
        target.pay_fg += source.pay_fg
        target.fg_triggered |= source.fg_triggered
        target.fg_spins += source.fg_spins
        target.retriggers += source.retriggers
        target.cascades_bg += source.cascades_bg
        target.cascades_fg += source.cascades_fg
        target.max_multiplier = max(target.max_multiplier, source.max_multiplier)
        target.golden_converted += source.golden_converted
        target.w2_events += source.w2_events
        target.bg_w2_events += source.bg_w2_events
        target.fg_w2_events += source.fg_w2_events
        target.bg_w2_counts.update(source.bg_w2_counts)
        target.fg_w2_counts.update(source.fg_w2_counts)
        target.bg_m1_spins += source.bg_m1_spins
        target.fg_m1_spins += source.fg_m1_spins
        target.bg_hit_spins += source.bg_hit_spins
        target.fg_hit_spins += source.fg_hit_spins
        target.bg_gold_spins += source.bg_gold_spins
        target.fg_gold_spins += source.fg_gold_spins
        target.bg_gold_symbols += source.bg_gold_symbols
        target.fg_gold_symbols += source.fg_gold_symbols
        target.special_symbol_cnt += source.special_symbol_cnt
        target.bg_trigger_fg_cnt += source.bg_trigger_fg_cnt
        target.bg_trigger_fg_pay += source.bg_trigger_fg_pay
        target.combo_fg.update(source.combo_fg)
        target.symbol_hits.update(source.symbol_hits)
        target.symbol_pay.update(source.symbol_pay)
        target.bg_symbol_length_hits.update(source.bg_symbol_length_hits)
        target.bg_symbol_length_pay.update(source.bg_symbol_length_pay)
        target.fg_symbol_length_hits.update(source.fg_symbol_length_hits)
        target.fg_symbol_length_pay.update(source.fg_symbol_length_pay)
        target.bg_initial_symbols.update(source.bg_initial_symbols)
        target.bg_drop_symbols.update(source.bg_drop_symbols)
        target.fg_initial_symbols.update(source.fg_initial_symbols)
        target.fg_drop_symbols.update(source.fg_drop_symbols)

    def round(self, bet_mode: int) -> RoundResult:
        if bet_mode in (MODE_FEATUREBUY, MODE_SUPERBUY):
            super_mode = bet_mode == MODE_SUPERBUY
            entry = self.board("buy")
            if sum(symbol == C1 for reel in entry for symbol in reel) < 3:
                raise RuntimeError("BF_Symbol weights must guarantee at least 3 C1")
            result = self.card_feature("super_feature" if super_mode else "buy_feature", super_mode) if self.card_enabled else self.free_session(super_mode)
            # The paid entry board is the round's Base/entry Spin and contains SC.
            result.special_symbol_cnt += 1
            return result
        result = RoundResult()
        if self.card_enabled:
            spin = self.card_spin(self.pick_card("base_game"))
        else:
            spin = self.natural_base_spin()
        result.pay_bg = spin.pay
        result.cascades_bg = spin.cascades
        result.max_multiplier = spin.max_multiplier
        result.golden_converted = spin.golden_converted
        result.w2_events = spin.w2_events
        result.bg_w2_events = spin.w2_events
        result.bg_w2_counts.update(spin.w2_counts)
        result.bg_m1_spins = int(spin.m1_present)
        result.bg_hit_spins = int(spin.pay > 0)
        result.bg_gold_spins = int(spin.initial_gold_count > 0)
        result.bg_gold_symbols = spin.initial_gold_count
        result.special_symbol_cnt = int(spin.scatter_count > 0)
        result.symbol_hits.update(spin.symbol_hits)
        result.symbol_pay.update(spin.symbol_pay)
        result.bg_symbol_length_hits.update(spin.symbol_length_hits)
        result.bg_symbol_length_pay.update(spin.symbol_length_pay)
        result.bg_initial_symbols.update(spin.initial_symbols)
        result.bg_drop_symbols.update(spin.drop_symbols)
        if spin.scatter_count >= 3:
            result.bg_trigger_fg_cnt = 1
            # Natural-probability reports use the same cap eligibility as the
            # Card System.  Trigger count remains the natural trigger count;
            # only eligible triggering-BG awards contribute to this pay field.
            result.bg_trigger_fg_pay = spin.pay if self.bg_trigger_within_cap(spin.pay) else 0.0
            feature = self.card_feature("free_game") if self.card_enabled else self.free_session()
            self.merge(result, feature)
        return result


def wager_for_mode(mode: int, config: dict[str, Any] | None = None) -> float:
    active = config or CFG
    factor = 1.0 if mode == MODE_NORMALBET else float(active["buy_price"] if mode == MODE_FEATUREBUY else active["super_buy_price"])
    return BASE_BET * BET_MULTI * factor


def _empty_stats() -> dict[str, Any]:
    return {
        "rounds": 0,
        "coin_in": 0.0,
        "pay_bg": 0.0,
        "pay_fg": 0.0,
        "hit_rounds": 0,
        "fg_triggers": 0,
        "fg_spins": 0,
        "retriggers": 0,
        "cascades_bg": 0,
        "cascades_fg": 0,
        "golden_converted": 0,
        "w2_events": 0,
        "bg_w2_events": 0,
        "fg_w2_events": 0,
        "bg_w2_counts": Counter(),
        "fg_w2_counts": Counter(),
        "bg_m1_spins": 0,
        "fg_m1_spins": 0,
        "bg_hit_spins": 0,
        "fg_hit_spins": 0,
        "bg_gold_spins": 0,
        "fg_gold_spins": 0,
        "bg_gold_symbols": 0,
        "fg_gold_symbols": 0,
        "special_symbol_cnt": 0,
        "bg_trigger_fg_cnt": 0,
        "bg_trigger_fg_pay": 0.0,
        "bg_trigger_fg_bucket_count": Counter(),
        "bg_trigger_fg_bucket_pay": Counter(),
        "max_multiplier": 1,
        "win_x_sum": 0.0,
        "win_x_square": 0.0,
        "retry_total": 0,
        "retry_limit_exceeded": 0,
        "retry_limit_bg_range": 0,
        "retry_limit_bg_freegame": 0,
        "retry_limit_fg": 0,
        "card_draws": Counter(),
        "combo_bg": Counter(),
        "combo_fg": Counter(),
        "buckets": Counter(),
        "multiplier_bg_count": Counter(),
        "multiplier_bg_pay": Counter(),
        "multiplier_fg_count": Counter(),
        "multiplier_fg_pay": Counter(),
        "multiplier_overall_count": Counter(),
        "multiplier_overall_pay": Counter(),
        "interval_bg_hits": Counter(),
        "interval_bg_combo": Counter(),
        "interval_bg_w2": Counter(),
        "interval_bg_gold_symbols": Counter(),
        "interval_fg_spins": Counter(),
        "interval_fg_hits": Counter(),
        "interval_fg_combo": Counter(),
        "interval_fg_w2": Counter(),
        "interval_fg_gold_symbols": Counter(),
        "symbol_hits": Counter(),
        "symbol_pay": Counter(),
        "bg_symbol_length_hits": Counter(),
        "bg_symbol_length_pay": Counter(),
        "fg_symbol_length_hits": Counter(),
        "fg_symbol_length_pay": Counter(),
        "bg_initial_symbols": Counter(),
        "bg_drop_symbols": Counter(),
        "fg_initial_symbols": Counter(),
        "fg_drop_symbols": Counter(),
    }


def _simulate_chunk(rounds: int, bet_mode: int, seed: int, config: dict[str, Any] | None = None) -> dict[str, Any]:
    active = config or CFG
    game = LuckyAce(active, seed, CARD_SYSTEM_ENABLED, CARD_SYSTEM_IS_NEWBIE)
    stats = _empty_stats()
    wager = wager_for_mode(bet_mode, active)
    for _ in range(rounds):
        result = game.round(bet_mode)
        ratio = result.pay / wager if wager else 0.0
        stats["rounds"] += 1
        stats["coin_in"] += wager
        stats["pay_bg"] += result.pay_bg
        stats["pay_fg"] += result.pay_fg
        stats["hit_rounds"] += int(result.pay > 0)
        stats["fg_triggers"] += int(result.fg_triggered)
        stats["fg_spins"] += result.fg_spins
        stats["retriggers"] += result.retriggers
        stats["cascades_bg"] += result.cascades_bg
        stats["cascades_fg"] += result.cascades_fg
        stats["golden_converted"] += result.golden_converted
        stats["w2_events"] += result.w2_events
        stats["bg_w2_events"] += result.bg_w2_events
        stats["fg_w2_events"] += result.fg_w2_events
        stats["bg_w2_counts"].update(result.bg_w2_counts)
        stats["fg_w2_counts"].update(result.fg_w2_counts)
        stats["bg_m1_spins"] += result.bg_m1_spins
        stats["fg_m1_spins"] += result.fg_m1_spins
        stats["bg_hit_spins"] += result.bg_hit_spins
        stats["fg_hit_spins"] += result.fg_hit_spins
        stats["bg_gold_spins"] += result.bg_gold_spins
        stats["fg_gold_spins"] += result.fg_gold_spins
        stats["bg_gold_symbols"] += result.bg_gold_symbols
        stats["fg_gold_symbols"] += result.fg_gold_symbols
        stats["special_symbol_cnt"] += result.special_symbol_cnt
        stats["bg_trigger_fg_cnt"] += result.bg_trigger_fg_cnt
        stats["bg_trigger_fg_pay"] += result.bg_trigger_fg_pay
        stats["max_multiplier"] = max(stats["max_multiplier"], result.max_multiplier)
        stats["win_x_sum"] += ratio
        stats["win_x_square"] += ratio * ratio
        stats["combo_bg"][min(result.cascades_bg, 5)] += 1
        stats["combo_fg"].update(result.combo_fg)
        multiplier_coin_in = BASE_BET * BET_MULTI
        bg_bucket = min(bisect.bisect_left(MULTIPLIER_THRESHOLDS, result.pay_bg / multiplier_coin_in), len(MULTIPLIER_THRESHOLDS) - 1)
        overall_bucket = min(bisect.bisect_left(MULTIPLIER_THRESHOLDS, result.pay / multiplier_coin_in), len(MULTIPLIER_THRESHOLDS) - 1)
        if bet_mode == MODE_NORMALBET:
            stats["multiplier_bg_count"][bg_bucket] += 1
            stats["multiplier_bg_pay"][bg_bucket] += result.pay_bg
            if result.bg_trigger_fg_cnt:
                # Keep the raw trigger-spin BG distribution independent of the
                # active Profile cap.  The report converts these exact buckets
                # to cumulative <= upper-bound values for any future cap.
                stats["bg_trigger_fg_bucket_count"][bg_bucket] += result.bg_trigger_fg_cnt
                stats["bg_trigger_fg_bucket_pay"][bg_bucket] += result.pay_bg
            stats["interval_bg_hits"][bg_bucket] += result.bg_hit_spins
            stats["interval_bg_combo"][(bg_bucket, min(result.cascades_bg, 5))] += 1
            for ghost_count in (2, 3, 4):
                stats["interval_bg_w2"][(bg_bucket, ghost_count)] += result.bg_w2_counts[ghost_count]
            stats["interval_bg_gold_symbols"][bg_bucket] += result.bg_gold_symbols
        if result.fg_triggered:
            fg_bucket = min(bisect.bisect_left(MULTIPLIER_THRESHOLDS, result.pay_fg / multiplier_coin_in), len(MULTIPLIER_THRESHOLDS) - 1)
            stats["multiplier_fg_count"][fg_bucket] += 1
            stats["multiplier_fg_pay"][fg_bucket] += result.pay_fg
            stats["interval_fg_spins"][fg_bucket] += result.fg_spins
            stats["interval_fg_hits"][fg_bucket] += result.fg_hit_spins
            for combo_count in range(1, 6):
                stats["interval_fg_combo"][(fg_bucket, combo_count)] += result.combo_fg[combo_count]
            for ghost_count in (2, 3, 4):
                stats["interval_fg_w2"][(fg_bucket, ghost_count)] += result.fg_w2_counts[ghost_count]
            stats["interval_fg_gold_symbols"][fg_bucket] += result.fg_gold_symbols
        stats["multiplier_overall_count"][overall_bucket] += 1
        stats["multiplier_overall_pay"][overall_bucket] += result.pay
        label = "0" if ratio == 0 else "(0,1)" if ratio < 1 else "[1,10)" if ratio < 10 else "[10,100)" if ratio < 100 else "100+"
        stats["buckets"][label] += 1
        stats["symbol_hits"].update(result.symbol_hits)
        stats["symbol_pay"].update(result.symbol_pay)
        stats["bg_symbol_length_hits"].update(result.bg_symbol_length_hits)
        stats["bg_symbol_length_pay"].update(result.bg_symbol_length_pay)
        stats["fg_symbol_length_hits"].update(result.fg_symbol_length_hits)
        stats["fg_symbol_length_pay"].update(result.fg_symbol_length_pay)
        stats["bg_initial_symbols"].update(result.bg_initial_symbols)
        stats["bg_drop_symbols"].update(result.bg_drop_symbols)
        stats["fg_initial_symbols"].update(result.fg_initial_symbols)
        stats["fg_drop_symbols"].update(result.fg_drop_symbols)
    stats["retry_total"] = game.retry_total
    stats["retry_limit_exceeded"] = game.retry_limit_exceeded
    stats["retry_limit_bg_range"] = game.retry_limit_bg_range
    stats["retry_limit_bg_freegame"] = game.retry_limit_bg_freegame
    stats["retry_limit_fg"] = game.retry_limit_fg
    stats["card_draws"].update(game.card_draws)
    return stats


def _merge_stats(target: dict[str, Any], source: dict[str, Any]) -> None:
    counter_fields = {
        "combo_bg",
        "combo_fg",
        "buckets",
        "symbol_hits",
        "symbol_pay",
        "bg_symbol_length_hits",
        "bg_symbol_length_pay",
        "fg_symbol_length_hits",
        "fg_symbol_length_pay",
        "bg_initial_symbols",
        "bg_drop_symbols",
        "fg_initial_symbols",
        "fg_drop_symbols",
        "bg_w2_counts",
        "fg_w2_counts",
        "multiplier_bg_count",
        "multiplier_bg_pay",
        "multiplier_fg_count",
        "multiplier_fg_pay",
        "multiplier_overall_count",
        "multiplier_overall_pay",
        "bg_trigger_fg_bucket_count",
        "bg_trigger_fg_bucket_pay",
        "interval_bg_hits",
        "interval_bg_combo",
        "interval_bg_w2",
        "interval_bg_gold_symbols",
        "interval_fg_spins",
        "interval_fg_hits",
        "interval_fg_combo",
        "interval_fg_w2",
        "interval_fg_gold_symbols",
        "card_draws",
    }
    for key, value in source.items():
        if key in counter_fields:
            target[key].update(value)
        elif key == "max_multiplier":
            target[key] = max(target[key], value)
        else:
            target[key] += value


def _simulation_result(
    stats: dict[str, Any],
    duration: float,
    bet_mode: int,
    bet_multi: int,
    active: dict[str, Any],
    seed: int,
) -> dict[str, Any]:
    return {
        "stats": stats,
        "duration": float(duration),
        "bet_mode": int(bet_mode),
        "bet_multi": int(bet_multi),
        "seed": int(seed),
        "config_file": CONFIG_FILE,
        "config_rtp_file": CONFIG_RTP_FILE,
        "math_version": MATH_VERSION,
        "base_math_version": str(CFG_NATURAL.get("excel_version", "")),
        "base_parsheet_id": str(CFG_NATURAL.get("parsheet_id", "")),
        "base_source_xlsx": str(CFG_NATURAL.get("source_xlsx", "")),
        "game_id": str(active.get("game_id", GAME_ID)),
        "parsheet_id": str(active.get("parsheet_id", PARSHEET_ID)),
        "source_multiplier_xlsx": str(active.get("source_multiplier_xlsx", "")),
        "game_name": str(active.get("name_zh", GAME_NAME)),
        "card_system_enabled": bool(CARD_SYSTEM_ENABLED),
        "card_system_is_newbie": bool(CARD_SYSTEM_IS_NEWBIE),
        "card_retry_limit": int(CARD_RETRY_LIMIT),
        "card_system_config": active.get("card_system") or {},
    }


def run_simulation(
    total_rounds: int = TOTAL_ROUNDS,
    bet_mode: int = BET_MODE,
    bet_multi: int = BET_MULTI,
    threads: int = THREADS,
    config: dict[str, Any] | None = None,
    seed: int = RNG_SEED,
) -> dict[str, Any]:
    global BET_MULTI
    if bet_mode not in SUPPORTED_BET_MODES:
        raise ValueError(f"Unsupported bet mode: {bet_mode}")
    BET_MULTI = int(bet_multi)
    threads = max(1, min(int(threads), int(total_rounds)))
    base, extra = divmod(int(total_rounds), threads)
    chunks = [base + (1 if i < extra else 0) for i in range(threads)]
    active = config or CFG
    fast_api_version = getattr(fast_simulator, "FAST_SIMULATOR_API_VERSION", 1) if fast_simulator is not None else 0
    fast_card_capable = fast_api_version >= 2
    if FAST_SIMULATION and fast_simulator is not None and (not CARD_SYSTEM_ENABLED or fast_card_capable):
        packed = fast_simulator.prepare_config(active)
        if fast_card_capable:
            fast_simulator.warm(
                packed,
                int(bet_mode),
                int(bet_multi),
                seed=int(seed),
                card_enabled=bool(CARD_SYSTEM_ENABLED),
                card_newbie=bool(CARD_SYSTEM_IS_NEWBIE),
            )
        else:
            fast_simulator.warm(packed, int(bet_mode), int(bet_multi), int(seed))
        started = time.perf_counter()
        if fast_card_capable:
            packed_result = fast_simulator.run_prepared(
                packed,
                int(total_rounds),
                int(bet_mode),
                int(bet_multi),
                threads,
                seed=int(seed),
                card_enabled=bool(CARD_SYSTEM_ENABLED),
                card_newbie=bool(CARD_SYSTEM_IS_NEWBIE),
            )
        else:
            packed_result = fast_simulator.run_prepared(
                packed,
                int(total_rounds),
                int(bet_mode),
                int(bet_multi),
                threads,
                seed=int(seed),
            )
        return _simulation_result(
            fast_simulator.to_stats(packed_result),
            time.perf_counter() - started,
            bet_mode,
            bet_multi,
            active,
            seed,
        )
    started = time.perf_counter()
    merged = _empty_stats()
    if threads == 1:
        _merge_stats(merged, _simulate_chunk(chunks[0], bet_mode, int(seed), active))
    else:
        with ThreadPoolExecutor(max_workers=threads) as pool:
            futures = [pool.submit(_simulate_chunk, count, bet_mode, int(seed) + i * 100003, active) for i, count in enumerate(chunks) if count]
            for future in futures:
                _merge_stats(merged, future.result())
    return _simulation_result(
        merged,
        time.perf_counter() - started,
        bet_mode,
        bet_multi,
        active,
        seed,
    )


simulate = run_simulation


def mode_name(mode: int) -> str:
    return {0: "Normal Bet", 2: "Buy Feature", 3: "Buy Super Feature"}.get(mode, f"Mode {mode}")


def calculated_metrics(result: dict[str, Any]) -> dict[str, Any]:
    s = result["stats"]
    rounds, coin_in = max(1, s["rounds"]), max(1.0, s["coin_in"])
    pay_total = s["pay_bg"] + s["pay_fg"]
    mean = s["win_x_sum"] / rounds
    variance = max(0.0, s["win_x_square"] / rounds - mean * mean)
    fg_triggers = s["fg_triggers"]
    fg_spins = s["fg_spins"]
    retriggers = s["retriggers"]
    special_symbol_cnt = int(s.get("special_symbol_cnt", 0))
    # SCR is normalized by paid rounds only.  FG Spins contribute to the
    # special-symbol numerator, but are not added to the denominator.
    special_symbol_rate = special_symbol_cnt / rounds
    return {
        "rounds": s["rounds"],
        "coin_in_per_round": s["coin_in"] / rounds,
        "rtp_total": pay_total / coin_in,
        "rtp_bg": s["pay_bg"] / coin_in,
        "rtp_fg": s["pay_fg"] / coin_in,
        "hit_rate_bg": s["bg_hit_spins"] / rounds,
        "hit_rate_fg": s["fg_hit_spins"] / max(1, fg_spins),
        "fg_trigger_rate": fg_triggers / rounds,
        "fg_trigger_cycle": rounds / fg_triggers if fg_triggers else math.inf,
        "retrigger_trigger_rate": retriggers / max(1, fg_spins),
        "retrigger_trigger_cycle": fg_spins / retriggers if retriggers else math.inf,
        "avg_fg_spins": fg_spins / fg_triggers if fg_triggers else 0.0,
        "special_symbol_cnt": special_symbol_cnt,
        "special_symbol_rate": special_symbol_rate,
        "SCR": special_symbol_rate * 10_000_000_000,
        "volatility_std": math.sqrt(variance),
        "standard_error": math.sqrt(variance) / math.sqrt(rounds),
    }


def common_summary_sections(result: dict[str, Any]) -> list[list[tuple[str, Any]]]:
    """Return the fixed Console sections required by specification section 3.3."""
    s = result["stats"]
    metrics = calculated_metrics(result)
    sections: list[list[tuple[str, Any]]] = [
        [
            ("game_name", result["game_name"]),
            ("game_id", result["game_id"]),
        ],
        [
            ("config_file", result["config_file"]),
            ("config_rtp_file", result["config_rtp_file"]),
            ("math_version", result["math_version"]),
            ("card_system", "on" if result["card_system_enabled"] else "off"),
        ],
        [
            ("bet_mode", mode_name(result["bet_mode"])),
            ("bet_multi", result["bet_multi"]),
            ("coin_in", metrics["coin_in_per_round"]),
            ("total_rounds", s["rounds"]),
            ("duration", result["duration"]),
        ],
    ]
    sections.extend(
        [
            [
                ("rtp_total", metrics["rtp_total"]),
                ("rtp_bg", metrics["rtp_bg"]),
                ("rtp_fg", metrics["rtp_fg"]),
                ("hit_rate_bg", metrics["hit_rate_bg"]),
                ("hit_rate_fg", metrics["hit_rate_fg"]),
                ("fg_trigger_rate", (metrics["fg_trigger_rate"], metrics["fg_trigger_cycle"])),
                ("retrigger_trigger_rate", (metrics["retrigger_trigger_rate"], metrics["retrigger_trigger_cycle"])),
                ("avg_fg_spins", metrics["avg_fg_spins"]),
            ],
            [
                ("bg_trigger_fg_cnt", s.get("bg_trigger_fg_cnt", 0)),
                ("bg_trigger_fg_pay", s.get("bg_trigger_fg_pay", 0.0)),
                ("special_symbol_cnt", metrics["special_symbol_cnt"]),
                ("SCR", metrics["SCR"]),
            ],
            [
                ("volatility_std", metrics["volatility_std"]),
                ("standard_error", metrics["standard_error"]),
            ],
        ]
    )
    if result["card_system_enabled"]:
        sections.extend(
            [
                [
                    ("card_system_profile", "newbie" if result["card_system_is_newbie"] else "oldhand"),
                    ("card_retry_limit", result["card_retry_limit"]),
                    ("retry_total", s.get("retry_total", 0)),
                    ("avg_retry", s.get("retry_total", 0) / max(1, s["rounds"])),
                ],
                [
                    ("retry_limit_exceeded", s.get("retry_limit_exceeded", 0)),
                    ("retry_limit_bg_range", s.get("retry_limit_bg_range", 0)),
                    ("retry_limit_bg_freegame", s.get("retry_limit_bg_freegame", 0)),
                    ("retry_limit_fg", s.get("retry_limit_fg", 0)),
                ],
            ]
        )
    return sections


def common_summary_rows(result: dict[str, Any]) -> list[tuple[str, Any]]:
    """Flatten the fixed Console sections without changing their order."""
    return [row for section in common_summary_sections(result) for row in section]


def game_info_rows(result: dict[str, Any]) -> list[tuple[str, Any]]:
    s = result["stats"]
    rounds = max(1, s["rounds"])
    fg_spins = max(1, s["fg_spins"])
    return [
        ("avg_cascades_bg", s["cascades_bg"] / rounds),
        ("avg_cascades_fg", s["cascades_fg"] / fg_spins),
        ("golden_converted", s["golden_converted"]),
        ("w2_events", s["w2_events"]),
        ("w2_bg_event_rate", s["bg_w2_events"] / rounds),
        ("w2_fg_event_rate", s["fg_w2_events"] / fg_spins),
        ("w2_bg_count_2", s["bg_w2_counts"][2]),
        ("w2_bg_count_3", s["bg_w2_counts"][3]),
        ("w2_bg_count_4", s["bg_w2_counts"][4]),
        ("w2_fg_count_2", s["fg_w2_counts"][2]),
        ("w2_fg_count_3", s["fg_w2_counts"][3]),
        ("w2_fg_count_4", s["fg_w2_counts"][4]),
        ("m1_bg_spin_rate", s["bg_m1_spins"] / rounds),
    ]


def _display_value(key: str, value: Any) -> str:
    if key == "duration":
        return f"{float(value):05.2f} sec"
    if key in {
        "total_rounds",
        "bg_trigger_fg_cnt",
        "special_symbol_cnt",
        "SCR",
        "retry_total",
        "retry_limit_exceeded",
        "retry_limit_bg_range",
        "retry_limit_bg_freegame",
        "retry_limit_fg",
        "golden_converted",
        "w2_events",
        "w2_bg_count_2",
        "w2_bg_count_3",
        "w2_bg_count_4",
        "w2_fg_count_2",
        "w2_fg_count_3",
        "w2_fg_count_4",
    }:
        return f"{int(value):,}"
    if key == "card_retry_limit":
        return str(int(value))
    if key == "bg_trigger_fg_pay":
        return f"{float(value):,.0f}"
    if key == "coin_in":
        return f"{float(value):.1f}"
    if key in {"fg_trigger_rate", "retrigger_trigger_rate"}:
        rate, cycle = value
        unit = "spins" if key == "fg_trigger_rate" else "free spins"
        cycle_text = "N/A" if not math.isfinite(float(cycle)) else f"{float(cycle):.2f} {unit}"
        return f"{float(rate):.4%} (cycle {cycle_text})"
    if key.startswith("rtp_") or key in {"hit_rate_bg", "hit_rate_fg", "special_symbol_rate"}:
        return f"{float(value):.4%}"
    if key.endswith("_rate"):
        return f"{float(value):.6%}"
    if key == "avg_fg_spins":
        return f"{float(value):.2f} spins"
    if key in {"avg_retry", "volatility_std", "standard_error"}:
        return f"{float(value):05.2f}"
    if isinstance(value, float):
        return f"{value:.6f}"
    return str(value)


def overview_rows(result: dict[str, Any]) -> list[tuple[str, str]]:
    return [(key, _display_value(key, value)) for key, value in common_summary_rows(result) + game_info_rows(result)]


def summary_rows(result: dict[str, Any]) -> list[tuple[str, Any]]:
    """Backward-compatible numeric rows used by existing analysis scripts."""
    metrics = calculated_metrics(result)
    rows = [
        ("parsheet_id", result["parsheet_id"]),
        ("game_id", result["game_id"]),
        ("game_name", result["game_name"]),
        ("config_file", result["config_file"]),
        ("config_rtp_file", result["config_rtp_file"]),
        ("math_version", result["math_version"]),
        ("card_system", "on" if result["card_system_enabled"] else "off"),
        ("card_system_profile", "newbie" if result["card_system_is_newbie"] else "oldhand"),
        ("card_retry_limit", result["card_retry_limit"] if result["card_system_enabled"] else 0),
        ("retry_total", result["stats"].get("retry_total", 0)),
        ("avg_retry", result["stats"].get("retry_total", 0) / max(1, result["stats"]["rounds"])),
        ("retry_limit_exceeded", result["stats"].get("retry_limit_exceeded", 0)),
        ("retry_limit_bg_range", result["stats"].get("retry_limit_bg_range", 0)),
        ("retry_limit_bg_freegame", result["stats"].get("retry_limit_bg_freegame", 0)),
        ("retry_limit_fg", result["stats"].get("retry_limit_fg", 0)),
        ("bet_mode", mode_name(result["bet_mode"])),
        ("bet_multi", result["bet_multi"]),
        ("coin_in", metrics["coin_in_per_round"]),
        ("total_rounds", result["stats"]["rounds"]),
        ("duration_sec", round(result["duration"], 3)),
        ("rtp_total", metrics["rtp_total"]),
        ("rtp_bg", metrics["rtp_bg"]),
        ("rtp_fg", metrics["rtp_fg"]),
        ("hit_rate", result["stats"]["hit_rounds"] / max(1, result["stats"]["rounds"])),
        ("bg_hit_rate", metrics["hit_rate_bg"]),
        ("fg_hit_rate", metrics["hit_rate_fg"]),
        ("fg_trigger_rate", metrics["fg_trigger_rate"]),
        ("fg_trigger_cycle", metrics["fg_trigger_cycle"]),
        ("retrigger_trigger_rate", metrics["retrigger_trigger_rate"]),
        ("retrigger_trigger_cycle", metrics["retrigger_trigger_cycle"]),
        ("avg_fg_spins", metrics["avg_fg_spins"]),
        ("bg_trigger_fg_cnt", result["stats"].get("bg_trigger_fg_cnt", 0)),
        ("bg_trigger_fg_pay", result["stats"].get("bg_trigger_fg_pay", 0.0)),
        ("special_symbol_cnt", metrics["special_symbol_cnt"]),
        ("SCR", metrics["SCR"]),
        ("volatility_std", metrics["volatility_std"]),
        ("standard_error", metrics["standard_error"]),
    ]
    rows.extend(game_info_rows(result))
    rows.extend(
        [
            ("m1_fg_spin_rate", result["stats"]["fg_m1_spins"] / max(1, result["stats"]["fg_spins"])),
            ("bg_gold_spin_rate", result["stats"]["bg_gold_spins"] / max(1, result["stats"]["rounds"])),
            ("fg_gold_spin_rate", result["stats"]["fg_gold_spins"] / max(1, result["stats"]["fg_spins"])),
            ("avg_bg_gold_symbols", result["stats"]["bg_gold_symbols"] / max(1, result["stats"]["rounds"])),
            ("avg_fg_gold_symbols", result["stats"]["fg_gold_symbols"] / max(1, result["stats"]["fg_spins"])),
            ("max_win_multiplier", result["stats"]["max_multiplier"]),
            ("rounds_per_second", result["stats"]["rounds"] / max(result["duration"], 1e-9)),
        ]
    )
    return rows


def symbol_ratio_df(counter: Counter) -> pd.DataFrame:
    reel_totals = [sum(counter[(reel, symbol)] for symbol in SYMBOL_STR) for reel in range(5)]
    symbols = [symbol for symbol in sorted(SYMBOL_STR) if any(counter[(reel, symbol)] for reel in range(5))]
    rows = []
    for symbol in symbols:
        row: dict[str, Any] = {"Symbol": SYMBOL_STR[symbol]}
        for reel in range(5):
            total = reel_totals[reel]
            row[f"R{reel + 1}"] = counter[(reel, symbol)] / total if total else 0.0
        rows.append(row)
    return pd.DataFrame(rows, columns=["Symbol", "R1", "R2", "R3", "R4", "R5"])


def symbol_ratio_tables(result: dict[str, Any]) -> list[tuple[str, pd.DataFrame]]:
    s = result["stats"]
    return [
        ("BG 初始 R1-R5", symbol_ratio_df(s["bg_initial_symbols"])),
        ("BG 掉落 R1-R5", symbol_ratio_df(s["bg_drop_symbols"])),
        ("FG 初始 R1-R5", symbol_ratio_df(s["fg_initial_symbols"])),
        ("FG 掉落 R1-R5", symbol_ratio_df(s["fg_drop_symbols"])),
    ]


def print_symbol_ratio_tables(result: dict[str, Any]) -> None:
    formatters = {f"R{reel}": (lambda value: f"{value:.4%}") for reel in range(1, 6)}
    for title, frame in symbol_ratio_tables(result):
        print(f"\n=== {title} ===")
        print("No samples" if frame.empty else frame.to_string(index=False, formatters=formatters))


def print_console(result: dict[str, Any]) -> None:
    if SHOW_CONSOLE_SUMMARY:
        print()
        for section_index, section in enumerate(common_summary_sections(result)):
            if section_index:
                print()
            for key, value in section:
                print(f"{key:24s}: {_display_value(key, value)}")
        print("\n<< By Game Info >>\n")
        for key, value in game_info_rows(result):
            print(f"{key:24s}: {_display_value(key, value)}")
    if SHOW_CONSOLE_DETAIL:
        print_symbol_ratio_tables(result)
        print("\nPay buckets:", dict(result["stats"]["buckets"]))


def format_rounds_tag(total_rounds: int) -> str:
    """Format exact powers of ten as 10N, for example 10**7 -> 107."""
    total_rounds = int(total_rounds)
    value = total_rounds
    exponent = 0
    while value > 0 and value % 10 == 0:
        value //= 10
        exponent += 1
    if value == 1 and exponent > 0:
        return f"10{exponent}"
    return str(total_rounds)


def format_version_tag(version: Any) -> str:
    """Convert a four-part math version to the fixed eight-digit report tag."""
    parts = str(version or "").strip().split(".")
    if len(parts) != 4 or any(not part.isdigit() for part in parts):
        raise ValueError(f"math_version must contain four numeric parts, got {version!r}")
    values = [int(part) for part in parts]
    if any(value < 0 or value > 99 for value in values):
        raise ValueError(f"math_version parts must be between 0 and 99, got {version!r}")
    return "".join(f"{value:02d}" for value in values)


def format_base_version_tag(version: Any) -> str:
    """Convert the one-part base math version to the fixed two-digit tag."""
    raw = str(version or "").strip()
    if not raw.isdigit():
        raise ValueError(f"base_math_version must contain one numeric part, got {version!r}")
    value = int(raw)
    if value < 0 or value > 99:
        raise ValueError(f"base_math_version must be between 0 and 99, got {version!r}")
    return f"{value:02d}"


def format_rtp_tag(rtp_ratio: Any) -> str:
    """Convert an RTP ratio to its percentage tag: 0.9201 -> 9201."""
    value = float(rtp_ratio)
    if not math.isfinite(value) or value < 0:
        raise ValueError(f"rtp_total must be a non-negative finite value, got {rtp_ratio!r}")
    return f"{value * 100:.2f}".replace(".", "")


def report_game_id(result: dict[str, Any]) -> str:
    """Resolve the full RTP/Variant game id used by the report filename."""
    candidates = (
        Path(str(result.get("source_multiplier_xlsx") or "")).stem,
        str(result.get("parsheet_id") or ""),
        str(result.get("game_id") or ""),
    )
    for candidate in candidates:
        if re.fullmatch(r"H\d{3}1\d{2}[A-Z]", candidate, flags=re.IGNORECASE):
            return candidate.upper()
    parsheet_id = str(result.get("parsheet_id") or "")
    if re.fullmatch(r"H\d{3}1\d{2}", parsheet_id, flags=re.IGNORECASE):
        return f"{parsheet_id.upper()}A"
    base_game_id = str(result.get("game_id") or "").upper()
    config_match = re.search(
        r"config_(\d{2})([A-Z]?)",
        Path(str(result.get("config_rtp_file") or "")).stem,
        flags=re.IGNORECASE,
    )
    if re.fullmatch(r"H\d{3}", base_game_id) and config_match:
        variant = (config_match.group(2) or "A").upper()
        return f"{base_game_id}1{config_match.group(1)}{variant}"
    raise ValueError("Unable to resolve the full RTP/Variant game id for report output")


def report_base_game_id(result: dict[str, Any]) -> str:
    """Resolve the complete base-game id used when Card System is off."""
    candidates = (
        Path(str(result.get("base_source_xlsx") or "")).stem,
        str(result.get("base_parsheet_id") or ""),
    )
    for candidate in candidates:
        if re.fullmatch(r"H\d{3}1", candidate, flags=re.IGNORECASE):
            return candidate.upper()
    base_game_id = str(result.get("game_id") or "").upper()
    if re.fullmatch(r"H\d{3}", base_game_id):
        return f"{base_game_id}1"
    raise ValueError("Unable to resolve the complete base game id for report output")


def report_filename(
    result: dict[str, Any],
    rtp_total: Any,
    timestamp: str | None = None,
) -> str:
    """Build the exact report filename required by the slot specification."""
    timestamp = timestamp or datetime.now().strftime("%y%m%d%H%M")
    card_enabled = bool(result["card_system_enabled"])
    if card_enabled:
        game_id = report_game_id(result)
        version_tag = format_version_tag(result["math_version"])
    else:
        game_id = report_base_game_id(result)
        version_tag = format_base_version_tag(result["base_math_version"])
    parts = [
        game_id,
        version_tag,
        timestamp,
        f"betmode{result['bet_mode']}",
        format_rounds_tag(result["stats"]["rounds"]),
    ]
    if card_enabled:
        parts.append(format_rtp_tag(rtp_total))
        if int(result["bet_mode"]) == MODE_NORMALBET:
            parts.append("newbie" if result["card_system_is_newbie"] else "oldhand")
        parts.append("card")
    return "_".join(parts) + ".xlsx"


def card_selection_df(result: dict[str, Any]) -> pd.DataFrame:
    card_system = result.get("card_system_config") or {}
    profiles = card_system.get("profiles") or {}
    draws: Counter = result["stats"].get("card_draws", Counter())
    profile_names = ("weight_1", "weight_2")
    profile_labels = ("newbie", "oldhand")
    section_names = ("base_game", "free_game", "buy_feature", "super_feature")
    rows: list[dict[str, Any]] = []
    for profile_index, (profile_name, profile_label) in enumerate(zip(profile_names, profile_labels)):
        profile = profiles.get(profile_name) or {}
        for section_index, section_name in enumerate(section_names):
            cards = list(profile.get(section_name) or [])
            total_weight = sum(max(0.0, float(card.get("weight", 0.0))) for card in cards)
            total_draws = sum(draws[(profile_index, section_index, index)] for index in range(len(cards)))
            for card_index, card in enumerate(cards):
                draw_count = draws[(profile_index, section_index, card_index)]
                weight = max(0.0, float(card.get("weight", 0.0)))
                rows.append(
                    {
                        "Profile": profile_label,
                        "Config_Profile": profile_name,
                        "Section": section_name,
                        "Card_Index": card_index,
                        "Type": str(card.get("type", "range")),
                        "Min": card.get("min"),
                        "Max": card.get("max"),
                        "Table": card.get("table"),
                        "Weight": weight,
                        "Expected_Rate": weight / total_weight if total_weight else 0.0,
                        "Draw_Count": int(draw_count),
                        "Actual_Rate": draw_count / total_draws if total_draws else 0.0,
                    }
                )
    return pd.DataFrame(rows)


def feature_summary_df(result: dict[str, Any]) -> pd.DataFrame:
    s = result["stats"]
    rounds = max(1, s["rounds"])
    fg_spins = max(1, s["fg_spins"])
    return pd.DataFrame(
        [
            {
                "Bet_Mode": mode_name(result["bet_mode"]),
                "FG_Triggers": s["fg_triggers"],
                "FG_Trigger_Rate": s["fg_triggers"] / rounds,
                "FG_Spins": s["fg_spins"],
                "Avg_FG_Spins": s["fg_spins"] / s["fg_triggers"] if s["fg_triggers"] else 0.0,
                "Retriggers": s["retriggers"],
                "Retrigger_Rate": s["retriggers"] / fg_spins,
            }
        ]
    )


def output_report(result: dict[str, Any], output_dir: Path | None = None) -> Path:
    report_dir = Path(output_dir) if output_dir is not None else OUTPUT_DIR
    report_dir.mkdir(parents=True, exist_ok=True)
    s = result["stats"]
    summary = dict(summary_rows(result))
    path = report_dir / report_filename(result, summary["rtp_total"])
    overview_df = pd.DataFrame(overview_rows(result), columns=["field", "value"])
    base_df = pd.DataFrame(summary_rows(result), columns=["field", "value"])
    symbols = SCORE_SYMBOLS
    hits_df = pd.DataFrame({"symbol": [SYMBOL_STR[symbol] for symbol in symbols], "hits": [s["symbol_hits"][symbol] for symbol in symbols], "pay": [s["symbol_pay"][symbol] for symbol in symbols]})
    symbol_length_df = pd.DataFrame(
        [
            {
                "scene": scene,
                "symbol": SYMBOL_STR[symbol],
                "length": length,
                "hits": s[f"{scene.lower()}_symbol_length_hits"][(symbol, length)],
                "pay": s[f"{scene.lower()}_symbol_length_pay"][(symbol, length)],
            }
            for scene in ("BG", "FG")
            for symbol in symbols
            for length in (3, 4, 5)
        ]
    )
    combo_df = pd.DataFrame({"combo": ["0", "1", "2", "3", "4", "5+"], "BG": [s["combo_bg"][i] for i in range(6)], "FG": [s["combo_fg"][i] for i in range(6)]})
    interval_labels = ["0" if index == 0 else f"{MULTIPLIER_THRESHOLDS[index - 1]:.1f} < X <= {threshold:.1f}" for index, threshold in enumerate(MULTIPLIER_THRESHOLDS)]

    def interval_rate(counter: Counter, denominator: Counter, key_builder) -> list[float]:
        return [counter[key_builder(index)] / denominator[index] if denominator[index] else 0.0 for index in range(len(MULTIPLIER_THRESHOLDS))]

    bg_interval_count = s["multiplier_bg_count"]
    fg_interval_spins = s["interval_fg_spins"]
    zero_counts = [0 for _ in MULTIPLIER_THRESHOLDS]
    zero_pay = [0.0 for _ in MULTIPLIER_THRESHOLDS]
    fg_counts = [s["multiplier_fg_count"][i] for i in range(len(MULTIPLIER_THRESHOLDS))]
    fg_pay = [s["multiplier_fg_pay"][i] for i in range(len(MULTIPLIER_THRESHOLDS))]
    natural_fg_counts = fg_counts if result["bet_mode"] == MODE_NORMALBET else zero_counts
    natural_fg_pay = fg_pay if result["bet_mode"] == MODE_NORMALBET else zero_pay
    buy_fg_counts = fg_counts if result["bet_mode"] == MODE_FEATUREBUY else zero_counts
    buy_fg_pay = fg_pay if result["bet_mode"] == MODE_FEATUREBUY else zero_pay
    super_fg_counts = fg_counts if result["bet_mode"] == MODE_SUPERBUY else zero_counts
    super_fg_pay = fg_pay if result["bet_mode"] == MODE_SUPERBUY else zero_pay
    trigger_count_lte: list[int] = []
    trigger_pay_lte: list[float] = []
    running_trigger_count = 0
    running_trigger_pay = 0.0
    for index in range(len(MULTIPLIER_THRESHOLDS)):
        running_trigger_count += int(s["bg_trigger_fg_bucket_count"][index])
        running_trigger_pay += float(s["bg_trigger_fg_bucket_pay"][index])
        trigger_count_lte.append(running_trigger_count)
        trigger_pay_lte.append(running_trigger_pay)
    multiplier_df = pd.DataFrame(
        {
            "Interval": interval_labels,
            "base_game_cnt": [s["multiplier_bg_count"][i] for i in range(len(MULTIPLIER_THRESHOLDS))],
            "base_game_pay": [s["multiplier_bg_pay"][i] for i in range(len(MULTIPLIER_THRESHOLDS))],
            "free_game_cnt": natural_fg_counts,
            "free_game_pay": natural_fg_pay,
            "free_game_cnt_BF": buy_fg_counts,
            "free_game_pay_BF": buy_fg_pay,
            "free_game_cnt_SF": super_fg_counts,
            "free_game_pay_SF": super_fg_pay,
            "Interval_Upper": list(MULTIPLIER_THRESHOLDS),
            "bg_trigger_fg_cnt_lte_upper": trigger_count_lte,
            "bg_trigger_fg_pay_lte_upper": trigger_pay_lte,
            "FG_Hit_Rate": interval_rate(s["interval_fg_hits"], fg_interval_spins, lambda i: i),
            "FG_Spin_Count": [s["interval_fg_spins"][i] for i in range(len(MULTIPLIER_THRESHOLDS))],
            "BG_Combo_1_Rate": interval_rate(s["interval_bg_combo"], bg_interval_count, lambda i: (i, 1)),
            "BG_Combo_2_Rate": interval_rate(s["interval_bg_combo"], bg_interval_count, lambda i: (i, 2)),
            "BG_Combo_3_Rate": interval_rate(s["interval_bg_combo"], bg_interval_count, lambda i: (i, 3)),
            "BG_Combo_4_Rate": interval_rate(s["interval_bg_combo"], bg_interval_count, lambda i: (i, 4)),
            "BG_Combo_5+_Rate": interval_rate(s["interval_bg_combo"], bg_interval_count, lambda i: (i, 5)),
            "FG_Combo_1_Rate": interval_rate(s["interval_fg_combo"], fg_interval_spins, lambda i: (i, 1)),
            "FG_Combo_2_Rate": interval_rate(s["interval_fg_combo"], fg_interval_spins, lambda i: (i, 2)),
            "FG_Combo_3_Rate": interval_rate(s["interval_fg_combo"], fg_interval_spins, lambda i: (i, 3)),
            "FG_Combo_4_Rate": interval_rate(s["interval_fg_combo"], fg_interval_spins, lambda i: (i, 4)),
            "FG_Combo_5+_Rate": interval_rate(s["interval_fg_combo"], fg_interval_spins, lambda i: (i, 5)),
            "BG_Big_Ghost_2_Rate": interval_rate(s["interval_bg_w2"], bg_interval_count, lambda i: (i, 2)),
            "BG_Big_Ghost_3_Rate": interval_rate(s["interval_bg_w2"], bg_interval_count, lambda i: (i, 3)),
            "BG_Big_Ghost_4_Rate": interval_rate(s["interval_bg_w2"], bg_interval_count, lambda i: (i, 4)),
            "FG_Big_Ghost_2_Rate": interval_rate(s["interval_fg_w2"], fg_interval_spins, lambda i: (i, 2)),
            "FG_Big_Ghost_3_Rate": interval_rate(s["interval_fg_w2"], fg_interval_spins, lambda i: (i, 3)),
            "FG_Big_Ghost_4_Rate": interval_rate(s["interval_fg_w2"], fg_interval_spins, lambda i: (i, 4)),
            "BG_Avg_Gold_Frames": interval_rate(s["interval_bg_gold_symbols"], bg_interval_count, lambda i: i),
            "FG_Avg_Gold_Frames": interval_rate(s["interval_fg_gold_symbols"], fg_interval_spins, lambda i: i),
        }
    )
    record_df = pd.DataFrame({"field": list(s.keys()), "value": [dict(v) if isinstance(v, Counter) else v for v in s.values()]})
    ratio_sheets = [
        ("BG Initial Symbol", symbol_ratio_df(s["bg_initial_symbols"])),
        ("BG Drop Symbol", symbol_ratio_df(s["bg_drop_symbols"])),
        ("FG Initial Symbol", symbol_ratio_df(s["fg_initial_symbols"])),
        ("FG Drop Symbol", symbol_ratio_df(s["fg_drop_symbols"])),
    ]
    feature_df = feature_summary_df(result)
    card_df = card_selection_df(result) if result["card_system_enabled"] else pd.DataFrame()
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        overview_df.to_excel(writer, sheet_name="Overview", index=False)
        # Legacy numeric alias retained for existing H016 analysis scripts.
        base_df.to_excel(writer, sheet_name="Base Info", index=False)
        feature_df.to_excel(writer, sheet_name="Feature", index=False)
        hits_df.to_excel(writer, sheet_name="Hits", index=False)
        symbol_length_df.to_excel(writer, sheet_name="Symbol Length", index=False)
        combo_df.to_excel(writer, sheet_name="Eliminate", index=False)
        multiplier_df.to_excel(writer, sheet_name="Multiplier Line", index=False)
        record_df.to_excel(writer, sheet_name="Record Data", index=False)
        if not card_df.empty:
            card_df.to_excel(writer, sheet_name="Card Selection", index=False)
        for sheet_name, frame in ratio_sheets:
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes = "B2"
            worksheet.column_dimensions["A"].width = 14
            for column in "BCDEF":
                worksheet.column_dimensions[column].width = 13
                for row in range(2, len(frame) + 2):
                    worksheet[f"{column}{row}"].number_format = "0.0000%"
        worksheet = writer.sheets["Multiplier Line"]
        worksheet.freeze_panes = "B2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.column_dimensions["A"].width = 22
        for column_index, column_name in enumerate(multiplier_df.columns, start=1):
            column_letter = worksheet.cell(row=1, column=column_index).column_letter
            worksheet.column_dimensions[column_letter].width = max(14, min(28, len(str(column_name)) + 3))
            if column_name.endswith("_Rate"):
                number_format = "0.0000%"
            elif column_name.endswith("_Count") or "_cnt" in column_name or "_pay" in column_name:
                number_format = "#,##0"
            elif column_name.endswith("_Frames"):
                number_format = "0.0000"
            else:
                continue
            for row in range(2, len(multiplier_df) + 2):
                worksheet.cell(row=row, column=column_index).number_format = number_format
        for sheet_name in ("Overview", "Base Info"):
            worksheet = writer.sheets[sheet_name]
            worksheet.freeze_panes = "A2"
            worksheet.column_dimensions["A"].width = 28
            worksheet.column_dimensions["B"].width = 44
        feature_sheet = writer.sheets["Feature"]
        for column in ("C", "G"):
            feature_sheet[f"{column}2"].number_format = "0.0000%"
        if "Card Selection" in writer.sheets:
            card_sheet = writer.sheets["Card Selection"]
            card_sheet.freeze_panes = "A2"
            for column in ("J", "L"):
                for row in range(2, len(card_df) + 2):
                    card_sheet[f"{column}{row}"].number_format = "0.0000%"
    return path


def format_board(board: list[list[int]]) -> str:
    return "\n".join(" | ".join(f"{SYMBOL_STR[board[reel][row]]:>3}" for reel in range(5)) for row in range(3, -1, -1))


def run_single_spin_debug() -> None:
    game = LuckyAce(CFG, RNG_SEED, False, CARD_SYSTEM_IS_NEWBIE)
    spin = game.spin("bg_high")
    print("\nInitial board:\n" + format_board(spin.initial_board))
    print("\nFinal board:\n" + format_board(spin.final_board))
    print(f"pay={spin.pay}, cascades={spin.cascades}, scatter={spin.scatter_count}, max=x{spin.max_multiplier}")


BATCH_REQUIRED_FIELDS = (
    "config_file",
    "config_rtp_file",
    "bet_mode",
    "total_rounds",
    "card_system_enabled",
    "card_system_is_newbie",
)


def validate_batch_run(combo: dict[str, Any]) -> dict[str, Any]:
    missing = [field for field in BATCH_REQUIRED_FIELDS if field not in combo]
    if missing:
        raise ValueError(f"BATCH_RUNS entry missing required fields: {', '.join(missing)}")
    normalized = {field: combo[field] for field in BATCH_REQUIRED_FIELDS}
    for field in ("config_file", "config_rtp_file"):
        if not isinstance(normalized[field], str) or not normalized[field].strip():
            raise TypeError(f"{field} must be a non-empty string")
    if isinstance(normalized["bet_mode"], bool) or not isinstance(normalized["bet_mode"], int):
        raise TypeError("bet_mode must be an integer")
    if normalized["bet_mode"] not in SUPPORTED_BET_MODES:
        raise ValueError(f"Unsupported bet_mode: {normalized['bet_mode']}")
    if isinstance(normalized["total_rounds"], bool) or not isinstance(normalized["total_rounds"], int):
        raise TypeError("total_rounds must be a positive integer")
    if normalized["total_rounds"] <= 0:
        raise ValueError("total_rounds must be a positive integer")
    for field in ("card_system_enabled", "card_system_is_newbie"):
        if not isinstance(normalized[field], bool):
            raise TypeError(f"{field} must be Boolean")
    return normalized


def run_all_combinations() -> None:
    global CONFIG_FILE, CONFIG_RTP_FILE, CFG_NATURAL, CFG_RTP, CFG
    global CARD_SYSTEM_ENABLED, CARD_SYSTEM_IS_NEWBIE
    global CARD_RETRY_LIMIT, BET_MODE, TOTAL_ROUNDS
    global GAME_ID, PARSHEET_ID, GAME_NAME, MATH_VERSION, SYMBOL_STR

    for index, combo in enumerate(BATCH_RUNS, 1):
        batch = validate_batch_run(combo)
        print(f"\n=== Batch {index}/{len(BATCH_RUNS)}: {batch} ===", flush=True)
        CONFIG_FILE = batch["config_file"]
        CONFIG_RTP_FILE = batch["config_rtp_file"]
        CFG_NATURAL = _load_config(BASE_DIR / CONFIG_FILE)
        CFG_RTP = CFG_NATURAL if Path(CONFIG_RTP_FILE) == Path(CONFIG_FILE) else _load_config(BASE_DIR / CONFIG_RTP_FILE)
        validate_config_pair(CFG_NATURAL, CFG_RTP, CONFIG_FILE, CONFIG_RTP_FILE)
        CFG = compose_runtime_config(CFG_NATURAL, CFG_RTP)
        BET_MODE = batch["bet_mode"]
        TOTAL_ROUNDS = batch["total_rounds"]
        CARD_SYSTEM_IS_NEWBIE = batch["card_system_is_newbie"]
        CARD_SYSTEM_ENABLED = bool(batch["card_system_enabled"] and CFG.get("card_system", {}).get("enabled"))
        CARD_RETRY_LIMIT = max(1, int(CFG.get("card_system", {}).get("retry_limit", CARD_RETRY_LIMIT)))
        GAME_ID = str(CFG.get("game_id", "H016"))
        PARSHEET_ID = str(CFG.get("parsheet_id", "H016192"))
        GAME_NAME = str(CFG.get("name_zh", "幸運王牌"))
        MATH_VERSION = str(CFG_RTP.get("excel_version", CFG.get("excel_version", "")))
        SYMBOL_STR = {int(key): value for key, value in CFG["symbol_names"].items()}

        result = run_simulation(
            TOTAL_ROUNDS,
            BET_MODE,
            BET_MULTI,
            THREADS,
            config=CFG,
            seed=RNG_SEED,
        )
        print_console(result)
        if OUTPUT_REPORT:
            print(f"\nReport: {output_report(result)}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="H016 幸運王牌 simulator")
    parser.add_argument("rounds", nargs="?", type=int)
    parser.add_argument("bet_mode", nargs="?", type=int, choices=SUPPORTED_BET_MODES)
    parser.add_argument("--bet-multi", type=int, default=BET_MULTI)
    parser.add_argument("--threads", type=int, default=THREADS)
    parser.add_argument("--seed", type=int, default=RNG_SEED)
    parser.add_argument("--no-report", action="store_true")
    parser.add_argument("--debug-spin", action="store_true")
    # Notebook kernels append their own command-line flags. H015-style IDE and
    # Notebook execution should ignore those unrelated arguments.
    return parser.parse_known_args()[0]


def main() -> None:
    args = parse_args()
    explicit_single_run = args.rounds is not None or args.bet_mode is not None or args.debug_spin
    if RUN_ALL_COMBINATIONS and not explicit_single_run and os.environ.get("H016_BATCH_CHILD") != "1":
        run_all_combinations()
        return
    if RUN_SINGLE_SPIN_DEBUG or args.debug_spin:
        run_single_spin_debug()
        return
    result = run_simulation(
        args.rounds or TOTAL_ROUNDS,
        BET_MODE if args.bet_mode is None else args.bet_mode,
        args.bet_multi,
        args.threads,
        seed=args.seed,
    )
    print_console(result)
    if OUTPUT_REPORT and not args.no_report:
        print(f"\nReport: {output_report(result)}")


if __name__ == "__main__":
    main()
