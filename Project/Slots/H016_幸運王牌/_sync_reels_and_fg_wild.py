from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


TARGET = Path(__file__).resolve().parent / "Source" / "H0161.xlsx"
REEL_SOURCES = {
    "BG_Symbol": ("BG_Symbol (2)", "BG_Symbol (3)"),
    "FG_Symbol": ("FG_Symbol (2)", "FG_Symbol (3)"),
}
FG_RANDOM_WILD = {
    "FG_Symbol": [28956, 2000, 500, 200],
    "FG_Symbol (2)": [13128, 2000, 500, 200],
    "FG_Symbol (3)": [1, 0, 0, 0],
}
BG3_SELECTED_STOPS = {0: 31, 1: 47, 2: 1}


def main() -> None:
    workbook = load_workbook(TARGET, data_only=False)
    for source_name, target_names in REEL_SOURCES.items():
        source = workbook[source_name]
        for target_name in target_names:
            target = workbook[target_name]
            for row in range(4, 404):
                for column in range(11, 16):
                    target.cell(row, column).value = source.cell(row, column).value

    for sheet_name, weights in FG_RANDOM_WILD.items():
        worksheet = workbook[sheet_name]
        for row, weight in enumerate(weights, start=4):
            worksheet.cell(row, 30).value = weight

    bg3 = workbook["BG_Symbol (3)"]
    bg1 = workbook["BG_Symbol"]
    for reel in range(5):
        for stop in range(200):
            weight = 1 if reel >= 3 else int(stop == BG3_SELECTED_STOPS[reel])
            bg3.cell(4 + stop, 23 + reel).value = weight
    for row in range(4, 23):
        for column in range(33, 38):
            bg3.cell(row, column).value = bg1.cell(row, column).value

    parameter = workbook["Parameter"]
    for row, weight in zip((11, 12, 13), (1, 1, 0)):
        parameter.cell(row, 3).value = weight
    for row, weight in zip((18, 19, 20), (1, 1, 0)):
        parameter.cell(row, 3).value = weight

    workbook.save(TARGET)
    workbook.close()

    checked = load_workbook(TARGET, read_only=False, data_only=True)
    result: dict[str, object] = {"file": str(TARGET), "reels": {}, "random_wild": {}}
    for sheet_name in [
        "BG_Symbol", "BG_Symbol (2)", "BG_Symbol (3)",
        "FG_Symbol", "FG_Symbol (2)", "FG_Symbol (3)",
    ]:
        worksheet = checked[sheet_name]
        lengths = [
            sum(worksheet.cell(row, column).value not in (None, "") for row in range(4, 404))
            for column in range(11, 16)
        ]
        if lengths != [200] * 5:
            raise ValueError(f"{sheet_name}: reel lengths are {lengths}, expected 5 x 200")
        result["reels"][sheet_name] = lengths
        result["random_wild"][sheet_name] = [worksheet.cell(row, 30).value for row in range(4, 8)]

    for source_name, target_names in REEL_SOURCES.items():
        source = checked[source_name]
        source_values = [
            [source.cell(row, column).value for row in range(4, 204)]
            for column in range(11, 16)
        ]
        for target_name in target_names:
            target = checked[target_name]
            target_values = [
                [target.cell(row, column).value for row in range(4, 204)]
                for column in range(11, 16)
            ]
            if target_values != source_values:
                raise ValueError(f"{target_name}: physical reels differ from {source_name}")

    result["free_table_weights"] = [checked["Parameter"].cell(row, 3).value for row in (11, 12, 13)]
    result["retrigger_table_weights"] = [checked["Parameter"].cell(row, 3).value for row in (18, 19, 20)]
    checked.close()
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
