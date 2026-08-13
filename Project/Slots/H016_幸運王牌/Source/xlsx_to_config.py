from __future__ import annotations

import argparse
import importlib.util
import json
import os
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
PROJECT_DIR = HERE.parent
DEFAULT_SOURCE = HERE / "H0161.xlsx"
DEFAULT_OUTPUT = PROJECT_DIR / "config_92.js"
FRONTEND_NAMES = {
    "0": "WW1", "1": "WW2", "2": "C1",
    "3": "M1", "4": "M2", "5": "M3", "6": "M4",
    "7": "A", "8": "K", "9": "Q", "10": "J",
    "11": "M1G", "12": "M2G", "13": "M3G", "14": "M4G",
    "15": "AG", "16": "KG", "17": "QG", "18": "JG",
}


def load_simulator_module():
    os.environ.setdefault("H016_BASE_DIR", str(PROJECT_DIR))
    os.environ.setdefault("H016_RUN_ALL_COMBINATIONS", "false")
    spec = importlib.util.spec_from_file_location("h016_simulator_for_config", PROJECT_DIR / "Simulator.py")
    if spec is None or spec.loader is None:
        raise RuntimeError("Cannot load Simulator.py")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _parse_range(label: str) -> tuple[float, float] | None:
    match = re.fullmatch(r"\(\s*(-?[0-9.]+)\s*,\s*([0-9.]+)\s*\]", label)
    return (float(match.group(1)), float(match.group(2))) if match else None


def load_card_system(variant: Path) -> dict[str, Any]:
    workbook = load_workbook(variant, read_only=True, data_only=True)
    try:
        sheet = workbook["Multiplier_Weight"]
        headers = {str(sheet.cell(2, column).value or "").strip(): column for column in range(2, sheet.max_column + 1)}
        required = {
            "Weight_NB_BG_Newbie", "Weight_NB_FG_Newbie", "Weight_NB_BG",
            "Weight_NB_FG", "Weight_BF", "Weight_SF",
        }
        missing = required - set(headers)
        if missing:
            raise ValueError(f"{variant.name}: missing Multiplier_Weight columns {sorted(missing)}")

        ranges = []
        free_game_row = None
        for row in range(3, sheet.max_row + 1):
            label = str(sheet.cell(row, 1).value or "").strip()
            parsed = _parse_range(label)
            if parsed is not None:
                ranges.append((row, parsed))
            elif label.lower() == "free game":
                free_game_row = row
        if len(ranges) != 64 or free_game_row is None:
            raise ValueError(f"{variant.name}: expected 64 ranges plus Free Game")

        def range_cards(header: str, table: str) -> list[dict[str, Any]]:
            column = headers[header]
            return [
                {
                    "type": "range", "min": minimum, "max": maximum, "table": table,
                    "weight": int(sheet.cell(row, column).value or 0),
                }
                for row, (minimum, maximum) in ranges
            ]

        def base_cards(header: str) -> list[dict[str, Any]]:
            cards = range_cards(header, "B")
            cards.append({
                "type": "free_game", "table": "A",
                "weight": int(sheet.cell(free_game_row, headers[header]).value or 0),
            })
            return cards

        buy = range_cards("Weight_BF", "E")
        super_cards = range_cards("Weight_SF", "G")
        profiles = {
            "weight_1": {
                "base_game": base_cards("Weight_NB_BG_Newbie"),
                "free_game": range_cards("Weight_NB_FG_Newbie", "E"),
                "buy_feature": [dict(card) for card in buy],
                "super_feature": [dict(card) for card in super_cards],
            },
            "weight_2": {
                "base_game": base_cards("Weight_NB_BG"),
                "free_game": range_cards("Weight_NB_FG", "E"),
                "buy_feature": buy,
                "super_feature": super_cards,
            },
        }
        for profile in profiles.values():
            for section, cards in profile.items():
                if sum(card["weight"] for card in cards) != 1_000_000_000:
                    raise ValueError(f"{variant.name}: {section} weights do not sum to 1,000,000,000")
        return {"enabled": True, "retry_limit": 200_000, "default_profile": "weight_2", "profiles": profiles}
    finally:
        workbook.close()


def frontend_config(source: Path, variant: Path | None = None) -> dict[str, Any]:
    simulator = load_simulator_module()
    config = simulator._load_xlsx_config(source)
    config["excel_version"] = str(config.get("excel_version") or config.get("version") or "1.0.0.0")
    config.pop("version", None)
    config["name_en"] = "Lucky Ace"
    rtp_label = int(variant.stem[-2:]) if variant is not None and variant.stem[-2:].isdigit() else 92
    config["parsheet_id"] = f"H0161{rtp_label}"
    config["rtp_label"] = rtp_label
    config["reel_num"] = 5
    config["window_size"] = 4
    config["max_ways"] = 1024
    config["symbol_names"] = FRONTEND_NAMES
    config["base_table_weights"] = {"high": 1, "low": 0}
    config["card_system"] = load_card_system(variant) if variant is not None else {"enabled": False, "profiles": {}}
    if variant is not None:
        config["source_xlsx"] = variant.name
    for table in config["tables"].values():
        normalized = []
        for reel in table["weights"]:
            if any(float(weight) != int(weight) for weight in reel):
                raise ValueError("H0161.xlsx contains a non-integer reel weight")
            normalized.append([int(weight) for weight in reel])
        table["weights"] = normalized
        normalized_drop = []
        for reel in table["drop_weights"]:
            if any(float(weight) != int(weight) for weight in reel):
                raise ValueError("H0161.xlsx contains a non-integer Symbol Drop Weight")
            normalized_drop.append([int(weight) for weight in reel])
        table["drop_weights"] = normalized_drop
        random_weights = table["random_wild"]["weights"]
        if any(float(weight) != int(weight) for weight in random_weights):
            raise ValueError("H0161.xlsx contains a non-integer Random Wild weight")
        table["random_wild"]["weights"] = [int(weight) for weight in random_weights]
    return config


def validate(config: dict[str, Any]) -> dict[str, Any]:
    tables = config["tables"]
    bg = tables["bg_high"]
    fg = tables["fg_high_a"]
    gold_ids = set(range(11, 19))

    def gold_counts(table: dict[str, Any]) -> list[int]:
        return [sum(symbol in gold_ids for symbol in reel) for reel in table["reels"]]

    for table in tables.values():
        if [len(reel) for reel in table["reels"]] != [200] * 5:
            raise ValueError("Every frontend reel must contain exactly 200 stops")
        if any(any(type(weight) is not int or weight < 0 for weight in reel) for reel in table["weights"]):
            raise ValueError("Every frontend initial stop weight must be a non-negative integer")
        if any(sum(reel) <= 0 for reel in table["weights"]):
            raise ValueError("Every frontend initial stop-weight reel must have a positive total")
        if any(any(type(weight) is not int or weight < 0 for weight in reel) for reel in table["drop_weights"]):
            raise ValueError("Every frontend Symbol Drop Weight must be a non-negative integer")
        if any(sum(reel) <= 0 for reel in table["drop_weights"]):
            raise ValueError("Every frontend Symbol Drop Weight reel must have a positive total")
    result = {
        "bg_lengths": [len(reel) for reel in bg["reels"]],
        "fg_lengths": [len(reel) for reel in fg["reels"]],
        "bg_gold_counts": gold_counts(bg),
        "fg_gold_counts": gold_counts(fg),
        "bg_random_wild": bg["random_wild"],
        "fg_random_wild": fg["random_wild"],
        "bg_drop_weight_sums": [sum(reel) for reel in bg["drop_weights"]],
        "fg_drop_weight_sums": [sum(reel) for reel in fg["drop_weights"]],
    }
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert H0161.xlsx to the index frontend JS config or pure JSON")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--variant", type=Path, help="H016192/H016194 multiplier workbook")
    args = parser.parse_args()
    source = args.source.resolve()
    output = args.output.resolve()
    variant = args.variant.resolve() if args.variant else HERE / f"H0161{output.stem[-2:]}.xlsx"
    if not variant.is_file():
        variant = None
    config = frontend_config(source, variant)
    result = validate(config)
    if output.suffix.lower() == ".json":
        output.write_text(json.dumps(config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    else:
        payload = json.dumps(config, ensure_ascii=False, separators=(",", ":"))
        output.write_text(
            "// Generated from Source/H0161.xlsx by Source/xlsx_to_config.py.\n"
            f"window.H016_CONFIG={payload};\n",
            encoding="utf-8",
        )
    result.update({"source": str(source), "output": str(output)})
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
