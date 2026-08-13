from __future__ import annotations

import argparse
import importlib.util
import json
import os
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
PROJECT = HERE.parent
DEFAULT_CONFIG = PROJECT / "config_92.js"
DEFAULT_XLSX = HERE / "H0161.xlsx"

SHEET_TABLES = {
    "BG_Symbol": "bg_1",
    "BG_Symbol (2)": "bg_2",
    "BG_Symbol (3)": "bg_3",
    "FG_Symbol": "fg_1",
    "FG_Symbol (2)": "fg_2",
    "FG_Symbol (3)": "fg_3",
}
MASTER_SHEETS = {"BG_Symbol", "FG_Symbol"}


def load_config(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8-sig")
    match = re.search(r"\{.*\}", text, flags=re.DOTALL)
    if not match:
        raise ValueError(f"Cannot find JSON payload in {path}")
    return json.loads(match.group(0))


def integer(value: Any, label: str, *, allow_zero: bool = True) -> int:
    if type(value) is not int:
        raise ValueError(f"{label} must be an integer, got {value!r}")
    if value < 0 or (not allow_zero and value == 0):
        raise ValueError(f"{label} has invalid value {value}")
    return value


def validate_table(table: dict[str, Any], table_name: str) -> None:
    for key in ("reels", "weights", "drop_values", "drop_weights"):
        if len(table[key]) != 5:
            raise ValueError(f"{table_name}.{key} must define five reels")
    for reel, (symbols, weights) in enumerate(zip(table["reels"], table["weights"]), start=1):
        if len(symbols) != 200 or len(weights) != 200:
            raise ValueError(f"{table_name} R{reel} must contain exactly 200 stops")
        for index, symbol in enumerate(symbols):
            integer(symbol, f"{table_name}.reels[{reel}][{index}]")
        for index, weight in enumerate(weights):
            integer(weight, f"{table_name}.weights[{reel}][{index}]")
        positive_weights = [weight for weight in weights if weight > 0]
        if not positive_weights:
            raise ValueError(f"{table_name} R{reel} must have at least one enabled stop")
        if max(positive_weights) / min(positive_weights) > 10 + 1e-12:
            raise ValueError(f"{table_name} R{reel} stop-weight ratio exceeds 10x")
    for reel, (values, weights) in enumerate(zip(table["drop_values"], table["drop_weights"]), start=1):
        if len(values) != 19 or len(weights) != 19:
            raise ValueError(f"{table_name} drop R{reel} must contain 19 symbols")
        for index, value in enumerate(values):
            integer(value, f"{table_name}.drop_values[{reel}][{index}]")
        for index, weight in enumerate(weights):
            integer(weight, f"{table_name}.drop_weights[{reel}][{index}]")
        if sum(weights) <= 0:
            raise ValueError(f"{table_name} drop R{reel} must have positive total weight")
    random_wild = table["random_wild"]
    if random_wild["values"] != [0, 2, 3, 4]:
        raise ValueError(f"{table_name} Random Wild values must be 0/2/3/4")
    for index, weight in enumerate(random_wild["weights"]):
        integer(weight, f"{table_name}.random_wild.weights[{index}]")


def write_table(ws, table: dict[str, Any], id_to_name: dict[int, str], *, write_drop_labels: bool) -> None:
    # Clear the full parser range first, then write the canonical 200-stop data.
    for row in range(4, 404):
        for column in (*range(11, 16), *range(23, 28)):
            ws.cell(row, column).value = None
    for reel in range(5):
        for offset, (symbol, weight) in enumerate(zip(table["reels"][reel], table["weights"][reel])):
            row = 4 + offset
            ws.cell(row, 11 + reel).value = id_to_name[int(symbol)]
            ws.cell(row, 23 + reel).value = int(weight)

    for offset, (value, weight) in enumerate(
        zip(table["random_wild"]["values"], table["random_wild"]["weights"])
    ):
        row = 4 + offset
        ws.cell(row, 29).value = int(value)
        ws.cell(row, 30).value = int(weight)

    if write_drop_labels:
        reference_values = table["drop_values"][0]
        if any(values != reference_values for values in table["drop_values"]):
            raise ValueError(f"{ws.title}: drop symbol order must be the same on R1-R5")
        for offset, symbol in enumerate(reference_values):
            ws.cell(4 + offset, 32).value = id_to_name[int(symbol)]
    for reel in range(5):
        for offset, weight in enumerate(table["drop_weights"][reel]):
            ws.cell(4 + offset, 33 + reel).value = int(weight)


def load_converter():
    path = HERE / "xlsx_to_config.py"
    module_name = "h016_xlsx_to_config_roundtrip"
    spec = importlib.util.spec_from_file_location(module_name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    try:
        spec.loader.exec_module(module)
        return module
    finally:
        sys.modules.pop(module_name, None)


def assert_roundtrip(expected: dict[str, Any], xlsx_path: Path) -> None:
    converter = load_converter()
    variant_name = str(expected.get("source_xlsx") or "")
    variant = DEFAULT_XLSX.parent / variant_name if variant_name else None
    actual = converter.frontend_config(xlsx_path, variant if variant and variant.is_file() else None)
    for table_name in expected["tables"]:
        if actual["tables"].get(table_name) != expected["tables"][table_name]:
            raise ValueError(f"Round-trip mismatch in table {table_name}")
    if actual["table_selection"] != expected["table_selection"]:
        raise ValueError("Round-trip mismatch in table_selection")
    if actual["pays"] != expected["pays"]:
        raise ValueError("Round-trip mismatch in pays")


def main() -> None:
    parser = argparse.ArgumentParser(description="Write H016 config_92.js tables back to H0161.xlsx")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    args = parser.parse_args()

    config_path = args.config.resolve()
    xlsx_path = args.xlsx.resolve()
    config = load_config(config_path)
    for table_name in SHEET_TABLES.values():
        validate_table(config["tables"][table_name], table_name)

    workbook = load_workbook(xlsx_path, read_only=False, data_only=False, keep_links=True)
    missing = set(SHEET_TABLES).difference(workbook.sheetnames)
    if missing:
        raise ValueError(f"Workbook is missing sheets: {sorted(missing)}")
    overview = workbook["Overview"]
    id_to_name = {
        int(overview.cell(row, 8).value): str(overview.cell(row, 1).value)
        for row in range(29, 48)
        if overview.cell(row, 1).value not in (None, "")
        and overview.cell(row, 8).value not in (None, "")
    }
    if set(id_to_name) != set(range(19)):
        raise ValueError("Overview symbol ids must be exactly 0..18")
    for sheet_name, table_name in SHEET_TABLES.items():
        write_table(
            workbook[sheet_name],
            config["tables"][table_name],
            id_to_name,
            write_drop_labels=sheet_name in MASTER_SHEETS,
        )

    parameter = workbook["Parameter"]
    selection_rows = {"base": (4, 5, 6), "free": (11, 12, 13), "retrigger": (18, 19, 20)}
    for group, rows in selection_rows.items():
        selections = config["table_selection"][group]
        if len(selections) != 3:
            raise ValueError(f"table_selection.{group} must contain three rows")
        for row, item in zip(rows, selections):
            parameter.cell(row, 3).value = integer(item["weight"], f"table_selection.{group}.{item['table']}")

    if workbook.calculation is not None:
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True

    temporary = xlsx_path.with_name(xlsx_path.stem + ".config-sync.tmp" + xlsx_path.suffix)
    try:
        workbook.save(temporary)
        workbook.close()
        assert_roundtrip(config, temporary)
        os.replace(temporary, xlsx_path)
    finally:
        workbook.close()
        if temporary.exists():
            temporary.unlink()

    print(json.dumps({
        "xlsx": str(xlsx_path),
        "config": str(config_path),
        "sheets_written": list(SHEET_TABLES),
        "roundtrip": "PASS",
        "table_selection": config["table_selection"],
        "bg_random_wild": config["tables"]["bg_1"]["random_wild"],
        "fg_random_wild": config["tables"]["fg_1"]["random_wild"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
