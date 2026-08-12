from __future__ import annotations

import argparse
import bisect
import json
import math
import random
import statistics
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
DEFAULT_TARGET = HERE / "H0161.xlsx"
STOP_WEIGHT_PROFILE = HERE / "initial_stop_weights.json"
DEFAULT_SOURCE = Path(
    "C:/Users/rhinshen/Mine/個人工作區/市場資訊/H5/遊戲資源/JILI/"
    "JILI - Super Ace/遊戲資料/StripTable_SuperAce_還原.xlsx"
)
TARGETS = {"BG_Symbol": "BG_Strip", "FG_Symbol": "FG_Strip"}
VARIANT_REEL_TARGETS = {
    "BG_Symbol": ("BG_Symbol (2)", "BG_Symbol (3)"),
    "FG_Symbol": ("FG_Symbol (2)", "FG_Symbol (3)"),
}
COMPETITOR_JSONS = (
    "SuperAce_BG_Combined_NoJP.jsonl",
    "SuperAce_BG_3.jsonl",
    "Super_Ace_BG_4.jsonl",
)
COMPETITOR_SYMBOL_MAP = {
    "Bonus": "C1", "Symbol1": "M1", "Symbol2": "M2", "Symbol3": "M3",
    "Symbol4": "M4", "Symbol5": "A", "Symbol6": "K", "Symbol7": "Q", "Symbol8": "J",
}
EDITABLE_COLUMNS = {*range(11, 16), *range(23, 28), 30, *range(32, 38)}  # K:O, W:AA, AD, AF:AK
REEL_LENGTH = 200
GOLD_RATES = {
    "BG_Symbol": [0.0, 0.15023788015657935, 0.11346281240590184, 0.07600421559771153, 0.0],
    "FG_Symbol": [0.0, 0.21209989447766445, 0.18279985930355258, 0.15239183960604993, 0.0],
}
RANDOM_WILD_WEIGHTS = {
    "BG_Symbol": [34600, 1401, 235, 18],
    # Preserve 101003's conditional 2/3/4 distribution while halving its
    # non-zero draw probability: 2700 / 31656 = (2700 / 15828) / 2.
    "FG_Symbol": [28956, 2000, 500, 200],
}
FG_VARIANT_RANDOM_WILD_WEIGHTS = {
    "FG_Symbol (2)": [13128, 2000, 500, 200],
    "FG_Symbol (3)": [1, 0, 0, 0],
}
GOLD_SYMBOLS = {
    "M1": "G1", "M2": "G2", "M3": "G3", "M4": "G4",
    "A": "GA", "K": "GK", "Q": "GQ", "J": "GJ",
}
BASE_SYMBOLS = {gold: symbol for symbol, gold in GOLD_SYMBOLS.items()}
MIN_SC_GAP = 4
MAX_SYMBOL_STACK = 4
SCORE_SYMBOLS = ("M1", "M2", "M3", "M4", "A", "K", "Q", "J")
SCORE_BITS = {symbol: 1 << index for index, symbol in enumerate(SCORE_SYMBOLS)}
INITIAL_HIT_TARGETS = {"BG_Symbol": 0.237699, "FG_Symbol": 0.405628}
INITIAL_SCATTER_FACTORS = {"BG_Symbol": -0.245, "FG_Symbol": 0.0}
STOP_WEIGHT_BASE = 100
CALIBRATION_SEEDS = {"BG_Symbol": 16016, "FG_Symbol": 16017}
CALIBRATION_TRIALS = 50_000
DROP_WEIGHT_TOTAL = 1_000_000
DROP_GOLD_RATES = {
    "BG_Symbol": [0.0, 0.153181, 0.111969, 0.079140, 0.0],
    "FG_Symbol": [0.0, 0.158077, 0.127544, 0.098582, 0.0],
}
EXACT_DROP_PERCENTAGES = {
    "BG_Symbol": [
        [1.9997, 4.9120, 9.8476, 9.8374, 9.4918, 14.5730, 14.7668, 17.1100, 17.4617],
        [1.9703, 10.0272, 4.8583, 7.1978, 12.5699, 17.3589, 16.7137, 14.6561, 14.6478],
        [2.1121, 5.1527, 10.0534, 7.9044, 7.5098, 15.5995, 15.2210, 18.2615, 18.1856],
        [2.0405, 9.9382, 4.9609, 7.1703, 12.1557, 16.9259, 17.2746, 14.6920, 14.8419],
        [1.9876, 4.9449, 10.2715, 8.0899, 7.6415, 15.3921, 16.3011, 17.3070, 18.0645],
    ],
    "FG_Symbol": [
        [0.7756, 4.9351, 9.9334, 10.0655, 9.8070, 15.0063, 15.0580, 17.3331, 17.0861],
        [0.8222, 9.6449, 4.7741, 7.4340, 12.4845, 17.6316, 17.4934, 14.6815, 15.0339],
        [1.0372, 5.1792, 9.9383, 7.4373, 7.5358, 16.5026, 15.1963, 18.9379, 18.2355],
        [0.8936, 9.9149, 5.0922, 7.3901, 12.3404, 17.6312, 16.9078, 15.1206, 14.7092],
        [1.0577, 4.8353, 10.5470, 7.9480, 7.9480, 15.3521, 14.3850, 19.0692, 18.8577],
    ],
}
DROP_SYMBOLS = ("WW", "W2", "C1", *SCORE_SYMBOLS, "G1", "G2", "G3", "G4", "GA", "GK", "GQ", "GJ")


def stable_value(value: Any) -> Any:
    if hasattr(value, "text"):
        return (value.text, value.ref)
    return value


def protected_values(workbook) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for worksheet in workbook.worksheets:
        cells: dict[str, Any] = {}
        for row in worksheet.iter_rows():
            for cell in row:
                variant_reel = worksheet.title in {
                    name for names in VARIANT_REEL_TARGETS.values() for name in names
                } and cell.column in range(11, 16)
                variant_random_wild = (
                    worksheet.title in FG_VARIANT_RANDOM_WILD_WEIGHTS
                    and cell.column == 30
                    and cell.row in range(4, 8)
                )
                if (worksheet.title in TARGETS and cell.column in EDITABLE_COLUMNS) or variant_reel or variant_random_wild:
                    continue
                if cell.value is not None:
                    cells[cell.coordinate] = stable_value(cell.value)
        result[worksheet.title] = cells
    return result


def read_reels(worksheet) -> tuple[list[list[str]], list[list[float]]]:
    reels: list[list[str]] = [[] for _ in range(5)]
    weights: list[list[float]] = [[] for _ in range(5)]
    for row in worksheet.iter_rows(min_row=3, max_col=12, values_only=True):
        for reel in range(5):
            symbol = row[1 + reel]
            stop_weight = row[7 + reel]
            if symbol in (None, ""):
                continue
            weight = float(stop_weight)
            if weight <= 0:
                raise ValueError(f"{worksheet.title} R{reel + 1}: stop weight must be positive")
            reels[reel].append(str(symbol))
            weights[reel].append(weight)
    if any(not reel for reel in reels):
        raise ValueError(f"{worksheet.title}: all five reels must be non-empty")
    return reels, weights


def resample_to_200(symbols: list[str], weights: list[float]) -> list[str]:
    cumulative: list[float] = []
    running = 0.0
    for weight in weights:
        running += weight
        cumulative.append(running)
    return [
        symbols[min(bisect.bisect_left(cumulative, running * (index + 0.5) / REEL_LENGTH), len(symbols) - 1)]
        for index in range(REEL_LENGTH)
    ]


def proportional_gold_counts(symbols: list[str], total_gold: int) -> dict[str, int]:
    counts = Counter(symbol for symbol in symbols if symbol in GOLD_SYMBOLS)
    eligible = sum(counts.values())
    if total_gold > eligible:
        raise ValueError(f"gold target {total_gold} exceeds {eligible} eligible symbols")
    exact = {symbol: total_gold * count / eligible for symbol, count in counts.items()}
    allocated = {symbol: int(value) for symbol, value in exact.items()}
    remainder = total_gold - sum(allocated.values())
    order = sorted(counts, key=lambda symbol: (exact[symbol] - allocated[symbol], counts[symbol], symbol), reverse=True)
    for symbol in order[:remainder]:
        allocated[symbol] += 1
    return allocated


def evenly_spaced(items: list[int], count: int) -> list[int]:
    if count <= 0:
        return []
    return [items[min(int((index + 0.5) * len(items) / count), len(items) - 1)] for index in range(count)]


def apply_gold(symbols: list[str], rate: float, gold_profile: dict[str, int]) -> tuple[list[str], int]:
    result = list(symbols)
    target = round(REEL_LENGTH * rate)
    by_symbol: dict[str, list[int]] = {
        symbol: [index for index, value in enumerate(result) if value == symbol]
        for symbol in GOLD_SYMBOLS
    }
    allocation = (
        largest_remainder({symbol: gold_profile[symbol] for symbol in SCORE_SYMBOLS}, target)
        if target else {symbol: 0 for symbol in SCORE_SYMBOLS}
    )
    for symbol, count in allocation.items():
        if count > len(by_symbol[symbol]):
            raise ValueError(f"{symbol}: gold target {count} exceeds {len(by_symbol[symbol])} stops")
        for index in evenly_spaced(by_symbol[symbol], count):
            result[index] = GOLD_SYMBOLS[symbol]
    return result, target


def base_symbol(symbol: str) -> str:
    return BASE_SYMBOLS.get(symbol, symbol)


def cyclic_runs(symbols: list[str]) -> list[list[int]]:
    """Return canonical-symbol runs, treating the reel as a closed loop."""
    size = len(symbols)
    canonical = [base_symbol(symbol) for symbol in symbols]
    if len(set(canonical)) == 1:
        return [list(range(size))]
    start = next(index for index in range(size) if canonical[index] != canonical[index - 1])
    runs: list[list[int]] = []
    current: list[int] = []
    current_symbol: str | None = None
    for offset in range(size):
        index = (start + offset) % size
        if canonical[index] != current_symbol:
            if current:
                runs.append(current)
            current = [index]
            current_symbol = canonical[index]
        else:
            current.append(index)
    if current:
        runs.append(current)
    return runs


def reel_constraint_state(symbols: list[str]) -> tuple[tuple[int, int, int], list[int]]:
    """Return a sortable penalty and positions that directly violate constraints."""
    size = len(symbols)
    scatter_positions = [index for index, symbol in enumerate(symbols) if symbol == "C1"]
    sc_penalty = 0
    violating: set[int] = set()
    if scatter_positions:
        for offset, left in enumerate(scatter_positions):
            right = scatter_positions[(offset + 1) % len(scatter_positions)]
            gap = (right - left - 1) % size
            if gap < MIN_SC_GAP:
                sc_penalty += MIN_SC_GAP - gap
                violating.add(right)

    stack_penalty = 0
    for run in cyclic_runs(symbols):
        if len(run) > MAX_SYMBOL_STACK:
            stack_penalty += len(run) - MAX_SYMBOL_STACK
            violating.update(run[MAX_SYMBOL_STACK:])

    return (sc_penalty + stack_penalty, sc_penalty, stack_penalty), sorted(violating)


def repair_reel_constraints(symbols: list[str]) -> tuple[list[str], list[tuple[int, int]]]:
    """Repair constraints with deterministic, minimum-penalty swaps within one reel."""
    result = list(symbols)
    swaps: list[tuple[int, int]] = []
    for _ in range(len(result) * 2):
        penalty, violating = reel_constraint_state(result)
        if penalty[0] == 0:
            return result, swaps

        best: tuple[tuple[int, int, int], int, int, int] | None = None
        for left in violating:
            for right in range(len(result)):
                if left == right or result[left] == result[right]:
                    continue
                result[left], result[right] = result[right], result[left]
                candidate_penalty, _ = reel_constraint_state(result)
                result[left], result[right] = result[right], result[left]
                if candidate_penalty >= penalty:
                    continue
                distance = min((right - left) % len(result), (left - right) % len(result))
                candidate = (candidate_penalty, distance, left, right)
                if best is None or candidate < best:
                    best = candidate

        if best is None:
            raise RuntimeError(f"Unable to repair reel constraints; remaining penalty={penalty}")
        _, _, left, right = best
        result[left], result[right] = result[right], result[left]
        swaps.append((left, right))

    raise RuntimeError("Reel constraint repair exceeded iteration limit")


def repair_reel_pairs(symbols: list[str], weights: list[float]) -> tuple[list[str], list[float], list[tuple[int, int]]]:
    repaired, swaps = repair_reel_constraints(symbols)
    repaired_weights = list(weights)
    for left, right in swaps:
        repaired_weights[left], repaired_weights[right] = repaired_weights[right], repaired_weights[left]
    return repaired, repaired_weights, swaps


def read_fill_weights(workbook) -> dict[str, list[dict[str, float]]]:
    """Read the competitor's R1補..R5補 canonical-symbol percentages."""
    result = {"BG_Symbol": [dict() for _ in range(5)], "FG_Symbol": [dict() for _ in range(5)]}
    worksheet = workbook["FillWeight"]
    sheet_names = {"BG": "BG_Symbol", "FG": "FG_Symbol"}
    for row in worksheet.iter_rows(min_row=3, max_col=13, values_only=True):
        scene, symbol = row[0], row[1]
        if scene not in sheet_names or symbol not in (*SCORE_SYMBOLS, "C1"):
            continue
        for reel, column in enumerate((4, 6, 8, 10, 12)):
            result[sheet_names[scene]][reel][str(symbol)] = float(row[column])
    for sheet_name, reels in result.items():
        if any(set(reel) != {*SCORE_SYMBOLS, "C1"} for reel in reels):
            raise ValueError(f"FillWeight is incomplete for {sheet_name}")
    return result


def read_competitor_gold_profiles(source_path: Path) -> dict[str, dict[str, list[dict[str, int]]]]:
    """Read joint gold counts by scene/stage/reel/symbol from competitor JSONL."""
    result = {
        sheet: {stage: [Counter() for _ in range(5)] for stage in ("initial", "drop")}
        for sheet in TARGETS
    }
    paths = [source_path.parent / name for name in COMPETITOR_JSONS]
    missing = [path for path in paths if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing competitor JSONL: {missing}")
    for path in paths:
        with path.open(encoding="utf-8") as handle:
            for line in handle:
                obj = json.loads(line)
                for plate_index, plate in enumerate(obj["plate"]["plate"]):
                    sheet = "BG_Symbol" if plate_index == 0 else "FG_Symbol"
                    for reel, column in enumerate(plate["column"]):
                        for raw_symbol, gold in zip(column["row"], column["isGold"]):
                            symbol = COMPETITOR_SYMBOL_MAP[raw_symbol]
                            if symbol in SCORE_SYMBOLS and gold in (1, 2):
                                result[sheet]["initial"][reel][symbol] += 1
                    for combo in plate.get("combo", []):
                        for change in combo.get("change", []):
                            if "symbol" not in change:
                                continue
                            symbol = COMPETITOR_SYMBOL_MAP[change["symbol"]]
                            reel = int(change.get("column", 0))
                            if symbol in SCORE_SYMBOLS and change.get("isGold") in (1, 2):
                                result[sheet]["drop"][reel][symbol] += 1
    for sheet in TARGETS:
        for stage in ("initial", "drop"):
            for reel in range(5):
                if sum(result[sheet][stage][reel].values()) == 0 and reel not in (0, 4):
                    raise ValueError(f"Missing {sheet} {stage} R{reel + 1} gold observations")
                for symbol in SCORE_SYMBOLS:
                    result[sheet][stage][reel].setdefault(symbol, 0)
    return result


def exact_drop_weights(sheet_name: str) -> list[dict[str, float]]:
    symbols = ("C1", *SCORE_SYMBOLS)
    return [dict(zip(symbols, percentages)) for percentages in EXACT_DROP_PERCENTAGES[sheet_name]]


def largest_remainder(values: dict[str, float], total: int) -> dict[str, int]:
    normalized = sum(values.values())
    exact = {key: total * value / normalized for key, value in values.items()}
    allocated = {key: int(value) for key, value in exact.items()}
    for key in sorted(values, key=lambda item: (exact[item] - allocated[item], item), reverse=True)[: total - sum(allocated.values())]:
        allocated[key] += 1
    return allocated


def spread_total(count: int, total: int) -> list[int]:
    if count <= 0:
        if total:
            raise ValueError("Cannot allocate weight to an empty stop group")
        return []
    if total < count:
        raise ValueError(f"Positive integer allocation needs at least {count}, got {total}")
    base, remainder = divmod(total, count)
    result = [base] * count
    for index in evenly_spaced(list(range(count)), remainder):
        result[index] += 1
    return result


def drop_weights(symbols: list[str], percentages: dict[str, float], gold_rate: float) -> list[int]:
    canonical_totals = largest_remainder(percentages, DROP_WEIGHT_TOTAL)
    positions: dict[str, list[int]] = {symbol: [] for symbol in (*SCORE_SYMBOLS, "C1")}
    for index, symbol in enumerate(symbols):
        positions[base_symbol(symbol)].append(index)

    weights = [0] * len(symbols)
    eligible_totals = {symbol: canonical_totals[symbol] for symbol in SCORE_SYMBOLS}
    gold_total = round(DROP_WEIGHT_TOTAL * gold_rate)
    gold_by_symbol = largest_remainder(eligible_totals, gold_total) if gold_total else {symbol: 0 for symbol in SCORE_SYMBOLS}
    for symbol, indexes in positions.items():
        gold_indexes = [index for index in indexes if symbols[index] in BASE_SYMBOLS]
        plain_indexes = [index for index in indexes if index not in set(gold_indexes)]
        symbol_total = canonical_totals[symbol]
        symbol_gold = gold_by_symbol.get(symbol, 0)
        if not gold_indexes:
            symbol_gold = 0
        for index, weight in zip(gold_indexes, spread_total(len(gold_indexes), symbol_gold)):
            weights[index] = weight
        for index, weight in zip(plain_indexes, spread_total(len(plain_indexes), symbol_total - symbol_gold)):
            weights[index] = weight
    if any(type(weight) is not int or weight <= 0 for weight in weights):
        raise ValueError("Every drop stop weight must be a positive integer")
    return weights


def symbol_drop_weights(
    percentages: dict[str, float], gold_rate: float, gold_profile: dict[str, int]
) -> dict[str, int]:
    """Split canonical FillWeight into plain/gold rows in AF:AK."""
    totals = largest_remainder(percentages, DROP_WEIGHT_TOTAL)
    gold_total = round(DROP_WEIGHT_TOTAL * gold_rate)
    gold_by_symbol = (
        largest_remainder({symbol: gold_profile[symbol] for symbol in SCORE_SYMBOLS}, gold_total)
        if gold_total else {symbol: 0 for symbol in SCORE_SYMBOLS}
    )
    result = {symbol: 0 for symbol in DROP_SYMBOLS}
    result["C1"] = totals["C1"]
    for symbol in SCORE_SYMBOLS:
        result[symbol] = totals[symbol] - gold_by_symbol[symbol]
        result[GOLD_SYMBOLS[symbol]] = gold_by_symbol[symbol]
    if sum(result.values()) != DROP_WEIGHT_TOTAL or any(type(weight) is not int or weight < 0 for weight in result.values()):
        raise ValueError("Invalid Symbol Drop Weight allocation")
    return result


def window_mask(symbols: list[str], stop: int) -> int:
    mask = 0
    for offset in range(4):
        symbol = base_symbol(symbols[(stop + offset) % len(symbols)])
        mask |= SCORE_BITS.get(symbol, 0)
    return mask


def initial_hit_rate(reels: list[list[str]]) -> float:
    histograms = [Counter(window_mask(reels[reel], stop) for stop in range(REEL_LENGTH)) for reel in range(3)]
    pair_intersections = Counter()
    for mask1, count1 in histograms[0].items():
        for mask2, count2 in histograms[1].items():
            pair_intersections[mask1 & mask2] += count1 * count2
    wins = sum(
        pair_count * count3
        for pair_mask, pair_count in pair_intersections.items()
        for mask3, count3 in histograms[2].items()
        if pair_mask & mask3
    )
    return wins / REEL_LENGTH**3


def weighted_initial_hit_rate(reels: list[list[str]], weights: list[list[int]]) -> float:
    histograms: list[Counter] = []
    for reel, reel_weights in zip(reels[:3], weights[:3]):
        histogram = Counter()
        for stop, weight in enumerate(reel_weights):
            histogram[window_mask(reel, stop)] += weight
        histograms.append(histogram)
    wins = sum(
        weight1 * weight2 * weight3
        for mask1, weight1 in histograms[0].items()
        for mask2, weight2 in histograms[1].items()
        for mask3, weight3 in histograms[2].items()
        if mask1 & mask2 & mask3
    )
    return wins / math.prod(sum(reel_weights) for reel_weights in weights[:3])


def initial_stop_propensity(reels: list[list[str]]) -> list[list[float]]:
    """Score each R1-R3 stop by its uniform chance of completing a 3-reel win."""
    masks = [[window_mask(reel, stop) for stop in range(REEL_LENGTH)] for reel in reels[:3]]
    histograms = [Counter(values) for values in masks]
    scores: list[list[float]] = []
    for reel in range(3):
        others = [index for index in range(3) if index != reel]
        denominator = REEL_LENGTH**2
        by_mask = {}
        for current_mask in histograms[reel]:
            wins = sum(
                count1 * count2
                for mask1, count1 in histograms[others[0]].items()
                for mask2, count2 in histograms[others[1]].items()
                if current_mask & mask1 & mask2
            )
            by_mask[current_mask] = wins / denominator
        values = [by_mask[mask] for mask in masks[reel]]
        mean = statistics.fmean(values)
        deviation = statistics.pstdev(values) or 1.0
        scores.append([(value - mean) / deviation for value in values])
    return scores


def window_scatter_count(symbols: list[str], stop: int) -> int:
    return sum(symbols[(stop + offset) % REEL_LENGTH] == "C1" for offset in range(4))


def calibrated_stop_weights(
    reels: list[list[str]], target_hit: float, scatter_factor: float
) -> tuple[list[list[int]], float, float]:
    """Create positive integer W:AA weights without changing any physical stop."""
    propensities = initial_stop_propensity(reels)

    def build(hit_factor: float) -> list[list[int]]:
        result: list[list[int]] = []
        for reel, symbols in enumerate(reels):
            result.append([
                max(
                    1,
                    round(STOP_WEIGHT_BASE * math.exp(
                        (hit_factor * propensities[reel][stop] if reel < 3 else 0.0)
                        + scatter_factor * window_scatter_count(symbols, stop)
                    )),
                )
                for stop in range(REEL_LENGTH)
            ])
        return result

    low, high = -4.0, 4.0
    best: tuple[float, float, list[list[int]], float] | None = None
    for _ in range(32):
        hit_factor = (low + high) / 2
        weights = build(hit_factor)
        rate = weighted_initial_hit_rate(reels, weights)
        candidate = (abs(rate - target_hit), hit_factor, weights, rate)
        if best is None or candidate[0] < best[0]:
            best = candidate
        if rate < target_hit:
            low = hit_factor
        else:
            high = hit_factor
    assert best is not None
    return best[2], best[3], best[1]


def calibrate_initial_hit(reels: list[list[str]], target: float, seed: int) -> tuple[float, int]:
    """Reorder within R1-R3 only; counts, gold totals and reel constraints stay unchanged."""
    rng = random.Random(seed)
    current = initial_hit_rate(reels)
    error = abs(current - target)
    accepted = 0
    for _ in range(CALIBRATION_TRIALS):
        reel = rng.randrange(3)
        left, right = rng.sample(range(REEL_LENGTH), 2)
        if reels[reel][left] == reels[reel][right]:
            continue
        reels[reel][left], reels[reel][right] = reels[reel][right], reels[reel][left]
        if reel_constraint_state(reels[reel])[0][0]:
            reels[reel][left], reels[reel][right] = reels[reel][right], reels[reel][left]
            continue
        candidate = initial_hit_rate(reels)
        candidate_error = abs(candidate - target)
        if candidate_error < error:
            current, error = candidate, candidate_error
            accepted += 1
            if error <= 1 / REEL_LENGTH**3:
                break
        else:
            reels[reel][left], reels[reel][right] = reels[reel][right], reels[reel][left]
    return current, accepted


def write_reels(worksheet, reels: list[list[str]], weights: list[list[int]]) -> None:
    for row in range(4, 404):
        for column in EDITABLE_COLUMNS:
            worksheet.cell(row, column).value = None
    for reel, symbols in enumerate(reels):
        for row, (symbol, weight) in enumerate(zip(symbols, weights[reel]), start=4):
            worksheet.cell(row, 11 + reel).value = symbol
            worksheet.cell(row, 23 + reel).value = weight


def write_drop_table(worksheet, tables: list[dict[str, int]]) -> None:
    for row, symbol in enumerate(DROP_SYMBOLS, start=4):
        worksheet.cell(row, 32).value = symbol
        for reel in range(5):
            worksheet.cell(row, 33 + reel).value = tables[reel][symbol]


def write_random_wild_weights(worksheet, weights: list[int]) -> None:
    if len(weights) != 4 or any(type(weight) is not int or weight < 0 for weight in weights):
        raise ValueError("Random Wild weights must be four non-negative integers")
    for row, weight in enumerate(weights, start=4):
        worksheet.cell(row, 30).value = weight


def copy_physical_reels(source, target) -> None:
    """Copy only K:O physical reels; each variant keeps its own W:AA/AF:AK."""
    for row in range(4, 404):
        for column in range(11, 16):
            target.cell(row, column).value = source.cell(row, column).value


def read_stop_weight_profile(path: Path = STOP_WEIGHT_PROFILE) -> dict[str, list[list[int]]]:
    if not path.exists():
        return {}
    raw = json.loads(path.read_text(encoding="utf-8"))
    result: dict[str, list[list[int]]] = {}
    for sheet_name in TARGETS:
        reels = raw.get(sheet_name)
        if reels is None:
            continue
        if len(reels) != 5 or any(len(reel) != REEL_LENGTH for reel in reels):
            raise ValueError(f"{path.name} {sheet_name}: expected 5 reels × {REEL_LENGTH} weights")
        normalized = [[int(weight) for weight in reel] for reel in reels]
        if any(any(type(weight) is not int or weight <= 0 for weight in reel) for reel in normalized):
            raise ValueError(f"{path.name} {sheet_name}: every stop weight must be a positive integer")
        result[sheet_name] = normalized
    return result


def update_workbook(target_path: Path, source_path: Path) -> dict[str, Any]:
    target = load_workbook(target_path, data_only=False)
    source = load_workbook(source_path, read_only=True, data_only=True)
    fill_weights = read_fill_weights(source)
    gold_profiles = read_competitor_gold_profiles(source_path)
    stop_weight_profile = read_stop_weight_profile()
    before = protected_values(target)
    result: dict[str, Any] = {}
    for target_name, source_name in TARGETS.items():
        source_reels, source_weights = read_reels(source[source_name])
        reels: list[list[str]] = []
        weights: list[list[int]] = []
        reel_swaps: list[int] = []
        for reel in range(5):
            # Expand the competitor's weighted source strip to exactly 200
            # physical stops.  W:AA is calibrated separately after all five
            # physical strips are finalized; no strip symbol is changed here.
            expanded = resample_to_200(source_reels[reel], source_weights[reel])
            with_gold, _ = apply_gold(
                expanded, GOLD_RATES[target_name][reel], gold_profiles[target_name]["initial"][reel]
            )
            repaired, swaps = repair_reel_constraints(with_gold)
            reels.append(repaired)
            reel_swaps.append(len(swaps))
        if target_name in stop_weight_profile:
            weights = stop_weight_profile[target_name]
            calibrated_hit = weighted_initial_hit_rate(reels, weights)
            hit_factor = None
        else:
            weights, calibrated_hit, hit_factor = calibrated_stop_weights(
                reels, INITIAL_HIT_TARGETS[target_name], INITIAL_SCATTER_FACTORS[target_name]
            )
        drop_tables = [
            symbol_drop_weights(
                exact_drop_weights(target_name)[reel],
                DROP_GOLD_RATES[target_name][reel],
                gold_profiles[target_name]["drop"][reel],
            )
            for reel in range(5)
        ]
        write_reels(target[target_name], reels, weights)
        write_drop_table(target[target_name], drop_tables)
        write_random_wild_weights(target[target_name], RANDOM_WILD_WEIGHTS[target_name])
        result[target_name] = {
            "reel_lengths": [len(reel) for reel in reels],
            "integer_weight_sums": [sum(reel) for reel in weights],
            "weighted_initial_hit_rate": calibrated_hit,
            "initial_hit_factor": hit_factor,
            "stop_weight_profile": STOP_WEIGHT_PROFILE.name if target_name in stop_weight_profile else None,
            "initial_scatter_factor": INITIAL_SCATTER_FACTORS[target_name],
            "source_weight_sums": [sum(reel) for reel in source_weights],
            "constraint_repair_swaps": reel_swaps,
            "r1_has_gold": any(symbol in GOLD_SYMBOLS.values() for symbol in reels[0]),
            "r5_has_gold": any(symbol in GOLD_SYMBOLS.values() for symbol in reels[4]),
            "all_weights_are_positive_integers": True,
            "drop_weight_sums": [sum(table.values()) for table in drop_tables],
            "drop_canonical_percentages": [{
                symbol: (table[symbol] + (table[GOLD_SYMBOLS[symbol]] if symbol in GOLD_SYMBOLS else 0)) / DROP_WEIGHT_TOTAL
                for symbol in ("C1", *SCORE_SYMBOLS)
            } for table in drop_tables],
            "drop_gold_rates": [
                sum(table[symbol] for symbol in BASE_SYMBOLS) / DROP_WEIGHT_TOTAL for table in drop_tables
            ],
            "initial_gold_rates": [
                sum(symbol in BASE_SYMBOLS for symbol in reel) / REEL_LENGTH for reel in reels
            ],
            "random_wild_weights": RANDOM_WILD_WEIGHTS[target_name],
        }
    for source_name, variant_names in VARIANT_REEL_TARGETS.items():
        for variant_name in variant_names:
            copy_physical_reels(target[source_name], target[variant_name])
            result[variant_name] = {"reel_lengths": [REEL_LENGTH] * 5, "reel_source": source_name}
    for sheet_name, weights in FG_VARIANT_RANDOM_WILD_WEIGHTS.items():
        write_random_wild_weights(target[sheet_name], weights)
        result[sheet_name]["random_wild_weights"] = weights
    source.close()
    target.save(target_path)

    checked = load_workbook(target_path, data_only=False)
    checked_protected = protected_values(checked)
    if checked_protected != before:
        changed = sorted(name for name in before if checked_protected[name] != before[name])
        raise RuntimeError(f"Values outside K:O/W:AA/AD/AF:AK changed: {changed}")
    result["protected_values_unchanged"] = True
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--target", type=Path, default=DEFAULT_TARGET)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    args = parser.parse_args()
    result = update_workbook(args.target.resolve(), args.source.resolve())
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
