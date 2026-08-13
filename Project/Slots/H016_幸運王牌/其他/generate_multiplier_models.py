"""Generate H016192/H016194 multiplier-line workbooks and config card weights.

The workbook layout follows H028's RTP-variant split, while the natural
distribution comes from the non-overlapping Super Ace JSONL samples used by
H016's competitor report.  Card weights use H028's (min, max] intervals and
0.1% minimum natural-rate rule.  The 92/94 variants share the current H016
math tables; only their card weights and RTP metadata differ.
"""
from __future__ import annotations

import argparse
import copy
import json
import math
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT = Path(__file__).resolve().parent.parent
SOURCE_DIR = Path(__file__).resolve().parent
H028_SOURCE = PROJECT.parent / "H028_雷神爆金 1000" / "Source"
H028_TEMPLATE = H028_SOURCE / "H028192A.xlsx"
COMPETITOR_ROOT = Path(
    "C:/Users/rhinshen/Mine/個人工作區/市場資訊/H5/遊戲資源/JILI"
)
COMPETITOR_FILES = (
    "SuperAce_BG_Combined_NoJP.jsonl",
    "SuperAce_BG_3.jsonl",
    "Super_Ace_BG_4.jsonl",
)
WEIGHT_TOTAL = 1_000_000_000
MIN_NATURAL_RATE = 0.001
BUY_PRICE = 40.5
SUPER_BUY_PRICE = 250.0


@dataclass
class Bucket:
    label: str
    minimum: float
    maximum: float
    count: int = 0
    pay: float = 0.0

    @property
    def average(self) -> float:
        return self.pay / self.count if self.count else 0.0


def competitor_dir() -> Path:
    for name in ("JILI - Super Ace", "JILI - Super Ace - m"):
        candidate = COMPETITOR_ROOT / name / "遊戲資料"
        if all((candidate / filename).is_file() for filename in COMPETITOR_FILES):
            return candidate
    raise FileNotFoundError("Cannot locate the three Super Ace JSONL files")


def load_config(path: Path) -> dict[str, Any]:
    raw = path.read_text(encoding="utf-8-sig")
    start, end = raw.find("{"), raw.rfind("}")
    if start < 0 or end <= start:
        raise ValueError(f"No JSON payload in {path}")
    return json.loads(raw[start : end + 1])


def save_config(path: Path, config: dict[str, Any]) -> None:
    payload = (
        "// Generated from Source/H0161.xlsx and Super Ace multiplier samples.\n"
        "window.H016_CONFIG="
        + json.dumps(config, ensure_ascii=False, separators=(",", ":"))
        + ";\n"
    )
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(payload, encoding="utf-8")
    os.replace(temporary, path)


def parse_range(label: str) -> tuple[float, float] | None:
    match = re.fullmatch(r"\(\s*(-?[0-9.]+)\s*,\s*([0-9.]+)\s*\]", label)
    return (float(match.group(1)), float(match.group(2))) if match else None


def h028_ranges() -> list[Bucket]:
    workbook = load_workbook(H028_TEMPLATE, read_only=True, data_only=False)
    try:
        sheet = workbook["Multiplier_Weight"]
        rows: list[Bucket] = []
        for row in range(3, sheet.max_row + 1):
            label = str(sheet.cell(row, 1).value or "").strip()
            parsed = parse_range(label)
            if parsed is not None:
                rows.append(Bucket(label, parsed[0], parsed[1]))
        if len(rows) != 64:
            raise ValueError(f"Expected 64 H028 multiplier ranges, got {len(rows)}")
        return rows
    finally:
        workbook.close()


def bucket_index(buckets: list[Bucket], value: float) -> int:
    for index, bucket in enumerate(buckets):
        if bucket.minimum < value <= bucket.maximum:
            return index
    raise ValueError(f"Multiplier {value} is outside H028 range coverage")


def collect_competitor(buckets: list[Bucket]) -> dict[str, Any]:
    bg = copy.deepcopy(buckets)
    fg = copy.deepcopy(buckets)
    trigger_count = 0
    trigger_bg_pay = 0.0
    rounds = 0
    total_bet = 0.0
    source = competitor_dir()
    for filename in COMPETITOR_FILES:
        with (source / filename).open(encoding="utf-8") as handle:
            for line in handle:
                obj = json.loads(line)
                bet = float(obj["bet"])
                if bet <= 0:
                    raise ValueError("Competitor bet must be positive")
                plates = obj["plate"]["plate"]
                rounds += 1
                total_bet += bet
                bg_multiplier = float(plates[0].get("win", 0.0)) / bet
                if len(plates) > 1:
                    trigger_count += 1
                    trigger_bg_pay += bg_multiplier
                    fg_multiplier = sum(float(plate.get("win", 0.0)) for plate in plates[1:]) / bet
                    target = fg[bucket_index(fg, fg_multiplier)]
                    target.count += 1
                    target.pay += fg_multiplier
                else:
                    target = bg[bucket_index(bg, bg_multiplier)]
                    target.count += 1
                    target.pay += bg_multiplier
    bg_rtp = (sum(bucket.pay for bucket in bg) + trigger_bg_pay) / rounds
    fg_rtp = sum(bucket.pay for bucket in fg) / rounds
    return {
        "directory": str(source),
        "rounds": rounds,
        "total_bet": total_bet,
        "trigger_count": trigger_count,
        "trigger_rate": trigger_count / rounds,
        "trigger_bg_pay": trigger_bg_pay,
        "bg": bg,
        "fg": fg,
        "bg_rtp": bg_rtp,
        "fg_rtp": fg_rtp,
        "total_rtp": bg_rtp + fg_rtp,
    }


def scale_weights(raw: list[float], total: int = WEIGHT_TOTAL) -> list[int]:
    source_total = sum(raw)
    if source_total <= 0:
        raise ValueError("Cannot scale an empty weight vector")
    exact = [value * total / source_total for value in raw]
    result = [int(value) for value in exact]
    remainder = total - sum(result)
    order = sorted(range(len(raw)), key=lambda i: exact[i] - result[i], reverse=True)
    for index in order[:remainder]:
        result[index] += 1
    return result


def natural_raw(buckets: list[Bucket], denominator: int) -> list[float]:
    return [
        float(bucket.count) if bucket.count / denominator >= MIN_NATURAL_RATE else 0.0
        for bucket in buckets
    ]


def tilted_weights(buckets: list[Bucket], target_mean: float) -> list[int]:
    denominator = sum(bucket.count for bucket in buckets)
    raw = natural_raw(buckets, denominator)
    active = [i for i, value in enumerate(raw) if value > 0 and buckets[i].count > 0]
    if not active:
        raise ValueError("No multiplier ranges pass the natural-rate threshold")
    low_mean = min(buckets[i].average for i in active)
    high_mean = max(buckets[i].average for i in active)
    if not low_mean <= target_mean <= high_mean:
        raise ValueError(f"Target mean {target_mean:.6f} outside {low_mean:.6f}..{high_mean:.6f}")

    def distribution(lam: float) -> list[float]:
        logs = [math.log(raw[i]) + lam * buckets[i].average for i in active]
        pivot = max(logs)
        values = [math.exp(value - pivot) for value in logs]
        total = sum(values)
        result = [0.0] * len(buckets)
        for index, value in zip(active, values):
            result[index] = value / total
        return result

    lower, upper = -1.0, 1.0
    for _ in range(120):
        middle = (lower + upper) / 2
        probs = distribution(middle)
        mean = sum(prob * bucket.average for prob, bucket in zip(probs, buckets))
        if mean < target_mean:
            lower = middle
        else:
            upper = middle
    return scale_weights(distribution((lower + upper) / 2))


def base_weights(stats: dict[str, Any]) -> list[int]:
    raw = natural_raw(stats["bg"], stats["rounds"])
    raw.append(float(stats["trigger_count"]))
    return scale_weights(raw)


def base_model_metrics(stats: dict[str, Any], weights: list[int]) -> tuple[float, float]:
    trigger_average = stats["trigger_bg_pay"] / stats["trigger_count"]
    means = [bucket.average for bucket in stats["bg"]] + [trigger_average]
    rtp = sum(weight / WEIGHT_TOTAL * mean for weight, mean in zip(weights, means))
    return rtp, weights[-1] / WEIGHT_TOTAL


def cards_from_weights(
    buckets: list[Bucket], weights: list[int], table: str, *, include_free_game: bool = False
) -> list[dict[str, Any]]:
    cards = [
        {
            "type": "range", "min": bucket.minimum, "max": bucket.maximum,
            "table": table, "weight": int(weight),
        }
        for bucket, weight in zip(buckets, weights[: len(buckets)])
    ]
    if include_free_game:
        cards.append({"type": "free_game", "table": "A", "weight": int(weights[-1])})
    return cards


def version_cards(stats: dict[str, Any], target_rtp: float) -> dict[str, Any]:
    base = base_weights(stats)
    bg_model_rtp, trigger_rate = base_model_metrics(stats, base)
    fg_contribution = target_rtp - bg_model_rtp
    if fg_contribution <= 0:
        raise ValueError("Target RTP is lower than competitor BG RTP")
    fg_target_mean = fg_contribution / trigger_rate
    free_weights = tilted_weights(stats["fg"], fg_target_mean)
    buy_weights = tilted_weights(stats["fg"], target_rtp * BUY_PRICE)
    super_weights = tilted_weights(stats["fg"], target_rtp * SUPER_BUY_PRICE)
    profile = {
        "base_game": cards_from_weights(stats["bg"], base, "B", include_free_game=True),
        "free_game": cards_from_weights(stats["fg"], free_weights, "E"),
        "buy_feature": cards_from_weights(stats["fg"], buy_weights, "E"),
        "super_feature": cards_from_weights(stats["fg"], super_weights, "G"),
    }
    return {
        "enabled": True,
        "retry_limit": 200_000,
        "default_profile": "weight_2",
        "profiles": {"weight_1": copy.deepcopy(profile), "weight_2": profile},
    }


def style_workbook(workbook: Workbook) -> None:
    dark = PatternFill("solid", fgColor="666666")
    light = PatternFill("solid", fgColor="D9EAF7")
    border = Border(*(Side(style="thin", color="999999") for _ in range(4)))
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.border = border
                    cell.alignment = Alignment(vertical="center")
        for cell in sheet[1]:
            cell.fill = dark
            cell.font = Font(color="FFFFFF", bold=True)
        if sheet.title != "Overview":
            for cell in sheet[2]:
                cell.fill = dark
                cell.font = Font(color="FFFFFF", bold=True)
            for row in range(3, sheet.max_row + 1):
                if row % 2:
                    for cell in sheet[row]:
                        cell.fill = light


def write_detail_block(
    sheet, start: int, title: str, buckets: list[Bucket], weights: list[int], denominator: int
) -> None:
    sheet.cell(start, 1, title)
    headers = ["Range", "Count", "Pay (x)", "Natural Rate", "Avg. Multiplier", "", "Range", "Fix Num", "Final Rate", "Weight", "RTP"]
    for column, header in enumerate(headers, 1):
        sheet.cell(start + 1, column, header)
    for offset, (bucket, weight) in enumerate(zip(buckets, weights), start + 2):
        natural = bucket.count / denominator if denominator else 0.0
        final = weight / WEIGHT_TOTAL
        sheet.cell(offset, 1, bucket.label)
        sheet.cell(offset, 2, bucket.count)
        sheet.cell(offset, 3, bucket.pay)
        sheet.cell(offset, 4, natural)
        sheet.cell(offset, 5, bucket.average)
        sheet.cell(offset, 7, bucket.label)
        sheet.cell(offset, 8, final / natural if natural else 0.0)
        sheet.cell(offset, 9, final)
        sheet.cell(offset, 10, weight)
        sheet.cell(offset, 11, final * bucket.average)
        for column in (4, 8, 9, 11):
            sheet.cell(offset, column).number_format = "0.000000%"


def create_workbook(path: Path, stats: dict[str, Any], config: dict[str, Any], target: float) -> None:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.append(["H016 Multiplier Line Model", "Value"])
    overview_rows = [
        ("PARsheet ID", config["parsheet_id"]), ("Target RTP", target),
        ("Reference", "JILI Super Ace JSONL"), ("Reference rounds", stats["rounds"]),
        ("Reference FG sessions", stats["trigger_count"]),
        ("Reference BG RTP", stats["bg_rtp"]), ("Reference FG RTP", stats["fg_rtp"]),
        ("Reference Total RTP", stats["total_rtp"]),
        ("Range source", "H028192A.xlsx / Multiplier_Weight"),
        ("Natural-rate threshold", MIN_NATURAL_RATE),
        ("Normal bet", "BG fixed to competitor; FG tilted to target Total RTP"),
        ("Buy Feature", f"FG-session distribution tilted to {target:.2%} at {BUY_PRICE}x price"),
        ("Super Feature", f"Competitor FG-session proxy tilted to {target:.2%} at {SUPER_BUY_PRICE}x price"),
    ]
    for key, value in overview_rows:
        overview.append([key, value])
    for row in range(3, overview.max_row + 1):
        if isinstance(overview.cell(row, 2).value, float) and overview.cell(row, 1).value in {
            "Target RTP", "Reference BG RTP", "Reference FG RTP", "Reference Total RTP", "Natural-rate threshold"
        }:
            overview.cell(row, 2).number_format = "0.0000%"

    multiplier = workbook.create_sheet("Multiplier_Weight")
    multiplier.append(["H016 / H028-style Multiplier Weight", None, None, None, None, None, None])
    multiplier.append(["Range", "Weight_NB_BG_Newbie", "Weight_NB_FG_Newbie", "Weight_NB_BG", "Weight_NB_FG", "Weight_BF", "Weight_SF"])
    profiles = config["card_system"]["profiles"]
    columns = [
        profiles["weight_1"]["base_game"], profiles["weight_1"]["free_game"],
        profiles["weight_2"]["base_game"], profiles["weight_2"]["free_game"],
        profiles["weight_2"]["buy_feature"], profiles["weight_1"]["super_feature"],
    ]
    labels = [bucket.label for bucket in stats["bg"]] + ["Free Game"]
    for index, label in enumerate(labels):
        row = [label]
        for cards in columns:
            row.append(int(cards[index]["weight"]) if index < len(cards) else 0)
        multiplier.append(row)

    for sheet_name, profile_name in (("Detail", "weight_2"), ("Detail_Newbie", "weight_1")):
        detail = workbook.create_sheet(sheet_name)
        detail.append([f"H016 {profile_name} multiplier detail", None])
        detail.append(["Threshold", WEIGHT_TOTAL])
        detail.append(["Coin in", 1])
        profile = profiles[profile_name]
        bg_weights = [int(card["weight"]) for card in profile["base_game"][:-1]]
        write_detail_block(detail, 5, "Normal Bet / Base Game", stats["bg"], bg_weights, stats["rounds"])
        write_detail_block(detail, 73, "Normal Bet / Free Game Session", stats["fg"], [int(card["weight"]) for card in profile["free_game"]], stats["trigger_count"])
        if sheet_name == "Detail":
            write_detail_block(detail, 141, "Buy Feature / Free Game Session", stats["fg"], [int(card["weight"]) for card in profile["buy_feature"]], stats["trigger_count"])
            write_detail_block(detail, 209, "Super Feature / Free Game Session (competitor proxy)", stats["fg"], [int(card["weight"]) for card in profiles["weight_1"]["super_feature"]], stats["trigger_count"])

    style_workbook(workbook)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A3" if sheet.title == "Multiplier_Weight" else "A2"
        for column in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(column)].width = 18
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    temporary = path.with_suffix(".tmp.xlsx")
    workbook.save(temporary)
    workbook.close()
    os.replace(temporary, path)


def validate(config: dict[str, Any], target: float, stats: dict[str, Any]) -> dict[str, float]:
    card_system = config["card_system"]
    for profile in card_system["profiles"].values():
        for cards in profile.values():
            if sum(int(card["weight"]) for card in cards) != WEIGHT_TOTAL:
                raise ValueError("Every card table must sum to 1,000,000,000")
            if any(int(card["weight"]) < 0 for card in cards):
                raise ValueError("Card weights must be non-negative integers")
    profile = card_system["profiles"]["weight_2"]
    base_cards = profile["base_game"]
    bg_means = [bucket.average for bucket in stats["bg"]] + [stats["trigger_bg_pay"] / stats["trigger_count"]]
    bg_rtp = sum(int(card["weight"]) / WEIGHT_TOTAL * mean for card, mean in zip(base_cards, bg_means))
    fg_mean = sum(int(card["weight"]) / WEIGHT_TOTAL * bucket.average for card, bucket in zip(profile["free_game"], stats["fg"]))
    trigger = int(base_cards[-1]["weight"]) / WEIGHT_TOTAL
    total = bg_rtp + trigger * fg_mean
    if abs(total - target) > 2e-6:
        raise ValueError(f"Card model RTP {total:.9f} does not match target {target:.9f}")
    return {"bg_rtp": bg_rtp, "fg_session_mean": fg_mean, "trigger_rate": trigger, "total_rtp": total}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--check", action="store_true", help="Calculate and validate without writing")
    args = parser.parse_args()
    buckets = h028_ranges()
    stats = collect_competitor(buckets)
    base = load_config(PROJECT / "config_92.js")
    outputs = {}
    for rtp, suffix in ((0.92, "92"), (0.94, "94")):
        config = copy.deepcopy(base)
        config["parsheet_id"] = f"H0161{suffix}"
        config["rtp_label"] = int(round(rtp * 100))
        config["source_xlsx"] = f"H0161{suffix}.xlsx"
        config["card_system"] = version_cards(stats, rtp)
        metrics = validate(config, rtp, stats)
        outputs[suffix] = metrics
        if not args.check:
            save_config(PROJECT / f"config_{suffix}.js", config)
            create_workbook(SOURCE_DIR / f"H0161{suffix}.xlsx", stats, config, rtp)
    print(json.dumps({
        "reference": stats["directory"], "rounds": stats["rounds"],
        "fg_sessions": stats["trigger_count"], "competitor_rtp": stats["total_rtp"],
        "outputs": outputs, "written": not args.check,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
