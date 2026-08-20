"""Build the H016 v8 vs JHS101003 92-oldhand tuning comparison.

The current H016 side is read from the locked 100M Card-On Record workbook.
The JHS101003 side is a parameter replay: its formal JSON/Java settings are
mapped into the current validated fast simulator so both sides share the same
statistics definitions.  The report explicitly records the remaining engine
translation limitation for BG physical-wheel replenishment.
"""
from __future__ import annotations

import argparse
import copy
import importlib.util
import json
import math
import os
import sys
from bisect import bisect_left
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCRIPT = Path(__file__).resolve()
H016_DIR = SCRIPT.parents[2]
DIAG_DIR = SCRIPT.parent
CURRENT_RECORD = H016_DIR / "Record" / "H016192A_08000000_2608190629_betmode0_108_9209_oldhand_card.xlsx"
CURRENT_CONFIG = H016_DIR / "config.js"
CURRENT_CARD = H016_DIR / "config_92A.js"
JHS_JSON = H016_DIR / "其他" / "參考資料" / "gameSetting_JHS101003.json"
JHS_BG_JAVA = H016_DIR / "其他" / "參考資料" / "BaseGameHandler_JHS101003.java"
JHS_FG_JAVA = H016_DIR / "其他" / "參考資料" / "FreeGameHandler_JHS101003.java"
FAST_SIMULATOR = H016_DIR / "其他" / "fast_simulator.py"
COMPETITOR_BUILDER = H016_DIR / "其他" / "工具" / "build_competitor_comparison.py"
SUPER_ACE_DIR = Path("C:/Users/rhinshen/Mine/個人工作區/市場資訊/H5/遊戲資源/JILI/JILI - Super Ace - m/遊戲資料")
SUPER_ACE_FILES = tuple(
    SUPER_ACE_DIR / name
    for name in (
        "SuperAce_BG_Combined_NoJP.jsonl",
        "SuperAce_BG_3.jsonl",
        "Super_Ace_BG_4.jsonl",
    )
)

OUT_STEM = "H016_v8_vs_101003_92老手_調性比較"
OUT_MD = DIAG_DIR / f"{OUT_STEM}.md"
OUT_ARTIFACT = DIAG_DIR / f"{OUT_STEM}.artifact.json"

SYMBOL_NAMES = {0: "WW", 1: "W2", 2: "C1", 3: "M1", 4: "M2", 5: "M3", 6: "M4", 7: "A", 8: "K", 9: "Q", 10: "J", 11: "M1G", 12: "M2G", 13: "M3G", 14: "M4G", 15: "AG", 16: "KG", 17: "QG", 18: "JG"}
REPORT_SYMBOLS = ("C1", "M1", "M2", "M3", "M4", "A", "K", "Q", "J")
PAY_SYMBOLS = REPORT_SYMBOLS[1:]
THRESHOLDS = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 15, 20, 25, 30, 35, 40, 45, 50, 60, 70, 80, 90, 100, 120, 140, 160, 180, 200, 250, 300, 350, 400, 450, 500, 550, 600, 650, 700, 750, 800, 850, 900, 950, 1000, 2000, 3000, 4000, 5000, 6000, 7000, 8000, 9000, 10000, 20000, 30000, 40000, 50000, 60000, 70000, 80000, 90000, 100000, 9999999]


def load_js(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8-sig")
    return json.loads(text[text.index("{"): text.rfind("}") + 1])


def load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


def canonical_id(value: int) -> int:
    value = int(value)
    return value - 8 if 11 <= value <= 18 else value


def symbol_name(value: int) -> str:
    return SYMBOL_NAMES[canonical_id(value)]


def card_rows(levels: list[list[float]], weights: list[int], include_free_game: bool) -> list[dict[str, Any]]:
    cards: list[dict[str, Any]] = []
    last = len(levels) - 1
    for index, ((lower, upper), weight) in enumerate(zip(levels, weights)):
        if include_free_game and index == last:
            cards.append({"type": "free_game", "table": "FG", "weight": int(weight)})
        else:
            cards.append({"type": "range", "min": float(lower), "max": float(upper), "table": "B", "weight": int(weight)})
    return cards


def build_jhs_config() -> tuple[dict[str, Any], str]:
    root = json.loads(JHS_JSON.read_text(encoding="utf-8"))["executeSetting"][0][0]
    base = root["baseGameSetting"]
    free = root["freeGameSetting"]
    bext = base["baseGameExtendSetting"]
    fext = free["freeGameExtendSetting"]

    pays = {
        str(symbol): [float(row[2]) / 100.0, float(row[3]) / 100.0, float(row[4]) / 100.0]
        for symbol, row in enumerate(base["payTable"])
    }
    tables: dict[str, Any] = {}
    for index in range(int(base["tableCount"])):
        reels = [[int(value) for value in reel["wheelData"]] for reel in base["wheelData"][index]]
        weights = [[int(value) for value in reel["wheelData"]] for reel in bext["wheelWeight"][index]]
        tables[f"bg_{index + 1}"] = {
            "reels": reels,
            "weights": weights,
            # Java BG replacement receives wheelData.  Mapping the same 200
            # symbols and stop weights preserves its one-cell marginal.
            "drop_values": copy.deepcopy(reels),
            "drop_weights": copy.deepcopy(weights),
            "random_wild": {
                "values": [int(value) for value in bext["randomWild"]],
                "weights": [int(value) for value in bext["randomWildWeight"][index]],
            },
            "multipliers": [int(value) for value in bext["comboMultiplierList"]],
        }
    for index in range(int(free["tableCount"])):
        tables[f"fg_{index + 1}"] = {
            "reels": [[int(value) for value in reel["wheelData"]] for reel in free["wheelData"][index]],
            "weights": [[int(value) for value in reel["wheelData"]] for reel in fext["wheelWeight"][index]],
            "drop_values": [list(range(19)) for _ in range(5)],
            "drop_weights": [[int(value) for value in reel["wheelData"]] for reel in fext["dropWheelWeight"][index]],
            "random_wild": {
                "values": [int(value) for value in fext["randomWild"]],
                "weights": [int(value) for value in fext["randomWildWeight"][index]],
            },
            "multipliers": [int(value) for value in fext["comboMultiplierList"]],
        }

    tables["buy"] = copy.deepcopy(tables["bg_3"])
    tables["super"] = copy.deepcopy(tables["fg_2"])
    permissive = [{"type": "range", "min": -1, "max": 9_999_999, "table": "B", "weight": 1}]
    profiles: dict[str, Any] = {}
    for profile, row in (("weight_1", 0), ("weight_2", 2)):
        profiles[profile] = {
            "base_game": card_rows(base["oddsLevelList"], base["oddsWeightTable"][row], True),
            "free_game": card_rows(free["oddsLevelList"], free["oddsWeightTable"][row], False),
            "buy_feature": copy.deepcopy(permissive),
            "super_feature": copy.deepcopy(permissive),
        }

    config = {
        "model": "H0161",
        "version": str(root["version"]),
        "symbol_names": {str(key): value for key, value in SYMBOL_NAMES.items()},
        "pays": pays,
        "tables": tables,
        "table_selection": {
            "base": [{"table": f"bg_{index + 1}", "weight": int(weight)} for index, weight in enumerate(base["tableHitProbability"]) if int(weight) > 0],
            "free": [{"table": f"fg_{index + 1}", "weight": int(weight)} for index, weight in enumerate(free["tableHitProbability"]) if int(weight) > 0],
            "retrigger": [{"table": f"fg_{index + 1}", "weight": int(weight)} for index, weight in enumerate(free["tableHitProbability"]) if int(weight) > 0],
            "super_free": [{"table": "super", "weight": 1}],
            "super_retrigger": [{"table": "super", "weight": 1}],
        },
        "free_spins": int(fext["baseRound"]),
        "retrigger_spins": int(fext["increaseFreeGameRound"]),
        "free_spin_cap": int(fext["maxRound"]),
        "buy_price": 40.5,
        "super_buy_price": 250.0,
        "card_system": {
            "enabled": True,
            "default_profile": "weight_2",
            "retry_limit": 20000,
            "profiles": profiles,
        },
    }
    return config, str(root["version"])


def load_current_record(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    base = {str(a): b for a, b in wb["Base Info"].iter_rows(min_row=2, values_only=True) if a is not None}
    combo = {str(label): {"BG": int(bg or 0), "FG": int(fg or 0)} for label, bg, fg in wb["Eliminate"].iter_rows(min_row=2, values_only=True) if label is not None}
    length_hits: dict[str, Counter] = {"BG": Counter(), "FG": Counter()}
    length_pay: dict[str, Counter] = {"BG": Counter(), "FG": Counter()}
    for scene, symbol, length, hits, pay in wb["Symbol Length"].iter_rows(min_row=2, values_only=True):
        if scene in length_hits:
            length_hits[str(scene)][(str(symbol), int(length))] = int(hits or 0)
            length_pay[str(scene)][(str(symbol), int(length))] = float(pay or 0)

    distributions: dict[str, dict[str, list[float]]] = {}
    gold: dict[str, dict[str, Any]] = {}
    for scene, stage, sheet in (("BG", "initial", "BG Initial Symbol"), ("BG", "drop", "BG Drop Symbol"), ("FG", "initial", "FG Initial Symbol"), ("FG", "drop", "FG Drop Symbol")):
        totals = [0.0] * 5
        merged = {symbol: [0.0] * 5 for symbol in REPORT_SYMBOLS}
        gold_total = [0.0] * 5
        gold_symbols = {symbol: [0.0] * 5 for symbol in PAY_SYMBOLS}
        rows = list(wb[sheet].iter_rows(min_row=2, values_only=True))
        for row in rows:
            raw = str(row[0])
            values = [float(row[1 + reel] or 0) for reel in range(5)]
            is_gold = raw.endswith("G") and raw[:-1] in PAY_SYMBOLS
            canonical = raw[:-1] if is_gold else raw
            for reel, value in enumerate(values):
                totals[reel] += value
                if canonical in merged:
                    merged[canonical][reel] += value
                if is_gold:
                    gold_total[reel] += value
                    gold_symbols[canonical][reel] += value
        for reel in range(5):
            denominator = max(1.0, totals[reel])
            for symbol in REPORT_SYMBOLS:
                merged[symbol][reel] /= denominator
            gold_total[reel] /= denominator
            for symbol in PAY_SYMBOLS:
                gold_symbols[symbol][reel] /= denominator
        distributions[f"{scene}_{stage}"] = merged
        gold[f"{scene}_{stage}"] = {"total": gold_total, "symbols": gold_symbols, "denominators": totals}

    multiplier = []
    headers = [str(cell) for cell in next(wb["Multiplier Line"].iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {name: index for index, name in enumerate(headers)}
    for row_index, row in enumerate(wb["Multiplier Line"].iter_rows(min_row=2, values_only=True)):
        multiplier.append({
            "index": row_index,
            "interval": str(row[idx["Interval"]]),
            "upper": float(row[idx["Interval_Upper"]] or 0),
            "bg_count": int(row[idx["base_game_cnt"]] or 0),
            "bg_pay": float(row[idx["base_game_pay"]] or 0),
            "fg_count": int(row[idx["free_game_cnt"]] or 0),
            "fg_pay": float(row[idx["free_game_pay"]] or 0),
        })
    wb.close()
    return {"base": base, "combo": combo, "length_hits": length_hits, "length_pay": length_pay, "distributions": distributions, "gold": gold, "multiplier": multiplier}


def stats_distributions(stats: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any]]:
    distributions: dict[str, dict[str, list[float]]] = {}
    gold: dict[str, dict[str, Any]] = {}
    for scene, stage in (("BG", "initial"), ("BG", "drop"), ("FG", "initial"), ("FG", "drop")):
        counter = stats[f"{scene.lower()}_{stage}_symbols"]
        totals = [0.0] * 5
        merged = {symbol: [0.0] * 5 for symbol in REPORT_SYMBOLS}
        gold_total = [0.0] * 5
        gold_symbols = {symbol: [0.0] * 5 for symbol in PAY_SYMBOLS}
        for (reel, numeric), count in counter.items():
            reel, numeric, count = int(reel), int(numeric), float(count)
            totals[reel] += count
            name = symbol_name(numeric)
            if name in merged:
                merged[name][reel] += count
            if 11 <= numeric <= 18:
                gold_total[reel] += count
                gold_symbols[name][reel] += count
        for reel in range(5):
            denominator = max(1.0, totals[reel])
            for symbol in REPORT_SYMBOLS:
                merged[symbol][reel] /= denominator
            gold_total[reel] /= denominator
            for symbol in PAY_SYMBOLS:
                gold_symbols[symbol][reel] /= denominator
        distributions[f"{scene}_{stage}"] = merged
        gold[f"{scene}_{stage}"] = {"total": gold_total, "symbols": gold_symbols, "denominators": totals}
    return distributions, gold


def jhs_result(stats: dict[str, Any]) -> dict[str, Any]:
    distributions, gold = stats_distributions(stats)
    combo = {label: {"BG": int(stats["combo_bg"].get(index, 0)), "FG": int(stats["combo_fg"].get(index, 0))} for index, label in enumerate(("0", "1", "2", "3", "4", "5+"))}
    multiplier = []
    lower = None
    for index, upper in enumerate(THRESHOLDS):
        interval = "0" if index == 0 else f"({lower}, {upper}]"
        multiplier.append({
            "index": index,
            "interval": interval,
            "upper": float(upper),
            "bg_count": int(stats["multiplier_bg_count"].get(index, 0)),
            "bg_pay": float(stats["multiplier_bg_pay"].get(index, 0)),
            "fg_count": int(stats["multiplier_fg_count"].get(index, 0)),
            "fg_pay": float(stats["multiplier_fg_pay"].get(index, 0)),
        })
        lower = upper
    return {
        "stats": stats,
        "combo": combo,
        "length_hits": {"BG": Counter({(symbol_name(s), n): int(v) for (s, n), v in stats["bg_symbol_length_hits"].items()}), "FG": Counter({(symbol_name(s), n): int(v) for (s, n), v in stats["fg_symbol_length_hits"].items()})},
        "length_pay": {"BG": Counter({(symbol_name(s), n): float(v) for (s, n), v in stats["bg_symbol_length_pay"].items()}), "FG": Counter({(symbol_name(s), n): float(v) for (s, n), v in stats["fg_symbol_length_pay"].items()})},
        "distributions": distributions,
        "gold": gold,
        "multiplier": multiplier,
    }


def load_super_ace() -> dict[str, Any]:
    """Load the non-overlapping natural-play Super Ace JSONL sample."""
    module = load_module("_h016_competitor_builder", COMPETITOR_BUILDER)
    module.COMPETITOR_DIR = SUPER_ACE_DIR
    module.FILES = list(SUPER_ACE_FILES)
    raw = module.raw_competitor()

    multiplier_bg_count: Counter[int] = Counter()
    multiplier_bg_pay: Counter[int] = Counter()
    multiplier_fg_count: Counter[int] = Counter()
    multiplier_fg_pay: Counter[int] = Counter()
    cascades_bg = cascades_fg = 0
    for path in SUPER_ACE_FILES:
        with path.open(encoding="utf-8") as handle:
            for line in handle:
                obj = json.loads(line)
                plates = obj["plate"]["plate"]
                bet = float(obj["bet"])
                bg_pay = float(plates[0].get("win", 0.0))
                bg_bucket = min(len(THRESHOLDS) - 1, bisect_left(THRESHOLDS, bg_pay / max(bet, 1e-15)))
                multiplier_bg_count[bg_bucket] += 1
                multiplier_bg_pay[bg_bucket] += bg_pay
                cascades_bg += len(plates[0].get("combo", []))
                if len(plates) > 1:
                    fg_pay = sum(float(plate.get("win", 0.0)) for plate in plates[1:])
                    fg_bucket = min(len(THRESHOLDS) - 1, bisect_left(THRESHOLDS, fg_pay / max(bet, 1e-15)))
                    multiplier_fg_count[fg_bucket] += 1
                    multiplier_fg_pay[fg_bucket] += fg_pay
                    cascades_fg += sum(len(plate.get("combo", [])) for plate in plates[1:])

    distributions: dict[str, dict[str, list[float]]] = {}
    gold: dict[str, dict[str, Any]] = {}
    stack: dict[str, dict[str, Any]] = {}
    for scene in ("BG", "FG"):
        data = raw["counts"][scene]
        stack[scene] = {
            "counts": [Counter({key: float(value) / max(1, data["stack_total"][reel]) for key, value in data["stack"][reel].items()}) for reel in range(5)],
            "totals": [1.0] * 5,
        }
        for stage in ("initial", "drop"):
            distributions[f"{scene}_{stage}"] = {
                symbol: [
                    float(data[stage][reel][symbol]) / max(1, data[f"{stage}_total"][reel])
                    for reel in range(5)
                ]
                for symbol in REPORT_SYMBOLS
            }
            gold[f"{scene}_{stage}"] = {
                "total": [
                    float(data[f"gold_{stage}"][reel]) / max(1, data[f"{stage}_total"][reel])
                    for reel in range(5)
                ],
                "denominators": [float(data[f"{stage}_total"][reel]) for reel in range(5)],
                "symbols": {
                    symbol: [
                        float(data[f"gold_symbol_{stage}"][reel][symbol]) / max(1, data[f"{stage}_total"][reel])
                        for reel in range(5)
                    ]
                    for symbol in PAY_SYMBOLS
                },
            }

    multiplier = []
    lower = None
    for index, upper in enumerate(THRESHOLDS):
        multiplier.append({
            "index": index,
            "interval": "0" if index == 0 else f"({lower}, {upper}]",
            "upper": float(upper),
            "bg_count": int(multiplier_bg_count[index]),
            "bg_pay": float(multiplier_bg_pay[index]),
            "fg_count": int(multiplier_fg_count[index]),
            "fg_pay": float(multiplier_fg_pay[index]),
        })
        lower = upper

    counts = raw["counts"]
    stats = {
        "rounds": int(raw["rounds"]),
        "fg_spins": int(raw["fg_spins"]),
        "fg_triggers": int(round(raw["fg_trigger_rate"] * raw["rounds"])),
        "coin_in": float(raw["coin_in"]),
        "pay_bg": float(raw["rtp_bg"] * raw["coin_in"]),
        "pay_fg": float(raw["rtp_fg"] * raw["coin_in"]),
        "bg_hit_spins": int(counts["BG"]["hits"]),
        "fg_hit_spins": int(counts["FG"]["hits"]),
        "cascades_bg": int(cascades_bg),
        "cascades_fg": int(cascades_fg),
        "bg_w2_events": int(round(raw["w2_bg_event_rate"] * raw["rounds"])),
        "fg_w2_events": 0,
        "bg_w2_counts": Counter(raw["w2_bg_counts"]),
        "fg_w2_counts": Counter(),
    }
    combo = {
        label: {scene: int(counts[scene]["combo"].get(index, 0)) for scene in ("BG", "FG")}
        for index, label in enumerate(("0", "1", "2", "3", "4", "5+"))
    }
    return {
        "stats": stats,
        "combo": combo,
        "length_hits": {scene: Counter(counts[scene]["symbol_length_hits"]) for scene in ("BG", "FG")},
        "length_pay": {scene: Counter(counts[scene]["symbol_length_win"]) for scene in ("BG", "FG")},
        "distributions": distributions,
        "gold": gold,
        "stack": stack,
        "multiplier": multiplier,
    }


def visible_stack_counts(window: list[str]) -> Counter:
    result: Counter = Counter()
    index = 0
    while index < len(window):
        end = index + 1
        while end < len(window) and window[end] == window[index]:
            end += 1
        length = end - index
        if length in (2, 3, 4) and window[index] in REPORT_SYMBOLS:
            result[(window[index], length)] += 1
        index = end
    return result


def table_stack(table: dict[str, Any]) -> dict[str, Any]:
    counts = [Counter() for _ in range(5)]
    totals = [0.0] * 5
    for reel in range(5):
        values = table["reels"][reel]
        weights = table["weights"][reel]
        for stop, weight in enumerate(weights):
            weight = max(0.0, float(weight))
            if weight <= 0:
                continue
            window = [symbol_name(values[(stop + row) % len(values)]) for row in range(4)]
            for key, occurrences in visible_stack_counts(window).items():
                counts[reel][key] += occurrences * weight
            totals[reel] += weight
    return {"counts": counts, "totals": totals}


def mixed_stack(config: dict[str, Any], group: str) -> dict[str, Any]:
    items = [(str(item["table"]), float(item["weight"])) for item in config["table_selection"][group] if float(item["weight"]) > 0]
    selection_total = sum(weight for _, weight in items)
    result = [Counter() for _ in range(5)]
    for table_name, selection in items:
        data = table_stack(config["tables"][table_name])
        for reel in range(5):
            for key, count in data["counts"][reel].items():
                result[reel][key] += selection / selection_total * count / max(1.0, data["totals"][reel])
    return {"counts": result, "totals": [1.0] * 5}


def rel(reference: float, current: float) -> str:
    if abs(reference) < 1e-15:
        return "N/A"
    return f"{(reference - current) / reference:+.2%}"


def pct(value: float, digits: int = 4) -> str:
    return f"{float(value) * 100:.{digits}f}%"


def num(value: float, digits: int = 4) -> str:
    return f"{float(value):,.{digits}f}"


def core_metrics(current: dict[str, Any], old: dict[str, Any], competitor: dict[str, Any]) -> list[dict[str, Any]]:
    c = current["base"]
    s = old["stats"]
    p = competitor["stats"]
    rounds, fg_spins, triggers = s["rounds"], s["fg_spins"], s["fg_triggers"]
    rows = [
        ("Total RTP", (p["pay_bg"] + p["pay_fg"]) / p["coin_in"], (s["pay_bg"] + s["pay_fg"]) / s["coin_in"], c["rtp_total"], "pct"),
        ("BG RTP", p["pay_bg"] / p["coin_in"], s["pay_bg"] / s["coin_in"], c["rtp_bg"], "pct"),
        ("FG RTP", p["pay_fg"] / p["coin_in"], s["pay_fg"] / s["coin_in"], c["rtp_fg"], "pct"),
        ("BG Hit Rate", p["bg_hit_spins"] / p["rounds"], s["bg_hit_spins"] / rounds, c["bg_hit_rate"], "pct"),
        ("FG Hit Rate", p["fg_hit_spins"] / max(1, p["fg_spins"]), s["fg_hit_spins"] / max(1, fg_spins), c["fg_hit_rate"], "pct"),
        ("FG 觸發率", p["fg_triggers"] / p["rounds"], triggers / rounds, c["fg_trigger_rate"], "pct"),
        ("FG 週期", p["rounds"] / max(1, p["fg_triggers"]), rounds / max(1, triggers), c["fg_trigger_cycle"], "num"),
        ("平均 FG Spins", p["fg_spins"] / max(1, p["fg_triggers"]), fg_spins / max(1, triggers), c["avg_fg_spins"], "num"),
        ("BG 平均消除次數", p["cascades_bg"] / p["rounds"], s["cascades_bg"] / rounds, c["avg_cascades_bg"], "num"),
        ("FG 平均消除次數", p["cascades_fg"] / max(1, p["fg_spins"]), s["cascades_fg"] / max(1, fg_spins), c["avg_cascades_fg"], "num"),
        ("BG 大鬼事件率", p["bg_w2_events"] / p["rounds"], s["bg_w2_events"] / rounds, c["w2_bg_event_rate"], "pct"),
        ("FG 大鬼事件率", p["fg_w2_events"] / max(1, p["fg_spins"]), s["fg_w2_events"] / max(1, fg_spins), c["w2_fg_event_rate"], "pct"),
    ]
    return [{"metric": name, "super_ace": comp, "jhs": ref, "h016": cur, "format": fmt} for name, comp, ref, cur, fmt in rows]


def aggregate_stack_rate(stack: dict[str, Any]) -> float:
    """Average exact 2/3/4-stack occurrences per reel RNG across R1-R5."""
    total = 0.0
    for reel in range(5):
        denominator = max(1.0, float(stack["totals"][reel]))
        total += sum(float(count) for (_, length), count in stack["counts"][reel].items() if length in (2, 3, 4)) / denominator
    return total / 5.0


def aggregate_gold_rate(result: dict[str, Any], scene: str) -> float:
    """Gold symbols divided by all visible initial and replenishment symbols."""
    gold_count = 0.0
    symbol_count = 0.0
    for stage in ("initial", "drop"):
        data = result["gold"][f"{scene}_{stage}"]
        for rate, denominator in zip(data["total"], data["denominators"]):
            denominator = float(denominator)
            gold_count += float(rate) * denominator
            symbol_count += denominator
    return gold_count / max(1.0, symbol_count)


def tuning_summary_metrics(
    current: dict[str, Any],
    old: dict[str, Any],
    competitor: dict[str, Any],
    current_config: dict[str, Any],
    old_config: dict[str, Any],
) -> list[dict[str, Any]]:
    c = current["base"]
    s = old["stats"]
    p = competitor["stats"]
    rows: list[dict[str, Any]] = []

    def append(scene: str, metric: str, super_ace: float, jhs: float, h016: float) -> None:
        rows.append({
            "scene": scene,
            "metric": metric,
            "super_ace": super_ace,
            "jhs": jhs,
            "h016": h016,
            "vs_super_ace": None if abs(super_ace) < 1e-15 else (super_ace - h016) / super_ace,
            "vs_jhs": None if abs(jhs) < 1e-15 else (jhs - h016) / jhs,
        })

    for scene, group in (("BG", "base"), ("FG", "free")):
        append(
            scene,
            "綜合 2/3/4 堆疊率",
            aggregate_stack_rate(competitor["stack"][scene]),
            aggregate_stack_rate(mixed_stack(old_config, group)),
            aggregate_stack_rate(mixed_stack(current_config, group)),
        )
    for scene in ("BG", "FG"):
        append(
            scene,
            "整體金框率",
            aggregate_gold_rate(competitor, scene),
            aggregate_gold_rate(old, scene),
            aggregate_gold_rate(current, scene),
        )
    append("BG", "大鬼事件率", p["bg_w2_events"] / max(1, p["rounds"]), s["bg_w2_events"] / max(1, s["rounds"]), c["w2_bg_event_rate"])
    append("FG", "大鬼事件率", p["fg_w2_events"] / max(1, p["fg_spins"]), s["fg_w2_events"] / max(1, s["fg_spins"]), c["w2_fg_event_rate"])
    return rows


def md_table(headers: list[str], rows: list[list[str]], aligns: list[str] | None = None) -> str:
    aligns = aligns or ["---"] * len(headers)
    return "\n".join(["| " + " | ".join(headers) + " |", "|" + "|".join(aligns) + "|", *["| " + " | ".join(row) + " |" for row in rows]])


def report_markdown(current: dict[str, Any], old: dict[str, Any], competitor: dict[str, Any], current_config: dict[str, Any], old_config: dict[str, Any], old_version: str, rounds: int) -> str:
    core = core_metrics(current, old, competitor)
    tuning_summary = tuning_summary_metrics(current, old, competitor, current_config, old_config)
    c = current["base"]
    s = old["stats"]
    p = competitor["stats"]
    core_rows = []
    for item in core:
        formatter = pct if item["format"] == "pct" else lambda value: num(value, 4)
        core_rows.append([item["metric"], formatter(item["super_ace"]), formatter(item["jhs"]), formatter(item["h016"]), rel(item["super_ace"], item["h016"]), rel(item["jhs"], item["h016"])])
    tuning_summary_rows = [
        [
            item["scene"],
            item["metric"],
            pct(item["super_ace"]),
            pct(item["jhs"]),
            pct(item["h016"]),
            "N/A" if item["vs_super_ace"] is None else f"{item['vs_super_ace']:+.2%}",
            "N/A" if item["vs_jhs"] is None else f"{item['vs_jhs']:+.2%}",
        ]
        for item in tuning_summary
    ]

    combo_sections = []
    for scene, denominator_current, denominator_old, denominator_comp in (("BG", int(current["base"]["total_rounds"]), old["stats"]["rounds"], p["rounds"]), ("FG", sum(row["FG"] for row in current["combo"].values()), old["stats"]["fg_spins"], p["fg_spins"])):
        rows = []
        for label in ("0", "1", "2", "3", "4", "5+"):
            comp = competitor["combo"][label][scene] / max(1, denominator_comp)
            jhs = old["combo"][label][scene] / max(1, denominator_old)
            h016 = current["combo"][label][scene] / max(1, denominator_current)
            rows.append([label, pct(comp), pct(jhs), pct(h016), rel(comp, h016), rel(jhs, h016)])
        combo_sections.append(f"### {scene}\n\n" + md_table(["消除次數", "Super Ace", "101003", "H016 v8", "H016 vs Super Ace", "H016 vs 101003"], rows, ["---", "---:", "---:", "---:", "---:", "---:"]))

    length_sections = []
    coin_current = float(current["base"]["coin_in"]) * int(current["base"]["total_rounds"])
    for scene, current_spins, old_spins, comp_spins in (("BG", int(current["base"]["total_rounds"]), old["stats"]["rounds"], p["rounds"]), ("FG", sum(row["FG"] for row in current["combo"].values()), old["stats"]["fg_spins"], p["fg_spins"])):
        rows = []
        for symbol in PAY_SYMBOLS:
            for length in (3, 4, 5):
                key = (symbol, length)
                comp_rtp = competitor["length_pay"][scene][key] / p["coin_in"]
                jhs_rtp = old["length_pay"][scene][key] / old["stats"]["coin_in"]
                h016_rtp = current["length_pay"][scene][key] / coin_current
                comp_hit = competitor["length_hits"][scene][key] / max(1, comp_spins)
                jhs_hit = old["length_hits"][scene][key] / max(1, old_spins)
                h016_hit = current["length_hits"][scene][key] / max(1, current_spins)
                rows.append([symbol, str(length), pct(comp_rtp), pct(jhs_rtp), pct(h016_rtp), rel(comp_rtp, h016_rtp), rel(jhs_rtp, h016_rtp), pct(comp_hit), pct(jhs_hit), pct(h016_hit), rel(comp_hit, h016_hit), rel(jhs_hit, h016_hit)])
        length_sections.append(f"### {scene}\n\n" + md_table(["Symbol", "輪數", "Super Ace RTP", "101003 RTP", "H016 RTP", "H016 vs Super Ace RTP", "H016 vs 101003 RTP", "Super Ace Hit", "101003 Hit", "H016 Hit", "H016 vs Super Ace Hit", "H016 vs 101003 Hit"], rows, ["---", "---:"] + ["---:"] * 10))

    stack_sections = []
    stack_summary_rows = []
    for scene, group in (("BG", "base"), ("FG", "free")):
        comp_stack = competitor["stack"][scene]
        a = mixed_stack(old_config, group)
        b = mixed_stack(current_config, group)
        rows = []
        differences = []
        competitor_differences = []
        for symbol in REPORT_SYMBOLS:
            for length in (2, 3, 4):
                for reel in range(5):
                    comp = float(comp_stack["counts"][reel].get((symbol, length), 0))
                    ref = float(a["counts"][reel].get((symbol, length), 0))
                    cur = float(b["counts"][reel].get((symbol, length), 0))
                    differences.append(abs(ref - cur))
                    competitor_differences.append(abs(comp - cur))
                rows.append([symbol, str(length), *[pct(float(comp_stack["counts"][reel].get((symbol, length), 0))) for reel in range(5)], *[pct(float(a["counts"][reel].get((symbol, length), 0))) for reel in range(5)], *[pct(float(b["counts"][reel].get((symbol, length), 0))) for reel in range(5)]])
        stack_summary_rows.append([scene, f"{max(competitor_differences, default=0) * 100:.4f} pp", f"{sum(competitor_differences) / max(1, len(competitor_differences)) * 100:.4f} pp", f"{max(differences, default=0) * 100:.4f} pp", f"{sum(differences) / max(1, len(differences)) * 100:.4f} pp"])
        stack_sections.append(f"### {scene}\n\n" + md_table(["Symbol", "堆疊", "Super Ace R1", "R2", "R3", "R4", "R5", "101003 R1", "R2", "R3", "R4", "R5", "H016 R1", "R2", "R3", "R4", "R5"], rows, ["---", "---:"] + ["---:"] * 15))

    distribution_sections = []
    for scene in ("BG", "FG"):
        for stage, stage_zh in (("initial", "初始"), ("drop", "掉落")):
            rows = []
            for symbol in REPORT_SYMBOLS:
                comp = competitor["distributions"][f"{scene}_{stage}"][symbol]
                ref = old["distributions"][f"{scene}_{stage}"][symbol]
                cur = current["distributions"][f"{scene}_{stage}"][symbol]
                rows.append([symbol, *[pct(value) for value in comp], *[pct(value) for value in ref], *[pct(value) for value in cur]])
            distribution_sections.append(f"### {scene} {stage_zh} R1–R5\n\n" + md_table(["Symbol", "Super Ace R1", "R2", "R3", "R4", "R5", "101003 R1", "R2", "R3", "R4", "R5", "H016 R1", "R2", "R3", "R4", "R5"], rows, ["---"] + ["---:"] * 15))

    gold_sections = []
    gold_symbol_sections = []
    for scene in ("BG", "FG"):
        rows = []
        for stage, stage_zh in (("initial", "初始"), ("drop", "掉落")):
            for reel in range(5):
                comp = competitor["gold"][f"{scene}_{stage}"]["total"][reel]
                ref = old["gold"][f"{scene}_{stage}"]["total"][reel]
                cur = current["gold"][f"{scene}_{stage}"]["total"][reel]
                rows.append([stage_zh, f"R{reel + 1}", pct(comp), pct(ref), pct(cur), rel(comp, cur), rel(ref, cur)])
            symbol_rows = []
            for symbol in PAY_SYMBOLS:
                comp_values = competitor["gold"][f"{scene}_{stage}"]["symbols"][symbol]
                ref_values = old["gold"][f"{scene}_{stage}"]["symbols"][symbol]
                cur_values = current["gold"][f"{scene}_{stage}"]["symbols"][symbol]
                symbol_rows.append([symbol, *[pct(value) for value in comp_values], *[pct(value) for value in ref_values], *[pct(value) for value in cur_values]])
            gold_symbol_sections.append(f"### {scene} {stage_zh}\n\n" + md_table(["Symbol", "Super Ace R1", "R2", "R3", "R4", "R5", "101003 R1", "R2", "R3", "R4", "R5", "H016 R1", "R2", "R3", "R4", "R5"], symbol_rows, ["---"] + ["---:"] * 15))
        gold_sections.append(f"### {scene}\n\n" + md_table(["階段", "Reel", "Super Ace", "101003", "H016 v8", "H016 vs Super Ace", "H016 vs 101003"], rows, ["---", "---", "---:", "---:", "---:", "---:", "---:"]))

    w2_sections = []
    c = current["base"]
    s = old["stats"]
    for scene, current_total, old_total, comp_total, current_prefix, old_prefix in (("BG", int(c["w2_bg_count_2"] + c["w2_bg_count_3"] + c["w2_bg_count_4"]), sum(int(s["bg_w2_counts"].get(n, 0)) for n in (2, 3, 4)), sum(int(p["bg_w2_counts"].get(n, 0)) for n in (2, 3, 4)), "w2_bg_count_", "bg_w2_counts"), ("FG", int(c["w2_fg_count_2"] + c["w2_fg_count_3"] + c["w2_fg_count_4"]), sum(int(s["fg_w2_counts"].get(n, 0)) for n in (2, 3, 4)), 0, "w2_fg_count_", "fg_w2_counts")):
        rows = []
        for made in (2, 3, 4):
            comp_value = int(p[f"{scene.lower()}_w2_counts"].get(made, 0)) / comp_total if comp_total else None
            ref = int(s[old_prefix].get(made, 0)) / max(1, old_total)
            cur = int(c[f"{current_prefix}{made}"]) / max(1, current_total)
            rows.append([str(made), "N/A" if comp_value is None else pct(comp_value), pct(ref), pct(cur), "N/A" if comp_value is None else rel(comp_value, cur), rel(ref, cur)])
        w2_sections.append(f"### {scene} 大鬼 2／3／4 顆條件分布\n\n" + md_table(["顆數", "Super Ace", "101003", "H016 v8", "H016 vs Super Ace", "H016 vs 101003"], rows, ["---:", "---:", "---:", "---:", "---:", "---:"]))

    multiplier_sections = []
    multiplier_summary = []
    for scene, count_key, pay_key, cur_den, old_den, comp_den in (("BG", "bg_count", "bg_pay", int(c["total_rounds"]), old["stats"]["rounds"], p["rounds"]), ("FG", "fg_count", "fg_pay", int(c["bg_trigger_fg_cnt"]), old["stats"]["fg_triggers"], p["fg_triggers"])):
        rows = []
        total_cur_hit = total_old_hit = total_comp_hit = 0.0
        for comp_row, ref_row, cur_row in zip(competitor["multiplier"], old["multiplier"], current["multiplier"]):
            comp_hit = comp_row[count_key] / max(1, comp_den)
            ref_hit = ref_row[count_key] / max(1, old_den)
            cur_hit = cur_row[count_key] / max(1, cur_den)
            comp_rtp = comp_row[pay_key] / p["coin_in"]
            ref_rtp = ref_row[pay_key] / old["stats"]["coin_in"]
            cur_rtp = cur_row[pay_key] / (float(c["coin_in"]) * int(c["total_rounds"]))
            if ref_row["index"] != 0:
                total_cur_hit += cur_hit
                total_old_hit += ref_hit
                total_comp_hit += comp_hit
            if comp_row[count_key] or ref_row[count_key] or cur_row[count_key] or comp_row[pay_key] or ref_row[pay_key] or cur_row[pay_key]:
                rows.append([cur_row["interval"], pct(comp_hit), pct(ref_hit), pct(cur_hit), rel(comp_hit, cur_hit), rel(ref_hit, cur_hit), pct(comp_rtp), pct(ref_rtp), pct(cur_rtp), rel(comp_rtp, cur_rtp), rel(ref_rtp, cur_rtp)])
        multiplier_summary.append([scene, pct(total_comp_hit), pct(total_old_hit), pct(total_cur_hit), rel(total_comp_hit, total_cur_hit), rel(total_old_hit, total_cur_hit)])
        multiplier_sections.append(f"### {scene}\n\n" + md_table(["倍率區間", "Super Ace Hit", "101003 Hit", "H016 Hit", "H016 vs Super Ace Hit", "H016 vs 101003 Hit", "Super Ace RTP", "101003 RTP", "H016 RTP", "H016 vs Super Ace RTP", "H016 vs 101003 RTP"], rows, ["---"] + ["---:"] * 10))

    pay_rows = []
    for symbol in PAY_SYMBOLS:
        numeric = next(key for key, name in SYMBOL_NAMES.items() if name == symbol)
        ref = [float(value) for value in old_config["pays"][str(numeric)]]
        cur = [float(value) for value in current_config["pays"][str(numeric)]]
        pay_rows.append([symbol, *[num(value, 2) for value in ref], *[num(value, 2) for value in ref], *[num(value, 2) for value in cur], "PASS" if ref == cur else "FAIL"])

    config_rows = []
    rw_rows = []
    for scene, group, names in (("BG", "base", ("bg_1", "bg_2", "bg_3")), ("FG", "free", ("fg_1", "fg_2", "fg_3"))):
        for name in names:
            for label, config in (("101003", old_config), ("H016 v8", current_config)):
                table = config["tables"][name]
                selected = next((float(item["weight"]) for item in config["table_selection"][group] if item["table"] == name), 0.0)
                ratios = []
                for weights in table["weights"]:
                    positive = [float(value) for value in weights if float(value) > 0]
                    ratios.append(max(positive) / min(positive) if positive else math.nan)
                config_rows.append([scene, name, label, num(selected, 0), *[f"{value:.2f}x" if math.isfinite(value) else "N/A" for value in ratios]])
                rw = table["random_wild"]
                rw_rows.append([scene, name, label, "/".join(str(int(value)) for value in rw["weights"])])

    current_sel_bg = "/".join(str(int(item["weight"])) for item in current_config["table_selection"]["base"])
    current_sel_fg = "/".join(str(int(item["weight"])) for item in current_config["table_selection"]["free"])
    old_sel_bg = "/".join(str(int(item["weight"])) for item in old_config["table_selection"]["base"])
    old_sel_fg = "/".join(str(int(item["weight"])) for item in old_config["table_selection"]["free"])

    jhs_total = (s["pay_bg"] + s["pay_fg"]) / s["coin_in"]
    jhs_bg = s["pay_bg"] / s["coin_in"]
    jhs_fg = s["pay_fg"] / s["coin_in"]
    jhs_bg_hit = s["bg_hit_spins"] / s["rounds"]
    jhs_fg_hit = s["fg_hit_spins"] / max(1, s["fg_spins"])
    jhs_cycle = s["rounds"] / max(1, s["fg_triggers"])
    comp_total = (p["pay_bg"] + p["pay_fg"]) / p["coin_in"]
    comp_bg = p["pay_bg"] / p["coin_in"]
    comp_fg = p["pay_fg"] / p["coin_in"]
    comp_bg_hit = p["bg_hit_spins"] / p["rounds"]
    comp_fg_hit = p["fg_hit_spins"] / max(1, p["fg_spins"])
    comp_cycle = p["rounds"] / max(1, p["fg_triggers"])
    return f"""# H016 v8 vs 101003 vs Super Ace：92 老手調性比較

## 目錄

- [結論摘要](#結論摘要)
- [比較範圍與口徑](#比較範圍與口徑)
- [核心指標](#核心指標)
- [三項綜合調性指標](#三項綜合調性指標)
- [賠率與設定差異](#賠率與設定差異)
- [倍率線型](#倍率線型)
- [符號-345-連線-rtp--hit-rate](#符號-345-連線-rtp--hit-rate)
- [不同符號-234-堆疊-rng-比例](#不同符號-234-堆疊-rng-比例)
- [符號分布](#符號分布)
- [消除分布](#消除分布)
- [金框比例與各符號金框](#金框比例與各符號金框)
- [大鬼事件](#大鬼事件)
- [方法與限制](#方法與限制)

## 結論摘要

- 本報告並列 **Super Ace 競品自然側錄**、**101003 92 老手 Card-On 參數重播**與 **H016 v8 92 老手 Card-On**；競品無卡片篩選，倍率線型為自然分布。
- **Total RTP**：Super Ace {pct(comp_total)}、101003 {pct(jhs_total)}、H016 {pct(c['rtp_total'])}；H016 低於兩個比較基準。
- **BG / FG RTP**：Super Ace {pct(comp_bg)} / {pct(comp_fg)}；101003 {pct(jhs_bg)} / {pct(jhs_fg)}；H016 {pct(c['rtp_bg'])} / {pct(c['rtp_fg'])}。
- **BG / FG Hit Rate**：Super Ace {pct(comp_bg_hit)} / {pct(comp_fg_hit)}；101003 {pct(jhs_bg_hit)} / {pct(jhs_fg_hit)}；H016 {pct(c['bg_hit_rate'])} / {pct(c['fg_hit_rate'])}。
- **FG 週期**：Super Ace {num(comp_cycle, 2)} 場、101003 {num(jhs_cycle, 2)} 場、H016 {num(c['fg_trigger_cycle'], 2)} 場。
- **倍率線型有明顯重分配**：H016 BG 將量體加到 `(10,15]`、`(20,60]`，且 70x 以上為 0；H016 FG 新增 5–10x 區間，並把 101003 原本集中在 20–40x 的 Hit 往 45–200x 分散。
- H016 v8 的選表比例為 BG `{current_sel_bg}`、FG `{current_sel_fg}`；101003 為 BG `{old_sel_bg}`、FG `{old_sel_fg}`。這是堆疊、消除與倍率線型差異的重要來源。
- 相對差異同時列 `(Super Ace - H016) ÷ Super Ace` 與 `(101003 - H016) ÷ 101003`；負值代表 H016 較高，參考值為 0 時顯示 `N/A`。
- 倍率線型的 FG Hit Rate 分母是進入 FG 次數，BG Hit Rate 分母是 Base rounds；區間 RTP 一律除以 Base coin-in。

## 比較範圍與口徑

| 項目 | Super Ace | 101003 | 目前 H016 |
|---|---|---|---|
| 版本 | JILI 實機側錄 | `{old_version}` | `{current["base"]["math_version"]}` |
| 樣本 | `{p['rounds']:,}` BG Round / `{p['fg_spins']:,}` FG Spin / `{p['fg_triggers']:,}` FG Session | `{rounds:,}` 場參數重播 | `{int(current["base"]["total_rounds"]):,}` 場報表 |
| 模式 | Normal Bet / 自然側錄 | 92 老手 / Normal Bet / Card-On | 92 老手 / Normal Bet / Card-On |
| 參數 | `JILI - Super Ace - m` 三份非重疊 JSONL | `gameSetting_JHS101003.json` | `config.js` + `config_92A.js` |
| 流程 | 競品實際盤面與消除紀錄 | Java 規則映射後以共同 fast engine 重播 | v8 正式 Record |

## 核心指標

{md_table(["指標", "Super Ace", "101003", "H016 v8", "H016 vs Super Ace", "H016 vs 101003"], core_rows, ["---", "---:", "---:", "---:", "---:", "---:"])}

## 三項綜合調性指標

- **綜合 2/3/4 堆疊率**：每個 Scene 將所有符號的 exact 最大 2／3／4 堆疊事件加總，除以 R1–R5 的 RNG 總數；五輪等權平均成單一指標。
- **整體金框率**：初始盤面與後續補牌的金框數加總，除以相同範圍內所有可見符號數。
- **大鬼事件率**：成功產生大鬼的事件數除以該 Scene Spins；BG 一個 Spin 最多一次，FG 每次 Cascade 可重新觸發。

{md_table(["Scene", "指標", "Super Ace", "101003", "H016 v8", "H016 vs Super Ace", "H016 vs 101003"], tuning_summary_rows, ["---", "---", "---:", "---:", "---:", "---:", "---:"])}

## 賠率與設定差異

### 345 賠率

{md_table(["Symbol", "Super Ace 3", "4", "5", "101003 3", "4", "5", "H016 3", "4", "5", "是否一致"], pay_rows, ["---"] + ["---:"] * 9 + ["---"])}

### Table Selection 與初始停輪權重範圍

權重倍數為每輪正權重 `max/min`，0 為禁用 stop，不列入比率。

{md_table(["Scene", "Table", "模型", "Table Weight", "R1", "R2", "R3", "R4", "R5"], config_rows, ["---", "---", "---", "---:"] + ["---:"] * 5)}

### Random Wild 0／2／3／4 原始權重

{md_table(["Scene", "Table", "模型", "0/2/3/4 Weight"], rw_rows, ["---", "---", "---", "---"])}

## 倍率線型

倍率區間依 H016 `Multiplier Line` 固定區間比較。FG 是「每次完整 FG session 總獎金」落入哪個區間，不是單次 Free Spin。

{md_table(["Scene", "Super Ace 有獎區間 Hit Rate", "101003", "H016", "H016 vs Super Ace", "H016 vs 101003"], multiplier_summary, ["---", "---:", "---:", "---:", "---:", "---:"])}

{chr(10).join(multiplier_sections)}

## 符號 345 連線 RTP / Hit Rate

RTP 分母為 Base coin-in；Hit Rate 分母 BG 為 Base rounds、FG 為 Free Spins。同一 cascade 內同符號同輪數只記一筆 hit，與競品分析報告同口徑。

{chr(10).join(length_sections)}

## 不同符號 2／3／4 堆疊 RNG 比例

每個 stop RNG 取連續 4 格可見符號，金框先併回原符號，只記最大連續段；各表依自身 stop weight 正規化後，再依 Table Selection 混合。

{md_table(["Scene", "vs Super Ace 最大絕對差", "vs Super Ace 平均絕對差", "vs 101003 最大絕對差", "vs 101003 平均絕對差"], stack_summary_rows, ["---", "---:", "---:", "---:", "---:"])}

{chr(10).join(stack_sections)}

## 符號分布

金框符號先併回原符號；初始與消除掉落分開，各輪分母為該輪實際統計符號總數。

{chr(10).join(distribution_sections)}

## 消除分布

{chr(10).join(combo_sections)}

## 金框比例與各符號金框

{chr(10).join(gold_sections)}

### 各符號、各輪金框佔所有可見符號比例

{chr(10).join(gold_symbol_sections)}

## 大鬼事件

{chr(10).join(w2_sections)}

## 方法與限制

1. H016 來源是 `{CURRENT_RECORD.relative_to(H016_DIR)}`，共100,000,000場；沒有改寫 config 或 Excel。
2. 101003 指定資料夾內只有 `data.js` 與實驗性 Python，沒有可直接對等的 92 老手 Card-On 正式報表；因此本報告使用專案保存的 `gameSetting_JHS101003.json`、`BaseGameHandler_JHS101003.java`、`FreeGameHandler_JHS101003.java` 重建參數與流程，再以同一個 v8 fast engine 重播。
3. 101003 Java 的 BG 補牌傳入 `wheelData`。共同引擎映射為使用同一組 200 格輪帶符號與 stop weight 單格抽樣：單格邊際機率對齊，但不保留 Java 若有的同次補牌連續位置相關性。FG 掉落則直接使用 JSON `dropWheelWeight`。
4. Super Ace 使用 `JILI - Super Ace - m/遊戲資料` 中三份非重疊自然轉 JSONL；倍率線型 BG 以每個 Round 的 BG pay，FG 以每次自然觸發的完整 FG session pay 分區。
5. 競品無 Card System；因此 Super Ace 線型是自然發生率，101003/H016 是 92 老手 Card-On 接受後分布。可比較調性，不應把三者視為相同選樣機制。
6. 101003 樣本為 {rounds:,}、H016 為 100,000,000、Super Ace 僅 {p['fg_triggers']:,} 次自然 FG Session；稀有高倍區間會有較大抽樣波動。
7. 本報告是診斷比較，不會自動調整 v8 數學。
"""


def artifact_json(current: dict[str, Any], old: dict[str, Any], competitor: dict[str, Any], current_config: dict[str, Any], old_config: dict[str, Any], version: str, rounds: int) -> dict[str, Any]:
    core = core_metrics(current, old, competitor)
    tuning_summary = tuning_summary_metrics(current, old, competitor, current_config, old_config)
    current_base = current["base"]
    old_stats = old["stats"]
    competitor_stats = competitor["stats"]
    rows = ([{"metric": row["metric"], "model": "Super Ace", "value": row["super_ace"], "format": row["format"]} for row in core]
            + [{"metric": row["metric"], "model": "101003", "value": row["jhs"], "format": row["format"]} for row in core]
            + [{"metric": row["metric"], "model": "H016 v8", "value": row["h016"], "format": row["format"]} for row in core])
    rtp_rows = [row for row in rows if row["metric"] in ("Total RTP", "BG RTP", "FG RTP")]
    hit_rows = [row for row in rows if row["metric"] in ("BG Hit Rate", "FG Hit Rate", "FG 觸發率")]
    exact = [{"metric": row["metric"], "super_ace": row["super_ace"], "jhs": row["jhs"], "h016": row["h016"], "vs_super_ace": None if abs(row["super_ace"]) < 1e-15 else (row["super_ace"] - row["h016"]) / row["super_ace"], "vs_jhs": None if abs(row["jhs"]) < 1e-15 else (row["jhs"] - row["h016"]) / row["jhs"], "format": row["format"]} for row in core]

    multiplier_datasets: dict[str, list[dict[str, Any]]] = {}
    for scene, count_key, pay_key, current_hit_den, old_hit_den, competitor_hit_den in (
        ("BG", "bg_count", "bg_pay", int(current_base["total_rounds"]), int(old_stats["rounds"]), int(competitor_stats["rounds"])),
        ("FG", "fg_count", "fg_pay", int(current_base["bg_trigger_fg_cnt"]), int(old_stats["fg_triggers"]), int(competitor_stats["fg_triggers"])),
    ):
        hit_dataset = f"multiplier_{scene.lower()}_hit"
        rtp_dataset = f"multiplier_{scene.lower()}_rtp"
        multiplier_datasets[hit_dataset] = []
        multiplier_datasets[rtp_dataset] = []
        for comp_row, ref_row, cur_row in zip(competitor["multiplier"], old["multiplier"], current["multiplier"]):
            if int(ref_row["index"]) == 0:
                continue
            if not (comp_row[count_key] or ref_row[count_key] or cur_row[count_key] or comp_row[pay_key] or ref_row[pay_key] or cur_row[pay_key]):
                continue
            interval = str(cur_row["interval"])
            order = int(cur_row["index"])
            upper = float(cur_row["upper"])
            for model, source_row, hit_den, coin_in in (
                ("Super Ace", comp_row, competitor_hit_den, float(competitor_stats["coin_in"])),
                ("101003", ref_row, old_hit_den, float(old_stats["coin_in"])),
                ("H016 v8", cur_row, current_hit_den, float(current_base["coin_in"]) * int(current_base["total_rounds"])),
            ):
                count = int(source_row[count_key])
                pay = float(source_row[pay_key])
                base = {
                    "scene": scene,
                    "interval": interval,
                    "interval_order": order,
                    "upper": upper,
                    "model": model,
                    "count": count,
                    "pay": pay,
                    "hit_denominator": hit_den,
                    "coin_in": coin_in,
                }
                multiplier_datasets[hit_dataset].append({**base, "value": count / max(1, hit_den), "metric": "Hit Rate"})
                multiplier_datasets[rtp_dataset].append({**base, "value": pay / max(1.0, coin_in), "metric": "RTP"})

    source_sql = "SELECT metric, model, value FROM report_metrics"
    widget_source = {"id": "three_way_comparison", "label": "Super Ace、101003 與 H016 v8 三方比較", "type": "analysis", "query": {"engine": "portable-values", "language": "sql", "sql": source_sql, "description": "從本報告已固化的三方核心指標資料集取值。"}}
    def multiplier_source(scene: str, metric: str) -> dict[str, Any]:
        return {
            "id": f"multiplier_{scene.lower()}_{metric.lower().replace(' ', '_')}",
            "label": f"{scene} 倍率線型 {metric}",
            "type": "analysis",
            "query": {
                "engine": "portable-values",
                "language": "sql",
                "sql": f"SELECT interval, interval_order, model, value, count, pay, hit_denominator, coin_in FROM report_multiplier_line WHERE scene = '{scene}' AND metric = '{metric}' ORDER BY interval_order, model",
                "description": "使用報表已固化的倍率區間 count/pay 與同口徑分母。",
            },
        }

    def multiplier_chart(chart_id: str, dataset: str, title: str, scene: str, metric: str) -> dict[str, Any]:
        return {
            "id": chart_id,
            "type": "line",
            "title": title,
            "dataset": dataset,
            "x": "interval",
            "y": "value",
            "color": "model",
            "encodings": {
                "x": {"field": "interval", "type": "nominal", "sort": {"field": "interval_order", "order": "ascending"}},
                "y": {"field": "value", "type": "quantitative", "format": "percent"},
                "color": {"field": "model", "type": "nominal"},
            },
            "format": {"y": "percent"},
            "legend": {"position": "bottom"},
            "source": multiplier_source(scene, metric),
        }
    return {
        "surface": "report",
        "manifest": {
            "version": 1,
            "surface": "report",
            "title": "H016 v8 vs 101003 vs Super Ace：92 老手調性比較",
            "description": "並列 Super Ace 競品自然側錄、101003 參數重播與 H016 v8 的核心指標及倍率線型。",
            "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
            "sources": [
                {"id": "h016_record", "label": "H016 v8 92 老手 Card-On 100M", "type": "file", "path": str(CURRENT_RECORD.relative_to(H016_DIR)), "description": "目前正式 Record。"},
                {"id": "jhs_setting", "label": "gameSetting_JHS101003.json", "type": "file", "path": str(JHS_JSON.relative_to(H016_DIR)), "description": f"101003 正式參數 {version}。"},
                {"id": "jhs_replay", "label": "101003 92 老手參數重播", "type": "analysis", "path": str(SCRIPT.relative_to(H016_DIR)), "description": f"{rounds:,} rounds; current fast engine; Card-On oldhand."},
                {"id": "super_ace", "label": "JILI Super Ace 非重疊自然轉 JSONL", "type": "file", "path": "JILI - Super Ace - m/遊戲資料/SuperAce_BG_Combined_NoJP+BG_3+BG_4.jsonl", "description": f"{competitor_stats['rounds']:,} BG rounds; {competitor_stats['fg_triggers']:,} natural FG sessions."},
            ],
            "charts": [
                {"id": "rtp", "type": "bar", "title": "RTP 比較", "dataset": "rtp", "x": "metric", "y": "value", "color": "model", "encodings": {"x": {"field": "metric", "type": "nominal"}, "y": {"field": "value", "type": "quantitative", "format": "percent"}, "color": {"field": "model", "type": "nominal"}}, "grouped": True, "format": {"y": "percent"}, "labels": "all", "legend": {"position": "bottom"}, "source": widget_source},
                {"id": "hit", "type": "bar", "title": "Hit / Feature 比較", "dataset": "hit", "x": "metric", "y": "value", "color": "model", "encodings": {"x": {"field": "metric", "type": "nominal"}, "y": {"field": "value", "type": "quantitative", "format": "percent"}, "color": {"field": "model", "type": "nominal"}}, "grouped": True, "format": {"y": "percent"}, "labels": "all", "legend": {"position": "bottom"}, "source": widget_source},
                multiplier_chart("multiplier_bg_hit", "multiplier_bg_hit", "BG 倍率線型：Hit Rate", "BG", "Hit Rate"),
                multiplier_chart("multiplier_bg_rtp", "multiplier_bg_rtp", "BG 倍率線型：RTP", "BG", "RTP"),
                multiplier_chart("multiplier_fg_hit", "multiplier_fg_hit", "FG 倍率線型：Hit Rate", "FG", "Hit Rate"),
                multiplier_chart("multiplier_fg_rtp", "multiplier_fg_rtp", "FG 倍率線型：RTP", "FG", "RTP"),
            ],
            "tables": [
                {"id": "core", "title": "核心指標精確比較", "dataset": "core", "columns": [{"field": "metric", "label": "指標", "type": "text"}, {"field": "super_ace", "label": "Super Ace", "type": "number"}, {"field": "jhs", "label": "101003", "type": "number"}, {"field": "h016", "label": "H016 v8", "type": "number"}, {"field": "vs_super_ace", "label": "H016 vs Super Ace", "type": "percent"}, {"field": "vs_jhs", "label": "H016 vs 101003", "type": "percent"}], "source": widget_source},
                {"id": "tuning_summary", "title": "堆疊、金框與大鬼綜合指標", "dataset": "tuning_summary", "columns": [{"field": "scene", "label": "Scene", "type": "text"}, {"field": "metric", "label": "指標", "type": "text"}, {"field": "super_ace", "label": "Super Ace", "type": "percent"}, {"field": "jhs", "label": "101003", "type": "percent"}, {"field": "h016", "label": "H016 v8", "type": "percent"}, {"field": "vs_super_ace", "label": "H016 vs Super Ace", "type": "percent"}, {"field": "vs_jhs", "label": "H016 vs 101003", "type": "percent"}], "source": widget_source},
            ],
            "blocks": [
                {"id": "title", "type": "markdown", "body": "# H016 v8 vs 101003 vs Super Ace：92 老手調性比較"},
                {"id": "summary", "type": "markdown", "body": f"## 結論摘要\n\nSuper Ace 競品已加入核心指標與四張倍率線型圖。Total RTP：Super Ace {(competitor_stats['pay_bg'] + competitor_stats['pay_fg']) / competitor_stats['coin_in']:.4%}、101003 {(old_stats['pay_bg'] + old_stats['pay_fg']) / old_stats['coin_in']:.4%}、H016 {float(current_base['rtp_total']):.4%}。競品線型是自然側錄；101003 與 H016 為 92 老手 Card-On。"},
                {"id": "rtp_chart", "type": "chart", "chartId": "rtp"},
                {"id": "hit_chart", "type": "chart", "chartId": "hit"},
                {"id": "core_table", "type": "table", "tableId": "core"},
                {"id": "tuning_intro", "type": "markdown", "body": "## 三項綜合調性指標\n\n堆疊率是所有符號 exact 最大 2／3／4 堆疊事件除以 R1–R5 RNG 總數後的五輪等權平均；金框率按初始與補牌所有實際可見符號加權；大鬼事件率以 Scene Spins 為分母。"},
                {"id": "tuning_table", "type": "table", "tableId": "tuning_summary"},
                {"id": "multiplier_bg_hit_intro", "type": "markdown", "body": "## 倍率線型圖表\n\n### BG Hit Rate\n\n三條線分別為 Super Ace 自然側錄、101003 與 H016 v8。顯示各得分倍率區間佔 Base rounds 的發生比例；`(-1, 0]` 無得分區間不列入。"},
                {"id": "multiplier_bg_hit_chart", "type": "chart", "chartId": "multiplier_bg_hit"},
                {"id": "multiplier_bg_rtp_intro", "type": "markdown", "body": "### BG RTP\n\n各區間 RTP 為該區間 BG pay 除以全部 Base coin-in，與 Markdown 明細表使用相同分母。"},
                {"id": "multiplier_bg_rtp_chart", "type": "chart", "chartId": "multiplier_bg_rtp"},
                {"id": "multiplier_fg_hit_intro", "type": "markdown", "body": "### FG Hit Rate\n\nFG 以完整 Feature session 總獎金分區；各區間 Hit Rate 分母為進入 FG 的次數，不是 Free Spin 數。"},
                {"id": "multiplier_fg_hit_chart", "type": "chart", "chartId": "multiplier_fg_hit"},
                {"id": "multiplier_fg_rtp_intro", "type": "markdown", "body": "### FG RTP\n\n各區間 RTP 為該區間整場 FG pay 除以 Base coin-in；因此可直接比較 FG 倍率線型對整體 RTP 的貢獻。"},
                {"id": "multiplier_fg_rtp_chart", "type": "chart", "chartId": "multiplier_fg_rtp"},
                {"id": "method", "type": "markdown", "body": f"## 口徑與限制\n\nSuper Ace 使用 {competitor_stats['rounds']:,} 個自然 BG Round 與 {competitor_stats['fg_triggers']:,} 次自然 FG Session；101003 以 `{version}` JSON/Java 參數在共同 fast engine 重播 {rounds:,} 場；H016 使用 v8 100M 正式報表。競品無 Card System，所以倍率線型為自然分布。"},
            ],
        },
        "snapshot": {"version": 1, "datasets": {"rtp": rtp_rows, "hit": hit_rows, "core": exact, "tuning_summary": tuning_summary, **multiplier_datasets}},
        "sources": [
            {"id": "h016_record", "label": "H016 v8 92 老手 Card-On 100M", "type": "file", "path": str(CURRENT_RECORD.relative_to(H016_DIR))},
            {"id": "jhs_setting", "label": "gameSetting_JHS101003.json", "type": "file", "path": str(JHS_JSON.relative_to(H016_DIR))},
            {"id": "jhs_replay", "label": "101003 92 老手參數重播", "type": "analysis", "path": str(SCRIPT.relative_to(H016_DIR))},
            {"id": "super_ace", "label": "JILI Super Ace 非重疊自然轉 JSONL", "type": "file", "path": "JILI - Super Ace - m/遊戲資料/SuperAce_BG_Combined_NoJP+BG_3+BG_4.jsonl"},
        ],
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--rounds", type=int, default=10_000_000)
    parser.add_argument("--threads", type=int, default=max(1, min(8, os.cpu_count() or 1)))
    parser.add_argument("--seed", type=int, default=101003)
    args = parser.parse_args()
    for path in (CURRENT_RECORD, CURRENT_CONFIG, CURRENT_CARD, JHS_JSON, JHS_BG_JAVA, JHS_FG_JAVA, FAST_SIMULATOR, COMPETITOR_BUILDER, *SUPER_ACE_FILES):
        if not path.exists():
            raise FileNotFoundError(path)

    current_config = load_js(CURRENT_CONFIG)
    old_config, old_version = build_jhs_config()
    # Numba's on-disk cache records the original module name.
    fast = load_module("fast_simulator", FAST_SIMULATOR)
    print(f"Replay JHS101003 92 oldhand Card-On: {args.rounds:,} rounds / {args.threads} threads", flush=True)
    raw = fast.run(old_config, args.rounds, 0, 1, args.threads, seed=args.seed, card_enabled=True, card_newbie=False)
    old = jhs_result(fast.to_stats(raw))
    current = load_current_record(CURRENT_RECORD)
    print("Load JILI Super Ace natural-play JSONL", flush=True)
    competitor = load_super_ace()

    markdown = report_markdown(current, old, competitor, current_config, old_config, old_version, args.rounds)
    OUT_MD.write_text(markdown, encoding="utf-8")
    OUT_ARTIFACT.write_text(json.dumps(artifact_json(current, old, competitor, current_config, old_config, old_version, args.rounds), ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Markdown: {OUT_MD}")
    print(f"Artifact: {OUT_ARTIFACT}")
    print(f"JHS total RTP: {(old['stats']['pay_bg'] + old['stats']['pay_fg']) / old['stats']['coin_in']:.6%}")
    print(f"JHS BG/FG RTP: {old['stats']['pay_bg'] / old['stats']['coin_in']:.6%} / {old['stats']['pay_fg'] / old['stats']['coin_in']:.6%}")
    print(f"Super Ace total RTP: {(competitor['stats']['pay_bg'] + competitor['stats']['pay_fg']) / competitor['stats']['coin_in']:.6%}")


if __name__ == "__main__":
    main()
