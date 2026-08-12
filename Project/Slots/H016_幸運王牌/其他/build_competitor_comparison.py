from __future__ import annotations

import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


H016_DIR = Path(__file__).resolve().parent.parent
OUTPUT = H016_DIR / "其他" / "競品參考數值比較.md"
RECORD_DIR = H016_DIR / "Record"
MODEL_XLSX = H016_DIR / "Source" / "H0161.xlsx"
COMPETITOR_DIR = Path(
    "C:/Users/rhinshen/Mine/個人工作區/市場資訊/H5/遊戲資源/JILI/"
    "JILI - Super Ace/遊戲資料"
)
COMPETITOR_STRIP_XLSX = COMPETITOR_DIR / "StripTable_SuperAce_還原.xlsx"
FILES = [
    COMPETITOR_DIR / "SuperAce_BG_Combined_NoJP.jsonl",
    COMPETITOR_DIR / "SuperAce_BG_3.jsonl",
    COMPETITOR_DIR / "Super_Ace_BG_4.jsonl",
]
SYMBOL_MAP = {
    "Bonus": "C1", "Symbol1": "M1", "Symbol2": "M2", "Symbol3": "M3",
    "Symbol4": "M4", "Symbol5": "A", "Symbol6": "K", "Symbol7": "Q", "Symbol8": "J",
}
SYMBOLS = ["C1", "M1", "M2", "M3", "M4", "A", "K", "Q", "J"]
GOLD_TO_BASE = {"G1": "M1", "G2": "M2", "G3": "M3", "G4": "M4", "GA": "A", "GK": "K", "GQ": "Q", "GJ": "J"}


def pct(value: float, digits: int = 4) -> str:
    return f"{value * 100:.{digits}f}%"


def pp(value: float) -> str:
    return f"{value * 100:+.4f} pp".replace("+0.0000", "0.0000")


def raw_competitor() -> dict[str, Any]:
    counts = {
        scene: {
            "initial": [Counter() for _ in range(5)], "drop": [Counter() for _ in range(5)],
            "gold_initial": Counter(), "gold_drop": Counter(),
            "gold_symbol_initial": [Counter() for _ in range(5)],
            "gold_symbol_drop": [Counter() for _ in range(5)],
            "initial_total": Counter(), "drop_total": Counter(),
            "combo": Counter(), "symbol_length_hits": Counter(),
            "symbol_length_win": Counter(), "spins": 0, "hits": 0,
        }
        for scene in ("BG", "FG")
    }
    coin_in = total_win = bg_win = fg_win = 0.0
    triggers = big_events = 0
    for path in FILES:
        with path.open(encoding="utf-8") as handle:
            for line in handle:
                obj = json.loads(line)
                plates = obj["plate"]["plate"]
                coin_in += float(obj["bet"])
                total_win += float(obj["win"])
                triggers += int(len(plates) > 1)
                for plate_index, plate in enumerate(plates):
                    scene = "BG" if plate_index == 0 else "FG"
                    data = counts[scene]
                    data["spins"] += 1
                    win = float(plate.get("win", 0.0))
                    data["hits"] += int(win > 0)
                    if scene == "BG":
                        bg_win += win
                    else:
                        fg_win += win
                    for reel, column in enumerate(plate["column"]):
                        for symbol, gold in zip(column["row"], column["isGold"]):
                            mapped = SYMBOL_MAP[symbol]
                            data["initial"][reel][mapped] += 1
                            data["initial_total"][reel] += 1
                            data["gold_initial"][reel] += int(gold in (1, 2))
                            data["gold_symbol_initial"][reel][mapped] += int(gold in (1, 2))
                    combos = plate.get("combo", [])
                    data["combo"][min(len(combos), 5)] += 1
                    for combo in combos:
                        combo_bonus = float(combo.get("comboBonus", 1.0))
                        for award in combo.get("award", []):
                            symbol = SYMBOL_MAP.get(str(award.get("symbol")))
                            length = int(award.get("maxLen", 0))
                            if symbol in SYMBOLS and length in (3, 4, 5):
                                key = (symbol, length)
                                data["symbol_length_hits"][key] += 1
                                data["symbol_length_win"][key] += float(award.get("win", 0.0)) * combo_bonus
                        if scene == "BG" and any(change.get("isGold") == 102 for change in combo.get("change", [])):
                            big_events += 1
                        for change in combo.get("change", []):
                            if "symbol" not in change:
                                continue
                            reel = int(change.get("column", 0))
                            mapped = SYMBOL_MAP[change["symbol"]]
                            data["drop"][reel][mapped] += 1
                            data["drop_total"][reel] += 1
                            data["gold_drop"][reel] += int(change.get("isGold") in (1, 2))
                            data["gold_symbol_drop"][reel][mapped] += int(change.get("isGold") in (1, 2))
    bg_spins = counts["BG"]["spins"]
    fg_spins = counts["FG"]["spins"]
    return {
        "counts": counts, "rounds": bg_spins, "fg_spins": fg_spins, "coin_in": coin_in,
        "rtp_total": total_win / coin_in, "rtp_bg": bg_win / coin_in, "rtp_fg": fg_win / coin_in,
        "bg_hit_rate": counts["BG"]["hits"] / bg_spins, "fg_hit_rate": counts["FG"]["hits"] / fg_spins,
        "fg_trigger_rate": triggers / bg_spins, "avg_fg_spins": fg_spins / triggers,
        "w2_bg_event_rate": big_events / bg_spins,
    }


def record_data(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    summary = {row[0]: row[1] for row in workbook["Base Info"].iter_rows(min_row=2, values_only=True)}
    combo = {str(row[0]): {"BG": int(row[1]), "FG": int(row[2])} for row in workbook["Eliminate"].iter_rows(min_row=2, values_only=True)}
    ratios: dict[str, dict[str, list[float]]] = {}
    gold_ratios: dict[str, list[float]] = {}
    for scene, stage, sheet_name in (
        ("BG", "initial", "BG Initial Symbol"), ("BG", "drop", "BG Drop Symbol"),
        ("FG", "initial", "FG Initial Symbol"), ("FG", "drop", "FG Drop Symbol"),
    ):
        merged = {symbol: [0.0] * 5 for symbol in SYMBOLS}
        gold = [0.0] * 5
        for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            symbol = GOLD_TO_BASE.get(str(row[0]), str(row[0]))
            if str(row[0]) in GOLD_TO_BASE:
                for reel in range(5):
                    gold[reel] += float(row[1 + reel] or 0)
            if symbol in merged:
                for reel in range(5):
                    merged[symbol][reel] += float(row[1 + reel] or 0)
        ratios[f"{scene}_{stage}"] = merged
        gold_ratios[f"{scene}_{stage}"] = gold
    fg_spins = sum(row["FG"] for row in combo.values())
    symbol_length = {scene: {"hits": Counter(), "pay": Counter()} for scene in ("BG", "FG")}
    if "Symbol Length" not in workbook.sheetnames:
        raise ValueError(f"{path.name}: missing Symbol Length; rerun Simulator.py with the current statistics schema")
    for scene, symbol, length, hits, pay in workbook["Symbol Length"].iter_rows(min_row=2, values_only=True):
        if scene in symbol_length and symbol in SYMBOLS and int(length) in (3, 4, 5):
            key = (str(symbol), int(length))
            symbol_length[str(scene)]["hits"][key] = int(hits or 0)
            symbol_length[str(scene)]["pay"][key] = float(pay or 0.0)
    workbook.close()
    return {
        "summary": summary, "combo": combo, "ratios": ratios, "gold_ratios": gold_ratios,
        "fg_spins": fg_spins, "symbol_length": symbol_length,
    }


def competitor_ratios(competitor: dict[str, Any], scene: str, stage: str) -> dict[str, list[float]]:
    data = competitor["counts"][scene]
    return {
        symbol: [data[stage][reel][symbol] / max(1, data[f"{stage}_total"][reel]) for reel in range(5)]
        for symbol in SYMBOLS
    }


def model_drop_ratios(sheet_name: str) -> dict[str, list[float]]:
    workbook = load_workbook(MODEL_XLSX, read_only=True, data_only=True)
    worksheet = workbook[sheet_name]
    values = {
        str(worksheet.cell(row, 32).value): [float(worksheet.cell(row, 33 + reel).value or 0) for reel in range(5)]
        for row in range(4, 23)
        if worksheet.cell(row, 32).value not in (None, "")
    }
    result = {symbol: [0.0] * 5 for symbol in SYMBOLS}
    for symbol in SYMBOLS:
        gold = {"M1": "G1", "M2": "G2", "M3": "G3", "M4": "G4", "A": "GA", "K": "GK", "Q": "GQ", "J": "GJ"}.get(symbol)
        for reel in range(5):
            total = sum(row[reel] for row in values.values())
            result[symbol][reel] = (values[symbol][reel] + (values[gold][reel] if gold else 0)) / total
    workbook.close()
    return result


def model_random_wild_weights(sheet_name: str) -> list[int]:
    workbook = load_workbook(MODEL_XLSX, read_only=False, data_only=True)
    worksheet = workbook[sheet_name]
    values = [int(worksheet.cell(row, 30).value or 0) for row in range(4, 8)]
    workbook.close()
    return values


def physical_reels(path: Path, sheet_name: str, start_row: int, start_column: int) -> list[list[str]]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    worksheet = workbook[sheet_name]
    reels = []
    for reel in range(5):
        values = [
            worksheet.cell(row, start_column + reel).value
            for row in range(start_row, worksheet.max_row + 1)
        ]
        reels.append([
            GOLD_TO_BASE.get(str(value), str(value))
            for value in values
            if value not in (None, "")
        ])
    workbook.close()
    return reels


def cyclic_stack_counts(reels: list[list[str]]) -> Counter[tuple[str, int]]:
    """Count maximal circular runs; every run of 5 or more is the 5+ bucket."""
    result: Counter[tuple[str, int]] = Counter()
    for reel in reels:
        if not reel:
            continue
        if len(set(reel)) == 1:
            result[(reel[0], 5 if len(reel) >= 5 else len(reel))] += 1
            continue
        start = next(index for index, symbol in enumerate(reel) if symbol != reel[index - 1])
        current = reel[start]
        length = 0
        for offset in range(len(reel)):
            symbol = reel[(start + offset) % len(reel)]
            if symbol == current:
                length += 1
                continue
            if length >= 2:
                result[(current, min(length, 5))] += 1
            current = symbol
            length = 1
        if length >= 2:
            result[(current, min(length, 5))] += 1
    return result


def stack_section(scene: str, competitor_reels: list[list[str]], h016_reels: list[list[str]]) -> str:
    competitor_counts = cyclic_stack_counts(competitor_reels)
    h016_counts = cyclic_stack_counts(h016_reels)
    competitor_stops = sum(map(len, competitor_reels))
    h016_stops = sum(map(len, h016_reels))

    def cell(count: int, stops: int) -> str:
        return f"{count} ({count / max(1, stops) * 1000:.2f}‰)"

    lines = [
        f"### {scene}", "",
        "| Symbol | 模型 | 2 堆疊 | 3 堆疊 | 4 堆疊 | 5 堆疊 |",
        "|---|---|---:|---:|---:|---:|",
    ]
    for symbol in SYMBOLS:
        lines.append("| " + " | ".join([
            symbol, "Super Ace",
            *[cell(competitor_counts[(symbol, length)], competitor_stops) for length in (2, 3, 4, 5)],
        ]) + " |")
        lines.append("| " + " | ".join([
            symbol, "H016",
            *[cell(h016_counts[(symbol, length)], h016_stops) for length in (2, 3, 4, 5)],
        ]) + " |")
    return "\n".join(lines)


def model_gold_ratios(sheet_name: str) -> dict[str, list[float]]:
    workbook = load_workbook(MODEL_XLSX, read_only=False, data_only=True)
    worksheet = workbook[sheet_name]
    gold_symbols = set(GOLD_TO_BASE)
    initial = []
    for reel in range(5):
        values = [worksheet.cell(row, 11 + reel).value for row in range(4, 404)]
        values = [str(value) for value in values if value not in (None, "")]
        initial.append(sum(value in gold_symbols for value in values) / max(1, len(values)))
    drop_rows = {
        str(worksheet.cell(row, 32).value): [float(worksheet.cell(row, 33 + reel).value or 0) for reel in range(5)]
        for row in range(4, 23)
        if worksheet.cell(row, 32).value not in (None, "")
    }
    drop = [
        sum(values[reel] for symbol, values in drop_rows.items() if symbol in gold_symbols)
        / max(1.0, sum(values[reel] for values in drop_rows.values()))
        for reel in range(5)
    ]
    workbook.close()
    return {"initial": initial, "drop": drop}


def model_gold_symbol_ratios(sheet_name: str) -> dict[str, dict[str, list[float]]]:
    workbook = load_workbook(MODEL_XLSX, read_only=False, data_only=True)
    worksheet = workbook[sheet_name]
    initial = {symbol: [0.0] * 5 for symbol in SYMBOLS if symbol != "C1"}
    for reel in range(5):
        values = [str(worksheet.cell(row, 11 + reel).value) for row in range(4, 404) if worksheet.cell(row, 11 + reel).value not in (None, "")]
        total = max(1, len(values))
        for gold, base in GOLD_TO_BASE.items():
            initial[base][reel] = values.count(gold) / total
    drop_rows = {
        str(worksheet.cell(row, 32).value): [float(worksheet.cell(row, 33 + reel).value or 0) for reel in range(5)]
        for row in range(4, 23)
        if worksheet.cell(row, 32).value not in (None, "")
    }
    drop = {symbol: [0.0] * 5 for symbol in initial}
    for reel in range(5):
        total = max(1.0, sum(values[reel] for values in drop_rows.values()))
        for gold, base in GOLD_TO_BASE.items():
            drop[base][reel] = drop_rows[gold][reel] / total
    workbook.close()
    return {"initial": initial, "drop": drop}


def distribution_section(title: str, competitor_rows: dict[str, list[float]], h016_rows: dict[str, list[float]]) -> str:
    lines = [f"### {title}", "", "| Symbol | 模型 | R1 | R2 | R3 | R4 | R5 |", "|---|---|---:|---:|---:|---:|---:|"]
    for symbol in SYMBOLS:
        lines.append("| " + " | ".join([symbol, "Super Ace", *[pct(value) for value in competitor_rows[symbol]]]) + " |")
        lines.append("| " + " | ".join([symbol, "H016", *[pct(value) for value in h016_rows[symbol]]]) + " |")
    return "\n".join(lines)


def symbol_length_section(
    competitor: dict[str, Any], h016: dict[str, Any], scene: str
) -> str:
    competitor_scene = competitor["counts"][scene]
    competitor_spins = competitor_scene["spins"]
    h016_spins = int(h016["summary"]["total_rounds"]) if scene == "BG" else h016["fg_spins"]
    coin_in = float(h016["summary"]["coin_in"]) * int(h016["summary"]["total_rounds"])
    lines = [
        f"### {scene}", "",
        "| Symbol | 輪數 | Super Ace RTP | H016 RTP | RTP 差異 | Super Ace Hit Rate | H016 Hit Rate | Hit Rate 差異 |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for symbol in (item for item in SYMBOLS if item != "C1"):
        for length in (3, 4, 5):
            key = (symbol, length)
            competitor_rtp = competitor_scene["symbol_length_win"][key] / competitor["coin_in"]
            h016_rtp = h016["symbol_length"][scene]["pay"][key] / coin_in
            competitor_hit = competitor_scene["symbol_length_hits"][key] / competitor_spins
            h016_hit = h016["symbol_length"][scene]["hits"][key] / max(1, h016_spins)
            lines.append(
                f"| {symbol} | {length} | {pct(competitor_rtp)} | {pct(h016_rtp)} | {pp(h016_rtp - competitor_rtp)} | "
                f"{pct(competitor_hit)} | {pct(h016_hit)} | {pp(h016_hit - competitor_hit)} |"
            )
    return "\n".join(lines)


def gold_table(
    competitor: dict[str, Any], model: dict[str, float], observed: dict[str, Any], scene: str, stage: str
) -> str:
    data = competitor["counts"][scene]
    lines = [
        "| 階段 | Reel | Super Ace | H016 設定 | H016 10 萬場 | 設定差異 | 實測差異 |",
        "|---|---|---:|---:|---:|---:|---:|",
    ]
    for reel in range(5):
        comp = data[f"gold_{stage}"][reel] / max(1, data[f"{stage}_total"][reel])
        setting = model[stage][reel]
        actual = observed["gold_ratios"][f"{scene}_{stage}"][reel]
        lines.append(
            f"| {'初始' if stage == 'initial' else '掉落'} | R{reel + 1} | {pct(comp)} | {pct(setting)} | "
            f"{pct(actual)} | {pp(setting - comp)} | {pp(actual - comp)} |"
        )
    return "\n".join(lines)


def gold_symbol_table(
    competitor: dict[str, Any], model: dict[str, dict[str, list[float]]], scene: str, stage: str
) -> str:
    data = competitor["counts"][scene]
    title = "初始" if stage == "initial" else "掉落"
    lines = [
        f"#### {scene} {title}", "",
        "| Symbol | 模型 | R1 | R2 | R3 | R4 | R5 |",
        "|---|---|---:|---:|---:|---:|---:|",
    ]
    for symbol in (item for item in SYMBOLS if item != "C1"):
        comp = [
            data[f"gold_symbol_{stage}"][reel][symbol] / max(1, data[f"{stage}_total"][reel])
            for reel in range(5)
        ]
        lines.append("| " + " | ".join([symbol, "Super Ace", *[pct(value) for value in comp]]) + " |")
        lines.append("| " + " | ".join([symbol, "H016 設定", *[pct(value) for value in model[stage][symbol]]]) + " |")
    return "\n".join(lines)


def main() -> None:
    record = max(RECORD_DIR.glob("H0161_*_betmode0_100000_*.xlsx"), key=lambda path: path.stat().st_mtime)
    h016 = record_data(record)
    hs = h016["summary"]
    competitor = raw_competitor()
    bg_model_gold = model_gold_ratios("BG_Symbol")
    fg_model_gold = model_gold_ratios("FG_Symbol")
    bg_model_gold_symbols = model_gold_symbol_ratios("BG_Symbol")
    fg_model_gold_symbols = model_gold_symbol_ratios("FG_Symbol")
    bg_random_wild = model_random_wild_weights("BG_Symbol")
    fg1_random_wild = model_random_wild_weights("FG_Symbol")
    fg2_random_wild = model_random_wild_weights("FG_Symbol (2)")
    competitor_w2 = [1401, 235, 18]
    competitor_w2_total = sum(competitor_w2)
    bg_w2_total = sum(int(hs[f"w2_bg_count_{count}"]) for count in (2, 3, 4))
    fg_w2_total = sum(int(hs[f"w2_fg_count_{count}"]) for count in (2, 3, 4))
    w2_distribution_lines = [
        "| 大鬼顆數 | Super Ace 條件分布 | BG 設定分布 | BG 10 萬場 | FG1 設定分布 | FG2 設定分布 | FG 10 萬場 |",
        "|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for index, count in enumerate((2, 3, 4), start=1):
        comp_ratio = competitor_w2[index - 1] / competitor_w2_total
        bg_setting = bg_random_wild[index] / sum(bg_random_wild[1:])
        fg1_setting = fg1_random_wild[index] / sum(fg1_random_wild[1:])
        fg2_setting = fg2_random_wild[index] / sum(fg2_random_wild[1:])
        bg_actual = int(hs[f"w2_bg_count_{count}"]) / max(1, bg_w2_total)
        fg_actual = int(hs[f"w2_fg_count_{count}"]) / max(1, fg_w2_total)
        w2_distribution_lines.append(
            f"| {count} | {pct(comp_ratio)} | {pct(bg_setting)} | {pct(bg_actual)} | "
            f"{pct(fg1_setting)} | {pct(fg2_setting)} | {pct(fg_actual)} |"
        )
    competitor_bg_reels = physical_reels(COMPETITOR_STRIP_XLSX, "BG_Strip", 3, 2)
    competitor_fg_reels = physical_reels(COMPETITOR_STRIP_XLSX, "FG_Strip", 3, 2)
    h016_bg_reels = physical_reels(MODEL_XLSX, "BG_Symbol", 4, 11)
    h016_fg_reels = physical_reels(MODEL_XLSX, "FG_Symbol", 4, 11)
    core = [
        ("RTP", "Total RTP", competitor["rtp_total"], float(hs["rtp_total"])),
        ("RTP", "BG RTP", competitor["rtp_bg"], float(hs["rtp_bg"])),
        ("RTP", "FG RTP", competitor["rtp_fg"], float(hs["rtp_fg"])),
        ("Hit Rate", "BG Hit Rate", competitor["bg_hit_rate"], float(hs["bg_hit_rate"])),
        ("Hit Rate", "FG Hit Rate", competitor["fg_hit_rate"], float(hs["fg_hit_rate"])),
        ("FG", "FG Trigger Rate", competitor["fg_trigger_rate"], float(hs["fg_trigger_rate"])),
    ]
    core_lines = ["| 類別 | 指標 | Super Ace | H016 | 差異 |", "|---|---|---:|---:|---:|"]
    for category, metric, comp, model in core:
        core_lines.append(f"| {category} | {metric} | {pct(comp)} | {pct(model)} | {pp(model - comp)} |")
    core_lines.extend([
        f"| FG | 平均觸發週期 | {1 / competitor['fg_trigger_rate']:.2f} 局／次 | {float(hs['fg_trigger_cycle']):.2f} 局／次 | H016 {'慢' if float(hs['fg_trigger_cycle']) > 1 / competitor['fg_trigger_rate'] else '快'} {abs(float(hs['fg_trigger_cycle']) - 1 / competitor['fg_trigger_rate']):.2f} 局 |",
        f"| FG | 平均 Free Spins | {competitor['avg_fg_spins']:.4f} | {float(hs['avg_fg_spins']):.4f} | {float(hs['avg_fg_spins']) - competitor['avg_fg_spins']:+.4f} |",
    ])

    pay_lines = [
        "| Symbol | 3 輪 | 4 輪 | 5 輪 | 比較結果 |", "|---|---:|---:|---:|---|",
        "| M1 | 0.50 | 1.50 | 2.50 | 相同 |", "| M2 | 0.40 | 1.20 | 2.00 | 相同 |",
        "| M3 | 0.30 | 0.90 | 1.50 | 相同 |", "| M4 | 0.20 | 0.60 | 1.00 | 相同 |",
        "| A | 0.10 | 0.30 | 0.50 | 相同 |", "| K | 0.10 | 0.30 | 0.50 | 相同 |",
        "| Q | 0.05 | 0.15 | 0.25 | 相同 |", "| J | 0.05 | 0.15 | 0.25 | 相同 |",
    ]

    combo_sections = []
    for scene in ("BG", "FG"):
        h_den = int(hs["total_rounds"]) if scene == "BG" else sum(row[scene] for row in h016["combo"].values())
        c_den = competitor["rounds"] if scene == "BG" else competitor["fg_spins"]
        lines = [f"### {scene}", "", f"| 消除次數 | Super Ace | H016 | 差異 |", "|---|---:|---:|---:|"]
        for index, label in enumerate(("0", "1", "2", "3", "4", "5+")):
            comp = competitor["counts"][scene]["combo"][index] / c_den
            model = h016["combo"][label][scene] / h_den
            lines.append(f"| {label} | {pct(comp)} | {pct(model)} | {pp(model - comp)} |")
        combo_sections.append("\n".join(lines))

    document = f"""# 競品參考數值比較

## 目錄

- [Overview](#overview)
  - [核心指標比較](#核心指標比較)
- [賠率](#賠率)
- [符號 345 連線 RTP / Hit Rate 占比](#符號-345-連線-rtp--hit-rate-占比)
  - [BG](#bg)
  - [FG](#fg)
- [不同符號 2／3／4／5 堆疊競品比較](#不同符號-2345-堆疊競品比較)
- [符號分布](#符號分布)
  - [BG 初始 R1-R5](#bg-初始-r1-r5)
  - [BG 掉落 R1-R5](#bg-掉落-r1-r5)
  - [FG 初始 R1-R5](#fg-初始-r1-r5)
  - [FG 掉落 R1-R5](#fg-掉落-r1-r5)
- [消除率](#消除率)
- [金框比例](#金框比例)
- [大鬼事件率](#大鬼事件率)

## Overview

### 比較基準

| 項目 | Super Ace | H016 現在版本 |
|---|---|---|
| 來源 | JILI 實機非重複 JSONL | `{record.name}` + `Simulator.py` |
| 樣本 | {competitor['rounds']:,} 個 BG Round、{competitor['fg_spins']:,} 個 FG Spin | {int(hs['total_rounds']):,} 個 BG Round、{h016['fg_spins']:,} 個 FG Spin |
| Card System | 競品實際遊玩資料 | 關閉 |
| Bet Mode | 一般投注 | Normal Bet，Bet Multi 1 |
| 輪帶 | `StripTable_SuperAce_還原.xlsx` | 六張表的 `K:O` 均各自存有 5 輪 × 200 格輪帶；各表使用獨立非負整數停輪權重 |

同一 Scene 的三張表使用相同輪帶排列，但資料完整複製到各工作表 `K:O`，Simulator 直接讀取該次所選工作表。停輪權重讀取各表 `W:AA`，補牌讀取各表 `AF:AK`；`Parameter` 直接控制實際選表。

### 多表用途與權重

| Scene | 表 | Table Weight | 用途 |
|---|---|---:|---|
| BG | `BG_Symbol` | 891,563 | 競品表；R3、R5 初始與掉落不出 SC，且程式禁止觸發 FG |
| BG | `BG_Symbol (2)` | 100,000 | 大鬼能見度表；R1、R4 初始與掉落不出 SC，得分集中 M4／A／Q／J，且程式禁止觸發 FG |
| BG | `BG_Symbol (3)` | 8,437 | FG 觸發表；固定三輪 SC，實測 100% 觸發且進場得分 0x |
| FG | `FG_Symbol` | 1 | 競品表；保留 101003 的 2／3／4 顆條件分布，非 0 抽中率為 FG(2) 的一半 |
| FG | `FG_Symbol (2)` | 1 | 大鬼能見度表；使用 101003 Random Wild 參數，得分集中 M4／A／Q／J |
| FG | `FG_Symbol (3)` | 0 | 未啟用 |

分表各自獨立跑 10 萬 Spin 的 Hit Rate 倍率如下。倍率分母為同 Scene 的競品表；目標群為約 2 倍、其他一般付費符號為約 1/2。因實體輪帶固定，僅用整數停輪與補牌權重無法讓每個符號同時精確等於目標，以下為目前校準結果。

| 符號 | BG(2) ÷ BG 競品表 | FG(2) ÷ FG 競品表 | 目標 |
|---|---:|---:|---:|
| M1 | 0.58x | 0.52x | 0.5x |
| M2 | 0.35x | 0.58x | 0.5x |
| M3 | 0.33x | 0.47x | 0.5x |
| M4 | 2.30x | 2.44x | 2.0x |
| A | 1.51x | 1.47x | 2.0x |
| K | 0.59x | 0.45x | 0.5x |
| Q | 1.79x | 1.97x | 2.0x |
| J | 1.82x | 1.81x | 2.0x |

最新 10 萬場 FG 觸發率為 {pct(float(hs['fg_trigger_rate']))}、週期 {float(hs['fg_trigger_cycle']):.2f} 局；競品為 {pct(competitor['fg_trigger_rate'])}、{1 / competitor['fg_trigger_rate']:.2f} 局。高能見度表會提高整體 RTP，這是本版按指定表用途產生的直接取捨。

### 核心指標比較

{chr(10).join(core_lines)}

H016 的 BG、FG 與總 RTP 高於競品，主因為本版啟用 BG(2)／FG(2) 高能見度表，且其指定符號與大鬼頻率都高於競品表；金框折算進實體輪帶後也會參與連續 4 格視窗與 Cascade 補牌，因此不能把兩者視為完全等價的生成機制。

## 賠率

H016 與 Super Ace 參考賠率相同；表內數字為相對投注額倍數。

{chr(10).join(pay_lines)}

## 符號 345 連線 RTP / Hit Rate 占比

RTP 分母為該次模擬或競品樣本的總投注額；Hit Rate 分母為該 Scene 的 Spin 數。每次 Cascade 中，同一符號同一輪數的一筆 award 計為一次 Hit；同一 Spin 可同時命中多符號或多次 Cascade，所以各列 Hit Rate 不應加總為整體 Spin Hit Rate。競品 RTP 已將 `award.win` 乘上該次 `comboBonus`；H016 RTP 使用模擬器實際派彩，含 BG／FG Cascade 倍率。表內的「所有符號」指具有 3／4／5 輪賠率的 `M1～M4、A、K、Q、J`；`C1`、`WW`、`W2` 不屬於獨立付費符號，因此不列入。

{symbol_length_section(competitor, h016, 'BG')}

{symbol_length_section(competitor, h016, 'FG')}

## 不同符號 2／3／4／5 堆疊競品比較

堆疊以單條循環輪帶上「同一符號的最大連續段」計算；金框符號併回原符號，例如 `A／GA／A` 算同一段。表格顯示 `段數（每千格段數）`，用每千格標準化競品原始帶長與 H016 200 格帶長的差異。表內「5 堆疊」代表 5 格以上；H016 六張表各自存有輪帶資料，同 Scene 三張表的排列相同，因此以 BG／FG 競品表代表 H016。

{stack_section('BG', competitor_bg_reels, h016_bg_reels)}

{stack_section('FG', competitor_fg_reels, h016_fg_reels)}

## 符號分布

### 套用方法

1. 依 `BG_Strip`／`FG_Strip` 原排列與 stopW 累積機率，等距重採樣成每輪 200 格後填入模型 `K:O`。
2. 六張工作表的 `K:O` 都各自保存 200 格輪帶；每一格停輪權重直接讀取本次所選工作表的 `W:AA`。
3. 消除後補牌讀取本次所選工作表的 `AF:AK Symbol Drop Weight`。
4. 初始盤面依輪帶連續取 4 格；Cascade 在被消除的原位置依 Symbol Drop Weight 獨立補牌，不做重力掉落。

{distribution_section('BG 初始 R1-R5', competitor_ratios(competitor, 'BG', 'initial'), h016['ratios']['BG_initial'])}

{distribution_section('BG 掉落 R1-R5', competitor_ratios(competitor, 'BG', 'drop'), h016['ratios']['BG_drop'])}

{distribution_section('FG 初始 R1-R5', competitor_ratios(competitor, 'FG', 'initial'), h016['ratios']['FG_initial'])}

{distribution_section('FG 掉落 R1-R5', competitor_ratios(competitor, 'FG', 'drop'), h016['ratios']['FG_drop'])}

以上 H016 初始與掉落符號分布皆為最新 10 萬場模擬實測。FG `Symbol Drop Weight` 維持競品分布；BG 為校準 FG 觸發週期，僅降低掉落 C1 並按原比例回分至一般符號，金框權重不變。掉落分母為各輪實際補牌顆數。

## 消除率

比較口徑為每個 Spin 實際發生的消除次數，統一合併為 `0、1、2、3、4、5+`；FG 分母為實際 Free Spins。

{combo_sections[0]}

{combo_sections[1]}

## 金框比例

競品金框是盤面生成後的獨立疊加層；H016 依需求將其折算進既有輪帶。200 格輪帶的最小刻度為 0.5%，因此採最接近競品比例的整數格數；R1、R5 不放金框，C1 不轉金框。

### BG

{gold_table(competitor, bg_model_gold, h016, 'BG', 'initial')}

{gold_table(competitor, bg_model_gold, h016, 'BG', 'drop')}

### FG

{gold_table(competitor, fg_model_gold, h016, 'FG', 'initial')}

{gold_table(competitor, fg_model_gold, h016, 'FG', 'drop')}

### 各符號金框分布

比例口徑為「該符號的金框數 ÷ 該輪全部 Symbol 數」。H016 初始輪帶受 200 格、最小 0.5 個百分點限制；掉落使用總權重 1,000,000，因此可近乎精確對齊競品逐符號分布。

{gold_symbol_table(competitor, bg_model_gold_symbols, 'BG', 'initial')}

{gold_symbol_table(competitor, bg_model_gold_symbols, 'BG', 'drop')}

{gold_symbol_table(competitor, fg_model_gold_symbols, 'FG', 'initial')}

{gold_symbol_table(competitor, fg_model_gold_symbols, 'FG', 'drop')}

## 大鬼事件率

事件率口徑為「成功啟動大鬼的事件數 ÷ 該 Scene Spins」。BG 同一個 Spin 最多成功啟動一次 WW2；FG 每次 Cascade 有新金框轉換都會重新判定，因此同一個 FG Spin 可成功啟動多次。抽到 0 時該次金框只轉 WW1，後續金框仍會再判定。

| Scene | Super Ace | H016 多表混合實測 | 差異 |
|---|---:|---:|---:|
| BG | {pct(competitor['w2_bg_event_rate'])} | {pct(float(hs['w2_bg_event_rate']))} | {pp(float(hs['w2_bg_event_rate']) - competitor['w2_bg_event_rate'])} |
| FG | 0.0000% | {pct(float(hs['w2_fg_event_rate']))} | {pp(float(hs['w2_fg_event_rate']))} |

BG 競品表權重為 `0/2/3/4 = {'/'.join(map(str, bg_random_wild))}`。101003 的 FG 參數為 `13128/2000/500/200`，非 0 抽中率為 `2700/15828`。FG1 使用 `{'/'.join(map(str, fg1_random_wild))}`，保留相同 2／3／4 顆條件分布，並將非 0 抽中率精確減半為 `2700/31656`；FG2 完整使用 `{'/'.join(map(str, fg2_random_wild))}`；FG3 的 Table Selection Weight 為 0，暫不使用。

{chr(10).join(w2_distribution_lines)}
"""
    OUTPUT.write_text(document, encoding="utf-8")
    print(json.dumps({"output": str(OUTPUT), "record": record.name, "competitor_bg": competitor["rounds"], "competitor_fg": competitor["fg_spins"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
