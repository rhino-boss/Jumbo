"""Build H016 v6.4 with both newbie Normal profiles locked to 93% RTP.

Oldhand, BF, and the approved v6.3 SF line are preserved.  Newbie BG stays
at 65%, while newbie FG is recalculated to 28% for both 92A and 94A.
"""
from __future__ import annotations

import importlib.util
import json
from copy import deepcopy
from pathlib import Path

from openpyxl import load_workbook


TOOL_DIR = Path(__file__).resolve().parent
PROJECT = TOOL_DIR.parents[1]
SOURCE = TOOL_DIR / "calculate_v5_1_multiplier_payload.py"
CURRENT_PAYLOAD = PROJECT / "其他" / "診斷" / "H016_v6_3_sf_payload.json"
OUTPUT = PROJECT / "其他" / "診斷" / "H016_v6_4_multiplier_payload.json"
VERSION = "6.4.0.0"


def main() -> None:
    spec = importlib.util.spec_from_file_location("h016_v6_4_solver", SOURCE)
    if spec is None or spec.loader is None:
        raise ImportError(f"Cannot load {SOURCE}")
    solver = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(solver)

    current = json.loads(CURRENT_PAYLOAD.read_text(encoding="utf-8"))
    archived_bf = current["bf_source_report"]
    bf_report = {
        "path": f"{CURRENT_PAYLOAD.resolve()}#bf_source_report",
        "rounds": int(archived_bf["rounds"]),
        "base_bet": float(archived_bf["base_bet"]),
        "bf_count": [int(value) for value in archived_bf["bf_count"]],
        "bf_pay": [float(value) for value in archived_bf["bf_pay"]],
    }
    if sum(bf_report["bf_count"]) != bf_report["rounds"]:
        raise ValueError("Archived BF counts do not match rounds")

    solver.VERSION = VERSION
    solver.NORMAL_REPORT = PROJECT / "Record" / "H0161_06_2608191134_betmode0_109.xlsx"
    solver.BF_REPORT = CURRENT_PAYLOAD
    solver.load_bf_report = lambda _path: bf_report
    solver.TARGET_NEWBIE_BG_RTP = 0.65
    solver.TARGET_NEWBIE_FG_RTP_BY_VERSION = {"92": 0.28, "94": 0.28}
    solver.OUTPUT_PATH = OUTPUT
    solver.main()

    payload = json.loads(OUTPUT.read_text(encoding="utf-8"))
    for key in ("92", "94"):
        recalculated = payload["versions"][key]
        preserved = deepcopy(current["versions"][key])
        archived_book = load_workbook(
            PROJECT / "Versions" / "6.3" / "Source" / f"H0161{key}A.xlsx",
            read_only=True,
            data_only=True,
            keep_links=False,
        )
        try:
            detail = archived_book["Detail"]
            for section, start_row in (("bg", 15), ("fg", 86), ("bf", 163), ("sf", 234)):
                block = preserved[section]
                block["fix"] = [
                    float(detail.cell(row, 8).value or 0.0)
                    for row in range(start_row, start_row + 64)
                ]
                block["weights"] = [
                    int(detail.cell(row, 11).value or 0)
                    for row in range(start_row, start_row + 64)
                ]
                for item in block.get("audit", []):
                    row = start_row + int(item["index"])
                    weight = int(detail.cell(row, 11).value or 0)
                    rtp = float(detail.cell(row, 13).value or 0.0)
                    item["after_weight"] = weight
                    item["after_hit_rate"] = weight / 1_000_000_000
                    item["after_rtp"] = rtp
                    item["target_scene_rtp"] = rtp
        finally:
            archived_book.close()
        preserved["newbie"] = recalculated["newbie"]
        for metric in (
            "newbie_target_rtp",
            "newbie_trigger_bg_count",
            "newbie_trigger_bg_pay",
            "newbie_trigger_bg_average",
            "newbie_trigger_bg_cap",
            "newbie_rtp",
            "newbie_bg_rtp",
            "newbie_fg_rtp",
        ):
            if metric in recalculated["metrics"]:
                preserved["metrics"][metric] = recalculated["metrics"][metric]
        preserved["metrics"]["version"] = VERSION
        payload["versions"][key] = preserved
    for key in ("sf_source_report", "sf_competitor_reference"):
        payload[key] = deepcopy(current[key])
    payload["rules"]["sf"] = "unchanged from v6.3"
    payload["rules"]["sf_rule"] = current["rules"]["sf_rule"]
    payload["rules"]["targets"] = deepcopy(current["rules"].get("targets", {}))
    payload["version"] = VERSION
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(json.dumps({
        "output": str(OUTPUT.resolve()),
        "version": VERSION,
        "newbie_targets": payload["rules"]["newbie_targets"],
        "newbie_metrics": {
            key: {
                "rtp": payload["versions"][key]["metrics"]["newbie_rtp"],
                "bg": payload["versions"][key]["metrics"]["newbie_bg_rtp"],
                "fg": payload["versions"][key]["metrics"]["newbie_fg_rtp"],
            }
            for key in ("92", "94")
        },
        "non_newbie_preserved": all(
            all(
                payload["versions"][key][section] == current["versions"][key][section]
                for section in ("bg", "fg", "bf", "sf")
            )
            for key in ("92", "94")
        ),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
