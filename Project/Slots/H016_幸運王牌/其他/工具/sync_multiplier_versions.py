from __future__ import annotations

import argparse
import json
import re
from copy import deepcopy
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[2]
SOURCE_DIR = PROJECT_DIR / "Source"
VERSIONS_DIR = PROJECT_DIR / "Versions"
VERSION = "3.0.0.0"
MATH_KEY = "3.0"
BASE_VERSION = "3"
CONFIGS = {
    "92": (PROJECT_DIR / "config_92A.js", SOURCE_DIR / "H016192A.xlsx"),
    "94": (PROJECT_DIR / "config_94A.js", SOURCE_DIR / "H016194A.xlsx"),
}
CONFIG_PATTERN = re.compile(r"window\.H016_CONFIG\s*=\s*(\{.*\})\s*;", re.DOTALL)


def load_config(path: Path) -> tuple[str, dict[str, Any]]:
    source = path.read_text(encoding="utf-8")
    match = CONFIG_PATTERN.search(source)
    if match is None:
        raise ValueError(f"Cannot find window.H016_CONFIG in {path}")
    header = source[: match.start()].rstrip()
    return header, json.loads(match.group(1))


def atomic_write(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(text, encoding="utf-8")
    temporary.replace(path)


def config_text(header: str, config: dict[str, Any]) -> str:
    payload = json.dumps(config, ensure_ascii=False, separators=(",", ":"))
    prefix = header + "\n" if header else ""
    return f"{prefix}window.H016_CONFIG={payload};\n"


def range_cards(
    sheet, start: int, table: str, override_weights: list[int] | None = None
) -> list[dict[str, Any]]:
    cards: list[dict[str, Any]] = []
    for row in range(start, start + 64):
        minimum = float(sheet.cell(row, 15).value)
        maximum = float(sheet.cell(row, 16).value)
        index = row - start
        weight = int(
            override_weights[index]
            if override_weights is not None
            else sheet.cell(row, 11).value or 0
        )
        if weight <= 0:
            continue
        cards.append({
            "type": "range",
            "min": minimum,
            "max": maximum,
            "table": table,
            "weight": weight,
        })
    return cards


def card_system(
    workbook_path: Path, override: dict[str, Any] | None = None
) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        detail = workbook["Detail"]
        newbie = workbook["Detail_Newbie"]

        def base_cards(sheet, bg_weights, entry_weight_override) -> list[dict[str, Any]]:
            cards = range_cards(sheet, 15, "B", bg_weights)
            free_weight = int(
                entry_weight_override
                if entry_weight_override is not None
                else (
                    bg_weights[64]
                    if bg_weights is not None and len(bg_weights) > 64
                    else sheet.cell(79, 11).value or 0
                )
            )
            if free_weight > 0:
                cards.append({"type": "free_game", "table": "A", "weight": free_weight})
            return cards

        old_bg_weights = override["bg"]["weights"] if override is not None else None
        old_fg_weights = override["fg"]["weights"] if override is not None else None
        newbie_bg_weights = override["newbie"]["bg"]["weights"] if override is not None else None
        newbie_fg_weights = override["newbie"]["fg"]["weights"] if override is not None else None
        bf_weights = override["bf"]["weights"] if override is not None else None
        sf_weights = override["sf"]["weights"] if override is not None else None
        entry_weight_override = None
        if override is not None:
            trigger_rate = float(override["metrics"]["trigger_rate"])
            entry_weight_override = round(
                trigger_rate / (1.0 - trigger_rate) * 1_000_000_000
            )
        buy = range_cards(detail, 163, "E", bf_weights)
        super_cards = range_cards(detail, 234, "G", sf_weights)
        profiles = {
            "weight_1": {
                "base_game": base_cards(newbie, newbie_bg_weights, entry_weight_override),
                "free_game": range_cards(newbie, 86, "E", newbie_fg_weights),
                "buy_feature": deepcopy(buy),
                "super_feature": deepcopy(super_cards),
            },
            "weight_2": {
                "base_game": base_cards(detail, old_bg_weights, entry_weight_override),
                "free_game": range_cards(detail, 86, "E", old_fg_weights),
                "buy_feature": buy,
                "super_feature": super_cards,
            },
        }

        for profile_name, profile in profiles.items():
            for section in ("free_game", "buy_feature", "super_feature"):
                cards = profile[section]
                total = sum(card["weight"] for card in cards)
                delta = 1_000_000_000 - total
                if abs(delta) > 1:
                    raise ValueError(
                        f"{workbook_path.name}: {profile_name}.{section} sums to {total:,}, expected 1,000,000,000"
                    )
                if delta:
                    cards[-1]["weight"] += delta
            base_range_total = sum(
                card["weight"] for card in profile["base_game"] if card["type"] == "range"
            )
            if base_range_total != 1_000_000_000:
                raise ValueError(
                    f"{workbook_path.name}: {profile_name}.base_game range cards sum to "
                    f"{base_range_total:,}, expected 1,000,000,000"
                )
        return {
            "enabled": True,
            "retry_limit": 10_000,
            "default_profile": "weight_2",
            "profiles": profiles,
        }
    finally:
        workbook.close()


def update_manifest(config_paths: dict[str, str]) -> None:
    manifest_path = VERSIONS_DIR / "version_manifest.json"
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["current"] = VERSION
    entry = {
        "version": VERSION,
        "math_key": MATH_KEY,
        "date": date.today().isoformat(),
        "configs": config_paths,
        "changes": [
            "H0161 基底版號改為單碼 3；RTP/Variant 版號同步為 3.0.0.0。",
            "新增 SF_Symbol、SF_Symbol (2)、SF_Symbol (3) 與 SF 初始／Retrigger 選表參數。",
            "Super Feature 依 JHS101003 使用獨立 SF 表，並套用 Super Buy 金框位置排除規則。",
            "Console 與 Overview 依 slot_development_specification.md §3.3 固定順序輸出。",
            "Card Retry Limit 正式統一為 10,000。",
        ],
    }
    versions = [item for item in manifest.get("versions", []) if item.get("version") != VERSION]
    for item in versions:
        if item.get("version") in {"1.1", "1.2", "1.3", "1.4", "1.5", "1.6", "1.7", "1.8", "1.9", "1.10", "1.11", "1.12"}:
            item["status"] = "superseded"
            item["superseded_by"] = VERSION
    versions.append(entry)
    manifest["versions"] = versions
    atomic_write(manifest_path, json.dumps(manifest, ensure_ascii=False, indent=2) + "\n")
    atomic_write(
        VERSIONS_DIR / "version_manifest.js",
        "window.H016_VERSION_MANIFEST = "
        + json.dumps(manifest, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
    )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--payload", type=Path, help="Validated multiplier-rule payload override")
    parser.add_argument(
        "--versions-only", action="store_true",
        help="Only synchronize Overview version cells before config generation",
    )
    args = parser.parse_args()
    workbook_versions = {
        SOURCE_DIR / "H0161.xlsx": BASE_VERSION,
        SOURCE_DIR / "H016192A.xlsx": VERSION,
        SOURCE_DIR / "H016194A.xlsx": VERSION,
    }
    for workbook_path, version in workbook_versions.items():
        workbook = load_workbook(workbook_path, read_only=True, data_only=False)
        try:
            if str(workbook["Overview"]["B3"].value or "").strip() != version:
                raise ValueError(
                    f"{workbook_path.name}: Overview!B3 must be {version}; "
                    "update it with Excel so formula caches are preserved"
                )
        finally:
            workbook.close()
    if args.versions_only:
        print(json.dumps({path.name: version for path, version in workbook_versions.items()}, ensure_ascii=False, indent=2))
        return
    payload = (
        json.loads(args.payload.read_text(encoding="utf-8"))
        if args.payload is not None
        else None
    )
    version_configs: dict[str, str] = {}
    base_config = PROJECT_DIR / "config.js"
    if not base_config.is_file():
        raise FileNotFoundError(base_config)
    _, base_runtime = load_config(base_config)
    archived_base = VERSIONS_DIR / MATH_KEY / base_config.name
    atomic_write(archived_base, base_config.read_text(encoding="utf-8"))
    version_configs["base"] = archived_base.relative_to(PROJECT_DIR).as_posix()
    for label, (config_path, workbook_path) in CONFIGS.items():
        if not workbook_path.is_file():
            raise FileNotFoundError(workbook_path)
        header, previous_config = load_config(config_path)
        # Runtime variants must use exactly the same game math as config.js.
        # Only version/provenance fields and Card System multiplier weights may differ.
        config = deepcopy(base_runtime)
        for key in ("parsheet_id", "excel_version", "rtp_label"):
            if key in previous_config:
                config[key] = previous_config[key]
        config["runtime_version"] = VERSION
        config["source_multiplier_xlsx"] = workbook_path.name
        override = payload["versions"][label] if payload is not None else None
        config["card_system"] = card_system(workbook_path, override)
        rendered = config_text(header, config)
        atomic_write(config_path, rendered)
        archived = VERSIONS_DIR / MATH_KEY / config_path.name
        atomic_write(archived, rendered)
        version_configs[label] = archived.relative_to(PROJECT_DIR).as_posix()

    update_manifest(version_configs)
    print(json.dumps({"version": VERSION, "configs": version_configs}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
