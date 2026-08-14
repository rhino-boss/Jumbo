from __future__ import annotations

import argparse
import importlib.util
import json
import os
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
PROJECT_DIR = HERE.parents[1]
SOURCE_DIR = PROJECT_DIR / "Source"
DEFAULT_SOURCE = SOURCE_DIR / "H0161.xlsx"
DEFAULT_OUTPUT = PROJECT_DIR / "config.js"
FRONTEND_NAMES = {
    "0": "WW1", "1": "WW2", "2": "C1",
    "3": "M1", "4": "M2", "5": "M3", "6": "M4",
    "7": "A", "8": "K", "9": "Q", "10": "J",
    "11": "M1G", "12": "M2G", "13": "M3G", "14": "M4G",
    "15": "AG", "16": "KG", "17": "QG", "18": "JG",
}


def load_simulator_module():
    os.environ.setdefault("H016_BASE_DIR", str(PROJECT_DIR))
    os.environ.setdefault("H016_RUN_ALL_COMBINATIONS", "false")
    # Config generation must parse the source workbook independently of any
    # previously generated RTP variant. During a major-version transition the
    # old variant is intentionally incompatible until it is regenerated.
    os.environ["H016_CONFIG_FILE"] = "config.js"
    os.environ["H016_CONFIG_RTP_FILE"] = "config.js"
    spec = importlib.util.spec_from_file_location("h016_simulator_for_config", PROJECT_DIR / "Simulator.py")
    if spec is None or spec.loader is None:
        raise RuntimeError("Cannot load Simulator.py")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def _parse_range(label: str) -> tuple[float, float] | None:
    match = re.fullmatch(r"\(\s*(-?[0-9.]+)\s*,\s*([0-9.]+)\s*\]", label)
    return (float(match.group(1)), float(match.group(2))) if match else None


def load_card_system(variant: Path) -> dict[str, Any]:
    workbook = load_workbook(variant, read_only=True, data_only=True)
    try:
        detail = workbook["Detail"]
        newbie = workbook["Detail_Newbie"]
        threshold = int(detail["B2"].value or 0)
        if threshold != 1_000_000_000:
            raise ValueError(f"{variant.name}: Detail!B2 must be 1,000,000,000")
        ranges = [
            (float(detail.cell(row, 15).value), float(detail.cell(row, 16).value))
            for row in range(15, 79)
        ]

        def normalized_weights(sheet, start_row: int) -> list[int]:
            raw = [
                float(sheet.cell(row, 2).value or 0)
                * float(sheet.cell(row, 8).value or 0)
                for row in range(start_row, start_row + 64)
            ]
            total = sum(raw)
            if total <= 0:
                raise ValueError(
                    f"{variant.name}: {sheet.title} row {start_row} has no weighted mass"
                )
            exact = [value * threshold / total for value in raw]
            weights = [int(value) for value in exact]
            remainder = threshold - sum(weights)
            order = sorted(
                range(64), key=lambda index: exact[index] - weights[index], reverse=True
            )
            for index in order[:remainder]:
                weights[index] += 1
            return weights

        def range_cards(weights: list[int], table: str) -> list[dict[str, Any]]:
            return [
                {
                    "type": "range", "min": minimum, "max": maximum,
                    "table": table, "weight": int(weight),
                }
                for (minimum, maximum), weight in zip(ranges, weights)
            ]

        def base_cards(sheet) -> list[dict[str, Any]]:
            cards = range_cards(normalized_weights(sheet, 15), "B")
            rounds = float(sheet["B13"].value or 0)
            count = float(sheet["B79"].value or 0)
            fix = float(sheet["H79"].value or 0)
            if rounds <= 0:
                raise ValueError(f"{variant.name}: {sheet.title}!B13 must be positive")
            entry_weight = int(count / rounds * fix * threshold)
            cards.append({
                "type": "free_game", "table": "A", "weight": entry_weight,
            })
            return cards

        buy = range_cards(normalized_weights(detail, 163), "E")
        super_cards = range_cards(normalized_weights(detail, 234), "G")
        profiles = {
            "weight_1": {
                "base_game": base_cards(newbie),
                "free_game": range_cards(normalized_weights(newbie, 86), "E"),
                "buy_feature": [dict(card) for card in buy],
                "super_feature": [dict(card) for card in super_cards],
            },
            "weight_2": {
                "base_game": base_cards(detail),
                "free_game": range_cards(normalized_weights(detail, 86), "E"),
                "buy_feature": buy,
                "super_feature": super_cards,
            },
        }
        for profile in profiles.values():
            base_range_total = sum(
                card["weight"] for card in profile["base_game"]
                if card["type"] == "range"
            )
            if base_range_total != 1_000_000_000:
                raise ValueError(
                    f"{variant.name}: base_game range weights sum to "
                    f"{base_range_total:,}, expected 1,000,000,000"
                )
            for section in ("free_game", "buy_feature", "super_feature"):
                cards = profile[section]
                total = sum(card["weight"] for card in cards)
                if total != 1_000_000_000:
                    raise ValueError(
                        f"{variant.name}: {section} weights sum to {total:,}, "
                        "expected 1,000,000,000"
                    )
        return {"enabled": True, "retry_limit": 10_000, "default_profile": "weight_2", "profiles": profiles}
    finally:
        workbook.close()


def workbook_identity(path: Path) -> tuple[str, str]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheet = workbook["Overview"]
        return str(sheet["B2"].value or "").strip(), str(sheet["B3"].value or "").strip()
    finally:
        workbook.close()


def version_major(value: Any) -> str:
    return str(value or "").strip().split(".", 1)[0]


def variant_rtp_label(variant: Path) -> int:
    match = re.fullmatch(r"H0161(\d{2})[A-Za-z]?", variant.stem)
    if match is None:
        raise ValueError(f"Cannot determine RTP from variant workbook {variant.name!r}")
    return int(match.group(1))


def frontend_config(source: Path, variant: Path | None = None) -> dict[str, Any]:
    simulator = load_simulator_module()
    config = simulator._load_xlsx_config(source)
    source_model, source_version = workbook_identity(source)
    if source_model != "H0161":
        raise ValueError(f"{source.name}: Overview!B2 must be H0161, got {source_model!r}")
    base_version = version_major(source_version)
    if not re.fullmatch(r"\d+", base_version):
        raise ValueError(f"{source.name}: invalid base version {source_version!r}")
    config["excel_version"] = base_version
    config.pop("version", None)
    config["name_en"] = "Lucky Ace"
    config["parsheet_id"] = "H0161"
    config.pop("rtp_label", None)
    config["reel_num"] = 5
    config["window_size"] = 4
    config["max_ways"] = 1024
    config["symbol_names"] = FRONTEND_NAMES
    config["base_table_weights"] = {"high": 1, "low": 0}
    config["source_xlsx"] = source.name
    config.pop("source_multiplier_xlsx", None)
    config.pop("runtime_version", None)
    config["card_system"] = {"enabled": False, "profiles": {}}
    if variant is not None:
        variant_model, variant_version = workbook_identity(variant)
        expected_model = variant.stem[:-1] if variant.stem[-1:].isalpha() else variant.stem
        if variant_model != expected_model:
            raise ValueError(
                f"{variant.name}: Overview!B2 must be {expected_model}, got {variant_model!r}"
            )
        if not re.fullmatch(r"\d+\.\d+\.\d+\.\d+", variant_version):
            raise ValueError(f"{variant.name}: invalid RTP/Variant version {variant_version!r}")
        if version_major(variant_version) != base_version:
            raise ValueError(
                f"Version mismatch: {source.name}={source_version!r}, "
                f"{variant.name}={variant_version!r}"
            )
        rtp_label = variant_rtp_label(variant)
        config["excel_version"] = variant_version
        config["parsheet_id"] = f"H0161{rtp_label}"
        config["rtp_label"] = rtp_label
        config["source_multiplier_xlsx"] = variant.name
        config["card_system"] = load_card_system(variant)
    for table in config["tables"].values():
        normalized = []
        for reel in table["weights"]:
            if any(float(weight) != int(weight) for weight in reel):
                raise ValueError("H0161.xlsx contains a non-integer reel weight")
            normalized.append([int(weight) for weight in reel])
        table["weights"] = normalized
        normalized_drop = []
        for reel in table["drop_weights"]:
            if any(float(weight) != int(weight) for weight in reel):
                raise ValueError("H0161.xlsx contains a non-integer Symbol Drop Weight")
            normalized_drop.append([int(weight) for weight in reel])
        table["drop_weights"] = normalized_drop
        random_weights = table["random_wild"]["weights"]
        if any(float(weight) != int(weight) for weight in random_weights):
            raise ValueError("H0161.xlsx contains a non-integer Random Wild weight")
        table["random_wild"]["weights"] = [int(weight) for weight in random_weights]
    return config


def validate(config: dict[str, Any]) -> dict[str, Any]:
    tables = config["tables"]
    bg = tables["bg_high"]
    fg = tables["fg_high_a"]
    gold_ids = set(range(11, 19))

    def gold_counts(table: dict[str, Any]) -> list[int]:
        return [sum(symbol in gold_ids for symbol in reel) for reel in table["reels"]]

    for table in tables.values():
        if [len(reel) for reel in table["reels"]] != [200] * 5:
            raise ValueError("Every frontend reel must contain exactly 200 stops")
        if any(any(type(weight) is not int or weight < 0 for weight in reel) for reel in table["weights"]):
            raise ValueError("Every frontend initial stop weight must be a non-negative integer")
        if any(sum(reel) <= 0 for reel in table["weights"]):
            raise ValueError("Every frontend initial stop-weight reel must have a positive total")
        if any(any(type(weight) is not int or weight < 0 for weight in reel) for reel in table["drop_weights"]):
            raise ValueError("Every frontend Symbol Drop Weight must be a non-negative integer")
        if any(sum(reel) <= 0 for reel in table["drop_weights"]):
            raise ValueError("Every frontend Symbol Drop Weight reel must have a positive total")
    result = {
        "bg_lengths": [len(reel) for reel in bg["reels"]],
        "fg_lengths": [len(reel) for reel in fg["reels"]],
        "bg_gold_counts": gold_counts(bg),
        "fg_gold_counts": gold_counts(fg),
        "bg_random_wild": bg["random_wild"],
        "fg_random_wild": fg["random_wild"],
        "bg_drop_weight_sums": [sum(reel) for reel in bg["drop_weights"]],
        "fg_drop_weight_sums": [sum(reel) for reel in fg["drop_weights"]],
    }
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert H0161.xlsx to the index frontend JS config or pure JSON")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--variant", type=Path, help="H016192A/H016194A multiplier workbook")
    args = parser.parse_args()
    source = args.source.resolve()
    output = args.output.resolve()
    variant = args.variant.resolve() if args.variant else None
    if variant is None and output.stem != "config":
        match = re.fullmatch(r"config_(\d{2})([A-Za-z]?)", output.stem)
        if match is not None:
            variant_suffix = match.group(1) + (match.group(2).upper() or "A")
            candidate = SOURCE_DIR / f"H0161{variant_suffix}.xlsx"
            if candidate.is_file():
                variant = candidate
    config = frontend_config(source, variant)
    result = validate(config)
    if output.suffix.lower() == ".json":
        output.write_text(json.dumps(config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    else:
        payload = json.dumps(config, ensure_ascii=False, separators=(",", ":"))
        output.write_text(
            f"// Generated from Source/{source.name} by 其他/工具/xlsx_to_config.py.\n"
            f"window.H016_CONFIG={payload};\n",
            encoding="utf-8",
        )
    result.update({"source": str(source), "output": str(output)})
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
