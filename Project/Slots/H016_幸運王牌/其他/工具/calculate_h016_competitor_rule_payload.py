"""Build H016 multiplier lines from the Super Ace relative Hit Rate shape."""
from __future__ import annotations

import argparse
import json
import math
import re
from copy import deepcopy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WEIGHT_TOTAL = 1_000_000_000
MIN_NATURAL_RATE = 0.001
MIN_MISSING_RTP = 0.002
MAX_RULE_UPPER = 2000.0
FG_SPECIAL_MULTIPLIERS = {"(10, 15]": 1.0, "(50, 60]": 1.5}
FEATURE_SPECIAL_MULTIPLIERS = {"(10, 15]": 1.6, "(50, 60]": 1.5}
BG_SPECIAL_MULTIPLIERS = {"(10, 15]": 1.6, "(50, 60]": 3.0}
HIT_MATCH_RANGES = {
    "(20, 25]", "(25, 30]", "(30, 35]",
    "(35, 40]", "(40, 45]", "(45, 50]", "(60, 70]",
}

BG_HIT_RATE_BOOSTS = {"(10, 15]": 2.0, "(50, 60]": 2.0}
FG_HIT_RATE_BOOSTS = {"(50, 60]": 2.0}
PROJECT_DIR = Path(__file__).resolve().parents[2]


def parse_upper(label: str) -> float:
    match = re.fullmatch(r"\(\s*-?[0-9.]+\s*,\s*([0-9.]+)\s*\]", label)
    if match is None:
        raise ValueError(f"Invalid range label: {label!r}")
    return float(match.group(1))


def parse_lower(label: str) -> float:
    match = re.fullmatch(r"\(\s*(-?[0-9.]+)\s*,\s*[0-9.]+\s*\]", label)
    if match is None:
        raise ValueError(f"Invalid range label: {label!r}")
    return float(match.group(1))


def enabled_bg_cap(sheet, labels: list[str]) -> float:
    """Return the highest enabled BG range upper bound for one Profile."""
    enabled = [
        parse_upper(label)
        for offset, label in enumerate(labels)
        if float(sheet.cell(row=15 + offset, column=11).value or 0) > 0
    ]
    if not enabled:
        raise ValueError(f"{sheet.title}: no positive BG range weight")
    return max(enabled)


def trigger_stats_at_cap(report: dict[str, Any], cap: float) -> dict[str, Any]:
    """Select cumulative trigger count/pay at the exact configured cap."""
    matches = [
        index
        for index, upper in enumerate(report["interval_upper"])
        if math.isclose(float(upper), float(cap), rel_tol=0.0, abs_tol=1e-9)
    ]
    if len(matches) != 1:
        raise ValueError(
            f"Report has no unique Multiplier Line upper bound for BG cap {cap:g}x"
        )
    index = matches[0]
    count = int(report["trigger_count_lte"][index])
    pay = float(report["trigger_pay_lte"][index])
    if count <= 0:
        raise ValueError(f"Report has no eligible FG-trigger BG samples at cap {cap:g}x")
    return {
        "cap": float(cap),
        "bucket_index": index,
        "count": count,
        "pay": pay,
        "average": pay / count / float(report["coin_in"]),
    }


def fix_numbers(weights: list[int], counts: list[int]) -> list[float]:
    denominator = sum(counts)
    if denominator <= 0:
        raise ValueError("Natural count denominator must be positive")
    result = []
    for weight, count in zip(weights, counts):
        if count <= 0:
            if weight:
                raise ValueError("Positive weight assigned to a range with no H016 samples")
            result.append(0.0)
        else:
            result.append((weight / WEIGHT_TOTAL) / (count / denominator))
    return result


def preserve_paying_hit_rate_shape(
    *,
    name: str,
    labels: list[str],
    baseline_weights: list[int],
    natural_counts: list[int],
    means: list[float],
    probability_denominator: int,
    target_scene_rtp: float,
) -> dict[str, Any]:
    """Scale every paying bucket equally and let the zero bucket absorb mass.

    This keeps the already-approved relative Hit Rate relationship intact while
    correcting only the scene RTP that changed when the trigger-spin BG average
    was replaced with the actual H016 value.
    """
    if len(baseline_weights) != 64 or sum(baseline_weights) != WEIGHT_TOTAL:
        raise ValueError(f"{name}: invalid 64-range baseline")
    if len(natural_counts) != 64 or len(means) != 64 or len(labels) != 64:
        raise ValueError(f"{name}: expected 64 natural buckets")
    if probability_denominator <= WEIGHT_TOTAL:
        raise ValueError(f"{name}: BG probability denominator must include FG entry weight")

    paying = [
        index for index in range(1, 64)
        if baseline_weights[index] > 0 and means[index] > 0
    ]
    current_rtp = sum(
        baseline_weights[index] / probability_denominator * means[index]
        for index in paying
    )
    if current_rtp <= 0 or target_scene_rtp <= 0:
        raise ValueError(f"{name}: invalid current/target paying RTP")
    scale = target_scene_rtp / current_rtp
    weights = [0] * 64
    exact = {
        index: baseline_weights[index] * scale
        for index in paying
    }
    for index in paying:
        weights[index] = round(exact[index])
    if sum(weights) > WEIGHT_TOTAL:
        raise ValueError(f"{name}: scaled paying weights exceed 1,000,000,000")
    weights[0] = WEIGHT_TOTAL - sum(weights)

    scene_rtp_after = sum(
        weight / probability_denominator * mean
        for weight, mean in zip(weights, means)
    )
    audits = []
    natural_total = sum(natural_counts)
    for index in range(64):
        after_rtp = weights[index] / probability_denominator * means[index]
        audits.append({
            "index": index,
            "range": labels[index],
            "rule": (
                "Zero-result bucket absorbs RTP correction"
                if index == 0
                else "Preserve approved relative Hit Rate shape"
            ),
            "natural_rate": natural_counts[index] / natural_total,
            "before_weight": baseline_weights[index],
            "after_weight": weights[index],
            "target_scene_rtp": after_rtp,
            "target_rtp": after_rtp,
            "after_hit_rate": weights[index] / probability_denominator,
            "after_rtp": after_rtp,
            "shape_scale": (
                weights[index] / baseline_weights[index]
                if baseline_weights[index] > 0 else None
            ),
        })
    return {
        "weights": weights,
        "fix": fix_numbers(weights, natural_counts),
        "audit": audits,
        "zero_bucket_before": baseline_weights[0],
        "zero_bucket_after": weights[0],
        "unrequested_mass_factor": scale,
        "calibration": "approved paying Hit Rate shape scaled uniformly; zero bucket absorbs mass",
        "target_scene_rtp": target_scene_rtp,
        "scene_rtp_before": current_rtp,
        "scene_rtp_after": scene_rtp_after,
    }


def workbook_scene(sheet, start_row: int, counts: list[int], means: list[float]) -> dict[str, Any]:
    weights = [int(sheet.cell(row, 11).value or 0) for row in range(start_row, start_row + 64)]
    delta = WEIGHT_TOTAL - sum(weights)
    if abs(delta) > 1:
        raise ValueError(f"{sheet.title}!K{start_row}:K{start_row + 63} does not sum to 1,000,000,000")
    if delta:
        last_active = max(index for index, weight in enumerate(weights) if weight > 0)
        weights[last_active] += delta
    labels = [str(sheet.cell(row, 1).value or "").strip() for row in range(start_row, start_row + 64)]
    scene_rtp = sum(weight / WEIGHT_TOTAL * mean for weight, mean in zip(weights, means))
    return {
        "weights": weights,
        "fix": fix_numbers(weights, counts),
        "audit": [],
        "zero_bucket_before": weights[0],
        "zero_bucket_after": weights[0],
        "unrequested_mass_factor": 1.0,
        "calibration": "preserved from current workbook",
        "target_scene_rtp": scene_rtp,
        "scene_rtp_before": scene_rtp,
        "scene_rtp_after": scene_rtp,
        "labels": labels,
    }


def load_h016_report(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        fields = {
            str(row[0]): row[1]
            for row in workbook["Base Info"].iter_rows(min_row=2, values_only=True)
            if row[0] is not None
        }
        table = list(
            workbook["Multiplier Line"].iter_rows(
                min_row=1, max_row=65, values_only=True
            )
        )
        if len(table) != 65:
            raise ValueError(f"{path.name}: expected header plus 64 multiplier buckets")
        header = {
            str(value): index
            for index, value in enumerate(table[0])
            if value is not None
        }
        required = {
            "Interval",
            "base_game_cnt",
            "base_game_pay",
            "free_game_cnt",
            "free_game_pay",
            "Interval_Upper",
            "bg_trigger_fg_cnt_lte_upper",
            "bg_trigger_fg_pay_lte_upper",
        }
        missing = required.difference(header)
        if missing:
            raise ValueError(
                f"{path.name}: missing Multiplier Line columns {sorted(missing)}; "
                "rerun the natural report with the current Simulator"
            )
        rows = table[1:]
        if len(rows) != 64:
            raise ValueError(f"{path.name}: expected 64 multiplier buckets")
        result = {
            "path": str(path.resolve()),
            "rounds": int(fields["total_rounds"]),
            "coin_in": float(fields["coin_in"]),
            "trigger_count": int(fields["bg_trigger_fg_cnt"]),
            "trigger_pay_for_active_profile": float(fields["bg_trigger_fg_pay"]),
            "interval_upper": [float(row[header["Interval_Upper"]]) for row in rows],
            "trigger_count_lte": [int(row[header["bg_trigger_fg_cnt_lte_upper"]] or 0) for row in rows],
            "trigger_pay_lte": [float(row[header["bg_trigger_fg_pay_lte_upper"]] or 0) for row in rows],
            "bg_count": [int(row[header["base_game_cnt"]] or 0) for row in rows],
            "bg_pay": [float(row[header["base_game_pay"]] or 0) for row in rows],
            "fg_count": [int(row[header["free_game_cnt"]] or 0) for row in rows],
            "fg_pay": [float(row[header["free_game_pay"]] or 0) for row in rows],
        }
    finally:
        workbook.close()
    if sum(result["bg_count"]) != result["rounds"]:
        raise ValueError(f"{path.name}: BG bucket count does not equal total rounds")
    if sum(result["fg_count"]) != result["trigger_count"]:
        raise ValueError(f"{path.name}: FG bucket count does not equal trigger count")
    if any(
        current < previous
        for previous, current in zip(result["trigger_count_lte"], result["trigger_count_lte"][1:])
    ):
        raise ValueError(f"{path.name}: cumulative trigger count is not monotonic")
    if any(
        current < previous
        for previous, current in zip(result["trigger_pay_lte"], result["trigger_pay_lte"][1:])
    ):
        raise ValueError(f"{path.name}: cumulative trigger pay is not monotonic")
    if result["trigger_count_lte"][-1] != result["trigger_count"]:
        raise ValueError(f"{path.name}: final cumulative trigger count differs from Base Info")
    result["trigger_pay"] = result["trigger_pay_lte"][-1]
    return result


def build_report_correction_version(
    key: str,
    workbook_path: Path,
    report: dict[str, Any],
) -> dict[str, Any]:
    workbook = load_workbook(
        workbook_path, read_only=True, data_only=True, keep_links=False
    )
    try:
        detail = workbook["Detail"]
        newbie_sheet = workbook["Detail_Newbie"]
        labels = [str(detail.cell(row, 1).value or "").strip() for row in range(15, 79)]
        bg_means = [
            pay / count / report["coin_in"] if count else 0.0
            for count, pay in zip(report["bg_count"], report["bg_pay"])
        ]
        fg_means = [
            pay / count / report["coin_in"] if count else 0.0
            for count, pay in zip(report["fg_count"], report["fg_pay"])
        ]
        for sheet in (detail, newbie_sheet):
            existing_bg_count = [int(sheet.cell(row, 2).value or 0) for row in range(15, 79)]
            existing_bg_pay = [float(sheet.cell(row, 3).value or 0) for row in range(15, 79)]
            existing_fg_count = [int(sheet.cell(row, 2).value or 0) for row in range(86, 150)]
            existing_fg_pay = [float(sheet.cell(row, 3).value or 0) for row in range(86, 150)]
            if existing_bg_count != report["bg_count"] or existing_fg_count != report["fg_count"]:
                raise ValueError(f"{workbook_path.name}/{sheet.title}: natural counts differ from report")
            if any(abs(a - b) > 1e-6 for a, b in zip(existing_bg_pay, report["bg_pay"])):
                raise ValueError(f"{workbook_path.name}/{sheet.title}: BG pay differs from report")
            if any(abs(a - b) > 1e-6 for a, b in zip(existing_fg_pay, report["fg_pay"])):
                raise ValueError(f"{workbook_path.name}/{sheet.title}: FG pay differs from report")

        entry_weight = int(detail["K79"].value or 0)
        if entry_weight != int(newbie_sheet["K79"].value or 0):
            raise ValueError(f"{workbook_path.name}: oldhand/newbie FG entry weights differ")
        denominator = WEIGHT_TOTAL + entry_weight
        trigger_rate = entry_weight / denominator
        oldhand_trigger = trigger_stats_at_cap(
            report, enabled_bg_cap(detail, labels)
        )
        newbie_trigger = trigger_stats_at_cap(
            report, enabled_bg_cap(newbie_sheet, labels)
        )
        oldhand_trigger_bg_rtp = trigger_rate * oldhand_trigger["average"]
        newbie_trigger_bg_rtp = trigger_rate * newbie_trigger["average"]
        oldhand_regular_bg_target = 0.70 - oldhand_trigger_bg_rtp
        newbie_regular_bg_target = 0.70 - newbie_trigger_bg_rtp
        if oldhand_regular_bg_target <= 0 or newbie_regular_bg_target <= 0:
            raise ValueError(f"{workbook_path.name}: trigger BG RTP exceeds the 70% BG target")

        old_bg = preserve_paying_hit_rate_shape(
            name=f"{key} oldhand BG",
            labels=labels,
            baseline_weights=[int(detail.cell(row, 11).value or 0) for row in range(15, 79)],
            natural_counts=report["bg_count"],
            means=bg_means,
            probability_denominator=denominator,
            target_scene_rtp=oldhand_regular_bg_target,
        )
        newbie_bg = preserve_paying_hit_rate_shape(
            name=f"{key} newbie BG",
            labels=labels,
            baseline_weights=[int(newbie_sheet.cell(row, 11).value or 0) for row in range(15, 79)],
            natural_counts=report["bg_count"],
            means=bg_means,
            probability_denominator=denominator,
            target_scene_rtp=newbie_regular_bg_target,
        )
        fg = workbook_scene(detail, 86, report["fg_count"], fg_means)
        newbie_fg = workbook_scene(newbie_sheet, 86, report["fg_count"], fg_means)
        bf = workbook_scene(detail, 163, report["fg_count"], fg_means)
        # SF owns a dedicated natural report and is outside this correction.
        # Read its existing workbook population so report-mode BG/FG/BF updates
        # cannot replace SF with ordinary FG samples.
        sf_counts = [int(detail.cell(row, 2).value or 0) for row in range(234, 298)]
        sf_pay = [float(detail.cell(row, 3).value or 0) for row in range(234, 298)]
        sf_means = [
            pay / count / report["coin_in"] if count else 0.0
            for count, pay in zip(sf_counts, sf_pay)
        ]
        sf = workbook_scene(detail, 234, sf_counts, sf_means)
    finally:
        workbook.close()

    fg_rtp = trigger_rate * fg["scene_rtp_after"]
    newbie_fg_rtp = trigger_rate * newbie_fg["scene_rtp_after"]
    expected_fg = 0.22 if key == "92" else 0.24
    if abs(fg_rtp - expected_fg) > 0.000003:
        raise ValueError(f"{workbook_path.name}: preserved FG RTP is {fg_rtp}, expected {expected_fg}")
    if abs(newbie_fg_rtp - 0.23) > 0.000003:
        raise ValueError(f"{workbook_path.name}: preserved newbie FG RTP is {newbie_fg_rtp}, expected 0.23")
    return {
        "bg": old_bg,
        "fg": fg,
        "newbie": {"bg": newbie_bg, "fg": newbie_fg},
        "bf": bf,
        "sf": sf,
        "metrics": {
            "target_rtp": float(key) / 100.0,
            "entry_weight": entry_weight,
            "trigger_rate": trigger_rate,
            "trigger_bg_cap": oldhand_trigger["cap"],
            "trigger_bg_count": oldhand_trigger["count"],
            "trigger_bg_pay": oldhand_trigger["pay"],
            "trigger_bg_average": oldhand_trigger["average"],
            "trigger_bg_rtp": oldhand_trigger_bg_rtp,
            "newbie_trigger_bg_cap": newbie_trigger["cap"],
            "newbie_trigger_bg_count": newbie_trigger["count"],
            "newbie_trigger_bg_pay": newbie_trigger["pay"],
            "newbie_trigger_bg_average": newbie_trigger["average"],
            "newbie_trigger_bg_rtp": newbie_trigger_bg_rtp,
            "normal_rtp": 0.70 + fg_rtp,
            "bg_rtp": 0.70,
            "fg_rtp": fg_rtp,
            "newbie_rtp": 0.70 + newbie_fg_rtp,
            "newbie_bg_rtp": 0.70,
            "newbie_fg_rtp": newbie_fg_rtp,
            "bf_rtp": bf["scene_rtp_after"] / 40.5,
            "sf_rtp": sf["scene_rtp_after"] / 250.0,
            "buy_price": 40.5,
            "super_price": 250.0,
        },
    }


def scale_integer_weights(weights: list[int], target: int) -> list[int]:
    total = sum(weights)
    if target < 0 or total <= 0:
        raise ValueError("Cannot scale free weights to the requested total")
    exact = [weight * target / total for weight in weights]
    result = [math.floor(value) for value in exact]
    remainder = target - sum(result)
    order = sorted(
        range(len(weights)), key=lambda index: exact[index] - result[index], reverse=True
    )
    for index in order[:remainder]:
        result[index] += 1
    return result


def tilted_weights(raw: list[int], means: list[float], target_mean: float, total: int) -> list[int]:
    active = [index for index, value in enumerate(raw) if value > 0]
    if not active:
        raise ValueError("No supported free ranges remain")
    low = min(means[index] for index in active)
    high = max(means[index] for index in active)
    if not low <= target_mean <= high:
        raise ValueError(f"Free target mean {target_mean} is outside {low}..{high}")

    def probabilities(lam: float) -> list[float]:
        logs = [math.log(raw[index]) + lam * means[index] for index in active]
        pivot = max(logs)
        values = [math.exp(value - pivot) for value in logs]
        denominator = sum(values)
        result = [0.0] * len(raw)
        for index, value in zip(active, values):
            result[index] = value / denominator
        return result

    lower, upper = -1.0, 1.0
    while sum(p * m for p, m in zip(probabilities(lower), means)) > target_mean:
        lower *= 2
    while sum(p * m for p, m in zip(probabilities(upper), means)) < target_mean:
        upper *= 2
    for _ in range(180):
        middle = (lower + upper) / 2
        mean = sum(p * m for p, m in zip(probabilities(middle), means))
        if mean < target_mean:
            lower = middle
        else:
            upper = middle
    probabilities_final = probabilities((lower + upper) / 2)
    exact = [value * total for value in probabilities_final]
    result = [math.floor(value) for value in exact]
    remainder = total - sum(result)
    order = sorted(
        range(len(result)), key=lambda index: exact[index] - result[index], reverse=True
    )
    for index in order[:remainder]:
        result[index] += 1
    return result


def relative_hit_rate_scene(
    *,
    name: str,
    labels: list[str],
    baseline_weights: list[int],
    natural_counts: list[int],
    means: list[float],
    competitor_counts: list[int],
    competitor_denominator: int,
    probability_denominator: int,
    target_scene_rtp: float,
    hit_rate_boosts: dict[str, float],
    maximum_supported_upper: float | None = None,
    minimum_supported_upper: float | None = None,
    minimum_bucket_rtp_share: float = 0.0,
    allow_zero_result_weight: bool = True,
) -> dict[str, Any]:
    """Scale one competitor Hit Rate shape to an exact scene RTP target."""
    if len(baseline_weights) != 64 or sum(baseline_weights) != WEIGHT_TOTAL:
        raise ValueError(f"{name}: invalid 64-range baseline")
    natural_total = sum(natural_counts)
    supported: list[tuple[int, float]] = []
    audits: list[dict[str, Any]] = []
    for index in range(1, 64):
        label = labels[index]
        natural_rate = natural_counts[index] / natural_total
        competitor_rate = competitor_counts[index] / competitor_denominator
        boost = hit_rate_boosts.get(label, 1.0)
        reason = None
        if natural_rate < MIN_NATURAL_RATE:
            reason = "H016 natural probability below 0.1%; weight disabled"
        elif minimum_supported_upper is not None and parse_upper(label) < minimum_supported_upper:
            reason = f"Minimum weighted range starts at {minimum_supported_upper:g}x; weight disabled"
        elif maximum_supported_upper is not None and parse_upper(label) > maximum_supported_upper:
            reason = f"Multiplier upper limit {maximum_supported_upper:g}x; weight disabled"
        elif means[index] <= 0:
            reason = "H016 range cannot pay; weight disabled"
        elif competitor_counts[index] <= 0 and minimum_bucket_rtp_share <= 0:
            reason = "Super Ace Hit Rate is zero; weight disabled"
        else:
            supported.append((index, competitor_counts[index] * boost))
        if reason is not None:
            audits.append({
                "index": index,
                "range": label,
                "rule": reason,
                "natural_rate": natural_rate,
                "competitor_rate": competitor_rate,
                "hit_rate_boost": boost,
                "before_weight": baseline_weights[index],
                "after_weight": 0,
                "target_scene_rtp": 0.0,
                "target_rtp": 0.0,
            })
    minimum_bucket_rtp = target_scene_rtp * minimum_bucket_rtp_share
    floor_weights = {
        index: math.ceil(minimum_bucket_rtp * probability_denominator / means[index])
        for index, _score in supported
    }
    floor_scene_rtp = sum(
        floor_weights[index] / probability_denominator * means[index]
        for index, _score in supported
    )
    if floor_scene_rtp > target_scene_rtp:
        raise ValueError(
            f"{name}: minimum bucket RTP shares require {floor_scene_rtp}, "
            f"above target {target_scene_rtp}"
        )
    score_rtp = sum(score * means[index] for index, score in supported)
    if score_rtp <= 0:
        raise ValueError(f"{name}: no supported competitor Hit Rate shape remains")

    weights = [0] * 64
    floor_active_indices: set[int] = set()
    if allow_zero_result_weight:
        def rtp_at(candidate_scale: float) -> float:
            return sum(
                max(floor_weights[index], candidate_scale * score)
                / probability_denominator
                * means[index]
                for index, score in supported
            )

        lower, upper = 0.0, target_scene_rtp * probability_denominator / score_rtp
        while rtp_at(upper) < target_scene_rtp:
            upper *= 2
        for _ in range(160):
            middle = (lower + upper) / 2
            if rtp_at(middle) < target_scene_rtp:
                lower = middle
            else:
                upper = middle
        scale = (lower + upper) / 2
        for index, score in supported:
            weights[index] = max(floor_weights[index], round(scale * score))
            if weights[index] == floor_weights[index]:
                floor_active_indices.add(index)
        weights[0] = WEIGHT_TOTAL - sum(weights[1:])
        if weights[0] < 0:
            raise ValueError(f"{name}: target requires more than 100% paying weight")
        calibration = "locked RTP with per-bucket floor, then closest scaled competitor Hit Rate shape"
    else:
        score_by_index = dict(supported)
        floor_active_indices = {
            index for index, score in supported if score <= 0
        }
        while True:
            free_indices = [
                index for index, _score in supported
                if index not in floor_active_indices
            ]
            fixed_weight = sum(floor_weights[index] for index in floor_active_indices)
            remaining_weight = WEIGHT_TOTAL - fixed_weight
            fixed_pay = sum(
                floor_weights[index] * means[index]
                for index in floor_active_indices
            )
            if remaining_weight <= 0:
                raise ValueError(f"{name}: no weight remains after bucket floors")
            target_free_mean = (
                target_scene_rtp * probability_denominator - fixed_pay
            ) / remaining_weight
            scaled = tilted_weights(
                [score_by_index[index] for index in free_indices],
                [means[index] for index in free_indices],
                target_free_mean,
                remaining_weight,
            )
            violations = {
                index for index, weight in zip(free_indices, scaled)
                if weight < floor_weights[index]
            }
            if not violations:
                for index in floor_active_indices:
                    weights[index] = floor_weights[index]
                for index, weight in zip(free_indices, scaled):
                    weights[index] = weight
                break
            floor_active_indices.update(violations)
        weights[0] = 0
        scale = 1.0
        calibration = "minimum-KL competitor Hit Rate shape with zero-result disabled, locked RTP and bucket floors"
    scene_rtp_after = sum(
        weight / probability_denominator * mean
        for weight, mean in zip(weights, means)
    )
    for index, score in supported:
        label = labels[index]
        competitor_rate = competitor_counts[index] / competitor_denominator
        boost = hit_rate_boosts.get(label, 1.0)
        after_hit_rate = weights[index] / probability_denominator
        after_rtp = after_hit_rate * means[index]
        floor_active = index in floor_active_indices
        if floor_active:
            rule = "Minimum 0.2% share of total FG RTP"
        else:
            rule = "Super Ace relative Hit Rate shape" + (f" x {boost:g}" if boost != 1 else "")
        audits.append({
            "index": index,
            "range": label,
            "rule": rule,
            "natural_rate": natural_counts[index] / natural_total,
            "competitor_rate": competitor_rate,
            "hit_rate_boost": boost,
            "before_weight": baseline_weights[index],
            "after_weight": weights[index],
            "target_scene_rtp": after_rtp,
            "target_rtp": after_rtp,
            "after_hit_rate": after_hit_rate,
            "after_rtp": after_rtp,
            "bucket_rtp_share": after_rtp / target_scene_rtp,
            "minimum_bucket_rtp_share": minimum_bucket_rtp_share,
            "floor_active": floor_active,
            "shape_scale": (
                after_hit_rate / (competitor_rate * boost)
                if competitor_rate > 0
                else None
            ),
        })
    audits.sort(key=lambda item: item["index"])
    return {
        "weights": weights,
        "fix": fix_numbers(weights, natural_counts),
        "audit": audits,
        "zero_bucket_before": baseline_weights[0],
        "zero_bucket_after": weights[0],
        "unrequested_mass_factor": scale,
        "calibration": calibration,
        "target_scene_rtp": target_scene_rtp,
        "scene_rtp_before": sum(
            weight / probability_denominator * mean
            for weight, mean in zip(baseline_weights, means)
        ),
        "scene_rtp_after": scene_rtp_after,
    }


def adjust_scene(
    *,
    name: str,
    labels: list[str],
    baseline_weights: list[int],
    natural_counts: list[int],
    means: list[float],
    competitor_counts: list[int],
    competitor_pay: list[float],
    competitor_denominator: int,
    probability_denominator: int,
    target_scene_rtp: float | None,
    special_multipliers: dict[str, float],
    minimum_supported_natural_rate: float | None = None,
    maximum_supported_upper: float | None = None,
    missing_rtp_floor: float = MIN_MISSING_RTP,
    overall_contribution_scale: float = 1.0,
    cycle_adjusted_floor_minimum_lower: float | None = None,
) -> dict[str, Any]:
    if len(baseline_weights) != 64:
        raise ValueError(f"{name}: expected 64 baseline weights")
    if sum(baseline_weights) != WEIGHT_TOTAL:
        raise ValueError(f"{name}: baseline weights do not sum to 1,000,000,000")
    if means[0] != 0:
        raise ValueError(f"{name}: first bucket must be the zero-win bucket")

    natural_total = sum(natural_counts)
    weights = list(baseline_weights)
    audits: list[dict[str, Any]] = []
    for index, label in enumerate(labels):
        natural_rate = natural_counts[index] / natural_total
        mean = means[index]
        current_rtp = baseline_weights[index] / probability_denominator * mean
        competitor_rate = competitor_counts[index] / competitor_denominator
        competitor_rtp = competitor_pay[index] / competitor_denominator
        rule = None
        target_rtp = current_rtp

        natural_unsupported = (
            index > 0
            and minimum_supported_natural_rate is not None
            and natural_rate < minimum_supported_natural_rate
        )
        cap_unsupported = (
            index > 0
            and maximum_supported_upper is not None
            and parse_upper(label) > maximum_supported_upper
        )
        unsupported = natural_unsupported or cap_unsupported
        if natural_unsupported:
            target_rtp = 0.0
            rule = "H016 natural probability below 0.1%; weight disabled"
        elif cap_unsupported:
            target_rtp = 0.0
            rule = f"Multiplier upper limit {maximum_supported_upper:g}x; weight disabled"
        elif label in special_multipliers:
            if competitor_counts[index] <= 0 or competitor_rtp <= 0:
                raise ValueError(
                    f"{name} {label}: competitor has no RTP for the "
                    f"{special_multipliers[label]:g}x rule"
                )
            target_rtp = competitor_rtp * special_multipliers[label]
            rule = f"Super Ace RTP x {special_multipliers[label]:g}"
        elif label in HIT_MATCH_RANGES:
            target_rtp = competitor_rate * mean
            rule = "Super Ace Hit Rate match"
        elif (
            parse_upper(label) <= MAX_RULE_UPPER
            and natural_rate >= MIN_NATURAL_RATE
            and mean > 0
            and baseline_weights[index] == 0
        ):
            cycle_adjusted_floor = (
                cycle_adjusted_floor_minimum_lower is not None
                and parse_lower(label) >= cycle_adjusted_floor_minimum_lower
            )
            target_rtp = max(
                current_rtp,
                missing_rtp_floor / overall_contribution_scale
                if cycle_adjusted_floor
                else missing_rtp_floor,
            )
            rule = (
                "H016 baseline absent; overall RTP floor 0.2% after FG cycle"
                if cycle_adjusted_floor
                else "H016 baseline absent; RTP floor 0.2%"
            )

        if rule is None:
            continue
        if unsupported:
            target_weight = 0
        else:
            exact_weight = target_rtp / mean * probability_denominator
            target_weight = (
                math.ceil(exact_weight)
                if "RTP floor 0.2%" in rule
                else round(exact_weight)
            )
        if target_weight < 0:
            raise ValueError(f"{name} {label}: negative target weight")
        weights[index] = target_weight
        audits.append({
            "index": index,
            "range": label,
            "rule": rule,
            "natural_rate": natural_rate,
            "competitor_count": competitor_counts[index],
            "competitor_rate": competitor_rate,
            "competitor_rtp": competitor_rtp,
            "before_weight": baseline_weights[index],
            "after_weight": target_weight,
            "before_rtp": current_rtp,
            "target_rtp": target_rtp,
            "target_scene_rtp": target_rtp,
            "target_overall_rtp": target_rtp * overall_contribution_scale,
            "target_hit_rate": target_rtp / mean if mean else 0.0,
        })

    fixed_indices = {item["index"] for item in audits}
    fixed_total = sum(weights[index] for index in fixed_indices)
    remaining = WEIGHT_TOTAL - fixed_total
    if remaining < 0:
        raise ValueError(f"{name}: requested fixed buckets alone exceed all available mass")
    if target_scene_rtp is None:
        free_indices = [index for index in range(1, 64) if index not in fixed_indices]
        free_total = sum(baseline_weights[index] for index in free_indices)
        weights[0] = remaining - free_total
        if weights[0] < 0:
            raise ValueError(f"{name}: zero-win bucket cannot absorb the requested rules")
        normalization_factor = 1.0
        calibration = "unrequested paying buckets unchanged; zero-win bucket absorbs mass"
    else:
        free_indices = [index for index in range(64) if index not in fixed_indices]
        fixed_scene_rtp = sum(
            weights[index] / probability_denominator * means[index]
            for index in fixed_indices
        )
        free_target_mean = (
            target_scene_rtp - fixed_scene_rtp
        ) * probability_denominator / remaining
        raw_free = [
            baseline_weights[index]
            if baseline_weights[index] > 0
            else natural_counts[index]
            for index in free_indices
        ]
        means_free = [means[index] for index in free_indices]
        scaled = tilted_weights(raw_free, means_free, free_target_mean, remaining)
        for index, weight in zip(free_indices, scaled):
            weights[index] = weight
        normalization_factor = remaining / sum(raw_free)
        calibration = "minimum-KL tilt of unrequested buckets to the cycle-adjusted scene RTP"
    if sum(weights) != WEIGHT_TOTAL:
        raise AssertionError(f"{name}: adjusted weights do not sum to 1,000,000,000")
    for item in audits:
        index = item["index"]
        item["after_scene_rtp"] = weights[index] / probability_denominator * means[index]
        item["after_rtp"] = item["after_scene_rtp"]
        item["after_overall_rtp"] = item["after_scene_rtp"] * overall_contribution_scale
        item["after_hit_rate"] = weights[index] / probability_denominator
        item["rounding_error"] = item["after_rtp"] - item["target_rtp"]

    return {
        "weights": weights,
        "fix": fix_numbers(weights, natural_counts),
        "audit": audits,
        "zero_bucket_before": baseline_weights[0],
        "zero_bucket_after": weights[0],
        "unrequested_mass_factor": normalization_factor,
        "calibration": calibration,
        "target_scene_rtp": target_scene_rtp,
        "scene_rtp_before": sum(
            weight / probability_denominator * mean
            for weight, mean in zip(baseline_weights, means)
        ),
        "scene_rtp_after": sum(
            weight / probability_denominator * mean
            for weight, mean in zip(weights, means)
        ),
    }


def build_version(
    key: str, source: dict[str, Any], previous_version: dict[str, Any]
) -> dict[str, Any]:
    base = source["versions"][key]
    shared = source["versions"]["92"]
    h016 = source["h016"]
    competitor = source["competitor"]
    labels = source["ranges"]
    entry_weight = int(base["bg_weights"][64])
    trigger_rate = entry_weight / (WEIGHT_TOTAL + entry_weight)
    competitor_trigger_rate = float(competitor["trigger_count"]) / float(competitor["rounds"])
    if abs(trigger_rate - competitor_trigger_rate) > 1e-9:
        raise ValueError(
            f"{key}: FG cycle {trigger_rate:.12%} does not match Super Ace "
            f"{competitor_trigger_rate:.12%}"
        )
    buy_price = 40.5
    super_price = 250.0

    def normal_line(
        name: str,
        target_bg_rtp: float,
        target_fg_rtp: float,
        preserved_bg: dict[str, Any],
        fg_baseline: list[int],
        *,
        fg_maximum: float | None = None,
    ) -> dict[str, Any]:
        if "trigger_bg_average" not in h016:
            raise ValueError(
                "Baseline is missing h016.trigger_bg_average; do not substitute "
                "the Super Ace trigger-spin average"
            )
        trigger_bg_average = float(h016["trigger_bg_average"])
        trigger_bg_rtp = trigger_rate * trigger_bg_average
        bg = preserve_paying_hit_rate_shape(
            name=f"{name} BG",
            labels=labels,
            baseline_weights=[int(value) for value in preserved_bg["weights"]],
            natural_counts=h016["bg_count"],
            means=h016["bg_mean"],
            probability_denominator=WEIGHT_TOTAL + entry_weight,
            target_scene_rtp=target_bg_rtp - trigger_bg_rtp,
        )
        fg_target = target_fg_rtp / trigger_rate
        fg = relative_hit_rate_scene(
            name=f"{name} FG",
            labels=labels,
            baseline_weights=[int(value) for value in fg_baseline],
            natural_counts=h016["fg_count"],
            means=h016["fg_mean"],
            competitor_counts=competitor["fg_count"],
            competitor_denominator=int(competitor["trigger_count"]),
            probability_denominator=WEIGHT_TOTAL,
            target_scene_rtp=fg_target,
            hit_rate_boosts=FG_HIT_RATE_BOOSTS,
            maximum_supported_upper=fg_maximum,
            minimum_supported_upper=6.0,
            minimum_bucket_rtp_share=0.002,
            allow_zero_result_weight=False,
        )
        actual_bg = bg["scene_rtp_after"] + trigger_bg_rtp
        actual_fg = trigger_rate * fg["scene_rtp_after"]
        actual = actual_bg + actual_fg
        return {
            "bg": bg,
            "fg": fg,
            "metrics": {
                "target_rtp": target_bg_rtp + target_fg_rtp,
                "normal_rtp": actual,
                "target_bg_rtp": target_bg_rtp,
                "target_fg_rtp": target_fg_rtp,
                "bg_rtp": actual_bg,
                "fg_rtp": actual_fg,
                "fg_target_session_rtp": fg_target,
                "trigger_bg_rtp": trigger_bg_rtp,
            },
        }

    oldhand = normal_line(
        f"{key} oldhand",
        0.70,
        0.22 if key == "92" else 0.24,
        previous_version["bg"],
        base["fg_weights"],
    )
    newbie = normal_line(
        f"{key} newbie",
        0.70,
        0.23,
        previous_version["newbie"]["bg"],
        shared["fg_weights"],
        fg_maximum=120.0,
    )
    bf = relative_hit_rate_scene(
        name=f"{key} BF",
        labels=labels,
        baseline_weights=[int(value) for value in previous_version["bf"]["weights"]],
        natural_counts=h016["fg_count"],
        means=h016["fg_mean"],
        competitor_counts=competitor["fg_count"],
        competitor_denominator=int(competitor["trigger_count"]),
        probability_denominator=WEIGHT_TOTAL,
        target_scene_rtp=0.925 * buy_price,
        hit_rate_boosts=FG_HIT_RATE_BOOSTS,
        minimum_supported_upper=6.0,
        minimum_bucket_rtp_share=0.002,
        allow_zero_result_weight=False,
    )
    sf = deepcopy(previous_version["sf"])
    bf_rtp = bf["scene_rtp_after"] / buy_price
    sf_rtp = sf["scene_rtp_after"] / super_price
    return {
        "bg": oldhand["bg"],
        "fg": oldhand["fg"],
        "newbie": newbie,
        "bf": bf,
        "sf": sf,
        "metrics": {
            "target_rtp": float(key) / 100.0,
            "trigger_rate": trigger_rate,
            "competitor_trigger_rate": competitor_trigger_rate,
            "normal_rtp": oldhand["metrics"]["normal_rtp"],
            "bg_rtp": oldhand["metrics"]["bg_rtp"],
            "fg_rtp": oldhand["metrics"]["fg_rtp"],
            "newbie_rtp": newbie["metrics"]["normal_rtp"],
            "newbie_bg_rtp": newbie["metrics"]["bg_rtp"],
            "newbie_fg_rtp": newbie["metrics"]["fg_rtp"],
            "bf_rtp": bf_rtp,
            "sf_rtp": sf_rtp,
            "buy_price": buy_price,
            "super_price": super_price,
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--baseline", type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--previous-payload", type=Path)
    parser.add_argument(
        "--report",
        type=Path,
        help="Correct current 92A/94A workbooks from one H016 card-off 10^9 report",
    )
    args = parser.parse_args()
    source_report = None
    if args.report is not None:
        if args.baseline is not None or args.previous_payload is not None:
            parser.error("--report cannot be combined with --baseline/--previous-payload")
        report = load_h016_report(args.report)
        source_report = {
            "path": report["path"],
            "rounds": report["rounds"],
            "coin_in": report["coin_in"],
            "trigger_count": report["trigger_count"],
            "trigger_pay": report["trigger_pay"],
            "trigger_average": (
                report["trigger_pay"] / report["trigger_count"] / report["coin_in"]
            ),
            "bg_count": report["bg_count"],
            "bg_pay": report["bg_pay"],
            "fg_count": report["fg_count"],
            "fg_pay": report["fg_pay"],
            "interval_upper": report["interval_upper"],
            "trigger_count_lte": report["trigger_count_lte"],
            "trigger_pay_lte": report["trigger_pay_lte"],
        }
        version_94 = build_report_correction_version(
            "94", PROJECT_DIR / "Source" / "H016194A.xlsx", report
        )
        version_92 = build_report_correction_version(
            "92",
            PROJECT_DIR / "Source" / "H016192A.xlsx",
            report,
        )
        versions = {"92": version_92, "94": version_94}
    else:
        if args.baseline is None or args.previous_payload is None:
            parser.error("--baseline and --previous-payload are required without --report")
        source = json.loads(args.baseline.read_text(encoding="utf-8"))
        previous = json.loads(args.previous_payload.read_text(encoding="utf-8"))
        versions = {
            key: build_version(key, source, previous["versions"][key])
            for key in ("92", "94")
        }
    payload = {
        "rules": {
            "line_shape": "FG locks RTP and per-bucket floor first, then preserves the closest possible Super Ace relative Hit Rate shape after boosts",
            "minimum_h016_natural_rate": MIN_NATURAL_RATE,
            "hit_rate_boosts": {"bg": BG_HIT_RATE_BOOSTS, "fg": FG_HIT_RATE_BOOSTS},
            "minimum_weighted_fg_range": "(5, 6]; the (-1, 0] zero-result bucket is also disabled",
            "minimum_eligible_fg_bucket_rtp_share": 0.002,
            "bg_rule": (
                "Approved paying-bucket Hit Rate shape is preserved proportionally; "
                "zero bucket absorbs the correction required by actual H016 trigger-spin BG pay"
            ),
            "bf_rule": "BF independently uses the same zero-free FG line-shape method at 92.5% RTP",
            "sf_rule": "SF weights are preserved from the previous payload",
            "targets": {
                "oldhand_bg": 0.70,
                "oldhand_fg": {"92": 0.22, "94": 0.24},
                "newbie_bg": 0.70,
                "newbie_fg": 0.23,
                "newbie_normal": 0.93,
                "bf": 0.925,
                "sf": "preserved",
            },
            "newbie_multiplier_caps": {"bg": 30.0, "fg": 120.0},
            "mass_balance": "BG and cycle-adjusted FG are independently scaled to their locked RTP components",
            "fg_cycle": "matches the Super Ace trigger rate",
        },
        "versions": versions,
    }
    if source_report is not None:
        payload["source_report"] = source_report
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "output": str(args.output.resolve()),
        "versions": {
            key: {
                scene: {
                    "changed_rule_buckets": len(versions[key][scene]["audit"]),
                    "rtp_before": versions[key][scene]["scene_rtp_before"],
                    "rtp_after": versions[key][scene]["scene_rtp_after"],
                    "zero_weight_before": versions[key][scene]["zero_bucket_before"],
                    "zero_weight_after": versions[key][scene]["zero_bucket_after"],
                    "unrequested_mass_factor": versions[key][scene]["unrequested_mass_factor"],
                }
                for scene in ("bg", "fg", "bf", "sf")
            }
            for key in ("92", "94")
        },
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
