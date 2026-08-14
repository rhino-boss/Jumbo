from __future__ import annotations

import importlib.util
import json
import math
import os
import sys
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


H016_DIR = Path(__file__).resolve().parents[2]
OUTPUT = H016_DIR / "其他" / "競品參考數值比較.md"
RECORD_DIR = H016_DIR / "Record"
MODEL_CONFIG = H016_DIR / "config.js"
SIMULATOR = H016_DIR / "Simulator.py"
SOURCE_APPLIER = H016_DIR / "其他" / "工具" / "apply_super_ace_claude.py"
SOURCE_REELS = H016_DIR / "其他" / "參考資料" / "Super Ace_claude.txt"
SIMULATION_ROUNDS = 100_000
COMPETITOR_ROOT = Path(
    "C:/Users/rhinshen/Mine/個人工作區/市場資訊/H5/遊戲資源/JILI"
)
COMPETITOR_FILENAMES = (
    "SuperAce_BG_Combined_NoJP.jsonl",
    "SuperAce_BG_3.jsonl",
    "Super_Ace_BG_4.jsonl",
)
COMPETITOR_DIR = next(
    (
        candidate / "遊戲資料"
        for candidate in (
            COMPETITOR_ROOT / "JILI - Super Ace",
            COMPETITOR_ROOT / "JILI - Super Ace - m",
        )
        if all((candidate / "遊戲資料" / name).is_file() for name in COMPETITOR_FILENAMES)
    ),
    COMPETITOR_ROOT / "JILI - Super Ace" / "遊戲資料",
)
FILES = [COMPETITOR_DIR / name for name in COMPETITOR_FILENAMES]
SYMBOL_MAP = {
    "Bonus": "C1", "Symbol1": "M1", "Symbol2": "M2", "Symbol3": "M3",
    "Symbol4": "M4", "Symbol5": "A", "Symbol6": "K", "Symbol7": "Q", "Symbol8": "J",
}
SYMBOLS = ["C1", "M1", "M2", "M3", "M4", "A", "K", "Q", "J"]
GOLD_TO_BASE = {
    "G1": "M1", "G2": "M2", "G3": "M3", "G4": "M4",
    "GA": "A", "GK": "K", "GQ": "Q", "GJ": "J",
    "M1G": "M1", "M2G": "M2", "M3G": "M3", "M4G": "M4",
    "AG": "A", "KG": "K", "QG": "Q", "JG": "J",
}
PRIMARY_TABLES = {
    "BG_Symbol": "bg_1", "BG_Symbol (2)": "bg_2", "BG_Symbol (3)": "bg_3",
    "FG_Symbol": "fg_1", "FG_Symbol (2)": "fg_2", "FG_Symbol (3)": "fg_3",
}
PRIMARY_BY_SCENE = {
    "BG": ("bg_1", "bg_2", "bg_3"),
    "FG": ("fg_1", "fg_2", "fg_3"),
}
ALL_TABLES_BY_SCENE = {
    "BG": ("bg_1", "bg_2", "bg_3", "bg_high", "bg_low", "buy"),
    "FG": (
        "fg_1", "fg_2", "fg_3", "fg_high_a", "fg_high_k", "fg_high_q",
        "fg_high_j", "fg_low", "super",
    ),
}
ALIAS_TO_PRIMARY = {
    "bg_high": "bg_1", "bg_low": "bg_2", "buy": "bg_3",
    "fg_high_a": "fg_1", "fg_high_k": "fg_2", "fg_high_q": "fg_3",
    "fg_high_j": "fg_1", "fg_low": "fg_1", "super": "fg_2",
}
SUPER_ACE_PAYS = {
    "M1": (0.50, 1.50, 2.50), "M2": (0.40, 1.20, 2.00),
    "M3": (0.30, 0.90, 1.50), "M4": (0.20, 0.60, 1.00),
    "A": (0.10, 0.30, 0.50), "K": (0.10, 0.30, 0.50),
    "Q": (0.05, 0.15, 0.25), "J": (0.05, 0.15, 0.25),
}


def load_model_config() -> dict[str, Any]:
    raw = MODEL_CONFIG.read_text(encoding="utf-8-sig")
    start, end = raw.find("{"), raw.rfind("}")
    if start < 0 or end <= start:
        raise ValueError(f"Config does not contain a JSON object: {MODEL_CONFIG}")
    config = json.loads(raw[start : end + 1])
    if config.get("game_id") != "H016" or not isinstance(config.get("tables"), dict):
        raise ValueError(f"Not a valid H016 config: {MODEL_CONFIG}")
    missing = set(PRIMARY_TABLES.values()).difference(config["tables"])
    if missing:
        raise ValueError(f"{MODEL_CONFIG.name}: missing tables {sorted(missing)}")
    return config


def table_key(sheet_or_table: str) -> str:
    return PRIMARY_TABLES.get(sheet_or_table, sheet_or_table)


def symbol_name(config: dict[str, Any], symbol_id: int, merge_gold: bool = False) -> str:
    name = str(config["symbol_names"][str(int(symbol_id))])
    return GOLD_TO_BASE.get(name, name) if merge_gold else name


def pct(value: float, digits: int = 4) -> str:
    return f"{value * 100:.{digits}f}%"


def pp(value: float) -> str:
    return f"{value * 100:+.4f} pp".replace("+0.0000", "0.0000")


def relative_difference(reference: float, actual: float) -> str:
    """Return (A-B)/A with A=Super Ace and B=H016."""
    if math.isclose(reference, 0.0, rel_tol=0.0, abs_tol=1e-12):
        return "N/A"
    return f"{(reference - actual) / reference * 100:+.4f}%".replace(
        "+0.0000", "0.0000"
    )


def visible_stack_counts(symbols: list[str]) -> Counter[tuple[str, int]]:
    """Count exact maximal vertical runs in one visible reel window."""
    result: Counter[tuple[str, int]] = Counter()
    if not symbols:
        return result
    current = symbols[0]
    length = 1
    for symbol in symbols[1:]:
        if symbol == current:
            length += 1
            continue
        if length >= 2:
            result[(current, length)] += 1
        current = symbol
        length = 1
    if length >= 2:
        result[(current, length)] += 1
    return result


def raw_competitor() -> dict[str, Any]:
    counts = {
        scene: {
            "initial": [Counter() for _ in range(5)], "drop": [Counter() for _ in range(5)],
            "gold_initial": Counter(), "gold_drop": Counter(),
            "gold_symbol_initial": [Counter() for _ in range(5)],
            "gold_symbol_drop": [Counter() for _ in range(5)],
            "initial_total": Counter(), "drop_total": Counter(),
            "combo": Counter(), "symbol_length_hits": Counter(),
            "symbol_length_win": Counter(), "spins": 0, "hits": 0,
            "stack": [Counter() for _ in range(5)], "stack_total": Counter(),
        }
        for scene in ("BG", "FG")
    }
    coin_in = total_win = bg_win = fg_win = 0.0
    triggers = big_events = 0
    bg_w2_counts: Counter[int] = Counter()
    for path in FILES:
        with path.open(encoding="utf-8") as handle:
            for line in handle:
                obj = json.loads(line)
                plates = obj["plate"]["plate"]
                coin_in += float(obj["bet"])
                total_win += float(obj["win"])
                triggers += int(len(plates) > 1)
                for plate_index, plate in enumerate(plates):
                    scene = "BG" if plate_index == 0 else "FG"
                    data = counts[scene]
                    data["spins"] += 1
                    win = float(plate.get("win", 0.0))
                    data["hits"] += int(win > 0)
                    if scene == "BG":
                        bg_win += win
                    else:
                        fg_win += win
                    for reel, column in enumerate(plate["column"]):
                        visible_symbols = []
                        for symbol, gold in zip(column["row"], column["isGold"]):
                            mapped = SYMBOL_MAP[symbol]
                            visible_symbols.append(mapped)
                            data["initial"][reel][mapped] += 1
                            data["initial_total"][reel] += 1
                            data["gold_initial"][reel] += int(gold in (1, 2))
                            data["gold_symbol_initial"][reel][mapped] += int(gold in (1, 2))
                        data["stack"][reel].update(visible_stack_counts(visible_symbols))
                        data["stack_total"][reel] += 1
                    combos = plate.get("combo", [])
                    data["combo"][min(len(combos), 5)] += 1
                    for combo in combos:
                        combo_bonus = float(combo.get("comboBonus", 1.0))
                        for award in combo.get("award", []):
                            symbol = SYMBOL_MAP.get(str(award.get("symbol")))
                            length = int(award.get("maxLen", 0))
                            if symbol in SYMBOLS and length in (3, 4, 5):
                                key = (symbol, length)
                                data["symbol_length_hits"][key] += 1
                                data["symbol_length_win"][key] += float(award.get("win", 0.0)) * combo_bonus
                        if scene == "BG":
                            made = sum(change.get("isGold") == 102 for change in combo.get("change", []))
                            if made:
                                # One source gold becomes W2 plus 2/3/4 additional WW,
                                # so the config-facing Random Wild value is made - 1.
                                big_events += 1
                                bg_w2_counts[made - 1] += 1
                        for change in combo.get("change", []):
                            if "symbol" not in change:
                                continue
                            reel = int(change.get("column", 0))
                            mapped = SYMBOL_MAP[change["symbol"]]
                            data["drop"][reel][mapped] += 1
                            data["drop_total"][reel] += 1
                            data["gold_drop"][reel] += int(change.get("isGold") in (1, 2))
                            data["gold_symbol_drop"][reel][mapped] += int(change.get("isGold") in (1, 2))
    bg_spins = counts["BG"]["spins"]
    fg_spins = counts["FG"]["spins"]
    return {
        "counts": counts, "rounds": bg_spins, "fg_spins": fg_spins, "coin_in": coin_in,
        "rtp_total": total_win / coin_in, "rtp_bg": bg_win / coin_in, "rtp_fg": fg_win / coin_in,
        "bg_hit_rate": counts["BG"]["hits"] / bg_spins, "fg_hit_rate": counts["FG"]["hits"] / fg_spins,
        "fg_trigger_rate": triggers / bg_spins, "avg_fg_spins": fg_spins / triggers,
        "w2_bg_event_rate": big_events / bg_spins, "w2_bg_counts": bg_w2_counts,
    }


def record_data(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    summary = {row[0]: row[1] for row in workbook["Base Info"].iter_rows(min_row=2, values_only=True)}
    combo = {str(row[0]): {"BG": int(row[1]), "FG": int(row[2])} for row in workbook["Eliminate"].iter_rows(min_row=2, values_only=True)}
    ratios: dict[str, dict[str, list[float]]] = {}
    gold_ratios: dict[str, list[float]] = {}
    for scene, stage, sheet_name in (
        ("BG", "initial", "BG Initial Symbol"), ("BG", "drop", "BG Drop Symbol"),
        ("FG", "initial", "FG Initial Symbol"), ("FG", "drop", "FG Drop Symbol"),
    ):
        merged = {symbol: [0.0] * 5 for symbol in SYMBOLS}
        gold = [0.0] * 5
        for row in workbook[sheet_name].iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            symbol = GOLD_TO_BASE.get(str(row[0]), str(row[0]))
            if str(row[0]) in GOLD_TO_BASE:
                for reel in range(5):
                    gold[reel] += float(row[1 + reel] or 0)
            if symbol in merged:
                for reel in range(5):
                    merged[symbol][reel] += float(row[1 + reel] or 0)
        ratios[f"{scene}_{stage}"] = merged
        gold_ratios[f"{scene}_{stage}"] = gold
    fg_spins = sum(row["FG"] for row in combo.values())
    symbol_length = {scene: {"hits": Counter(), "pay": Counter()} for scene in ("BG", "FG")}
    if "Symbol Length" not in workbook.sheetnames:
        raise ValueError(f"{path.name}: missing Symbol Length; rerun Simulator.py with the current statistics schema")
    for scene, symbol, length, hits, pay in workbook["Symbol Length"].iter_rows(min_row=2, values_only=True):
        if scene in symbol_length and symbol in SYMBOLS and int(length) in (3, 4, 5):
            key = (str(symbol), int(length))
            symbol_length[str(scene)]["hits"][key] = int(hits or 0)
            symbol_length[str(scene)]["pay"][key] = float(pay or 0.0)
    workbook.close()
    return {
        "summary": summary, "combo": combo, "ratios": ratios, "gold_ratios": gold_ratios,
        "fg_spins": fg_spins, "symbol_length": symbol_length,
    }


def load_simulator_module():
    module_name = "h016_simulator_for_competitor_report"
    environment = {
        "H016_BASE_DIR": str(H016_DIR),
        "H016_CONFIG_FILE": MODEL_CONFIG.name,
        "H016_CONFIG_RTP_FILE": MODEL_CONFIG.name,
        "H016_RUN_ALL_COMBINATIONS": "false",
        "H016_CARD_SYSTEM_ENABLED": "false",
        "H016_OUTPUT_REPORT": "false",
        "H016_SHOW_CONSOLE_SUMMARY": "false",
        "H016_SHOW_CONSOLE_DETAIL": "false",
    }
    previous = {name: os.environ.get(name) for name in environment}
    previous_dont_write = sys.dont_write_bytecode
    os.environ.update(environment)
    sys.dont_write_bytecode = True
    try:
        spec = importlib.util.spec_from_file_location(module_name, SIMULATOR)
        if spec is None or spec.loader is None:
            raise RuntimeError(f"Cannot load {SIMULATOR}")
        module = importlib.util.module_from_spec(spec)
        sys.modules[module_name] = module
        spec.loader.exec_module(module)
        return module
    finally:
        sys.dont_write_bytecode = previous_dont_write
        for name, value in previous.items():
            if value is None:
                os.environ.pop(name, None)
            else:
                os.environ[name] = value


def observed_ratios(config: dict[str, Any], counter: Counter) -> tuple[dict[str, list[float]], list[float]]:
    merged = {symbol: [0.0] * 5 for symbol in SYMBOLS}
    gold_counts = [0.0] * 5
    totals = [0.0] * 5
    for (reel, symbol_id), count in counter.items():
        reel = int(reel)
        symbol_id = int(symbol_id)
        count = float(count)
        totals[reel] += count
        if 11 <= symbol_id <= 18:
            gold_counts[reel] += count
        name = symbol_name(config, symbol_id, merge_gold=True)
        if name in merged:
            merged[name][reel] += count
    for reel in range(5):
        denominator = max(1.0, totals[reel])
        gold_counts[reel] /= denominator
        for symbol in SYMBOLS:
            merged[symbol][reel] /= denominator
    return merged, gold_counts


def simulation_data(config: dict[str, Any]) -> dict[str, Any]:
    """Run config.js in memory and expose the same shape as record_data()."""
    simulator = load_simulator_module()
    threads = max(1, int(os.environ.get("H016_REPORT_THREADS", simulator.THREADS)))
    result = simulator.run_simulation(
        total_rounds=SIMULATION_ROUNDS,
        bet_mode=simulator.MODE_NORMALBET,
        bet_multi=1,
        threads=threads,
        config=config,
    )
    stats = result["stats"]
    summary = dict(simulator.summary_rows(result))
    combo = {
        label: {"BG": int(stats["combo_bg"][index]), "FG": int(stats["combo_fg"][index])}
        for index, label in enumerate(("0", "1", "2", "3", "4", "5+"))
    }
    ratios: dict[str, dict[str, list[float]]] = {}
    gold_ratios: dict[str, list[float]] = {}
    for scene, stage, field in (
        ("BG", "initial", "bg_initial_symbols"), ("BG", "drop", "bg_drop_symbols"),
        ("FG", "initial", "fg_initial_symbols"), ("FG", "drop", "fg_drop_symbols"),
    ):
        ratios[f"{scene}_{stage}"], gold_ratios[f"{scene}_{stage}"] = observed_ratios(
            config, stats[field]
        )
    symbol_length = {scene: {"hits": Counter(), "pay": Counter()} for scene in ("BG", "FG")}
    for scene in ("BG", "FG"):
        for metric, suffix in (("hits", "hits"), ("pay", "pay")):
            counter = stats[f"{scene.lower()}_symbol_length_{suffix}"]
            for (symbol_id, length), value in counter.items():
                name = symbol_name(config, int(symbol_id), merge_gold=True)
                if name in SYMBOLS and int(length) in (3, 4, 5):
                    symbol_length[scene][metric][(name, int(length))] = value
    return {
        "summary": summary,
        "combo": combo,
        "ratios": ratios,
        "gold_ratios": gold_ratios,
        "fg_spins": int(stats["fg_spins"]),
        "symbol_length": symbol_length,
    }


def h016_data(config: dict[str, Any]) -> tuple[dict[str, Any], str, str | None]:
    # This report is the config-only calibration view.  A historical Record may
    # have been produced from Source/H0161.xlsx or from an older config, so it is
    # not a trustworthy cache for the current config.js.
    return (
        simulation_data(config),
        "`config.js` + `Simulator.py`（記憶體 100,000 場；未產生 Record xlsx）",
        None,
    )


def competitor_ratios(competitor: dict[str, Any], scene: str, stage: str) -> dict[str, list[float]]:
    data = competitor["counts"][scene]
    return {
        symbol: [data[stage][reel][symbol] / max(1, data[f"{stage}_total"][reel]) for reel in range(5)]
        for symbol in SYMBOLS
    }


def model_drop_ratios(config: dict[str, Any], sheet_or_table: str) -> dict[str, list[float]]:
    table = config["tables"][table_key(sheet_or_table)]
    result = {symbol: [0.0] * 5 for symbol in SYMBOLS}
    for reel, (values, weights) in enumerate(zip(table["drop_values"], table["drop_weights"])):
        total = sum(max(0.0, float(weight)) for weight in weights)
        for value, weight in zip(values, weights):
            name = symbol_name(config, int(value), merge_gold=True)
            if name in result and total > 0:
                result[name][reel] += max(0.0, float(weight)) / total
    return result


def model_random_wild_weights(config: dict[str, Any], sheet_or_table: str) -> list[int]:
    table = config["tables"][table_key(sheet_or_table)]
    random_wild = table["random_wild"]
    values = list(map(int, random_wild["values"]))
    weights = list(map(int, random_wild["weights"]))
    if values != [0, 2, 3, 4] or len(values) != len(weights):
        raise ValueError(f"{table_key(sheet_or_table)} Random Wild must define 0/2/3/4")
    return weights


def model_rng_stack_data(config: dict[str, Any], sheet_or_table: str) -> dict[str, Any]:
    """Enumerate four-cell screens using each stop's integer RNG weight."""
    table_name = table_key(sheet_or_table)
    table = config["tables"][table_name]
    counts = [Counter() for _ in range(5)]
    totals = [0.0] * 5
    for reel in range(5):
        symbols = [symbol_name(config, int(value), merge_gold=True) for value in table["reels"][reel]]
        weights = list(map(float, table["weights"][reel]))
        if len(symbols) != len(weights):
            raise ValueError(f"{table_name} R{reel + 1}: reel and stop-weight lengths differ")
        for stop, weight in enumerate(weights):
            if weight <= 0:
                continue
            window = [symbols[(stop + row) % len(symbols)] for row in range(4)]
            for key, occurrences in visible_stack_counts(window).items():
                counts[reel][key] += occurrences * weight
            totals[reel] += weight
    return {"counts": counts, "totals": totals}


def selection_items(config: dict[str, Any], group: str) -> list[tuple[str, float]]:
    result = []
    for item in config.get("table_selection", {}).get(group, []):
        table_name = str(item.get("table", ""))
        weight = float(item.get("weight", 0))
        if table_name in config["tables"] and weight > 0:
            result.append((table_name, weight))
    return result


def mixed_rng_stack_data(config: dict[str, Any], group: str) -> dict[str, Any]:
    """Mix each table's normalized stop RNG rates by Table Selection weight."""
    items = selection_items(config, group)
    selection_total = sum(weight for _, weight in items)
    counts = [Counter() for _ in range(5)]
    totals = [1.0 if selection_total > 0 else 0.0 for _ in range(5)]
    for table_name, selection in items:
        table_data = model_rng_stack_data(config, table_name)
        table_probability = selection / selection_total
        for reel in range(5):
            denominator = table_data["totals"][reel]
            if denominator <= 0:
                continue
            for key, count in table_data["counts"][reel].items():
                counts[reel][key] += table_probability * count / denominator
    return {"counts": counts, "totals": totals}


def stack_section(
    scene: str,
    competitor: dict[str, Any],
    h016_models: list[tuple[str, dict[str, Any]]],
) -> str:
    competitor_data = competitor["counts"][scene]
    lines = [
        f"### {scene}", "",
        "| Symbol | 模型 | 堆疊 | R1 | R2 | R3 | R4 | R5 |",
        "|---|---|---:|---:|---:|---:|---:|---:|",
    ]
    for symbol in SYMBOLS:
        for length in (2, 3, 4):
            competitor_rates = [
                competitor_data["stack"][reel][(symbol, length)]
                / max(1, competitor_data["stack_total"][reel])
                for reel in range(5)
            ]
            lines.append("| " + " | ".join([
                symbol, "Super Ace", str(length), *[pct(value) for value in competitor_rates],
            ]) + " |")
            for label, h016 in h016_models:
                h016_rates = [
                    h016["counts"][reel][(symbol, length)] / max(1.0, h016["totals"][reel])
                    for reel in range(5)
                ]
                lines.append("| " + " | ".join([
                    symbol, label, str(length), *[pct(value) for value in h016_rates],
                ]) + " |")
    return "\n".join(lines)


def stack_models(config: dict[str, Any], scene: str) -> list[tuple[str, dict[str, Any]]]:
    group = "base" if scene == "BG" else "free"
    return [("H016 綜合", mixed_rng_stack_data(config, group))]


def stack_comparison_section(config: dict[str, Any], competitor: dict[str, Any]) -> str:
    return f"""## 不同符號 2／3／4 堆疊 RNG 比例

每個初始停輪 RNG 分別檢查 R1～R5 的 4 格可見符號，由上到下計算「同一符號的最大連續段」。例如 `A／A／A／K` 只記 1 次 A 的 3 堆疊，不另外重複記 2 堆疊；`A／A／K／K` 則分別記 1 次 A 與 K 的 2 堆疊。金框符號先併回原符號，例如 `A／GA／A／K` 記為 A 的 3 堆疊。

各欄比例公式為 `該 Scene、該 Reel、該符號、該堆疊長度的出現次數 ÷ 該 Reel 總 RNG 數`。Super Ace 使用實際初始盤面，每個 Spin 在每輪各提供 1 筆 RNG。H016 不再逐張顯示 primary table；BG 依 `table_selection.base`、FG 依 `table_selection.free` 綜合全部啟用表，各表先正規化 stop RNG，再按 Table Selection 權重混合，最後只以一列「H016 綜合」和 Super Ace 比較。畫面只有 4 格，所以只列 2／3／4 堆疊；消除後補牌不是輪帶停輪 RNG，不納入本表。

{stack_section('BG', competitor, stack_models(config, 'BG'))}

{stack_section('FG', competitor, stack_models(config, 'FG'))}"""


def model_gold_ratios(config: dict[str, Any], sheet_or_table: str) -> dict[str, list[float]]:
    table = config["tables"][table_key(sheet_or_table)]
    initial = []
    for values, weights in zip(table["reels"], table["weights"]):
        positive_total = sum(max(0.0, float(weight)) for weight in weights)
        gold_exposure = 0.0
        if values:
            for stop, weight in enumerate(weights):
                weight = max(0.0, float(weight))
                gold_exposure += weight * sum(
                    11 <= int(values[(stop + row) % len(values)]) <= 18 for row in range(4)
                )
        initial.append(gold_exposure / max(1.0, positive_total * 4))
    drop = []
    for values, weights in zip(table["drop_values"], table["drop_weights"]):
        total = sum(max(0.0, float(weight)) for weight in weights)
        gold = sum(
            max(0.0, float(weight))
            for value, weight in zip(values, weights)
            if 11 <= int(value) <= 18
        )
        drop.append(gold / max(1.0, total))
    return {"initial": initial, "drop": drop}


def model_gold_symbol_ratios(config: dict[str, Any], sheet_or_table: str) -> dict[str, dict[str, list[float]]]:
    table = config["tables"][table_key(sheet_or_table)]
    initial = {symbol: [0.0] * 5 for symbol in SYMBOLS if symbol != "C1"}
    for reel, (values, weights) in enumerate(zip(table["reels"], table["weights"])):
        positive_total = sum(max(0.0, float(weight)) for weight in weights)
        if not values:
            continue
        for stop, weight in enumerate(weights):
            weight = max(0.0, float(weight))
            for row in range(4):
                numeric = int(values[(stop + row) % len(values)])
                if 11 <= numeric <= 18:
                    initial[symbol_name(config, numeric, merge_gold=True)][reel] += (
                        weight / max(1.0, positive_total * 4)
                    )
    drop = {symbol: [0.0] * 5 for symbol in initial}
    for reel, (values, weights) in enumerate(zip(table["drop_values"], table["drop_weights"])):
        total = max(1.0, sum(max(0.0, float(weight)) for weight in weights))
        for value, weight in zip(values, weights):
            numeric = int(value)
            if 11 <= numeric <= 18:
                drop[symbol_name(config, numeric, merge_gold=True)][reel] += max(0.0, float(weight)) / total
    return {"initial": initial, "drop": drop}


def mixed_gold_ratios(config: dict[str, Any], group: str) -> dict[str, list[float]]:
    items = selection_items(config, group)
    total = sum(weight for _, weight in items)
    result = {"initial": [0.0] * 5, "drop": [0.0] * 5}
    for table_name, weight in items:
        model = model_gold_ratios(config, table_name)
        probability = weight / total
        for stage in ("initial", "drop"):
            for reel in range(5):
                result[stage][reel] += probability * model[stage][reel]
    return result


def mixed_gold_symbol_ratios(
    config: dict[str, Any], group: str
) -> dict[str, dict[str, list[float]]]:
    items = selection_items(config, group)
    total = sum(weight for _, weight in items)
    result = {
        stage: {symbol: [0.0] * 5 for symbol in SYMBOLS if symbol != "C1"}
        for stage in ("initial", "drop")
    }
    for table_name, weight in items:
        model = model_gold_symbol_ratios(config, table_name)
        probability = weight / total
        for stage in ("initial", "drop"):
            for symbol in result[stage]:
                for reel in range(5):
                    result[stage][symbol][reel] += probability * model[stage][symbol][reel]
    return result


def distribution_section(title: str, competitor_rows: dict[str, list[float]], h016_rows: dict[str, list[float]]) -> str:
    lines = [f"### {title}", "", "| Symbol | 模型 | R1 | R2 | R3 | R4 | R5 |", "|---|---|---:|---:|---:|---:|---:|"]
    for symbol in SYMBOLS:
        lines.append("| " + " | ".join([symbol, "Super Ace", *[pct(value) for value in competitor_rows[symbol]]]) + " |")
        lines.append("| " + " | ".join([symbol, "H016", *[pct(value) for value in h016_rows[symbol]]]) + " |")
    return "\n".join(lines)


def symbol_length_section(
    competitor: dict[str, Any], h016: dict[str, Any], scene: str
) -> str:
    competitor_scene = competitor["counts"][scene]
    competitor_spins = competitor_scene["spins"]
    h016_spins = int(h016["summary"]["total_rounds"]) if scene == "BG" else h016["fg_spins"]
    coin_in = float(h016["summary"]["coin_in"]) * int(h016["summary"]["total_rounds"])
    lines = [
        f"### {scene}", "",
        "| Symbol | 輪數 | Super Ace RTP | H016 RTP | RTP 相對差異 (A-B)/A | Super Ace Hit Rate | H016 Hit Rate | Hit Rate 相對差異 (A-B)/A |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for symbol in (item for item in SYMBOLS if item != "C1"):
        for length in (3, 4, 5):
            key = (symbol, length)
            competitor_rtp = competitor_scene["symbol_length_win"][key] / competitor["coin_in"]
            h016_rtp = h016["symbol_length"][scene]["pay"][key] / coin_in
            competitor_hit = competitor_scene["symbol_length_hits"][key] / competitor_spins
            h016_hit = h016["symbol_length"][scene]["hits"][key] / max(1, h016_spins)
            lines.append(
                f"| {symbol} | {length} | {pct(competitor_rtp)} | {pct(h016_rtp)} | "
                f"{relative_difference(competitor_rtp, h016_rtp)} | {pct(competitor_hit)} | "
                f"{pct(h016_hit)} | {relative_difference(competitor_hit, h016_hit)} |"
            )
    return "\n".join(lines)


def gold_table(
    competitor: dict[str, Any], model: dict[str, list[float]], observed: dict[str, Any],
    scene: str, stage: str, setting_label: str,
) -> str:
    data = competitor["counts"][scene]
    lines = [
        f"| 階段 | Reel | Super Ace | {setting_label} | H016 10 萬場 | 設定差異 | 實測差異 |",
        "|---|---|---:|---:|---:|---:|---:|",
    ]
    for reel in range(5):
        comp = data[f"gold_{stage}"][reel] / max(1, data[f"{stage}_total"][reel])
        setting = model[stage][reel]
        actual = observed["gold_ratios"][f"{scene}_{stage}"][reel]
        lines.append(
            f"| {'初始' if stage == 'initial' else '掉落'} | R{reel + 1} | {pct(comp)} | {pct(setting)} | "
            f"{pct(actual)} | {pp(setting - comp)} | {pp(actual - comp)} |"
        )
    return "\n".join(lines)


def gold_symbol_table(
    competitor: dict[str, Any], model: dict[str, dict[str, list[float]]], scene: str,
    stage: str, setting_label: str,
) -> str:
    data = competitor["counts"][scene]
    title = "初始" if stage == "initial" else "掉落"
    lines = [
        f"#### {scene} {title}", "",
        "| Symbol | 模型 | R1 | R2 | R3 | R4 | R5 |",
        "|---|---|---:|---:|---:|---:|---:|",
    ]
    for symbol in (item for item in SYMBOLS if item != "C1"):
        comp = [
            data[f"gold_symbol_{stage}"][reel][symbol] / max(1, data[f"{stage}_total"][reel])
            for reel in range(5)
        ]
        lines.append("| " + " | ".join([symbol, "Super Ace", *[pct(value) for value in comp]]) + " |")
        lines.append("| " + " | ".join([symbol, setting_label, *[pct(value) for value in model[stage][symbol]]]) + " |")
    return "\n".join(lines)


def gold_symbol_error_summary(
    competitor: dict[str, Any], model: dict[str, dict[str, list[float]]], scene: str,
) -> str:
    data = competitor["counts"][scene]
    rows = [
        "| Scene | 階段 | 最大絕對差 | 平均絕對差 |",
        "|---|---|---:|---:|",
    ]
    for stage, label in (("initial", "初始停輪"), ("drop", "消除補牌")):
        differences: list[float] = []
        for symbol in (item for item in SYMBOLS if item != "C1"):
            for reel in range(5):
                reference = (
                    data[f"gold_symbol_{stage}"][reel][symbol]
                    / max(1, data[f"{stage}_total"][reel])
                )
                differences.append(abs(model[stage][symbol][reel] - reference))
        rows.append(
            f"| {scene} | {label} | {max(differences, default=0.0) * 100:.4f} pp | "
            f"{(sum(differences) / max(1, len(differences))) * 100:.4f} pp |"
        )
    return "\n".join(rows)


def conditional_random_wild_ratio(weights: list[int], index: int) -> float:
    total = sum(max(0, int(weight)) for weight in weights[1:])
    return max(0, int(weights[index])) / total if total else 0.0


def random_wild_config_section(config: dict[str, Any]) -> str:
    rows = [
        "| Scene | config table | Table Selection | Random Wild 0/2/3/4 | 非 0 抽中率 | 2 條件占比 | 3 條件占比 | 4 條件占比 |",
        "|---|---|---:|---|---:|---:|---:|---:|",
    ]
    for scene in ("BG", "FG"):
        group = "base" if scene == "BG" else "free"
        for table_name in PRIMARY_BY_SCENE[scene]:
            weights = model_random_wild_weights(config, table_name)
            total = sum(max(0, int(weight)) for weight in weights)
            nonzero = sum(max(0, int(weight)) for weight in weights[1:])
            rows.append(
                f"| {scene} | `{table_name}` | {format_weight(selection_weight(config, group, table_name))} | "
                f"`{'/'.join(map(str, weights))}` | {pct(nonzero / total if total else 0.0)} | "
                f"{pct(conditional_random_wild_ratio(weights, 1))} | "
                f"{pct(conditional_random_wild_ratio(weights, 2))} | "
                f"{pct(conditional_random_wild_ratio(weights, 3))} |"
            )
    return "\n".join(rows)


def selection_weight(config: dict[str, Any], group: str, table_name: str) -> float:
    for item in config.get("table_selection", {}).get(group, []):
        if str(item.get("table")) == table_name:
            return float(item.get("weight", 0))
    return 0.0


def format_weight(value: float) -> str:
    return f"{int(value):,}" if float(value).is_integer() else f"{value:,.6g}"


def stop_weight_ranges(table: dict[str, Any]) -> str:
    result = []
    for reel, weights in enumerate(table["weights"], start=1):
        positive = [float(weight) for weight in weights if float(weight) > 0]
        if not positive:
            result.append(f"R{reel} 無正權重")
            continue
        minimum, maximum = min(positive), max(positive)
        result.append(f"R{reel} {format_weight(minimum)}–{format_weight(maximum)} ({maximum / minimum:.2f}x)")
    return "；".join(result)


def _weight_leaves(value: Any, path: str) -> list[tuple[str, Any]]:
    if isinstance(value, dict):
        result = []
        for key, item in value.items():
            result.extend(_weight_leaves(item, f"{path}.{key}"))
        return result
    if isinstance(value, list):
        result = []
        for index, item in enumerate(value):
            result.extend(_weight_leaves(item, f"{path}[{index}]"))
        return result
    return [(path, value)]


def config_weight_values(config: dict[str, Any]) -> list[tuple[str, Any]]:
    result: list[tuple[str, Any]] = []

    def visit(value: Any, path: str) -> None:
        if isinstance(value, dict):
            for key, item in value.items():
                child = f"{path}.{key}" if path else str(key)
                if "weight" in str(key).lower():
                    result.extend(_weight_leaves(item, child))
                else:
                    visit(item, child)
        elif isinstance(value, list):
            for index, item in enumerate(value):
                visit(item, f"{path}[{index}]")

    visit(config, "")
    return result


def is_nonnegative_integer(value: Any) -> bool:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return False
    numeric = float(value)
    return math.isfinite(numeric) and numeric >= 0 and numeric.is_integer()


def validation_status(ok: bool) -> str:
    return "PASS" if ok else "FAIL"


def config_constraint_section(
    config: dict[str, Any], competitor: dict[str, Any], summary: dict[str, Any]
) -> str:
    rows = [
        "| 檢查 | 範圍 | 結果 | 詳細 |",
        "|---|---|---:|---|",
    ]
    for scene in ("BG", "FG"):
        for table_name in PRIMARY_BY_SCENE[scene]:
            lengths = [len(reel) for reel in config["tables"][table_name]["reels"]]
            ok = len(lengths) == 5 and all(length == 200 for length in lengths)
            rows.append(
                f"| Primary reel 長度 = 200 | `{table_name}` | {validation_status(ok)} | "
                f"{'/'.join(map(str, lengths))} |"
            )

    weight_values = config_weight_values(config)
    invalid = [(path, value) for path, value in weight_values if not is_nonnegative_integer(value)]
    invalid_detail = "；".join(f"`{path}`={value!r}" for path, value in invalid[:5])
    if len(invalid) > 5:
        invalid_detail += f"；另 {len(invalid) - 5} 筆"
    rows.append(
        f"| 所有 config 權重為非負整數 | 全 config | {validation_status(not invalid)} | "
        f"共 {len(weight_values):,} 筆" + (f"；{invalid_detail}" if invalid_detail else "") + " |"
    )

    for scene in ("BG", "FG"):
        for table_name in PRIMARY_BY_SCENE[scene]:
            table = config["tables"][table_name]
            ratios = []
            valid = len(table.get("weights", [])) == 5
            for reel in range(5):
                weights = table.get("weights", [])[reel] if reel < len(table.get("weights", [])) else []
                reel_values = table.get("reels", [])[reel] if reel < len(table.get("reels", [])) else []
                positive = [float(weight) for weight in weights if float(weight) > 0]
                shape_ok = len(weights) == len(reel_values)
                ratio = max(positive) / min(positive) if positive else math.inf
                valid &= shape_ok and bool(positive) and ratio <= 10.0
                ratios.append(f"R{reel + 1} {'無正權重' if not positive else f'{ratio:.2f}x'}")
            rows.append(
                f"| 正 stop max/min ≤ 10x（排除 0） | `{table_name}` | "
                f"{validation_status(valid)} | {'；'.join(ratios)} |"
            )

    for label, competitor_key, model_key in (
        ("Total RTP", "rtp_total", "rtp_total"),
        ("BG RTP", "rtp_bg", "rtp_bg"),
        ("FG RTP", "rtp_fg", "rtp_fg"),
    ):
        target = float(competitor[competitor_key]) * 2
        actual = float(summary[model_key])
        rows.append(
            f"| RTP ≤ Super Ace 2x | {label} | {validation_status(actual <= target)} | "
            f"H016 {pct(actual)}；上限 {pct(target)} |"
        )
    return "\n".join(rows)


def config_table_section(config: dict[str, Any]) -> str:
    rows = [
        "| Scene | config table | 初始 Table Weight | Retrigger Weight | Reel 長度 | Cascade 倍數 | 正停輪權重 min–max | Random Wild 0/2/3/4 |",
        "|---|---|---:|---:|---|---|---|---|",
    ]
    for scene, table_name in (
        ("BG", "bg_1"), ("BG", "bg_2"), ("BG", "bg_3"),
        ("FG", "fg_1"), ("FG", "fg_2"), ("FG", "fg_3"),
    ):
        table = config["tables"][table_name]
        group = "base" if scene == "BG" else "free"
        initial_weight = selection_weight(config, group, table_name)
        retrigger = "—" if scene == "BG" else format_weight(selection_weight(config, "retrigger", table_name))
        lengths = "/".join(str(len(reel)) for reel in table["reels"])
        random_wild = "/".join(map(str, model_random_wild_weights(config, table_name)))
        multipliers = "/".join(f"x{int(value)}" for value in table["multipliers"])
        rows.append(
            f"| {scene} | `{table_name}` | {format_weight(initial_weight)} | {retrigger} | {lengths} | `{multipliers}` | "
            f"{stop_weight_ranges(table)} | `{random_wild}` |"
        )
    return "\n".join(rows)


def load_source_applier():
    module_name = "h016_super_ace_claude_source_for_report"
    spec = importlib.util.spec_from_file_location(module_name, SOURCE_APPLIER)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load {SOURCE_APPLIER}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[module_name] = module
    try:
        spec.loader.exec_module(module)
        return module
    finally:
        sys.modules.pop(module_name, None)


def source_application_section(config: dict[str, Any]) -> str:
    applier = load_source_applier()
    parsed = applier.parse_source(SOURCE_REELS)
    source_summary = applier.summary(config, parsed)
    rows = [
        "| Scene | 主表底層輪帶 exact match | 主表 stopW ×10,000 exact match | Alias 對應正確 | 輪帶長度 | 倍數 |",
        "|---|---:|---:|---:|---|---|",
    ]
    for scene, master in (("BG", "bg_1"), ("FG", "fg_1")):
        table = config["tables"][master]
        canonical_reels = [
            [symbol - 8 if 11 <= symbol <= 18 else symbol for symbol in reel]
            for reel in table["reels"]
        ]
        reels_match = canonical_reels == parsed[scene]["reels"]
        weights_match = table["weights"] == parsed[scene]["weights"]
        aliases_match = all(
            config["tables"].get(alias) == config["tables"].get(primary)
            for alias, primary in ALIAS_TO_PRIMARY.items()
            if alias in ALL_TABLES_BY_SCENE[scene]
        )
        lengths = "/".join(str(len(reel)) for reel in table["reels"])
        multipliers = "/".join(f"x{int(value)}" for value in table["multipliers"])
        rows.append(
            f"| {scene} | {validation_status(reels_match)} | {validation_status(weights_match)} | "
            f"{validation_status(aliases_match)} | {lengths} | `{multipliers}` |"
        )

    bg_c1 = [parsed["BG"]["reels"][reel].count(2) for reel in range(5)]
    fg_c1 = [parsed["FG"]["reels"][reel].count(2) for reel in range(5)]

    def cyclic_max_run(reel: list[int]) -> int:
        if not reel:
            return 0
        best = run = 1
        doubled = reel + reel
        for index in range(1, len(doubled)):
            run = run + 1 if doubled[index] == doubled[index - 1] else 1
            best = max(best, min(run, len(reel)))
        return best

    bg_runs = [cyclic_max_run(reel) for reel in parsed["BG"]["reels"]]
    fg_runs = [cyclic_max_run(reel) for reel in parsed["FG"]["reels"]]
    return "\n".join([
        f"來源：`{SOURCE_REELS.name}`；SHA-256 `{source_summary['source_sha256']}`；原始 stopW 為四位小數，統一乘以 10,000 無損轉成整數。",
        "",
        *rows,
        "",
        "> 底層輪帶 exact match 的判定會先把金框 ID 合併回原符號；PASS 表示只覆蓋競品金框、不改動文字檔的底層符號排列與數量。",
        "",
        f"來源輪帶的 C1 格數為 BG R1～R5 = `{'/'.join(map(str, bg_c1))}`、FG R1～R5 = `{'/'.join(map(str, fg_c1))}`。",
        "",
        "> 表別角色：BG_Symbol 與 BG_Symbol (2) 以停輪權重及 C1 掉落權重硬性阻止 FG；BG_Symbol (3) 專門提高 FG 觸發。FG_Symbol 與 FG_Symbol (2) 的大鬼只允許 2 顆，FG_Symbol (3) 關閉。各 alias 必須與其對應 primary table 完全一致。",
        "",
        f"> 輪帶連段警告：來源輪帶各輪最大循環連段為 BG `{'/'.join(map(str, bg_runs))}`、FG `{'/'.join(map(str, fg_runs))}`；最高分別為 {max(bg_runs)} 與 {max(fg_runs)}，不符合先前提出的實體輪帶最多 4 顆同符號連段限制。若拆段，就不再是文字檔的 exact source。",
        "",
        "> 範圍限制：本次依需求只套用文字檔的輪帶與停輪權重；文字檔另述的「BG 恰有 2 個 SC 時，37.1% 抑制其中 1 個」動態層未加入，因其屬於額外玩法規則。故 FG Trigger Rate 是目前 H016 規則下的比較值，不代表已還原該抑制層。",
    ])


def pay_table_lines(config: dict[str, Any]) -> list[str]:
    lines = [
        "| Symbol | Super Ace 3/4/5 | H016 3/4/5 | 比較結果 |",
        "|---|---|---|---|",
    ]
    for symbol_id in range(3, 11):
        name = symbol_name(config, symbol_id)
        pays = tuple(float(value) for value in config["pays"][str(symbol_id)])
        reference = SUPER_ACE_PAYS[name]
        same = len(pays) == len(reference) and all(
            math.isclose(actual, expected, rel_tol=0.0, abs_tol=1e-12)
            for actual, expected in zip(pays, reference)
        )
        lines.append(
            f"| {name} | `{'/'.join(f'{value:g}' for value in reference)}` | "
            f"`{'/'.join(f'{value:g}' for value in pays)}` | {validation_status(same)} |"
        )
    return lines


def main() -> None:
    config = load_model_config()
    h016, h016_source, record_name = h016_data(config)
    hs = h016["summary"]
    competitor = raw_competitor()
    bg_model_gold = mixed_gold_ratios(config, "base")
    fg_model_gold = mixed_gold_ratios(config, "free")
    bg_model_gold_symbols = mixed_gold_symbol_ratios(config, "base")
    fg_model_gold_symbols = mixed_gold_symbol_ratios(config, "free")
    random_wild_settings = random_wild_config_section(config)
    competitor_w2 = [int(competitor["w2_bg_counts"][count]) for count in (2, 3, 4)]
    competitor_w2_total = sum(competitor_w2)
    bg_w2_total = sum(int(hs[f"w2_bg_count_{count}"]) for count in (2, 3, 4))
    fg_w2_total = sum(int(hs[f"w2_fg_count_{count}"]) for count in (2, 3, 4))
    w2_distribution_lines = [
        "| 隨機 WW 顆數 | Super Ace 條件分布 | H016 BG 10 萬場 | H016 FG 10 萬場 |",
        "|---:|---:|---:|---:|",
    ]
    for count in (2, 3, 4):
        comp_ratio = int(competitor["w2_bg_counts"][count]) / max(1, competitor_w2_total)
        bg_actual = int(hs[f"w2_bg_count_{count}"]) / max(1, bg_w2_total)
        fg_actual = int(hs[f"w2_fg_count_{count}"]) / max(1, fg_w2_total)
        w2_distribution_lines.append(
            f"| {count} | {pct(comp_ratio)} | {pct(bg_actual)} | {pct(fg_actual)} |"
        )
    core = [
        ("RTP", "Total RTP", competitor["rtp_total"], float(hs["rtp_total"])),
        ("RTP", "BG RTP", competitor["rtp_bg"], float(hs["rtp_bg"])),
        ("RTP", "FG RTP", competitor["rtp_fg"], float(hs["rtp_fg"])),
        ("Hit Rate", "BG Hit Rate", competitor["bg_hit_rate"], float(hs["bg_hit_rate"])),
        ("Hit Rate", "FG Hit Rate", competitor["fg_hit_rate"], float(hs["fg_hit_rate"])),
        ("FG", "FG Trigger Rate", competitor["fg_trigger_rate"], float(hs["fg_trigger_rate"])),
    ]
    core_lines = ["| 類別 | 指標 | Super Ace | H016 | 差異 |", "|---|---|---:|---:|---:|"]
    for category, metric, comp, model in core:
        core_lines.append(f"| {category} | {metric} | {pct(comp)} | {pct(model)} | {pp(model - comp)} |")
    core_lines.extend([
        f"| FG | 平均觸發週期 | {1 / competitor['fg_trigger_rate']:.2f} 局／次 | {float(hs['fg_trigger_cycle']):.2f} 局／次 | H016 {'慢' if float(hs['fg_trigger_cycle']) > 1 / competitor['fg_trigger_rate'] else '快'} {abs(float(hs['fg_trigger_cycle']) - 1 / competitor['fg_trigger_rate']):.2f} 局 |",
        f"| FG | 平均 Free Spins | {competitor['avg_fg_spins']:.4f} | {float(hs['avg_fg_spins']):.4f} | {float(hs['avg_fg_spins']) - competitor['avg_fg_spins']:+.4f} |",
    ])

    pay_lines = pay_table_lines(config)
    config_tables = config_table_section(config)
    constraint_table = config_constraint_section(config, competitor, hs)
    source_application = source_application_section(config)

    combo_sections = []
    for scene in ("BG", "FG"):
        h_den = int(hs["total_rounds"]) if scene == "BG" else sum(row[scene] for row in h016["combo"].values())
        c_den = competitor["rounds"] if scene == "BG" else competitor["fg_spins"]
        lines = [f"### {scene}", "", f"| 消除次數 | Super Ace | H016 | 差異 |", "|---|---:|---:|---:|"]
        for index, label in enumerate(("0", "1", "2", "3", "4", "5+")):
            comp = competitor["counts"][scene]["combo"][index] / c_den
            model = h016["combo"][label][scene] / max(1, h_den)
            lines.append(f"| {label} | {pct(comp)} | {pct(model)} | {pp(model - comp)} |")
        combo_sections.append("\n".join(lines))

    document = f"""# 競品參考數值比較

## 目錄

- [Overview](#overview)
  - [Super Ace_claude 套用驗證](#super-ace_claude-套用驗證)
  - [Config-only 約束驗證](#config-only-約束驗證)
  - [核心指標比較](#核心指標比較)
- [賠率](#賠率)
- [符號 345 連線 RTP / Hit Rate 占比](#符號-345-連線-rtp--hit-rate-占比)
  - [BG](#bg)
  - [FG](#fg)
- [不同符號 2／3／4 堆疊 RNG 比例](#不同符號-234-堆疊-rng-比例)
- [符號分布](#符號分布)
  - [BG 初始 R1-R5](#bg-初始-r1-r5)
  - [BG 掉落 R1-R5](#bg-掉落-r1-r5)
  - [FG 初始 R1-R5](#fg-初始-r1-r5)
  - [FG 掉落 R1-R5](#fg-掉落-r1-r5)
- [消除率](#消除率)
- [金框比例](#金框比例)
- [大鬼事件率](#大鬼事件率)

## Overview

### 比較基準

| 項目 | Super Ace | H016 現在版本 |
|---|---|---|
| 來源 | JILI 實機非重複 JSONL | {h016_source} |
| 樣本 | {competitor['rounds']:,} 個 BG Round、{competitor['fg_spins']:,} 個 FG Spin | {int(hs['total_rounds']):,} 個 BG Round、{h016['fg_spins']:,} 個 FG Spin |
| Card System | 競品實際遊玩資料 | 關閉 |
| Bet Mode | 一般投注 | Normal Bet，Bet Multi 1 |
| 數值設定 | 實機 JSONL 盤面與消除紀錄 | `config.js`；本報告不讀取 `Source/H0161.xlsx` |

H016 的輪帶排列、停輪權重、補牌權重、Random Wild 與 Table Selection 全部直接讀取 `config.js`。本產生器固定以同一份 config 在記憶體中模擬 100,000 場，不讀取歷史 Record，也不建立或修改任何 xlsx。

### Super Ace_claude 套用驗證

{source_application}

### 多表用途與權重

{config_tables}

最新 config-only 10 萬場 FG 觸發率為 {pct(float(hs['fg_trigger_rate']))}、週期 {float(hs['fg_trigger_cycle']):.2f} 局；競品為 {pct(competitor['fg_trigger_rate'])}、{1 / competitor['fg_trigger_rate']:.2f} 局。

### Config-only 約束驗證

權重整數檢查涵蓋 `config.js` 內所有名稱含 `weight` 的設定。Stop max/min 只比較正權重；`0` 代表停用 stop，因此明確排除。RTP 以本次 10 萬場實測和 Super Ace 各 Scene RTP 的 2 倍上限比較。

{constraint_table}

### 核心指標比較

{chr(10).join(core_lines)}

以上差異直接取自 `config.js` 的本次 10 萬場結果；調整 config 後必須重新執行本產生器，才可把新結果與 Super Ace JSONL 使用相同口徑比較。

## 賠率

表內數字為相對投注額倍數；比較結果由 `config.js` 與 Super Ace 參考 pay table 動態計算，不預設一定相同。

{chr(10).join(pay_lines)}

## 符號 345 連線 RTP / Hit Rate 占比

RTP 分母為該次模擬或競品樣本的總投注額；Hit Rate 分母為該 Scene 的 Spin 數。每次 Cascade 中，同一符號同一輪數的一筆 award 計為一次 Hit；同一 Spin 可同時命中多符號或多次 Cascade，所以各列 Hit Rate 不應加總為整體 Spin Hit Rate。競品 RTP 已將 `award.win` 乘上該次 `comboBonus`；H016 RTP 使用模擬器實際派彩，含 BG／FG Cascade 倍率。相對差異統一以 `A = Super Ace`、`B = H016`，公式為 `(A-B)／A`；負值代表 H016 高於 Super Ace，正值代表 H016 低於 Super Ace，A 為 0 時顯示 `N/A`。表內的「所有符號」指具有 3／4／5 輪賠率的 `M1～M4、A、K、Q、J`；`C1`、`WW`、`W2` 不屬於獨立付費符號，因此不列入。

{symbol_length_section(competitor, h016, 'BG')}

{symbol_length_section(competitor, h016, 'FG')}

{stack_comparison_section(config, competitor)}

## 符號分布

### 套用方法

1. 初始輪帶直接讀取 `config.js` 中本次所選 `tables.*.reels`，停輪權重讀取同表的 `weights`。
2. 消除後補牌直接讀取本次所選 table 的 `drop_values`／`drop_weights`。
3. `table_selection.base/free/retrigger` 控制每次 BG、FG 與 Retrigger 實際選用哪張 config table。
4. 初始盤面依輪帶連續取 4 格；Cascade 在被消除的原位置依 Symbol Drop Weight 獨立補牌，不做重力掉落。

{distribution_section('BG 初始 R1-R5', competitor_ratios(competitor, 'BG', 'initial'), h016['ratios']['BG_initial'])}

{distribution_section('BG 掉落 R1-R5', competitor_ratios(competitor, 'BG', 'drop'), h016['ratios']['BG_drop'])}

{distribution_section('FG 初始 R1-R5', competitor_ratios(competitor, 'FG', 'initial'), h016['ratios']['FG_initial'])}

{distribution_section('FG 掉落 R1-R5', competitor_ratios(competitor, 'FG', 'drop'), h016['ratios']['FG_drop'])}

以上 H016 初始與掉落符號分布皆為 `config.js` 最新 10 萬場模擬實測；掉落分母為各輪實際補牌顆數。

## 消除率

比較口徑為每個 Spin 實際發生的消除次數，統一合併為 `0、1、2、3、4、5+`；FG 分母為實際 Free Spins。

{combo_sections[0]}

{combo_sections[1]}

## 金框比例

競品金框是盤面生成後的獨立疊加層；H016 設定值直接由 `config.js` 的金框 symbol ID（11～18）、停輪權重、補牌權重及 Table Selection 計算。BG 設定欄是 `table_selection.base` 混合；FG 設定欄是 `table_selection.free` 混合。初始設定值已依每個 stop 的 4 格可見視窗計算，不用單一 bg1／fg1 代表多表；10 萬場欄位是同一份 config 的實測。

### BG

{gold_table(competitor, bg_model_gold, h016, 'BG', 'initial', 'H016 base 混合設定')}

{gold_table(competitor, bg_model_gold, h016, 'BG', 'drop', 'H016 base 混合設定')}

### FG

{gold_table(competitor, fg_model_gold, h016, 'FG', 'initial', 'H016 free 混合設定')}

{gold_table(competitor, fg_model_gold, h016, 'FG', 'drop', 'H016 free 混合設定')}

### 各符號金框分布

比例口徑為「該金框符號的理論可見量或掉落權重 ÷ 該輪全部 Symbol 可見量或掉落權重」，並依 Scene 的 Table Selection 混合。

初始停輪受限於每輪固定 200 格與固定停輪權重，採可達組合中最接近競品者；消除補牌使用整數權重直接校準。以下誤差涵蓋 8 種符號 × R1～R5，R1／R5 的競品 0% 也納入驗證。

{gold_symbol_error_summary(competitor, bg_model_gold_symbols, 'BG')}

{gold_symbol_error_summary(competitor, fg_model_gold_symbols, 'FG')}

{gold_symbol_table(competitor, bg_model_gold_symbols, 'BG', 'initial', 'H016 base 混合設定')}

{gold_symbol_table(competitor, bg_model_gold_symbols, 'BG', 'drop', 'H016 base 混合設定')}

{gold_symbol_table(competitor, fg_model_gold_symbols, 'FG', 'initial', 'H016 free 混合設定')}

{gold_symbol_table(competitor, fg_model_gold_symbols, 'FG', 'drop', 'H016 free 混合設定')}

## 大鬼事件率

事件率口徑為「成功啟動大鬼的事件數 ÷ 該 Scene Spins」。BG 同一個 Spin 最多成功啟動一次 WW2；FG 每次 Cascade 有新金框轉換都會重新判定，因此同一個 FG Spin 可成功啟動多次。抽到 0 時該次金框只轉 WW1，後續金框仍會再判定。

| Scene | Super Ace | H016 10 萬場實測 | 差異 |
|---|---:|---:|---:|
| BG | {pct(competitor['w2_bg_event_rate'])} | {pct(float(hs['w2_bg_event_rate']))} | {pp(float(hs['w2_bg_event_rate']) - competitor['w2_bg_event_rate'])} |
| FG | 0.0000% | {pct(float(hs['w2_fg_event_rate']))} | {pp(float(hs['w2_fg_event_rate']))} |

以下逐表列出全部 primary table 的設定與 Table Selection。各表依角色保留獨立停輪、補牌與大鬼權重；`非 0 抽中率` 的分母含 0 權重，`2／3／4 條件占比` 的分母只含非 0 權重。後一張表的 H016 欄是實際 10 萬場混合結果。

{random_wild_settings}

{chr(10).join(w2_distribution_lines)}
"""
    OUTPUT.write_text(document, encoding="utf-8")
    print(json.dumps({
        "output": str(OUTPUT),
        "config": str(MODEL_CONFIG),
        "record": record_name,
        "simulated_in_memory": record_name is None,
        "competitor_bg": competitor["rounds"],
        "competitor_fg": competitor["fg_spins"],
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
