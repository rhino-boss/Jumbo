from __future__ import annotations

import argparse
import json
import math
import os
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[2]
DEFAULT_CONFIG = PROJECT_DIR / "config.js"
DEFAULT_XLSX = PROJECT_DIR / "Source" / "H0161.xlsx"
CONFIG_PATTERN = re.compile(r"window\.H016_CONFIG\s*=\s*(\{.*\})\s*;", re.DOTALL)
SCORE_SYMBOLS = tuple(range(3, 11))
WILDS = {0, 1}
SCATTER = 2


def load_config(path: Path) -> dict[str, Any]:
    match = CONFIG_PATTERN.search(path.read_text(encoding="utf-8-sig"))
    if match is None:
        raise ValueError(f"Cannot find window.H016_CONFIG in {path}")
    return json.loads(match.group(1))


def canonical(symbol: int) -> int:
    return symbol - 8 if 11 <= symbol <= 18 else symbol


def visible_window(reel: list[int], stop: int) -> tuple[int, ...]:
    return tuple(int(reel[(stop + offset) % len(reel)]) for offset in range(4))


def score_set(window: tuple[int, ...]) -> frozenset[int]:
    return frozenset(
        canonical(symbol)
        for symbol in window
        if canonical(symbol) in SCORE_SYMBOLS
    )


def scatter_count(window: tuple[int, ...]) -> int:
    return sum(symbol == SCATTER for symbol in window)


def candidates(reel: list[int], minimum_scatter: int, allowed: frozenset[int] | None) -> list[int]:
    result: list[int] = []
    for stop in range(len(reel)):
        window = visible_window(reel, stop)
        if any(symbol in WILDS for symbol in window):
            continue
        if scatter_count(window) < minimum_scatter:
            continue
        if allowed is not None and not score_set(window).issubset(allowed):
            continue
        result.append(stop)
    return result


def distinct_windows(reel: list[int], stops: list[int]) -> int:
    return len({visible_window(reel, stop) for stop in stops})


def solve(reels: list[list[int]], old_weights: list[list[int]]) -> tuple[list[list[int]], dict[str, Any]]:
    if len(reels) != 5 or any(len(reel) != 200 for reel in reels):
        raise ValueError("BF reels must be exactly 5 x 200")
    if len(old_weights) != 5 or any(len(weights) != 200 for weights in old_weights):
        raise ValueError("BF weights must be exactly 5 x 200")

    all_symbols = frozenset(SCORE_SYMBOLS)
    best: tuple[tuple[int, ...], list[list[int]], frozenset[int]] | None = None
    # R1 and R2 are assigned disjoint score-symbol groups. This alone proves
    # that no Ways award can start from the left, irrespective of R3-R5.
    for mask in range(1 << len(SCORE_SYMBOLS)):
        r1_allowed = frozenset(
            symbol for index, symbol in enumerate(SCORE_SYMBOLS) if mask & (1 << index)
        )
        r2_allowed = all_symbols.difference(r1_allowed)
        for sc1 in range(5):
            r1 = candidates(reels[0], sc1, r1_allowed)
            if len(r1) < 2 or distinct_windows(reels[0], r1) < 2:
                continue
            for sc2 in range(5):
                r2 = candidates(reels[1], sc2, r2_allowed)
                if len(r2) < 2 or distinct_windows(reels[1], r2) < 2:
                    continue
                for sc3 in range(5):
                    r3 = candidates(reels[2], sc3, None)
                    if len(r3) < 2 or distinct_windows(reels[2], r3) < 2:
                        continue
                    r4 = [index for index, weight in enumerate(old_weights[3]) if weight > 0]
                    r5 = [index for index, weight in enumerate(old_weights[4]) if weight > 0]
                    if not r4 or not r5:
                        continue
                    minimum_total_scatter = sum(
                        min(scatter_count(visible_window(reels[index], stop)) for stop in stops)
                        for index, stops in enumerate((r1, r2, r3, r4, r5))
                    )
                    if minimum_total_scatter < 3:
                        continue
                    stops = [r1, r2, r3, r4, r5]
                    diversity = tuple(distinct_windows(reels[index], reel_stops) for index, reel_stops in enumerate(stops))
                    original_overlap = sum(
                        old_weights[index][stop] > 0
                        for index, reel_stops in enumerate(stops[:3])
                        for stop in reel_stops
                    )
                    # Prefer many independently selectable stops and visible
                    # windows; overlap is only a final tie-breaker.
                    objective = (
                        min(diversity[:3]),
                        sum(diversity[:3]),
                        len(r1) + len(r2) + len(r3),
                        original_overlap,
                    )
                    if best is None or objective > best[0]:
                        best = (objective, stops, r1_allowed)

    if best is None:
        raise ValueError("No BF stop-weight solution satisfies no-win + 3 SC + visual diversity")

    _, stops, r1_allowed = best
    new_weights: list[list[int]] = []
    for reel_index, reel_stops in enumerate(stops):
        if reel_index < 3:
            enabled = set(reel_stops)
            new_weights.append([100 if stop in enabled else 0 for stop in range(200)])
        else:
            new_weights.append([int(weight) for weight in old_weights[reel_index]])

    # Exact proof checks over every enabled R1/R2 window pair.
    enabled = [[index for index, weight in enumerate(weights) if weight > 0] for weights in new_weights]
    for r1 in enabled[0]:
        w1 = visible_window(reels[0], r1)
        if any(symbol in WILDS for symbol in w1):
            raise AssertionError("Enabled BF R1 window contains Wild")
        for r2 in enabled[1]:
            w2 = visible_window(reels[1], r2)
            if any(symbol in WILDS for symbol in w2):
                raise AssertionError("Enabled BF R2 window contains Wild")
            if score_set(w1).intersection(score_set(w2)):
                raise AssertionError("Enabled BF R1/R2 pair can form a Ways award")

    minimum_scatter = [
        min(scatter_count(visible_window(reels[index], stop)) for stop in reel_stops)
        for index, reel_stops in enumerate(enabled)
    ]
    if sum(minimum_scatter) < 3:
        raise AssertionError("Enabled BF windows do not guarantee at least 3 SC")
    positive = [[weight for weight in weights if weight > 0] for weights in new_weights]
    if any(max(values) / min(values) > 10 for values in positive):
        raise AssertionError("BF positive stop-weight ratio exceeds 10x")

    report = {
        "enabled_stops": [len(stops) for stops in enabled],
        "distinct_windows": [distinct_windows(reels[index], stops) for index, stops in enumerate(enabled)],
        "minimum_scatter_by_reel": minimum_scatter,
        "minimum_total_scatter": sum(minimum_scatter),
        "possible_stop_combinations": math.prod(len(stops) for stops in enabled),
        "r1_score_symbols": sorted(r1_allowed),
        "r2_score_symbols": sorted(all_symbols.difference(r1_allowed)),
        "no_win_r1_r2_pairs_checked": len(enabled[0]) * len(enabled[1]),
        "positive_weight_ratio": [max(values) / min(values) for values in positive],
    }
    return new_weights, report


def write_xlsx(path: Path, weights: list[list[int]]) -> None:
    workbook = load_workbook(path, read_only=False, data_only=False, keep_links=True)
    try:
        if "BF_Symbol" not in workbook.sheetnames:
            raise ValueError("H0161.xlsx is missing BF_Symbol")
        sheet = workbook["BF_Symbol"]
        for reel in range(5):
            for stop, weight in enumerate(weights[reel]):
                sheet.cell(4 + stop, 23 + reel).value = int(weight)
        if workbook.calculation is not None:
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
        temporary = path.with_name(path.stem + ".bf-no-win.tmp" + path.suffix)
        try:
            workbook.save(temporary)
            check = load_workbook(temporary, read_only=True, data_only=False)
            try:
                sheet = check["BF_Symbol"]
                actual = [
                    [int(sheet.cell(4 + stop, 23 + reel).value or 0) for stop in range(200)]
                    for reel in range(5)
                ]
                if actual != weights:
                    raise ValueError("BF weight round-trip mismatch")
            finally:
                check.close()
            os.replace(temporary, path)
        finally:
            if temporary.exists():
                temporary.unlink()
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="Configure BF stops for varied, guaranteed no-win entry boards")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--write", action="store_true")
    args = parser.parse_args()

    config = load_config(args.config.resolve())
    table = config["tables"]["buy"]
    weights, report = solve(table["reels"], table["weights"])
    if args.write:
        write_xlsx(args.xlsx.resolve(), weights)
    report.update({"xlsx": str(args.xlsx.resolve()), "written": bool(args.write)})
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
