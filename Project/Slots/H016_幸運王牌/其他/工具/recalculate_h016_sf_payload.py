from __future__ import annotations

import argparse
import json
import math
import re
from copy import deepcopy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WEIGHT_TOTAL = 1_000_000_000
TARGET_SF_RTP = 0.925
SUPER_PRICE = 250.0
MIN_NATURAL_RATE = 0.0007
MIN_SF_WEIGHTED_INDEX = 24  # (100, 120]
PROFIT_START_INDEX = 30  # (250, 300]
ABOVE_500_START_INDEX = 35  # (500, 550]
BELOW_100_END_INDEX = 23  # (90, 100]
TARGET_PROFIT_HIT_RATE = 0.30
TARGET_ABOVE_500_HIT_RATE = 0.06
MIN_ABOVE_500_HIT_RATE = 0.05
MAX_ABOVE_500_HIT_RATE = 0.07
MAX_BELOW_100_HIT_RATE = 0.50
PER_RANGE_RTP_SHARE_CAP = 0.15
REFERENCE_PSEUDOCOUNT = 0.25
HEAD_DECAY_RATIO = 0.97
BOOST_BY_LABEL = {"(350, 400]": 1.5, "(450, 500]": 1.5}
TAIL_RATIO_MIN = 0.65
TAIL_RATIO_MAX = 0.90
PROJECT_DIR = Path(__file__).resolve().parents[2]


def base_info(workbook) -> dict[str, Any]:
    return {
        str(row[0]): row[1]
        for row in workbook["Base Info"].iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    }


def load_sf_report(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        fields = base_info(workbook)
        rows = list(workbook["Multiplier Line"].iter_rows(min_row=1, max_row=65, values_only=True))
        if len(rows) != 65:
            raise ValueError(f"{path.name}: expected header plus 64 multiplier buckets")
        header = {str(value): index for index, value in enumerate(rows[0]) if value is not None}
        required = {"Interval", "free_game_cnt_SF", "free_game_pay_SF"}
        missing = required.difference(header)
        if missing:
            raise ValueError(f"{path.name}: missing Multiplier Line columns {sorted(missing)}")
        data = rows[1:]
        result = {
            "path": str(path.resolve()),
            "rounds": int(fields["total_rounds"]),
            "coin_in": float(fields["coin_in"]),
            "rtp_total": float(fields["rtp_total"]),
            "rtp_fg": float(fields["rtp_fg"]),
            "bet_mode": str(fields["bet_mode"]),
            "card_system": str(fields["card_system"]),
            "intervals": [str(row[header["Interval"]]) for row in data],
            "sf_count": [int(row[header["free_game_cnt_SF"]] or 0) for row in data],
            "sf_pay": [float(row[header["free_game_pay_SF"]] or 0) for row in data],
        }
    finally:
        workbook.close()

    if result["bet_mode"] != "Buy Super Feature":
        raise ValueError(f"{path.name}: expected Buy Super Feature, got {result['bet_mode']!r}")
    if result["card_system"].lower() != "off":
        raise ValueError(f"{path.name}: SF natural report must have Card System off")
    if sum(result["sf_count"]) != result["rounds"]:
        raise ValueError(f"{path.name}: SF bucket count does not equal total rounds")
    base_bet = result["coin_in"] / SUPER_PRICE
    observed_rtp = sum(result["sf_pay"]) / result["rounds"] / result["coin_in"]
    if abs(observed_rtp - result["rtp_total"]) > 1e-10:
        raise ValueError(
            f"{path.name}: SF bucket pay RTP {observed_rtp} differs from Base Info {result['rtp_total']}"
        )
    if not math.isclose(base_bet, 100.0, rel_tol=0.0, abs_tol=1e-9):
        raise ValueError(f"Unexpected SF base bet: {base_bet}")
    result["base_bet"] = base_bet
    return result


def workbook_labels(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        labels = [str(workbook["Detail"].cell(row, 1).value or "").strip() for row in range(234, 298)]
    finally:
        workbook.close()
    if len(labels) != 64 or any(not label for label in labels):
        raise ValueError(f"{path.name}: Detail SF range labels are incomplete")
    return labels


def interval_bounds(report_labels: list[str]) -> list[tuple[float, float]]:
    bounds = [(-1.0, 0.0)]
    for label in report_labels[1:]:
        match = re.fullmatch(r"\s*([-0-9.]+)\s*<\s*X\s*<=\s*([-0-9.]+)\s*", label)
        if match is None:
            raise ValueError(f"Cannot parse SF interval {label!r}")
        bounds.append((float(match.group(1)), float(match.group(2))))
    if len(bounds) != 64:
        raise ValueError("Expected 64 SF interval bounds")
    return bounds


def load_competitor_sf(path: Path, bounds: list[tuple[float, float]]) -> dict[str, Any]:
    counts = [0] * len(bounds)
    pays = [0.0] * len(bounds)
    samples = 0
    bonus_samples = 0
    bets: set[float] = set()
    with path.open("r", encoding="utf-8") as source:
        for line_number, line in enumerate(source, 1):
            if not line.strip():
                continue
            item = json.loads(line)
            bet = float(item.get("bet", 0.0))
            if bet <= 0:
                raise ValueError(f"{path.name}:{line_number}: invalid bet {bet}")
            multiplier = float(item.get("win", 0.0)) / bet
            if multiplier < 0:
                raise ValueError(f"{path.name}:{line_number}: negative multiplier {multiplier}")
            bets.add(bet)
            samples += 1
            bonus_samples += int(bool(item.get("is_bonus")))
            if multiplier == 0:
                index = 0
            else:
                index = next(
                    (index for index, (lower, upper) in enumerate(bounds) if lower < multiplier <= upper),
                    None,
                )
                if index is None:
                    raise ValueError(f"{path.name}:{line_number}: multiplier {multiplier} is out of range")
            counts[index] += 1
            pays[index] += multiplier
    if samples <= 0:
        raise ValueError(f"{path.name}: no competitor SF samples")
    return {
        "path": str(path.resolve()),
        "samples": samples,
        "bonus_samples": bonus_samples,
        "bets": sorted(bets),
        "counts": counts,
        "pay_multipliers": pays,
        "observed_rtp_vs_h016_price": sum(pays) / samples / SUPER_PRICE,
        "observed_profit_hit_rate": sum(counts[PROFIT_START_INDEX:]) / samples,
    }


def fix_numbers(weights: list[int], counts: list[int]) -> list[float]:
    total = sum(counts)
    result: list[float] = []
    for weight, count in zip(weights, counts):
        if count <= 0:
            if weight:
                raise ValueError("Positive SF weight assigned to a range absent from the 10^8 report")
            result.append(0.0)
        else:
            result.append((weight / WEIGHT_TOTAL) / (count / total))
    return result


def smooth_log_values(values: list[float]) -> list[float]:
    """Triangular five-point smoothing for sparse competitor buckets."""
    kernel = ((-2, 1.0), (-1, 2.0), (0, 3.0), (1, 2.0), (2, 1.0))
    smoothed: list[float] = []
    for index in range(len(values)):
        numerator = 0.0
        denominator = 0.0
        for offset, weight in kernel:
            neighbor = index + offset
            if 0 <= neighbor < len(values):
                numerator += values[neighbor] * weight
                denominator += weight
        smoothed.append(numerator / denominator)
    return smoothed


def decreasing_isotonic(values: list[float]) -> list[float]:
    """Least-squares non-increasing projection using pooled adjacent blocks."""
    blocks: list[list[float]] = []
    for index, value in enumerate(values):
        blocks.append([float(index), float(index), value, 1.0])
        while (
            len(blocks) >= 2
            and blocks[-2][2] / blocks[-2][3] < blocks[-1][2] / blocks[-1][3]
        ):
            right = blocks.pop()
            left = blocks.pop()
            blocks.append([
                left[0],
                right[1],
                left[2] + right[2],
                left[3] + right[3],
            ])
    result = [0.0] * len(values)
    for start, end, total, count in blocks:
        for index in range(int(start), int(end) + 1):
            result[index] = total / count
    return result


def competitor_reference(
    eligible: list[int], means: list[float], competitor_counts: list[int]
) -> tuple[list[float], float, int]:
    """Build a smooth competitor line and distance-decayed zero-sample tail."""
    observed_last = max(index for index in eligible if competitor_counts[index] > 0)
    observed = [index for index in eligible if index <= observed_last]
    logs = [math.log(competitor_counts[index] + REFERENCE_PSEUDOCOUNT) for index in observed]
    fitted = decreasing_isotonic(smooth_log_values(logs))
    normalized_slopes = [
        (fitted[position] - fitted[position - 1])
        / max(1.0, (means[observed[position]] - means[observed[position - 1]]) / 50.0)
        for position in range(1, len(fitted))
        if fitted[position] < fitted[position - 1] - 1e-12
    ]
    if not normalized_slopes:
        raise ValueError("Competitor SF samples do not establish a declining multiplier line")
    tail_ratio = min(
        TAIL_RATIO_MAX,
        max(TAIL_RATIO_MIN, math.exp(sum(normalized_slopes) / len(normalized_slopes))),
    )
    observed_position = {index: position for position, index in enumerate(observed)}
    reference = [0.0] * len(means)
    for index in eligible:
        if index <= observed_last:
            reference[index] = math.exp(fitted[observed_position[index]])
        else:
            distance = (means[index] - means[observed_last]) / 50.0
            reference[index] = math.exp(fitted[-1]) * tail_ratio**distance
    return reference, tail_ratio, observed_last


def solve_competitor_probabilities(
    eligible: list[int],
    means: list[float],
    reference: list[float],
    labels: list[str],
    target_scene_rtp: float,
) -> tuple[list[float], list[float], float]:
    """Power-tilt the smooth competitor line while locking the target mean."""
    boost_indices = {
        index for index in eligible if labels[index] in BOOST_BY_LABEL
    }
    if len(boost_indices) != len(BOOST_BY_LABEL):
        raise ValueError("One or more requested SF boost ranges are not naturally eligible")

    def probabilities(power: float) -> tuple[list[float], list[float]]:
        base_logs = decreasing_isotonic([
            math.log(reference[index]) + power * math.log(means[index])
            for index in eligible
        ])
        pivot = max(base_logs)
        base_values = [math.exp(value - pivot) for value in base_logs]
        base_total = sum(base_values)
        base_probabilities = [value / base_total for value in base_values]
        adjusted = [
            value * BOOST_BY_LABEL.get(labels[index], 1.0)
            for value, index in zip(base_values, eligible)
        ]
        adjusted_total = sum(adjusted)
        return [value / adjusted_total for value in adjusted], base_probabilities

    lower, upper = -1.0, 1.0
    while sum(
        probability * means[index]
        for probability, index in zip(probabilities(lower)[0], eligible)
    ) > target_scene_rtp:
        lower *= 2.0
    while sum(
        probability * means[index]
        for probability, index in zip(probabilities(upper)[0], eligible)
    ) < target_scene_rtp:
        upper *= 2.0
    for _ in range(220):
        power = (lower + upper) / 2.0
        current, _base = probabilities(power)
        observed = sum(
            probability * means[index]
            for probability, index in zip(current, eligible)
        )
        if observed < target_scene_rtp:
            lower = power
        else:
            upper = power
    power = (lower + upper) / 2.0
    final, base = probabilities(power)
    return final, base, power


def integerize_total(probabilities: list[float], eligible: list[int]) -> list[int]:
    exact = [probability * WEIGHT_TOTAL for probability in probabilities]
    rounded = [math.floor(value) for value in exact]
    remainder = WEIGHT_TOTAL - sum(rounded)
    order = sorted(
        range(len(rounded)),
        key=lambda index: exact[index] - rounded[index],
        reverse=True,
    )
    for index in order[:remainder]:
        rounded[index] += 1
    weights = [0] * 64
    for index, weight in zip(eligible, rounded):
        weights[index] = weight
    return weights


def geometric_group_distribution(
    indices: list[int],
    mass: float,
    means: list[float],
    ratio: float,
    scene_cap: float,
) -> list[float]:
    if not indices:
        raise ValueError("Empty SF optimization group")
    if not 0.0 < ratio <= 1.0:
        raise ValueError(f"Invalid SF geometric decay ratio: {ratio}")
    if indices != list(range(indices[0], indices[-1] + 1)):
        raise ValueError("Weighted SF ranges must be contiguous for a monotonic Hit Rate line")
    factors = [ratio ** rank for rank in range(len(indices))]
    factor_total = sum(factors)
    result = [0.0] * len(means)
    for index, factor in zip(indices, factors):
        result[index] = mass * factor / factor_total
        if result[index] * means[index] > scene_cap + 1e-12:
            raise ValueError(f"SF range {index} exceeds the per-range RTP cap")
    return result


def solve_probabilities(eligible: list[int], means: list[float], target_scene_rtp: float) -> tuple[list[float], float]:
    scene_cap = target_scene_rtp * PER_RANGE_RTP_SHARE_CAP
    non_profit = [index for index in eligible if index < PROFIT_START_INDEX]
    profit_to_500 = [
        index for index in eligible if PROFIT_START_INDEX <= index < ABOVE_500_START_INDEX
    ]
    above_500 = [index for index in eligible if index >= ABOVE_500_START_INDEX]
    profit_to_500_mass = TARGET_PROFIT_HIT_RATE - TARGET_ABOVE_500_HIT_RATE
    expected = list(range(MIN_SF_WEIGHTED_INDEX, above_500[-1] + 1))
    if eligible != expected:
        raise ValueError(
            "Naturally eligible SF ranges must be contiguous from (100,120] through the last weighted range"
        )

    head = geometric_group_distribution(
        non_profit, 1.0 - TARGET_PROFIT_HIT_RATE, means, HEAD_DECAY_RATIO, scene_cap
    )
    middle = geometric_group_distribution(
        profit_to_500, profit_to_500_mass, means, HEAD_DECAY_RATIO, scene_cap
    )

    def for_tail_ratio(ratio: float) -> list[float]:
        tail = geometric_group_distribution(
            above_500, TARGET_ABOVE_500_HIT_RATE, means, ratio, scene_cap
        )
        result = [0.0] * len(means)
        for group in (head, middle, tail):
            result = [left + right for left, right in zip(result, group)]
        return result

    lower, upper = 0.01, 1.0
    lower_value = sum(p * mean for p, mean in zip(for_tail_ratio(lower), means))
    upper_value = sum(p * mean for p, mean in zip(for_tail_ratio(upper), means))
    if not lower_value <= target_scene_rtp <= upper_value:
        raise ValueError(
            f"Monotonic SF target scene RTP {target_scene_rtp} is infeasible: "
            f"{lower_value}..{upper_value}"
        )
    for _ in range(220):
        ratio = (lower + upper) / 2.0
        probabilities = for_tail_ratio(ratio)
        observed = sum(p * mean for p, mean in zip(probabilities, means))
        if observed < target_scene_rtp:
            lower = ratio
        else:
            upper = ratio
    tail_ratio = (lower + upper) / 2.0
    probabilities = for_tail_ratio(tail_ratio)

    observed = sum(p * mean for p, mean in zip(probabilities, means))
    if abs(observed - target_scene_rtp) > 1e-8:
        raise ValueError(f"Floating SF target mismatch: {observed} vs {target_scene_rtp}")
    if any(
        probabilities[index] < probabilities[index + 1] - 1e-15
        for index in range(MIN_SF_WEIGHTED_INDEX, eligible[-1])
    ):
        raise ValueError("Floating SF Hit Rate is not monotonically decreasing")
    return probabilities, tail_ratio


def integerize_group(
    probabilities: list[float], indices: list[int], target_weight: int, means: list[float], scene_cap: float
) -> dict[int, int]:
    exact = {index: probabilities[index] * WEIGHT_TOTAL for index in indices}
    caps = {index: math.floor(scene_cap / means[index] * WEIGHT_TOTAL + 1e-9) for index in indices}
    result = {index: min(math.floor(exact[index]), caps[index]) for index in indices}
    remainder = target_weight - sum(result.values())
    order = sorted(indices, key=lambda index: exact[index] - result[index], reverse=True)
    while remainder > 0:
        changed = False
        for index in order:
            if result[index] < caps[index]:
                result[index] += 1
                remainder -= 1
                changed = True
                if remainder == 0:
                    break
        if not changed:
            raise ValueError("Integer SF weights cannot satisfy per-range cap")
    return result


def integerize_probabilities(probabilities: list[float], means: list[float], target_scene_rtp: float) -> list[int]:
    scene_cap = target_scene_rtp * PER_RANGE_RTP_SHARE_CAP
    below_100 = [
        index for index, value in enumerate(probabilities)
        if value > 0 and index <= BELOW_100_END_INDEX
    ]
    from_100_to_250 = [
        index for index, value in enumerate(probabilities)
        if value > 0 and BELOW_100_END_INDEX < index < PROFIT_START_INDEX
    ]
    profit_to_500 = [
        index for index, value in enumerate(probabilities)
        if value > 0 and PROFIT_START_INDEX <= index < ABOVE_500_START_INDEX
    ]
    above_500 = [
        index for index, value in enumerate(probabilities)
        if value > 0 and index >= ABOVE_500_START_INDEX
    ]
    below_100_target = min(
        500_000_000,
        round(sum(probabilities[index] for index in below_100) * WEIGHT_TOTAL),
    )
    weights = [0] * len(probabilities)
    groups = (
        (below_100, below_100_target),
        (from_100_to_250, 700_000_000 - below_100_target),
        (profit_to_500, 240_000_000),
        (above_500, 60_000_000),
    )
    for indices, target in groups:
        for index, value in integerize_group(
            probabilities, indices, target, means, scene_cap
        ).items():
            weights[index] = value
    return weights


def recalibrate_sf(
    baseline: dict[str, Any],
    labels: list[str],
    report: dict[str, Any],
    competitor: dict[str, Any],
) -> dict[str, Any]:
    baseline_weights = [int(value) for value in baseline["weights"]]
    if len(baseline_weights) != 64 or sum(baseline_weights) != WEIGHT_TOTAL:
        raise ValueError("Previous SF weights must contain 64 rows summing to 1,000,000,000")
    counts = report["sf_count"]
    pays = report["sf_pay"]
    base_bet = report["base_bet"]
    means = [pay / count / base_bet if count else 0.0 for count, pay in zip(counts, pays)]
    natural_rates = [count / report["rounds"] for count in counts]
    eligible = [
        index
        for index in range(MIN_SF_WEIGHTED_INDEX, 64)
        if natural_rates[index] > MIN_NATURAL_RATE and counts[index] > 0 and means[index] > 0
    ]
    if MIN_SF_WEIGHTED_INDEX not in eligible:
        raise ValueError("The required minimum SF range (100,120] is not naturally supported")
    if not any(index >= PROFIT_START_INDEX for index in eligible):
        raise ValueError("No naturally supported profitable SF ranges")
    target_scene_rtp = TARGET_SF_RTP * SUPER_PRICE
    references, tail_decay_ratio, observed_last = competitor_reference(
        eligible, means, competitor["counts"]
    )
    probabilities, base_probabilities, power = solve_competitor_probabilities(
        eligible, means, references, labels, target_scene_rtp
    )
    weights = integerize_total(probabilities, eligible)
    after = sum(weight / WEIGHT_TOTAL * mean for weight, mean in zip(weights, means))
    profit_hit = sum(weights[PROFIT_START_INDEX:]) / WEIGHT_TOTAL
    above_500_hit = sum(weights[ABOVE_500_START_INDEX:]) / WEIGHT_TOTAL
    below_100_hit = sum(weights[:BELOW_100_END_INDEX + 1]) / WEIGHT_TOTAL
    range_shares = [
        (weight / WEIGHT_TOTAL * mean / after) if after else 0.0
        for weight, mean in zip(weights, means)
    ]

    if sum(weights) != WEIGHT_TOTAL:
        raise ValueError(f"SF integer weight sum is {sum(weights)}")
    if abs(after - target_scene_rtp) > 0.00075:
        raise ValueError(f"SF integerized target mismatch: {after} vs {target_scene_rtp}")
    if any(weights[:MIN_SF_WEIGHTED_INDEX]):
        raise ValueError("SF has weight below the required (100,120] minimum range")
    if weights[MIN_SF_WEIGHTED_INDEX] <= 0:
        raise ValueError("SF minimum range (100,120] has no weight")
    if any(weight and natural_rates[index] <= MIN_NATURAL_RATE for index, weight in enumerate(weights)):
        raise ValueError("SF has weight in a range with natural probability at or below 0.07%")
    if max(range_shares) > PER_RANGE_RTP_SHARE_CAP + 1e-8:
        raise ValueError(f"SF single-range RTP share exceeds 15%: {max(range_shares)}")
    boost_indices = sorted(
        index for index in eligible if labels[index] in BOOST_BY_LABEL
    )
    monotonic_violations = [
        index
        for index in range(MIN_SF_WEIGHTED_INDEX, 63)
        if weights[index] < weights[index + 1] and index + 1 not in boost_indices
    ]
    if monotonic_violations:
        raise ValueError(
            f"SF non-boost Hit Rate increases at range indices {monotonic_violations}"
        )

    audits = []
    competitor_eligible_total = sum(competitor["counts"][index] for index in eligible)
    for index, label in enumerate(labels):
        after_rtp = weights[index] / WEIGHT_TOTAL * means[index]
        if index < MIN_SF_WEIGHTED_INDEX:
            rule = "Below minimum (100,120]: disabled"
        elif natural_rates[index] <= MIN_NATURAL_RATE:
            rule = "H016 natural probability at or below 0.07%: disabled"
        elif label in BOOST_BY_LABEL:
            rule = f"Competitor-smoothed reference x{BOOST_BY_LABEL[label]:g}"
        elif index in eligible:
            rule = "Competitor-smoothed SF Hit Rate with a shared RTP tilt"
        else:
            rule = "No naturally supported SF payout"
        audits.append({
            "index": index,
            "range": label,
            "rule": rule,
            "natural_rate": natural_rates[index],
            "natural_count": counts[index],
            "natural_mean": means[index],
            "competitor_count": competitor["counts"][index],
            "competitor_hit_rate_all": competitor["counts"][index] / competitor["samples"],
            "competitor_hit_rate_eligible": (
                competitor["counts"][index] / competitor_eligible_total
                if competitor_eligible_total and index in eligible
                else 0.0
            ),
            "reference_prior": references[index],
            "reference_base_hit_rate": (
                base_probabilities[eligible.index(index)] if index in eligible else 0.0
            ),
            "boost_factor": BOOST_BY_LABEL.get(label, 1.0),
            "before_weight": baseline_weights[index],
            "after_weight": weights[index],
            "target_scene_rtp": after_rtp,
            "target_rtp": after_rtp,
            "after_hit_rate": weights[index] / WEIGHT_TOTAL,
            "after_rtp": after_rtp,
            "after_rtp_share": range_shares[index],
        })

    return {
        "weights": weights,
        "fix": fix_numbers(weights, counts),
        "audit": audits,
        "zero_bucket_before": baseline_weights[0],
        "zero_bucket_after": weights[0],
        "calibration": (
            "Super Ace feature_buy_2 smoothed Hit Rate; (350,400] and (450,500] use x1.5 "
            "reference factors; all other ranges use a non-increasing shared power tilt; "
            "minimum (100,120]; H016 natural >0.07%; hard 92.5% RTP"
        ),
        "minimum_weighted_index": MIN_SF_WEIGHTED_INDEX,
        "minimum_weighted_range": labels[MIN_SF_WEIGHTED_INDEX],
        "target_scene_rtp": target_scene_rtp,
        "scene_rtp_before": sum(
            baseline_weights[index] / WEIGHT_TOTAL * means[index] for index in range(64)
        ),
        "scene_rtp_after": after,
        "profit_hit_rate": profit_hit,
        "above_500_hit_rate": above_500_hit,
        "below_100_hit_rate": below_100_hit,
        "head_decay_ratio": None,
        "tail_decay_ratio": tail_decay_ratio,
        "competitor_observed_last_index": observed_last,
        "competitor_observed_last_range": labels[observed_last],
        "rtp_power_tilt": power,
        "boost_indices": boost_indices,
        "boost_ranges": {label: factor for label, factor in BOOST_BY_LABEL.items()},
        "monotonic_hit_rate": not monotonic_violations,
        "monotonic_violations": monotonic_violations,
        "max_range_rtp_share": max(range_shares),
        "max_range_rtp_index": max(range(64), key=range_shares.__getitem__),
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Recalculate H016192 SF multiplier weights from H016 natural and Super Ace SF data"
    )
    parser.add_argument("--report", required=True, type=Path)
    parser.add_argument("--competitor", required=True, type=Path)
    parser.add_argument("--previous-payload", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--version", default="6.2.0.0")
    args = parser.parse_args()

    report = load_sf_report(args.report.resolve())
    previous = json.loads(args.previous_payload.read_text(encoding="utf-8"))
    labels = workbook_labels(PROJECT_DIR / "Source" / "H016192A.xlsx")
    competitor = load_competitor_sf(args.competitor.resolve(), interval_bounds(report["intervals"]))

    payload = deepcopy(previous)
    payload["version"] = args.version
    sf = recalibrate_sf(payload["versions"]["92"]["sf"], labels, report, competitor)
    for version in ("92", "94"):
        payload["versions"][version]["sf"] = deepcopy(sf)
        payload["versions"][version]["metrics"]["version"] = args.version
        payload["versions"][version]["metrics"]["sf_rtp"] = (
            sf["scene_rtp_after"] / SUPER_PRICE
        )
    payload.setdefault("rules", {})["sf_rule"] = (
        "H016192A/H016194A SF use Super_Ace_feature_buy_2 as a smoothed Hit Rate reference; "
        "(350,400] and (450,500] receive x1.5 reference factors; all other eligible ranges "
        "use one smooth non-increasing power tilt; minimum weighted range (100,120]; H016 "
        "natural probability must be greater than 0.07%; SF RTP is hard-locked to 92.5%"
    )
    payload["rules"].setdefault("targets", {})["sf"] = TARGET_SF_RTP
    payload["sf_source_report"] = {
        "path": report["path"],
        "rounds": report["rounds"],
        "coin_in": report["coin_in"],
        "base_bet": report["base_bet"],
        "observed_rtp": report["rtp_total"],
        "sf_count": report["sf_count"],
        "sf_pay": report["sf_pay"],
    }
    payload["sf_competitor_reference"] = competitor
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "output": str(args.output.resolve()),
        "source_rounds": report["rounds"],
        "source_rtp": report["rtp_total"],
        "competitor_samples": competitor["samples"],
        "target_sf_rtp": TARGET_SF_RTP,
        "version": args.version,
        "scene_rtp_after": sf["scene_rtp_after"],
        "sf_rtp_after": payload["versions"]["92"]["metrics"]["sf_rtp"],
        "profit_hit_rate": sf["profit_hit_rate"],
        "above_500_hit_rate": sf["above_500_hit_rate"],
        "below_100_hit_rate": sf["below_100_hit_rate"],
        "monotonic_hit_rate": sf["monotonic_hit_rate"],
        "tail_decay_ratio": sf["tail_decay_ratio"],
        "rtp_power_tilt": sf["rtp_power_tilt"],
        "boost_ranges": sf["boost_ranges"],
        "max_range_rtp_share": sf["max_range_rtp_share"],
        "max_range": labels[sf["max_range_rtp_index"]],
        "first_weighted_range": labels[next(i for i, value in enumerate(sf["weights"]) if value)],
        "weight_sum": sum(sf["weights"]),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
