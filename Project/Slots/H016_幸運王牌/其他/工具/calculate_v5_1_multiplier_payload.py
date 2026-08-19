"""Build the H016 v5.1 card multiplier payload.

This recalculation intentionally leaves SF untouched.  Normal BG/FG are
locked to 65%/27%, Buy Feature is locked to 92.5%, and the natural FG-entry
cycle is 130 base-game rounds for every 92/94 and oldhand/newbie profile.
"""
from __future__ import annotations

import argparse
import importlib.util
import json
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[2]
TOOL_DIR = Path(__file__).resolve().parent
CORE_PATH = TOOL_DIR / "calculate_h016_competitor_rule_payload.py"
NORMAL_REPORT = PROJECT_DIR / "Record" / "H0161_05_2608190940_betmode0_108.xlsx"
BF_REPORT = PROJECT_DIR / "Record" / "H0161_05_2608190944_betmode2_107.xlsx"
OUTPUT_PATH = PROJECT_DIR / "其他" / "診斷" / "H016_v5_1_multiplier_payload.json"
COMPETITOR_DIR = Path(
    r"C:\Users\rhinshen\Mine\個人工作區\市場資訊\H5\遊戲資源\JILI"
) / "JILI - Super Ace - m" / "遊戲資料"
COMPETITOR_FILES = (
    "SuperAce_BG_Combined_NoJP.jsonl",
    "SuperAce_BG_3.jsonl",
    "Super_Ace_BG_4.jsonl",
)

TARGET_BG_RTP = 0.65
TARGET_FG_RTP = 0.27
TARGET_BF_RTP = 0.925
TARGET_CYCLE = 130.0
BUY_PRICE = 40.5
NEWBIE_BG_CAP = 30.0
NEWBIE_FG_CAP = 120.0
VERSION = "5.1.0.0"


def load_core():
    spec = importlib.util.spec_from_file_location("h016_multiplier_core", CORE_PATH)
    if spec is None or spec.loader is None:
        raise ImportError(f"Cannot load {CORE_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


core = load_core()


def report_table(path: Path) -> tuple[dict[str, Any], dict[str, int], list[tuple[Any, ...]]]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        fields = {
            str(row[0]): row[1]
            for row in workbook["Base Info"].iter_rows(min_row=2, values_only=True)
            if row[0] is not None
        }
        table = list(workbook["Multiplier Line"].iter_rows(min_row=1, max_row=65, values_only=True))
    finally:
        workbook.close()
    if len(table) != 65:
        raise ValueError(f"{path.name}: expected header plus 64 rows")
    header = {str(value): index for index, value in enumerate(table[0]) if value is not None}
    return fields, header, table[1:]


def load_bf_report(path: Path) -> dict[str, Any]:
    fields, header, rows = report_table(path)
    required = {"Interval", "free_game_cnt_BF", "free_game_pay_BF"}
    missing = required.difference(header)
    if missing:
        raise ValueError(f"{path.name}: missing BF columns {sorted(missing)}")
    count = [int(row[header["free_game_cnt_BF"]] or 0) for row in rows]
    pay = [float(row[header["free_game_pay_BF"]] or 0) for row in rows]
    rounds = int(fields["total_rounds"])
    if sum(count) != rounds:
        raise ValueError(f"{path.name}: BF count {sum(count)} != rounds {rounds}")
    return {
        "path": str(path.resolve()),
        "rounds": rounds,
        "base_bet": 100.0,
        "bf_count": count,
        "bf_pay": pay,
    }


def bucket_index(value: float, labels: list[str]) -> int:
    if value <= 0:
        return 0
    for index, label in enumerate(labels[1:], start=1):
        if core.parse_lower(label) < value <= core.parse_upper(label):
            return index
    raise ValueError(f"Multiplier {value:g} is outside workbook ranges")


def load_competitor_fg(labels: list[str], directory: Path) -> dict[str, Any]:
    files = [directory / name for name in COMPETITOR_FILES]
    missing = [str(path) for path in files if not path.is_file()]
    if missing:
        raise FileNotFoundError(f"Missing Super Ace input: {missing}")
    counts = [0] * 64
    sessions = 0
    for path in files:
        with path.open(encoding="utf-8") as handle:
            for line in handle:
                obj = json.loads(line)
                plates = obj["plate"]["plate"]
                if len(plates) <= 1:
                    continue
                bet = float(obj["bet"])
                if bet <= 0:
                    raise ValueError(f"{path.name}: non-positive bet")
                fg_win = sum(float(plate.get("win", 0.0)) for plate in plates[1:])
                counts[bucket_index(fg_win / bet, labels)] += 1
                sessions += 1
    if sessions <= 0 or sum(counts) != sessions:
        raise ValueError("Super Ace FG session parsing failed")
    return {
        "directory": str(directory.resolve()),
        "files": [str(path.resolve()) for path in files],
        "fg_sessions": sessions,
        "fg_count": counts,
    }


def natural_means(counts: list[int], pay: list[float], base_bet: float) -> list[float]:
    return [
        value / count / base_bet if count else 0.0
        for count, value in zip(counts, pay)
    ]


def profile_line(
    *,
    name: str,
    sheet,
    labels: list[str],
    report: dict[str, Any],
    bg_means: list[float],
    fg_means: list[float],
    competitor: dict[str, Any],
    entry_weight: int,
    bg_cap: float,
    fg_cap: float | None,
) -> dict[str, Any]:
    denominator = core.WEIGHT_TOTAL + entry_weight
    trigger_rate = entry_weight / denominator
    trigger = core.trigger_stats_at_cap(report, bg_cap)
    trigger_bg_rtp = trigger_rate * trigger["average"]
    regular_bg_target = TARGET_BG_RTP - trigger_bg_rtp
    if regular_bg_target <= 0:
        raise ValueError(f"{name}: cap-qualified trigger BG exceeds target")
    bg = core.preserve_paying_hit_rate_shape(
        name=f"{name} BG",
        labels=labels,
        baseline_weights=[int(sheet.cell(row, 11).value or 0) for row in range(15, 79)],
        natural_counts=report["bg_count"],
        means=bg_means,
        probability_denominator=denominator,
        target_scene_rtp=regular_bg_target,
    )
    fg_session_target = TARGET_FG_RTP / trigger_rate
    fg = core.relative_hit_rate_scene(
        name=f"{name} FG",
        labels=labels,
        baseline_weights=[int(sheet.cell(row, 11).value or 0) for row in range(86, 150)],
        natural_counts=report["fg_count"],
        means=fg_means,
        competitor_counts=competitor["fg_count"],
        competitor_denominator=competitor["fg_sessions"],
        probability_denominator=core.WEIGHT_TOTAL,
        target_scene_rtp=fg_session_target,
        hit_rate_boosts=core.FG_HIT_RATE_BOOSTS,
        maximum_supported_upper=fg_cap,
        minimum_supported_upper=6.0,
        minimum_bucket_rtp_share=0.002,
        allow_zero_result_weight=False,
    )
    return {
        "bg": bg,
        "fg": fg,
        "trigger": trigger,
        "trigger_bg_rtp": trigger_bg_rtp,
        "normal_rtp": bg["scene_rtp_after"] + trigger_bg_rtp + trigger_rate * fg["scene_rtp_after"],
        "bg_rtp": bg["scene_rtp_after"] + trigger_bg_rtp,
        "fg_rtp": trigger_rate * fg["scene_rtp_after"],
        "fg_session_target": fg_session_target,
    }


def build_version(
    key: str,
    workbook_path: Path,
    report: dict[str, Any],
    bf_report: dict[str, Any],
    competitor: dict[str, Any],
    entry_weight: int,
) -> dict[str, Any]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True, keep_links=False)
    try:
        detail = workbook["Detail"]
        newbie_sheet = workbook["Detail_Newbie"]
        labels = [str(detail.cell(row, 1).value or "").strip() for row in range(15, 79)]
        bg_means = natural_means(report["bg_count"], report["bg_pay"], report["coin_in"])
        fg_means = natural_means(report["fg_count"], report["fg_pay"], report["coin_in"])
        bf_means = natural_means(bf_report["bf_count"], bf_report["bf_pay"], bf_report["base_bet"])

        oldhand_cap = core.enabled_bg_cap(detail, labels)
        if not math.isclose(oldhand_cap, 70.0):
            raise ValueError(f"{workbook_path.name}: expected oldhand BG cap 70x, got {oldhand_cap:g}x")
        detected_newbie_cap = core.enabled_bg_cap(newbie_sheet, labels)
        if not math.isclose(detected_newbie_cap, NEWBIE_BG_CAP):
            raise ValueError(
                f"{workbook_path.name}: expected newbie BG cap {NEWBIE_BG_CAP:g}x, "
                f"got {detected_newbie_cap:g}x"
            )
        oldhand = profile_line(
            name=f"{key} oldhand",
            sheet=detail,
            labels=labels,
            report=report,
            bg_means=bg_means,
            fg_means=fg_means,
            competitor=competitor,
            entry_weight=entry_weight,
            bg_cap=oldhand_cap,
            fg_cap=None,
        )
        newbie = profile_line(
            name=f"{key} newbie",
            sheet=newbie_sheet,
            labels=labels,
            report=report,
            bg_means=bg_means,
            fg_means=fg_means,
            competitor=competitor,
            entry_weight=entry_weight,
            bg_cap=detected_newbie_cap,
            fg_cap=NEWBIE_FG_CAP,
        )
        bf = core.relative_hit_rate_scene(
            name=f"{key} BF",
            labels=labels,
            baseline_weights=[int(detail.cell(row, 11).value or 0) for row in range(163, 227)],
            natural_counts=bf_report["bf_count"],
            means=bf_means,
            competitor_counts=competitor["fg_count"],
            competitor_denominator=competitor["fg_sessions"],
            probability_denominator=core.WEIGHT_TOTAL,
            target_scene_rtp=TARGET_BF_RTP * BUY_PRICE,
            hit_rate_boosts=core.FG_HIT_RATE_BOOSTS,
            minimum_supported_upper=6.0,
            minimum_bucket_rtp_share=0.002,
            allow_zero_result_weight=False,
        )

        # SF is copied byte-for-byte at the workbook/config synchronization stage.
        # These values only allow the existing apply script to validate it read-only.
        sf_counts = [int(detail.cell(row, 2).value or 0) for row in range(234, 298)]
        sf_pay = [float(detail.cell(row, 3).value or 0) for row in range(234, 298)]
        sf_means = natural_means(sf_counts, sf_pay, 100.0)
        sf = core.workbook_scene(detail, 234, sf_counts, sf_means)
        sf["audit"] = []

        trigger_rate = entry_weight / (core.WEIGHT_TOTAL + entry_weight)
        return {
            "bg": oldhand["bg"],
            "fg": oldhand["fg"],
            "newbie": {"bg": newbie["bg"], "fg": newbie["fg"]},
            "bf": bf,
            "sf": sf,
            "metrics": {
                "version": VERSION,
                "target_rtp": TARGET_BG_RTP + TARGET_FG_RTP,
                "entry_weight": entry_weight,
                "trigger_rate": trigger_rate,
                "fg_cycle": 1.0 / trigger_rate,
                "trigger_bg_count": oldhand["trigger"]["count"],
                "trigger_bg_pay": oldhand["trigger"]["pay"],
                "trigger_bg_average": oldhand["trigger"]["average"],
                "trigger_bg_cap": oldhand_cap,
                "newbie_trigger_bg_count": newbie["trigger"]["count"],
                "newbie_trigger_bg_pay": newbie["trigger"]["pay"],
                "newbie_trigger_bg_average": newbie["trigger"]["average"],
                "newbie_trigger_bg_cap": detected_newbie_cap,
                "normal_rtp": oldhand["normal_rtp"],
                "bg_rtp": oldhand["bg_rtp"],
                "fg_rtp": oldhand["fg_rtp"],
                "newbie_rtp": newbie["normal_rtp"],
                "newbie_bg_rtp": newbie["bg_rtp"],
                "newbie_fg_rtp": newbie["fg_rtp"],
                "fg_target_session_rtp": oldhand["fg_session_target"],
                "bf_rtp": bf["scene_rtp_after"] / BUY_PRICE,
                "sf_rtp": sf["scene_rtp_after"] / 250.0,
                "buy_price": BUY_PRICE,
                "super_price": 250.0,
            },
        }
    finally:
        workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--normal-report", type=Path, default=NORMAL_REPORT)
    parser.add_argument("--bf-report", type=Path, default=BF_REPORT)
    parser.add_argument("--competitor-dir", type=Path, default=COMPETITOR_DIR)
    parser.add_argument("--output", type=Path, default=OUTPUT_PATH)
    args = parser.parse_args()

    report = core.load_h016_report(args.normal_report)
    bf_report = load_bf_report(args.bf_report)
    workbook_92 = PROJECT_DIR / "Source" / "H016192A.xlsx"
    workbook = load_workbook(workbook_92, read_only=True, data_only=True, keep_links=False)
    try:
        labels = [str(workbook["Detail"].cell(row, 1).value or "").strip() for row in range(15, 79)]
    finally:
        workbook.close()
    competitor = load_competitor_fg(labels, args.competitor_dir)
    entry_weight = round(core.WEIGHT_TOTAL / (TARGET_CYCLE - 1.0))
    actual_cycle = (core.WEIGHT_TOTAL + entry_weight) / entry_weight
    if abs(actual_cycle - TARGET_CYCLE) > 0.00001:
        raise AssertionError(f"Integer entry weight gives unexpected cycle {actual_cycle}")

    versions = {
        key: build_version(
            key,
            PROJECT_DIR / "Source" / f"H0161{key}A.xlsx",
            report,
            bf_report,
            competitor,
            entry_weight,
        )
        for key in ("92", "94")
    }
    payload = {
        "version": VERSION,
        "rules": {
            "sf": "untouched",
            "normal_targets": {"bg": TARGET_BG_RTP, "fg": TARGET_FG_RTP, "total": 0.92},
            "bf_target": TARGET_BF_RTP,
            "fg_cycle": TARGET_CYCLE,
            "fg_minimum_weighted_range": "(5, 6]",
            "minimum_h016_natural_rate": core.MIN_NATURAL_RATE,
            "minimum_eligible_fg_bucket_rtp_share": 0.002,
            "fg_hit_rate_shape": "Super Ace relative session Hit Rate; (50, 60] x2",
            "profile_caps": {
                "oldhand_bg": 70.0,
                "newbie_bg": NEWBIE_BG_CAP,
                "newbie_fg": NEWBIE_FG_CAP,
            },
        },
        "source_report": {
            "path": report["path"],
            "rounds": report["rounds"],
            "coin_in": report["coin_in"],
            "trigger_count": report["trigger_count"],
            "trigger_pay": report["trigger_pay"],
            "bg_count": report["bg_count"],
            "bg_pay": report["bg_pay"],
            "fg_count": report["fg_count"],
            "fg_pay": report["fg_pay"],
            "interval_upper": report["interval_upper"],
            "trigger_count_lte": report["trigger_count_lte"],
            "trigger_pay_lte": report["trigger_pay_lte"],
        },
        "bf_source_report": bf_report,
        "competitor": competitor,
        "versions": versions,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "output": str(args.output.resolve()),
        "version": VERSION,
        "entry_weight": entry_weight,
        "fg_cycle": actual_cycle,
        "competitor_fg_sessions": competitor["fg_sessions"],
        "versions": {
            key: {
                field: versions[key]["metrics"][field]
                for field in (
                    "normal_rtp", "bg_rtp", "fg_rtp",
                    "newbie_rtp", "newbie_bg_rtp", "newbie_fg_rtp",
                    "bf_rtp", "sf_rtp",
                )
            }
            for key in ("92", "94")
        },
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
