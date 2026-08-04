#!/usr/bin/env python3
"""Apply Lucky Neko BG/FG symbol, silver-frame, drop, and R7 ratios to H028."""

from __future__ import annotations

import argparse
import json
import random
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_ANALYSIS = ROOT / "其他" / "參考資料" / "analysis_lucky_neko.xlsx"
DEFAULT_CONFIG = ROOT / "config_92A.js"
SYMBOL_COUNT = 26
BASE_SYMBOL_COUNT = 13
DROP_WEIGHT_TOTAL = 1_000_000


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Apply Lucky Neko ratios to active H028 BG/FG config sets."
    )
    parser.add_argument("--analysis", type=Path, default=DEFAULT_ANALYSIS)
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--seed", type=int, default=20260803)
    parser.add_argument("--check", action="store_true")
    parser.add_argument("--keep-version", action="store_true")
    return parser.parse_args()


def load_config(path: Path) -> tuple[dict, str]:
    text = path.read_text(encoding="utf-8-sig")
    match = re.fullmatch(r"\s*const\s+data\s*=\s*(\{.*\})\s*;\s*", text, re.S)
    if not match:
        raise ValueError(f"Unsupported config format: {path}")
    return json.loads(match.group(1)), text


def section_rows(ws, label: str) -> list[int]:
    start = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == label:
            start = row + 2
            break
    if start is None:
        raise ValueError(f"{ws.title}: missing section {label}")
    rows = []
    for row in range(start, ws.max_row + 1):
        if ws.cell(row, 1).value is None:
            break
        rows.append(row)
    return rows


def read_symbol_ratios(wb, sheet: str, mode: str) -> list[list[float]]:
    ws = wb[sheet]
    ratios = [[0.0] * BASE_SYMBOL_COUNT for _ in range(6)]
    for row in section_rows(ws, mode):
        symbol_id = int(ws.cell(row, 1).value)
        if not 0 <= symbol_id < BASE_SYMBOL_COUNT:
            raise ValueError(f"{sheet}/{mode}: invalid symbol ID {symbol_id}")
        for reel in range(6):
            ratios[reel][symbol_id] = float(ws.cell(row, reel + 2).value or 0)
    return ratios


def read_count_distributions(wb, sheet: str, mode: str) -> list[dict[int, float]]:
    ws = wb[sheet]
    distributions = [dict() for _ in range(6)]
    for row in section_rows(ws, mode):
        count = int(ws.cell(row, 1).value)
        for reel in range(6):
            distributions[reel][count] = float(ws.cell(row, reel + 2).value or 0)
    return distributions


def expected_value(distribution: dict[int, float]) -> float:
    return sum(value * probability for value, probability in distribution.items())


def normalize(values: list[float]) -> list[float]:
    total = sum(values)
    if total <= 0:
        raise ValueError("Cannot normalize an empty distribution")
    return [value / total for value in values]


def split_silver(base_probabilities: list[float], silver_share: float) -> list[float]:
    base = normalize(base_probabilities)
    eligible_total = sum(base[2:13])
    silver_share = min(max(float(silver_share), 0.0), eligible_total)
    silver_fraction = silver_share / eligible_total if eligible_total else 0.0
    result = [0.0] * SYMBOL_COUNT
    result[0] = base[0]
    result[1] = base[1]
    for symbol_id in range(2, 13):
        silver_probability = base[symbol_id] * silver_fraction
        result[symbol_id] = base[symbol_id] - silver_probability
        result[symbol_id + 11] = silver_probability
    return normalize(result)


def largest_remainder(probabilities: list[float], total: int) -> list[int]:
    normalized = normalize(probabilities)
    exact = [value * total for value in normalized]
    result = [int(value) for value in exact]
    remaining = total - sum(result)
    order = sorted(range(len(exact)), key=lambda i: (exact[i] - result[i], -i), reverse=True)
    for index in order[:remaining]:
        result[index] += 1
    return result


def largest_remainder_keep_positive(probabilities: list[float], total: int) -> list[int]:
    normalized = normalize(probabilities)
    result = largest_remainder(normalized, total)
    exact = [value * total for value in normalized]
    for index, probability in enumerate(normalized):
        if probability <= 0 or result[index] > 0:
            continue
        donors = [i for i, count in enumerate(result) if count > 1]
        donor = max(donors, key=lambda i: (result[i] - exact[i], result[i], -i))
        result[donor] -= 1
        result[index] = 1
    return result


def shuffled_reel(counts: list[int], seed: int) -> list[int]:
    reel = [symbol_id for symbol_id, count in enumerate(counts) for _ in range(count)]
    random.Random(seed).shuffle(reel)
    return reel


def allocate_initial_counts(probabilities: list[float], total: int) -> list[int]:
    collapsed = probabilities[:BASE_SYMBOL_COUNT]
    for symbol_id in range(2, 13):
        collapsed[symbol_id] += probabilities[symbol_id + 11]
    base_counts = largest_remainder(collapsed, total)

    silver_probabilities = [probabilities[symbol_id + 11] for symbol_id in range(2, 13)]
    silver_total = min(round(sum(silver_probabilities) * total), sum(base_counts[2:13]))
    if silver_total:
        normalized = normalize(silver_probabilities)
        exact = [value * silver_total for value in normalized]
        silver_counts = [min(int(value), base_counts[index + 2]) for index, value in enumerate(exact)]
        while sum(silver_counts) < silver_total:
            candidates = [
                index for index in range(11) if silver_counts[index] < base_counts[index + 2]
            ]
            best = max(candidates, key=lambda index: (exact[index] - silver_counts[index], -index))
            silver_counts[best] += 1
    else:
        silver_counts = [0] * 11

    result = [0] * SYMBOL_COUNT
    result[0] = base_counts[0]
    result[1] = base_counts[1]
    for index, symbol_id in enumerate(range(2, 13)):
        result[symbol_id] = base_counts[symbol_id] - silver_counts[index]
        result[symbol_id + 11] = silver_counts[index]
    return result


def read_r7_ratios(wb, mode: str) -> list[float]:
    ws = wb["ExtraReelSame"]
    rate_column = 6 if mode == "BG" else 9
    values = [0.0] * SYMBOL_COUNT
    for row in range(2, ws.max_row + 1):
        symbol_id = ws.cell(row, 3).value
        if symbol_id is None:
            continue
        values[int(symbol_id)] = float(ws.cell(row, rate_column).value or 0)
    return normalize(values)


def bump_version(version: str) -> str:
    parts = version.split(".")
    if not parts or not parts[-1].isdigit():
        raise ValueError(f"Cannot bump version: {version}")
    parts[-1] = str(int(parts[-1]) + 1)
    return ".".join(parts)


def build_mode_targets(wb, mode: str) -> dict:
    initial_base = read_symbol_ratios(wb, "SymbolOcc_Init", mode)
    drop_base = read_symbol_ratios(wb, "SymbolOcc_Drop", mode)
    silver_initial = read_count_distributions(wb, "Silver_Init", mode)
    silver_drop = read_count_distributions(wb, "Silver_Drop", mode)
    reel_height = read_count_distributions(wb, "ReelHigh_Init", mode)

    initial_probabilities = []
    drop_probabilities = []
    initial_silver_shares = []
    drop_silver_shares = []
    for reel in range(6):
        expected_blocks = expected_value(reel_height[reel])
        initial_share = (
            expected_value(silver_initial[reel]) / expected_blocks if expected_blocks else 0.0
        )
        # Cascade replacements are 1x1 and the main reel is five cells high.
        drop_share = expected_value(silver_drop[reel]) / 5.0
        initial_silver_shares.append(initial_share)
        drop_silver_shares.append(drop_share)
        initial_probabilities.append(split_silver(initial_base[reel], initial_share))
        drop_probabilities.append(split_silver(drop_base[reel], drop_share))

    r7 = read_r7_ratios(wb, mode)
    return {
        "initial": initial_probabilities + [r7],
        "drop": drop_probabilities + [r7],
        "initial_silver_shares": initial_silver_shares,
        "drop_silver_shares": drop_silver_shares,
    }


def update_config(config: dict, wb, seed: int, keep_version: bool = False) -> dict:
    targets = {mode: build_mode_targets(wb, mode) for mode in ("BG", "FG")}
    mode_settings = {
        "BG": ("BaseGame", "BaseGameSymbol1", "BaseGameSymbolWeight1"),
        "FG": ("FreeGame", "FreeGameSymbol1", "FreeGameSymbolWeight1"),
    }

    for mode, (prefix, symbol_key, weight_key) in mode_settings.items():
        lengths = [len(reel) for reel in config[symbol_key]]
        if len(lengths) != 7 or any(length <= 0 for length in lengths):
            raise ValueError(f"{symbol_key}: expected seven non-empty reels")

        new_reels = []
        for reel, length in enumerate(lengths):
            counts = (
                allocate_initial_counts(targets[mode]["initial"][reel], length)
                if reel < 6
                else largest_remainder_keep_positive(targets[mode]["initial"][reel], length)
            )
            new_reels.append(shuffled_reel(counts, seed + (0 if mode == "BG" else 10_000) + reel))
        config[symbol_key] = new_reels
        config[weight_key] = [[1] * length for length in lengths]

        drop_weights = [
            largest_remainder(targets[mode]["drop"][reel], DROP_WEIGHT_TOTAL)
            for reel in range(7)
        ]
        for combo in range(1, 6):
            config[f"{prefix}1Drop{combo}"] = [row[:] for row in drop_weights]

    if not keep_version:
        config["excel_version"] = bump_version(str(config["excel_version"]))
    return targets


def main() -> int:
    args = parse_args()
    config, original_text = load_config(args.config)
    original_version = str(config["excel_version"])
    wb = load_workbook(args.analysis, read_only=True, data_only=True)
    try:
        targets = update_config(config, wb, args.seed, args.keep_version)
    finally:
        wb.close()

    new_text = "const data = " + json.dumps(config, ensure_ascii=False, indent=2) + ";\n"
    changed = new_text != original_text
    print(f"Analysis: {args.analysis}")
    print(f"Config: {args.config}")
    print(f"Version: {original_version} -> {config['excel_version']}")
    print(f"Changed: {changed}")
    for mode in ("BG", "FG"):
        initial = ", ".join(f"{value:.6%}" for value in targets[mode]["initial_silver_shares"])
        drop = ", ".join(f"{value:.6%}" for value in targets[mode]["drop_silver_shares"])
        print(f"{mode} Silver_Init shares R1-R6: {initial}")
        print(f"{mode} Silver_Drop shares R1-R6: {drop}")

    if args.check:
        return 1 if changed else 0

    temporary = args.config.with_suffix(args.config.suffix + ".tmp")
    temporary.write_text(new_text, encoding="utf-8", newline="\n")
    temporary.replace(args.config)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
