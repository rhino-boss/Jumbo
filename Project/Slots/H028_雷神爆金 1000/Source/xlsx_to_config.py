import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_SOURCE = BASE_DIR / "H028192A.xlsx"
DEFAULT_OUTPUT = BASE_DIR.parent / "config_92.js"

METADATA = {
    "game_id": "101016",
    "parsheet_id": "H0281",
    "display_name": "Thunder Boost 1000",
    "game_name_zh": "雷神爆金1000",
    "default_coin_in": 100,
    "mode_normalbet": 0,
    "mode_extrabet": 1,
    "mode_featurebuy": 2,
    "normalbet": 1,
    "featurebuy": 75,
    "supported_bet_modes": [0, 2],
}

SYMBOL_SHEET_GROUPS = (
    ("BaseGame", 1, "BG_Symbol", 150),
    ("BaseGame", 2, "BG_Symbol (2)", 121),
    ("FreeGame", 1, "FG_Symbol", 121),
    ("FreeGame", 2, "FG_Symbol (2)", 121),
    ("FreeGame", 3, "FG_Symbol (3)", 121),
)

# H028192A current layout. Keep this as the single source of truth for both
# xlsx_to_config.py and config_to_xlsx.py.
SYMBOL_SHEET_RANGES = {
    "MegaWay": "C33:H47",
    "MY": "C51:C63",
    "PostC1": "B67:C74",
}
DROP_START_ROWS = (4, 33, 62, 91, 120)

PARAMETER_RANGES = {
    "ReelWeight": "C5:C6",
    "FreeReelWeight": "C11:C13",
    "FreeTriggerReel": "C18:C20",
}


def extract_transposed(sheet, range_text):
    rows = [[cell.value for cell in row] for row in sheet[range_text]]
    columns = [list(column) for column in zip(*rows)]
    if len(columns) == 1:
        return columns[0]
    return columns


def extract_no_transpose(sheet, range_text):
    return [[cell.value for cell in row] for row in sheet[range_text]]


def extract_symbol_ids(sheet, range_text):
    symbol_to_id = {
        str(sheet.cell(row, 1).value): int(sheet.cell(row, 10).value)
        for row in range(4, 30)
        if sheet.cell(row, 1).value is not None and sheet.cell(row, 10).value is not None
    }
    symbol_reels = extract_transposed(sheet, range_text)
    result = []
    for reel_index, reel in enumerate(symbol_reels, start=1):
        try:
            result.append([symbol_to_id[str(symbol)] for symbol in reel])
        except KeyError as error:
            raise ValueError(
                f"Unknown symbol {error.args[0]!r} in {sheet.title} reel R{reel_index}"
            ) from error
    return result


def get_sheet(workbook, *names):
    for name in names:
        if name in workbook.sheetnames:
            return workbook[name]
    raise KeyError(f"Worksheet not found; tried {names}. Available: {workbook.sheetnames}")


def extract_linkpoint(overview):
    for row_index in range(1, overview.max_row + 1):
        if overview.cell(row_index, 1).value == "M1":
            return extract_no_transpose(overview, f"C{row_index}:F{row_index + 10}")
    raise ValueError("Pay-table start symbol M1 was not found in Overview")


def parse_card_range(label):
    if label is None:
        return None
    match = re.fullmatch(r"\(\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*\]", str(label).strip())
    if not match:
        return None
    return float(match.group(1)), float(match.group(2))


def parse_card_system(ws):
    profiles = {
        "newbie": {"normal_bet": {"weight_bg": [], "weight_fg": []}},
        "oldhand": {
            "normal_bet": {"weight_bg": [], "weight_fg": []},
            "buy_feature": {"weight_fg": []},
        },
    }
    header_row = next(
        row for row in range(1, ws.max_row + 1)
        if str(ws.cell(row, 2).value).strip().lower() == "range"
    )
    profile_columns = {}
    for col in range(3, ws.max_column + 1):
        player = str(ws.cell(header_row - 1, col).value or "").strip().lower()
        header = str(ws.cell(header_row, col).value or "").strip()
        if player == "newbie" and header == "Weight_NB_BG":
            profile_columns[col] = ("newbie", "normal_bet", "weight_bg")
        elif player == "newbie" and header == "Weight_NB_FG":
            profile_columns[col] = ("newbie", "normal_bet", "weight_fg")
        elif player == "oldhand" and header == "Weight_NB_BG":
            profile_columns[col] = ("oldhand", "normal_bet", "weight_bg")
        elif player == "oldhand" and header == "Weight_NB_FG":
            profile_columns[col] = ("oldhand", "normal_bet", "weight_fg")
        elif player == "oldhand" and header == "Weight_BF_FG":
            profile_columns[col] = ("oldhand", "buy_feature", "weight_fg")

    row = header_row + 1
    while ws.cell(row, 2).value is not None:
        label = str(ws.cell(row, 2).value).strip()
        range_pair = parse_card_range(label)
        for col, (player, mode, segment) in profile_columns.items():
            weight = int(ws.cell(row, col).value or 0)
            if label.lower() == "free game":
                card = {"type": "free_game", "weight": weight}
            elif range_pair is not None:
                card = {"type": "range", "min": range_pair[0], "max": range_pair[1], "weight": weight}
            else:
                continue
            profiles[player][mode][segment].append(card)
        row += 1
    return {"enabled": True, "retry_limit": 5000, **profiles}


def build_config(source_path):
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    output = dict(METADATA)

    overview = workbook["Overview"]
    excel_version = overview["B3"].value
    if excel_version is None or not str(excel_version).strip():
        workbook.close()
        raise ValueError("Overview!B3 (Version) is empty")
    output["excel_version"] = str(excel_version).strip()
    output["linkpoint"] = extract_linkpoint(overview)
    output["card_system"] = parse_card_system(workbook["Multiplier_Weight"])

    for scene, group_index, sheet_name, reel_length in SYMBOL_SHEET_GROUPS:
        sheet = workbook[sheet_name]
        last_row = 3 + reel_length
        output[f"{scene}Symbol{group_index}"] = extract_symbol_ids(sheet, f"M4:S{last_row}")
        output[f"{scene}SymbolWeight{group_index}"] = extract_transposed(sheet, f"AC4:AI{last_row}")
        for field, range_text in SYMBOL_SHEET_RANGES.items():
            if field == "PostC1":
                key = f"{scene}{group_index}PostC1"
            else:
                key = f"{scene}{field}{group_index}"
            output[key] = extract_transposed(sheet, range_text)
        for drop_index, first_row in enumerate(DROP_START_ROWS, start=1):
            key = f"{scene}{group_index}Drop{drop_index}"
            output[key] = extract_transposed(sheet, f"AL{first_row}:AR{first_row + 25}")

    parameter = workbook["Parameter"]
    for key, range_text in PARAMETER_RANGES.items():
        output[key] = extract_transposed(parameter, range_text)

    workbook.close()
    return output


def load_js_config(path):
    text = path.read_text(encoding="utf-8").strip()
    if not text.startswith("const data = "):
        raise ValueError(f"Unsupported config header: {path}")
    return json.loads(text[len("const data = ") :].rstrip(";"))


def write_js_config(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "const data = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )


def derive_output_path(source_path):
    match = re.fullmatch(r"H0281(?P<rtp>\d{2})(?P<variant>[A-Za-z0-9_-]*)", source_path.stem, re.IGNORECASE)
    if not match:
        raise ValueError(
            f"Cannot derive config name from {source_path.name}; expected H0281 + two-digit RTP + optional variant, "
            "for example H028192A.xlsx"
        )
    return DEFAULT_OUTPUT.parent / f"config_{match.group('rtp')}{match.group('variant')}.js"


def compare_config(generated, output_path):
    current = load_js_config(output_path)
    changed_keys = [key for key in generated if generated[key] != current.get(key)]
    extra_keys = [key for key in current if key not in generated]
    if changed_keys or extra_keys:
        raise ValueError(f"Config mismatch for {output_path}. Changed keys: {changed_keys}; extra keys: {extra_keys}")


def process_source(source_path, output_path, check=False):
    generated = build_config(source_path)
    if check:
        compare_config(generated, output_path)
        print(f"Config is current: {output_path}")
    else:
        write_js_config(output_path, generated)
        compare_config(generated, output_path)
        print(f"Config written: {output_path} <- {source_path.name}")
    return generated


def main():
    parser = argparse.ArgumentParser(description="Build H028 config files from xlsx sources")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--auto-output", action="store_true", help="Derive config_RTPVARIANT.js from the xlsx name")
    parser.add_argument("--all", action="store_true", help="Convert every H0281*.xlsx file in the Source folder")
    parser.add_argument(
        "--sync-default",
        action="store_true",
        help=f"Also update {DEFAULT_OUTPUT.name} when converting {DEFAULT_SOURCE.name}",
    )
    parser.add_argument("--check", action="store_true", help="Compare generated data with the existing output")
    args = parser.parse_args()

    if args.all:
        sources = sorted(
            path for path in BASE_DIR.glob("H0281*.xlsx") if not path.name.startswith("~$")
        )
        if not sources:
            raise FileNotFoundError(f"No H0281*.xlsx files found in {BASE_DIR}")
        for source_path in sources:
            output_path = derive_output_path(source_path)
            generated = process_source(source_path, output_path, args.check)
            if args.sync_default and source_path.name.casefold() == DEFAULT_SOURCE.name.casefold():
                if args.check:
                    compare_config(generated, DEFAULT_OUTPUT)
                    print(f"Config is current: {DEFAULT_OUTPUT}")
                else:
                    write_js_config(DEFAULT_OUTPUT, generated)
                    compare_config(generated, DEFAULT_OUTPUT)
                    print(f"Default config written: {DEFAULT_OUTPUT} <- {source_path.name}")
        print(f"Processed {len(sources)} xlsx file(s).")
        return

    source_path = args.source.resolve()
    output_path = (
        args.output.resolve()
        if args.output is not None
        else derive_output_path(source_path)
        if args.auto_output
        else DEFAULT_OUTPUT
    )
    generated = process_source(source_path, output_path, args.check)
    if args.sync_default and source_path.name.casefold() == DEFAULT_SOURCE.name.casefold() and output_path != DEFAULT_OUTPUT:
        if args.check:
            compare_config(generated, DEFAULT_OUTPUT)
            print(f"Config is current: {DEFAULT_OUTPUT}")
        else:
            write_js_config(DEFAULT_OUTPUT, generated)
            compare_config(generated, DEFAULT_OUTPUT)
            print(f"Default config written: {DEFAULT_OUTPUT} <- {source_path.name}")


if __name__ == "__main__":
    main()
