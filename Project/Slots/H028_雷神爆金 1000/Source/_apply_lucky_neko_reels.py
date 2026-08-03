from __future__ import annotations

import argparse
import json
import random
import re
import subprocess
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
CONFIG_PATH = PROJECT_DIR / "config_92A.js"
SOURCE_PATH = Path(
    r"C:\Users\rhinshen\Mine\個人工作區\市場資訊\H5\遊戲資源\PG"
    r"\PG - Lucky Neko\遊玩資料\analysis_lucky_neko.xlsx"
)
SYMBOL_KEYS = {
    "BaseGameSymbolWeight1",
    "BaseGameSymbolWeight2",
    "FreeGameSymbolWeight1",
    "FreeGameSymbolWeight2",
    "FreeGameSymbolWeight3",
}
REEL_KEYS = {
    "BG": [
        ("BaseGameSymbol1", "BaseGameSymbolWeight1"),
        ("BaseGameSymbol2", "BaseGameSymbolWeight2"),
    ],
    "FG": [
        ("FreeGameSymbol1", "FreeGameSymbolWeight1"),
        ("FreeGameSymbol2", "FreeGameSymbolWeight2"),
        ("FreeGameSymbol3", "FreeGameSymbolWeight3"),
    ],
}


def load_config(path: Path) -> dict:
    text = path.read_text(encoding="utf-8")
    match = re.fullmatch(r"\s*const\s+data\s*=\s*(\{.*\})\s*;\s*", text, re.S)
    if not match:
        raise ValueError(f"Unsupported config wrapper: {path}")
    return json.loads(match.group(1))


def load_head_config(path: Path) -> dict:
    repository = Path(__file__).resolve().parents[4]
    relative_path = path.relative_to(repository).as_posix()
    result = subprocess.run(
        ["git", "show", f"HEAD:{relative_path}"],
        cwd=repository,
        check=True,
        capture_output=True,
    )
    text = result.stdout.decode("utf-8")
    match = re.fullmatch(r"\s*const\s+data\s*=\s*(\{.*\})\s*;\s*", text, re.S)
    if not match:
        raise ValueError(f"Unsupported config wrapper in HEAD: {relative_path}")
    return json.loads(match.group(1))


def load_targets(path: Path) -> dict[str, np.ndarray]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["SymbolOcc_Init"]
    targets: dict[str, np.ndarray] = {}
    for label, first_row in (("BG", 3), ("FG", 19)):
        values = np.array(
            [
                [float(sheet.cell(row=row, column=column).value) for column in range(2, 8)]
                for row in range(first_row, first_row + 13)
            ],
            dtype=float,
        )
        values /= values.sum(axis=0, keepdims=True)
        targets[label] = values
    workbook.close()
    return targets


def make_strip(target: np.ndarray, seed: int, length: int = 121) -> np.ndarray:
    raw_counts = target * length
    counts = np.floor(raw_counts).astype(int)
    remaining = length - int(counts.sum())
    for symbol in np.argsort(-(raw_counts - counts))[:remaining]:
        counts[symbol] += 1

    # Very rare positive symbols still need at least one anchor in the strip.
    for symbol in np.where((target > 0) & (counts == 0))[0]:
        donors = [index for index in range(13) if counts[index] > 1]
        donor = max(donors, key=lambda index: counts[index] - raw_counts[index])
        counts[donor] -= 1
        counts[symbol] += 1

    strip = [symbol for symbol, count in enumerate(counts) for _ in range(int(count))]
    random.Random(seed).shuffle(strip)
    return np.array(strip, dtype=int)


def window_matrix(strip: np.ndarray, height: int = 5) -> np.ndarray:
    matrix = np.zeros((13, len(strip)), dtype=float)
    for stop in range(len(strip)):
        for offset in range(height):
            matrix[strip[(stop + offset) % len(strip)], stop] += 1.0 / height
    return matrix


def solve_stop_weights(strip: np.ndarray, target: np.ndarray) -> tuple[np.ndarray, float]:
    matrix = window_matrix(strip)
    # The first 12 symbol equations plus total weight form an independent system.
    constraints = np.vstack([matrix[:12], np.ones(len(strip))])
    expected = np.r_[target[:12], 1.0]
    active = np.ones(len(strip), dtype=bool)
    baseline = np.ones(len(strip), dtype=float) / len(strip)

    for _ in range(len(strip)):
        active_constraints = constraints[:, active]
        active_baseline = baseline[active]
        weights = active_baseline + active_constraints.T @ np.linalg.pinv(
            active_constraints @ active_constraints.T, rcond=1e-13
        ) @ (expected - active_constraints @ active_baseline)
        residual = float(np.max(np.abs(active_constraints @ weights - expected)))
        if weights.min() >= -1e-11 and residual < 1e-10:
            result = np.zeros(len(strip), dtype=float)
            result[active] = np.maximum(weights, 0.0)
            result /= result.sum()
            error = float(np.max(np.abs(matrix @ result - target)))
            return result, error
        active_index = int(np.argmin(weights))
        active[np.where(active)[0][active_index]] = False

    raise RuntimeError("Unable to solve non-negative stop weights")


def integerize_weights(weights: np.ndarray, total: int = 1_000_000_000_000) -> np.ndarray:
    raw = weights * total
    result = np.floor(raw).astype(np.int64)
    remaining = total - int(result.sum())
    for index in np.argsort(-(raw - result))[:remaining]:
        result[index] += 1
    return result


def zero_all_numbers(value):
    if isinstance(value, dict):
        return {key: zero_all_numbers(item) for key, item in value.items()}
    if isinstance(value, list):
        return [zero_all_numbers(item) for item in value]
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return 0
    return value


def zero_named_weight(value):
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return 0
    if isinstance(value, list):
        if all(isinstance(item, (int, float)) and not isinstance(item, bool) for item in value):
            return [0 for _ in value]
        return [zero_other_weights(item) for item in value]
    if isinstance(value, dict):
        return {key: zero_other_weights(item, key) for key, item in value.items()}
    return value


def zero_other_weights(value, key: str = ""):
    if key not in SYMBOL_KEYS and "weight" in key.lower():
        return zero_named_weight(value)
    if re.match(r"^(?:Base|Free)Game(?:\d+)?(?:MegaWay|MY|PostC1|Drop\d+)$", key):
        return zero_all_numbers(value)
    if isinstance(value, dict):
        return {child_key: zero_other_weights(child, child_key) for child_key, child in value.items()}
    if isinstance(value, list):
        return [zero_other_weights(child, key) for child in value]
    return value


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--base-from-head", action="store_true")
    args = parser.parse_args()

    config = load_head_config(CONFIG_PATH) if args.base_from_head else load_config(CONFIG_PATH)
    targets = load_targets(SOURCE_PATH)
    reports = []

    for mode_index, (mode, pairs) in enumerate(REEL_KEYS.items()):
        for reel_index in range(6):
            target = targets[mode][:, reel_index]
            strip = make_strip(target, 28000 + mode_index * 100 + reel_index)
            weights, error = solve_stop_weights(strip, target)
            if error > 1e-10:
                raise RuntimeError(f"{mode} R{reel_index + 1} ratio error: {error}")
            integer_weights = integerize_weights(weights)
            integer_error = float(
                np.max(np.abs(window_matrix(strip) @ (integer_weights / integer_weights.sum()) - target))
            )
            if integer_error > 1e-10:
                raise RuntimeError(
                    f"{mode} R{reel_index + 1} integer ratio error: {integer_error}"
                )
            reports.append(
                (
                    mode,
                    reel_index + 1,
                    integer_error,
                    int(np.count_nonzero(integer_weights)),
                )
            )
            for symbol_key, weight_key in pairs:
                config[symbol_key][reel_index] = strip.tolist()
                config[weight_key][reel_index] = integer_weights.tolist()

    config = zero_other_weights(config)

    for mode, reel, error, positive_stops in reports:
        print(
            f"{mode} R{reel}: max_error={error:.3e}, "
            f"positive_stops={positive_stops}/121"
        )

    if args.apply:
        payload = "const data = " + json.dumps(config, ensure_ascii=False, indent=2) + ";\n"
        CONFIG_PATH.write_text(payload, encoding="utf-8")
        print(f"Updated: {CONFIG_PATH}")
    else:
        print("Check only; use --apply to update config_92A.js")


if __name__ == "__main__":
    main()
