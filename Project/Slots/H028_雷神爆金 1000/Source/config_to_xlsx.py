import argparse
import json
import math
import os
import posixpath
import re
import warnings
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter, range_boundaries

from xlsx_to_config import (
    DROP_START_ROWS,
    PARAMETER_RANGES,
    SYMBOL_SHEET_GROUPS,
    SYMBOL_SHEET_RANGES,
    build_config,
)


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_SOURCE = BASE_DIR / "H028192A.xlsx"
DEFAULT_CONFIG = BASE_DIR.parent / "config_92A.js"

CELL_PATTERN = re.compile(
    r'<c(?P<selfattrs>[^>]*?\br="(?P<selfref>[A-Z]{1,3}[1-9][0-9]*)"[^>]*?)\s*/>'
    r'|<c(?P<attrs>[^>]*?\br="(?P<ref>[A-Z]{1,3}[1-9][0-9]*)"[^>]*?)>'
    r'(?P<body>.*?)</c>',
    re.DOTALL,
)
ROW_PATTERN = re.compile(
    r'<row(?P<attrs>[^>]*\br="(?P<row>[1-9][0-9]*)"[^>]*)>'
    r'(?P<body>.*?)</row>',
    re.DOTALL,
)


def load_js_config(path):
    text = path.read_text(encoding="utf-8").strip()
    prefix = "const data = "
    if not text.startswith(prefix):
        raise ValueError(f"Unsupported config header: {path}")
    return json.loads(text[len(prefix):].rstrip(";"))


def get_sheet_name(workbook, *names):
    for name in names:
        if name in workbook.sheetnames:
            return name
    raise KeyError(f"Worksheet not found; tried {names}. Available: {workbook.sheetnames}")


def require_key(config, key):
    if key not in config:
        raise KeyError(f"Config key is missing: {key}")
    return config[key]


def cell_addresses(range_text):
    min_col, min_row, max_col, max_row = range_boundaries(range_text)
    return [
        [f"{get_column_letter(col)}{row}" for row in range(min_row, max_row + 1)]
        for col in range(min_col, max_col + 1)
    ]


def add_update(updates, sheet_name, address, value, key):
    sheet_updates = updates.setdefault(sheet_name, {})
    if address in sheet_updates and sheet_updates[address][0] != value:
        previous_value, previous_key = sheet_updates[address]
        raise ValueError(
            f"Conflicting writes for {sheet_name}!{address}: "
            f"{previous_key}={previous_value!r}, {key}={value!r}"
        )
    sheet_updates[address] = (value, key)


def add_transposed_range(updates, sheet_name, range_text, values, key):
    columns = cell_addresses(range_text)
    if len(columns) == 1:
        if not isinstance(values, list) or len(values) != len(columns[0]):
            raise ValueError(f"{key} must contain {len(columns[0])} values")
        value_columns = [values]
    else:
        if not isinstance(values, list) or len(values) != len(columns):
            raise ValueError(f"{key} must contain {len(columns)} columns")
        value_columns = values
    for column_index, addresses in enumerate(columns):
        column_values = value_columns[column_index]
        if not isinstance(column_values, list) or len(column_values) != len(addresses):
            raise ValueError(
                f"{key} column {column_index + 1} must contain {len(addresses)} values"
            )
        for address, value in zip(addresses, column_values):
            add_update(updates, sheet_name, address, value, key)


def add_row_major_range(updates, sheet_name, range_text, values, key):
    columns = cell_addresses(range_text)
    row_count = len(columns[0])
    column_count = len(columns)
    if not isinstance(values, list) or len(values) != row_count:
        raise ValueError(f"{key} must contain {row_count} rows")
    for row_index in range(row_count):
        row_values = values[row_index]
        if not isinstance(row_values, list) or len(row_values) != column_count:
            raise ValueError(f"{key} row {row_index + 1} must contain {column_count} values")
        for column_index in range(column_count):
            add_update(
                updates,
                sheet_name,
                columns[column_index][row_index],
                row_values[column_index],
                key,
            )


def parse_range_label(label):
    match = re.fullmatch(
        r"\(\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*\]",
        str(label).strip(),
    )
    return None if not match else (float(match.group(1)), float(match.group(2)))


def card_signature(card):
    if card.get("type") == "free_game":
        return ("free_game",)
    if card.get("type") == "range":
        return ("range", float(card["min"]), float(card["max"]))
    raise ValueError(f"Unsupported card entry: {card!r}")


def label_signature(label):
    if str(label).strip().lower() == "free game":
        return ("free_game",)
    pair = parse_range_label(label)
    if pair is None:
        raise ValueError(f"Unsupported Multiplier_Weight range label: {label!r}")
    return ("range", pair[0], pair[1])


def discover_layout(source_path):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        missing_symbol_sheets = [
            sheet_name for _scene, _group_index, sheet_name, _reel_length in SYMBOL_SHEET_GROUPS
            if sheet_name not in workbook.sheetnames
        ]
        if missing_symbol_sheets:
            raise KeyError(f"Missing symbol worksheets: {missing_symbol_sheets}")
        symbol_by_id = {}
        for _scene, _group_index, sheet_name, _reel_length in SYMBOL_SHEET_GROUPS:
            sheet = workbook[sheet_name]
            symbol_by_id[sheet_name] = {
                int(sheet.cell(row, 10).value): str(sheet.cell(row, 1).value)
                for row in range(4, 30)
                if sheet.cell(row, 1).value is not None and sheet.cell(row, 10).value is not None
            }
        overview = workbook["Overview"]
        linkpoint_row = next(
            row for row in range(1, overview.max_row + 1)
            if overview.cell(row, 1).value == "M1"
        )

        multiplier = workbook["Multiplier_Weight"]
        header_row = next(
            row for row in range(1, multiplier.max_row + 1)
            if str(multiplier.cell(row, 2).value).strip().lower() == "range"
        )
        profile_columns = {}
        for col in range(3, multiplier.max_column + 1):
            player = str(multiplier.cell(header_row - 1, col).value or "").strip().lower()
            header = str(multiplier.cell(header_row, col).value or "").strip()
            mapping = None
            if player == "newbie" and header == "Weight_NB_BG":
                mapping = ("newbie", "normal_bet", "weight_bg")
            elif player == "newbie" and header == "Weight_NB_FG":
                mapping = ("newbie", "normal_bet", "weight_fg")
            elif player == "oldhand" and header == "Weight_NB_BG":
                mapping = ("oldhand", "normal_bet", "weight_bg")
            elif player == "oldhand" and header == "Weight_NB_FG":
                mapping = ("oldhand", "normal_bet", "weight_fg")
            elif player == "oldhand" and header == "Weight_BF_FG":
                mapping = ("oldhand", "buy_feature", "weight_fg")
            if mapping is not None:
                profile_columns[col] = mapping

        labels = []
        row = header_row + 1
        while multiplier.cell(row, 2).value is not None:
            labels.append((row, multiplier.cell(row, 2).value))
            row += 1
        return {
            "linkpoint_row": linkpoint_row,
            "symbol_by_id": symbol_by_id,
            "profile_columns": profile_columns,
            "card_labels": labels,
        }
    finally:
        workbook.close()


def add_card_system_updates(updates, config, layout):
    card_system = require_key(config, "card_system")
    labels = layout["card_labels"]
    for col, (player, mode, segment) in layout["profile_columns"].items():
        try:
            cards = card_system[player][mode][segment]
        except KeyError as error:
            raise KeyError(f"Missing card system path: {player}.{mode}.{segment}") from error
        if len(cards) != len(labels):
            raise ValueError(
                f"card_system.{player}.{mode}.{segment} has {len(cards)} entries; "
                f"xlsx has {len(labels)} range rows"
            )
        for card, (row, label) in zip(cards, labels):
            if card_signature(card) != label_signature(label):
                raise ValueError(
                    f"Card range mismatch at Multiplier_Weight!{get_column_letter(col)}{row}: "
                    f"config={card!r}, xlsx label={label!r}"
                )
            add_update(
                updates,
                "Multiplier_Weight",
                f"{get_column_letter(col)}{row}",
                int(card.get("weight", 0)),
                f"card_system.{player}.{mode}.{segment}",
            )


def build_updates(source_path, config):
    layout = discover_layout(source_path)
    updates = {}
    add_update(updates, "Overview", "B3", require_key(config, "excel_version"), "excel_version")
    start_row = layout["linkpoint_row"]
    add_row_major_range(
        updates,
        "Overview",
        f"C{start_row}:F{start_row + 10}",
        require_key(config, "linkpoint"),
        "linkpoint",
    )
    for scene, group_index, sheet_name, reel_length in SYMBOL_SHEET_GROUPS:
        last_row = 3 + reel_length
        symbol_key = f"{scene}Symbol{group_index}"
        symbol_map = layout["symbol_by_id"][sheet_name]
        symbol_values = []
        for reel_index, reel in enumerate(require_key(config, symbol_key), start=1):
            try:
                symbol_values.append([symbol_map[int(symbol_id)] for symbol_id in reel])
            except (KeyError, TypeError, ValueError) as error:
                raise ValueError(
                    f"{symbol_key} reel R{reel_index} contains an unknown Symbol ID"
                ) from error
        add_transposed_range(
            updates,
            sheet_name,
            f"M4:S{last_row}",
            symbol_values,
            symbol_key,
        )
        weight_key = f"{scene}SymbolWeight{group_index}"
        add_transposed_range(
            updates,
            sheet_name,
            f"AC4:AI{last_row}",
            require_key(config, weight_key),
            weight_key,
        )
        for field, range_text in SYMBOL_SHEET_RANGES.items():
            if field == "PostC1":
                key = f"{scene}{group_index}PostC1"
            else:
                key = f"{scene}{field}{group_index}"
            add_transposed_range(
                updates,
                sheet_name,
                range_text,
                require_key(config, key),
                key,
            )
        for drop_index, first_row in enumerate(DROP_START_ROWS, start=1):
            key = f"{scene}{group_index}Drop{drop_index}"
            add_transposed_range(
                updates,
                sheet_name,
                f"AL{first_row}:AR{first_row + 25}",
                require_key(config, key),
                key,
            )
    for key, range_text in PARAMETER_RANGES.items():
        add_transposed_range(
            updates,
            "Parameter",
            range_text,
            require_key(config, key),
            key,
        )
    add_card_system_updates(updates, config, layout)
    return updates


def values_equal(left, right):
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return float(left) == float(right)
    return left == right


def changed_cells(source_path, updates):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(source_path, read_only=False, data_only=True)
    try:
        changed = []
        for sheet_name, cells in updates.items():
            sheet = workbook[sheet_name]
            for address, (new_value, key) in cells.items():
                old_value = sheet[address].value
                if not values_equal(old_value, new_value):
                    changed.append((sheet_name, address, old_value, new_value, key))
        return changed
    finally:
        workbook.close()


def workbook_sheet_parts(archive):
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_root
    }
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    parts = {}
    for sheet in workbook_root.findall(f".//{{{main_ns}}}sheet"):
        target = rel_targets[sheet.attrib[f"{{{rel_ns}}}id"]]
        if target.startswith("/"):
            part = target.lstrip("/")
        else:
            part = posixpath.normpath(posixpath.join("xl", target))
        parts[sheet.attrib["name"]] = part
    return parts


def number_text(value):
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ValueError(f"Non-finite numeric value cannot be written to xlsx: {value}")
        return format(value, ".15g")
    raise TypeError(f"Unsupported numeric value: {value!r}")


def render_cell_xml(attrs, value):
    attrs = re.sub(r'\s+t="[^"]*"', "", attrs.rstrip())
    if value is None:
        return f"<c{attrs}></c>"
    if isinstance(value, str):
        space = ' xml:space="preserve"' if value != value.strip() else ""
        return f'<c{attrs} t="inlineStr"><is><t{space}>{escape(value)}</t></is></c>'
    return f"<c{attrs}><v>{number_text(value)}</v></c>"


def insert_missing_cells(sheet_xml, sheet_name, remaining, cell_updates):
    by_row = {}
    for address in remaining:
        match = re.fullmatch(r"([A-Z]{1,3})([1-9][0-9]*)", address)
        by_row.setdefault(int(match.group(2)), []).append(address)
    found_rows = set()

    def replace_row(match):
        row_number = int(match.group("row"))
        if row_number not in by_row:
            return match.group(0)
        found_rows.add(row_number)
        body = match.group("body")
        addresses = sorted(
            by_row[row_number],
            key=lambda address: column_index_from_string(re.match(r"[A-Z]+", address).group(0)),
        )
        for address in addresses:
            new_col = column_index_from_string(re.match(r"[A-Z]+", address).group(0))
            insert_at = len(body)
            for cell_match in CELL_PATTERN.finditer(body):
                existing_address = cell_match.group("ref") or cell_match.group("selfref")
                existing_col = column_index_from_string(
                    re.match(r"[A-Z]+", existing_address).group(0)
                )
                if existing_col > new_col:
                    insert_at = cell_match.start()
                    break
            value, _key = cell_updates[address]
            cell_xml = render_cell_xml(f' r="{address}"', value)
            body = body[:insert_at] + cell_xml + body[insert_at:]
            remaining.discard(address)
        return f'<row{match.group("attrs")}>{body}</row>'

    patched = ROW_PATTERN.sub(replace_row, sheet_xml)
    missing_rows = sorted(set(by_row) - found_rows)
    if missing_rows:
        sheet_data_pattern = re.compile(
            r"(?P<open><sheetData[^>]*>)(?P<body>.*?)(?P<close></sheetData>)",
            re.DOTALL,
        )

        def add_rows(match):
            body = match.group("body")
            for row_number in missing_rows:
                addresses = sorted(
                    by_row[row_number],
                    key=lambda address: column_index_from_string(
                        re.match(r"[A-Z]+", address).group(0)
                    ),
                )
                cells = []
                for address in addresses:
                    value, _key = cell_updates[address]
                    cells.append(render_cell_xml(f' r="{address}"', value))
                    remaining.discard(address)
                row_xml = f'<row r="{row_number}">{"".join(cells)}</row>'
                insert_at = len(body)
                for row_match in ROW_PATTERN.finditer(body):
                    if int(row_match.group("row")) > row_number:
                        insert_at = row_match.start()
                        break
                body = body[:insert_at] + row_xml + body[insert_at:]
            return match.group("open") + body + match.group("close")

        patched, count = sheet_data_pattern.subn(add_rows, patched, count=1)
        if count != 1:
            raise ValueError(f"sheetData element not found in {sheet_name}")
    return patched


def patch_sheet_xml(xml_bytes, sheet_name, cell_updates, overwrite_formulas=False):
    text = xml_bytes.decode("utf-8")
    remaining = set(cell_updates)

    def replace_cell(match):
        address = match.group("ref") or match.group("selfref")
        if address not in cell_updates:
            return match.group(0)
        body = match.group("body") or ""
        if re.search(r"<f(?:\s|>)", body) and not overwrite_formulas:
            raise ValueError(f"Refusing to overwrite formula cell {sheet_name}!{address}")
        attrs = match.group("attrs") or match.group("selfattrs") or ""
        value, _key = cell_updates[address]
        remaining.discard(address)
        return render_cell_xml(attrs, value)

    patched = CELL_PATTERN.sub(replace_cell, text)
    if remaining:
        patched = insert_missing_cells(patched, sheet_name, remaining, cell_updates)
    if remaining:
        sample = ", ".join(sorted(remaining)[:10])
        raise ValueError(f"Cells could not be inserted in {sheet_name}: {sample}")
    payload = patched.encode("utf-8")
    try:
        ET.fromstring(payload)
    except ET.ParseError as error:
        raise ValueError(f"Invalid worksheet XML after patching {sheet_name}: {error}") from error
    return payload


def restore_symbol_id_formulas(xml_bytes, sheet_name, last_row):
    """Restore U:AA formulas after older tool versions replaced them with values."""
    text = xml_bytes.decode("utf-8")
    restored = set()

    def replace_cell(match):
        address = match.group("ref") or match.group("selfref")
        column_match = re.fullmatch(r"([A-Z]+)([1-9][0-9]*)", address)
        column = column_index_from_string(column_match.group(1))
        row = int(column_match.group(2))
        if not (column_index_from_string("U") <= column <= column_index_from_string("AA")):
            return match.group(0)
        if not (4 <= row <= last_row):
            return match.group(0)
        source_column = get_column_letter(column - 8)
        attrs = match.group("attrs") or match.group("selfattrs") or ""
        attrs = re.sub(r'\s+t="[^"]*"', "", attrs.rstrip())
        restored.add(address)
        return (
            f"<c{attrs}><f>VLOOKUP({source_column}{row}, "
            "$A$4:$J$29, 10, 0)</f></c>"
        )

    patched = CELL_PATTERN.sub(replace_cell, text)

    missing_by_row = {}
    for row in range(4, last_row + 1):
        for column in range(column_index_from_string("U"), column_index_from_string("AA") + 1):
            address = f"{get_column_letter(column)}{row}"
            if address not in restored:
                missing_by_row.setdefault(row, []).append(address)

    def insert_formula_cells(match):
        row = int(match.group("row"))
        if row not in missing_by_row:
            return match.group(0)
        body = match.group("body")
        for address in missing_by_row[row]:
            column = column_index_from_string(re.match(r"[A-Z]+", address).group(0))
            source_column = get_column_letter(column - 8)
            insert_at = len(body)
            for cell_match in CELL_PATTERN.finditer(body):
                existing_address = cell_match.group("ref") or cell_match.group("selfref")
                existing_column = column_index_from_string(
                    re.match(r"[A-Z]+", existing_address).group(0)
                )
                if existing_column > column:
                    insert_at = cell_match.start()
                    break
            formula_xml = (
                f'<c r="{address}"><f>VLOOKUP({source_column}{row}, '
                "$A$4:$J$29, 10, 0)</f></c>"
            )
            body = body[:insert_at] + formula_xml + body[insert_at:]
            restored.add(address)
        return f'<row{match.group("attrs")}>{body}</row>'

    if missing_by_row:
        patched = ROW_PATTERN.sub(insert_formula_cells, patched)
    expected = 7 * (last_row - 3)
    if len(restored) != expected:
        raise ValueError(
            f"Could not restore every Symbol ID formula in {sheet_name}: "
            f"restored {len(restored)} of {expected}"
        )
    payload = patched.encode("utf-8")
    ET.fromstring(payload)
    return payload


def remove_calc_chain_metadata(filename, payload):
    if filename == "xl/_rels/workbook.xml.rels":
        text = payload.decode("utf-8")
        text = re.sub(
            r'<Relationship\b[^>]*\bType="[^"]*/calcChain"[^>]*/>',
            "",
            text,
        )
        return text.encode("utf-8")
    if filename == "[Content_Types].xml":
        text = payload.decode("utf-8")
        text = re.sub(
            r'<Override\b[^>]*\bPartName="/xl/calcChain\.xml"[^>]*/>',
            "",
            text,
        )
        return text.encode("utf-8")
    if filename == "xl/workbook.xml":
        text = payload.decode("utf-8")

        def update_calc_properties(match):
            attrs = match.group(1)
            for name, value in (
                ("calcMode", "auto"),
                ("fullCalcOnLoad", "1"),
                ("forceFullCalc", "1"),
            ):
                if re.search(rf'\b{name}="[^"]*"', attrs):
                    attrs = re.sub(rf'\b{name}="[^"]*"', f'{name}="{value}"', attrs)
                else:
                    attrs += f' {name}="{value}"'
            return f"<calcPr{attrs}/>"

        text, count = re.subn(r"<calcPr\b([^>]*)/>", update_calc_properties, text, count=1)
        if count == 0:
            text = text.replace(
                "</workbook>",
                '<calcPr calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"/></workbook>',
            )
        return text.encode("utf-8")
    return payload


def write_patched_workbook(
    source_path,
    output_path,
    updates,
    force=False,
    overwrite_formulas=False,
):
    source_path = source_path.resolve()
    output_path = output_path.resolve()
    if output_path.exists() and output_path != source_path and not force:
        raise FileExistsError(f"Output already exists; pass --force to replace it: {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(output_path.name + ".tmp")
    if temporary_path.exists():
        temporary_path.unlink()
    try:
        with zipfile.ZipFile(source_path, "r") as source_zip:
            parts = workbook_sheet_parts(source_zip)
            part_updates = {}
            for sheet_name, cells in updates.items():
                if sheet_name not in parts:
                    raise KeyError(f"Worksheet part not found: {sheet_name}")
                part_updates[parts[sheet_name]] = (sheet_name, cells)
            symbol_formula_repairs = {
                parts[sheet_name]: (sheet_name, 3 + reel_length)
                for _scene, _group_index, sheet_name, reel_length in SYMBOL_SHEET_GROUPS
            }
            with zipfile.ZipFile(temporary_path, "w") as output_zip:
                for info in source_zip.infolist():
                    if info.filename == "xl/calcChain.xml":
                        continue
                    payload = source_zip.read(info.filename)
                    if info.filename in part_updates:
                        sheet_name, cells = part_updates[info.filename]
                        payload = patch_sheet_xml(
                            payload,
                            sheet_name,
                            cells,
                            overwrite_formulas=overwrite_formulas,
                        )
                    if info.filename in symbol_formula_repairs:
                        sheet_name, last_row = symbol_formula_repairs[info.filename]
                        payload = restore_symbol_id_formulas(payload, sheet_name, last_row)
                    payload = remove_calc_chain_metadata(info.filename, payload)
                    output_zip.writestr(info, payload)
        os.replace(temporary_path, output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def verify_output(output_path, config):
    generated = build_config(output_path)
    changed_keys = [key for key, value in generated.items() if config.get(key) != value]
    extra_keys = [key for key in config if key not in generated]
    if changed_keys or extra_keys:
        raise ValueError(
            f"Round-trip verification failed. Changed keys: {changed_keys}; extra keys: {extra_keys}"
        )


def default_output_path(source_path, config_path):
    return source_path.with_name(f"{source_path.stem}_from_{config_path.stem}.xlsx")


def main():
    parser = argparse.ArgumentParser(
        description="Write H028 config_*.js data back into the mapped cells of an xlsx workbook"
    )
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, help="Output xlsx; defaults to *_from_config_*.xlsx")
    parser.add_argument("--in-place", action="store_true", help="Replace the source xlsx atomically")
    parser.add_argument(
        "--overwrite-formulas",
        action="store_true",
        help="Allow mapped formula cells to be replaced by config values (required with --in-place)",
    )
    parser.add_argument("--force", action="store_true", help="Replace an existing output file")
    parser.add_argument("--check", action="store_true", help="Report differences without writing a file")
    args = parser.parse_args()

    source_path = args.source.resolve()
    config_path = args.config.resolve()
    if not source_path.exists():
        raise FileNotFoundError(source_path)
    if not config_path.exists():
        raise FileNotFoundError(config_path)
    if args.in_place and args.output is not None:
        parser.error("--in-place and --output cannot be used together")
    if args.in_place and not args.overwrite_formulas:
        parser.error("--in-place requires --overwrite-formulas because mapped ranges contain formulas")

    config = load_js_config(config_path)
    updates = build_updates(source_path, config)
    changed = changed_cells(source_path, updates)
    print(f"Mapped cells: {sum(len(cells) for cells in updates.values()):,}")
    print(f"Changed cells: {len(changed):,}")
    for sheet_name, address, old_value, new_value, key in changed[:20]:
        print(f"  {sheet_name}!{address}: {old_value!r} -> {new_value!r} ({key})")
    if len(changed) > 20:
        print(f"  ... and {len(changed) - 20:,} more")
    if args.check:
        print("Check only; no xlsx was written.")
        return

    output_path = (
        source_path
        if args.in_place
        else args.output.resolve()
        if args.output is not None
        else default_output_path(source_path, config_path)
    )
    overwrite_formulas = args.overwrite_formulas or output_path != source_path
    if overwrite_formulas:
        print(
            "Mapped formula cells may be replaced by fixed config values; "
            "Symbol ID U:AA formulas are restored and Excel will rebuild the calculation chain."
        )
    write_patched_workbook(
        source_path,
        output_path,
        updates,
        force=args.force or args.in_place,
        overwrite_formulas=overwrite_formulas,
    )
    verify_output(output_path, config)
    print(f"Xlsx written and round-trip verified: {output_path}")


if __name__ == "__main__":
    main()
