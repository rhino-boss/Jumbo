"""H028 model <-> config sync (single entry point).

    model_sync.py export ...   xlsx  -> config_*.js
    model_sync.py import ...   config_*.js -> H0281.xlsx or an RTP workbook

The model is split in two: H0281.xlsx carries everything shared by every RTP
variant (pay table, reel strips, symbol weights, Parameter), and each
H0281<rtp><variant>.xlsx carries only that variant's version stamp and its
Multiplier_Weight card table.  Multiplier_Weight itself is derived from
Detail!K / Detail_Newbie!K.  Variant imports therefore preserve those formulas
and backfill their Fix Num plus cached results instead of replacing formulas
with constants.
"""
import argparse
import json
import math
import os
import posixpath
import re
import sys
import warnings
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape

from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, get_column_letter, range_boundaries

# ------------------------------------------------------------------------
# xlsx -> config
# ------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent
# The model is split in two: H0281.xlsx holds everything shared by every RTP
# variant (pay table, reel strips, symbol weights, Parameter), while each
# H0281<rtp><variant>.xlsx holds only that variant's version stamp and
# Multiplier_Weight card table.
BASE_WORKBOOK = BASE_DIR / "H0281.xlsx"
DEFAULT_SOURCE = BASE_DIR / "H028192A.xlsx"
DEFAULT_OUTPUT = BASE_DIR.parent / "config_92A.js"

METADATA = {
    "game_id": "101016",
    "parsheet_id": "H0281",
    "display_name": "Thunder Boost 1000",
    "game_name_zh": "雷神爆金1000",
    "default_coin_in": 100,
    "mode_normalbet": 0,
    "mode_extrabet": 1,
    "mode_featurebuy": 2,
    "normalbet": 1,
    "featurebuy": 75,
    "supported_bet_modes": [0, 2],
}

SYMBOL_SHEET_GROUPS = (
    ("BaseGame", 1, "BG_Symbol", 200),
    ("BaseGame", 2, "BG_Symbol (2)", 200),
    ("BaseGame", 3, "BF_Symbol", 200),
    ("FreeGame", 1, "FG_Symbol", 200),
    ("FreeGame", 2, "FG_Symbol (2)", 200),
    ("FreeGame", 3, "FG_Symbol (3)", 200),
)

# H028192A current layout. Keep this as the single source of truth for both
# xlsx_to_config.py and config_to_xlsx.py.
SYMBOL_SHEET_RANGES = {
    "MegaWay": "C33:H47",
    "MY": "C51:C63",
    "PostC1": "B67:C74",
}
DROP_START_ROWS = (4, 33, 62, 91, 120)

PARAMETER_RANGES = {
    "ReelWeight": "C5:C6",
    "FreeReelWeight": "C11:C13",
    "FreeTriggerReel": "C18:C20",
}


def extract_transposed(sheet, range_text):
    rows = [[cell.value for cell in row] for row in sheet[range_text]]
    columns = [list(column) for column in zip(*rows)]
    if len(columns) == 1:
        return columns[0]
    return columns


def extract_no_transpose(sheet, range_text):
    return [[cell.value for cell in row] for row in sheet[range_text]]


def extract_symbol_ids(sheet, range_text):
    symbol_to_id = {
        str(sheet.cell(row, 1).value): int(sheet.cell(row, 10).value)
        for row in range(4, 30)
        if sheet.cell(row, 1).value is not None and sheet.cell(row, 10).value is not None
    }
    symbol_reels = extract_transposed(sheet, range_text)
    result = []
    for reel_index, reel in enumerate(symbol_reels, start=1):
        try:
            result.append([symbol_to_id[str(symbol)] for symbol in reel])
        except KeyError as error:
            raise ValueError(
                f"Unknown symbol {error.args[0]!r} in {sheet.title} reel R{reel_index}"
            ) from error
    return result



def extract_linkpoint(overview):
    for row_index in range(1, overview.max_row + 1):
        if overview.cell(row_index, 1).value == "M1":
            return extract_no_transpose(overview, f"C{row_index}:F{row_index + 10}")
    raise ValueError("Pay-table start symbol M1 was not found in Overview")


def parse_card_range(label):
    if label is None:
        return None
    match = re.fullmatch(r"\(\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*\]", str(label).strip())
    if not match:
        return None
    return float(match.group(1)), float(match.group(2))


# Multiplier_Weight column header -> (profile, bet mode, segment).
# The sheet no longer carries a Newbie/Oldhand grouping row above the headers;
# the profile is encoded in the header itself via the _Newbie suffix.
# "oldhand" is the non-newbie (normal) profile, and only it has a buy feature.
CARD_COLUMNS = {
    "Weight_NB_BG_Newbie": ("newbie", "normal_bet", "weight_bg"),
    "Weight_NB_FG_Newbie": ("newbie", "normal_bet", "weight_fg"),
    "Weight_NB_BG": ("oldhand", "normal_bet", "weight_bg"),
    "Weight_NB_FG": ("oldhand", "normal_bet", "weight_fg"),
    "Weight_BF": ("oldhand", "buy_feature", "weight_fg"),
}

# card profile -> (worksheet, first row, last row)
CARD_DETAIL_RANGES = {
    ("newbie", "normal_bet", "weight_bg"): ("Detail_Newbie", 15, 79),
    ("newbie", "normal_bet", "weight_fg"): ("Detail_Newbie", 86, 149),
    ("oldhand", "normal_bet", "weight_bg"): ("Detail", 15, 79),
    ("oldhand", "normal_bet", "weight_fg"): ("Detail", 86, 149),
    ("oldhand", "buy_feature", "weight_fg"): ("Detail", 163, 226),
}


def parse_card_system(ws):
    profiles = {
        "newbie": {"normal_bet": {"weight_bg": [], "weight_fg": []}},
        "oldhand": {
            "normal_bet": {"weight_bg": [], "weight_fg": []},
            "buy_feature": {"weight_fg": []},
        },
    }
    header_row = next(
        (row for row in range(1, ws.max_row + 1)
         if str(ws.cell(row, 1).value).strip().lower() == "range"),
        None,
    )
    if header_row is None:
        raise ValueError(
            "Multiplier_Weight: no header row found (expected 'Range' in column A)"
        )
    profile_columns = {}
    for col in range(2, ws.max_column + 1):
        header = str(ws.cell(header_row, col).value or "").strip()
        if header in CARD_COLUMNS:
            profile_columns[col] = CARD_COLUMNS[header]
    # A silent miss here used to emit a config with empty card tables, which the
    # simulator happily accepts and then runs unweighted. Fail loudly instead.
    missing = set(CARD_COLUMNS.values()) - set(profile_columns.values())
    if missing:
        found = [str(ws.cell(header_row, c).value) for c in range(2, ws.max_column + 1)]
        raise ValueError(
            f"Multiplier_Weight: missing weight column(s) {sorted(missing)}. "
            f"Headers found on row {header_row}: {found}"
        )

    row = header_row + 1
    while ws.cell(row, 1).value is not None:
        label = str(ws.cell(row, 1).value).strip()
        range_pair = parse_card_range(label)
        for col, (player, mode, segment) in profile_columns.items():
            weight = int(ws.cell(row, col).value or 0)
            if label.lower() == "free game":
                card = {"type": "free_game", "weight": weight}
            elif range_pair is not None:
                card = {"type": "range", "min": range_pair[0], "max": range_pair[1], "weight": weight}
            else:
                continue
            profiles[player][mode][segment].append(card)
        row += 1
    return {"enabled": True, "retry_limit": 5000, **profiles}


def build_config(source_path, base_path=None):
    """source_path = the RTP variant workbook (version + Multiplier_Weight).
    base_path    = H0281.xlsx, holding everything shared across variants."""
    base_path = Path(base_path) if base_path is not None else BASE_WORKBOOK
    if not base_path.exists():
        raise FileNotFoundError(f"Base workbook not found: {base_path}")

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    base = load_workbook(base_path, read_only=True, data_only=True)
    output = dict(METADATA)

    excel_version = workbook["Overview"]["B3"].value
    if excel_version is None or not str(excel_version).strip():
        workbook.close(); base.close()
        raise ValueError("Overview!B3 (Version) is empty")
    output["excel_version"] = str(excel_version).strip()
    output["linkpoint"] = extract_linkpoint(base["Overview"])
    output["card_system"] = parse_card_system(workbook["Multiplier_Weight"])

    for scene, group_index, sheet_name, reel_length in SYMBOL_SHEET_GROUPS:
        sheet = base[sheet_name]
        last_row = 3 + reel_length
        output[f"{scene}Symbol{group_index}"] = extract_symbol_ids(sheet, f"M4:S{last_row}")
        output[f"{scene}SymbolWeight{group_index}"] = extract_transposed(sheet, f"AC4:AI{last_row}")
        for field, range_text in SYMBOL_SHEET_RANGES.items():
            if field == "PostC1":
                key = f"{scene}{group_index}PostC1"
            else:
                key = f"{scene}{field}{group_index}"
            output[key] = extract_transposed(sheet, range_text)
        for drop_index, first_row in enumerate(DROP_START_ROWS, start=1):
            key = f"{scene}{group_index}Drop{drop_index}"
            output[key] = extract_transposed(sheet, f"AL{first_row}:AR{first_row + 25}")

    parameter = base["Parameter"]
    for key, range_text in PARAMETER_RANGES.items():
        output[key] = extract_transposed(parameter, range_text)

    workbook.close()
    base.close()
    return output


def load_js_config(path):
    text = path.read_text(encoding="utf-8").strip()
    if not text.startswith("const data = "):
        raise ValueError(f"Unsupported config header: {path}")
    return json.loads(text[len("const data = ") :].rstrip(";"))


def write_js_config(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "const data = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )


def derive_output_path(source_path):
    match = re.fullmatch(r"H0281(?P<rtp>\d{2})(?P<variant>[A-Za-z0-9_-]*)", source_path.stem, re.IGNORECASE)
    if not match:
        raise ValueError(
            f"Cannot derive config name from {source_path.name}; expected H0281 + two-digit RTP + optional variant, "
            "for example H028192A.xlsx"
        )
    return DEFAULT_OUTPUT.parent / f"config_{match.group('rtp')}{match.group('variant')}.js"


def compare_config(generated, output_path):
    current = load_js_config(output_path)
    changed_keys = [key for key in generated if generated[key] != current.get(key)]
    extra_keys = [key for key in current if key not in generated]
    if changed_keys or extra_keys:
        raise ValueError(f"Config mismatch for {output_path}. Changed keys: {changed_keys}; extra keys: {extra_keys}")


def process_source(source_path, output_path, check=False, base_path=None):
    generated = build_config(source_path, base_path)
    if check:
        compare_config(generated, output_path)
        print(f"Config is current: {output_path}")
    else:
        write_js_config(output_path, generated)
        compare_config(generated, output_path)
        print(f"Config written: {output_path} <- {source_path.name}")
    return generated


# ------------------------------------------------------------------------
# config -> xlsx
# ------------------------------------------------------------------------
DEFAULT_BASE_XLSX = BASE_DIR / "H0281.xlsx"
DEFAULT_VARIANT = BASE_DIR / "H028192A.xlsx"
DEFAULT_CONFIG = BASE_DIR.parent / "config_92A.js"

CELL_PATTERN = re.compile(
    r'<c(?P<selfattrs>[^>]*?\br="(?P<selfref>[A-Z]{1,3}[1-9][0-9]*)"[^>]*?)\s*/>'
    r'|<c(?P<attrs>[^>]*?\br="(?P<ref>[A-Z]{1,3}[1-9][0-9]*)"[^>]*?)>'
    r'(?P<body>.*?)</c>',
    re.DOTALL,
)
ROW_PATTERN = re.compile(
    r'<row(?P<attrs>[^>]*\br="(?P<row>[1-9][0-9]*)"[^>]*)>'
    r'(?P<body>.*?)</row>',
    re.DOTALL,
)




def require_key(config, key):
    if key not in config:
        raise KeyError(f"Config key is missing: {key}")
    return config[key]


def cell_addresses(range_text):
    min_col, min_row, max_col, max_row = range_boundaries(range_text)
    return [
        [f"{get_column_letter(col)}{row}" for row in range(min_row, max_row + 1)]
        for col in range(min_col, max_col + 1)
    ]


def add_update(updates, sheet_name, address, value, key):
    sheet_updates = updates.setdefault(sheet_name, {})
    if address in sheet_updates and sheet_updates[address][0] != value:
        previous_value, previous_key = sheet_updates[address]
        raise ValueError(
            f"Conflicting writes for {sheet_name}!{address}: "
            f"{previous_key}={previous_value!r}, {key}={value!r}"
        )
    sheet_updates[address] = (value, key)


def add_transposed_range(updates, sheet_name, range_text, values, key):
    columns = cell_addresses(range_text)
    if len(columns) == 1:
        if not isinstance(values, list) or len(values) != len(columns[0]):
            raise ValueError(f"{key} must contain {len(columns[0])} values")
        value_columns = [values]
    else:
        if not isinstance(values, list) or len(values) != len(columns):
            raise ValueError(f"{key} must contain {len(columns)} columns")
        value_columns = values
    for column_index, addresses in enumerate(columns):
        column_values = value_columns[column_index]
        if not isinstance(column_values, list) or len(column_values) != len(addresses):
            raise ValueError(
                f"{key} column {column_index + 1} must contain {len(addresses)} values"
            )
        for address, value in zip(addresses, column_values):
            add_update(updates, sheet_name, address, value, key)


def add_row_major_range(updates, sheet_name, range_text, values, key):
    columns = cell_addresses(range_text)
    row_count = len(columns[0])
    column_count = len(columns)
    if not isinstance(values, list) or len(values) != row_count:
        raise ValueError(f"{key} must contain {row_count} rows")
    for row_index in range(row_count):
        row_values = values[row_index]
        if not isinstance(row_values, list) or len(row_values) != column_count:
            raise ValueError(f"{key} row {row_index + 1} must contain {column_count} values")
        for column_index in range(column_count):
            add_update(
                updates,
                sheet_name,
                columns[column_index][row_index],
                row_values[column_index],
                key,
            )





def discover_layout(source_path):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        missing_symbol_sheets = [
            sheet_name for _scene, _group_index, sheet_name, _reel_length in SYMBOL_SHEET_GROUPS
            if sheet_name not in workbook.sheetnames
        ]
        if missing_symbol_sheets:
            raise KeyError(f"Missing symbol worksheets: {missing_symbol_sheets}")
        symbol_by_id = {}
        for _scene, _group_index, sheet_name, _reel_length in SYMBOL_SHEET_GROUPS:
            sheet = workbook[sheet_name]
            symbol_by_id[sheet_name] = {
                int(sheet.cell(row, 10).value): str(sheet.cell(row, 1).value)
                for row in range(4, 30)
                if sheet.cell(row, 1).value is not None and sheet.cell(row, 10).value is not None
            }
        overview = workbook["Overview"]
        linkpoint_row = next(
            row for row in range(1, overview.max_row + 1)
            if overview.cell(row, 1).value == "M1"
        )

        # Multiplier_Weight lives in the RTP-variant workbooks and is fully derived
        # (its columns are formulas reading Detail!K / Detail_Newbie!K), so this
        # direction neither reads nor writes it.
        return {
            "linkpoint_row": linkpoint_row,
            "symbol_by_id": symbol_by_id,
        }
    finally:
        workbook.close()



def build_base_updates(source_path, config):
    layout = discover_layout(source_path)
    updates = {}
    start_row = layout["linkpoint_row"]
    add_row_major_range(
        updates,
        "Overview",
        f"C{start_row}:F{start_row + 10}",
        require_key(config, "linkpoint"),
        "linkpoint",
    )
    for scene, group_index, sheet_name, reel_length in SYMBOL_SHEET_GROUPS:
        last_row = 3 + reel_length
        symbol_key = f"{scene}Symbol{group_index}"
        symbol_map = layout["symbol_by_id"][sheet_name]
        symbol_values = []
        for reel_index, reel in enumerate(require_key(config, symbol_key), start=1):
            try:
                symbol_values.append([symbol_map[int(symbol_id)] for symbol_id in reel])
            except (KeyError, TypeError, ValueError) as error:
                raise ValueError(
                    f"{symbol_key} reel R{reel_index} contains an unknown Symbol ID"
                ) from error
        add_transposed_range(
            updates,
            sheet_name,
            f"M4:S{last_row}",
            symbol_values,
            symbol_key,
        )
        weight_key = f"{scene}SymbolWeight{group_index}"
        add_transposed_range(
            updates,
            sheet_name,
            f"AC4:AI{last_row}",
            require_key(config, weight_key),
            weight_key,
        )
        for field, range_text in SYMBOL_SHEET_RANGES.items():
            if field == "PostC1":
                key = f"{scene}{group_index}PostC1"
            else:
                key = f"{scene}{field}{group_index}"
            add_transposed_range(
                updates,
                sheet_name,
                range_text,
                require_key(config, key),
                key,
            )
        for drop_index, first_row in enumerate(DROP_START_ROWS, start=1):
            key = f"{scene}{group_index}Drop{drop_index}"
            add_transposed_range(
                updates,
                sheet_name,
                f"AL{first_row}:AR{first_row + 25}",
                require_key(config, key),
                key,
            )
    for key, range_text in PARAMETER_RANGES.items():
        add_transposed_range(
            updates,
            "Parameter",
            range_text,
            require_key(config, key),
            key,
        )
    return updates


def is_variant_workbook(path):
    return re.fullmatch(
        r"H0281\d{2}[A-Za-z0-9_-]+\.xlsx",
        Path(path).name,
        re.IGNORECASE,
    ) is not None


def config_cards(config, profile, mode, segment):
    try:
        cards = config["card_system"][profile][mode][segment]
    except KeyError as error:
        raise KeyError(
            f"Config card table is missing: card_system.{profile}.{mode}.{segment}"
        ) from error
    if not isinstance(cards, list) or not cards:
        raise ValueError(f"Card table is empty: {profile}.{mode}.{segment}")
    return cards


def card_label(card):
    if card.get("type") == "free_game":
        return "Free Game"
    if card.get("type") != "range":
        raise ValueError(f"Unsupported card type: {card.get('type')!r}")
    def format_number(value):
        number = float(value)
        return str(int(number)) if number.is_integer() else format(number, ".15g")

    return f"({format_number(card['min'])}, {format_number(card['max'])}]"


def build_variant_updates(source_path, config):
    """Backfill version/cards while preserving the variant workbook formulas."""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        required = {"Overview", "Multiplier_Weight", "Detail", "Detail_Newbie"}
        missing = required - set(workbook.sheetnames)
        if missing:
            raise KeyError(f"Variant workbook is missing worksheets: {sorted(missing)}")

        updates = {}
        add_update(
            updates,
            "Overview",
            "B3",
            str(require_key(config, "excel_version")),
            "excel_version",
        )

        multiplier = workbook["Multiplier_Weight"]
        header_row = next(
            (row for row in range(1, multiplier.max_row + 1)
             if str(multiplier.cell(row, 1).value or "").strip().lower() == "range"),
            None,
        )
        if header_row is None:
            raise ValueError("Multiplier_Weight: Range header not found")
        header_columns = {
            str(multiplier.cell(header_row, col).value or "").strip(): col
            for col in range(2, multiplier.max_column + 1)
        }

        profile_metrics = {}
        for header, profile_key in CARD_COLUMNS.items():
            if header not in header_columns:
                raise ValueError(f"Multiplier_Weight column is missing: {header}")
            profile, mode, segment = profile_key
            detail_name, first_row, last_row = CARD_DETAIL_RANGES[profile_key]
            cards = config_cards(config, profile, mode, segment)
            expected = last_row - first_row + 1
            detail_cards = cards[:expected]
            placeholder_cards = cards[expected:]
            if len(detail_cards) != expected or any(
                card.get("type") != "free_game" or int(card.get("weight", 0) or 0) != 0
                for card in placeholder_cards
            ):
                raise ValueError(
                    f"{profile}.{mode}.{segment} has {len(cards)} cards; expected "
                    f"{expected} Detail cards plus optional zero-weight Free Game placeholder"
                )
            weights = [int(card.get("weight", 0) or 0) for card in cards]
            if any(weight < 0 for weight in weights):
                raise ValueError(f"Negative card weight in {profile}.{mode}.{segment}")
            total_weight = sum(weights)
            if total_weight <= 0:
                raise ValueError(f"No positive card weight in {profile}.{mode}.{segment}")

            detail = workbook[detail_name]
            rtp = 0.0
            final_rates = []
            for index, (card, weight) in enumerate(zip(detail_cards, weights[:expected])):
                row = first_row + index
                expected_label = card_label(card)
                actual_label = str(detail.cell(row, 7).value or "").strip()
                if actual_label != expected_label:
                    raise ValueError(
                        f"{detail_name}!G{row} is {actual_label!r}; expected {expected_label!r}"
                    )
                natural_rate = float(detail.cell(row, 4).value or 0)
                average_multiplier = float(detail.cell(row, 5).value or 0)
                final_rate = weight / total_weight
                if weight > 0 and natural_rate <= 0:
                    raise ValueError(
                        f"Positive card weight has zero natural rate: {detail_name}!G{row}"
                    )
                fix_num = final_rate / natural_rate if natural_rate > 0 else 0.0
                rtp_value = final_rate * average_multiplier
                final_rates.append(final_rate)
                rtp += rtp_value
                for column, value, key_suffix in (
                    ("H", fix_num, "Fix Num"),
                    ("I", final_rate, "Fix Rate"),
                    ("J", final_rate, "Final Rate"),
                    ("K", weight, "Weight"),
                    ("L", final_rate, "Hit Rate"),
                    ("M", rtp_value, "Simulator RTP"),
                ):
                    add_update(
                        updates,
                        detail_name,
                        f"{column}{row}",
                        value,
                        f"{header} {key_suffix}",
                    )

            for index, (card, weight) in enumerate(zip(cards, weights)):
                expected_label = card_label(card)
                multiplier_row = header_row + 1 + index
                multiplier_label = str(multiplier.cell(multiplier_row, 1).value or "").strip()
                if multiplier_label != expected_label:
                    raise ValueError(
                        f"Multiplier_Weight!A{multiplier_row} is {multiplier_label!r}; "
                        f"expected {expected_label!r}"
                    )
                add_update(
                    updates,
                    "Multiplier_Weight",
                    f"{get_column_letter(header_columns[header])}{multiplier_row}",
                    weight,
                    header,
                )
            profile_metrics[profile_key] = {
                "rtp": rtp,
                "final_rates": final_rates,
            }

        old_bg = profile_metrics[("oldhand", "normal_bet", "weight_bg")]
        old_fg = profile_metrics[("oldhand", "normal_bet", "weight_fg")]
        newbie_bg = profile_metrics[("newbie", "normal_bet", "weight_bg")]
        newbie_fg = profile_metrics[("newbie", "normal_bet", "weight_fg")]
        buy_fg = profile_metrics[("oldhand", "buy_feature", "weight_fg")]
        old_trigger = old_bg["final_rates"][-1]
        newbie_trigger = newbie_bg["final_rates"][-1]

        summary_updates = (
            ("Detail", "C7", old_bg["rtp"], "Weight_NB_BG RTP"),
            ("Detail", "D7", old_fg["rtp"] * old_trigger, "Weight_NB_FG RTP"),
            ("Detail", "E7", old_bg["rtp"] + old_fg["rtp"] * old_trigger, "Weight_NB total RTP"),
            ("Detail", "F7", old_trigger, "Weight_NB FG trigger"),
            ("Detail", "G7", 1 / old_trigger, "Weight_NB FG period"),
            ("Detail", "H7", old_fg["rtp"], "Weight_NB FG average"),
            ("Detail", "C8", 0.0, "Weight_BF BG RTP"),
            ("Detail", "D8", buy_fg["rtp"] / 75.0, "Weight_BF FG RTP"),
            ("Detail", "E8", buy_fg["rtp"] / 75.0, "Weight_BF total RTP"),
            ("Detail", "F8", 1.0, "Weight_BF trigger"),
            ("Detail", "G8", 1.0, "Weight_BF period"),
            ("Detail", "H8", buy_fg["rtp"], "Weight_BF average"),
            ("Detail_Newbie", "C7", newbie_bg["rtp"], "Weight_NB_Newbie BG RTP"),
            ("Detail_Newbie", "D7", newbie_fg["rtp"] * newbie_trigger, "Weight_NB_Newbie FG RTP"),
            ("Detail_Newbie", "E7", newbie_bg["rtp"] + newbie_fg["rtp"] * newbie_trigger, "Weight_NB_Newbie total RTP"),
            ("Detail_Newbie", "F7", newbie_trigger, "Weight_NB_Newbie FG trigger"),
            ("Detail_Newbie", "G7", 1 / newbie_trigger, "Weight_NB_Newbie FG period"),
            ("Detail_Newbie", "H7", newbie_fg["rtp"], "Weight_NB_Newbie FG average"),
        )
        for sheet_name, address, value, key in summary_updates:
            add_update(updates, sheet_name, address, value, key)
        return updates
    finally:
        workbook.close()


def build_updates(source_path, config):
    if is_variant_workbook(source_path):
        return build_variant_updates(source_path, config)
    return build_base_updates(source_path, config)


def values_equal(left, right):
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return math.isclose(float(left), float(right), rel_tol=1e-12, abs_tol=1e-12)
    return left == right


def changed_cells(source_path, updates):
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        workbook = load_workbook(source_path, read_only=False, data_only=True)
    try:
        changed = []
        for sheet_name, cells in updates.items():
            sheet = workbook[sheet_name]
            for address, (new_value, key) in cells.items():
                old_value = sheet[address].value
                if not values_equal(old_value, new_value):
                    changed.append((sheet_name, address, old_value, new_value, key))
        return changed
    finally:
        workbook.close()


def workbook_sheet_parts(archive):
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    rels_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels_root
    }
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    parts = {}
    for sheet in workbook_root.findall(f".//{{{main_ns}}}sheet"):
        target = rel_targets[sheet.attrib[f"{{{rel_ns}}}id"]]
        if target.startswith("/"):
            part = target.lstrip("/")
        else:
            part = posixpath.normpath(posixpath.join("xl", target))
        parts[sheet.attrib["name"]] = part
    return parts


def number_text(value):
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ValueError(f"Non-finite numeric value cannot be written to xlsx: {value}")
        return format(value, ".15g")
    raise TypeError(f"Unsupported numeric value: {value!r}")


def render_cell_xml(attrs, value):
    attrs = re.sub(r'\s+t="[^"]*"', "", attrs.rstrip())
    if value is None:
        return f"<c{attrs}></c>"
    if isinstance(value, str):
        space = ' xml:space="preserve"' if value != value.strip() else ""
        return f'<c{attrs} t="inlineStr"><is><t{space}>{escape(value)}</t></is></c>'
    return f"<c{attrs}><v>{number_text(value)}</v></c>"


def insert_missing_cells(sheet_xml, sheet_name, remaining, cell_updates):
    by_row = {}
    for address in remaining:
        match = re.fullmatch(r"([A-Z]{1,3})([1-9][0-9]*)", address)
        by_row.setdefault(int(match.group(2)), []).append(address)
    found_rows = set()

    def replace_row(match):
        row_number = int(match.group("row"))
        if row_number not in by_row:
            return match.group(0)
        found_rows.add(row_number)
        body = match.group("body")
        addresses = sorted(
            by_row[row_number],
            key=lambda address: column_index_from_string(re.match(r"[A-Z]+", address).group(0)),
        )
        for address in addresses:
            new_col = column_index_from_string(re.match(r"[A-Z]+", address).group(0))
            insert_at = len(body)
            for cell_match in CELL_PATTERN.finditer(body):
                existing_address = cell_match.group("ref") or cell_match.group("selfref")
                existing_col = column_index_from_string(
                    re.match(r"[A-Z]+", existing_address).group(0)
                )
                if existing_col > new_col:
                    insert_at = cell_match.start()
                    break
            value, _key = cell_updates[address]
            cell_xml = render_cell_xml(f' r="{address}"', value)
            body = body[:insert_at] + cell_xml + body[insert_at:]
            remaining.discard(address)
        return f'<row{match.group("attrs")}>{body}</row>'

    patched = ROW_PATTERN.sub(replace_row, sheet_xml)
    missing_rows = sorted(set(by_row) - found_rows)
    if missing_rows:
        sheet_data_pattern = re.compile(
            r"(?P<open><sheetData[^>]*>)(?P<body>.*?)(?P<close></sheetData>)",
            re.DOTALL,
        )

        def add_rows(match):
            body = match.group("body")
            for row_number in missing_rows:
                addresses = sorted(
                    by_row[row_number],
                    key=lambda address: column_index_from_string(
                        re.match(r"[A-Z]+", address).group(0)
                    ),
                )
                cells = []
                for address in addresses:
                    value, _key = cell_updates[address]
                    cells.append(render_cell_xml(f' r="{address}"', value))
                    remaining.discard(address)
                row_xml = f'<row r="{row_number}">{"".join(cells)}</row>'
                insert_at = len(body)
                for row_match in ROW_PATTERN.finditer(body):
                    if int(row_match.group("row")) > row_number:
                        insert_at = row_match.start()
                        break
                body = body[:insert_at] + row_xml + body[insert_at:]
            return match.group("open") + body + match.group("close")

        patched, count = sheet_data_pattern.subn(add_rows, patched, count=1)
        if count != 1:
            raise ValueError(f"sheetData element not found in {sheet_name}")
    return patched


def patch_sheet_xml(
    xml_bytes,
    sheet_name,
    cell_updates,
    overwrite_formulas=False,
    preserve_formula_cache=False,
):
    text = xml_bytes.decode("utf-8")
    remaining = set(cell_updates)

    def replace_cell(match):
        address = match.group("ref") or match.group("selfref")
        if address not in cell_updates:
            return match.group(0)
        body = match.group("body") or ""
        has_formula = re.search(r"<f(?:\s|>)", body) is not None
        if has_formula and preserve_formula_cache:
            value, _key = cell_updates[address]
            if not isinstance(value, (int, float)) or isinstance(value, bool):
                raise TypeError(
                    f"Formula cache must be numeric: {sheet_name}!{address}={value!r}"
                )
            attrs = match.group("attrs") or match.group("selfattrs") or ""
            attrs = re.sub(r'\s+t="[^"]*"', "", attrs.rstrip())
            value_xml = f"<v>{number_text(value)}</v>"
            if re.search(r"<v(?:\s[^>]*)?>.*?</v>", body, re.DOTALL):
                body = re.sub(
                    r"<v(?:\s[^>]*)?>.*?</v>",
                    value_xml,
                    body,
                    count=1,
                    flags=re.DOTALL,
                )
            else:
                formula_end = body.find("</f>")
                if formula_end < 0:
                    raise ValueError(f"Malformed formula cell {sheet_name}!{address}")
                formula_end += len("</f>")
                body = body[:formula_end] + value_xml + body[formula_end:]
            remaining.discard(address)
            return f"<c{attrs}>{body}</c>"
        if has_formula and not overwrite_formulas:
            raise ValueError(f"Refusing to overwrite formula cell {sheet_name}!{address}")
        attrs = match.group("attrs") or match.group("selfattrs") or ""
        value, _key = cell_updates[address]
        remaining.discard(address)
        return render_cell_xml(attrs, value)

    patched = CELL_PATTERN.sub(replace_cell, text)
    if remaining:
        patched = insert_missing_cells(patched, sheet_name, remaining, cell_updates)
    if remaining:
        sample = ", ".join(sorted(remaining)[:10])
        raise ValueError(f"Cells could not be inserted in {sheet_name}: {sample}")
    payload = patched.encode("utf-8")
    try:
        ET.fromstring(payload)
    except ET.ParseError as error:
        raise ValueError(f"Invalid worksheet XML after patching {sheet_name}: {error}") from error
    return payload


def restore_symbol_id_formulas(xml_bytes, sheet_name, last_row):
    """Restore U:AA formulas after older tool versions replaced them with values."""
    text = xml_bytes.decode("utf-8")
    restored = set()

    def replace_cell(match):
        address = match.group("ref") or match.group("selfref")
        column_match = re.fullmatch(r"([A-Z]+)([1-9][0-9]*)", address)
        column = column_index_from_string(column_match.group(1))
        row = int(column_match.group(2))
        if not (column_index_from_string("U") <= column <= column_index_from_string("AA")):
            return match.group(0)
        if not (4 <= row <= last_row):
            return match.group(0)
        source_column = get_column_letter(column - 8)
        attrs = match.group("attrs") or match.group("selfattrs") or ""
        attrs = re.sub(r'\s+t="[^"]*"', "", attrs.rstrip())
        restored.add(address)
        return (
            f"<c{attrs}><f>VLOOKUP({source_column}{row}, "
            "$A$4:$J$29, 10, 0)</f></c>"
        )

    patched = CELL_PATTERN.sub(replace_cell, text)

    missing_by_row = {}
    for row in range(4, last_row + 1):
        for column in range(column_index_from_string("U"), column_index_from_string("AA") + 1):
            address = f"{get_column_letter(column)}{row}"
            if address not in restored:
                missing_by_row.setdefault(row, []).append(address)

    def insert_formula_cells(match):
        row = int(match.group("row"))
        if row not in missing_by_row:
            return match.group(0)
        body = match.group("body")
        for address in missing_by_row[row]:
            column = column_index_from_string(re.match(r"[A-Z]+", address).group(0))
            source_column = get_column_letter(column - 8)
            insert_at = len(body)
            for cell_match in CELL_PATTERN.finditer(body):
                existing_address = cell_match.group("ref") or cell_match.group("selfref")
                existing_column = column_index_from_string(
                    re.match(r"[A-Z]+", existing_address).group(0)
                )
                if existing_column > column:
                    insert_at = cell_match.start()
                    break
            formula_xml = (
                f'<c r="{address}"><f>VLOOKUP({source_column}{row}, '
                "$A$4:$J$29, 10, 0)</f></c>"
            )
            body = body[:insert_at] + formula_xml + body[insert_at:]
            restored.add(address)
        return f'<row{match.group("attrs")}>{body}</row>'

    if missing_by_row:
        patched = ROW_PATTERN.sub(insert_formula_cells, patched)
    expected = 7 * (last_row - 3)
    if len(restored) != expected:
        raise ValueError(
            f"Could not restore every Symbol ID formula in {sheet_name}: "
            f"restored {len(restored)} of {expected}"
        )
    payload = patched.encode("utf-8")
    ET.fromstring(payload)
    return payload


def remove_calc_chain_metadata(filename, payload):
    if filename == "xl/_rels/workbook.xml.rels":
        text = payload.decode("utf-8")
        text = re.sub(
            r'<Relationship\b[^>]*\bType="[^"]*/calcChain"[^>]*/>',
            "",
            text,
        )
        return text.encode("utf-8")
    if filename == "[Content_Types].xml":
        text = payload.decode("utf-8")
        text = re.sub(
            r'<Override\b[^>]*\bPartName="/xl/calcChain\.xml"[^>]*/>',
            "",
            text,
        )
        return text.encode("utf-8")
    if filename == "xl/workbook.xml":
        text = payload.decode("utf-8")

        def update_calc_properties(match):
            attrs = match.group(1)
            for name, value in (
                ("calcMode", "auto"),
                ("fullCalcOnLoad", "1"),
                ("forceFullCalc", "1"),
            ):
                if re.search(rf'\b{name}="[^"]*"', attrs):
                    attrs = re.sub(rf'\b{name}="[^"]*"', f'{name}="{value}"', attrs)
                else:
                    attrs += f' {name}="{value}"'
            return f"<calcPr{attrs}/>"

        text, count = re.subn(r"<calcPr\b([^>]*)/>", update_calc_properties, text, count=1)
        if count == 0:
            text = text.replace(
                "</workbook>",
                '<calcPr calcMode="auto" fullCalcOnLoad="1" forceFullCalc="1"/></workbook>',
            )
        return text.encode("utf-8")
    return payload


def write_patched_workbook(
    source_path,
    output_path,
    updates,
    force=False,
    overwrite_formulas=False,
    preserve_formula_cache=False,
):
    source_path = source_path.resolve()
    output_path = output_path.resolve()
    if output_path.exists() and output_path != source_path and not force:
        raise FileExistsError(f"Output already exists; pass --force to replace it: {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(output_path.name + ".tmp")
    if temporary_path.exists():
        temporary_path.unlink()
    try:
        with zipfile.ZipFile(source_path, "r") as source_zip:
            parts = workbook_sheet_parts(source_zip)
            part_updates = {}
            for sheet_name, cells in updates.items():
                if sheet_name not in parts:
                    raise KeyError(f"Worksheet part not found: {sheet_name}")
                part_updates[parts[sheet_name]] = (sheet_name, cells)
            symbol_formula_repairs = {
                parts[sheet_name]: (sheet_name, 3 + reel_length)
                for _scene, _group_index, sheet_name, reel_length in SYMBOL_SHEET_GROUPS
                if sheet_name in parts
            }
            with zipfile.ZipFile(temporary_path, "w") as output_zip:
                for info in source_zip.infolist():
                    if info.filename == "xl/calcChain.xml":
                        continue
                    payload = source_zip.read(info.filename)
                    if info.filename in part_updates:
                        sheet_name, cells = part_updates[info.filename]
                        payload = patch_sheet_xml(
                            payload,
                            sheet_name,
                            cells,
                            overwrite_formulas=overwrite_formulas,
                            preserve_formula_cache=preserve_formula_cache,
                        )
                    if info.filename in symbol_formula_repairs:
                        sheet_name, last_row = symbol_formula_repairs[info.filename]
                        payload = restore_symbol_id_formulas(payload, sheet_name, last_row)
                    payload = remove_calc_chain_metadata(info.filename, payload)
                    output_zip.writestr(info, payload)
        os.replace(temporary_path, output_path)
    finally:
        if temporary_path.exists():
            temporary_path.unlink()


def verify_output(output_path, config, variant_path=None):
    if is_variant_workbook(output_path):
        generated = build_config(output_path, base_path=BASE_WORKBOOK)
        ignored = set()
    else:
        # Shared workbook verification deliberately ignores the variant-owned
        # version/card fields.
        generated = build_config(variant_path or DEFAULT_VARIANT, base_path=output_path)
        ignored = {"excel_version", "card_system"}
    changed_keys = [
        key for key, value in generated.items()
        if key not in ignored and config.get(key) != value
    ]
    extra_keys = [key for key in config if key not in generated and key not in ignored]
    if changed_keys or extra_keys:
        raise ValueError(
            f"Round-trip verification failed. Changed keys: {changed_keys}; extra keys: {extra_keys}"
        )


def default_output_path(source_path, config_path):
    return source_path.with_name(f"{source_path.stem}_from_{config_path.stem}.xlsx")


def run_export(argv):
    parser = argparse.ArgumentParser(description="Build H028 config files from xlsx sources")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--auto-output", action="store_true", help="Derive config_RTPVARIANT.js from the xlsx name")
    parser.add_argument("--all", action="store_true", help="Convert every H0281*.xlsx file in the Source folder")
    parser.add_argument(
        "--sync-default",
        action="store_true",
        help=f"Also update {DEFAULT_OUTPUT.name} when converting {DEFAULT_SOURCE.name}",
    )
    parser.add_argument("--check", action="store_true", help="Compare generated data with the existing output")
    parser.add_argument(
        "--base", type=Path, default=BASE_WORKBOOK,
        help=f"Shared workbook holding pay table / reels / Parameter (default {BASE_WORKBOOK.name})",
    )
    args = parser.parse_args(argv)
    base_path = args.base.resolve()

    if args.all:
        # H0281.xlsx is the shared base, not an RTP variant -> never a config source.
        sources = sorted(
            path for path in BASE_DIR.glob("H0281*.xlsx")
            if not path.name.startswith("~$")
            and path.resolve() != base_path.resolve()
        )
        if not sources:
            raise FileNotFoundError(f"No H0281*.xlsx files found in {BASE_DIR}")
        for source_path in sources:
            output_path = derive_output_path(source_path)
            generated = process_source(source_path, output_path, args.check, base_path)
            if args.sync_default and source_path.name.casefold() == DEFAULT_SOURCE.name.casefold():
                if args.check:
                    compare_config(generated, DEFAULT_OUTPUT)
                    print(f"Config is current: {DEFAULT_OUTPUT}")
                else:
                    write_js_config(DEFAULT_OUTPUT, generated)
                    compare_config(generated, DEFAULT_OUTPUT)
                    print(f"Default config written: {DEFAULT_OUTPUT} <- {source_path.name}")
        print(f"Processed {len(sources)} xlsx file(s).")
        return

    source_path = args.source.resolve()
    output_path = (
        args.output.resolve()
        if args.output is not None
        else derive_output_path(source_path)
        if args.auto_output
        else DEFAULT_OUTPUT
    )
    generated = process_source(source_path, output_path, args.check, base_path)
    if args.sync_default and source_path.name.casefold() == DEFAULT_SOURCE.name.casefold() and output_path != DEFAULT_OUTPUT:
        if args.check:
            compare_config(generated, DEFAULT_OUTPUT)
            print(f"Config is current: {DEFAULT_OUTPUT}")
        else:
            write_js_config(DEFAULT_OUTPUT, generated)
            compare_config(generated, DEFAULT_OUTPUT)
            print(f"Default config written: {DEFAULT_OUTPUT} <- {source_path.name}")


def run_import(argv):
    parser = argparse.ArgumentParser(
        description=(
            "Write H028 config_*.js data back into H0281.xlsx or an "
            "H0281<RTP><variant>.xlsx workbook"
        )
    )
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--source", type=Path, default=DEFAULT_BASE_XLSX)
    parser.add_argument(
        "--all-variants",
        action="store_true",
        help="Backfill every existing H0281<RTP><variant>.xlsx from its config_*.js",
    )
    parser.add_argument("--output", type=Path, help="Output xlsx; defaults to *_from_config_*.xlsx")
    parser.add_argument("--in-place", action="store_true", help="Replace the source xlsx atomically")
    parser.add_argument(
        "--overwrite-formulas",
        action="store_true",
        help="Allow mapped formula cells to be replaced by config values (required with --in-place)",
    )
    parser.add_argument("--force", action="store_true", help="Replace an existing output file")
    parser.add_argument("--check", action="store_true", help="Report differences without writing a file")
    args = parser.parse_args(argv)

    if args.all_variants:
        if args.output is not None:
            parser.error("--all-variants cannot be combined with --output")
        variants = sorted(
            path for path in BASE_DIR.glob("H0281*.xlsx")
            if not path.name.startswith("~$")
            and path.resolve() != BASE_WORKBOOK.resolve()
            and is_variant_workbook(path)
        )
        if not variants:
            raise FileNotFoundError(f"No RTP variant workbooks found in {BASE_DIR}")
        for variant in variants:
            config_path = derive_output_path(variant)
            if not config_path.exists():
                raise FileNotFoundError(
                    f"Config for {variant.name} was not found: {config_path}"
                )
            child_args = [
                "--config", str(config_path),
                "--source", str(variant),
                "--in-place",
            ]
            if args.check:
                child_args.append("--check")
            run_import(child_args)
        print(f"Processed {len(variants)} RTP variant workbook(s).")
        return

    source_path = args.source.resolve()
    config_path = args.config.resolve()
    if not source_path.exists():
        raise FileNotFoundError(source_path)
    if not config_path.exists():
        raise FileNotFoundError(config_path)
    if args.in_place and args.output is not None:
        parser.error("--in-place and --output cannot be used together")
    variant_target = is_variant_workbook(source_path)
    if args.in_place and not variant_target and not args.overwrite_formulas:
        parser.error("--in-place requires --overwrite-formulas because mapped ranges contain formulas")

    config = load_js_config(config_path)
    updates = build_updates(source_path, config)
    changed = changed_cells(source_path, updates)
    print(f"Mapped cells: {sum(len(cells) for cells in updates.values()):,}")
    print(f"Changed cells: {len(changed):,}")
    for sheet_name, address, old_value, new_value, key in changed[:20]:
        print(f"  {sheet_name}!{address}: {old_value!r} -> {new_value!r} ({key})")
    if len(changed) > 20:
        print(f"  ... and {len(changed) - 20:,} more")
    if args.check:
        print("Check only; no xlsx was written.")
        return

    output_path = (
        source_path
        if args.in_place
        else args.output.resolve()
        if args.output is not None
        else default_output_path(source_path, config_path)
    )
    overwrite_formulas = (args.overwrite_formulas or output_path != source_path) and not variant_target
    preserve_formula_cache = variant_target
    if preserve_formula_cache:
        print(
            "Variant card formulas are preserved; Fix Num and cached formula "
            "results are updated."
        )
    elif overwrite_formulas:
        print(
            "Mapped formula cells may be replaced by fixed config values; "
            "Symbol ID U:AA formulas are restored and Excel will rebuild the calculation chain."
        )
    write_patched_workbook(
        source_path,
        output_path,
        updates,
        force=args.force or args.in_place,
        overwrite_formulas=overwrite_formulas,
        preserve_formula_cache=preserve_formula_cache,
    )
    verify_output(output_path, config)
    print(f"Xlsx written and round-trip verified: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        prog="model_sync.py",
        description="H028 model/config sync",
        usage="model_sync.py {export,import} [options]   (see --help on each)",
    )
    parser.add_argument("command", choices=["export", "import"],
                        help="export: xlsx -> config_*.js ; import: config_*.js -> xlsx")
    args, rest = parser.parse_known_args()
    if args.command == "export":
        run_export(rest)
    else:
        run_import(rest)


if __name__ == "__main__":
    main()
