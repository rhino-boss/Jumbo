"""Rebuild H028 card weights from fresh Simulator multiplier reports.

The script keeps the current low/mid multiplier line shape, applies the design
constraints, solves each profile to its theoretical RTP target, writes config,
then backfills the corresponding RTP workbook without removing formulas.
"""
from __future__ import annotations

import argparse
import json
import math
from pathlib import Path

from openpyxl import load_workbook

import model_sync


BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
COIN_IN = 100.0
WEIGHT_TOTAL = 1_000_000_000
NATURAL_RATE_MIN = 0.001


def read_js(path: Path):
    return model_sync.load_js_config(path)


def report_data(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        base_sheet = workbook["Base Info"]
        base = {
            str(base_sheet.cell(row, 1).value): base_sheet.cell(row, 2).value
            for row in range(2, base_sheet.max_row + 1)
            if base_sheet.cell(row, 1).value is not None
        }
        sheet = workbook["Multiplier Line"]
        headers = {
            str(sheet.cell(1, column).value): column
            for column in range(1, sheet.max_column + 1)
        }
        rows = []
        for row in range(2, sheet.max_row + 1):
            rows.append({name: sheet.cell(row, column).value for name, column in headers.items()})
        return base, rows
    finally:
        workbook.close()


def interval_average(pay, count):
    return float(pay or 0) / float(count or 1) / COIN_IN if count else 0.0


def integer_weights(probabilities, total=WEIGHT_TOTAL):
    raw = [max(0.0, float(value)) * total for value in probabilities]
    result = [int(math.floor(value)) for value in raw]
    missing = total - sum(result)
    order = sorted(range(len(raw)), key=lambda index: raw[index] - result[index], reverse=True)
    for index in order[:missing]:
        result[index] += 1
    if sum(result) != total:
        raise AssertionError("Integer card weights do not sum to the requested total")
    return result


def tilted_probabilities(base_weights, averages, mask, mass, expected):
    indices = [index for index, allowed in enumerate(mask) if allowed and base_weights[index] > 0]
    if not indices:
        raise ValueError("No eligible intervals remain for tilted distribution")
    target_mean = expected / mass
    min_mean = min(averages[index] for index in indices)
    max_mean = max(averages[index] for index in indices)
    if not min_mean <= target_mean <= max_mean:
        raise ValueError(
            f"Target mean {target_mean:.6f} is outside eligible interval means "
            f"[{min_mean:.6f}, {max_mean:.6f}]"
        )

    def distribution(parameter):
        logs = [
            math.log(base_weights[index]) + parameter * averages[index] / 100.0
            for index in indices
        ]
        maximum = max(logs)
        values = [math.exp(value - maximum) for value in logs]
        denominator = sum(values)
        return [value / denominator for value in values]

    low, high = -100.0, 100.0
    for _ in range(160):
        middle = (low + high) / 2.0
        shares = distribution(middle)
        mean = sum(share * averages[index] for share, index in zip(shares, indices))
        if mean < target_mean:
            low = middle
        else:
            high = middle
    shares = distribution((low + high) / 2.0)
    result = [0.0] * len(base_weights)
    for share, index in zip(shares, indices):
        result[index] = share * mass
    return result


def build_bg_weights(cards, natural_rates, averages, target_rtp, trigger_rate, trigger_average, newbie=False):
    range_cards = cards[:-1]
    base_weights = [int(card.get("weight", 0) or 0) for card in range_cards]
    mask = []
    for card, natural_rate, base_weight in zip(range_cards, natural_rates[:-1], base_weights):
        upper = float(card["max"])
        allowed = natural_rate >= NATURAL_RATE_MIN and base_weight > 0
        if newbie:
            allowed = allowed and upper <= 30.0
        mask.append(allowed)
    range_mass = 1.0 - trigger_rate
    range_expected = target_rtp - trigger_rate * trigger_average
    probabilities = tilted_probabilities(
        base_weights,
        averages[:-1],
        mask,
        range_mass,
        range_expected,
    )
    probabilities.append(trigger_rate)
    return integer_weights(probabilities)


def build_fg_weights(
    cards,
    natural_rates,
    averages,
    target_average,
    tail_expected,
    newbie=False,
):
    base_weights = [int(card.get("weight", 0) or 0) for card in cards]
    probabilities = [0.0] * len(cards)
    if newbie:
        main_mask = [
            natural >= NATURAL_RATE_MIN
            and base > 0
            and float(card["min"]) >= 30.0
            and float(card["max"]) <= 100.0
            for card, natural, base in zip(cards, natural_rates, base_weights)
        ]
        return integer_weights(
            tilted_probabilities(
                base_weights,
                averages,
                main_mask,
                1.0,
                target_average,
            )
        )

    tail_indices = []
    tail_factors = []
    for index, (card, natural, base) in enumerate(zip(cards, natural_rates, base_weights)):
        lower, upper = float(card["min"]), float(card["max"])
        if (
            lower >= 200.0
            and upper <= 20_000.0
            and natural >= NATURAL_RATE_MIN
            and base > 0
            and averages[index] > 0
        ):
            tail_indices.append(index)
            tail_factors.append(2.0 if lower == 2000.0 and upper == 3000.0 else 1.0)
    if not tail_indices:
        raise ValueError("No eligible >200x FG intervals remain")
    factor_total = sum(tail_factors)
    for index, factor in zip(tail_indices, tail_factors):
        probabilities[index] = tail_expected * factor / factor_total / averages[index]
    tail_mass = sum(probabilities)

    main_mask = [
        natural >= NATURAL_RATE_MIN
        and base > 0
        and float(card["min"]) >= 20.0
        and float(card["max"]) <= 200.0
        for card, natural, base in zip(cards, natural_rates, base_weights)
    ]
    main = tilted_probabilities(
        base_weights,
        averages,
        main_mask,
        1.0 - tail_mass,
        target_average - tail_expected,
    )
    probabilities = [left + right for left, right in zip(probabilities, main)]
    return integer_weights(probabilities)


def card_table(config, profile, mode, segment):
    return config["card_system"][profile][mode][segment]


def replace_weights(cards, weights):
    if len(cards) != len(weights):
        raise ValueError("Card/weight length mismatch")
    for card, weight in zip(cards, weights):
        card["weight"] = int(weight)


def natural_arrays(normal_base, normal_rows, bf_base, bf_rows):
    total_rounds = int(normal_base["total_rounds"])
    trigger_count = int(normal_base["trigger_fg_bg_count"])
    trigger_pay = int(normal_base["trigger_fg_bg_pay"])
    fg_total = sum(int(row["FG_Session_Count"] or 0) for row in normal_rows)
    bf_total = sum(int(row["FG_Session_Count"] or 0) for row in bf_rows)
    if fg_total != trigger_count:
        raise ValueError(f"Normal FG rows total {fg_total}, trigger count {trigger_count}")
    if bf_total != int(bf_base["total_rounds"]):
        raise ValueError(f"BF FG rows total {bf_total}, rounds {bf_base['total_rounds']}")

    bg_counts = [int(row["BG_Count"] or 0) for row in normal_rows]
    bg_pays = [int(row["BG_Pay"] or 0) for row in normal_rows]
    bg_rates = [count / total_rounds for count in bg_counts]
    bg_averages = [interval_average(pay, count) for pay, count in zip(bg_pays, bg_counts)]
    bg_counts.append(trigger_count)
    bg_pays.append(trigger_pay)
    bg_rates.append(trigger_count / total_rounds)
    bg_averages.append(interval_average(trigger_pay, trigger_count))

    fg_counts = [int(row["FG_Session_Count"] or 0) for row in normal_rows]
    fg_pays = [int(row["FG_Pay"] or 0) for row in normal_rows]
    fg_rates = [count / fg_total for count in fg_counts]
    fg_averages = [interval_average(pay, count) for pay, count in zip(fg_pays, fg_counts)]

    bf_counts = [int(row["FG_Session_Count"] or 0) for row in bf_rows]
    bf_pays = [int(row["FG_Pay"] or 0) for row in bf_rows]
    bf_rates = [count / bf_total for count in bf_counts]
    bf_averages = [interval_average(pay, count) for pay, count in zip(bf_pays, bf_counts)]
    return {
        "bg": (bg_counts, bg_pays, bg_rates, bg_averages),
        "fg": (fg_counts, fg_pays, fg_rates, fg_averages),
        "bf": (bf_counts, bf_pays, bf_rates, bf_averages),
    }


def natural_updates(arrays):
    updates = {}
    for sheet_name in ("Detail", "Detail_Newbie"):
        for key, first_row in (("bg", 15), ("fg", 86)):
            counts, pays, rates, averages = arrays[key]
            for offset, values in enumerate(zip(counts, pays, rates, averages)):
                row = first_row + offset
                for column, value, label in zip("BCDE", values, ("Cnt", "Pay", "Hit Rate", "Avg. Multi.")):
                    model_sync.add_update(
                        updates,
                        sheet_name,
                        f"{column}{row}",
                        value,
                        f"Natural {key.upper()} {label}",
                    )
    counts, pays, rates, averages = arrays["bf"]
    for offset, values in enumerate(zip(counts, pays, rates, averages)):
        row = 163 + offset
        for column, value, label in zip("BCDE", values, ("Cnt", "Pay", "Hit Rate", "Avg. Multi.")):
            model_sync.add_update(
                updates,
                "Detail",
                f"{column}{row}",
                value,
                f"Natural BF {label}",
            )
    return updates


def theoretical_metrics(config, arrays):
    def expectation(cards, averages):
        total = sum(int(card.get("weight", 0) or 0) for card in cards)
        return sum(int(card.get("weight", 0) or 0) / total * average for card, average in zip(cards, averages))

    bg_avg = arrays["bg"][3]
    fg_avg = arrays["fg"][3]
    bf_avg = arrays["bf"][3]
    result = {}
    for profile in ("newbie", "oldhand"):
        bg = card_table(config, profile, "normal_bet", "weight_bg")
        fg = card_table(config, profile, "normal_bet", "weight_fg")
        trigger = bg[-1]["weight"] / sum(card["weight"] for card in bg)
        bg_rtp = expectation(bg, bg_avg)
        fg_rtp = expectation(fg, fg_avg) * trigger
        result[profile] = (bg_rtp, fg_rtp, bg_rtp + fg_rtp, trigger)
    bf = card_table(config, "oldhand", "buy_feature", "weight_fg")
    result["buy_feature"] = expectation(bf, bf_avg) / 75.0
    return result


def tune_variant(config_path, workbook_path, rtp_code, arrays, version):
    config = read_js(config_path)
    config["excel_version"] = version
    bg_rates, bg_averages = arrays["bg"][2], arrays["bg"][3]
    fg_rates, fg_averages = arrays["fg"][2], arrays["fg"][3]
    bf_rates, bf_averages = arrays["bf"][2], arrays["bf"][3]

    old_bg = card_table(config, "oldhand", "normal_bet", "weight_bg")
    newbie_bg = card_table(config, "newbie", "normal_bet", "weight_bg")
    old_trigger = 1.0 / 300.0
    newbie_trigger = newbie_bg[-1]["weight"] / sum(card["weight"] for card in newbie_bg)
    replace_weights(
        old_bg,
        build_bg_weights(old_bg, bg_rates, bg_averages, 0.72, old_trigger, bg_averages[-1]),
    )
    replace_weights(
        newbie_bg,
        build_bg_weights(newbie_bg, bg_rates, bg_averages, 0.72, newbie_trigger, bg_averages[-1], newbie=True),
    )

    fg_target = 0.20 if rtp_code == 92 else 0.22
    old_fg = card_table(config, "oldhand", "normal_bet", "weight_fg")
    newbie_fg = card_table(config, "newbie", "normal_bet", "weight_fg")
    buy_fg = card_table(config, "oldhand", "buy_feature", "weight_fg")
    replace_weights(
        old_fg,
        build_fg_weights(old_fg, fg_rates, fg_averages, fg_target / old_trigger, 0.02 / old_trigger),
    )
    replace_weights(
        newbie_fg,
        build_fg_weights(newbie_fg, fg_rates, fg_averages, 0.21 / newbie_trigger, 0.0, newbie=True),
    )
    replace_weights(
        buy_fg,
        build_fg_weights(buy_fg, bf_rates, bf_averages, 0.925 * 75.0, 0.02 * 75.0),
    )

    model_sync.write_js_config(config_path, config)
    raw_updates = natural_updates(arrays)
    model_sync.write_patched_workbook(
        workbook_path,
        workbook_path,
        raw_updates,
        force=True,
        preserve_formula_cache=True,
    )
    variant_updates = model_sync.build_variant_updates(workbook_path, config)
    model_sync.write_patched_workbook(
        workbook_path,
        workbook_path,
        variant_updates,
        force=True,
        preserve_formula_cache=True,
    )
    model_sync.verify_output(workbook_path, config)
    return theoretical_metrics(config, arrays)


def validate_constraints(config, arrays, rtp_code):
    issues = []
    for profile, mode, segment in model_sync.CARD_DETAIL_RANGES:
        cards = card_table(config, profile, mode, segment)
        rates = arrays["bf" if mode == "buy_feature" else "bg" if segment == "weight_bg" else "fg"][2]
        for card, natural in zip(cards, rates):
            weight = int(card.get("weight", 0) or 0)
            if weight > 0 and natural < NATURAL_RATE_MIN:
                issues.append(f"{profile}.{mode}.{segment} {card_label(card)} natural={natural:.6%}")
            if card.get("type") == "range" and weight > 0 and float(card["max"]) > 20_000:
                issues.append(f"{profile}.{mode}.{segment} {card_label(card)} exceeds 20000x")
    if issues:
        raise ValueError("Constraint violations:\n" + "\n".join(issues))


def card_label(card):
    if card.get("type") == "free_game":
        return "Free Game"
    return f"({card['min']:g}, {card['max']:g}]"


def main():
    parser = argparse.ArgumentParser(description="Tune H028 multiplier card weights")
    parser.add_argument("--normal-report", type=Path, required=True)
    parser.add_argument("--bf-report", type=Path, required=True)
    parser.add_argument("--version", default="2.0.0.37")
    args = parser.parse_args()
    normal_base, normal_rows = report_data(args.normal_report.resolve())
    bf_base, bf_rows = report_data(args.bf_report.resolve())
    arrays = natural_arrays(normal_base, normal_rows, bf_base, bf_rows)

    targets = (
        (92, PROJECT_DIR / "config_92A.js", BASE_DIR / "H028192A.xlsx"),
        (94, PROJECT_DIR / "config_94A.js", BASE_DIR / "H028194A.xlsx"),
    )
    output = {}
    for rtp_code, config_path, workbook_path in targets:
        output[str(rtp_code)] = tune_variant(
            config_path,
            workbook_path,
            rtp_code,
            arrays,
            args.version,
        )
        validate_constraints(read_js(config_path), arrays, rtp_code)
    print(json.dumps(output, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
