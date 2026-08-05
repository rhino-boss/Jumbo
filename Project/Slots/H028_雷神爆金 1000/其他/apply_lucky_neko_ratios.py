#!/usr/bin/env python3
"""Apply Lucky Neko ratios with initial Scatter from PostC1 and Drop Scatter weights."""

from __future__ import annotations

import argparse
import heapq
import json
import math
import random
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_ANALYSIS = ROOT / "其他" / "參考資料" / "analysis_lucky_neko_final.xlsx"
DEFAULT_CONFIG = ROOT / "config_92A.js"
SYMBOL_COUNT = 26
BASE_SYMBOL_COUNT = 13
DROP_WEIGHT_TOTAL = 1_000_000
POST_WEIGHT_TOTAL = 1_000_000
POST_SCATTER_RARE_RATE = 1 / 10_000
MEGAWAY_WEIGHT_TOTAL = 10_000
R7_OPTIMIZATION_STEPS = 400_000
MEGAWAY_PATTERNS = [
    [4, 1], [1, 4], [3, 2], [2, 3],
    [3, 1, 1], [1, 3, 1], [1, 1, 3],
    [2, 2, 1], [2, 1, 2], [1, 2, 2],
    [2, 1, 1, 1], [1, 2, 1, 1], [1, 1, 2, 1], [1, 1, 1, 2],
    [1, 1, 1, 1, 1],
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Apply Lucky Neko ratios to active H028 BG/FG config sets."
    )
    parser.add_argument("--analysis", type=Path, default=DEFAULT_ANALYSIS)
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG)
    parser.add_argument("--seed", type=int, default=20260803)
    parser.add_argument("--reel-length", type=int, default=200)
    parser.add_argument("--check", action="store_true")
    parser.add_argument("--keep-version", action="store_true")
    return parser.parse_args()


def load_config(path: Path) -> tuple[dict, str]:
    text = path.read_text(encoding="utf-8-sig")
    match = re.fullmatch(r"\s*const\s+data\s*=\s*(\{.*\})\s*;\s*", text, re.S)
    if not match:
        raise ValueError(f"Unsupported config format: {path}")
    return json.loads(match.group(1)), text


def section_rows(ws, label: str) -> list[int]:
    start = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == label:
            start = row + 2
            break
    if start is None:
        raise ValueError(f"{ws.title}: missing section {label}")
    rows = []
    for row in range(start, ws.max_row + 1):
        if ws.cell(row, 1).value is None:
            break
        rows.append(row)
    return rows


def read_symbol_ratios(wb, sheet: str, mode: str) -> list[list[float]]:
    ws = wb[sheet]
    ratios = [[0.0] * BASE_SYMBOL_COUNT for _ in range(6)]
    for row in section_rows(ws, mode):
        symbol_id = int(ws.cell(row, 1).value)
        if not 0 <= symbol_id < BASE_SYMBOL_COUNT:
            raise ValueError(f"{sheet}/{mode}: invalid symbol ID {symbol_id}")
        for reel in range(6):
            ratios[reel][symbol_id] = float(ws.cell(row, reel + 2).value or 0)
    return ratios


def read_count_distributions(wb, sheet: str, mode: str) -> list[dict[int, float]]:
    ws = wb[sheet]
    distributions = [dict() for _ in range(6)]
    for row in section_rows(ws, mode):
        count = int(ws.cell(row, 1).value)
        for reel in range(6):
            distributions[reel][count] = float(ws.cell(row, reel + 2).value or 0)
    return distributions


def expected_value(distribution: dict[int, float]) -> float:
    return sum(value * probability for value, probability in distribution.items())


def normalize(values: list[float]) -> list[float]:
    total = sum(values)
    if total <= 0:
        raise ValueError("Cannot normalize an empty distribution")
    return [value / total for value in values]


def without_scatter(values: list[float]) -> list[float]:
    result = list(values)
    result[1] = 0.0
    return normalize(result)


def split_silver(
    base_probabilities: list[float],
    silver_share: float,
    *,
    remove_scatter: bool,
) -> list[float]:
    base = without_scatter(base_probabilities) if remove_scatter else normalize(base_probabilities)
    eligible_total = sum(base[2:13])
    silver_share = min(max(float(silver_share), 0.0), eligible_total)
    silver_fraction = silver_share / eligible_total if eligible_total else 0.0
    result = [0.0] * SYMBOL_COUNT
    result[0] = base[0]
    result[1] = base[1]
    for symbol_id in range(2, 13):
        silver_probability = base[symbol_id] * silver_fraction
        result[symbol_id] = base[symbol_id] - silver_probability
        result[symbol_id + 11] = silver_probability
    return normalize(result)


def boost_absolute_share(
    probabilities: list[float], symbol_ids: tuple[int, ...], factor: float
) -> list[float]:
    """Multiply the selected symbols' absolute share and rescale all other symbols."""
    result = normalize(list(probabilities))
    selected_share = sum(result[symbol_id] for symbol_id in symbol_ids)
    if selected_share <= 0 or selected_share >= 1:
        return result
    target_share = min(1.0, selected_share * factor)
    selected_scale = target_share / selected_share
    other_scale = (1.0 - target_share) / (1.0 - selected_share)
    for symbol_id in range(len(result)):
        result[symbol_id] *= selected_scale if symbol_id in symbol_ids else other_scale
    return normalize(result)


def max_entropy_low_counts(total_mass: float, target_mean: float) -> list[float]:
    """Maximum-entropy probabilities for counts 0..3 with fixed mass and mean."""
    if total_mass <= 0:
        return [0.0] * 4
    conditional_mean = min(3.0, max(0.0, target_mean / total_mass))
    low, high = -40.0, 40.0
    for _ in range(160):
        midpoint = (low + high) / 2.0
        values = [math.exp(midpoint * count) for count in range(4)]
        mean = sum(count * value for count, value in enumerate(values)) / sum(values)
        if mean < conditional_mean:
            low = midpoint
        else:
            high = midpoint
    values = [math.exp(((low + high) / 2.0) * count) for count in range(4)]
    scale = total_mass / sum(values)
    return [value * scale for value in values]


def overview_metric(wb, name: str) -> float:
    ws = wb["Overview"]
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 1).value == name:
            return float(ws.cell(row, 2).value)
    raise ValueError(f"Overview: missing metric {name}")


def rounds_scatter_tail(wb) -> tuple[dict[int, int], int]:
    """Return inferred BG initial SC counts and FG retrigger event count."""
    ws = wb["Rounds"]
    headers = {str(cell.value): index for index, cell in enumerate(next(ws.iter_rows(values_only=False)))}
    required = {"has_free_spin_trigger", "free_spin_awarded"}
    if not required.issubset(headers):
        raise ValueError(f"Rounds: missing columns {sorted(required - headers.keys())}")
    bg_counts = {count: 0 for count in range(4, 8)}
    retrigger_events = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not bool(row[headers["has_free_spin_trigger"]]):
            continue
        awarded = int(row[headers["free_spin_awarded"]] or 0)
        remainder = awarded % 10
        initial_award = 10 if remainder == 0 else 12 if remainder == 2 else 14 if remainder == 4 else 10
        initial_count = min(7, 4 + max(0, (initial_award - 10) // 2))
        bg_counts[initial_count] += 1
        retrigger_events += max(0, (awarded - initial_award) // 10)
    return bg_counts, retrigger_events


def build_post_scatter_distribution(
    wb,
    mode: str,
    initial_base: list[list[float]],
    reel_height: list[dict[int, float]],
    r7_initial: list[float],
) -> list[float]:
    expected_main = sum(
        initial_base[reel][1] * expected_value(reel_height[reel])
        for reel in range(6)
    )
    target_mean = expected_main + r7_initial[1] * 4.0
    bg_tail_counts, retrigger_events = rounds_scatter_tail(wb)
    probabilities = [0.0] * 8
    if mode == "BG":
        denominator = overview_metric(wb, "total_rounds")
        for count, occurrences in bg_tail_counts.items():
            probabilities[count] = occurrences / denominator
    else:
        denominator = overview_metric(wb, "fg_count")
        probabilities[4] = retrigger_events / denominator

    # Every zero-probability count receives the explicit H028 floor of 1/10,000.
    # BG parameter group 2 applies its own 0..3-only override in update_config().
    for count in range(8):
        if probabilities[count] <= 0:
            probabilities[count] = POST_SCATTER_RARE_RATE

    tail_mass = sum(probabilities[4:])
    tail_mean = sum(count * probabilities[count] for count in range(4, 8))
    residual_mass = max(0.0, 1.0 - tail_mass)
    residual_mean = max(0.0, target_mean - tail_mean)
    probabilities[:4] = max_entropy_low_counts(residual_mass, residual_mean)
    return normalize(probabilities)


def largest_remainder(probabilities: list[float], total: int) -> list[int]:
    normalized = normalize(probabilities)
    exact = [value * total for value in normalized]
    result = [int(value) for value in exact]
    remaining = total - sum(result)
    order = sorted(range(len(exact)), key=lambda i: (exact[i] - result[i], -i), reverse=True)
    for index in order[:remaining]:
        result[index] += 1
    return result


def largest_remainder_keep_positive(probabilities: list[float], total: int) -> list[int]:
    normalized = normalize(probabilities)
    result = largest_remainder(normalized, total)
    exact = [value * total for value in normalized]
    for index, probability in enumerate(normalized):
        if probability <= 0 or result[index] > 0:
            continue
        donors = [i for i, count in enumerate(result) if count > 1]
        donor = max(donors, key=lambda i: (result[i] - exact[i], result[i], -i))
        result[donor] -= 1
        result[index] = 1
    return result


def shuffled_reel(counts: list[int], seed: int) -> list[int]:
    reel = [symbol_id for symbol_id, count in enumerate(counts) for _ in range(count)]
    random.Random(seed).shuffle(reel)
    return reel


def allocate_initial_counts(probabilities: list[float], total: int) -> list[int]:
    collapsed = probabilities[:BASE_SYMBOL_COUNT]
    for symbol_id in range(2, 13):
        collapsed[symbol_id] += probabilities[symbol_id + 11]
    # Every SymbolOcc_Init source symbol with a non-zero ratio must remain
    # visible as at least one ordinary (non-silver) symbol on the reel.
    base_counts = largest_remainder_keep_positive(collapsed, total)

    silver_probabilities = [probabilities[symbol_id + 11] for symbol_id in range(2, 13)]
    silver_total = min(round(sum(silver_probabilities) * total), sum(base_counts[2:13]))
    if silver_total:
        normalized = normalize(silver_probabilities)
        exact = [value * silver_total for value in normalized]
        silver_caps = [
            max(0, base_counts[index + 2] - (1 if collapsed[index + 2] > 0 else 0))
            for index in range(11)
        ]
        silver_total = min(silver_total, sum(silver_caps))
        exact = [value * silver_total for value in normalized]
        silver_counts = [min(int(value), silver_caps[index]) for index, value in enumerate(exact)]
        while sum(silver_counts) < silver_total:
            candidates = [
                index for index in range(11) if silver_counts[index] < silver_caps[index]
            ]
            best = max(candidates, key=lambda index: (exact[index] - silver_counts[index], -index))
            silver_counts[best] += 1
    else:
        silver_counts = [0] * 11

    result = [0] * SYMBOL_COUNT
    result[0] = base_counts[0]
    result[1] = base_counts[1]
    for index, symbol_id in enumerate(range(2, 13)):
        result[symbol_id] = base_counts[symbol_id] - silver_counts[index]
        result[symbol_id + 11] = silver_counts[index]
    return result


def force_nearest_symbol_total(
    counts: list[int],
    probabilities: list[float],
    symbol_ids: tuple[int, ...],
    total: int,
) -> list[int]:
    """Force a combined symbol count to the nearest representable reel share."""
    result = list(counts)
    target_count = round(sum(probabilities[symbol_id] for symbol_id in symbol_ids) * total)
    current_count = sum(result[symbol_id] for symbol_id in symbol_ids)
    while current_count > target_count:
        candidates = [
            symbol_id for symbol_id in symbol_ids
            if result[symbol_id] > (1 if symbol_id < BASE_SYMBOL_COUNT and probabilities[symbol_id] > 0 else 0)
        ]
        remove_id = max(
            candidates,
            key=lambda symbol_id: result[symbol_id] - probabilities[symbol_id] * total,
        )
        donors = [
            symbol_id for symbol_id in range(len(result))
            if symbol_id not in symbol_ids and probabilities[symbol_id] > 0
        ]
        add_id = max(
            donors,
            key=lambda symbol_id: probabilities[symbol_id] * total - result[symbol_id],
        )
        result[remove_id] -= 1
        result[add_id] += 1
        current_count -= 1
    while current_count < target_count:
        donors = [
            symbol_id for symbol_id in range(len(result))
            if symbol_id not in symbol_ids
            and result[symbol_id] > (1 if symbol_id < BASE_SYMBOL_COUNT and probabilities[symbol_id] > 0 else 0)
        ]
        remove_id = max(
            donors,
            key=lambda symbol_id: result[symbol_id] - probabilities[symbol_id] * total,
        )
        add_id = max(
            symbol_ids,
            key=lambda symbol_id: probabilities[symbol_id] * total - result[symbol_id],
        )
        result[remove_id] -= 1
        result[add_id] += 1
        current_count += 1
    return result


def read_extra_reel_ratios(wb, sheet: str, mode: str) -> list[float]:
    ws = wb[sheet]
    rate_column = 3 if mode == "BG" else 4
    values = [0.0] * SYMBOL_COUNT
    for row in range(3, ws.max_row + 1):
        symbol_id = ws.cell(row, 1).value
        if symbol_id is None or not isinstance(symbol_id, (int, float)):
            continue
        values[int(symbol_id)] = float(ws.cell(row, rate_column).value or 0)
    return normalize(values)


def read_extra_reel_same(wb, mode: str) -> dict[str, list[float]]:
    ws = wb["ExtraReelSame"]
    all_column, more_column, denominator_column = (
        (4, 5, 4) if mode == "BG" else (6, 7, 6)
    )
    denominator = float(ws.cell(31, denominator_column).value or 0)
    if denominator <= 0:
        raise ValueError(f"ExtraReelSame/{mode}: invalid denominator")
    all_rates = [0.0] * SYMBOL_COUNT
    more_rates = [0.0] * SYMBOL_COUNT
    for row in range(19, 31):
        symbol_id = int(ws.cell(row, 3).value)
        # The published ExtraReelSame comparison starts at M1 (config ID 2).
        # ID 1 is Scatter and is not part of the Extra Reel same-symbol target.
        if symbol_id == 1:
            continue
        all_rates[symbol_id] = float(ws.cell(row, all_column).value or 0) / denominator
        more_rates[symbol_id] = float(ws.cell(row, more_column).value or 0) / denominator
    return {"all4": all_rates, "more2": more_rates}


def extra_reel_same_counts(reel: list[int]) -> dict[str, list[int]]:
    all_counts = [0] * SYMBOL_COUNT
    more_counts = [0] * SYMBOL_COUNT
    length = len(reel)
    for start in range(length):
        counts: dict[int, int] = {}
        for offset in range(4):
            symbol_id = reel[(start + offset) % length]
            counts[symbol_id] = counts.get(symbol_id, 0) + 1
        for symbol_id, count in counts.items():
            if count == 4:
                all_counts[symbol_id] += 1
            if count >= 2:
                more_counts[symbol_id] += 1
    return {"all4": all_counts, "more2": more_counts}


def optimize_extra_reel_order(
    reel: list[int], targets: dict[str, list[float]], seed: int
) -> list[int]:
    """Reorder a fixed R7 multiset to fit Lucky Neko's cyclic four-symbol windows."""
    rng = random.Random(seed)
    length = len(reel)
    target_all = [rate * length for rate in targets["all4"]]
    target_more = [rate * length for rate in targets["more2"]]
    # Prefer the upper adjacent 0.5 pp step so the reel contains slightly more
    # four-connected same-symbol windows than the fractional competitor target.
    target_all_total = math.ceil(sum(target_all) - 1e-12)

    symbol_totals = [reel.count(symbol_id) for symbol_id in range(SYMBOL_COUNT)]
    desired_all = [
        min(int(target_all[symbol_id]), symbol_totals[symbol_id] // 4)
        for symbol_id in range(SYMBOL_COUNT)
    ]
    desired_total = target_all_total
    while sum(desired_all) < desired_total:
        candidates = [
            symbol_id for symbol_id in range(SYMBOL_COUNT)
            if desired_all[symbol_id] < symbol_totals[symbol_id] // 4
        ]
        if not candidates:
            break
        best_symbol = min(
            candidates,
            key=lambda symbol_id: (
                (desired_all[symbol_id] + 1 - target_all[symbol_id]) ** 2
                - (desired_all[symbol_id] - target_all[symbol_id]) ** 2,
                symbol_id,
            ),
        )
        desired_all[best_symbol] += 1

    remaining = symbol_totals[:]
    cluster_symbols: list[int] = []
    for symbol_id, all_count in enumerate(desired_all):
        cluster_symbols.extend([symbol_id] * all_count)
        remaining[symbol_id] -= 4 * all_count

    def constraints_ok(sequence: list[int]) -> bool:
        boundary = next(
            (index for index in range(length) if sequence[index] != sequence[index - 1]),
            None,
        )
        if boundary is None:
            return False
        runs: list[int] = []
        traversed = 0
        index = boundary
        while traversed < length:
            symbol_id = sequence[index]
            run_length = 0
            while traversed < length and sequence[index] == symbol_id:
                run_length += 1
                traversed += 1
                index = (index + 1) % length
            runs.append(run_length)
        return max(runs) <= 4 and not any(
            runs[index] >= 2 and runs[(index + 1) % len(runs)] >= 2
            for index in range(len(runs))
        )

    def arrange_singletons() -> list[int] | None:
        for _ in range(500):
            heap = [
                (-count, rng.random(), symbol_id)
                for symbol_id, count in enumerate(remaining)
                if count > 0
            ]
            heapq.heapify(heap)
            result: list[int] = []
            while heap:
                first = heapq.heappop(heap)
                if result and first[2] == result[-1]:
                    if not heap:
                        result = []
                        break
                    second = heapq.heappop(heap)
                    heapq.heappush(heap, first)
                    first = second
                count, _, symbol_id = first
                result.append(symbol_id)
                count += 1
                if count < 0:
                    heapq.heappush(heap, (count, rng.random(), symbol_id))
            if result and result[0] != result[-1]:
                return result
        return None

    candidate = None
    for _ in range(500):
        singletons = arrange_singletons()
        if not singletons:
            continue
        clusters = cluster_symbols[:]
        rng.shuffle(clusters)
        available_gaps = set(range(len(singletons)))
        placed: dict[int, int] = {}
        failed = False
        for symbol_id in clusters:
            compatible = [
                gap for gap in available_gaps
                if singletons[gap] != symbol_id
                and singletons[(gap + 1) % len(singletons)] != symbol_id
            ]
            if not compatible:
                failed = True
                break
            gap = rng.choice(compatible)
            placed[gap] = symbol_id
            available_gaps.remove(gap)
        if failed:
            continue
        built: list[int] = []
        for gap, symbol_id in enumerate(singletons):
            built.append(symbol_id)
            if gap in placed:
                built.extend([placed[gap]] * 4)
        if len(built) == length and constraints_ok(built):
            candidate = built
            break
    if candidate is None:
        raise ValueError("Could not build a valid R7 sequence with separated runs")

    def loss(metrics: dict[str, list[int]]) -> float:
        all_counts, more_counts = metrics["all4"], metrics["more2"]
        symbol_error = sum(
            2.0 * (all_counts[symbol_id] - target_all[symbol_id]) ** 2
            + (more_counts[symbol_id] - target_more[symbol_id]) ** 2
            for symbol_id in range(SYMBOL_COUNT)
        )
        aggregate_error = (
            5.0 * (sum(all_counts) - target_all_total) ** 2
            + 1.5 * (sum(more_counts) - sum(target_more)) ** 2
        )
        return symbol_error + aggregate_error

    def window_contributions(start: int) -> tuple[int | None, tuple[int, ...]]:
        counts: dict[int, int] = {}
        for offset in range(4):
            symbol_id = candidate[(start + offset) % length]
            counts[symbol_id] = counts.get(symbol_id, 0) + 1
        all_symbol = next((symbol_id for symbol_id, count in counts.items() if count == 4), None)
        more_symbols = tuple(symbol_id for symbol_id, count in counts.items() if count >= 2)
        return all_symbol, more_symbols

    def local_constraints_ok(centers: tuple[int, int]) -> bool:
        for center in centers:
            for offset in range(-6, 7):
                index = (center + offset) % length
                symbol_id = candidate[index]
                run_length = 1
                step = 1
                while step <= 4 and candidate[(index - step) % length] == symbol_id:
                    run_length += 1
                    step += 1
                step = 1
                while step <= 4 and candidate[(index + step) % length] == symbol_id:
                    run_length += 1
                    step += 1
                if run_length > 4:
                    return False
                next_index = (index + 1) % length
                if candidate[index] == candidate[next_index]:
                    continue
                left_repeated = candidate[(index - 1) % length] == candidate[index]
                right_repeated = candidate[(next_index + 1) % length] == candidate[next_index]
                if left_repeated and right_repeated:
                    return False
        return True

    metrics = extra_reel_same_counts(candidate)
    current_loss = loss(metrics)
    best = candidate[:]
    best_loss = current_loss
    for step in range(R7_OPTIMIZATION_STEPS):
        left = rng.randrange(length)
        right = rng.randrange(length - 1)
        if right >= left:
            right += 1
        if candidate[left] == candidate[right]:
            continue
        affected = {
            (position - offset) % length
            for position in (left, right)
            for offset in range(4)
        }
        old_contributions = {start: window_contributions(start) for start in affected}
        old_all = metrics["all4"][:]
        old_more = metrics["more2"][:]
        for all_symbol, more_symbols in old_contributions.values():
            if all_symbol is not None:
                metrics["all4"][all_symbol] -= 1
            for symbol_id in more_symbols:
                metrics["more2"][symbol_id] -= 1
        candidate[left], candidate[right] = candidate[right], candidate[left]
        if not local_constraints_ok((left, right)):
            candidate[left], candidate[right] = candidate[right], candidate[left]
            metrics["all4"] = old_all
            metrics["more2"] = old_more
            continue
        for start in affected:
            all_symbol, more_symbols = window_contributions(start)
            if all_symbol is not None:
                metrics["all4"][all_symbol] += 1
            for symbol_id in more_symbols:
                metrics["more2"][symbol_id] += 1
        if sum(metrics["all4"]) < target_all_total:
            candidate[left], candidate[right] = candidate[right], candidate[left]
            metrics["all4"] = old_all
            metrics["more2"] = old_more
            continue
        new_loss = loss(metrics)
        temperature = max(0.02, 2.5 * (1.0 - step / R7_OPTIMIZATION_STEPS))
        accept = new_loss <= current_loss or rng.random() < math.exp((current_loss - new_loss) / temperature)
        if accept:
            current_loss = new_loss
            if new_loss < best_loss:
                best_loss = new_loss
                best = candidate[:]
        else:
            candidate[left], candidate[right] = candidate[right], candidate[left]
            metrics["all4"] = old_all
            metrics["more2"] = old_more
    if not constraints_ok(best):
        raise ValueError("Optimized R7 sequence violates consecutive-symbol constraints")
    return best


def read_symbol_size(wb, mode: str, symbol_id: int = 12) -> list[float]:
    ws = wb["SymbolSize"]
    for row in section_rows(ws, mode):
        if int(ws.cell(row, 1).value) == symbol_id:
            return normalize([float(ws.cell(row, column).value or 0) for column in range(2, 6)])
    raise ValueError(f"SymbolSize/{mode}: missing symbol ID {symbol_id}")


def build_megaway_weights(wb, mode: str, reel_height: list[dict[int, float]]) -> list[list[int]]:
    # H028's pattern describes only the main reel blocks. Lucky Neko's ReelHigh_Init
    # includes one Extra Reel symbol, so a source count of N maps to N-1 pattern parts.
    size_probabilities = read_symbol_size(wb, mode, symbol_id=12)
    result: list[list[int]] = []
    for reel in range(6):
        if reel in (0, 5):
            weights = [0] * len(MEGAWAY_PATTERNS)
            weights[-1] = 1
            result.append(weights)
            continue

        probabilities = [0.0] * len(MEGAWAY_PATTERNS)
        for part_count in range(2, 6):
            indices = [
                index for index, pattern in enumerate(MEGAWAY_PATTERNS)
                if len(pattern) == part_count
            ]
            raw = [
                math.prod(size_probabilities[size - 1] for size in MEGAWAY_PATTERNS[index])
                for index in indices
            ]
            conditional = normalize(raw)
            group_probability = reel_height[reel].get(part_count + 1, 0.0)
            for index, value in zip(indices, conditional):
                probabilities[index] = group_probability * value
        result.append(largest_remainder(probabilities, MEGAWAY_WEIGHT_TOTAL))
    return result


def bump_version(version: str) -> str:
    parts = version.split(".")
    if not parts or not parts[-1].isdigit():
        raise ValueError(f"Cannot bump version: {version}")
    parts[-1] = str(int(parts[-1]) + 1)
    return ".".join(parts)


def build_mode_targets(wb, mode: str) -> dict:
    initial_base = read_symbol_ratios(wb, "SymbolOcc_Init", mode)
    drop_base = read_symbol_ratios(wb, "SymbolOcc_Drop", mode)
    silver_initial = read_count_distributions(wb, "Silver_Init", mode)
    silver_drop = read_count_distributions(wb, "Silver_Drop", mode)
    reel_height = read_count_distributions(wb, "ReelHigh_Init", mode)

    initial_probabilities = []
    drop_probabilities = []
    initial_silver_shares = []
    drop_silver_shares = []
    for reel in range(6):
        expected_blocks = expected_value(reel_height[reel])
        initial_share = (
            expected_value(silver_initial[reel]) / expected_blocks if expected_blocks else 0.0
        )
        # Cascade replacements are 1x1 and the main reel is five cells high.
        drop_share = expected_value(silver_drop[reel]) / 5.0
        initial_silver_shares.append(initial_share)
        drop_silver_shares.append(drop_share)
        initial_probabilities.append(
            split_silver(initial_base[reel], initial_share, remove_scatter=True)
        )
        drop_probabilities.append(
            split_silver(drop_base[reel], drop_share, remove_scatter=False)
        )

    r7_initial_source = read_extra_reel_ratios(wb, "Extra Reel_Init", mode)
    r7_drop_source = read_extra_reel_ratios(wb, "Extra Reel_Drop", mode)
    post_scatter = build_post_scatter_distribution(
        wb,
        mode,
        initial_base,
        reel_height,
        r7_initial_source,
    )
    r7_initial = without_scatter(r7_initial_source)
    r7_drop = normalize(r7_drop_source)
    return {
        "initial": initial_probabilities + [r7_initial],
        "drop": drop_probabilities + [r7_drop],
        "post_scatter": post_scatter,
        "megaway": build_megaway_weights(wb, mode, reel_height),
        "extra_reel_same": read_extra_reel_same(wb, mode),
        "initial_silver_shares": initial_silver_shares,
        "drop_silver_shares": drop_silver_shares,
    }


def update_config(
    config: dict,
    wb,
    seed: int,
    reel_length: int = 200,
    keep_version: bool = False,
) -> dict:
    targets = {mode: build_mode_targets(wb, mode) for mode in ("BG", "FG")}
    mode_settings = {
        "BG": ("BaseGame", (1, 2)),
        "FG": ("FreeGame", (1, 2, 3)),
    }

    for mode, (prefix, group_indices) in mode_settings.items():
        if reel_length <= 0:
            raise ValueError("Reel length must be positive")
        table1_drop_targets = targets[mode]["drop"]
        if mode == "BG":
            table1_drop_targets = [
                boost_absolute_share(targets[mode]["drop"][reel], (2, 13), 0.5)
                for reel in range(7)
            ]
        base_drop_weights = [
            largest_remainder(table1_drop_targets[reel], DROP_WEIGHT_TOTAL)
            for reel in range(7)
        ]
        targets[mode]["table1_drop_weights"] = [row[:] for row in base_drop_weights]

        for group_index in group_indices:
            symbol_key = f"{prefix}Symbol{group_index}"
            weight_key = f"{prefix}SymbolWeight{group_index}"
            if len(config[symbol_key]) != 7:
                raise ValueError(f"{symbol_key}: expected seven non-empty reels")
            lengths = [
                reel_length
                for reel in range(7)
            ]
            group_initial_targets = targets[mode]["initial"]
            if mode == "BG" and group_index == 2:
                group_initial_targets = []
                for reel_symbols in config["BaseGameSymbol1"]:
                    copied_distribution = [
                        reel_symbols.count(symbol_id) / len(reel_symbols)
                        for symbol_id in range(SYMBOL_COUNT)
                    ]
                    group_initial_targets.append(
                        boost_absolute_share(copied_distribution, (2, 13), 2.0)
                    )
            if mode == "BG" and group_index == 2:
                drop_weights = [row[:] for row in base_drop_weights]
                for row in drop_weights:
                    row[1] = 0
                    row[2] *= 3
                    row[13] *= 3
            else:
                drop_weights = [row[:] for row in base_drop_weights]

            new_reels = []
            group_seed = seed + (0 if mode == "BG" else 10_000) + group_index * 1_000
            for reel, length in enumerate(lengths):
                counts = (
                    allocate_initial_counts(group_initial_targets[reel], length)
                    if reel < 6
                    else largest_remainder_keep_positive(group_initial_targets[reel], length)
                )
                if mode == "BG" and group_index == 2:
                    counts = force_nearest_symbol_total(
                        counts, group_initial_targets[reel], (2, 13), length
                    )
                new_reels.append(shuffled_reel(counts, group_seed + reel))
            if group_index == 1:
                new_reels[6] = optimize_extra_reel_order(
                    new_reels[6],
                    targets[mode]["extra_reel_same"],
                    seed + (30_000 if mode == "BG" else 40_000),
                )
            config[symbol_key] = new_reels
            config[weight_key] = [[1] * length for length in lengths]
            for combo in range(1, 6):
                config[f"{prefix}{group_index}Drop{combo}"] = [row[:] for row in drop_weights]
            post_weights = largest_remainder(targets[mode]["post_scatter"], POST_WEIGHT_TOTAL)
            if mode == "BG" and group_index == 2:
                post_weights[4:] = [0, 0, 0, 0]
            config[f"{prefix}{group_index}PostC1"] = [list(range(8)), post_weights]

            for reel, symbols in enumerate(new_reels):
                if len(symbols) != lengths[reel] or len(config[weight_key][reel]) != lengths[reel]:
                    raise ValueError(f"{mode} group {group_index} R{reel + 1}: invalid reel/weight length")
                source = group_initial_targets[reel]
                required_ids = range(BASE_SYMBOL_COUNT) if reel < 6 else range(SYMBOL_COUNT)
                missing = [
                    symbol_id for symbol_id in required_ids
                    if symbol_id != 1 and source[symbol_id] > 0 and symbol_id not in symbols
                ]
                if missing:
                    raise ValueError(
                        f"{mode} group {group_index} R{reel + 1}: missing non-zero source IDs {missing}"
                    )
                if 1 in symbols:
                    raise ValueError(f"{mode} group {group_index} R{reel + 1}: Scatter remains on reel")
            for reel, row in enumerate(drop_weights):
                expected_scatter = 0 if mode == "BG" and group_index == 2 else base_drop_weights[reel][1]
                if row[1] != expected_scatter:
                    raise ValueError(
                        f"{mode} group {group_index} R{reel + 1}: "
                        f"Drop Scatter {row[1]} != target {expected_scatter}"
                    )
            if any(config[f"{prefix}{group_index}Drop{combo}"] != drop_weights for combo in range(1, 6)):
                raise ValueError(f"{mode} group {group_index}: Drop1-Drop5 are not identical")

        config[f"{prefix}MegaWay1"] = [row[:] for row in targets[mode]["megaway"]]
        mega_sums = [sum(weights) for weights in config[f"{prefix}MegaWay1"]]
        if mega_sums != [1, MEGAWAY_WEIGHT_TOTAL, MEGAWAY_WEIGHT_TOTAL, MEGAWAY_WEIGHT_TOTAL, MEGAWAY_WEIGHT_TOTAL, 1]:
            raise ValueError(f"{mode}: invalid MegaWay totals {mega_sums}")

    # BG table 2 copies table 1 except for its M1 and Post Scatter overrides.
    config["BaseGameMegaWay2"] = [row[:] for row in config["BaseGameMegaWay1"]]
    config["BaseGameMY2"] = list(config["BaseGameMY1"])
    config["ReelWeight"] = [6000, 4000]

    if not keep_version:
        config["excel_version"] = bump_version(str(config["excel_version"]))
    return targets


def main() -> int:
    args = parse_args()
    config, original_text = load_config(args.config)
    original_version = str(config["excel_version"])
    wb = load_workbook(args.analysis, read_only=True, data_only=True)
    try:
        targets = update_config(
            config,
            wb,
            args.seed,
            reel_length=args.reel_length,
            keep_version=args.keep_version,
        )
    finally:
        wb.close()

    new_text = "const data = " + json.dumps(config, ensure_ascii=False, indent=2) + ";\n"
    changed = new_text != original_text
    print(f"Analysis: {args.analysis}")
    print(f"Config: {args.config}")
    print(f"Version: {original_version} -> {config['excel_version']}")
    print(f"Changed: {changed}")
    for mode in ("BG", "FG"):
        initial = ", ".join(f"{value:.6%}" for value in targets[mode]["initial_silver_shares"])
        drop = ", ".join(f"{value:.6%}" for value in targets[mode]["drop_silver_shares"])
        print(f"{mode} Silver_Init shares R1-R6: {initial}")
        print(f"{mode} Silver_Drop shares R1-R6: {drop}")
        post_weights = largest_remainder(targets[mode]["post_scatter"], POST_WEIGHT_TOTAL)
        print(f"{mode} Post Scatter counts 0-7: {post_weights}")
        drop_scatter = [row[1] for row in targets[mode]["table1_drop_weights"]]
        print(f"{mode} Drop Scatter weights R1-R7: {drop_scatter}")
        for reel, weights in enumerate(targets[mode]["megaway"], start=1):
            print(f"{mode} MegaWay R{reel}: {weights}")
        r7_metrics = extra_reel_same_counts(config[("BaseGame" if mode == "BG" else "FreeGame") + "Symbol1"][6])
        target_same = targets[mode]["extra_reel_same"]
        print(
            f"{mode} ExtraReelSame all4: target={sum(target_same['all4']):.6%}, "
            f"actual={sum(r7_metrics['all4']) / 200:.6%}"
        )
        print(
            f"{mode} ExtraReelSame more2: target={sum(target_same['more2']):.6%}, "
            f"actual={sum(r7_metrics['more2']) / 200:.6%}"
        )

    if args.check:
        return 1 if changed else 0

    temporary = args.config.with_suffix(args.config.suffix + ".tmp")
    temporary.write_text(new_text, encoding="utf-8", newline="\n")
    temporary.replace(args.config)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
