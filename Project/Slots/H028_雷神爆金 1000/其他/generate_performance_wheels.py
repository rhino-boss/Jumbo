from collections import Counter
from heapq import heappop, heappush
from pathlib import Path
import os
import random

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from scipy.optimize import Bounds, LinearConstraint, milp


PROJECT_DIR = Path(__file__).resolve().parents[1]
BOOK = PROJECT_DIR / "Source" / "H0281.xlsx"
TEMP = PROJECT_DIR / "Source" / "H0281.performance.tmp.xlsx"
PATTERNS = [
    [4, 1],
    [1, 4],
    [3, 2],
    [2, 3],
    [3, 1, 1],
    [1, 3, 1],
    [1, 1, 3],
    [2, 2, 1],
    [2, 1, 2],
    [1, 2, 2],
    [2, 1, 1, 1],
    [1, 2, 1, 1],
    [1, 1, 2, 1],
    [1, 1, 1, 2],
    [1, 1, 1, 1, 1],
]
SHEET_PAIRS = [
    ("BG_Symbol (2)", "BG_Performance Wheel", 280200),
    ("FG_Symbol (3)", "FG_Performance Wheel", 280300),
    ("BF_Symbol", "BF_Performance Wheel", 280400),
]


def height_basis(weights):
    raw = [
        sum(weight * pattern.count(height) for weight, pattern in zip(weights, PATTERNS))
        for height in range(1, 5)
    ]
    total_cells = 5 * sum(weights)
    if total_cells <= 0:
        raise ValueError(f"Invalid MegaWay weights: {weights}")

    expected = [200 * value / total_cells for value in raw]
    best = None
    for n4 in range(51):
        for n3 in range(67):
            for n2 in range(101):
                n1 = 200 - 4 * n4 - 3 * n3 - 2 * n2
                if n1 < 0:
                    break
                actual = (n1, n2, n3, n4)
                score = sum(
                    (actual[index] - expected[index]) ** 2 / (expected[index] + 0.25)
                    for index in range(4)
                )
                candidate = (score, sum(actual), actual)
                if best is None or candidate < best:
                    best = candidate
    return raw, total_cells, best[2]


def allocate_heights(symbols, target, raw, total_cells):
    counts = Counter(symbols)
    items = list(counts.items())
    symbol_count = len(items)
    x_count = symbol_count * 4
    variable_count = x_count * 3
    objective = np.zeros(variable_count, dtype=np.float64)
    integrality = np.zeros(variable_count, dtype=np.int32)
    integrality[:x_count] = 1
    rows = []
    bounds = []

    # Each symbol must still occupy exactly its original number of reel cells.
    for symbol_index, (_, cell_count) in enumerate(items):
        row = np.zeros(variable_count, dtype=np.float64)
        for height_index in range(4):
            row[symbol_index * 4 + height_index] = height_index + 1
        rows.append(row)
        bounds.append(float(cell_count))

    # Total number of 1x1/2x1/3x1/4x1 heads follows the rounded MegaWay basis.
    for height_index in range(4):
        row = np.zeros(variable_count, dtype=np.float64)
        for symbol_index in range(symbol_count):
            row[symbol_index * 4 + height_index] = 1
        rows.append(row)
        bounds.append(float(target[height_index]))

    # Minimize absolute per-symbol deviation from the independent MegaWay mix.
    for symbol_index, (_, cell_count) in enumerate(items):
        for height_index in range(4):
            x_index = symbol_index * 4 + height_index
            plus_index = x_count + x_index
            minus_index = 2 * x_count + x_index
            expected = cell_count * raw[height_index] / total_cells
            row = np.zeros(variable_count, dtype=np.float64)
            row[x_index] = 1
            row[plus_index] = -1
            row[minus_index] = 1
            rows.append(row)
            bounds.append(float(expected))
            penalty = 1.0 / (expected + 0.35)
            objective[plus_index] = penalty
            objective[minus_index] = penalty

    matrix = np.vstack(rows)
    result = milp(
        c=objective,
        integrality=integrality,
        bounds=Bounds(np.zeros(variable_count), np.full(variable_count, np.inf)),
        constraints=LinearConstraint(matrix, np.array(bounds), np.array(bounds)),
        options={"time_limit": 10.0},
    )
    if not result.success:
        raise ValueError(
            f"Unable to allocate height counts {target} to {dict(counts)}: {result.message}"
        )

    allocation = {}
    rounded = np.rint(result.x[:x_count]).astype(int).reshape(symbol_count, 4)
    for symbol_index, (symbol, _) in enumerate(items):
        allocation[symbol] = tuple(int(value) for value in rounded[symbol_index])

    blocks = []
    for symbol, _ in items:
        for height in range(1, 5):
            blocks.extend([(symbol, height)] * allocation[symbol][height - 1])
    return blocks


def arrange_blocks(blocks, seed):
    buckets = {}
    for symbol, height in blocks:
        buckets.setdefault(symbol, []).append(height)

    for attempt in range(1000):
        rng = random.Random(seed + attempt)
        shuffled = {symbol: heights.copy() for symbol, heights in buckets.items()}
        for heights in shuffled.values():
            rng.shuffle(heights)

        heap = []
        for symbol, heights in shuffled.items():
            heappush(heap, (-len(heights), rng.random(), symbol))

        arranged = []
        held = None
        while heap:
            count, _, symbol = heappop(heap)
            height = shuffled[symbol].pop()
            arranged.append((symbol, height))
            if held is not None:
                heappush(heap, held)
            count += 1
            held = (count, rng.random(), symbol) if count < 0 else None

        if len(arranged) == len(blocks) and all(
            arranged[index][0] != arranged[(index + 1) % len(arranged)][0]
            for index in range(len(arranged))
        ):
            return arranged
    raise ValueError("Unable to arrange blocks without adjacent identical symbols")


def expand_blocks(blocks):
    symbols = []
    heights = []
    for symbol, height in blocks:
        symbols.extend([symbol] * height)
        heights.append(height)
        heights.extend([0] * (height - 1))
    if len(symbols) != 200 or len(heights) != 200:
        raise ValueError(f"Expanded reel length is {len(symbols)}, expected 200")
    return symbols, heights


def validate_reel(source_symbols, target_symbols, weights, heights, label):
    if Counter(source_symbols) != Counter(target_symbols):
        raise ValueError(f"{label}: symbol distribution mismatch")
    if any(weight != 1 for weight in weights):
        raise ValueError(f"{label}: Symbol Weight must all be 1")

    position = 0
    while position < 200:
        height = heights[position]
        if height not in (1, 2, 3, 4):
            raise ValueError(f"{label}: invalid block head at row offset {position}")
        symbol = target_symbols[position]
        for continuation in range(1, height):
            if target_symbols[position + continuation] != symbol:
                raise ValueError(f"{label}: continuation symbol mismatch")
            if heights[position + continuation] != 0:
                raise ValueError(f"{label}: continuation height must be 0")
        position += height
    if position != 200:
        raise ValueError(f"{label}: expanded length must be 200")


def main():
    workbook = load_workbook(BOOK)
    results = []

    for source_name, target_name, seed_base in SHEET_PAIRS:
        source = workbook[source_name]
        target = workbook[target_name]

        for summary_row in range(4, 30):
            for reel_index in range(7):
                letter = get_column_letter(13 + reel_index)
                target.cell(summary_row, 3 + reel_index).value = (
                    f'=COUNTIF({letter}$4:{letter}$203,$A{summary_row})'
                )
        for row in range(4, 204):
            target.cell(row, 12).value = row - 4

        for reel_index in range(7):
            source_symbols = [
                source.cell(row, 13 + reel_index).value for row in range(4, 204)
            ]
            if any(symbol is None for symbol in source_symbols):
                raise ValueError(f"{source_name} R{reel_index + 1} is not a complete 200-cell reel")

            if reel_index in (0, 5, 6):
                blocks = [(symbol, 1) for symbol in source_symbols]
                height_counts = (200, 0, 0, 0)
            else:
                mega_weights = [
                    int(source.cell(row, 3 + reel_index).value or 0)
                    for row in range(33, 48)
                ]
                raw, total_cells, height_counts = height_basis(mega_weights)
                blocks = allocate_heights(source_symbols, height_counts, raw, total_cells)
                blocks = arrange_blocks(blocks, seed_base + reel_index)

            expanded_symbols, expanded_heights = expand_blocks(blocks)
            if Counter(expanded_symbols) != Counter(source_symbols):
                raise ValueError(f"{target_name} R{reel_index + 1}: symbol distribution mismatch")

            symbol_col = 13 + reel_index
            id_col = 21 + reel_index
            weight_col = 29 + reel_index
            height_col = 37 + reel_index
            for row, (symbol, height) in enumerate(
                zip(expanded_symbols, expanded_heights), start=4
            ):
                target.cell(row, symbol_col).value = symbol
                target.cell(row, id_col).value = (
                    f'=VLOOKUP({get_column_letter(symbol_col)}{row}, $A$4:$J$29, 10, 0)'
                )
                target.cell(row, weight_col).value = 1
                target.cell(row, height_col).value = height

            observed = tuple(
                sum(value == height for value in expanded_heights)
                for height in range(1, 5)
            )
            if observed != height_counts:
                raise ValueError(
                    f"{target_name} R{reel_index + 1}: {observed} != {height_counts}"
                )
            results.append((target_name, reel_index + 1, len(blocks), height_counts))

        for row in range(204, 305):
            target.cell(row, 12).value = None
            for base_col in (13, 21, 29, 37):
                for reel_index in range(7):
                    target.cell(row, base_col + reel_index).value = None

    if hasattr(workbook, "calculation"):
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
    workbook.save(TEMP)

    # Normal mode is intentional: repeated .cell() access on a read-only worksheet
    # reparses rows and turns this small validation into an O(n^2) scan.
    check = load_workbook(TEMP, data_only=False, read_only=False)
    for source_name, target_name, _ in SHEET_PAIRS:
        source = check[source_name]
        target = check[target_name]
        for reel_index in range(7):
            source_symbols = [
                source.cell(row, 13 + reel_index).value for row in range(4, 204)
            ]
            target_symbols = [
                target.cell(row, 13 + reel_index).value for row in range(4, 204)
            ]
            weights = [
                target.cell(row, 29 + reel_index).value for row in range(4, 204)
            ]
            heights = [
                target.cell(row, 37 + reel_index).value for row in range(4, 204)
            ]
            validate_reel(
                source_symbols,
                target_symbols,
                weights,
                heights,
                f"{target_name} R{reel_index + 1}",
            )
    check.close()

    os.replace(TEMP, BOOK)
    for target_name, reel, block_count, height_counts in results:
        print(
            f"{target_name} R{reel}: expanded=200, blocks={block_count}, "
            f"heights={height_counts}"
        )


if __name__ == "__main__":
    main()
