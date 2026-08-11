# -*- coding: utf-8 -*-
"""Make the H045 PARsheet fully English and tidy the symbol-sheet styling.

Run against each workbook in place:

    py localize_workbook.py H045192.xlsx H045194.xlsx

Every Chinese string in the workbook must have an entry in TRANSLATIONS; the
script fails loudly if it meets one that does not, so nothing is silently left
untranslated.
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Border, PatternFill

CJK = re.compile(r"[　-〿一-鿿＀-￯]")

TRANSLATIONS = {
    # ---- Overview -----------------------------------------------------
    "Pay Table：": "Pay Table:",
    "Feature：": "Feature:",
    "※ 金框符號（M1G~JG）判獎時視為對應一般符號，賠率相同。":
        "* Golden symbols (M1G-JG) pay as their base symbol; the pay table is identical.",
    "※ WW1 / WW2 可替代除 C1 外的所有一般符號，本身無賠率。":
        "* WW1 / WW2 substitute for every symbol except C1 and have no pay of their own.",
    "※ C1 僅統計出現總數用於觸發 Free Game，無 Scatter Pay。":
        "* C1 is only counted to trigger the Free Game; there is no scatter pay.",
    "1. Base Game 為 5 輪 × 4 列、1024 Ways 盤面。連線後消除所有連線符號，剩餘符號依輪帶順序落下補位。":
        "1. The Base Game is a 5 reel x 4 row, 1024 Ways board. Winning symbols are removed and the "
        "remaining symbols drop down, with new symbols drawn from the strip to refill the board.",
    "2. 同一次 Spin 每完成一次消除，得分倍率往上提升；BG 為 X1、X2、X3、X5、X10，FG 為 X2、X4、X6、X10、X20。":
        "2. Each cascade within a spin raises the win multiplier: X1, X2, X3, X5, X10 in the Base Game "
        "and X2, X4, X6, X10, X20 in the Free Game.",
    "3. 消除到金框符號時，該格原地轉為 WW1；或依 Random Wild 權重轉為 WW2 並額外複製 2~4 個 WW2。":
        "3. When a golden symbol is part of a win it turns into a WW1 in place, or - per the Random Wild "
        "weights - into a WW2 that copies itself to a further 2-4 positions.",
    "4. 每一輪皆有機會出現 C1，消除補位後也可能出現，同一輪可出現超過 1 顆 C1。":
        "4. C1 can appear on any reel, including on refills after a cascade, and more than one C1 may "
        "appear on the same reel.",
    "5. 所有消除結束後統計盤面 C1，達 3 顆以上觸發 10 場 Free Game；FG 內再達 3 顆固定加 5 場，上限 50 場。":
        "5. Once all cascades finish the C1 on the board are counted: 3 or more triggers 10 free spins. "
        "3 or more during the Free Game adds a fixed 5 spins, up to a maximum of 50.",
    "6. 本遊戲提供 Buy Feature（40.5 × Bet），不提供 Super Feature。":
        "6. A Buy Feature is available at 40.5 x Bet. There is no Super Feature.",
    # ---- Description --------------------------------------------------
    "1. Base Game 每次 Spin 依卡片系統（Multiplier_Weight）對應到的表持續抽取，直到抽出對應得分區間。":
        "1. Every Base Game spin draws repeatedly from the table selected by the card system "
        "(Multiplier_Weight) until the win falls inside the drawn range.",
    "種類": "Worksheet",
    "2. 初始盤面依選到的 BG 工作頁 5 條輪帶，並依該頁 Symbol Weight R1~R5 權重決定停輪位置與符號。":
        "2. The starting board uses the 5 strips of the selected BG worksheet; the stop positions are "
        "drawn using that worksheet's Symbol Weight R1-R5.",
    "3. 若有連線，消除連線符號，並依同輪權重抽取新符號補上空格。":
        "3. On a win the winning symbols are removed and the gaps are refilled by drawing new symbols "
        "with the same reel's weights.",
    "4. 當 Combo 隨消除上升時得分倍率跟著上升，倍數為 X1、X2、X3、X5、X10，達到 X10 後不再上升。":
        "4. The win multiplier rises with the combo count: X1, X2, X3, X5, X10. It stays at X10 "
        "thereafter.",
    "5. 隨機百搭判定：": "5. Random wild evaluation:",
    "(1) 當連線盤面消除到至少一個金框符號（M1G~JG）時進行以下判定。":
        "(1) The following is evaluated whenever at least one golden symbol (M1G-JG) is part of a win.",
    "(2) 每次 Spin 有一個「是否已出現隨機百搭」的 FLAG；FLAG 尚未觸發前，每次 Combo 只要消除到金框符號都會判定一次，"
    "FLAG 使用過後續 Combo 不再判定。":
        "(2) Each spin carries a flag for whether a random wild has already occurred. Until it is set, "
        "every combo that removes a golden symbol is evaluated once; after it is set no further combo "
        "in that spin is evaluated.",
    "(3) 消到金框符號時依 Parameter 的 Random Wild Weight，若骰到 2~4 則該格取代為 WW2，並額外複製 2~4 個 WW2 "
    "取代 R2~R5 中隨機格子，不會蓋掉原本是 C1、WW1、WW2 的格子。":
        "(3) When a golden symbol is removed the Random Wild Weight in Parameter is rolled. On a 2-4 the "
        "cell becomes a WW2 and a further 2-4 WW2 replace random positions on R2-R5, never overwriting "
        "an existing C1, WW1 or WW2.",
    "(4) 多個金框符號同時被消除時同樣走上述判定，但最多只有其中一個金框觸發隨機百搭，其餘轉成 WW1。":
        "(4) If several golden symbols are removed together the same evaluation applies, but at most one "
        "of them triggers the random wild; the rest become WW1.",
    "(5) 順序為：消除 → 補格子 → 所有金框符號瞇牌 → 金框格轉成 WW1 或 WW2 → 若有複製 WW2，最後再處理要蓋掉哪些格子。":
        "(5) Order of play: remove -> refill -> reveal all golden symbols -> golden cells become WW1 or "
        "WW2 -> if WW2 copies were rolled, resolve which positions they overwrite last.",
    "6. 每次連線消除後有機會掉落 C1，同一輪可出現超過 1 個 C1。":
        "6. C1 may drop in on any refill, and more than one C1 may appear on the same reel.",
    "7. C1 在消除過程中不參與連線消除，直到無法再連線時才統計盤面 C1，決定是否進入 Free Spins。":
        "7. C1 never takes part in a win. Only once no further win is possible are the C1 on the board "
        "counted to decide whether the Free Spins are entered.",
    "1. 進入 Free Game 後依卡片系統對應的組別取得指定場數的高表與低表，出現順序隨機排列。":
        "1. On entering the Free Game the card system's group decides how many high- and low-table "
        "spins are awarded; their order is shuffled.",
    "組別": "Group",
    "高表場數": "High Spins",
    "低表場數": "Low Spins",
    "※ 目前 Multiplier_Weight 中 Free Game 與 Buy Feature 的「使用的表」皆為 E，即固定 10 場高表。":
        "* Every Free Game and Buy Feature card in Multiplier_Weight currently uses group E, i.e. a "
        "fixed 10 high-table spins.",
    "2. 高表場依下列權重選擇本場使用的高表工作頁；低表場固定使用 FG_Symbol (5)。":
        "2. Each high-table spin picks its worksheet with the weights below; low-table spins always use "
        "FG_Symbol (5).",
    "工作頁": "Worksheet",
    "高表輪帶": "High Strip",
    "3. Free Game 中最終盤面出現 3 顆以上 C1 時固定增加 5 場，其中 1 場高表與 4 場低表，順序隨機排列。":
        "3. 3 or more C1 on a Free Game final board adds a fixed 5 spins - 1 high table and 4 low table "
        "- shuffled into the remaining order.",
    "4. 若有連線，消除連線符號並依同輪權重抽取新符號補上空格。":
        "4. On a win the winning symbols are removed and the gaps refilled with the same reel's weights.",
    "5. 當 Combo 隨消除上升時得分倍率跟著上升，倍數為 X2、X4、X6、X10、X20，達到 X20 後不再上升。":
        "5. The win multiplier rises with the combo count: X2, X4, X6, X10, X20. It stays at X20 "
        "thereafter.",
    "6. 隨機百搭判定同 Base Game 第 5 點；FG 不受「同一 Spin 只觸發一次」限制，每次 Combo 皆可判定。":
        "6. Random wild evaluation is as in Base Game point 5, except the Free Game has no once-per-spin "
        "limit: every combo is evaluated.",
    "7. 每次連線消除後有機會掉落 C1，同一輪可出現超過 1 個 C1。":
        "7. C1 may drop in on any refill, and more than one C1 may appear on the same reel.",
    "8. C1 在消除過程中不參與連線消除，直到無法再連線時才統計盤面 C1，決定是否 Retrigger。":
        "8. C1 never takes part in a win. Only once no further win is possible are the C1 on the board "
        "counted to decide whether the feature retriggers.",
    "9. Free Game 最大場次為 50 場。": "9. The Free Game is capped at 50 spins.",
    "1. 購買價格為 40.5 × Bet。購買後使用 BF_Symbol 輪帶與其 Symbol Weight 決定進場盤面 RNG。":
        "1. The price is 40.5 x Bet. The entry board is drawn from the BF_Symbol strips using that "
        "worksheet's Symbol Weight.",
    "2. 其餘流程皆與 Base Game 的 2~7 相同。": "2. Everything else follows Base Game points 2-7.",
    "3. 進場盤面統計 C1 觸發 Free Game 後，依卡片系統對應組別取得高表與低表場數，順序隨機排列。":
        "3. Once the entry board's C1 trigger the Free Game, the card system's group decides the high- "
        "and low-table spin counts, shuffled into a random order.",
    "4. 其餘流程皆與 Free Game 的 2~9 相同。": "4. Everything else follows Free Game points 2-9.",
    "注意": "Notes",
    "1. 本遊戲不提供 Super Feature / Super Free Game。":
        "1. This game has no Super Feature and no Super Free Game.",
    "2. 工作頁命名對應：BG_Symbol＝BG 高表、BG_Symbol (2)＝BG 低表、FG_Symbol~FG_Symbol (4)＝FG 高表 A/K/Q/J、"
    "FG_Symbol (5)＝FG 低表、BF_Symbol＝Buy Feature 進場表。":
        "2. Worksheet mapping: BG_Symbol = BG high table, BG_Symbol (2) = BG low table, "
        "FG_Symbol to FG_Symbol (4) = FG high tables A/K/Q/J, FG_Symbol (5) = FG low table, "
        "BF_Symbol = Buy Feature entry table.",
    "3. Random Wild Weight 與 Win Multiplier Ladder 集中於 Parameter 工作頁。":
        "3. Random Wild Weight and the Win Multiplier Ladder live on the Parameter worksheet.",
    "4. 卡片系統的區間權重集中於 Multiplier_Weight_Newbie 與 Multiplier_Weight_Oldhand。":
        "4. The card system's range weights live on Multiplier_Weight_Newbie and "
        "Multiplier_Weight_Oldhand.",
    "5. Overview 的 Pay Back / Hit% / Pulls-Hit 以公式連動 Multiplier_Weight_* 的彙總列，調整 Fix Num 後全部自動重算。":
        "5. Overview's Pay Back / Hit% / Pulls-Hit are linked by formula to the summary rows of "
        "Multiplier_Weight_*, so changing a Fix Num recalculates everything.",
    # ---- Parameter ----------------------------------------------------
    "Base Game 高表": "Base Game high table",
    "Base Game 低表": "Base Game low table",
    "※ 實際選表由 Multiplier_Weight_* 的「使用的表」欄位決定，不另外做表權重抽取。":
        "* The table actually used is set by the Table column of Multiplier_Weight_*; there is no "
        "separate table-selection draw.",
    "Worksheet Name \\ 複製數量": "Worksheet Name \\ Copies",
    "※ 骰到 0 代表本次金框只轉成 WW1，不複製 WW2。":
        "* Rolling 0 means the golden symbol only becomes a WW1, with no WW2 copies.",
    "Worksheet Name \\ 第 N 次消除": "Worksheet Name \\ Cascade N",
    "※ 倍率只在同一 Spin 的消除期間推進，新的 Spin 回到該模式第一階。":
        "* The multiplier only advances within a spin; a new spin returns to that mode's first step.",
    "Retrigger 高表場數": "Retrigger High Spins",
    "Retrigger 低表場數": "Retrigger Low Spins",
    # ---- card sheets ---------------------------------------------------
    "使用的表": "Table",
}

SYMBOL_SHEETS = ("BG_Symbol", "BG_Symbol (2)", "FG_Symbol", "FG_Symbol (2)", "FG_Symbol (3)",
                 "FG_Symbol (4)", "FG_Symbol (5)", "BF_Symbol")
NO_FILL = PatternFill()
NO_BORDER = Border()


def translate(wb) -> list[str]:
    missing = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or not CJK.search(cell.value):
                    continue
                replacement = TRANSLATIONS.get(cell.value)
                if replacement is None:
                    missing.append(f'{ws.title}!{cell.coordinate}: {cell.value}')
                else:
                    cell.value = replacement
    return missing


def restyle(wb) -> None:
    """Group captions and the strip-total row are labels, not table cells."""
    for name in SYMBOL_SHEETS:
        ws = wb[name]
        for col in (10, 17, 23):
            cell = ws.cell(2, col)
            cell.fill = NO_FILL
            cell.border = NO_BORDER
        # the "--" totals row sits directly under the symbol-count block
        for row in range(4, 40):
            if str(ws.cell(row, 1).value or "").strip() == "--":
                for col in range(1, 9):
                    cell = ws.cell(row, col)
                    cell.fill = NO_FILL
                    cell.border = NO_BORDER
                break


def main() -> None:
    parser = argparse.ArgumentParser(description="Translate the H045 workbook to English")
    parser.add_argument("workbooks", nargs="+", type=Path)
    args = parser.parse_args()

    failed = False
    for path in args.workbooks:
        wb = load_workbook(path)
        missing = translate(wb)
        restyle(wb)
        if missing:
            failed = True
            print(f'{path.name}: {len(missing)} untranslated string(s):')
            for item in missing:
                print('   ', item)
        wb.save(path)
        print(f'wrote {path}')
    if failed:
        sys.exit(1)


if __name__ == "__main__":
    main()
