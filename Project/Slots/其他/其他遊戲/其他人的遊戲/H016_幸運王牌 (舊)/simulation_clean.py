#%%
import re
import numpy as np
import math
from numba import jit, njit, types
from numba.typed import Dict
import numba

def extract_all_js_vars_numpy(filename, dtype=float):
    with open(filename, encoding='utf-8') as f:
        js = f.read()
    results = {}
    for m in re.finditer(r'const (\w+) = (.*?);', js, re.DOTALL):
        name = m.group(1)
        arr_str = m.group(2)
        arr_str = arr_str.replace('null', 'np.nan').replace('true', 'True').replace('false', 'False')
        try:
            arr = eval(arr_str, {"np": np, "True": True, "False": False})
            np_arr = np.array(arr, dtype=dtype)
            results[name] = np_arr
        except Exception:
            pass
    return results

# 使用
all_arrays = extract_all_js_vars_numpy('data.js')
globals().update(all_arrays)

# 預編譯和緩存常用的數據結構
@njit
def get_precompiled_data():
    """預編譯常用數據結構以提升性能"""
    symbol_to_lp = np.full(19, -1, dtype=np.int32)
    for i in range(4):
        symbol_to_lp[3+i] = i
        symbol_to_lp[11+i] = i
    for i in range(4):
        symbol_to_lp[7+i] = 4+i
        symbol_to_lp[15+i] = 4+i
    return symbol_to_lp

# 全局預編譯數據
PRECOMPILED_SYMBOL_TO_LP = get_precompiled_data()

@njit
def sample_sequences_weighted_numba(weight_array, content_array, n, seed=None):
    """
    Numba加速版本的加權抽樣函數
    """
    if seed is not None:
        np.random.seed(seed)
    
    lines, length = weight_array.shape
    output = np.zeros((lines, n), dtype=content_array.dtype)
    
    for i in range(lines):
        w = weight_array[i]
        w_sum = np.sum(w)
        if w_sum == 0:
            # 如果權重全為0，隨機選擇
            for j in range(n):
                start = np.random.randint(0, length)
                idx = (start + j) % length
                output[i, j] = content_array[i, idx]
        else:
            # 加權隨機抽一個位置
            prob = w / w_sum
            cumsum = np.cumsum(prob)
            r = np.random.random()
            start = 0
            for k in range(length):
                if r <= cumsum[k]:
                    start = k
                    break
            
            # 從start開始取n個連續元素
            for j in range(n):
                idx = (start + j) % length
                output[i, j] = content_array[i, idx]
    
    return output

@njit
def build_symbol_to_linkpoint_numba():
    """
    Numba加速版本的符號映射建立函數 - 使用預編譯版本
    """
    return PRECOMPILED_SYMBOL_TO_LP

@njit
def waygame_fullscore_numba(board, linkpoint):
    """
    Numba加速版本的計分函數
    """
    symbol_to_lp = build_symbol_to_linkpoint_numba()
    total_score = 0
    details = np.zeros((8,3), dtype=np.int32)
    rows, cols = board.shape
    
    for symbol_lp in range(8):  # 0~3:M1~M4, 4~7:A~J
        valid_symbol = np.where(symbol_to_lp == symbol_lp)[0]
        
        for c0 in range(cols):
            # 檢查起點是否有效
            symbol_idx = int(board[0, c0])
            can_start = False
            
            if symbol_idx in (0, 1):  # Wild符號可以作為任何符號的起點
                can_start = True
            elif symbol_idx == 2:  # C1符號不能作為起點
                can_start = False
            elif symbol_idx < len(symbol_to_lp) and symbol_to_lp[symbol_idx] == symbol_lp:
                can_start = True
            
            if not can_start:
                continue
            
            # 建立匹配遮罩
            mask_lengths = np.zeros(rows, dtype=np.int32)
            mask = np.zeros((rows, cols), dtype=np.int32)
            
            for r in range(rows):
                count = 0
                if r == 0:
                    if (board[0, c0] in valid_symbol) or (board[0, c0] in (0,1)):
                        mask[r, count] = c0
                        count = 1
                else:
                    for c in range(cols):
                        s = board[r, c]
                        if s in (0,1) or s in valid_symbol:
                            if s != 2:
                                mask[r, count] = c
                                count += 1
                    if count == 0:
                        break
                mask_lengths[r] = count
            
            # 計算最大連線長度和路徑數
            max_conn = 0
            max_ways = 0
            for conn in range(3, 6):
                if conn <= len(mask_lengths) and np.all(mask_lengths[:conn] > 0):
                    ways = 1
                    for layer in range(conn):
                        ways *= mask_lengths[layer]
                    if ways > 0 and conn > max_conn:
                        max_conn = conn
                        max_ways = ways
            
            if max_conn >= 3:
                details[symbol_lp, max_conn-3] += max_ways
    
    # 計算總分
    if linkpoint.shape == (3,8):
        linkpoint_t = linkpoint.T
    else:
        linkpoint_t = linkpoint
    
    total_score = np.sum(details * linkpoint_t)
    return total_score, details

@njit
def point_numba(weight_array, reel_array, initial_board, linkpoint, rdwild_values, rdwild_weights, seed=None):
    """
    Numba加速版本的遊戲主函數
    """
    if seed is not None:
        np.random.seed(seed)
    
    rows, cols = initial_board.shape
    board = initial_board.astype(np.int32).copy()
    total_score = 0
    golden_symbols_converted_this_round = 0
    cascade_count = 0
    random_wild_used = False
    
    # 追踪统计信息
    total_golden_symbols = 0  # 总共出现的黄金符号数
    total_golden_converted = 0  # 黄金符号转换为wild的总数
    total_wild_eliminated = 0  # wild被消除的总数
    cascades_with_golden = 0  # 有黄金符号消除的cascade次数
    ww2_triggered = 0  # ww2(randomwild)触发次数
    ww2_with_elimination = 0  # ww2触发后有消除的次数
    
    # 统计初始版面的黄金符号
    for r in range(rows):
        for c in range(cols):
            if 11 <= board[r, c] <= 18:
                total_golden_symbols += 1
    
    while True:
        # randomwild判定 - 一次spin最多觸發一次
        ww2_triggered_this_cascade = False
        if golden_symbols_converted_this_round > 0 and not random_wild_used:
            # 加權隨機選擇wild數量
            if np.sum(rdwild_weights) > 0:
                prob = rdwild_weights / np.sum(rdwild_weights)
                cumsum = np.cumsum(prob)
                r = np.random.random()
                chosen_idx = 0
                for k in range(len(prob)):
                    if r <= cumsum[k]:
                        chosen_idx = k
                        break
                wild_count = int(rdwild_values[chosen_idx])
                
                if wild_count > 0:
                    # 找出可替換的位置（R2~R5，非wild且非C1）
                    valid_positions = []
                    for r in range(1, rows):
                        for c in range(cols):
                            if board[r, c] not in (0, 1, 2):
                                valid_positions.append(r * cols + c)
                    
                    # 隨機選擇位置替換為WILD
                    if len(valid_positions) >= wild_count:
                        # Fisher-Yates shuffle前wild_count個元素
                        for i in range(min(wild_count, len(valid_positions))):
                            j = np.random.randint(i, len(valid_positions))
                            valid_positions[i], valid_positions[j] = valid_positions[j], valid_positions[i]
                        
                        for i in range(wild_count):
                            pos = valid_positions[i]
                            r, c = pos // cols, pos % cols
                            board[r, c] = 0  # 替換為WILD
                        
                        random_wild_used = True  # 標記已觸發
                        ww2_triggered = 1  # 记录ww2触发
                        ww2_triggered_this_cascade = True
        
        golden_symbols_converted_this_round = 0
        
        score, details = waygame_fullscore_numba(board, linkpoint)
        
        if score == 0:
            break
        
        # 如果ww2触发且有分数，记录ww2_with_elimination
        if ww2_triggered_this_cascade and score > 0:
            ww2_with_elimination = 1
        
        cascade_count += 1
        
        # 計算乘倍效果
        if cascade_count == 1:
            multiplier = 1
        elif cascade_count == 2:
            multiplier = 2
        elif cascade_count == 3:
            multiplier = 3
        else:
            multiplier = 5
            
        multiplied_score = score * multiplier
        total_score += multiplied_score
        
        # 找出有得分的符號位置並處理消除
        symbol_to_lp = build_symbol_to_linkpoint_numba()
        elimination_mask = np.zeros_like(board, dtype=np.bool_)
        golden_symbols_to_convert = []
        
        for symbol_lp in range(8):
            valid_symbol = np.where(symbol_to_lp == symbol_lp)[0]
            for c0 in range(cols):
                symbol_idx = int(board[0, c0])
                can_start = False
                
                if symbol_idx in (0, 1):
                    can_start = True
                elif symbol_idx == 2:
                    can_start = False
                elif symbol_idx < len(symbol_to_lp) and symbol_to_lp[symbol_idx] == symbol_lp:
                    can_start = True
                
                if not can_start:
                    continue
                
                # 建立匹配路徑
                mask_lengths = np.zeros(rows, dtype=np.int32)
                mask = np.zeros((rows, cols), dtype=np.int32)
                
                for r in range(rows):
                    count = 0
                    if r == 0:
                        if (board[0, c0] in valid_symbol) or (board[0, c0] in (0,1)):
                            mask[r, count] = c0
                            count = 1
                    else:
                        for c in range(cols):
                            s = board[r, c]
                            if s in (0,1) or s in valid_symbol:
                                if s != 2:
                                    mask[r, count] = c
                                    count += 1
                        if count == 0:
                            break
                    mask_lengths[r] = count
                
                # 找出最大連線長度
                max_conn = 0
                for conn in range(3, 6):
                    if conn <= len(mask_lengths) and np.all(mask_lengths[:conn] > 0):
                        ways = 1
                        for layer in range(conn):
                            ways *= mask_lengths[layer]
                        if ways > 0 and conn > max_conn:
                            max_conn = conn
                
                if max_conn >= 3:
                    for r in range(max_conn):
                        for idx in range(mask_lengths[r]):
                            c = mask[r, idx]
                            symbol = board[r, c]
                            if 11 <= symbol <= 18:
                                golden_symbols_to_convert.append(r * cols + c)
                            else:
                                elimination_mask[r, c] = True
        
        # 處理黃金符號轉換
        unique_golden = list(set(golden_symbols_to_convert))
        if len(unique_golden) > 0:
            cascades_with_golden += 1  # 记录有黄金符号消除的cascade
        
        for pos in unique_golden:
            r, c = pos // cols, pos % cols
            if board[r, c] >= 11 and board[r, c] <= 18:
                board[r, c] = 0
                golden_symbols_converted_this_round += 1
                total_golden_converted += 1  # 累计黄金转换数量
        
        # 消除符號，并统计wild被消除的数量
        wild_eliminated_this_cascade = 0
        for r in range(rows):
            for c in range(cols):
                if elimination_mask[r, c]:
                    if board[r, c] in (0, 1):  # Wild或Wild2被消除
                        wild_eliminated_this_cascade += 1
                    board[r, c] = -1
        total_wild_eliminated += wild_eliminated_this_cascade
        
        # 重力下降
        for r in range(rows):
            non_empty = []
            for c in range(cols):
                if board[r, c] != -1:
                    non_empty.append(board[r, c])
            
            empty_count = cols - len(non_empty)
            for c in range(len(non_empty)):
                board[r, c] = non_empty[c]
            for c in range(len(non_empty), cols):
                board[r, c] = -1
        
        # 補充新符號，并统计新出现的黄金符号
        for r in range(rows):
            for c in range(cols-1, -1, -1):
                if board[r, c] == -1:
                    new_symbols = sample_sequences_weighted_numba(
                        weight_array[r:r+1], reel_array[r:r+1], 1
                    )
                    new_symbol = int(new_symbols[0, 0])
                    board[r, c] = new_symbol
                    if 11 <= new_symbol <= 18:
                        total_golden_symbols += 1
    
    # 返回统计信息: [combo数, 黄金符号总数, 黄金转wild数, wild消除数, cascade详情[有黄金消除次数, ww2触发, ww2有消除]]
    cascade_details = np.array([cascades_with_golden, ww2_triggered, ww2_with_elimination], dtype=np.int32)
    return total_score, board, cascade_count, total_golden_symbols, total_golden_converted, total_wild_eliminated, cascade_details

@njit
def freegame_point_numba(weight_array, reel_array, initial_board, linkpoint, rdwild_values, rdwild_weights, seed=None):
    """
    Numba加速版本的免費遊戲函數
    """
    if seed is not None:
        np.random.seed(seed)
    
    rows, cols = initial_board.shape
    board = initial_board.astype(np.int32).copy()
    total_score = 0
    golden_symbols_converted_this_round = 0
    cascade_count = 0
    
    # 追踪统计信息
    total_golden_symbols = 0  # 总共出现的黄金符号数
    total_golden_converted = 0  # 黄金符号转换为wild的总数
    total_wild_eliminated = 0  # wild被消除的总数
    cascades_with_golden = 0  # 有黄金符号消除的cascade次数
    ww2_triggered_count = 0  # ww2(randomwild)触发次数（freegame可以多次触发）
    ww2_with_elimination_count = 0  # ww2触发后有消除的次数
    
    # 统计初始版面的黄金符号
    for r in range(rows):
        for c in range(cols):
            if 11 <= board[r, c] <= 18:
                total_golden_symbols += 1
    
    while True:
        # randomwild判定 - 每次有黃金符號轉換時都可能觸發
        ww2_triggered_this_cascade = False
        if golden_symbols_converted_this_round > 0:
            if np.sum(rdwild_weights) > 0:
                prob = rdwild_weights / np.sum(rdwild_weights)
                cumsum = np.cumsum(prob)
                r = np.random.random()
                chosen_idx = 0
                for k in range(len(prob)):
                    if r <= cumsum[k]:
                        chosen_idx = k
                        break
                wild_count = int(rdwild_values[chosen_idx])
                
                if wild_count > 0:
                    valid_positions = []
                    for r in range(1, rows):
                        for c in range(cols):
                            if board[r, c] not in (0, 1, 2):
                                valid_positions.append(r * cols + c)
                    
                    if len(valid_positions) >= wild_count:
                        for i in range(min(wild_count, len(valid_positions))):
                            j = np.random.randint(i, len(valid_positions))
                            valid_positions[i], valid_positions[j] = valid_positions[j], valid_positions[i]
                        
                        for i in range(wild_count):
                            pos = valid_positions[i]
                            r, c = pos // cols, pos % cols
                            board[r, c] = 0
                        
                        ww2_triggered_count += 1  # 记录ww2触发
                        ww2_triggered_this_cascade = True
        
        golden_symbols_converted_this_round = 0
        
        score, details = waygame_fullscore_numba(board, linkpoint)
        
        if score == 0:
            break
        
        # 如果ww2触发且有分数，记录ww2_with_elimination
        if ww2_triggered_this_cascade and score > 0:
            ww2_with_elimination_count += 1
        
        cascade_count += 1
        
        # Free Game乘倍係數
        if cascade_count == 1:
            multiplier = 2
        elif cascade_count == 2:
            multiplier = 4
        elif cascade_count == 3:
            multiplier = 6
        else:
            multiplier = 10
            
        multiplied_score = score * multiplier
        total_score += multiplied_score
        
        # 找出有得分的符號位置並處理消除
        symbol_to_lp = build_symbol_to_linkpoint_numba()
        elimination_mask = np.zeros_like(board, dtype=np.bool_)
        golden_symbols_to_convert = []
        
        for symbol_lp in range(8):
            valid_symbol = np.where(symbol_to_lp == symbol_lp)[0]
            for c0 in range(cols):
                symbol_idx = int(board[0, c0])
                can_start = False
                
                if symbol_idx in (0, 1):
                    can_start = True
                elif symbol_idx == 2:
                    can_start = False
                elif symbol_idx < len(symbol_to_lp) and symbol_to_lp[symbol_idx] == symbol_lp:
                    can_start = True
                
                if not can_start:
                    continue
                
                mask_lengths = np.zeros(rows, dtype=np.int32)
                mask = np.zeros((rows, cols), dtype=np.int32)
                
                for r in range(rows):
                    count = 0
                    if r == 0:
                        if (board[0, c0] in valid_symbol) or (board[0, c0] in (0,1)):
                            mask[r, count] = c0
                            count = 1
                    else:
                        for c in range(cols):
                            s = board[r, c]
                            if s in (0,1) or s in valid_symbol:
                                if s != 2:
                                    mask[r, count] = c
                                    count += 1
                        if count == 0:
                            break
                    mask_lengths[r] = count
                
                max_conn = 0
                for conn in range(3, 6):
                    if conn <= len(mask_lengths) and np.all(mask_lengths[:conn] > 0):
                        ways = 1
                        for layer in range(conn):
                            ways *= mask_lengths[layer]
                        if ways > 0 and conn > max_conn:
                            max_conn = conn
                
                if max_conn >= 3:
                    for r in range(max_conn):
                        for idx in range(mask_lengths[r]):
                            c = mask[r, idx]
                            symbol = board[r, c]
                            if 11 <= symbol <= 18:
                                golden_symbols_to_convert.append(r * cols + c)
                            else:
                                elimination_mask[r, c] = True
        
        # 處理黃金符號轉換
        unique_golden = list(set(golden_symbols_to_convert))
        if len(unique_golden) > 0:
            cascades_with_golden += 1  # 记录有黄金符号消除的cascade
        
        for pos in unique_golden:
            r, c = pos // cols, pos % cols
            if board[r, c] >= 11 and board[r, c] <= 18:
                board[r, c] = 0
                golden_symbols_converted_this_round += 1
                total_golden_converted += 1  # 累计黄金转换数量
        
        # 消除符號，并统计wild被消除的数量
        wild_eliminated_this_cascade = 0
        for r in range(rows):
            for c in range(cols):
                if elimination_mask[r, c]:
                    if board[r, c] in (0, 1):  # Wild或Wild2被消除
                        wild_eliminated_this_cascade += 1
                    board[r, c] = -1
        total_wild_eliminated += wild_eliminated_this_cascade
        
        # 重力下降
        for r in range(rows):
            non_empty = []
            for c in range(cols):
                if board[r, c] != -1:
                    non_empty.append(board[r, c])
            
            empty_count = cols - len(non_empty)
            for c in range(len(non_empty)):
                board[r, c] = non_empty[c]
            for c in range(len(non_empty), cols):
                board[r, c] = -1
        
        # 補充新符號，并统计新出现的黄金符号
        for r in range(rows):
            for c in range(cols-1, -1, -1):
                if board[r, c] == -1:
                    new_symbols = sample_sequences_weighted_numba(
                        weight_array[r:r+1], reel_array[r:r+1], 1
                    )
                    new_symbol = int(new_symbols[0, 0])
                    board[r, c] = new_symbol
                    if 11 <= new_symbol <= 18:
                        total_golden_symbols += 1
    
    # 返回统计信息
    return total_score, board, cascade_count, total_golden_symbols, total_golden_converted, total_wild_eliminated, cascades_with_golden, ww2_triggered_count, ww2_with_elimination_count

# 包裝函數以支持原有接口
def sample_sequences_weighted(weight_array, content_array, n):
    return sample_sequences_weighted_numba(weight_array, content_array, n)

def build_symbol_to_linkpoint():
    return build_symbol_to_linkpoint_numba()

def waygame_fullscore(board, linkpoint):
    return waygame_fullscore_numba(board.astype(np.int32), linkpoint)

def point(weight_array, reel_array, initial_board, linkpoint, rdwild):
    rdwild_values = np.array(rdwild[0], dtype=np.float64)
    rdwild_weights = np.array(rdwild[1], dtype=np.float64)
    return point_numba(weight_array, reel_array, initial_board.astype(np.int32), 
                       linkpoint, rdwild_values, rdwild_weights)

def freegame_point(weight_array, reel_array, initial_board, linkpoint, rdwild, is_freegame=True):
    rdwild_values = np.array(rdwild[0], dtype=np.float64)
    rdwild_weights = np.array(rdwild[1], dtype=np.float64)
    return freegame_point_numba(weight_array, reel_array, initial_board.astype(np.int32),
                               linkpoint, rdwild_values, rdwild_weights)

@njit
def freegame_numba(n, configs, config_probs, symbol_names_data, high_surface_probs,
                   high_symbol_weight, high_symbol_reels, high_randomwilds,
                   low_symbol_weight, low_symbol_reel, low_randomwild,
                   linkpoint, freecard, max_spins=50, seed=None):
    """
    Numba加速版本的免費遊戲批量執行
    """
    if seed is not None:
        np.random.seed(seed)
    
    total_scores = np.zeros(n, dtype=np.float64)
    non_zero_spin_counts = np.zeros(n, dtype=np.int32)  # 每场freegame中分数不为0的spin次数
    
    # 新增统计数组 - 累加所有spin的数据
    combo_counts = np.zeros(n, dtype=np.int32)  # 每个freegame的总combo数
    golden_symbol_counts = np.zeros(n, dtype=np.int32)  # 每个freegame的黄金符号总数
    golden_converted_counts = np.zeros(n, dtype=np.int32)  # 黄金符号转换为wild的总数
    wild_eliminated_counts = np.zeros(n, dtype=np.int32)  # wild被消除的总数
    cascade_details_matrix = np.zeros((n, 3), dtype=np.int32)  # [有黄金消除次数, ww2触发次数, ww2有消除次数]
    card_values = np.zeros(n, dtype=np.float64)  # 根据分数区间对应的卡片值
    
    for game_round in range(n):
        # 選擇場次配置
        r = np.random.random()
        cumsum = np.cumsum(config_probs)
        chosen_config_idx = 0
        for i in range(len(config_probs)):
            if r <= cumsum[i]:
                chosen_config_idx = i
                break
        
        high_spins, low_spins = configs[chosen_config_idx]
        
        # 建立場次序列並隨機排列
        spin_sequence = np.concatenate((np.ones(high_spins, dtype=np.int32),
                                      np.zeros(low_spins, dtype=np.int32)))
        # Fisher-Yates shuffle
        for i in range(len(spin_sequence)-1, 0, -1):
            j = np.random.randint(0, i+1)
            spin_sequence[i], spin_sequence[j] = spin_sequence[j], spin_sequence[i]
        
        total_score = 0
        current_spin = 0
        non_zero_count = 0  # 统计本场分数不为0的spin次数
        
        # 累加所有spin的统计数据
        game_combo_total = 0
        game_golden_total = 0
        game_golden_conv_total = 0
        game_wild_elim_total = 0
        game_cascades_with_golden = 0
        game_ww2_triggered = 0
        game_ww2_with_elim = 0
        
        while current_spin < len(spin_sequence) and current_spin < max_spins:
            spin_type = spin_sequence[current_spin]
            current_spin += 1
            if spin_type == 1:  # high table
                # 選擇高表符號
                r = np.random.random()
                cumsum = np.cumsum(high_surface_probs)
                chosen_symbol_idx = 0
                for i in range(len(high_surface_probs)):
                    if r <= cumsum[i]:
                        chosen_symbol_idx = i
                        break
                
                # 生成初始版面
                symbol_reel = high_symbol_reels[chosen_symbol_idx]
                initial_board = sample_sequences_weighted_numba(
                    high_symbol_weight, symbol_reel, 4
                ).astype(np.int32)
                
                # 執行遊戲
                randomwild_values = high_randomwilds[chosen_symbol_idx][0]
                randomwild_weights = high_randomwilds[chosen_symbol_idx][1]
                result = freegame_point_numba(
                    low_symbol_weight, symbol_reel, initial_board,
                    linkpoint, randomwild_values, randomwild_weights
                )
                spin_score = result[0]
                final_board = result[1]
                # 收集统计数据
                game_combo_total += result[2]
                game_golden_total += result[3]
                game_golden_conv_total += result[4]
                game_wild_elim_total += result[5]
                game_cascades_with_golden += result[6]
                game_ww2_triggered += result[7]
                game_ww2_with_elim += result[8]
                
            else:  # low table
                initial_board = sample_sequences_weighted_numba(
                    low_symbol_weight, low_symbol_reel, 4
                ).astype(np.int32)
                
                result = freegame_point_numba(
                    low_symbol_weight, low_symbol_reel, initial_board,
                    linkpoint, low_randomwild[0], low_randomwild[1]
                )
                spin_score = result[0]
                final_board = result[1]
                # 收集统计数据
                game_combo_total += result[2]
                game_golden_total += result[3]
                game_golden_conv_total += result[4]
                game_wild_elim_total += result[5]
                game_cascades_with_golden += result[6]
                game_ww2_triggered += result[7]
                game_ww2_with_elim += result[8]
            
            total_score += spin_score
            
            # 统计分数不为0的spin次数
            if spin_score > 0:
                non_zero_count += 1
            
            # 檢查Retrigger
            c1_count = np.sum(final_board == 2)
            if c1_count >= 3:
                # 增加5場: 1場高表 + 4場低表
                additional = np.array([1, 0, 0, 0, 0], dtype=np.int32)
                # 隨機排列
                for i in range(4, 0, -1):
                    j = np.random.randint(0, i+1)
                    additional[i], additional[j] = additional[j], additional[i]
                
                # 擴展序列
                new_sequence = np.concatenate((spin_sequence, additional))
                spin_sequence = new_sequence
        
        total_scores[game_round] = total_score
        non_zero_spin_counts[game_round] = non_zero_count
        
        # 保存累加的统计数据
        combo_counts[game_round] = game_combo_total
        golden_symbol_counts[game_round] = game_golden_total
        golden_converted_counts[game_round] = game_golden_conv_total
        wild_eliminated_counts[game_round] = game_wild_elim_total
        cascade_details_matrix[game_round, 0] = game_cascades_with_golden
        cascade_details_matrix[game_round, 1] = game_ww2_triggered
        cascade_details_matrix[game_round, 2] = game_ww2_with_elim
        
        # 根据分数判断所在区间，返回对应的卡片值
        score_divided = total_score / 100.0
        card_idx = 0
        for j in range(len(freecard[0])):
            if score_divided > freecard[0, j]:
                card_idx = j
            else:
                break
        card_values[game_round] = freecard[1, card_idx]
    
    return (total_scores, non_zero_spin_counts, combo_counts, golden_symbol_counts, 
            golden_converted_counts, wild_eliminated_counts, cascade_details_matrix, card_values)

@njit
def basegame_numba(n, surface_probs, high_symbol_weight, high_symbol_reel, high_randomwild,
                   low_symbol_weight, low_symbol_reel, low_randomwild, linkpoint, basecard, seed=None):
    """
    Numba加速版本的基本遊戲批量執行
    """
    if seed is not None:
        np.random.seed(seed)
    
    total_scores = np.zeros(n, dtype=np.float64)
    c1_trigger_count = 0  # 計算C1>=3的遊戲場次
    c1_trigger_vector = np.zeros(n, dtype=np.int32)  # 記錄每場是否觸發(0或1)
    
    # 新增统计数组
    combo_counts = np.zeros(n, dtype=np.int32)  # 每次spin的combo数
    golden_symbol_counts = np.zeros(n, dtype=np.int32)  # 每次spin出现的黄金符号总数
    golden_converted_counts = np.zeros(n, dtype=np.int32)  # 黄金符号转换为wild的数量
    wild_eliminated_counts = np.zeros(n, dtype=np.int32)  # wild被消除的数量
    cascade_details_matrix = np.zeros((n, 3), dtype=np.int32)  # [有黄金消除次数, ww2触发, ww2有消除]
    card_values = np.zeros(n, dtype=np.float64)  # 根据分数区间对应的卡片值
    
    for i in range(n):
        # 根據比例選擇高表或低表
        use_high_table = np.random.random() < surface_probs[0]
        
        if use_high_table:
            initial_board = sample_sequences_weighted_numba(
                high_symbol_weight, high_symbol_reel, 4
            ).astype(np.int32)
            total_score, final_board, combo_count, golden_total, golden_conv, wild_elim, cascade_det = point_numba(
                high_symbol_weight, high_symbol_reel, initial_board,
                linkpoint, high_randomwild[0], high_randomwild[1]
            )
        else:
            initial_board = sample_sequences_weighted_numba(
                low_symbol_weight, low_symbol_reel, 4
            ).astype(np.int32)
            total_score, final_board, combo_count, golden_total, golden_conv, wild_elim, cascade_det = point_numba(
                low_symbol_weight, low_symbol_reel, initial_board,
                linkpoint, low_randomwild[0], low_randomwild[1]
            )
        
        total_scores[i] = total_score
        combo_counts[i] = combo_count
        golden_symbol_counts[i] = golden_total
        golden_converted_counts[i] = golden_conv
        wild_eliminated_counts[i] = wild_elim
        cascade_details_matrix[i] = cascade_det
        
        # 根据分数判断所在区间，返回对应的卡片值
        score_divided = total_score / 100.0
        card_idx = 0
        for j in range(len(basecard[0])):
            if score_divided > basecard[0, j]:
                card_idx = j
            else:
                break
        card_values[i] = basecard[1, card_idx]
        
        # 檢查最終版面C1數量
        c1_count = np.sum(final_board == 2)
        if c1_count >= 3:
            c1_trigger_count += 1
            c1_trigger_vector[i] = 1
    
    return (total_scores, c1_trigger_count, c1_trigger_vector, combo_counts, 
            golden_symbol_counts, golden_converted_counts, wild_eliminated_counts, cascade_details_matrix, card_values)

# 包裝函數以支持原有接口
def freegame(n, trigger_c1_count=3, freeGameSurface=None, configs=None,
             freeGameHighSurface=None, linkpoint=None,
             freeGameHighSymbolWeight=None, freeGameLowSymbolWeight=None,
             freeGameSymbolHighA=None, freeGameSymbolHighK=None,
             freeGameSymbolHighQ=None, freeGameSymbolHighJ=None,
             freeGameSymbolLow=None,
             freeGameHighRandomwildA=None, freeGameHighRandomwildK=None,
             freeGameHighRandomwildQ=None, freeGameHighRandomwildJ=None,
             freeGameLowRandomwild=None, freecard=None):
    """
    原始接口的包裝函數
    """
    # 預設值處理
    if freeGameSurface is None:
        freeGameSurface = globals().get('freeGameSurface', [1, 1, 1, 1])[2:3] if 'freeGameSurface' in globals() else [1, 1, 1, 1]
    if freeGameHighSurface is None:
        freeGameHighSurface = globals().get('freeGameHighSurface', [1, 1, 1, 1])
    if linkpoint is None:
        linkpoint = globals().get('linkpoint', np.ones((8,3), dtype=np.float64))
    if freecard is None:
        freecard = np.array([[-1,0,1,2,3,4,5,6,7,8,9,10,15,20,25,30,35,40,45,50,60,70,80,90,100,120,140,160,180,200,250,300,350,400,450,500,550,600,650,700,750,800,850,900,950,1000,2000],
                            [0,0,0,0,0,0,0,0,0,0,0,1,1.035,1.2,1.2,1.2,1.2,0.5,0.1,0.4,0.1,0.02,0.01,0.01,0.02,0.001,0.0001,0.0001,0.0001,0.00001,0.000001,0.000001,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]])
    
    # 準備數據
    config_probs = np.array(freeGameSurface, dtype=np.float64)
    config_probs = config_probs / np.sum(config_probs)
    
    high_surface_probs = np.array(freeGameHighSurface, dtype=np.float64)
    high_surface_probs = high_surface_probs / np.sum(high_surface_probs)
    
    # 處理符號參數預設值
    if freeGameSymbolHighA is None:
        freeGameSymbolHighA = globals().get('freeGameSymbolHighA', np.zeros((5, 64), dtype=np.int32))
    if freeGameSymbolHighK is None:
        freeGameSymbolHighK = globals().get('freeGameSymbolHighK', np.zeros((5, 64), dtype=np.int32))
    if freeGameSymbolHighQ is None:
        freeGameSymbolHighQ = globals().get('freeGameSymbolHighQ', np.zeros((5, 64), dtype=np.int32))
    if freeGameSymbolHighJ is None:
        freeGameSymbolHighJ = globals().get('freeGameSymbolHighJ', np.zeros((5, 64), dtype=np.int32))
    
    # 高表符號數據
    high_symbol_reels = [
        freeGameSymbolHighA,
        freeGameSymbolHighK,
        freeGameSymbolHighQ,
        freeGameSymbolHighJ
    ]
    
    # 處理參數預設值
    if freeGameHighRandomwildA is None:
        freeGameHighRandomwildA = globals().get('freeGameHighRandomwildA', [[0], [1]])
    if freeGameHighRandomwildK is None:
        freeGameHighRandomwildK = globals().get('freeGameHighRandomwildK', [[0], [1]])
    if freeGameHighRandomwildQ is None:
        freeGameHighRandomwildQ = globals().get('freeGameHighRandomwildQ', [[0], [1]])
    if freeGameHighRandomwildJ is None:
        freeGameHighRandomwildJ = globals().get('freeGameHighRandomwildJ', [[0], [1]])
    
    high_randomwilds = [
        (np.array(freeGameHighRandomwildA[0], dtype=np.float64),
         np.array(freeGameHighRandomwildA[1], dtype=np.float64)),
        (np.array(freeGameHighRandomwildK[0], dtype=np.float64),
         np.array(freeGameHighRandomwildK[1], dtype=np.float64)),
        (np.array(freeGameHighRandomwildQ[0], dtype=np.float64),
         np.array(freeGameHighRandomwildQ[1], dtype=np.float64)),
        (np.array(freeGameHighRandomwildJ[0], dtype=np.float64),
         np.array(freeGameHighRandomwildJ[1], dtype=np.float64))
    ]
    
    # 處理其他參數的預設值
    if freeGameLowSymbolWeight is None:
        freeGameLowSymbolWeight = globals().get('freeGameLowSymbolWeight', np.ones((5, 64), dtype=np.float64))
    if freeGameSymbolLow is None:
        freeGameSymbolLow = globals().get('freeGameSymbolLow', np.zeros((5, 64), dtype=np.int32))
    if freeGameLowRandomwild is None:
        freeGameLowRandomwild = globals().get('freeGameLowRandomwild', [[0], [1]])
    if freeGameHighSymbolWeight is None:
        freeGameHighSymbolWeight = globals().get('freeGameHighSymbolWeight', np.ones((5, 64), dtype=np.float64))
    
    # 低表數據
    low_symbol_weight = freeGameLowSymbolWeight
    low_symbol_reel = freeGameSymbolLow
    low_randomwild = (
        np.array(freeGameLowRandomwild[0], dtype=np.float64),
        np.array(freeGameLowRandomwild[1], dtype=np.float64)
    )
    
    high_symbol_weight = freeGameHighSymbolWeight
    
    results = freegame_numba(
        n, configs, config_probs, None, high_surface_probs,
        high_symbol_weight, high_symbol_reels, high_randomwilds,
        low_symbol_weight, low_symbol_reel, low_randomwild,
        linkpoint, freecard
    )
    
    total_scores = results[0]
    non_zero_spin_counts = results[1]
    combo_counts = results[2]
    golden_symbol_counts = results[3]
    golden_converted_counts = results[4]
    wild_eliminated_counts = results[5]
    cascade_details_matrix = results[6]
    card_values = results[7]
    
    return (total_scores, non_zero_spin_counts, combo_counts, golden_symbol_counts,
            golden_converted_counts, wild_eliminated_counts, cascade_details_matrix, card_values)

def basegame(n, baseGameSurface, baseGameHighSymbolWeight=None, baseGameSymbolHigh=None, 
             baseGameHighRandomwild=None, baseGameLowSymbolWeight=None, baseGameSymbolLow=None,
             baseGameLowRandomwild=None, linkpoint=None, basecard=None):
    """
    原始接口的包裝函數 - 修正為優先使用輸入參數
    """
    surface_probs = np.array(baseGameSurface, dtype=np.float64)
    surface_probs = surface_probs / np.sum(surface_probs)
    
    # 處理參數預設值，優先使用輸入參數
    if baseGameHighSymbolWeight is None:
        baseGameHighSymbolWeight = globals().get('baseGameHighSymbolWeight', np.ones((5, 64), dtype=np.float64))
    if baseGameSymbolHigh is None:
        baseGameSymbolHigh = globals().get('baseGameSymbolHigh', np.zeros((5, 64), dtype=np.int32))
    if baseGameHighRandomwild is None:
        baseGameHighRandomwild = globals().get('baseGameHighRandomwild', [[0], [1]])
    if baseGameLowSymbolWeight is None:
        baseGameLowSymbolWeight = globals().get('baseGameLowSymbolWeight', np.ones((5, 64), dtype=np.float64))
    if baseGameSymbolLow is None:
        baseGameSymbolLow = globals().get('baseGameSymbolLow', np.zeros((5, 64), dtype=np.int32))
    if baseGameLowRandomwild is None:
        baseGameLowRandomwild = globals().get('baseGameLowRandomwild', [[0], [1]])
    if linkpoint is None:
        linkpoint = globals().get('linkpoint', np.ones((8,3), dtype=np.float64))
    if basecard is None:
        basecard = np.array([[-1,0,1,2,3,4,5,6,7,8,9,10,15,20,25,30,35,40,45,50,60,70,80,90,100,120,140,160,180,200,250,300,350,400,450,500,550,600,650,700,750,800,850,900,950,1000,2000],
                            [1.7108,0.865,0.58,0.52,0.53,0.5,0.3,0.3,0.3,0.28,0.32,0.14,0.1,0.07,0.05,0.02,0.01,0.03,0.05,0.0016,0.02,0.04,0.04,0.1,0.02,0.291,0.001,0.001,0.001,0.001,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]])
    
    # 使用輸入參數
    high_symbol_weight = baseGameHighSymbolWeight
    high_symbol_reel = baseGameSymbolHigh
    high_randomwild = (
        np.array(baseGameHighRandomwild[0], dtype=np.float64),
        np.array(baseGameHighRandomwild[1], dtype=np.float64)
    )
    
    low_symbol_weight = baseGameLowSymbolWeight
    low_symbol_reel = baseGameSymbolLow
    low_randomwild = (
        np.array(baseGameLowRandomwild[0], dtype=np.float64),
        np.array(baseGameLowRandomwild[1], dtype=np.float64)
    )
    
    results = basegame_numba(
        n, surface_probs, high_symbol_weight, high_symbol_reel, high_randomwild,
        low_symbol_weight, low_symbol_reel, low_randomwild, linkpoint, basecard
    )
    
    total_scores = results[0]
    c1_trigger_count = results[1]
    c1_trigger_vector = results[2]
    combo_counts = results[3]
    golden_symbol_counts = results[4]
    golden_converted_counts = results[5]
    wild_eliminated_counts = results[6]
    cascade_details_matrix = results[7]
    card_values = results[8]
    
    return (total_scores, c1_trigger_count, c1_trigger_vector, combo_counts,
            golden_symbol_counts, golden_converted_counts, wild_eliminated_counts, 
            cascade_details_matrix, card_values)

# 多線程加速函數


def fullgame(n, 
             # Base Game 參數
             baseGameSurface=None, baseGameHighSymbolWeight=None, baseGameSymbolHigh=None,
             baseGameHighRandomwild=None, baseGameLowSymbolWeight=None, baseGameSymbolLow=None,
             baseGameLowRandomwild=None,
             # Free Game 參數
             trigger_c1_count=3, freeGameSurface=None, configs=None,
             freeGameHighSurface=None, freeGameHighSymbolWeight=None, freeGameLowSymbolWeight=None,
             freeGameSymbolHighA=None, freeGameSymbolHighK=None, freeGameSymbolHighQ=None, 
             freeGameSymbolHighJ=None, freeGameSymbolLow=None,
             freeGameHighRandomwildA=None, freeGameHighRandomwildK=None,
             freeGameHighRandomwildQ=None, freeGameHighRandomwildJ=None,
             freeGameLowRandomwild=None,
             # 共用參數
             linkpoint=None):
    """
    完整遊戲模擬函數 - 簡化版，只返回核心統計
    
    Args: 
        n: 模擬的 base game 場次
        其他參數: 與 basegame 和 freegame 函數相同
    
    Returns:
        dict: {
            'total_scores': 每場的總分數 (base + free if triggered),
            'freegame_triggered': 觸發 free game 的總次數,
            'trigger_rate': 觸發率 (%),
            'avg_total': 平均總分
        }
    """
    # 處理參數預設值
    if baseGameSurface is None:
        baseGameSurface = globals().get('baseGameSurface', [1, 0])
    if linkpoint is None:
        linkpoint = globals().get('linkpoint', np.ones((8,3), dtype=np.float64))
    if configs is None:
        configs = globals().get('configs', np.array([[(1, 9), (4, 6), (5, 5), (6, 4)]], dtype=np.int32))
    
    # 執行 n 場 Base Game，一次性取得所有結果
    basegame_results = basegame(n, baseGameSurface, 
                                baseGameHighSymbolWeight=baseGameHighSymbolWeight,
                                baseGameSymbolHigh=baseGameSymbolHigh,
                                baseGameHighRandomwild=baseGameHighRandomwild,
                                baseGameLowSymbolWeight=baseGameLowSymbolWeight,
                                baseGameSymbolLow=baseGameSymbolLow,
                                baseGameLowRandomwild=baseGameLowRandomwild,
                                linkpoint=linkpoint)
    
    base_scores = basegame_results[0]
    total_c1_triggers = basegame_results[1]
    
    # 初始化結果
    total_scores = base_scores.copy()  # 先複製 base game 分數
    freegame_triggered = total_c1_triggers  # 使用 basegame 返回的觸發次數
    
    # 如果有觸發 Free Game，執行對應次數的 Free Game
    if freegame_triggered > 0:
        # 執行等量的 Free Game
        free_results = freegame(freegame_triggered, trigger_c1_count=trigger_c1_count,
                                freeGameSurface=freeGameSurface,
                                configs=configs,
                                freeGameHighSurface=freeGameHighSurface,
                                linkpoint=linkpoint,
                                freeGameHighSymbolWeight=freeGameHighSymbolWeight,
                                freeGameLowSymbolWeight=freeGameLowSymbolWeight,
                                freeGameSymbolHighA=freeGameSymbolHighA,
                                freeGameSymbolHighK=freeGameSymbolHighK,
                                freeGameSymbolHighQ=freeGameSymbolHighQ,
                                freeGameSymbolHighJ=freeGameSymbolHighJ,
                                freeGameSymbolLow=freeGameSymbolLow,
                                freeGameHighRandomwildA=freeGameHighRandomwildA,
                                freeGameHighRandomwildK=freeGameHighRandomwildK,
                                freeGameHighRandomwildQ=freeGameHighRandomwildQ,
                                freeGameHighRandomwildJ=freeGameHighRandomwildJ,
                                freeGameLowRandomwild=freeGameLowRandomwild)
        free_scores = free_results[0]
        
        # 找出哪些 base game 觸發了 free game，並加上對應的 free game 分數
        # 簡化處理：假設觸發是隨機分佈的
        triggered_indices = np.random.choice(n, freegame_triggered, replace=False)
        for i, idx in enumerate(triggered_indices):
            total_scores[idx] += free_scores[i]
    
    # 計算統計
    trigger_rate = freegame_triggered / n * 100
    avg_total = np.mean(total_scores)
    
    return {
        'total_scores': total_scores,
        'freegame_triggered': freegame_triggered,
        'trigger_rate': trigger_rate,
        'avg_total': avg_total
    }






#%%
b=np.array([[0,2,3,4],[13128,2000,500,200]])
d = freegame(1000000, trigger_c1_count=3, 
                    freeGameSurface=freeGameSurface[2], 
                    configs=[(10, 0), (10, 0), (10, 0), (10, 0),(10,0),(10,0)],
                    freeGameHighSurface=np.array(freeGameHighSurface), 
                    linkpoint=linkpoint,
                    freeGameHighSymbolWeight=freeGameLowSymbolWeight, 
                    freeGameLowSymbolWeight=freeGameLowSymbolWeight,
                    freeGameSymbolHighA=freeGameSymbolHighA, 
                    freeGameSymbolHighK=freeGameSymbolHighK,
                    freeGameSymbolHighQ=freeGameSymbolHighQ, 
                    freeGameSymbolHighJ=freeGameSymbolHighJ,
                    freeGameSymbolLow=freeGameSymbolLow,
                    freeGameHighRandomwildA=b, 
                    freeGameHighRandomwildK=b,
                    freeGameHighRandomwildQ=b, 
                    freeGameHighRandomwildJ=b,
                    freeGameLowRandomwild=b)
#%%
b=np.array([[0,2,3,4],[0,505,505,380]])
a = freegame(100000, trigger_c1_count=3, 
                    freeGameSurface=freeGameSurface[2], 
                    configs=[(10, 0), (10, 0), (10, 0), (10, 0),(10,0),(10,0)],
                    freeGameHighSurface=np.array(freeGameHighSurface), 
                    linkpoint=linkpoint,
                    freeGameHighSymbolWeight=freeGameLowSymbolWeight, 
                    freeGameLowSymbolWeight=freeGameLowSymbolWeight,
                    freeGameSymbolHighA=superfreeGameSymbol, 
                    freeGameSymbolHighK=superfreeGameSymbol,
                    freeGameSymbolHighQ=superfreeGameSymbol, 
                    freeGameSymbolHighJ=superfreeGameSymbol,
                    freeGameSymbolLow=superfreeGameSymbol,
                    freeGameHighRandomwildA=b, 
                    freeGameHighRandomwildK=b,
                    freeGameHighRandomwildQ=b,
                    freeGameHighRandomwildJ=b,
                    freeGameLowRandomwild=b)
# %%
a=basegame(10000000, np.array([1, 0]), baseGameHighSymbolWeight=baseGameHighSymbolWeight, baseGameSymbolHigh=baseGameSymbolHigh, 
             baseGameHighRandomwild=baseGameHighRandomwild, baseGameLowSymbolWeight=baseGameLowSymbolWeight, baseGameSymbolLow=baseGameSymbolLow,
             baseGameLowRandomwild=baseGameLowRandomwild, linkpoint=linkpoint)
# %%
# 分數區間定義: (下界, 上界]
score_bins = [
    -1, 0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10,  # 0-10: 每1單位
    15, 20, 25, 30, 35, 40, 45, 50,        # 10-50: 每5單位
    60, 70, 80, 90, 100,                    # 50-100: 每10單位
    120, 140, 160, 180, 200,                # 100-200: 每20單位
    250, 300, 350, 400, 450, 500,           # 200-500: 每50單位
    550, 600, 650, 700, 750, 800, 850, 900, 950, 1000,  # 500-1000: 每50單位
    2000                                     # 1000-2000
]

# 統計 a[0] 在各區間的分布（使用正确的左开右闭区间）
bin_indices_check = np.digitize(a[0]/100, bins=score_bins, right=True) - 1
bin_indices_check = np.clip(bin_indices_check, 0, len(score_bins) - 2)
counts = np.bincount(bin_indices_check, minlength=len(score_bins)-1)
total_count = len(a[0])
d1 = counts / total_count * 100  # 轉換為百分比

# 顯示每個區間的統計結果
print("區間統計 (左开右闭区间: 下界 < x <= 上界):")
for i in range(len(d1)):
    print(f"{score_bins[i]:6.0f} < x <= {score_bins[i+1]:6.0f}: {d1[i]:12.10f}%")

# %%
# 目标区间比例 (%)
target_proportions = np.array([
74.5420038000,
15.5622719000,
2.5957446000,
1.4797656000,
1.0335466000,
0.8921352000,
0.4284193000,
0.3878489000,
0.3287971000,
0.2981387000,
0.2951830000,
0.4994715000,
0.2562485000,
0.1351312000,
0.0761451000,
0.0134682000,
0.0050361000,
0.0168187000,
0.0351391000,
0.0011207000,
0.0166235000,
0.0252721000,
0.0193514000,
0.0376180000,
0.0106576000,
0.0102027000,
0.0022106000,
0.0014621000,
0.0010088000,
0.0013576000,
0.0000000000,
0.0000000000,
0.0000000000,
0.0000000000,
0.0000000000,
0.0000000000,
0.0000000000,
0.0000000000,
0.0000000000,
0.0000000000,
0.0000000000,
0.0000000000,
0.0000000000,
0.0000000000,
0.0000000000,
0.0000000000
]) / 100  # 转换为比例

# 将 a[0] 按区间分组
# right=True 表示区间为 (bins[i-1], bins[i]]，即下界 < x <= 上界
bin_indices = np.digitize(a[0]/100, bins=score_bins, right=True) - 1
bin_indices = np.clip(bin_indices, 0, len(score_bins) - 2)  # 确保索引在有效范围内

# 为每个区间创建索引列表
bin_groups = [np.where(bin_indices == i)[0] for i in range(len(target_proportions))]

# 调试：检查区间分组是否正确
print("区间分组调试信息:")
print(f"score_bins 长度: {len(score_bins)}")
print(f"target_proportions 长度: {len(target_proportions)}")
print(f"区间数量: {len(score_bins) - 1}")
sample_values = [a[0][0] if len(a[0]) > 0 else 0, 
                 np.min(a[0]) if len(a[0]) > 0 else 0,
                 np.max(a[0]) if len(a[0]) > 0 else 0]
print(f"样本数据: 第一个={sample_values[0]:.2f}, 最小={sample_values[1]:.2f}, 最大={sample_values[2]:.2f}")
print()

print("原始各区间数据量:")
for i in range(len(target_proportions)):
    if target_proportions[i] > 0:
        print(f"区间 {i} ({score_bins[i]:.0f}, {score_bins[i+1]:.0f}]: "
              f"可用={len(bin_groups[i])}, 目标比例={target_proportions[i]*100:.4f}%")

# 使用迭代方法找到最优的 target_total
# 从原始数据量开始，逐步降低直到所有区间都能满足
target_total = len(a[0])
max_iterations = 100

for iteration in range(max_iterations):
    all_satisfied = True
    for i, proportion in enumerate(target_proportions):
        if proportion > 1e-10:
            target_count = int(np.round(target_total * proportion))
            available = len(bin_groups[i])
            if target_count > available:
                # 这个区间不满足，需要降低 target_total
                all_satisfied = False
                # 根据这个区间的限制重新计算 target_total
                new_target = int(available / proportion)
                target_total = min(target_total, new_target)
                break
    
    if all_satisfied:
        break

print(f"\n原始数据量: {len(a[0])}")
print(f"最终确定抽样总数: {target_total}")

# 根据目标比例从各区间精确抽样（不重复）
a1_list = []
actual_counts = []
for i, proportion in enumerate(target_proportions):
    target_count = int(np.round(target_total * proportion))
    bin_data = bin_groups[i]
    
    if target_count > 0 and len(bin_data) > 0:
        # 此时应该能满足，但仍然做保护
        actual_count = min(target_count, len(bin_data))
        sampled_indices = np.random.choice(bin_data, size=actual_count, replace=False)
        a1_list.append(a[0][sampled_indices])
        actual_counts.append(actual_count)
        if actual_count != target_count:
            print(f"警告：区间 {i} 目标={target_count}, 实际={actual_count}")
    else:
        actual_counts.append(0)

print(f"\n各区间实际抽样数:")
for i, count in enumerate(actual_counts):
    if count > 0:
        print(f"区间 {i}: {count} 个")

# 合并所有区间的数据
a1 = np.concatenate(a1_list) if a1_list else np.array([])

# 随机打乱 a1 的顺序
if len(a1) > 0:
    np.random.shuffle(a1)

# 验证结果
print(f"\n调整后 a1 长度: {len(a1)}")
if len(a1) > 0:
    print(f"\n各区间比例验证:")
    counts_a1, _ = np.histogram(a1, bins=score_bins)
    actual_proportions = counts_a1 / len(a1) * 100
    
    print(f"{'区间':<4} {'下界':>8} {'上界':>8} {'目标比例%':>15} {'实际比例%':>15} {'误差':>10}")
    print("-" * 70)
    for i in range(len(target_proportions)):
        target_pct = target_proportions[i] * 100
        actual_pct = actual_proportions[i]
        error = actual_pct - target_pct
        if target_pct > 0 or actual_pct > 0:
            print(f"{i:>4} {score_bins[i]:>8.0f} {score_bins[i+1]:>8.0f} "
                  f"{target_pct:>15.10f} {actual_pct:>15.10f} {error:>10.6f}")
    
    print(f"\n总和验证: {np.sum(actual_proportions):.10f}% (应该是100%)")
    print(f"各区间数据总和: {np.sum(counts_a1)} (应该等于 {len(a1)})")


# %%
# 从 a1 随机抽样并计算平均值
# 每次抽取 100000 个数据，计算平均值，重复 600 次
sample_size = 100000
num_samples = 600

# 使用 a1 作为数据源（如果需要其他数据源，请修改这里）
c1 = a1  # 或者根据需要定义为其他数据

c1_means = np.zeros(num_samples)
for i in range(num_samples):
    # 从 c1 中随机抽取 sample_size 个数据（可重复抽样）
    sample = np.random.choice(c1, size=sample_size, replace=True)
    c1_means[i] = np.mean(sample)
    
    if (i + 1) % 100 == 0:
        print(f"已完成 {i+1}/{num_samples} 次抽样")

print(f"\n抽样完成！")
print(f"生成了 {len(c1_means)} 个平均值")
print(f"c1_means 的统计信息:")
print(f"  平均值: {np.mean(c1_means):.4f}")
print(f"  标准差: {np.std(c1_means):.4f}")
print(f"  最小值: {np.min(c1_means):.4f}")
print(f"  最大值: {np.max(c1_means):.4f}")

# 将结果写入 Excel 文件
import openpyxl

excel_path = r"C:\Users\lingyuho\Downloads\Telegram Desktop\Super_Ace_RTP_Analysis.xlsx"
sheet_name = "Bet_0.5_RTP"

print(f"\n正在写入 Excel 文件...")
try:
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(excel_path)
    
    # 选择工作表
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        print(f"警告：工作表 '{sheet_name}' 不存在，将创建新工作表")
        ws = wb.create_sheet(sheet_name)
    
    # 将 c1_means 写入 C2:C601
    for i, value in enumerate(c1_means):
        ws[f'C{i+2}'] = value
    
    # 保存文件
    wb.save(excel_path)
    wb.close()
    
    print(f"成功！已将 {len(c1_means)} 个数值写入 {sheet_name} 的 C2:C601")
    
except FileNotFoundError:
    print(f"错误：找不到文件 {excel_path}")
except Exception as e:
    print(f"错误：{e}")

# %%
