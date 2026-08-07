# -*- coding: utf-8 -*-
"""Write measured simulation results into the H045 Overview sheet.

Usage:
    py fill_overview.py <workbook.xlsx> <results.json>

The JSON holds one entry per bet type::

    {
      "Normal Bet (新手 / Newbie)": {"total": 0.93, "bg": 0.68, "fg": 0.25,
                                     "hit": 0.257, "cycle": 118.0},
      ...
    }
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


def fill(workbook_path: Path, results: dict[str, dict[str, float]]) -> None:
    wb = load_workbook(workbook_path)
    ws = wb["Overview"]

    # Total RTP column of the "Coin in / Price(x) / Total RTP / Bet Type" block.
    for row in range(11, 16):
        label = str(ws.cell(row, 4).value or "").strip()
        if label in results:
            ws.cell(row, 3, results[label]["total"])

    # Per-bet-type Pay Back / Hit% / Pulls-per-hit blocks.
    for row in range(17, 30):
        label = str(ws.cell(row, 1).value or "").strip()
        if label not in results:
            continue
        entry = results[label]
        ws.cell(row, 2, entry["bg"])
        ws.cell(row, 3, entry["hit"])
        ws.cell(row, 4, entry["cycle"])
        ws.cell(row + 1, 2, entry["fg"])

    wb.save(workbook_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Fill measured RTP into the H045 Overview")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("results", type=Path)
    args = parser.parse_args()
    fill(args.workbook, json.loads(args.results.read_text(encoding="utf-8")))
    print(f"updated {args.workbook}")


if __name__ == "__main__":
    main()
