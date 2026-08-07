"""Build H045 runtime data from the H045 PARsheet layout.

The source xlsx files remain authoritative.  This module is shared by
Simulator.py and the browser-config exporter so Python and index.html use the
same symbol tables and weights.

The workbook follows the H026-style layout:

    Overview / Description / Parameter /
    Multiplier_Weight / Multiplier_Weight_Newbie / Multiplier_Weight_Oldhand /
    OP Jackpot / BG_Symbol .. BF_Symbol

H045 has no Super Feature, so there is no ``super`` table and no
``super_buy_price``.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SYMBOL_TO_ID = {
    "WW1": 0,
    "WW2": 1,
    "C1": 2,
    "M1": 3,
    "M2": 4,
    "M3": 5,
    "M4": 6,
    "A": 7,
    "K": 8,
    "Q": 9,
    "J": 10,
    "M1G": 11,
    "M2G": 12,
    "M3G": 13,
    "M4G": 14,
    "AG": 15,
    "KG": 16,
    "QG": 17,
    "JG": 18,
}
ID_TO_SYMBOL = {value: key for key, value in SYMBOL_TO_ID.items()}

STRIP_LEN = 400
STRIP_FIRST_ROW = 4
STRIP_SYMBOL_COL = 11  # K
STRIP_WEIGHT_COL = 23  # W

# config table key -> worksheet name
TABLE_LAYOUT = {
    "bg_high": "BG_Symbol",
    "bg_low": "BG_Symbol (2)",
    "fg_high_a": "FG_Symbol",
    "fg_high_k": "FG_Symbol (2)",
    "fg_high_q": "FG_Symbol (3)",
    "fg_high_j": "FG_Symbol (4)",
    "fg_low": "FG_Symbol (5)",
    "buy": "BF_Symbol",
}

# Multiplier_Weight_* block title -> card-system section key
CARD_SECTION_TITLES = {
    "Base Game (Normal Bet)": "base_game",
    "Free Game": "free_game",
    "Buy Feature": "buy_feature",
}

PARAM_COL = 2  # Parameter/Multiplier_Weight blocks start in column B


def _number(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _parse_multiplier(value: Any) -> int:
    text = str(value or "").strip().upper().replace("×", "X")
    if text.startswith("X"):
        text = text[1:]
    return int(float(text))


def _find_block(sheet, title: str, column: int = PARAM_COL) -> int:
    """Return the row holding ``title`` in ``column``."""
    for row in range(1, sheet.max_row + 1):
        if str(sheet.cell(row, column).value or "").strip() == title:
            return row
    raise KeyError(f"{sheet.title}: block {title!r} not found")


def _read_keyed_rows(sheet, title: str, width: int) -> dict[str, list[Any]]:
    """Read a ``title`` block whose first data column is a worksheet name."""
    header = _find_block(sheet, title) + 1
    rows: dict[str, list[Any]] = {}
    row = header + 1
    while True:
        key = str(sheet.cell(row, PARAM_COL).value or "").strip()
        if not key or key.startswith("※"):
            break
        rows[key] = [sheet.cell(row, PARAM_COL + 1 + i).value for i in range(width)]
        row += 1
    return rows


def _read_parameter(workbook) -> dict[str, Any]:
    sheet = workbook["Parameter"]

    wild_header = _find_block(sheet, "Random Wild Weight (Weight/Total)") + 1
    wild_values = []
    col = PARAM_COL + 1
    while sheet.cell(wild_header, col).value not in (None, ""):
        wild_values.append(int(_number(sheet.cell(wild_header, col).value)))
        col += 1
    wild_rows = _read_keyed_rows(sheet, "Random Wild Weight (Weight/Total)", len(wild_values))

    ladder_rows = _read_keyed_rows(sheet, "Win Multiplier Ladder", 5)

    mix_header = _find_block(sheet, "Free Game Table Mix Weight") + 1
    groups: dict[str, list[dict[str, Any]]] = {}
    row = mix_header + 1
    while str(sheet.cell(row, PARAM_COL).value or "").strip():
        group = str(sheet.cell(row, PARAM_COL).value).strip()
        weight = max(0.0, _number(sheet.cell(row, PARAM_COL + 3).value))
        if weight:
            groups.setdefault(group, []).append(
                {
                    "high": int(_number(sheet.cell(row, PARAM_COL + 1).value)),
                    "low": int(_number(sheet.cell(row, PARAM_COL + 2).value)),
                    "weight": weight,
                }
            )
        else:
            groups.setdefault(group, [])
        row += 1

    high_rows = _read_keyed_rows(sheet, "Free Game High Table Weight", 2)
    high_variant_weights = [max(0.0, _number(value[1], 1.0)) for value in high_rows.values()]
    if not any(high_variant_weights):
        high_variant_weights = [1.0] * len(high_rows)

    setting_header = _find_block(sheet, "Feature Setting") + 1
    settings: dict[str, Any] = {}
    row = setting_header + 1
    while str(sheet.cell(row, PARAM_COL).value or "").strip():
        settings[str(sheet.cell(row, PARAM_COL).value).strip()] = sheet.cell(row, PARAM_COL + 1).value
        row += 1

    return {
        "wild_values": wild_values,
        "wild_rows": wild_rows,
        "ladder_rows": ladder_rows,
        "free_game_groups": groups,
        "high_variant_weights": high_variant_weights,
        "settings": settings,
    }


def _read_table(workbook, sheet_name: str, parameter: dict[str, Any]) -> dict[str, Any]:
    sheet = workbook[sheet_name]
    reels: list[list[int]] = [[] for _ in range(5)]
    weights: list[list[float]] = [[] for _ in range(5)]

    for offset in range(STRIP_LEN):
        row = STRIP_FIRST_ROW + offset
        for reel in range(5):
            raw_symbol = sheet.cell(row, STRIP_SYMBOL_COL + reel).value
            symbol_name = str(raw_symbol or "").strip()
            if symbol_name not in SYMBOL_TO_ID:
                raise ValueError(
                    f"{sheet_name}!{sheet.cell(row, STRIP_SYMBOL_COL + reel).coordinate}: "
                    f"unknown symbol {raw_symbol!r}"
                )
            reels[reel].append(SYMBOL_TO_ID[symbol_name])
            weights[reel].append(max(0.0, _number(sheet.cell(row, STRIP_WEIGHT_COL + reel).value)))

    wild_weights = parameter["wild_rows"][sheet_name]
    ladder = parameter["ladder_rows"][sheet_name]

    return {
        "reels": reels,
        "weights": weights,
        "random_wild": {
            "values": list(parameter["wild_values"]),
            "weights": [max(0.0, _number(weight)) for weight in wild_weights],
        },
        "multipliers": [_parse_multiplier(value) for value in ladder if value not in (None, "")],
    }


def _read_paytable(workbook) -> dict[str, list[float]]:
    sheet = workbook["Overview"]
    header = None
    for row in range(1, sheet.max_row + 1):
        if str(sheet.cell(row, 1).value or "").strip() == "Symbol":
            header = row
            break
    if header is None:
        raise KeyError("Overview: pay table header not found")

    pays: dict[str, list[float]] = {}
    row = header + 1
    while True:
        name = str(sheet.cell(row, 1).value or "").strip()
        if not name:
            break
        if name in SYMBOL_TO_ID and SYMBOL_TO_ID[name] in range(3, 11):
            pays[str(SYMBOL_TO_ID[name])] = [
                _number(sheet.cell(row, 3).value) / 100.0,
                _number(sheet.cell(row, 4).value) / 100.0,
                _number(sheet.cell(row, 5).value) / 100.0,
            ]
        row += 1
    return pays


def _read_card_sheet(workbook, sheet_name: str) -> dict[str, list[dict[str, Any]]]:
    sheet = workbook[sheet_name]
    sections: dict[str, list[dict[str, Any]]] = {}
    for title, key in CARD_SECTION_TITLES.items():
        header = _find_block(sheet, title) + 1
        cards: list[dict[str, Any]] = []
        row = header + 1
        while True:
            lower = sheet.cell(row, PARAM_COL).value
            if lower in (None, "") or str(lower).strip() == "Total":
                break
            table_code = str(sheet.cell(row, PARAM_COL + 3).value or "").strip()
            weight = max(0.0, _number(sheet.cell(row, PARAM_COL + 8).value))
            if weight:
                if str(lower).strip() == "FG Trigger":
                    cards.append({"type": "free_game", "table": table_code, "weight": weight})
                else:
                    cards.append(
                        {
                            "type": "range",
                            "min": _number(lower, -1.0),
                            "max": _number(sheet.cell(row, PARAM_COL + 1).value),
                            "table": table_code,
                            "weight": weight,
                        }
                    )
            row += 1
        sections[key] = cards
    return sections


def _read_card_system(workbook) -> dict[str, Any]:
    return {
        "enabled": True,
        "default_profile": "weight_2",
        "profiles": {
            "weight_1": _read_card_sheet(workbook, "Multiplier_Weight_Newbie"),
            "weight_2": _read_card_sheet(workbook, "Multiplier_Weight_Oldhand"),
        },
    }


def load_game_config(xlsx_path: str | Path) -> dict[str, Any]:
    path = Path(xlsx_path).resolve()
    # Normal mode is intentionally used here. The parser performs many indexed
    # cell lookups across a dozen sheets; read-only mode reparses rows
    # repeatedly and is dramatically slower for this access pattern.
    workbook = load_workbook(path, read_only=False, data_only=True)

    parameter = _read_parameter(workbook)
    tables = {
        key: _read_table(workbook, sheet_name, parameter)
        for key, sheet_name in TABLE_LAYOUT.items()
    }
    settings = parameter["settings"]

    parsheet = path.stem
    rtp = 92 if "192" in parsheet else 94 if "194" in parsheet else None
    return {
        "game_id": "H045",
        "parsheet_id": parsheet,
        "name_zh": "幸運王牌2",
        "name_en": "Lucky Ace",
        "rtp_label": rtp,
        "reel_num": int(_number(settings.get("Reel Num"), 5)),
        "window_size": int(_number(settings.get("Visible Window Size"), 4)),
        "max_ways": int(_number(settings.get("Max Ways"), 1024)),
        "symbol_names": {str(key): value for key, value in ID_TO_SYMBOL.items()},
        "pays": _read_paytable(workbook),
        "tables": tables,
        "base_table_weights": {"high": 1.0, "low": 0.0},
        "card_system": _read_card_system(workbook),
        "free_game_mix": {
            "groups": parameter["free_game_groups"],
            "high_variant_weights": parameter["high_variant_weights"],
        },
        "free_spins": int(_number(settings.get("Free Spins"), 10)),
        "retrigger_spins": int(_number(settings.get("Retrigger Spins"), 5)),
        "retrigger_high": int(_number(settings.get("Retrigger 高表場數"), 1)),
        "free_spin_cap": int(_number(settings.get("Free Spin Cap"), 50)),
        "scatter_trigger": int(_number(settings.get("Scatter Trigger Count"), 3)),
        "buy_price": _number(settings.get("Buy Feature Price (x Bet)"), 40.5),
        "source_xlsx": path.name,
    }


def write_js(config: dict[str, Any], output_path: str | Path) -> Path:
    output = Path(output_path)
    payload = json.dumps(config, ensure_ascii=False, separators=(",", ":"))
    output.write_text(
        "/* Generated from the H045 source xlsx. Do not edit by hand. */\n"
        f"window.H045_CONFIG={payload};\n",
        encoding="utf-8",
    )
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate H045 browser config from H045 xlsx")
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    write_js(load_game_config(args.xlsx), args.output)
    print(f"wrote {args.output}")


if __name__ == "__main__":
    main()
