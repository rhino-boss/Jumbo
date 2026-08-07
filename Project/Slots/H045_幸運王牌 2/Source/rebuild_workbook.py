# -*- coding: utf-8 -*-
"""Rebuild the H045 PARsheet workbook into the H026-style layout.

Reads the legacy H016-derived layout (Overview / Card / Symbol Weight /
"<Mode> Symbol" sheets) and writes the modern layout used by H026 / H028:

    Overview / Description / Parameter /
    Multiplier_Weight / Multiplier_Weight_Newbie / Multiplier_Weight_Oldhand /
    OP Jackpot / BG_Symbol .. BF_Symbol

Super Feature (Buy Super Feature / Super Free Game) is dropped: H045 has no SF.

Usage:  py rebuild_workbook.py <legacy.xlsx> <output.xlsx>
"""

from __future__ import annotations

import argparse
import shutil
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

STRIP_LEN = 400
SYMBOL_ROWS = range(4, 23)  # WW1 .. JG

# legacy sheet name -> (new sheet name, Symbol Weight start column)
TABLE_MAP = [
    ("Base Game Symbol - High", "BG_Symbol", 2),
    ("Base Game Symbol - Low", "BG_Symbol (2)", 9),
    ("Free Game Symbol - High - A", "FG_Symbol", 16),
    ("Free Game Symbol - High - K", "FG_Symbol (2)", 16),
    ("Free Game Symbol - High - Q", "FG_Symbol (3)", 16),
    ("Free Game Symbol - High - J", "FG_Symbol (4)", 16),
    ("Free Game Symbol - Low", "FG_Symbol (5)", 23),
    ("Buy Feature Symbol", "BF_Symbol", 30),
]

TABLE_NOTE = {
    "BG_Symbol": "Base Game 高表",
    "BG_Symbol (2)": "Base Game 低表",
    "FG_Symbol": "Free Game 高表 - A",
    "FG_Symbol (2)": "Free Game 高表 - K",
    "FG_Symbol (3)": "Free Game 高表 - Q",
    "FG_Symbol (4)": "Free Game 高表 - J",
    "FG_Symbol (5)": "Free Game 低表",
    "BF_Symbol": "Buy Feature 進場盤面",
}

# Card sheet section -> (first row, last row)
CARD_SECTIONS = [
    ("Base Game (Normal Bet)", 5, 57),
    ("Free Game", 64, 115),
    ("Buy Feature", 129, 180),
]

CARD_COLS = {
    "range": 16,
    "table": 17,
    "rate": 18,
    "fix_num": (19, 20, 21),
    "fix_rate": (22, 23, 24),
    "final_rate": (25, 26, 27),
    "weight": (28, 29, 30),
}

TITLE_FONT = Font(bold=True, size=12)
HEAD_FONT = Font(bold=True)
HEAD_FILL = PatternFill("solid", fgColor="DDEBF7")
THIN = Side(style="thin", color="B0B0B0")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _num(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _put(ws, row: int, col: int, value, *, head=False, title=False, fmt=None):
    cell = ws.cell(row, col, value)
    if title:
        cell.font = TITLE_FONT
    elif head:
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.border = BOX
        cell.alignment = Alignment(horizontal="center")
    if fmt:
        cell.number_format = fmt
    return cell


def _row(ws, row: int, col: int, values, *, head=False, fmt=None):
    for offset, value in enumerate(values):
        _put(ws, row, col + offset, value, head=head, fmt=fmt)


# --------------------------------------------------------------------------
# extraction
# --------------------------------------------------------------------------

def extract(src: Path) -> dict[str, Any]:
    wb = load_workbook(src, data_only=True)
    formulas = load_workbook(src, data_only=False)
    weight_sheet = formulas["Symbol Weight"]

    data: dict[str, Any] = {}

    ov = wb["Overview"]
    data["paytable"] = [
        (
            str(ov.cell(r, 1).value).strip(),
            _num(ov.cell(r, 4).value),
            _num(ov.cell(r, 5).value),
            _num(ov.cell(r, 6).value),
        )
        for r in range(41, 49)
    ]

    tables = {}
    for legacy, new_name, weight_col in TABLE_MAP:
        sheet = formulas[legacy]
        counts = [
            [sheet.cell(r, 1).value, sheet.cell(r, 2).value]
            + [_num(sheet.cell(r, 2 + i).value) for i in range(1, 6)]
            + [_num(sheet.cell(r, 8).value)]
            for r in SYMBOL_ROWS
        ]
        strip_symbols = [
            [sheet.cell(4 + i, 11 + reel).value for reel in range(5)] for i in range(STRIP_LEN)
        ]
        strip_ids = [
            [_num(sheet.cell(4 + i, 17 + reel).value) for reel in range(5)] for i in range(STRIP_LEN)
        ]
        weights = [
            [_num(weight_sheet.cell(5 + i, weight_col + reel).value) for reel in range(5)]
            for i in range(STRIP_LEN)
        ]
        random_wild = [
            (sheet.cell(r, 29).value, _num(sheet.cell(r, 30).value)) for r in range(3, 7)
        ]
        multipliers = [
            sheet.cell(r, 29).value for r in range(9, 14) if sheet.cell(r, 29).value not in (None, "")
        ]
        tables[new_name] = {
            "counts": counts,
            "strip_symbols": strip_symbols,
            "strip_ids": strip_ids,
            "weights": weights,
            "random_wild": random_wild,
            "multipliers": multipliers,
        }
    data["tables"] = tables

    desc = formulas["Description"]
    data["fg_mix"] = {
        "D": [
            [_num(desc.cell(r, 3).value), _num(desc.cell(r, 4).value), _num(desc.cell(r, 5).value)]
            for r in range(23, 29)
        ],
        "E": [
            [_num(desc.cell(r, 7).value), _num(desc.cell(r, 8).value), _num(desc.cell(r, 9).value)]
            for r in range(23, 29)
        ],
    }
    data["high_variant"] = [_num(desc.cell(r, 5).value, 1.0) for r in range(31, 35)]

    card = wb["Card"]
    sections = []
    for title, first, last in CARD_SECTIONS:
        rows = []
        for r in range(first, last + 1):
            rows.append(
                {
                    "lower": card.cell(r, 1).value,
                    "upper": card.cell(r, 2).value,
                    "range": card.cell(r, CARD_COLS["range"]).value,
                    "table": card.cell(r, CARD_COLS["table"]).value,
                    "rate": _num(card.cell(r, CARD_COLS["rate"]).value),
                    "fix_num": [_num(card.cell(r, c).value) for c in CARD_COLS["fix_num"]],
                    "fix_rate": [_num(card.cell(r, c).value) for c in CARD_COLS["fix_rate"]],
                    "final_rate": [_num(card.cell(r, c).value) for c in CARD_COLS["final_rate"]],
                    "weight": [_num(card.cell(r, c).value) for c in CARD_COLS["weight"]],
                }
            )
        sections.append({"title": title, "rows": rows})
    data["card_sections"] = sections
    data["final_rtp"] = [_num(card.cell(3, 32).value), _num(card.cell(3, 34).value)]

    op = wb["OP Jack Pot"]
    data["op"] = {
        "新手 (Normal Bet)": [
            [op.cell(r, 2).value, _num(op.cell(r, 3).value)] for r in list(range(3, 10)) + list(range(11, 18))
        ],
        "老手 (Normal Bet)": [
            [op.cell(r, 7).value, _num(op.cell(r, 8).value)] for r in list(range(3, 10)) + list(range(11, 18))
        ],
        "Buy Feature": [
            [op.cell(r, 12).value, _num(op.cell(r, 13).value)] for r in list(range(3, 10)) + list(range(11, 18))
        ],
    }
    data["op_sc_trig"] = [
        _num(op.cell(18, 3).value),
        _num(op.cell(18, 8).value),
        _num(op.cell(18, 13).value),
    ]

    wb.close()
    formulas.close()
    return data


# --------------------------------------------------------------------------
# sheet writers
# --------------------------------------------------------------------------

def write_overview(wb: Workbook, data: dict[str, Any], model: str) -> None:
    ws = wb.create_sheet("Overview")
    ws.column_dimensions["A"].width = 24
    for col in "BCDEF":
        ws.column_dimensions[col].width = 16
    ws.column_dimensions["E"].width = 26

    _put(ws, 2, 1, "Model:", head=True)
    _put(ws, 2, 2, model)
    _put(ws, 3, 1, "Version:", head=True)
    _put(ws, 3, 2, "1.0.0.0")

    _put(ws, 6, 1, "Base Bet", title=True)
    _put(ws, 7, 1, 100)

    _row(ws, 10, 1, ["Coin in", "Price(x)", "Total RTP", "Bet Type"], head=True)
    _row(ws, 11, 1, [100, 1, data["final_rtp"][0], "Normal Bet (新手 / Newbie)"])
    _row(ws, 12, 1, [100, 1, data["final_rtp"][1], "Normal Bet (老手 / Oldhand)"])
    _row(ws, 13, 1, [4050, 40.5, None, "Buy Feature (新手 / Newbie)"])
    _row(ws, 14, 1, [4050, 40.5, None, "Buy Feature (老手 / Oldhand)"])

    _put(ws, 16, 1, "Total Pay Back Percentage includes the following adjustments.")
    row = 17
    for label in (
        "Normal Bet (新手 / Newbie)",
        "Normal Bet (老手 / Oldhand)",
        "Buy Feature (新手 / Newbie)",
        "Buy Feature (老手 / Oldhand)",
    ):
        _row(ws, row, 1, ["Bet Type", "Pay Back", "Hit%", "Pulls/Hit", "Description"], head=True)
        _row(ws, row + 1, 1, [label, None, None, None, "Base Game"])
        _row(ws, row + 2, 1, [None, None, None, None, "Free Game"])
        row += 4

    _put(ws, row, 1, "※ Buy Feature 的 Base Game 欄為進場盤面得分；卡片系統以 Free Game 得分匹配區間，故進場局不另計分。")

    _row(ws, 30, 1, ["Reel #", 1, 2, 3, 4, 5], head=True)
    _row(ws, 31, 1, ["Visible Window Size", 4, 4, 4, 4, 4])

    _put(ws, 34, 1, "Free Spins Setting", title=True)
    _row(ws, 35, 1, ["C1 Num.", "Base Game", "Free Game"], head=True)
    for offset, count in enumerate((3, 4, 5, 6)):
        _row(ws, 36 + offset, 1, [count, 10, 5])
    _put(ws, 40, 1, "The maximum of free spins is 50.")

    _put(ws, 42, 1, "Pay Table：", title=True)
    _put(ws, 42, 2, "All wins show for 100 credit bet.")
    _row(ws, 43, 1, ["Symbol", "Description", "3", "4", "5", "Id"], head=True)
    ordered = [
        ("WW1", "Wild 1", 0, 0, 0, 0),
        ("WW2", "Wild 2 (Copy Wild)", 0, 0, 0, 1),
        ("C1", "Scatter", 0, 0, 0, 2),
    ]
    ids = {"M1": 3, "M2": 4, "M3": 5, "M4": 6, "A": 7, "K": 8, "Q": 9, "J": 10}
    gold_ids = {"M1": 11, "M2": 12, "M3": 13, "M4": 14, "A": 15, "K": 16, "Q": 17, "J": 18}
    row = 44
    for name, description, p3, p4, p5 in [(n, n, a, b, c) for n, a, b, c in data["paytable"]]:
        ordered.append((name, description, p3, p4, p5, ids[name]))
    for entry in ordered:
        _row(ws, row, 1, list(entry))
        row += 1
    for name, _d, p3, p4, p5 in [(n, n, a, b, c) for n, a, b, c in data["paytable"]]:
        _row(ws, row, 1, [f"{name}G", f"Golden {name}", p3, p4, p5, gold_ids[name]])
        row += 1
    _put(ws, row + 1, 1, "※ 金框符號（M1G~JG）判獎時視為對應一般符號，賠率相同。")
    _put(ws, row + 2, 1, "※ WW1 / WW2 可替代除 C1 外的所有一般符號，本身無賠率。")
    _put(ws, row + 3, 1, "※ C1 僅統計出現總數用於觸發 Free Game，無 Scatter Pay。")

    _put(ws, row + 5, 1, "Feature：", title=True)
    features = [
        "1. Base Game 為 5 輪 × 4 列、1024 Ways 盤面。連線後消除所有連線符號，剩餘符號依輪帶順序落下補位。",
        "2. 同一次 Spin 每完成一次消除，得分倍率往上提升；BG 為 X1、X2、X3、X5、X10，FG 為 X2、X4、X6、X10、X20。",
        "3. 消除到金框符號時，該格原地轉為 WW1；或依 Random Wild 權重轉為 WW2 並額外複製 2~4 個 WW2。",
        "4. 每一輪皆有機會出現 C1，消除補位後也可能出現，同一輪可出現超過 1 顆 C1。",
        "5. 所有消除結束後統計盤面 C1，達 3 顆以上觸發 10 場 Free Game；FG 內再達 3 顆固定加 5 場，上限 50 場。",
        "6. 本遊戲提供 Buy Feature（40.5 × Bet），不提供 Super Feature。",
    ]
    for offset, text in enumerate(features):
        _put(ws, row + 6 + offset, 1, text)


def write_description(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("Description")
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 118
    for col in "CDEFGHI":
        ws.column_dimensions[col].width = 13

    row = 2
    _put(ws, row, 2, "Base Game", title=True)
    for text in [
        "1. Base Game 每次 Spin 依卡片系統（Multiplier_Weight）對應到的表持續抽取，直到抽出對應得分區間。",
    ]:
        row += 1
        _put(ws, row, 2, text)
    row += 1
    _row(ws, row, 3, ["種類", "", "", "ID"], head=True)
    for name, code in (("BG_Symbol", "B"), ("BG_Symbol (2)", "A")):
        row += 1
        _put(ws, row, 3, name)
        _put(ws, row, 6, code)
    for text in [
        "2. 初始盤面依選到的 BG 工作頁 5 條輪帶，並依該頁 Symbol Weight R1~R5 權重決定停輪位置與符號。",
        "3. 若有連線，消除連線符號，並依同輪權重抽取新符號補上空格。",
        "4. 當 Combo 隨消除上升時得分倍率跟著上升，倍數為 X1、X2、X3、X5、X10，達到 X10 後不再上升。",
        "5. 隨機百搭判定：",
    ]:
        row += 1
        _put(ws, row, 2, text)
    for text in [
        "(1) 當連線盤面消除到至少一個金框符號（M1G~JG）時進行以下判定。",
        "(2) 每次 Spin 有一個「是否已出現隨機百搭」的 FLAG；FLAG 尚未觸發前，每次 Combo 只要消除到金框符號都會判定一次，"
        "FLAG 使用過後續 Combo 不再判定。",
        "(3) 消到金框符號時依 Parameter 的 Random Wild Weight，若骰到 2~4 則該格取代為 WW2，並額外複製 2~4 個 WW2 "
        "取代 R2~R5 中隨機格子，不會蓋掉原本是 C1、WW1、WW2 的格子。",
        "(4) 多個金框符號同時被消除時同樣走上述判定，但最多只有其中一個金框觸發隨機百搭，其餘轉成 WW1。",
        "(5) 順序為：消除 → 補格子 → 所有金框符號瞇牌 → 金框格轉成 WW1 或 WW2 → 若有複製 WW2，最後再處理要蓋掉哪些格子。",
    ]:
        row += 1
        _put(ws, row, 3, text)
    for text in [
        "6. 每次連線消除後有機會掉落 C1，同一輪可出現超過 1 個 C1。",
        "7. C1 在消除過程中不參與連線消除，直到無法再連線時才統計盤面 C1，決定是否進入 Free Spins。",
    ]:
        row += 1
        _put(ws, row, 2, text)

    row += 2
    _put(ws, row, 2, "Free Game", title=True)
    row += 1
    _put(ws, row, 2, "1. 進入 Free Game 後依卡片系統對應的組別取得指定場數的高表與低表，出現順序隨機排列。")
    row += 1
    _row(ws, row, 3, ["組別", "高表場數", "低表場數", "Weight"], head=True)
    for group in ("D", "E"):
        for high, low, weight in data["fg_mix"][group]:
            row += 1
            _row(ws, row, 3, [group, high, low, weight])
    row += 1
    _put(ws, row, 2, "※ 目前 Multiplier_Weight 中 Free Game 與 Buy Feature 的「使用的表」皆為 E，即固定 10 場高表。")
    row += 1
    _put(ws, row, 2, "2. 高表場依下列權重選擇本場使用的高表工作頁；低表場固定使用 FG_Symbol (5)。")
    row += 1
    _row(ws, row, 3, ["工作頁", "高表輪帶", "Weight"], head=True)
    for name, symbol, weight in zip(
        ("FG_Symbol", "FG_Symbol (2)", "FG_Symbol (3)", "FG_Symbol (4)"),
        ("A", "K", "Q", "J"),
        data["high_variant"],
    ):
        row += 1
        _row(ws, row, 3, [name, symbol, weight])
    for text in [
        "3. Free Game 中最終盤面出現 3 顆以上 C1 時固定增加 5 場，其中 1 場高表與 4 場低表，順序隨機排列。",
        "4. 若有連線，消除連線符號並依同輪權重抽取新符號補上空格。",
        "5. 當 Combo 隨消除上升時得分倍率跟著上升，倍數為 X2、X4、X6、X10、X20，達到 X20 後不再上升。",
        "6. 隨機百搭判定同 Base Game 第 5 點；FG 不受「同一 Spin 只觸發一次」限制，每次 Combo 皆可判定。",
        "7. 每次連線消除後有機會掉落 C1，同一輪可出現超過 1 個 C1。",
        "8. C1 在消除過程中不參與連線消除，直到無法再連線時才統計盤面 C1，決定是否 Retrigger。",
        "9. Free Game 最大場次為 50 場。",
    ]:
        row += 1
        _put(ws, row, 2, text)

    row += 2
    _put(ws, row, 2, "Buy Feature", title=True)
    for text in [
        "1. 購買價格為 40.5 × Bet。購買後使用 BF_Symbol 輪帶與其 Symbol Weight 決定進場盤面 RNG。",
        "2. 其餘流程皆與 Base Game 的 2~7 相同。",
        "3. 進場盤面統計 C1 觸發 Free Game 後，依卡片系統對應組別取得高表與低表場數，順序隨機排列。",
        "4. 其餘流程皆與 Free Game 的 2~9 相同。",
    ]:
        row += 1
        _put(ws, row, 2, text)

    row += 2
    _put(ws, row, 2, "注意", title=True)
    for text in [
        "1. 本遊戲不提供 Super Feature / Super Free Game。",
        "2. 工作頁命名對應：BG_Symbol＝BG 高表、BG_Symbol (2)＝BG 低表、FG_Symbol~FG_Symbol (4)＝FG 高表 A/K/Q/J、"
        "FG_Symbol (5)＝FG 低表、BF_Symbol＝Buy Feature 進場表。",
        "3. Random Wild Weight 與 Win Multiplier Ladder 集中於 Parameter 工作頁。",
        "4. 卡片系統的區間權重集中於 Multiplier_Weight_Newbie（新手）與 Multiplier_Weight_Oldhand（老手）。",
    ]:
        row += 1
        _put(ws, row, 2, text)


def write_parameter(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("Parameter")
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 30
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 14

    row = 2
    _put(ws, row, 2, "Table Selection - Base Game", title=True)
    row += 1
    _row(ws, row, 2, ["Worksheet Name", "ID", "Description"], head=True)
    for name, code in (("BG_Symbol", "B"), ("BG_Symbol (2)", "A")):
        row += 1
        _row(ws, row, 2, [name, code, TABLE_NOTE[name]])
    row += 1
    _put(ws, row, 2, "※ 實際選表由 Multiplier_Weight_* 的「使用的表」欄位決定，不另外做表權重抽取。")

    row += 2
    _put(ws, row, 2, "Random Wild Weight (Weight/Total)", title=True)
    row += 1
    values = [entry[0] for entry in data["tables"]["BG_Symbol"]["random_wild"]]
    _row(ws, row, 2, ["Worksheet Name \\ 複製數量"] + [str(v) for v in values], head=True)
    for _legacy, name, _col in TABLE_MAP:
        row += 1
        _row(ws, row, 2, [name] + [weight for _v, weight in data["tables"][name]["random_wild"]])
    row += 1
    _put(ws, row, 2, "※ 骰到 0 代表本次金框只轉成 WW1，不複製 WW2。")

    row += 2
    _put(ws, row, 2, "Win Multiplier Ladder", title=True)
    row += 1
    _row(ws, row, 2, ["Worksheet Name \\ 第 N 次消除", "1", "2", "3", "4", "5+"], head=True)
    for _legacy, name, _col in TABLE_MAP:
        row += 1
        _row(ws, row, 2, [name] + list(data["tables"][name]["multipliers"]))
    row += 1
    _put(ws, row, 2, "※ 倍率只在同一 Spin 的消除期間推進，新的 Spin 回到該模式第一階。")

    row += 2
    _put(ws, row, 2, "Free Game Table Mix Weight", title=True)
    row += 1
    _row(ws, row, 2, ["組別", "高表場數", "低表場數", "Weight"], head=True)
    for group in ("D", "E"):
        for high, low, weight in data["fg_mix"][group]:
            row += 1
            _row(ws, row, 2, [group, high, low, weight])

    row += 2
    _put(ws, row, 2, "Free Game High Table Weight", title=True)
    row += 1
    _row(ws, row, 2, ["Worksheet Name", "高表輪帶", "Weight"], head=True)
    for name, symbol, weight in zip(
        ("FG_Symbol", "FG_Symbol (2)", "FG_Symbol (3)", "FG_Symbol (4)"),
        ("A", "K", "Q", "J"),
        data["high_variant"],
    ):
        row += 1
        _row(ws, row, 2, [name, symbol, weight])

    row += 2
    _put(ws, row, 2, "Feature Setting", title=True)
    row += 1
    _row(ws, row, 2, ["Item", "Value"], head=True)
    for item, value in (
        ("Free Spins", 10),
        ("Retrigger Spins", 5),
        ("Free Spin Cap", 50),
        ("Retrigger 高表場數", 1),
        ("Retrigger 低表場數", 4),
        ("Buy Feature Price (x Bet)", 40.5),
        ("Scatter Trigger Count", 3),
        ("Reel Num", 5),
        ("Visible Window Size", 4),
        ("Max Ways", 1024),
    ):
        row += 1
        _row(ws, row, 2, [item, value])


def _write_card_sheet(wb: Workbook, data: dict[str, Any], title: str, index: int, label: str) -> None:
    ws = wb.create_sheet(title)
    ws.column_dimensions["A"].width = 4
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 17

    _put(ws, 2, 2, "Threshold", head=True)
    _put(ws, 2, 3, 1000000000)
    _put(ws, 3, 2, "Coin in", head=True)
    _put(ws, 3, 3, 100)
    _put(ws, 4, 2, "Profile", head=True)
    _put(ws, 4, 3, label)
    if index < len(data["final_rtp"]):
        _put(ws, 5, 2, "Final RTP", head=True)
        _put(ws, 5, 3, data["final_rtp"][index])

    row = 7
    for section in data["card_sections"]:
        _put(ws, row, 2, section["title"], title=True)
        row += 1
        _row(
            ws,
            row,
            2,
            ["Lower", "Upper", "Range", "使用的表", "Rate", "Fix Num", "Fix Rate", "Final Rate", "Weight"],
            head=True,
        )
        row += 1
        for entry in section["rows"]:
            weight = entry["weight"][index]
            _row(
                ws,
                row,
                2,
                [
                    entry["lower"],
                    entry["upper"],
                    entry["range"],
                    entry["table"],
                    entry["rate"],
                    entry["fix_num"][index],
                    entry["fix_rate"][index],
                    entry["final_rate"][index],
                    weight,
                ],
            )
            row += 1
        _put(ws, row, 2, "Total", head=True)
        _put(ws, row, 10, sum(e["weight"][index] for e in section["rows"]))
        row += 3


def write_multiplier_weight(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("Multiplier_Weight")
    ws.column_dimensions["A"].width = 4
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 18

    _put(ws, 2, 2, "Threshold", head=True)
    _put(ws, 2, 3, 1000000000)
    _put(ws, 3, 2, "Coin in", head=True)
    _put(ws, 3, 3, 100)

    row = 5
    for section in data["card_sections"]:
        _put(ws, row, 2, section["title"], title=True)
        row += 1
        _row(
            ws,
            row,
            2,
            ["Lower", "Upper", "Range", "使用的表", "Rate", "Weight_Newbie", "Weight_Oldhand"],
            head=True,
        )
        row += 1
        for entry in section["rows"]:
            _row(
                ws,
                row,
                2,
                [
                    entry["lower"],
                    entry["upper"],
                    entry["range"],
                    entry["table"],
                    entry["rate"],
                    entry["weight"][0],
                    entry["weight"][1],
                ],
            )
            row += 1
        _put(ws, row, 2, "Total", head=True)
        _put(ws, row, 7, sum(e["weight"][0] for e in section["rows"]))
        _put(ws, row, 8, sum(e["weight"][1] for e in section["rows"]))
        row += 3


def write_op_jackpot(wb: Workbook, data: dict[str, Any]) -> None:
    ws = wb.create_sheet("OP Jackpot")
    ws.column_dimensions["A"].width = 4
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 18

    col = 2
    for offset, (label, rows) in enumerate(data["op"].items()):
        _put(ws, 2, col, label, title=True)
        _row(ws, 3, col, ["Counter", "Num", "Count"], head=True)
        row = 4
        for index, (num, count) in enumerate(rows):
            counter = "bgC1Count" if index < 7 else "fgC1Count"
            _row(ws, row, col, [counter, num, count])
            row += 1
        _row(ws, row, col, ["SC Trig", None, data["op_sc_trig"][offset]])
        col += 4


def write_symbol_sheet(wb: Workbook, name: str, table: dict[str, Any]) -> None:
    ws = wb.create_sheet(name)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 8

    _put(ws, 2, 10, "Symbol", head=True)
    _put(ws, 2, 17, "Symbol ID", head=True)
    _put(ws, 2, 23, "Symbol Weight", head=True)
    _row(ws, 3, 1, ["Symbol", "Description", "R1", "R2", "R3", "R4", "R5", "ID"], head=True)
    _put(ws, 3, 10, "Line #", head=True)
    _row(ws, 3, 11, ["R1", "R2", "R3", "R4", "R5"], head=True)
    _row(ws, 3, 17, ["R1", "R2", "R3", "R4", "R5"], head=True)
    _row(ws, 3, 23, ["R1", "R2", "R3", "R4", "R5"], head=True)

    for offset, entry in enumerate(table["counts"]):
        _row(ws, 4 + offset, 1, entry)
    total_row = 4 + len(table["counts"])
    _put(ws, total_row, 1, "--")
    for reel in range(5):
        _put(ws, total_row, 3 + reel, sum(entry[2 + reel] for entry in table["counts"]))

    for index in range(STRIP_LEN):
        row = 4 + index
        _put(ws, row, 10, index)
        _row(ws, row, 11, table["strip_symbols"][index])
        _row(ws, row, 17, table["strip_ids"][index])
        _row(ws, row, 23, table["weights"][index])


def build(src: Path, dst: Path) -> None:
    data = extract(src)
    wb = Workbook()
    wb.remove(wb.active)

    write_overview(wb, data, dst.stem)
    write_description(wb, data)
    write_parameter(wb, data)
    write_multiplier_weight(wb, data)
    _write_card_sheet(wb, data, "Multiplier_Weight_Newbie", 0, "新手 / Newbie")
    _write_card_sheet(wb, data, "Multiplier_Weight_Oldhand", 1, "老手 / Oldhand")
    write_op_jackpot(wb, data)
    for _legacy, name, _col in TABLE_MAP:
        write_symbol_sheet(wb, name, data["tables"][name])

    wb.save(dst)


def main() -> None:
    parser = argparse.ArgumentParser(description="Rebuild H045 workbook into the H026-style layout")
    parser.add_argument("src", type=Path)
    parser.add_argument("dst", type=Path)
    args = parser.parse_args()
    build(args.src, args.dst)
    print(f"wrote {args.dst}")


if __name__ == "__main__":
    main()
