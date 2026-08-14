import ast
import asyncio
import html
import io
import json
import logging
import os
import re
import shutil
import socket
import subprocess
import sys
import tempfile
import time
import unicodedata
import urllib.request
import zipfile
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
from telegram import Bot, BotCommand, InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.constants import ParseMode
from telegram.error import Conflict
from telegram.ext import Application, CallbackQueryHandler, CommandHandler, MessageHandler, filters, ContextTypes
import rarfile

logging.basicConfig(
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger("tgbot")

r"""

# 在 VM 上執行：
python3 walking_encyclopedia_bot_vm.py

# 功能介紹（VM 版，移植自 walking_encyclopedia_bot.py）
  * 整理報表：支援 txt/zip/rar、多檔合併、巢狀壓縮檔，並自動轉成 Excel。
  * 壓測資料重點整理：抓取 RTP、Spin、Coin in、各 pool RTP 等重點欄位。
  * /simulator 開啟按鈕選單執行多遊戲模擬（僅限管理者）。
  * /help 說明本檔案完整功能。
  * 本版不包含 AI 問答功能（原版透過本機 Claude CLI 回答，VM 上未設置）。

"""

# ==================== 🔑 核心金鑰設定 ====================
TG_TOKEN = "8817922272:AAEzFERpAAlhLTf3bs1bFsdoevtUdzNU4cA"
ALLOWED_USER_ID = 5539551776  # 只有這個 Telegram 使用者 ID 可以下模擬器指令


# ==================== ⚙️ 機器人參數設定 ====================
BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "generated_reports"
STRESS_PROGRESS_FILE = BASE_DIR / "stress_test_progress.json"
STRESS_GAME_LIST_CANDIDATES = [
    BASE_DIR / "game_list.md",
    BASE_DIR.parent.parent / "Project" / "Slots" / "game_list.md",
]
STRESS_GAMES_PER_PAGE = 8
BET_TYPE_TO_STRESS_MODE = {
    "0": "NB",
    "1": "EX1",
    "2": "EX2",
    "3": "BF",
    "4": "SF",
}
STRESS_MODE_TO_COLUMN = {
    "NB": "押注-NB",
    "EX1": "押注-EX1",
    "EX2": "押注-EX2",
    "BF": "押注-BF",
    "SF": "押注-SF",
}
MAX_REPORT_FILES = 20
ARCHIVE_EXTENSIONS = {".zip", ".rar"}
REPORT_FILE_EXTENSIONS = ARCHIVE_EXTENSIONS | {".txt"}
MAX_ARCHIVE_DEPTH = 10
ARCHIVE_BATCH_DELAY_SECONDS = 2
BOT_INSTANCE_LOCK_PORT = 48726
SEVENZIP_CANDIDATES = [
    "/usr/bin/7z",
    "/usr/bin/7za",
    "/usr/local/bin/7z",
]

VALIDATOR_HEADER_PATTERNS = {
    "GameID": r"GameID\s*=\s*(\d+)",
    "optionID": r"optionID\s*=\s*(\d+)",
    "Coin in": r"Coin in\s*=\s*([0-9.]+)",
    "TotalSpin": r"TotalSpin\s*=\s*([0-9.]+)",
    "TotalRtp (including JP/bonus)": r"TotalRtp \(including JP/bonus\)\s*=\s*([0-9.]+)",
    "avgPeriod": r"avgPeriod\s*=\s*([0-9.]+)",
    "BaseGameRtp": r"BaseGameRtp\s*=\s*([0-9.]+)",
    "ScatterRtp": r"ScatterRtp\s*=\s*([0-9.]+)",
    "FreeGameRtp": r"FreeGameRtp\s*=\s*([0-9.]+)",
    "BonusJP Rtp": r"BonusJP\s+Rtp\s*=\s*([0-9.]+)|BonusJP\s+RTP\s*=\s*([0-9.]+)",
    "LinkJP Rtp": r"LinkJP\s+Rtp\s*=\s*([0-9.]+)|LinkJP\s+RTP\s*=\s*([0-9.]+)",
}

HIDDEN_VALIDATOR_COLUMNS = {
    "source_txt",
}

VALIDATOR_OUTPUT_COLUMNS = [
    ("source_txt", "source_txt"),
    ("txt在的資料夾名稱", "資料夾名稱"),
    ("GameID", "Game ID"),
    ("optionID", "Option ID"),
    ("Coin in", "Coin in"),
    ("TotalSpin", "Total Spin"),
    ("avgPeriod", "JP 週期"),
    ("TotalRtp (including JP/bonus)", "RTP Total"),
    ("LinkJP Rtp", "RTP JP Link"),
    ("BonusJP Rtp", "RTP JP Bonus"),
    ("RTP Game", "RTP Game"),
    ("BaseGameRtp", "RTP BG"),
    ("ScatterRtp", "RTP SC"),
    ("FreeGameRtp", "RTP FG"),
    ("BaseGame最大倍數", "最大倍數 BG"),
    ("FreeGame最大倍數", "最大倍數 FG"),
    ("BG進FG最大倍數", "最大倍數 B2F"),
]

RTP_OUTPUT_FIELDS = {
    "TotalRtp (including JP/bonus)",
    "LinkJP Rtp",
    "BonusJP Rtp",
    "RTP Game",
    "BaseGameRtp",
    "ScatterRtp",
    "FreeGameRtp",
}

THOUSANDS_OUTPUT_FIELDS = {
    "Coin in",
    "TotalSpin",
}

MULTIPLIER_SECTION_FIELDS = {
    "BaseGame 倍數": "BaseGame最大倍數",
    "FreeGame 倍數": "FreeGame最大倍數",
    "整場FreeGame 倍數": "FreeGame最大倍數",
    "整場 FreeGame 倍數": "FreeGame最大倍數",
    "BaseGame 進 FreeGame 倍數": "BG進FG最大倍數",
}


class NoValidatorReportError(ValueError):
    """上傳內容中找不到 RTP Validator 報表。"""

# ==================== 🎰 模擬器設定 ====================
GAMES_ROOT = Path("/root/Simulator")  # 每個子資料夾＝一個遊戲，例如 H026_彩罐熱舞
BET_MODES = {"0": "Normal Bet", "1": "Extra Bet", "2": "Feature Buy"}
MAX_ROUNDS = int(os.environ.get("MAX_ROUNDS", "1000000000"))
MAX_BATCH_COMBINATIONS = 20
SUMMARY_LINE_RE = re.compile(r"^\* .+$", re.MULTILINE)
REPORT_LINE_RE = re.compile(r"^Report: (.+)$", re.MULTILINE)


def discover_games() -> dict[str, Path]:
    """掃描 GAMES_ROOT 底下有 Simulator.py 的資料夾，即時抓目前有哪些遊戲。"""
    games = {}
    if not GAMES_ROOT.is_dir():
        return games
    for entry in sorted(GAMES_ROOT.iterdir()):
        if entry.is_dir() and (entry / "Simulator.py").exists():
            games[entry.name] = entry
    return games


def get_game_id(folder_name: str) -> str:
    return folder_name.split("_", 1)[0]


def find_game_by_id(game_id: str) -> tuple[str, Path] | None:
    game_id_upper = game_id.strip().upper()
    for name, path in discover_games().items():
        if get_game_id(name).upper() == game_id_upper:
            return name, path
    return None


def discover_configs(game_dir: Path) -> list[str]:
    return [f.stem[len("config_") :] for f in sorted(game_dir.glob("config_*.js"))]


def get_config_math_version(game_dir: Path, config: str) -> str | None:
    """從 config JS 讀取數學版本，優先採用 Excel 來源版本。"""
    config_path = game_dir / f"config_{config}.js"
    try:
        content = config_path.read_text(encoding="utf-8")
    except (OSError, UnicodeError):
        return None

    for key in ("excel_version", "game_version", "version"):
        match = re.search(
            rf"""["']{re.escape(key)}["']\s*:\s*["']([^"']+)["']""",
            content,
        )
        if match:
            return match.group(1).strip()
    return None


def build_batch_template(game_name: str, configs: list[str]) -> str:
    """建立可直接複製、修改並回傳給 bot 的批次參數範本。"""
    examples = configs[:2] or ["92A"]
    if len(examples) == 1:
        examples.append(examples[0])
    include_enabled = get_game_id(game_name).upper() != "H026"
    rows = []
    for index, config in enumerate(examples):
        enabled_part = ', "card_system_enabled": True' if include_enabled else ""
        rows.append(
            f'    {{"config_file": "config_{config}.js", "bet_mode": {index}, '
            f'"total_rounds": 10**4{enabled_part}, "card_system_is_newbie": False}},'
        )
    return "BATCH_COMBINATIONS = [\n" + "\n".join(rows) + "\n]"


def _safe_batch_value(node: ast.AST):
    """只解析批次範本需要的常值，允許 10**4，但不執行任意程式碼。"""
    if isinstance(node, ast.Constant) and isinstance(node.value, (str, int, bool)):
        return node.value
    if isinstance(node, ast.List):
        return [_safe_batch_value(item) for item in node.elts]
    if isinstance(node, ast.Dict):
        result = {}
        for key_node, value_node in zip(node.keys, node.values):
            if key_node is None:
                raise ValueError("不支援 **dict 展開語法")
            key = _safe_batch_value(key_node)
            if not isinstance(key, str):
                raise ValueError("參數欄位名稱必須是文字")
            result[key] = _safe_batch_value(value_node)
        return result
    if isinstance(node, ast.UnaryOp) and isinstance(node.op, (ast.UAdd, ast.USub)):
        value = _safe_batch_value(node.operand)
        if not isinstance(value, int) or isinstance(value, bool):
            raise ValueError("正負號只能套用在整數")
        return value if isinstance(node.op, ast.UAdd) else -value
    if isinstance(node, ast.BinOp) and isinstance(node.op, ast.Pow):
        base = _safe_batch_value(node.left)
        exponent = _safe_batch_value(node.right)
        if not isinstance(base, int) or isinstance(base, bool) or not isinstance(exponent, int) or isinstance(exponent, bool):
            raise ValueError("次方的底數與指數必須是整數")
        if exponent < 0 or exponent > 12:
            raise ValueError("次方指數必須介於 0 ~ 12")
        return base**exponent
    raise ValueError(f"不支援的語法: {type(node).__name__}")


def parse_batch_combinations(text: str) -> list[dict]:
    """解析 `BATCH_COMBINATIONS = [...]` 或單純 `[...]` 格式。"""
    cleaned = text.strip()
    if cleaned.startswith("```") and cleaned.endswith("```"):
        lines = cleaned.splitlines()
        cleaned = "\n".join(lines[1:-1]).strip()
    try:
        module = ast.parse(cleaned, mode="exec")
    except SyntaxError as exc:
        raise ValueError(f"第 {exc.lineno or '?'} 行語法錯誤: {exc.msg}") from exc
    if len(module.body) != 1:
        raise ValueError("請只傳送一份 BATCH_COMBINATIONS 清單")
    statement = module.body[0]
    if isinstance(statement, ast.Assign):
        if len(statement.targets) != 1 or not isinstance(statement.targets[0], ast.Name) or statement.targets[0].id != "BATCH_COMBINATIONS":
            raise ValueError("指定名稱必須是 BATCH_COMBINATIONS")
        value_node = statement.value
    elif isinstance(statement, ast.Expr):
        value_node = statement.value
    else:
        raise ValueError("格式必須是 BATCH_COMBINATIONS = [...] 或直接 [...]")
    combinations = _safe_batch_value(value_node)
    if not isinstance(combinations, list) or not combinations:
        raise ValueError("BATCH_COMBINATIONS 必須是非空白清單")
    if len(combinations) > MAX_BATCH_COMBINATIONS:
        raise ValueError(f"一次最多 {MAX_BATCH_COMBINATIONS} 組參數")
    if not all(isinstance(item, dict) for item in combinations):
        raise ValueError("清單中的每一組參數都必須是 {...}")
    return combinations


def normalize_batch_combinations(game_name: str, game_dir: Path, combinations: list[dict]) -> list[dict]:
    """驗證欄位並轉成 start_simulation_job 使用的標準格式。"""
    configs = {value.upper() for value in discover_configs(game_dir)}
    jobs = []
    allowed_keys = {"config_file", "bet_mode", "total_rounds", "card_system_enabled", "card_system_is_newbie"}
    for index, combo in enumerate(combinations, start=1):
        unknown = set(combo) - allowed_keys
        if unknown:
            raise ValueError(f"第 {index} 組有未知欄位: {', '.join(sorted(str(key) for key in unknown))}")
        missing = {"config_file", "bet_mode", "total_rounds"} - set(combo)
        if missing:
            raise ValueError(f"第 {index} 組缺少欄位: {', '.join(sorted(missing))}")

        config_file = combo["config_file"]
        if not isinstance(config_file, str):
            raise ValueError(f"第 {index} 組 config_file 必須是文字")
        config = config_file.strip()
        if config.lower().startswith("config_"):
            config = config[len("config_") :]
        if config.lower().endswith(".js"):
            config = config[:-3]
        config = config.upper()
        if config not in configs:
            raise ValueError(f"第 {index} 組 config_file 不存在；{game_name} 可用: {', '.join(sorted(configs))}")

        bet_mode = combo["bet_mode"]
        rounds = combo["total_rounds"]
        enabled = combo.get("card_system_enabled", True)
        newbie = combo.get("card_system_is_newbie", False)
        if not isinstance(bet_mode, int) or isinstance(bet_mode, bool) or str(bet_mode) not in BET_MODES:
            raise ValueError(f"第 {index} 組 bet_mode 必須是 {', '.join(BET_MODES)}")
        if not isinstance(rounds, int) or isinstance(rounds, bool) or not (1000 <= rounds <= MAX_ROUNDS):
            raise ValueError(f"第 {index} 組 total_rounds 必須介於 1,000 ~ {MAX_ROUNDS:,}")
        if not isinstance(enabled, bool) or not isinstance(newbie, bool):
            raise ValueError(f"第 {index} 組卡片系統欄位必須填 True 或 False")
        jobs.append(
            {
                "game_name": game_name,
                "game_dir": game_dir,
                "config": config,
                "bet_mode": str(bet_mode),
                "rounds": rounds,
                "card": "new" if newbie else "old",
                "card_enabled": enabled,
            }
        )
    return jobs


current_job = None  # dict: {"proc", "chat_id", "started_at", "label"}
current_batch = None  # dict: {"chat_id", "jobs", "total", "completed", "cancelled"}
job_lock = asyncio.Lock()
ROUNDS_PRESETS = [
    ("10 萬", 100_000),
    ("100 萬", 1_000_000),
    ("1000 萬", 10_000_000),
    ("1 億", 100_000_000),
    ("10 億（上限）", 1_000_000_000),
]

MENTION_WINDOW_SECONDS = 180  # 群組裡 @機器人 一次後，3 分鐘內符合條件的內容都會自動處理

# 記憶暫存區
archive_batches = {}
bot_instance_lock = None
simulator_wizard_state = {}  # chat_id -> {"config", "bet_mode", "card", "rounds", "awaiting_rounds"}
stress_test_search_state = {}  # chat_id -> 搜尋提示訊息 ID
mention_window_expiry = {}  # chat_id -> 這個時間之前都算「已 @ 過」


def activate_mention_window(chat_id: int):
    mention_window_expiry[chat_id] = time.time() + MENTION_WINDOW_SECONDS


def is_mention_window_active(chat_id: int) -> bool:
    return time.time() < mention_window_expiry.get(chat_id, 0)


def authorized(func):
    """限制只有 ALLOWED_USER_ID 能使用模擬器指令。"""

    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE):
        if update.effective_user is None or update.effective_user.id != ALLOWED_USER_ID:
            await update.message.reply_text("🚫 你沒有權限使用這個指令。")
            return
        await func(update, context)

    return wrapper


def require_mention_in_groups(func):
    """群組裡一定要 @機器人 才會回應指令；私訊則不受限制。"""

    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE):
        if update.message.chat.type != "private":
            if f"@{context.bot.username}" not in (update.message.text or ""):
                return
        await func(update, context)

    return wrapper

# ==================== 💾 本地文字檔儲存邏輯 ====================


def acquire_bot_instance_lock() -> bool:
    """避免同一支 bot 在本機重複啟動，造成 Telegram getUpdates Conflict。"""
    global bot_instance_lock

    lock_socket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        lock_socket.bind(("127.0.0.1", BOT_INSTANCE_LOCK_PORT))
        lock_socket.listen(1)
    except OSError:
        lock_socket.close()
        return False

    bot_instance_lock = lock_socket
    return True


async def check_telegram_polling_available() -> bool:
    try:
        async with Bot(TG_TOKEN) as bot:
            await bot.get_updates(timeout=0, limit=1)
        return True
    except Conflict:
        return False
    except Exception as e:
        print(f"[警告] Telegram polling 預檢失敗，將繼續嘗試啟動: {e}")
        return True


def ensure_output_dir():
    OUTPUT_DIR.mkdir(exist_ok=True)


def prune_old_reports():
    """只保留最新的 20 份產出文件。"""
    ensure_output_dir()
    files = [path for path in OUTPUT_DIR.iterdir() if path.is_file()]
    files.sort(key=lambda path: path.stat().st_mtime, reverse=True)

    for old_file in files[MAX_REPORT_FILES:]:
        try:
            old_file.unlink()
            print(f"[系統提示] 已刪除舊檔案: {old_file.name}")
        except Exception as e:
            print(f"[錯誤] 刪除舊檔案失敗 {old_file.name}: {e}")


def extract_validator_field(pattern: str, text: str) -> str:
    match = re.search(pattern, text)
    if not match:
        return ""
    groups = [group for group in match.groups() if group is not None]
    return groups[0] if groups else match.group(1)


def get_source_txt_filename(source_name: str) -> str:
    return Path(source_name.replace("\\", "/").split("!")[-1]).name


def get_source_txt_folder_name(source_name: str) -> str:
    txt_path = Path(source_name.replace("\\", "/").split("!")[-1])
    return txt_path.parent.name if str(txt_path.parent) != "." else ""


def decode_report_text(data: bytes) -> str:
    """兼容常見的 UTF-8、UTF-16 與繁中 Windows TXT 編碼。"""
    if data.startswith((b"\xff\xfe", b"\xfe\xff")):
        return data.decode("utf-16", errors="replace")

    try:
        return data.decode("utf-8-sig")
    except UnicodeDecodeError:
        return data.decode("cp950", errors="replace")


def normalize_report_heading(line: str) -> str:
    return re.sub(r"\s+", " ", line.strip().strip("=").strip())


def extract_max_multiplier_fields(text: str) -> dict[str, str]:
    """依報表區段抓出該列任一統計值大於 0 的最大倍數。"""
    maxima: dict[str, float] = {}
    current_field = None

    for line in text.splitlines():
        heading = normalize_report_heading(line)
        current_field = MULTIPLIER_SECTION_FIELDS.get(heading, current_field)

        # 遇到下一個 =====標題===== 時結束上一個倍數區段。
        if re.match(r"^\s*=+.*=+\s*$", line) and heading not in MULTIPLIER_SECTION_FIELDS:
            current_field = None
            continue

        if not current_field:
            continue

        match = re.match(
            r"^\s*(?:BaseGame|FreeGame)\s+([0-9]+(?:\.[0-9]+)?)x\s*=\s*(.*)$",
            line,
            re.IGNORECASE,
        )
        if not match:
            continue

        statistics = re.findall(
            r"(?:^|,\s*(?:Pay|\d+C\d+)\s*=\s*)([0-9]+(?:\.[0-9]+)?)",
            match.group(2),
            re.IGNORECASE,
        )
        if not statistics or not any(float(value) > 0 for value in statistics):
            continue

        multiplier = float(match.group(1))
        maxima[current_field] = max(maxima.get(current_field, multiplier), multiplier)

    return {
        field: str(int(value)) if value.is_integer() else str(value)
        for field, value in maxima.items()
    }


def has_stress_test_summary(text: str) -> bool:
    """圖 1 的 Newbie/Veteran Total RTP 摘要代表壓測報表。"""
    return bool(
        re.search(r"Taste\s*\(\s*Newbie\s*\)\s+TotalWin\s*=", text, re.IGNORECASE)
        or re.search(r"Relief\s*\(\s*Veteran\s*\)\s+TotalWin\s*=", text, re.IGNORECASE)
    )


def calculate_game_rtp(base_game_rtp: str, free_game_rtp: str) -> str:
    """RTP Game 按需求只加總 BG 與 FG，不包含 Scatter。"""
    values = [value for value in (base_game_rtp, free_game_rtp) if value]
    if not values:
        return ""

    try:
        total = sum((Decimal(value) for value in values), Decimal("0"))
    except InvalidOperation:
        return ""

    result = format(total, "f")
    if "." in result:
        result = result.rstrip("0").rstrip(".")
    return result or "0"


def is_rtp_output_field(source_name: str) -> bool:
    return source_name in RTP_OUTPUT_FIELDS or bool(re.fullmatch(r"pool\[\d+\]\s+rtp", source_name))


def to_excel_rtp_value(value):
    if value in ("", None):
        return ""
    try:
        return float(Decimal(str(value)))
    except (InvalidOperation, ValueError):
        return value


def to_excel_number_value(value):
    if value in ("", None):
        return ""
    try:
        number = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return value
    return int(number) if number == number.to_integral_value() else float(number)


def parse_validator_txt(text: str, source_name: str) -> dict[str, str]:
    row = {
        "source_txt": get_source_txt_filename(source_name),
        "txt在的資料夾名稱": get_source_txt_folder_name(source_name),
    }

    for field, pattern in VALIDATOR_HEADER_PATTERNS.items():
        row[field] = extract_validator_field(pattern, text)

    row["RTP Game"] = calculate_game_rtp(row["BaseGameRtp"], row["FreeGameRtp"])

    for pool_index, rtp in re.findall(r"pool\[(\d+)\].*?rtp\s*=\s*([0-9.]+)", text):
        row[f"pool[{pool_index}] rtp"] = rtp

    row.update(extract_max_multiplier_fields(text))
    row["_is_stress_test"] = "1" if has_stress_test_summary(text) else ""
    return row


def is_supported_archive_name(name: str) -> bool:
    return Path(name).suffix.lower() in ARCHIVE_EXTENSIONS


def is_supported_archive_bytes(data: bytes) -> bool:
    return data.startswith(b"PK\x03\x04") or data.startswith(b"Rar!\x1a\x07")


def collect_validator_rows_from_zip(zf: zipfile.ZipFile, archive_label: str, depth: int) -> list[dict[str, str]]:
    rows = []
    for info in zf.infolist():
        if info.is_dir():
            continue
        name = info.filename
        data = zf.read(info)
        rows.extend(collect_validator_rows_from_member(data, f"{archive_label}!{name}", depth + 1))

    return rows


def collect_validator_rows_from_rar(rf: rarfile.RarFile, archive_label: str, depth: int) -> list[dict[str, str]]:
    rows = []
    for info in rf.infolist():
        if info.isdir():
            continue
        name = info.filename
        data = rf.read(info)
        rows.extend(collect_validator_rows_from_member(data, f"{archive_label}!{name}", depth + 1))

    return rows


def collect_validator_rows_from_extracted_dir(extract_dir: Path, archive_label: str, depth: int) -> list[dict[str, str]]:
    rows = []
    for path in extract_dir.rglob("*"):
        if not path.is_file():
            continue
        relative_name = path.relative_to(extract_dir).as_posix()
        rows.extend(collect_validator_rows_from_member(path.read_bytes(), f"{archive_label}!{relative_name}", depth + 1))
    return rows


def find_sevenzip_tool() -> str | None:
    tool = shutil.which("7z") or shutil.which("7zz") or shutil.which("7za")
    if tool:
        return tool

    for candidate in SEVENZIP_CANDIDATES:
        if Path(candidate).exists():
            return candidate

    return None


def configure_rarfile_tools():
    sevenzip = find_sevenzip_tool()
    if sevenzip:
        rarfile.SEVENZIP_TOOL = sevenzip
    elif shutil.which("tar") and not shutil.which("bsdtar"):
        rarfile.BSDTAR_TOOL = "tar"

    try:
        rarfile.tool_setup(force=True)
    except rarfile.RarCannotExec:
        pass


def extract_archive_with_external_tool(archive_path: Path, extract_dir: Path):
    unrar = shutil.which("unrar")
    unar = shutil.which("unar")
    sevenzip = find_sevenzip_tool()
    bsdtar = shutil.which("bsdtar")
    tar = shutil.which("tar")

    commands = []
    if unrar:
        commands.append(("unrar", [unrar, "x", "-y", str(archive_path), f"{extract_dir}{os.sep}"]))
    if unar:
        commands.append(("unar", [unar, "-f", "-o", str(extract_dir), str(archive_path)]))
    if sevenzip:
        commands.append(("7-Zip", [sevenzip, "x", "-y", f"-o{extract_dir}", str(archive_path)]))
    if bsdtar:
        commands.append(("bsdtar", [bsdtar, "-xf", str(archive_path), "-C", str(extract_dir)]))
    elif tar:
        commands.append(("tar", [tar, "-xf", str(archive_path), "-C", str(extract_dir)]))

    if not commands:
        raise RuntimeError("找不到可用的 RAR 解壓工具，請安裝 unrar、unar、7-Zip 或 bsdtar。")

    errors = []
    for tool_name, command in commands:
        result = subprocess.run(command, capture_output=True, text=True, encoding="utf-8", errors="replace")
        if result.returncode == 0:
            return
        error_output = (result.stderr or result.stdout).strip() or f"exit code {result.returncode}"
        errors.append(f"{tool_name}: {error_output}")

    raise RuntimeError("RAR 解壓失敗；" + "；".join(errors))


def collect_validator_rows_from_rar_path(rar_path: Path, archive_label: str, depth: int) -> list[dict[str, str]]:
    configure_rarfile_tools()
    try:
        with rarfile.RarFile(rar_path) as rf:
            return collect_validator_rows_from_rar(rf, archive_label, depth)
    except Exception as rar_error:
        extract_dir = Path(tempfile.mkdtemp(prefix="telegram_bot_rar_"))
        try:
            try:
                extract_archive_with_external_tool(rar_path, extract_dir)
            except Exception as extract_error:
                raise RuntimeError(f"RAR 讀取失敗: {rar_error}; 外部解壓也失敗: {extract_error}") from extract_error
            return collect_validator_rows_from_extracted_dir(extract_dir, archive_label, depth)
        finally:
            shutil.rmtree(extract_dir, ignore_errors=True)


def collect_validator_rows_from_member(data: bytes, source_name: str, depth: int = 0) -> list[dict[str, str]]:
    if depth > MAX_ARCHIVE_DEPTH:
        raise ValueError(f"壓縮檔巢狀層數超過 {MAX_ARCHIVE_DEPTH} 層: {source_name}")

    lower_name = source_name.lower()
    if lower_name.endswith(".txt"):
        text = decode_report_text(data)
        row = parse_validator_txt(text, source_name)
        # 壓縮檔可能混有執行 log；只有帶 GameID 的 TXT 才是 Validator 報表。
        return [row] if row.get("GameID") else []

    if not is_supported_archive_name(source_name) and not is_supported_archive_bytes(data):
        return []

    archive_bytes = io.BytesIO(data)
    if lower_name.endswith(".zip") or data.startswith(b"PK\x03\x04"):
        with zipfile.ZipFile(archive_bytes) as zf:
            return collect_validator_rows_from_zip(zf, source_name, depth)

    if lower_name.endswith(".rar") or data.startswith(b"Rar!\x1a\x07"):
        ensure_output_dir()
        temp_rar_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".rar", dir=OUTPUT_DIR) as temp_rar:
                temp_rar.write(data)
                temp_rar_path = Path(temp_rar.name)

            return collect_validator_rows_from_rar_path(temp_rar_path, source_name, depth)
        finally:
            if temp_rar_path and temp_rar_path.exists():
                temp_rar_path.unlink()

    return []


def collect_validator_rows(archive_path: Path) -> tuple[list[dict[str, str]], list[str]]:
    lower_name = archive_path.name.lower()
    if lower_name.endswith(".zip"):
        with zipfile.ZipFile(archive_path) as zf:
            rows = collect_validator_rows_from_zip(zf, archive_path.name, 0)
    elif lower_name.endswith(".rar"):
        rows = collect_validator_rows_from_rar_path(archive_path, archive_path.name, 0)
    else:
        rows = collect_validator_rows_from_member(archive_path.read_bytes(), archive_path.name)

    pool_fields = {field for row in rows for field in row if field.startswith("pool[")}

    pool_headers = sorted(pool_fields, key=lambda value: int(re.search(r"\[(\d+)\]", value).group(1)))
    return rows, pool_headers


def collect_validator_rows_from_archives(archive_paths: list[Path]) -> tuple[list[dict[str, str]], list[str]]:
    rows = []
    pool_fields = set()

    for archive_path in archive_paths:
        archive_rows, _ = collect_validator_rows(archive_path)
        rows.extend(archive_rows)
        pool_fields.update(field for row in archive_rows for field in row if field.startswith("pool["))

    pool_headers = sorted(pool_fields, key=lambda value: int(re.search(r"\[(\d+)\]", value).group(1)))
    return rows, pool_headers


def write_validator_xlsx(rows: list[dict[str, str]], pool_headers: list[str], output_path: Path):
    if not rows:
        raise NoValidatorReportError("檔案內找不到含 GameID 的 RTP Validator 報表。")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Summary"

    output_columns = [*VALIDATOR_OUTPUT_COLUMNS, *((header, header) for header in pool_headers)]
    headers = [display_name for _, display_name in output_columns]

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row_index, row in enumerate(rows, start=2):
        sheet.append(
            [
                to_excel_rtp_value(row.get(source_name, ""))
                if is_rtp_output_field(source_name)
                else to_excel_number_value(row.get(source_name, ""))
                if source_name in THOUSANDS_OUTPUT_FIELDS
                else row.get(source_name, "")
                for source_name, _ in output_columns
            ]
        )
        for column_index, (source_name, _) in enumerate(output_columns, start=1):
            if is_rtp_output_field(source_name):
                sheet.cell(row=row_index, column=column_index).number_format = "0.00%"
            elif source_name in THOUSANDS_OUTPUT_FIELDS:
                sheet.cell(row=row_index, column=column_index).number_format = "#,##0"

    for column_cells, (source_name, _) in zip(sheet.columns, output_columns):
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_dimension = sheet.column_dimensions[column_cells[0].column_letter]
        column_dimension.width = min(max_length + 2, 32)
        column_dimension.hidden = source_name in HIDDEN_VALIDATOR_COLUMNS

    workbook.save(output_path)


def generate_xlsx_from_archive(archive_path: Path, output_path: Path):
    rows, pool_headers = collect_validator_rows(archive_path)
    write_validator_xlsx(rows, pool_headers, output_path)
    return output_path


def generate_xlsx_from_archives(archive_paths: list[Path], output_path: Path):
    rows, pool_headers = collect_validator_rows_from_archives(archive_paths)
    write_validator_xlsx(rows, pool_headers, output_path)
    return output_path


async def download_validator_archive(document, context: ContextTypes.DEFAULT_TYPE, timestamp: str, index: int) -> Path:
    """下載 Telegram 報表檔（TXT/ZIP/RAR）。"""
    mime_type = (document.mime_type or "").lower()
    fallback_suffix = ".txt" if mime_type.startswith("text/plain") else ".zip"
    source_name = document.file_name or f"validator_{timestamp}{fallback_suffix}"
    safe_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(source_name).stem).strip("._") or f"validator_{timestamp}"
    source_suffix = Path(source_name).suffix.lower()
    archive_suffix = source_suffix if source_suffix in REPORT_FILE_EXTENSIONS else ".zip"
    archive_path = OUTPUT_DIR / f"{timestamp}_{index:02d}_{safe_stem}{archive_suffix}"

    telegram_file = await context.bot.get_file(document.file_id)
    await telegram_file.download_to_drive(custom_path=str(archive_path))
    return archive_path


async def build_validator_report_from_documents(documents, context: ContextTypes.DEFAULT_TYPE, batch_name: str | None = None) -> Path:
    """下載 Telegram 報表檔，確認內容有效後合併轉成 xlsx。"""
    ensure_output_dir()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if batch_name:
        safe_output_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", batch_name).strip("._") or f"validator_{timestamp}"
    elif len(documents) == 1:
        source_name = documents[0].file_name or f"validator_{timestamp}.zip"
        safe_output_stem = re.sub(r"[^A-Za-z0-9._-]+", "_", Path(source_name).stem).strip("._") or f"validator_{timestamp}"
    else:
        safe_output_stem = f"validator_batch_{timestamp}"

    archive_paths = []
    xlsx_path = OUTPUT_DIR / f"{timestamp}_{safe_output_stem}.xlsx"
    try:
        for index, document in enumerate(documents, start=1):
            archive_paths.append(await download_validator_archive(document, context, timestamp, index))

        rows, pool_headers = collect_validator_rows_from_archives(archive_paths)
        if not rows:
            raise NoValidatorReportError("上傳檔案不是 RTP Validator 報表，或報表內缺少 GameID。")

        if any(row.get("_is_stress_test") for row in rows):
            xlsx_path = OUTPUT_DIR / f"{timestamp}_{safe_output_stem}_壓測.xlsx"

        write_validator_xlsx(rows, pool_headers, xlsx_path)
    finally:
        for archive_path in archive_paths:
            if archive_path.exists():
                archive_path.unlink()

    prune_old_reports()
    return xlsx_path


async def build_validator_report_from_document(document, context: ContextTypes.DEFAULT_TYPE) -> Path:
    return await build_validator_report_from_documents([document], context)


async def send_typing_while_waiting(chat_id: int, context: ContextTypes.DEFAULT_TYPE, stop_event: asyncio.Event):
    """在長時間處理完成前，持續顯示 Telegram typing 狀態。"""
    while not stop_event.is_set():
        try:
            await context.bot.send_chat_action(chat_id=chat_id, action="typing")
        except Exception as e:
            print(f"[系統日誌] 發送 typing 狀態失敗: {e}")
            return

        try:
            await asyncio.wait_for(stop_event.wait(), timeout=4)
        except asyncio.TimeoutError:
            continue


def is_supported_archive_document(document) -> bool:
    file_name = (document.file_name or "").lower()
    mime_type = (document.mime_type or "").lower()
    file_suffix = Path(file_name).suffix.lower()
    if file_suffix:
        return file_suffix in REPORT_FILE_EXTENSIONS

    is_supported_mime = "zip" in mime_type or "rar" in mime_type or mime_type.startswith("text/plain")
    return is_supported_mime


async def process_archive_batch_after_delay(batch_key, context: ContextTypes.DEFAULT_TYPE):
    try:
        await asyncio.sleep(ARCHIVE_BATCH_DELAY_SECONDS)
    except asyncio.CancelledError:
        return

    batch = archive_batches.pop(batch_key, None)
    if not batch:
        return

    message = batch["message"]
    documents = batch["documents"]
    chat_id = message.chat.id
    batch_name = f"validator_batch_{len(documents)}files"

    stop_typing_event = asyncio.Event()
    typing_task = asyncio.create_task(send_typing_while_waiting(chat_id, context, stop_typing_event))

    try:
        xlsx_path = await build_validator_report_from_documents(documents, context, batch_name=batch_name)
    except NoValidatorReportError as e:
        print(f"[系統日誌] 上傳檔案不是有效報表: {e}")
        await message.reply_text("找不到 RTP Validator 報表，請確認 TXT 內含 GameID，或 ZIP/RAR 內含有效報表 TXT。")
    except Exception as e:
        print(f"[系統日誌] 批次轉換報表檔為 xlsx 失敗: {e}")
        await message.reply_text("報表檔轉換失敗，請確認 txt/zip/rar 檔案格式是否正確。")
    else:
        with xlsx_path.open("rb") as report_file:
            await message.reply_document(document=report_file, filename=xlsx_path.name)
    finally:
        stop_typing_event.set()
        await typing_task


def parse_rtp_report(report_text: str):
    """解析 Formal RTP Report 文字，整理出摘要資料。"""
    if "Formal RTP Report" not in report_text or "Scenario RTP" not in report_text:
        return None

    report_id = ""
    generated_at = ""
    machine_type = ""
    extra_bet = ""
    bet = ""
    old_hand_init = ""
    section = None
    scenario_rows = []
    setting_rows = []
    current_row = None

    def finish_current_row():
        nonlocal current_row
        if not current_row:
            return
        if current_row["section"] == "scenario":
            scenario_rows.append(current_row)
        elif current_row["section"] == "setting":
            setting_rows.append(current_row)
        current_row = None

    for raw_line in report_text.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        if line.startswith("Report ID:"):
            report_id = line.split(":", 1)[1].strip()
            continue
        if line.startswith("Generated At:"):
            generated_at = line.split(":", 1)[1].strip()
            continue
        if line.startswith("- machineType ="):
            machine_type = line.split("=", 1)[1].strip()
            continue
        if line.startswith("- extraBet ="):
            extra_bet = line.split("=", 1)[1].strip()
            continue
        if line.startswith("- Bet ="):
            bet = line.split("=", 1)[1].strip()
            continue
        if line.startswith("- OldHand Init ="):
            old_hand_init = line.split("=", 1)[1].strip()
            continue

        if line == "Scenario RTP":
            finish_current_row()
            section = "scenario"
            continue
        if line == "SettingID RTP":
            finish_current_row()
            section = "setting"
            continue
        if line == "SettingID Trigger Count":
            finish_current_row()
            break

        if section not in {"scenario", "setting"}:
            continue

        is_scenario_row = section == "scenario" and not line.startswith("-")
        is_setting_row = section == "setting" and bool(re.match(r"^SettingID \d+", line))
        if is_scenario_row or is_setting_row:
            finish_current_row()
            current_row = {
                "label": line,
                "game": "",
                "jackpot": "",
                "bonus": "",
                "section": section,
            }
            continue

        if current_row and line.startswith("-"):
            match = re.match(r"-\s*(Game|Jackpot|Bonus Game) RTP\s*=\s*([0-9.]+%)", line)
            if not match:
                continue
            rtp_type, value = match.groups()
            if rtp_type == "Game":
                current_row["game"] = value
            elif rtp_type == "Jackpot":
                current_row["jackpot"] = value
            elif rtp_type == "Bonus Game":
                current_row["bonus"] = value

    finish_current_row()

    if not scenario_rows:
        return None

    return {
        "report_id": report_id,
        "generated_at": generated_at,
        "machine_type": machine_type,
        "extra_bet": extra_bet,
        "bet": bet,
        "old_hand_init": old_hand_init,
        "scenario_rows": scenario_rows,
        "setting_rows": setting_rows,
    }


def normalize_row_label(label: str) -> str:
    """將報表項目名稱整理成較精簡的顯示格式。"""
    if label == "整體 RTP":
        return "Total RTP"

    match = re.match(r"^(SettingID\s+\d+)\s+[^（(]+[（(]([^）)]+)[）)]$", label)
    if match:
        _, display_name = match.groups()
        short_name_map = {
            "一般": "沒有機制",
            "新手救援": "新手救援",
            "老手救援": "老手救援",
            "新手體驗 D": "新手體驗 FG",
            "新手 mini game 體驗 D": "新手體驗 JP",
        }
        return short_name_map.get(display_name, display_name)
    return label


def get_display_width(text: str) -> int:
    """估算字串在 monospace 環境中的顯示寬度。"""
    width = 0
    for char in text:
        width += 2 if unicodedata.east_asian_width(char) in {"W", "F"} else 1
    return width


def pad_display_text(text: str, target_width: int) -> str:
    """依顯示寬度補空白，避免中英混排時表格錯位。"""
    padding = max(0, target_width - get_display_width(text))
    return text + (" " * padding)


def pad_item_label(text: str, target_width: int) -> str:
    """微調項目欄位顯示。"""
    padded = pad_display_text(text, target_width)
    if text == "Total RTP" and padded.endswith(" "):
        return padded[:-1]
    return padded


def format_item_label(text: str) -> str:
    """調整項目欄位顯示文字。"""
    if text not in {"項目", "Total RTP", "Total RTP加總"}:
        return f"{text} "
    return text


def extract_bet_type(extra_bet: str) -> str:
    """從 extraBet 欄位取出數值型 BetType。"""
    match = re.match(r"^(\d+)", extra_bet.strip())
    return match.group(1) if match else extra_bet


BF_ZERO_LINK_FOUR_BONUS_BETS = {
    "3": {Decimal("40.5"), Decimal("50"), Decimal("75"), Decimal("100")},
    "4": {Decimal("250"), Decimal("500")},
}


def is_zero_link_four_bonus_bf_bet(bet_type: str, bet: str) -> bool:
    try:
        bet_value = Decimal(bet.strip())
    except (InvalidOperation, AttributeError):
        return False
    return bet_value in BF_ZERO_LINK_FOUR_BONUS_BETS.get(bet_type, set())


def percent_to_float(value: str) -> float:
    """將百分比字串轉成數值。"""
    return float(value.rstrip("%"))


def float_to_percent(value: float) -> str:
    """將數值轉成百分比字串。"""
    return f"{value:.2f}%"


SCENARIO_COLUMN_SPECS = [
    ("overall", "整體"),
    ("first_50", "~50"),
    ("first_200", "~200"),
    ("old_before_200", "OG ~200"),
    ("old_after_200", "OG 200~"),
]


def classify_scenario_row(label: str) -> str | None:
    compact_label = re.sub(r"\s+", "", label)
    if "整體RTP" in compact_label:
        return "overall"
    if "前50次Spin" in compact_label:
        return "first_50"
    if "前200次Spin" in compact_label or "前100次Spin" in compact_label:
        return "first_200"
    if "RTP" in compact_label and any(marker in compact_label for marker in ("<=200", "≤200", "<=100", "≤100")):
        return "old_before_200"
    if "RTP" in compact_label and (
        "以後" in compact_label or any(marker in compact_label for marker in (">200", ">100"))
    ):
        return "old_after_200"
    return None


def normalize_percent(value: str) -> str:
    if not value:
        return "-"
    try:
        return float_to_percent(percent_to_float(value))
    except ValueError:
        return "-"


def get_setting_game_rtp(setting_rows: list[dict], setting_kind: str) -> str:
    matched_values = []
    for row in setting_rows:
        label = row["label"]
        if setting_kind == "newbie":
            is_match = bool(re.search(r"SettingID\s+(?:2|4|5|6|7|8)\b", label))
        else:
            is_match = bool(re.search(r"SettingID\s+3\b", label))
        if is_match and row["game"]:
            matched_values.append(percent_to_float(row["game"]))
    return float_to_percent(sum(matched_values)) if matched_values else ""


def values_pass(values: list[str], targets: list[float], tolerance: float, skip_indexes: set[int] | None = None) -> bool | None:
    skip_indexes = skip_indexes or set()
    checked_count = 0
    for index, (value, target) in enumerate(zip(values, targets)):
        if index in skip_indexes:
            continue
        if not value:
            return False
        checked_count += 1
        if abs(percent_to_float(value) - target) > tolerance:
            return False
    return True if checked_count else None


def values_pass_with_tolerances(
    values: list[str],
    targets: list[float],
    tolerances: list[float],
    skip_indexes: set[int] | None = None,
) -> bool | None:
    """依各欄不同容許範圍判斷 RTP。"""
    skip_indexes = skip_indexes or set()
    checked_count = 0
    for index, (value, target, tolerance) in enumerate(zip(values, targets, tolerances)):
        if index in skip_indexes:
            continue
        if not value:
            return False
        checked_count += 1
        if abs(percent_to_float(value) - target) > tolerance:
            return False
    return True if checked_count else None


def game_values_pass(values: list[str], targets: list[float], tolerance: float) -> bool:
    """Game 前三格只需小於 100%，後兩格才依雙基準判斷。"""
    if len(values) < 5 or any(not value for value in values[:5]):
        return False

    if any(not 0 <= percent_to_float(value) < 100 for value in values[:3]):
        return False

    return all(
        abs(percent_to_float(values[index]) - targets[index]) <= tolerance
        for index in (3, 4)
    )


def format_check_result(result: bool | None) -> str:
    if result is None:
        return "—"
    return "✅" if result else "❌"


def format_rtp_value_line(values: list[str]) -> str:
    normalized = [normalize_percent(value) for value in values]
    return "|".join(normalized)


def format_rtp_table_value(value: str, standard: str = "") -> str:
    """格式化 Rich Message 表格中的 RTP 數值與該格判斷標準。"""
    normalized = normalize_percent(value)
    if not normalized:
        return "-"
    escaped_value = html.escape(normalized)
    return f"{escaped_value} ({html.escape(standard)})" if standard else escaped_value


def format_rtp_report_summary(parsed_report: dict) -> str:
    """依 Scenario/SettingID 輸出 Telegram Rich Message RTP 驗收表格。"""
    scenario_by_key = {}
    for row in parsed_report["scenario_rows"]:
        if key := classify_scenario_row(row["label"]):
            scenario_by_key[key] = row

    ordered_rows = [scenario_by_key.get(key, {}) for key, _ in SCENARIO_COLUMN_SPECS]
    game_values = [row.get("game", "") for row in ordered_rows]
    link_values = [row.get("jackpot", "") for row in ordered_rows]
    bonus_values = [row.get("bonus", "") for row in ordered_rows]
    bet_type = extract_bet_type(parsed_report["extra_bet"])
    is_bf = bet_type in {"3", "4"}
    old_hand_init = parsed_report.get("old_hand_init", "").strip().lower()
    is_old_hand = (
        old_hand_init == "enabled"
        if old_hand_init
        else any(value and percent_to_float(value) > 0 for value in link_values)
    )
    try:
        bet_value = Decimal(parsed_report["bet"].strip())
    except (InvalidOperation, AttributeError):
        bet_value = None
    # 一般模式直接看實際 Bet；Bet < 2 為低押注，Bet >= 2 為高押注。
    # 不套用 NB／EX1 等模式倍率。
    is_low_bet = bet_value is not None and bet_value < Decimal("2")
    mode = "老手" if is_old_hand else "新手"
    special_mode_name = BET_TYPE_TO_STRESS_MODE.get(bet_type, "")
    special_multiplier = None
    special_setting_missing = False
    if is_bf:
        game = next(
            (
                game
                for game in load_stress_games()
                if game["game_id"] == parsed_report.get("machine_type", "")
            ),
            None,
        )
        if game is not None:
            special_multiplier = parse_stress_multiplier(
                game["bets"].get(special_mode_name, "")
            )
        special_setting_missing = special_multiplier is None or bet_value is None

    if is_bf:
        game_targets = [92.5] * len(game_values)
        game_result = game_values_pass(game_values, game_targets, 1.0)
    else:
        if is_old_hand:
            game_before_target = 94.0 if is_low_bet else 92.0
        else:
            game_before_target = 93.0
        game_after_target = 94.0 if is_low_bet else 92.0
        game_targets = [0.0, 0.0, 0.0, game_before_target, game_after_target]
        game_result = game_values_pass(game_values, game_targets, 1.0)

    if is_bf:
        is_special_low_bet = (
            special_multiplier is not None
            and bet_value is not None
            and bet_value / special_multiplier < Decimal("2")
        )
        if is_special_low_bet:
            link_before_target = 0.0
            bonus_before_target = 4.0
            link_after_target = 0.0
            bonus_after_target = 4.0
        else:
            link_before_target = 2.0 if is_old_hand else 0.0
            bonus_before_target = 2.0 if is_old_hand else 4.0
            link_after_target = 2.0
            bonus_after_target = 2.0
        link_targets = [0.0, 0.0, 0.0, link_before_target, link_after_target]
        bonus_targets = [0.0, 0.0, 0.0, bonus_before_target, bonus_after_target]
    else:
        if is_old_hand:
            link_before_target = 0.0 if is_low_bet else 2.0
        else:
            link_before_target = 0.0
        link_after_target = 0.0 if is_low_bet else 2.0
        link_targets = [0.0, 0.0, 0.0, link_before_target, link_after_target]
        bonus_targets = [0.0, 0.0, 0.0, 2.0, 2.0]

    # 全域規則：Link 目標為 0%時必須精確等於 0%；
    # 只有非 0%的 Link 目標才套用 ±1.5%容許範圍。
    link_tolerances = [0.0 if target == 0.0 else 1.5 for target in link_targets]

    judged_scenario_skip_indexes = {0, 1, 2}
    link_result = values_pass_with_tolerances(
        link_values,
        link_targets,
        link_tolerances,
        skip_indexes=judged_scenario_skip_indexes,
    )
    bonus_result = values_pass(
        bonus_values,
        bonus_targets,
        0.5,
        skip_indexes=judged_scenario_skip_indexes,
    )
    if special_setting_missing:
        link_result = None
        bonus_result = None

    newbie_value = get_setting_game_rtp(parsed_report["setting_rows"], "newbie")
    veteran_value = get_setting_game_rtp(parsed_report["setting_rows"], "veteran")
    newbie_result = values_pass([newbie_value], [1.0], 1.0) if newbie_value else None
    veteran_result = values_pass([veteran_value], [1.0], 1.0) if veteran_value else None

    column_title = "|".join(label for _, label in SCENARIO_COLUMN_SPECS)
    column_labels = column_title.split("|")
    game_standards = ["<100%", "<100%", "<100%", f"{game_targets[3]:g}%", f"{game_targets[4]:g}%"]
    link_standards = ["", "", "", f"{link_targets[3]:g}%", f"{link_targets[4]:g}%"]
    bonus_standards = ["", "", "", f"{bonus_targets[3]:g}%", f"{bonus_targets[4]:g}%"]
    if special_setting_missing:
        missing_label = f"缺{special_mode_name}設定"
        if is_old_hand:
            link_standards[3] = missing_label
            bonus_standards[3] = missing_label
        link_standards[4] = missing_label
        bonus_standards[4] = missing_label

    def table_row(label: str, values: list[str], standards: list[str], result: bool | None) -> str:
        cells = "".join(
            f"<td>{format_rtp_table_value(value, standard)}</td>"
            for value, standard in zip(values, standards)
        )
        return f"<tr><th>{html.escape(label)}</th>{cells}<td>{format_check_result(result)}</td></tr>"

    rows = [
        table_row("Game", game_values, game_standards, game_result),
        table_row("Link", link_values, link_standards, link_result),
        table_row("Bonus", bonus_values, bonus_standards, bonus_result),
    ]
    if not is_bf and newbie_value:
        rows.append(
            table_row(
                "新手體驗",
                [newbie_value, "", "", newbie_value, ""],
                ["1%", "", "", "1%", ""],
                newbie_result,
            )
        )
    if not is_bf:
        rows.append(
            table_row(
                "老手救援",
                [veteran_value, "", "", "", veteran_value],
                ["1%", "", "", "", "1%"],
                veteran_result,
            )
        )

    headers = "".join(f"<th>{html.escape(label)}</th>" for label in column_labels)
    basic_info = "\n".join(
        [
            f"* Report ID: {parsed_report['report_id']}",
            f"* Time: {parsed_report['generated_at']}",
            f"* Game ID: {parsed_report['machine_type']}",
            f"* BetType: {bet_type}",
            f"* Bet: {parsed_report['bet']}",
            f"* Mode: {mode}",
        ]
    )
    report_blocks = [
        "<h3>📋 基本資訊</h3>",
        f"<pre>{html.escape(basic_info)}</pre>",
        "<p><br></p>",
        "<h3>📊 RTP 資訊</h3>",
        "<table bordered striped>",
        f"<tr><th>項目</th>{headers}<th>結果</th></tr>",
        *rows,
        "</table>",
    ]
    if special_setting_missing:
        report_blocks.append(
            f"<p>⚠️ game_list.md 缺少 Game ID {html.escape(parsed_report.get('machine_type', ''))}"
            f" 的{html.escape(special_mode_name)}押注設定，無法依倍率判斷特殊模式。</p>"
        )
    return "\n".join(report_blocks)


async def send_rtp_report_summary(message, parsed_report: dict) -> None:
    """透過 Bot API 傳送支援原生表格的 Rich Message。"""
    payload = json.dumps(
        {
            "chat_id": message.chat.id,
            "rich_message": {"html": format_rtp_report_summary(parsed_report)},
        },
        ensure_ascii=False,
    ).encode("utf-8")

    def send_request() -> None:
        request = urllib.request.Request(
            f"https://api.telegram.org/bot{TG_TOKEN}/sendRichMessage",
            data=payload,
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        with urllib.request.urlopen(request, timeout=30) as response:
            result = json.load(response)
        if not result.get("ok"):
            raise RuntimeError(f"sendRichMessage failed: {result.get('description', 'unknown error')}")

    await asyncio.to_thread(send_request)


def find_stress_game_list_path() -> Path | None:
    return next((path for path in STRESS_GAME_LIST_CANDIDATES if path.is_file()), None)


def load_stress_games() -> list[dict]:
    """從 game_list.md 讀取可供壓測選擇的遊戲及押注倍率。"""
    game_list_path = find_stress_game_list_path()
    if game_list_path is None:
        return []

    lines = game_list_path.read_text(encoding="utf-8").splitlines()
    headers = []
    games_by_id = {}
    for line in lines:
        if not line.lstrip().startswith("|"):
            continue
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
        if not headers and "Game ID" in cells and "遊戲中文名稱" in cells:
            headers = cells
            continue
        if not headers or len(cells) != len(headers) or all(set(cell) <= {"-", ":"} for cell in cells):
            continue
        row = dict(zip(headers, cells))
        game_id = row.get("Game ID", "").strip()
        chinese_name = row.get("遊戲中文名稱", "").strip()
        if not game_id.isdigit() or not chinese_name:
            continue
        games_by_id[game_id] = {
            "game_id": game_id,
            "parsheet_id": row.get("ParSheet ID", "").strip(),
            "english_name": row.get("遊戲英文名稱", "").strip(),
            "chinese_name": chinese_name,
            "bets": {
                mode: row.get(column, "").strip()
                for mode, column in STRESS_MODE_TO_COLUMN.items()
            },
        }
    return list(games_by_id.values())


def parse_stress_multiplier(value: str) -> Decimal | None:
    value = (value or "").strip()
    if not value or value == "-":
        return None
    try:
        multiplier = Decimal(value)
    except InvalidOperation:
        return None
    return multiplier if multiplier > 0 else None


def format_stress_number(value: Decimal) -> str:
    normalized = value.normalize()
    return format(normalized, "f")


def get_stress_bet_boundaries(mode: str, multiplier: Decimal) -> tuple[Decimal, Decimal]:
    """一般模式固定用2／100；BF、SF才乘上Game List倍率。"""
    if mode in {"NB", "EX1", "EX2"}:
        return Decimal("2"), Decimal("100")
    return multiplier * Decimal("2"), multiplier * Decimal("100")


def build_stress_test_items(game: dict) -> list[dict]:
    """依一般／特殊模式級距建立小／中／大押注與新／老手測項。"""
    items = []
    for mode in ("NB", "EX1", "EX2", "BF", "SF"):
        multiplier = parse_stress_multiplier(game["bets"].get(mode, ""))
        if multiplier is None:
            continue
        medium_start, large_start = get_stress_bet_boundaries(mode, multiplier)
        ranges = [
            ("small", "小押注", f"x < {format_stress_number(medium_start)}"),
            (
                "medium",
                "中押注",
                f"{format_stress_number(medium_start)} ≤ x ≤ {format_stress_number(large_start)}",
            ),
            ("large", "大押注", f"{format_stress_number(large_start)} < x"),
        ]
        for size_key, size_label, range_text in ranges:
            if mode in {"NB", "EX1", "EX2"}:
                if size_key == "small" and multiplier >= Decimal("2"):
                    continue
                if size_key == "medium" and multiplier > Decimal("100"):
                    continue
            for hand_key, hand_label in (("new", "新手"), ("old", "老手")):
                if mode in {"NB", "EX1", "EX2"} and size_key == "large" and hand_key == "new":
                    continue
                items.append(
                    {
                        "key": f"{mode}:{size_key}:{hand_key}",
                        "mode": mode,
                        "size": size_key,
                        "hand": hand_key,
                        "label": f"{mode} {size_label} {hand_label}",
                        "range": range_text,
                    }
                )
    return items


def load_stress_progress() -> dict:
    if not STRESS_PROGRESS_FILE.is_file():
        return {}
    try:
        data = json.loads(STRESS_PROGRESS_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        logger.exception("無法讀取壓測進度檔：%s", STRESS_PROGRESS_FILE)
        return {}
    return data if isinstance(data, dict) else {}


def save_stress_progress(progress: dict) -> None:
    STRESS_PROGRESS_FILE.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = STRESS_PROGRESS_FILE.with_suffix(".json.tmp")
    temporary_path.write_text(
        json.dumps(progress, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary_path.replace(STRESS_PROGRESS_FILE)


def build_stress_game_keyboard(games: list[dict], page: int = 0) -> InlineKeyboardMarkup:
    page_count = max(1, (len(games) + STRESS_GAMES_PER_PAGE - 1) // STRESS_GAMES_PER_PAGE)
    page = min(max(page, 0), page_count - 1)
    start = page * STRESS_GAMES_PER_PAGE
    rows = [[InlineKeyboardButton("🔍 直接輸入搜尋", callback_data="stress:search:x")]]
    game_buttons = [
        InlineKeyboardButton(
            game["chinese_name"],
            callback_data=f"stress:game:{game['game_id']}",
        )
        for game in games[start:start + STRESS_GAMES_PER_PAGE]
    ]
    rows.extend(
        game_buttons[index:index + 2]
        for index in range(0, len(game_buttons), 2)
    )
    if page > 0:
        rows.append([InlineKeyboardButton("⬅️ 上一頁", callback_data=f"stress:page:{page - 1}")])
    if page + 1 < page_count:
        rows.append([InlineKeyboardButton("下一頁 ➡️", callback_data=f"stress:page:{page + 1}")])
    rows.append([InlineKeyboardButton("關閉", callback_data="stress:close:x")])
    return InlineKeyboardMarkup(rows)


def search_stress_games(games: list[dict], query: str) -> list[dict]:
    keyword = re.sub(r"\s+", "", query).casefold()
    if not keyword:
        return []

    def searchable_values(game: dict) -> list[str]:
        return [
            game.get("game_id", ""),
            game.get("parsheet_id", ""),
            game.get("english_name", ""),
            game.get("chinese_name", ""),
        ]

    exact_matches = [
        game
        for game in games
        if any(re.sub(r"\s+", "", value).casefold() == keyword for value in searchable_values(game))
    ]
    if exact_matches:
        return exact_matches
    return [
        game
        for game in games
        if any(keyword in re.sub(r"\s+", "", value).casefold() for value in searchable_values(game))
    ]


def build_stress_search_results_keyboard(matches: list[dict]) -> InlineKeyboardMarkup:
    result_buttons = [
        InlineKeyboardButton(
            game["chinese_name"],
            callback_data=f"stress:game:{game['game_id']}",
        )
        for game in matches[:20]
    ]
    rows = [
        result_buttons[index:index + 2]
        for index in range(0, len(result_buttons), 2)
    ]
    rows.append([InlineKeyboardButton("重新搜尋", callback_data="stress:search:x")])
    rows.append([InlineKeyboardButton("返回所有遊戲", callback_data="stress:page:0")])
    return InlineKeyboardMarkup(rows)


def format_stress_test_list(game: dict) -> str:
    items = build_stress_test_items(game)
    if not items:
        return (
            f"🧪 {game['game_id']}_{game['chinese_name']}\n\n"
            "game_list.md 沒有寫 NB、EX1、EX2、BF 或 SF 的押注設定，請先更新。"
        )

    completed = load_stress_progress().get(game["game_id"], {})
    lines = [
        f"🧪 {game['game_id']}_{game['chinese_name']}",
        f"進度：{sum(item['key'] in completed for item in items)}/{len(items)}",
        "",
    ]
    current_mode = None
    for item in items:
        if item["mode"] != current_mode:
            if current_mode is not None:
                lines.append("")
            current_mode = item["mode"]
            lines.append(current_mode)
        marker = "✅" if item["key"] in completed else "⬜"
        lines.append(f"{marker} {item['label']}（{item['range']}）")
    return "\n".join(lines)


def stress_report_core_passed(parsed_report: dict) -> bool:
    summary = format_rtp_report_summary(parsed_report)
    for label in ("Game", "Link", "Bonus"):
        match = re.search(
            rf"<tr><th>{label}</th>.*?<td>(✅|❌|—)</td></tr>",
            summary,
            flags=re.DOTALL,
        )
        if not match or match.group(1) != "✅":
            return False
    return True


def is_stress_report_old_hand(parsed_report: dict) -> bool:
    old_hand_init = parsed_report.get("old_hand_init", "").strip().lower()
    if old_hand_init:
        return old_hand_init == "enabled"
    return any(
        row.get("jackpot") and percent_to_float(row["jackpot"]) > 0
        for row in parsed_report.get("scenario_rows", [])
    )


def update_stress_progress_from_report(parsed_report: dict) -> str | None:
    """通過的 RTP 報表自動更新對應壓測測項，回傳更新說明。"""
    if not stress_report_core_passed(parsed_report):
        return None

    games = {game["game_id"]: game for game in load_stress_games()}
    game_id = parsed_report.get("machine_type", "")
    game = games.get(game_id)
    bet_type = extract_bet_type(parsed_report.get("extra_bet", ""))
    mode = BET_TYPE_TO_STRESS_MODE.get(bet_type)
    if game is None or mode is None:
        return None

    multiplier = parse_stress_multiplier(game["bets"].get(mode, ""))
    if multiplier is None:
        return None
    try:
        bet_value = Decimal(parsed_report["bet"].strip())
    except (InvalidOperation, AttributeError):
        return None

    medium_start, large_start = get_stress_bet_boundaries(mode, multiplier)
    if bet_value < medium_start:
        size = "small"
    elif bet_value <= large_start:
        size = "medium"
    else:
        size = "large"
    hand = "old" if is_stress_report_old_hand(parsed_report) else "new"
    item_key = f"{mode}:{size}:{hand}"
    item = next((item for item in build_stress_test_items(game) if item["key"] == item_key), None)
    if item is None:
        return None

    progress = load_stress_progress()
    game_progress = progress.setdefault(game_id, {})
    already_completed = item_key in game_progress
    game_progress[item_key] = {
        "report_id": parsed_report.get("report_id", ""),
        "time": parsed_report.get("generated_at", ""),
        "bet": parsed_report.get("bet", ""),
    }
    save_stress_progress(progress)
    status = "已更新" if not already_completed else "已覆寫"
    return f"✅ /stress_test {status}：{game_id}_{game['chinese_name']}｜{item['label']}"


@require_mention_in_groups
async def stress_test_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user is None or update.effective_user.id != ALLOWED_USER_ID:
        await update.message.reply_text("🚫 你沒有權限使用這個指令。")
        return
    games = load_stress_games()
    if not games:
        await update.message.reply_text("找不到 game_list.md，請先更新並上傳檔案。")
        return
    if context.args:
        keyword = " ".join(context.args)
        matches = search_stress_games(games, keyword)
        if len(matches) == 1:
            await update.message.reply_text(
                format_stress_test_list(matches[0]),
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("返回遊戲列表", callback_data="stress:page:0")]]
                ),
            )
        elif matches:
            await update.message.reply_text(
                f"找到 {len(matches)} 個結果，請選擇：",
                reply_markup=build_stress_search_results_keyboard(matches),
            )
        else:
            await update.message.reply_text(f"找不到「{keyword}」，請確認 game_list.md 是否已更新。")
        return
    await update.message.reply_text(
        "請選擇要查看的遊戲：",
        reply_markup=build_stress_game_keyboard(games),
    )


async def stress_test_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if query.from_user is None or query.from_user.id != ALLOWED_USER_ID:
        await query.answer("🚫 你沒有權限", show_alert=True)
        return
    await query.answer()
    parts = (query.data or "").split(":")
    if len(parts) != 3 or parts[0] != "stress":
        return
    _, action, value = parts
    games = load_stress_games()
    if not games:
        await query.edit_message_text("找不到 game_list.md，請先更新並上傳檔案。")
        return
    if action == "page":
        stress_test_search_state.pop(query.message.chat.id, None)
        await query.edit_message_text(
            "請選擇要查看的遊戲：",
            reply_markup=build_stress_game_keyboard(games, int(value)),
        )
    elif action == "search":
        await query.edit_message_text(
            "請直接輸入 Game ID、遊戲中文／英文名稱或 ParSheet ID：",
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton("取消搜尋", callback_data="stress:page:0")]]
            ),
        )
        stress_test_search_state[query.message.chat.id] = query.message.message_id
    elif action == "game":
        stress_test_search_state.pop(query.message.chat.id, None)
        game = next((game for game in games if game["game_id"] == value), None)
        if game is None:
            await query.edit_message_text("game_list.md 找不到這個遊戲，請更新資料。")
            return
        await query.edit_message_text(
            format_stress_test_list(game),
            reply_markup=InlineKeyboardMarkup(
                [[InlineKeyboardButton("返回遊戲列表", callback_data="stress:page:0")]]
            ),
        )
    elif action == "close":
        stress_test_search_state.pop(query.message.chat.id, None)
        await query.message.delete()


async def document_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.document:
        return

    chat_id = update.message.chat.id
    if update.message.chat.type != "private":
        caption = update.message.caption or ""
        mentioned = f"@{context.bot.username}" in caption
        if mentioned:
            activate_mention_window(chat_id)
        elif not is_mention_window_active(chat_id):
            return

    document = update.message.document
    if not is_supported_archive_document(document):
        await update.message.reply_text("目前只支援上傳 RTP Validator 的 txt、zip、rar 檔案。")
        return

    media_group_id = update.message.media_group_id
    if media_group_id:
        batch_key = (update.message.chat.id, media_group_id)
        batch = archive_batches.setdefault(
            batch_key,
            {
                "documents": [],
                "message": update.message,
                "task": None,
            },
        )
        batch["documents"].append(document)
        batch["message"] = update.message

        if batch["task"] and not batch["task"].done():
            batch["task"].cancel()
        batch["task"] = asyncio.create_task(process_archive_batch_after_delay(batch_key, context))
        return

    stop_typing_event = asyncio.Event()
    typing_task = asyncio.create_task(send_typing_while_waiting(update.message.chat.id, context, stop_typing_event))

    try:
        xlsx_path = await build_validator_report_from_document(document, context)
    except NoValidatorReportError as e:
        print(f"[系統日誌] 上傳檔案不是有效報表: {e}")
        await update.message.reply_text("這個檔案不是 RTP Validator 報表，請確認 TXT 內含 GameID。")
    except Exception as e:
        print(f"[系統日誌] 轉換報表檔為 xlsx 失敗: {e}")
        await update.message.reply_text("報表檔轉換失敗，請確認 txt/zip/rar 檔案格式是否正確。")
    else:
        with xlsx_path.open("rb") as report_file:
            await update.message.reply_document(document=report_file, filename=xlsx_path.name)
    finally:
        stop_typing_event.set()
        await typing_task


# ==================== 🎰 H026 模擬器指令 ====================


def build_help_text() -> str:
    games = discover_games()
    game_list = ", ".join(get_game_id(name) for name in games) if games else "(尚未找到任何遊戲資料夾)"
    return (
        "🤖 *Jumbo 工具機器人 — 功能說明*\n\n"
        "📄 *RTP Validator 報表轉表格*\n"
        "私訊直接丟 txt/zip/rar 檔案（支援巢狀壓縮、多檔一起丟）；群組裡上傳要附文字 @我，我會先判斷是否為報表，再抓出各項 RTP、最大倍數與 pool RTP 等欄位，自動轉成 xlsx 傳回來。\n\n"
        "📊 *RTP Report 文字摘要*\n"
        "私訊我或在群組 @我，貼上 Formal RTP Report 文字，我會整理五個 Scenario 的 Game / Link / Bonus RTP、新手體驗與老手救援，並依容許範圍標示是否通過。\n\n"
        "🕐 *群組 3 分鐘窗口*\n"
        f"在群組裡 @我一次之後，{MENTION_WINDOW_SECONDS // 60} 分鐘內丟報表檔或貼 RTP Report 文字都會自動處理，不用每次都 @我。\n\n"
        "🧪 *壓測清單*（僅限管理者）\n"
        "傳送 `/stress_test` 後可從全部遊戲選擇，或直接輸入 Game ID、遊戲名稱、ParSheet ID 搜尋；通過的 Formal RTP Report 會自動更新對應壓測項目。\n\n"
        "🎰 *模擬器*（僅限管理者）\n"
        "傳送 `/simulator` 後選擇「開始模擬」；選好遊戲後，可在步驟 2 選擇單一 config 或 Batch 多組執行。跑完會自動把摘要跟報表路徑傳回來。\n"
        f"目前可用遊戲: {game_list}"
    )


def build_simulator_help_text() -> str:
    """Simulator 選單內的 Help：只說明模擬功能，顯示後直接結束選單。"""
    games = discover_games()
    game_list = ", ".join(get_game_id(name) for name in games) if games else "(尚未找到任何遊戲資料夾)"
    return (
        "🎰 Simulator 功能說明\n\n"
        "• 開始模擬：先選擇遊戲，再於步驟 2 選擇單一 config 或 Batch (多組)。\n"
        "• 單組：選擇 config 後，繼續設定 bet_mode、卡池與局數。\n"
        "• Batch (多組)：取得 BATCH_COMBINATIONS 範本；複製、修改並整段送出，即可依序執行多組參數。\n"
        "• Status：查看目前執行項目、耗時與批次進度。\n"
        "• Cancel：停止目前模擬，並清除尚未執行的批次。\n"
        "• 完成後會回傳模擬摘要與 VM 上的報表路徑。\n\n"
        f"目前可用遊戲：{game_list}\n"
        f"單組局數範圍：1,000 ~ {MAX_ROUNDS:,}\n"
        f"批次上限：一次 {MAX_BATCH_COMBINATIONS} 組\n\n"
        "說明結束；需要操作時請重新傳送 /simulator。"
    )


@require_mention_in_groups
async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(build_help_text(), parse_mode=ParseMode.MARKDOWN, reply_markup=build_menu_keyboard())


def build_status_text() -> str:
    if current_job is None:
        return "目前沒有模擬在跑。"
    elapsed = int(time.time() - current_job["started_at"])
    batch_text = ""
    if current_batch is not None:
        running_number = current_batch["completed"] + 1
        batch_text = f"\n批次進度: {running_number}/{current_batch['total']}"
    return f"⏳ 執行中: {current_job['label']}\n已耗時 {elapsed}s{batch_text}"


async def cancel_current_job() -> str:
    global current_job, current_batch
    async with job_lock:
        if current_job is None:
            current_batch = None
            return "目前沒有模擬在跑。"
        queued_count = len(current_batch["jobs"]) if current_batch is not None else 0
        if current_batch is not None:
            current_batch["cancelled"] = True
            current_batch["jobs"].clear()
            current_batch = None
        current_job["proc"].terminate()
        label = current_job["label"]
    queue_text = f"；並清除後續 {queued_count} 組批次" if queued_count else ""
    return f"🛑 已送出中止指令: {label}{queue_text}"


async def start_simulation_job(
    context: ContextTypes.DEFAULT_TYPE,
    chat_id: int,
    game_name: str,
    game_dir: Path,
    config: str,
    bet_mode: str,
    rounds: int,
    card: str,
    card_enabled: bool = True,
    batch_job: bool = False,
) -> str:
    """驗證參數並啟動模擬工作，回傳要回覆給使用者的訊息文字。"""
    global current_job

    game_script = game_dir / "Simulator.py"
    configs = discover_configs(game_dir)
    config = config.upper()
    if config not in configs:
        return f"❌ {game_name} 的 config 必須是: {', '.join(configs) if configs else '(找不到 config_*.js 檔案)'}"
    if bet_mode not in BET_MODES:
        return "❌ bet_mode 必須是 0, 1 或 2"
    if rounds < 1000 or rounds > MAX_ROUNDS:
        return f"❌ rounds 必須介於 1,000 ~ {MAX_ROUNDS:,} 之間"
    if card not in {"new", "old"}:
        return "❌ card 必須是 new 或 old"
    if not game_script.exists():
        return f"❌ 找不到模擬器腳本: {game_script}"

    async with job_lock:
        if current_batch is not None and not batch_job:
            return "⏳ 已經有批次在跑，請先取消或等它跑完。"
        if current_job is not None:
            elapsed = int(time.time() - current_job["started_at"])
            return f"⏳ 已經有模擬在跑了（{current_job['label']}，已耗時 {elapsed}s），請先用 /simulator 選單的 Cancel 或等它跑完。"

        card_label = card if card_enabled else "off"
        label = f"game={game_name} config={config} bet_mode={BET_MODES[bet_mode]} rounds={rounds:,} card={card_label}"
        env = os.environ.copy()
        env["PYTHONUNBUFFERED"] = "1"
        env_prefix = get_game_id(game_name).upper()
        env[f"{env_prefix}_RUN_ALL_COMBINATIONS"] = "false"
        env[f"{env_prefix}_CONFIG_FILE"] = f"config_{config}.js"
        env[f"{env_prefix}_BET_MODE"] = bet_mode
        env[f"{env_prefix}_TOTAL_ROUNDS"] = str(rounds)
        env[f"{env_prefix}_CARD_SYSTEM_ENABLED"] = "true" if card_enabled else "false"
        env[f"{env_prefix}_CARD_SYSTEM_IS_NEWBIE"] = "true" if card == "new" else "false"

        proc = await asyncio.create_subprocess_exec(
            sys.executable,
            str(game_script),
            cwd=str(game_dir),
            env=env,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE,
        )
        current_job = {"proc": proc, "chat_id": chat_id, "started_at": time.time(), "label": label}

    asyncio.create_task(watch_simulator_job(context, proc, chat_id, label))
    return f"🚀 開始模擬: {label}\n跑完會傳結果回來。"


async def start_simulation_batch(context: ContextTypes.DEFAULT_TYPE, chat_id: int, jobs: list[dict]) -> str:
    """建立批次佇列並啟動第一組；其餘工作由 watcher 依序接續。"""
    global current_batch
    async with job_lock:
        if current_job is not None or current_batch is not None:
            return "⏳ 已經有模擬或批次在跑，請先取消或等它跑完。"
        current_batch = {"chat_id": chat_id, "jobs": list(jobs), "total": len(jobs), "completed": 0, "cancelled": False}

    first_job = current_batch["jobs"].pop(0)
    reply = await start_simulation_job(context, chat_id, batch_job=True, **first_job)
    if not reply.startswith("🚀"):
        current_batch = None
        return reply
    return f"📚 已接受 {len(jobs)} 組批次參數，會依序執行。\n\n[1/{len(jobs)}] {reply}"


async def advance_simulation_batch(context: ContextTypes.DEFAULT_TYPE, chat_id: int) -> None:
    """目前工作結束後推進同一批次的下一組。"""
    global current_batch
    batch = current_batch
    if batch is None or batch["chat_id"] != chat_id or batch["cancelled"]:
        return
    batch["completed"] += 1
    if not batch["jobs"]:
        total = batch["total"]
        current_batch = None
        await context.bot.send_message(chat_id, f"🏁 批次執行完成，共 {total} 組。")
        return

    next_number = batch["completed"] + 1
    next_job = batch["jobs"].pop(0)
    reply = await start_simulation_job(context, chat_id, batch_job=True, **next_job)
    if not reply.startswith("🚀"):
        current_batch = None
        await context.bot.send_message(chat_id, f"❌ 批次在第 {next_number} 組停止：\n{reply}")
        return
    await context.bot.send_message(chat_id, f"[{next_number}/{batch['total']}] {reply}")


async def watch_simulator_job(context: ContextTypes.DEFAULT_TYPE, proc, chat_id, label):
    global current_job
    stdout, stderr = await proc.communicate()
    async with job_lock:
        if current_job is not None and current_job["proc"] is proc:
            current_job = None

    stdout_text = stdout.decode("utf-8", errors="replace")
    stderr_text = stderr.decode("utf-8", errors="replace")

    if proc.returncode != 0:
        tail = stderr_text[-1500:] if stderr_text else "(無錯誤輸出)"
        await context.bot.send_message(chat_id, f"❌ 模擬失敗（{label}）\nexit code={proc.returncode}\n```\n{tail}\n```", parse_mode=ParseMode.MARKDOWN)
    else:
        summary_lines = SUMMARY_LINE_RE.findall(stdout_text)
        report_match = REPORT_LINE_RE.search(stdout_text)

        if not summary_lines:
            await context.bot.send_message(chat_id, f"⚠️ 模擬跑完但沒有解析到摘要（{label}）\n輸出最後 1000 字:\n```\n{stdout_text[-1000:]}\n```", parse_mode=ParseMode.MARKDOWN)
        else:
            text = f"✅ 模擬完成: {label}\n\n" + "\n".join(summary_lines)
            if report_match:
                report_path = report_match.group(1).strip()
                text += f"\n\n報表檔案（存在 VM 上）: {report_path}"
            await context.bot.send_message(chat_id, text)

    await advance_simulation_batch(context, chat_id)


async def on_error(update, context: ContextTypes.DEFAULT_TYPE):
    logger.exception("Unhandled error", exc_info=context.error)


# ==================== 🕹️ /simulator 按鈕選單 ====================


def build_abort_button() -> InlineKeyboardButton:
    return InlineKeyboardButton("取消", callback_data="sim:abort:x")


def build_menu_keyboard() -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton("開始模擬", callback_data="sim:menu:run")],
        [InlineKeyboardButton("功能說明", callback_data="sim:menu:help")],
        [InlineKeyboardButton("Cancel", callback_data="sim:menu:close")],
    ]
    if current_job is not None:
        rows.extend(
            [
                [InlineKeyboardButton("⌛️顯示狀態", callback_data="sim:menu:status")],
                [InlineKeyboardButton("⌛️取消模擬", callback_data="sim:menu:cancel")],
            ]
        )
    return InlineKeyboardMarkup(rows)


def build_menu_title() -> str:
    return "請選擇你接下來要做什麼:"


def build_game_keyboard(stage: str = "game") -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton(name, callback_data=f"sim:{stage}:{get_game_id(name)}")] for name in discover_games()]
    rows.append([build_abort_button()])
    return InlineKeyboardMarkup(rows)


def build_config_keyboard(game_dir: Path, configs: list[str], include_batch: bool = False) -> InlineKeyboardMarkup:
    rows = []
    for config in configs:
        math_version = get_config_math_version(game_dir, config)
        version_label = math_version or "未知"
        rows.append(
            [
                InlineKeyboardButton(
                    f"config_{config}｜數學版本 {version_label}",
                    callback_data=f"sim:config:{config}",
                )
            ]
        )
    if include_batch:
        rows.append([InlineKeyboardButton("Batch (多組)", callback_data="sim:batchtemplate:x")])
    rows.append([build_abort_button()])
    return InlineKeyboardMarkup(rows)


def build_bet_keyboard() -> InlineKeyboardMarkup:
    buttons = [InlineKeyboardButton(label, callback_data=f"sim:bet:{key}") for key, label in BET_MODES.items()]
    rows = [[button] for button in buttons]
    rows.append([build_abort_button()])
    return InlineKeyboardMarkup(rows)


def build_card_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("新手卡池 (new)", callback_data="sim:card:new")],
            [InlineKeyboardButton("一般卡池 (old)", callback_data="sim:card:old")],
            [build_abort_button()],
        ]
    )


def build_rounds_keyboard() -> InlineKeyboardMarkup:
    preset_buttons = [InlineKeyboardButton(label, callback_data=f"sim:rounds:{value}") for label, value in ROUNDS_PRESETS]
    rows = [[button] for button in preset_buttons]
    rows.append([InlineKeyboardButton("其他（輸入數字）", callback_data="sim:rounds:custom")])
    rows.append([build_abort_button()])
    return InlineKeyboardMarkup(rows)


def build_confirm_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        [
            [InlineKeyboardButton("確認執行", callback_data="sim:confirm:yes")],
            [InlineKeyboardButton("取消", callback_data="sim:confirm:no")],
        ]
    )


def build_wizard_summary_text(state: dict) -> str:
    return (
        "請確認模擬參數:\n"
        f"遊戲: {state['game_name']}\n"
        f"config: {state['config']}\n"
        f"bet_mode: {BET_MODES[state['bet_mode']]}\n"
        f"rounds: {state['rounds']:,}\n"
        f"card: {state['card']}"
    )


@require_mention_in_groups
async def simulator_menu_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_user is None or update.effective_user.id != ALLOWED_USER_ID:
        await update.message.reply_text("🚫 你沒有權限使用這個指令。")
        return
    await update.message.reply_text(build_menu_title(), parse_mode=ParseMode.MARKDOWN, reply_markup=build_menu_keyboard())


async def simulator_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if query.from_user is None or query.from_user.id != ALLOWED_USER_ID:
        await query.answer("🚫 你沒有權限", show_alert=True)
        return
    await query.answer()

    parts = (query.data or "").split(":")
    if len(parts) != 3 or parts[0] != "sim":
        return
    _, stage, value = parts
    chat_id = query.message.chat.id

    if stage == "abort":
        simulator_wizard_state.pop(chat_id, None)
        await query.edit_message_text("已取消選單。")
        return

    if stage == "menu":
        if value == "help":
            simulator_wizard_state.pop(chat_id, None)
            await query.edit_message_text(build_simulator_help_text())
        elif value == "status":
            await query.edit_message_text(build_status_text(), reply_markup=build_menu_keyboard())
        elif value == "cancel":
            await query.edit_message_text(await cancel_current_job(), reply_markup=build_menu_keyboard())
        elif value == "run":
            simulator_wizard_state[chat_id] = {}
            await query.edit_message_text("步驟 1/5：請選擇遊戲", reply_markup=build_game_keyboard())
        elif value == "batch":
            simulator_wizard_state[chat_id] = {}
            await query.edit_message_text("批次模式：請先選擇遊戲", reply_markup=build_game_keyboard("batchgame"))
        elif value == "close":
            await query.message.delete()
            simulator_wizard_state.pop(chat_id, None)
        return

    state = simulator_wizard_state.setdefault(chat_id, {})

    if stage == "batchtemplate":
        game_name = state.get("game_name")
        game_dir_value = state.get("game_dir")
        if not game_name or not game_dir_value:
            simulator_wizard_state.pop(chat_id, None)
            await query.edit_message_text("❌ 選單狀態不完整，請重新 /simulator。")
            return
        game_dir = Path(game_dir_value)
        configs = discover_configs(game_dir)
        if not configs:
            simulator_wizard_state.pop(chat_id, None)
            await query.edit_message_text(f"❌ {game_name} 資料夾裡沒有 config_*.js 檔案。")
            return
        state["awaiting_batch"] = True
        template = build_batch_template(game_name, configs)
        await query.edit_message_text(
            f"已進入 {game_name} 的 Batch 模式。\n請複製下一則範本，直接修改後整段送出。\n"
            f"可增刪 {{...}}；局數可寫 10000 或 10**4；布林值請用 True / False；一次最多 {MAX_BATCH_COMBINATIONS} 組。"
        )
        await query.message.reply_text(f"<pre>{html.escape(template)}</pre>", parse_mode=ParseMode.HTML)
        return

    if stage == "batchgame":
        found = find_game_by_id(value)
        if not found:
            simulator_wizard_state.pop(chat_id, None)
            await query.edit_message_text("❌ 找不到這個遊戲，請重新 /simulator。")
            return
        game_name, game_dir = found
        configs = discover_configs(game_dir)
        if not configs:
            simulator_wizard_state.pop(chat_id, None)
            await query.edit_message_text(f"❌ {game_name} 資料夾裡沒有 config_*.js 檔案。")
            return
        state.update({"game_name": game_name, "game_dir": str(game_dir)})
        state["awaiting_batch"] = True
        template = build_batch_template(game_name, configs)
        await query.edit_message_text(
            f"已選擇 {game_name}。\n請複製下一則範本，直接修改後整段送出。\n"
            f"可增刪 {{...}}；局數可寫 10000 或 10**4；布林值請用 True / False；一次最多 {MAX_BATCH_COMBINATIONS} 組。"
        )
        await query.message.reply_text(f"<pre>{html.escape(template)}</pre>", parse_mode=ParseMode.HTML)
        return

    if stage == "game":
        found = find_game_by_id(value)
        if not found:
            simulator_wizard_state.pop(chat_id, None)
            await query.edit_message_text("❌ 找不到這個遊戲，請重新 /simulator。")
            return
        game_name, game_dir = found
        configs = discover_configs(game_dir)
        if not configs:
            simulator_wizard_state.pop(chat_id, None)
            await query.edit_message_text(f"❌ {game_name} 資料夾裡沒有 config_*.js 檔案。")
            return
        state["game_name"] = game_name
        state["game_dir"] = str(game_dir)
        await query.edit_message_text(
            f'步驟 2/5：請選擇 "{game_name}" 的配置',
            reply_markup=build_config_keyboard(game_dir, configs, include_batch=True),
        )
        return

    if stage == "config":
        state["config"] = value
        await query.edit_message_text("步驟 3/5：請選擇 bet_mode", reply_markup=build_bet_keyboard())
        return

    if stage == "bet":
        state["bet_mode"] = value
        await query.edit_message_text("步驟 4/5：請選擇卡池", reply_markup=build_card_keyboard())
        return

    if stage == "card":
        state["card"] = value
        await query.edit_message_text("步驟 5/5：請選擇模擬局數", reply_markup=build_rounds_keyboard())
        return

    if stage == "rounds":
        if value == "custom":
            state["awaiting_rounds"] = True
            await query.edit_message_text(f"請直接輸入模擬局數（純數字，1,000 ~ {MAX_ROUNDS:,}）")
            return
        state["rounds"] = int(value)
        state.pop("awaiting_rounds", None)
        await query.edit_message_text(build_wizard_summary_text(state), reply_markup=build_confirm_keyboard())
        return

    if stage == "confirm":
        if value == "no":
            simulator_wizard_state.pop(chat_id, None)
            await query.edit_message_text("已取消。")
            return
        final_state = simulator_wizard_state.pop(chat_id, None)
        if not final_state or not all(key in final_state for key in ("game_name", "game_dir", "config", "bet_mode", "rounds", "card")):
            await query.edit_message_text("❌ 選單狀態不完整，請重新 /simulator。")
            return
        reply = await start_simulation_job(
            context, chat_id, final_state["game_name"], Path(final_state["game_dir"]), final_state["config"], final_state["bet_mode"], final_state["rounds"], final_state["card"]
        )
        await query.edit_message_text(reply)
        return


# ==================== ⚡ Telegram 事件監聽 ====================


async def main_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.text:
        return

    chat_id = update.message.chat.id
    text = update.message.text.strip()
    chat_type = update.message.chat.type
    bot_username = context.bot.username

    if (
        chat_id in stress_test_search_state
        and update.effective_user
        and update.effective_user.id == ALLOWED_USER_ID
    ):
        keyword = text.replace(f"@{bot_username}", "").strip()
        games = load_stress_games()
        matches = search_stress_games(games, keyword)
        if not matches:
            await update.message.reply_text(
                f"找不到「{keyword}」。請重新輸入，或確認 game_list.md 是否已更新。",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("取消搜尋", callback_data="stress:page:0")]]
                ),
            )
            return
        prompt_message_id = stress_test_search_state.pop(chat_id, None)
        if prompt_message_id is not None:
            try:
                await context.bot.delete_message(chat_id=chat_id, message_id=prompt_message_id)
            except Exception as exc:
                logger.warning("無法刪除壓測搜尋提示訊息：%s", exc)
        if len(matches) == 1:
            await update.message.reply_text(
                format_stress_test_list(matches[0]),
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("返回遊戲列表", callback_data="stress:page:0")]]
                ),
            )
        else:
            await update.message.reply_text(
                f"找到 {len(matches)} 個結果，請選擇：",
                reply_markup=build_stress_search_results_keyboard(matches),
            )
        return

    # /simulator 選單正在等待使用者輸入自訂局數
    wizard_state = simulator_wizard_state.get(chat_id)
    if wizard_state and wizard_state.get("awaiting_batch") and update.effective_user and update.effective_user.id == ALLOWED_USER_ID:
        batch_text = text.replace(f"@{bot_username}", "").strip()
        try:
            combinations = parse_batch_combinations(batch_text)
            jobs = normalize_batch_combinations(wizard_state["game_name"], Path(wizard_state["game_dir"]), combinations)
        except ValueError as exc:
            await update.message.reply_text(f"❌ 批次格式有誤：{exc}\n\n請修改後重新整段送出；目前仍在等待批次參數。")
            return
        simulator_wizard_state.pop(chat_id, None)
        await update.message.reply_text(await start_simulation_batch(context, chat_id, jobs))
        return

    if wizard_state and wizard_state.get("awaiting_rounds") and update.effective_user and update.effective_user.id == ALLOWED_USER_ID:
        if not text.isdigit() or not (1000 <= int(text) <= MAX_ROUNDS):
            await update.message.reply_text(f"❌ 請輸入 1,000 ~ {MAX_ROUNDS:,} 之間的正整數。")
            return
        wizard_state["rounds"] = int(text)
        wizard_state.pop("awaiting_rounds", None)
        await update.message.reply_text(build_wizard_summary_text(wizard_state), reply_markup=build_confirm_keyboard())
        return

    is_private = chat_type == "private"
    mentioned = f"@{bot_username}" in text
    if not is_private and mentioned:
        activate_mention_window(chat_id)

    # 情況一：私訊，或群組裡直接 @機器人 → 完整回應（解析 RTP 報表，不然就回「你好」）
    if is_private or mentioned:
        clean_text = text.replace(f"@{bot_username}", "").strip()
        normalized_command = clean_text.lstrip("/").strip().lower()

        # 容錯：允許「@機器人 simulator」「@機器人 /simulator」這種先 @ 再打指令名稱的寫法
        if normalized_command == "simulator":
            await simulator_menu_cmd(update, context)
            return
        if normalized_command == "help":
            await help_cmd(update, context)
            return

        if parsed_report := parse_rtp_report(clean_text):
            await send_rtp_report_summary(update.message, parsed_report)
            if stress_update := update_stress_progress_from_report(parsed_report):
                await update.message.reply_text(stress_update)
            return

        await update.message.reply_text("你好")
        return

    # 情況二：群組裡在 3 分鐘的 @ 窗口內 → 只默默處理符合條件的內容，不然一律忽略
    if is_mention_window_active(chat_id):
        if parsed_report := parse_rtp_report(text):
            await send_rtp_report_summary(update.message, parsed_report)
            if stress_update := update_stress_progress_from_report(parsed_report):
                await update.message.reply_text(stress_update)


async def post_init(application: Application) -> None:
    """把指令註冊進 Telegram 的「/」選單，這樣可以直接點選，不用打字。"""
    await application.bot.set_my_commands(
        [
            BotCommand("help", "顯示機器人有哪些功能"),
            BotCommand("simulator", "開啟模擬器選單（按鈕操作）"),
            BotCommand("stress_test", "查看遊戲壓測項目與完成進度"),
        ]
    )


if __name__ == "__main__":
    if not acquire_bot_instance_lock():
        print("[錯誤] 機器人已經在本機執行中，請先關閉另一個 bot 程序後再啟動。")
        raise SystemExit(1)

    if not asyncio.run(check_telegram_polling_available()):
        print("[錯誤] Telegram Bot Token 正在被另一個 getUpdates 程序使用。")
        print("請關閉其他 bot 程序，或到 BotFather 重新產生 token 後再啟動。")
        raise SystemExit(1)

    print("正在啟動混合規則型機器人 中...")
    app = Application.builder().token(TG_TOKEN).post_init(post_init).build()
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(CommandHandler("simulator", simulator_menu_cmd))
    app.add_handler(CommandHandler("stress_test", stress_test_cmd))
    app.add_handler(CallbackQueryHandler(simulator_callback, pattern=r"^sim:"))
    app.add_handler(CallbackQueryHandler(stress_test_callback, pattern=r"^stress:"))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, main_handler))
    app.add_handler(MessageHandler(filters.Document.ALL, document_handler))
    app.add_error_handler(on_error)
    print("[系統提示] 機器人已成功上線！(請按 Ctrl + C 即可關閉)")
    app.run_polling()
