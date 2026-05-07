"""
add_result2.py
為既有的 JP*_Output.xlsx 新增 Result2 工作頁。
Result2 是 Result 的展開版（long format）：
  每個 (ID, Bet) 組合的 30 個 Bet Level 各展開為一列。

欄位（14 欄）:
  ID, Bet, Inc_JP1-4, Startup_JP1-4, JP1, JP2, JP3, JP4

用法:
  python add_result2.py                   # 處理同目錄下所有 JP*_Output.xlsx
  python add_result2.py JP0100A_Output.xlsx   # 指定檔案
Jupyter Notebook 中直接執行亦可。
"""

import os
import sys
import glob
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BET_LEVELS = 30

# Result 工作頁欄位配置（1-based）
# 1=ID, 2=Bet
# 3-6  = Inc_JP1-4
# 7-10 = Startup_JP1-4
# 11-40  = JP1 Level1-30
# 41-70  = JP2 Level1-30
# 71-100 = JP3 Level1-30
# 101-130= JP4 Level1-30
COL_ID      = 1
COL_BET     = 2
COL_INC     = list(range(3, 7))      # 3..6
COL_STARTUP = list(range(7, 11))     # 7..10
COL_JP1     = list(range(11, 41))    # 11..40
COL_JP2     = list(range(41, 71))    # 41..70
COL_JP3     = list(range(71, 101))   # 71..100
COL_JP4     = list(range(101, 131))  # 101..130


def make_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

HEADER_FONT  = Font(bold=True, color="FFFFFF")
HEADER_FILL  = make_fill("70AD47")
CENTER       = Alignment(horizontal="center", vertical="center")


def add_result2(wb):
    """讀取 Result 工作頁，新增 Result2；若已存在則先刪除再重建。"""
    if "Result" not in wb.sheetnames:
        print("  ✗ 找不到 Result 工作頁，略過")
        return False

    # 若 Result2 已存在，先刪除
    if "Result2" in wb.sheetnames:
        del wb["Result2"]

    ws_src = wb["Result"]
    ws2    = wb.create_sheet("Result2")

    # 標題列
    headers = ["ID", "Bet",
               "Inc_JP1", "Inc_JP2", "Inc_JP3", "Inc_JP4",
               "Startup_JP1", "Startup_JP2", "Startup_JP3", "Startup_JP4",
               "JP1", "JP2", "JP3", "JP4"]
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(1, c, h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER

    r2 = 2  # Result2 下一個寫入列
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if row[COL_ID - 1] is None:
            continue
        id_val  = row[COL_ID  - 1]
        bet_val = row[COL_BET - 1]
        inc     = [row[c - 1] for c in COL_INC]
        startup = [row[c - 1] for c in COL_STARTUP]
        jp1     = [row[c - 1] for c in COL_JP1]
        jp2     = [row[c - 1] for c in COL_JP2]
        jp3     = [row[c - 1] for c in COL_JP3]
        jp4     = [row[c - 1] for c in COL_JP4]

        for lvl in range(BET_LEVELS):
            out_row = ([id_val, bet_val] + inc + startup +
                       [jp1[lvl], jp2[lvl], jp3[lvl], jp4[lvl]])
            for c, v in enumerate(out_row, 1):
                ws2.cell(r2, c, v)
            r2 += 1

    ws2.freeze_panes = "A2"
    return True


def process_file(path):
    print(f"處理: {path}")
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        print(f"  ✗ 無法開啟: {e}")
        return

    ok = add_result2(wb)
    if not ok:
        return

    try:
        wb.save(path)
        print(f"  ✓ Result2 已寫入: {path}")
    except PermissionError:
        print(f"  ✗ 無法儲存（檔案可能已在 Excel 中開啟）: {path}")
    except Exception as e:
        print(f"  ✗ 儲存失敗: {e}")


def main():
    # 過濾 Jupyter 注入的 --f=kernel-xxx.json 等 "-" 開頭參數
    raw_args = [p for p in sys.argv[1:] if not p.startswith("-")]
    targets  = [p for p in raw_args if p.lower().endswith((".xlsx", ".xlsm"))]

    if not targets:
        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))
        except NameError:
            script_dir = ""
        cwd = os.getcwd()
        search_dirs = list(dict.fromkeys(d for d in [script_dir, cwd] if d))

        for d in search_dirs:
            for f in glob.glob(os.path.join(d, "JP*_Output.xlsx")):
                if f not in targets:
                    targets.append(f)

    if not targets:
        print("找不到符合條件的 JP*_Output.xlsx 檔案。")
        return

    for t in targets:
        process_file(t)

    print("\n全部完成。")


if __name__ == "__main__":
    main()
