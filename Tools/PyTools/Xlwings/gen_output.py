"""
gen_output.py
讀取 JP*.xlsm，依 Parameter_List 的 (ID, Bet) 組合計算 Output 值，
產生 JP*_Output.xlsx（含 Result 與 Result2 兩個工作頁）。

用法:
  python gen_output.py                  # 自動尋找同目錄下 JP*.xlsm
  python gen_output.py JP0100A.xlsm     # 指定檔案
  python gen_output.py JP*.xlsm         # 多檔 (shell glob)
Jupyter Notebook 中直接執行亦可。
"""

import os
import sys
import glob
import math
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

THRESHOLD = 10_000_000_000
BET_LEVELS = 30

# Parameter_List 欄位（1-based）
COL_ID        = 1
COL_DENOM     = 3
COL_MINBET    = 4
COL_BET       = 5
COL_RTP       = [6, 7, 8, 9]    # JP1-JP4
COL_INC       = [10, 11, 12, 13]
COL_STARTUP_D = [14, 15]         # JP1,JP2 in dollars
COL_STARTUP_X = [16, 17]         # JP3,JP4 multiplier

# Setting 工作頁欄位（1-based，資料從第 3 行起）
COL_S_LEVEL = 5   # E
COL_S_BCRED = 6   # F
COL_S_BMULT = 8   # H


def load_setting_levels(ws_setting):
    levels = {}
    for r in range(3, 33):
        lvl = ws_setting.cell(r, COL_S_LEVEL).value
        bc  = ws_setting.cell(r, COL_S_BCRED).value
        bm  = ws_setting.cell(r, COL_S_BMULT).value
        if lvl is not None:
            levels[int(lvl)] = {
                "bet_credit": float(bc) if bc is not None else 0.0,
                "bet_mult":   float(bm) if bm is not None else 0.0,
            }
    return levels


def read_param_list(ws_param):
    params = []
    for r in range(2, ws_param.max_row + 1):
        id_val = ws_param.cell(r, COL_ID).value
        if id_val is None:
            continue
        p = {
            "id":        id_val,
            "denom":     ws_param.cell(r, COL_DENOM).value,
            "minbet":    ws_param.cell(r, COL_MINBET).value,
            "bet":       ws_param.cell(r, COL_BET).value,
            "rtp":       [ws_param.cell(r, c).value for c in COL_RTP],
            "inc":       [ws_param.cell(r, c).value for c in COL_INC],
            "startup_d": [ws_param.cell(r, c).value for c in COL_STARTUP_D],
            "startup_x": [ws_param.cell(r, c).value for c in COL_STARTUP_X],
        }
        params.append(p)
    return params


def safe_float(v):
    return float(v) if v is not None else 0.0


def compute_row(param, levels):
    denom   = safe_float(param["denom"])
    min_bet = safe_float(param["minbet"])
    bet     = safe_float(param["bet"])
    rtp     = [safe_float(v) for v in param["rtp"]]
    inc     = [safe_float(v) for v in param["inc"]]
    sd      = [safe_float(v) for v in param["startup_d"]]
    sx      = [safe_float(v) for v in param["startup_x"]]

    inc_out = inc[:]
    startup_out = [
        int(round(sd[0] / denom)) if denom else 0,
        int(round(sd[1] / denom)) if denom else 0,
        int(sx[0] * min_bet),
        int(sx[1] * min_bet),
    ]

    jp_levels = [[], [], [], []]
    for lvl in range(1, BET_LEVELS + 1):
        lv = levels.get(lvl, {"bet_credit": 0.0, "bet_mult": 0.0})
        bc = lv["bet_credit"]
        bm = lv["bet_mult"]
        D4 = bc * denom * bet
        C = [
            sd[0],
            sd[1],
            sx[0] * min_bet * bm * denom,
            sx[1] * min_bet * bm * denom,
        ]
        for j in range(4):
            startup_rtp = rtp[j] - inc[j]
            if C[j] == 0 or startup_rtp == 0 or D4 == 0:
                jp_levels[j].append(0)
            else:
                val = round(startup_rtp / (C[j] / D4) * THRESHOLD)
                jp_levels[j].append(int(val))

    return inc_out, startup_out, jp_levels


def make_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

HEADER_FONT    = Font(bold=True, color="FFFFFF")
HEADER_FILL_R  = make_fill("4472C4")
HEADER_FILL_R2 = make_fill("70AD47")
CENTER = Alignment(horizontal="center", vertical="center")


def write_result_sheet(wb, rows_data):
    ws = wb.create_sheet("Result")
    headers = ["ID", "Bet"]
    for j in range(1, 5):
        headers.append(f"Inc_JP{j}")
    for j in range(1, 5):
        headers.append(f"Startup_JP{j}")
    for j in range(1, 5):
        for l in range(1, 31):
            headers.append(f"JP{j}_Level{l}")

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL_R
        cell.alignment = CENTER

    for r_idx, (id_val, bet_val, inc, startup, jp1, jp2, jp3, jp4) in enumerate(rows_data, 2):
        row = [id_val, bet_val] + inc + startup + jp1 + jp2 + jp3 + jp4
        for c, v in enumerate(row, 1):
            ws.cell(r_idx, c, v)

    ws.freeze_panes = "A2"
    return ws


def write_result2_sheet(wb, rows_data):
    ws = wb.create_sheet("Result2")
    headers = ["ID", "Bet",
               "Inc_JP1", "Inc_JP2", "Inc_JP3", "Inc_JP4",
               "Startup_JP1", "Startup_JP2", "Startup_JP3", "Startup_JP4",
               "JP1", "JP2", "JP3", "JP4"]

    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL_R2
        cell.alignment = CENTER

    r_idx = 2
    for (id_val, bet_val, inc, startup, jp1, jp2, jp3, jp4) in rows_data:
        for lvl in range(BET_LEVELS):
            row = [id_val, bet_val] + inc + startup + [jp1[lvl], jp2[lvl], jp3[lvl], jp4[lvl]]
            for c, v in enumerate(row, 1):
                ws.cell(r_idx, c, v)
            r_idx += 1

    ws.freeze_panes = "A2"
    return ws


def process_file(src_path):
    print(f"處理: {src_path}")
    try:
        wb_src = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  ✗ 無法開啟來源檔: {e}")
        return

    param_name   = next((s for s in wb_src.sheetnames if s.lower() == "parameter_list"), None)
    setting_name = next((s for s in wb_src.sheetnames if s.lower() == "setting"), None)

    if not param_name:
        print(f"  ✗ 找不到 Parameter_List 工作頁")
        wb_src.close()
        return
    if not setting_name:
        print(f"  ✗ 找不到 Setting 工作頁")
        wb_src.close()
        return

    ws_param   = wb_src[param_name]
    ws_setting = wb_src[setting_name]

    levels = load_setting_levels(ws_setting)
    params = read_param_list(ws_param)
    wb_src.close()

    if not params:
        print(f"  ✗ Parameter_List 無資料")
        return

    print(f"  共 {len(params)} 筆 (ID, Bet) 組合，計算中...")

    rows_data = []
    for p in params:
        inc_out, startup_out, jp_levels = compute_row(p, levels)
        rows_data.append((
            p["id"], p["bet"],
            inc_out, startup_out,
            jp_levels[0], jp_levels[1], jp_levels[2], jp_levels[3],
        ))

    base     = os.path.splitext(os.path.basename(src_path))[0]
    out_name = base + "_Output.xlsx"
    out_path = os.path.join(os.path.dirname(src_path), out_name)

    wb_out = openpyxl.Workbook()
    if "Sheet" in wb_out.sheetnames:
        del wb_out["Sheet"]

    write_result_sheet(wb_out, rows_data)
    write_result2_sheet(wb_out, rows_data)

    try:
        wb_out.save(out_path)
        print(f"  ✓ 已儲存: {out_path}")
    except PermissionError:
        print(f"  ✗ 無法儲存 {out_path}（檔案可能已在 Excel 中開啟）")
    except Exception as e:
        print(f"  ✗ 儲存失敗: {e}")


def main():
    # 過濾 Jupyter 注入的 --f=kernel-xxx.json 等 "-" 開頭參數
    raw_args = [p for p in sys.argv[1:] if not p.startswith("-")]
    targets  = [p for p in raw_args if p.lower().endswith((".xlsm", ".xlsx"))]

    if not targets:
        # 自動搜尋：優先腳本所在目錄，其次 cwd
        try:
            script_dir = os.path.dirname(os.path.abspath(__file__))
        except NameError:
            script_dir = ""
        cwd = os.getcwd()
        search_dirs = list(dict.fromkeys(d for d in [script_dir, cwd] if d))

        for d in search_dirs:
            for f in glob.glob(os.path.join(d, "JP*.xlsm")):
                if f not in targets:
                    targets.append(f)

    if not targets:
        print("找不到符合條件的 JP*.xlsm 檔案。")
        print("請將此腳本放在與 JP*.xlsm 相同目錄，或直接傳入檔案路徑。")
        return

    for t in targets:
        process_file(t)

    print("\n全部完成。")


if __name__ == "__main__":
    main()
