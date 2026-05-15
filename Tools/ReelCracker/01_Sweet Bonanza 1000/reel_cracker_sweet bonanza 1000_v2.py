# %%

# from selenium import webdriver  # 匯入 selenium 的 webdriver

from seleniumwire import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
import time
import pyautogui
import urllib.parse
import pandas as pd
import json
import os  # 新增匯入 os 模組，用來檢查檔案是否存在
from openpyxl import load_workbook
import keyboard
import threading
from pywinauto import Application
import win32gui
import win32api
import win32con
import time


# %% 開啟遊戲畫面

chrome_driver_path = "C:\\Users\\rhinshen\\Mine\\(個人工作區)\\3_Tools\\ReelCracker\\chromedriver-win64\\chromedriver-win64\\chromedriver.exe"
service = Service(chrome_driver_path)
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
options.add_argument("--force-device-scale-factor=1.0")  # 設置縮放比例為 70%

driver = webdriver.Chrome(service=service, options=options)
# driver.get("https://www.pragmaticplay.com/en/games/sweet-bonanza-1000/")
driver.get("https://www.pko99.ph/")
time.sleep(5)

# 關閉廣告業面
button = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CLASS_NAME, "dark-close")))

# 點擊按鈕
button.click()


# %% 登入

# 定位用戶名輸入框
username_field = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CLASS_NAME, "form-input")))  # 替換為實際 class
username_field.clear()
username_field.send_keys("jlmath3")  # 替換為你的用戶名


# 定位密碼輸入框（可選）
password_field = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.CLASS_NAME, "form-input.form-password")))  # 替換為實際 class
password_field.clear()
password_field.send_keys("abc8888")  # 替換為你的密碼

time.sleep(5)


# 定位並點擊登錄按鈕
login_button = WebDriverWait(driver, 3).until(
    EC.element_to_be_clickable(
        (
            By.XPATH,
            '//*[@id="__layout"]/div/div[2]/div[2]/header/div[1]/div/div/div[2]/div/div/div[2]/ul/li[1]/div',
        )
    )  # 替換為實際 class
)
login_button.click()

time.sleep(5)

# %% 選遊戲

# driver.get("https://www.pko99.ph/common/opengame?targetElementId=&isDemo=false&gameItem=%7B%22gameId%22%3A17328,%22onlineTime%22%3A%222024-06-04T17%3A39%3A33.000Z%22,%22gameOrder%22%3A999,%22brand%22%3A%22PPV2%22,%22gameKindId%22%3A1,%22gameCode%22%3A%22PP_vs20fruitswx_1%22,%22demoEngine%22%3A1,%22imagePath%22%3A%22PPV2%2FPlatformGameList.17328.3%22,%22version%22%3A%2214%22,%22gameName%22%3A%22Sweet%20Bonanza%201000%22,%22introduction%22%3Anull,%22playCount%22%3A486200,%22originStatus%22%3A1,%22gameDeviceType%22%3A2,%22status%22%3A1,%22isNew%22%3A0,%22addTime%22%3A%222024-06-04T17%3A39%3A33.000Z%22,%22lastModifyTime%22%3A%222024-07-26T04%3A59%3A00.000Z%22,%22lastModifyOperator%22%3A%22SYSTEM%22,%22isFavorite%22%3A0,%22gameTag%22%3A%221,3%22,%22gameTagName%22%3A%22NEW,promoteGames1%22,%22imgUrl%22%3A%22https%3A%2F%2Fdownload.ocms.cloud%2Fv2%2Fcommon%2FPPV2%2FPlatformGameList.17328.3.webp%3Fversion%3D14%22,%22minImgUrl%22%3A%22https%3A%2F%2Fdownload.ocms.cloud%2Fv2%2Fcommon%2FPPV2%2FPlatformGameList.17328.3.s.webp%3Fversion%3D14%22,%22isIframe%22%3A1%7D&skipCustomerService=true&gameopen=true&version=6.25.49-fc")
driver.get(
    "https://www.pko99.ph/common/opengame?targetElementId=&isDemo=false&gameItem=%7B%22gameId%22%3A2083,%22onlineTime%22%3A%222024-04-11T04%3A15%3A10.000Z%22,%22gameOrder%22%3A999,%22brand%22%3A%22PPV2%22,%22gameKindId%22%3A1,%22gameCode%22%3A%22PP_vs20fruitsw_1%22,%22demoEngine%22%3A1,%22imagePath%22%3A%22PPV2%2FPlatformGameList.2083.3%22,%22version%22%3A%2214%22,%22gameName%22%3A%22Sweet%20Bonanza%22,%22introduction%22%3Anull,%22playCount%22%3A390314,%22originStatus%22%3A1,%22gameDeviceType%22%3A2,%22status%22%3A1,%22isNew%22%3A0,%22addTime%22%3A%222024-04-11T04%3A15%3A10.000Z%22,%22lastModifyTime%22%3A%222024-04-11T04%3A15%3A10.000Z%22,%22lastModifyOperator%22%3A%22SYSTEM%22,%22isFavorite%22%3A0,%22gameTag%22%3A%223%22,%22gameTagName%22%3A%22promoteGames1%22,%22imgUrl%22%3A%22https%3A%2F%2Fdownload.ocms.cloud%2Fv2%2Fcommon%2FPPV2%2FPlatformGameList.2083.3.webp%3Fversion%3D14%22,%22minImgUrl%22%3A%22https%3A%2F%2Fdownload.ocms.cloud%2Fv2%2Fcommon%2FPPV2%2FPlatformGameList.2083.3.s.webp%3Fversion%3D14%22,%22isIframe%22%3A1%7D&skipCustomerService=true"
)
# # 進入slot
# game_button = WebDriverWait(driver, 3).until(
#         EC.element_to_be_clickable((By.XPATH, "//i/img[@alt='Slot game-button']"))
#     )
# driver.execute_script("arguments[0].scrollIntoView(true);", game_button)  # 滾動到元素
# game_button.click()

# # 遊戲名稱
# input_field = WebDriverWait(driver, 5).until(
#         EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Please enter game name.']"))
#     )
# # 清除並輸入文字
# input_field.clear()
# input_field.send_keys("Sweet Bonanza 1000")  # 替換為你想輸入的遊戲名稱

# # 搜尋按鈕
# sear_btn = WebDriverWait(driver,5).until(
#         EC.element_to_be_clickable((By.CLASS_NAME, "search-btn")))
# # 點擊按鈕
# sear_btn.click()

# driver.get("https://www.pko99.ph/common/opengame?targetElementId=&isDemo=false&gameItem=%7B%22gameId%22%3A17328,%22onlineTime%22%3A%222024-06-04T17%3A39%3A33.000Z%22,%22gameOrder%22%3A999,%22brand%22%3A%22PPV2%22,%22gameKindId%22%3A1,%22gameCode%22%3A%22PP_vs20fruitswx_1%22,%22demoEngine%22%3A1,%22imagePath%22%3A%22PPV2%2FPlatformGameList.17328.3%22,%22version%22%3A%2214%22,%22gameName%22%3A%22Sweet%20Bonanza%201000%22,%22introduction%22%3Anull,%22playCount%22%3A485417,%22originStatus%22%3A1,%22gameDeviceType%22%3A2,%22status%22%3A1,%22isNew%22%3A0,%22addTime%22%3A%222024-06-04T17%3A39%3A33.000Z%22,%22lastModifyTime%22%3A%222024-07-26T04%3A59%3A00.000Z%22,%22lastModifyOperator%22%3A%22SYSTEM%22,%22isFavorite%22%3A0,%22gameTag%22%3A%221,3%22,%22gameTagName%22%3A%22NEW,promoteGames1%22,%22imgUrl%22%3A%22https%3A%2F%2Fdownload.ocms.cloud%2Fv2%2Fcommon%2FPPV2%2FPlatformGameList.17328.3.webp%3Fversion%3D14%22,%22minImgUrl%22%3A%22https%3A%2F%2Fdownload.ocms.cloud%2Fv2%2Fcommon%2FPPV2%2FPlatformGameList.17328.3.s.webp%3Fversion%3D14%22,%22isIframe%22%3A1%7D&skipCustomerService=true&gameopen=true&version=6.25.36-fc")
time.sleep(30)  # 等待頁面載入

# %% 找到正確的點擊位置


# 獲取Chrome瀏覽器窗口
def get_chrome_window():
    try:
        app = Application(backend="uia").connect(title_re=".*Panaloko.*Vegas Site.*")
        chrome_window = app.top_window()
        return chrome_window
    except Exception as e:
        print(f"無法找到Chrome窗口：{e}")
        return None


def find_click_positions():
    """幫助找到正確的點擊位置"""
    print("=== 座標定位工具 ===")
    print("請按以下步驟操作：")
    print("1. 將滑鼠移到第一個按鈕上，然後輸入 '1' 並按Enter記錄位置")
    print("2. 將滑鼠移到第二個按鈕上，然後輸入 '2' 並按Enter記錄位置")
    print("3. 輸入 '3' 並按Enter完成設定")
    print("4. 輸入 '4' 並按Enter測試點擊")
    print("5. 輸入 'exit' 並按Enter退出")

    positions = {}

    while True:
        try:
            command = input("\n請輸入指令 (1/2/3/4/exit): ").strip().lower()

            if command == "1":
                pos = pyautogui.position()
                chrome_window = get_chrome_window()
                if chrome_window:
                    rect = chrome_window.rectangle()
                    rel_x = pos.x - rect.left
                    rel_y = pos.y - rect.top
                    positions["pos1"] = (rel_x, rel_y)
                    print(f"位置1已記錄: 相對座標 ({rel_x}, {rel_y}), 絕對座標 ({pos.x}, {pos.y})")
                else:
                    print("無法找到Chrome窗口")

            elif command == "2":
                pos = pyautogui.position()
                chrome_window = get_chrome_window()
                if chrome_window:
                    rect = chrome_window.rectangle()
                    rel_x = pos.x - rect.left
                    rel_y = pos.y - rect.top
                    positions["pos2"] = (rel_x, rel_y)
                    print(f"位置2已記錄: 相對座標 ({rel_x}, {rel_y}), 絕對座標 ({pos.x}, {pos.y})")
                else:
                    print("無法找到Chrome窗口")

            elif command == "3":
                if "pos1" in positions and "pos2" in positions:
                    print(f"\n=== 設定完成 ===")
                    print(f"位置1 (x1, y1): {positions['pos1']}")
                    print(f"位置2 (x2, y2): {positions['pos2']}")
                    print(f"\n請將以下座標複製到程式中：")
                    print(f"x1, y1 = {positions['pos1'][0]}, {positions['pos1'][1]}")
                    print(f"x2, y2 = {positions['pos2'][0]}, {positions['pos2'][1]}")
                else:
                    print("請先記錄兩個位置")

            elif command == "4":
                if "pos1" in positions and "pos2" in positions:
                    chrome_window = get_chrome_window()
                    if chrome_window:
                        print("測試點擊位置1...")
                        click_at_relative_position(chrome_window, positions["pos1"][0], positions["pos1"][1])
                        time.sleep(1)
                        print("測試點擊位置2...")
                        click_at_relative_position(chrome_window, positions["pos2"][0], positions["pos2"][1])
                    else:
                        print("無法找到Chrome窗口")
                else:
                    print("請先記錄兩個位置")

            elif command == "exit":
                break

            else:
                print("無效指令，請輸入 1, 2, 3, 4 或 exit")

        except KeyboardInterrupt:
            break

    return positions


# 如果需要找位置，取消下面這行的註解
find_click_positions()


def show_mouse_position():
    """即時顯示滑鼠座標"""
    print("=== 即時滑鼠座標顯示 ===")
    print("移動滑鼠查看座標，按 ESC 鍵退出")

    try:
        while True:
            pos = pyautogui.position()

            # 同時顯示絕對座標和相對於Chrome窗口的座標
            chrome_window = get_chrome_window()
            if chrome_window:
                rect = chrome_window.rectangle()
                rel_x = pos.x - rect.left
                rel_y = pos.y - rect.top
                print(f"\r滑鼠位置 - 絕對座標: ({pos.x}, {pos.y}) | 相對座標: ({rel_x}, {rel_y})", end="", flush=True)
            else:
                print(f"\r滑鼠位置 - 絕對座標: ({pos.x}, {pos.y}) | Chrome窗口未找到", end="", flush=True)

            time.sleep(0.1)
            if keyboard.is_pressed("esc"):
                break
    except KeyboardInterrupt:
        pass

    print("\n座標顯示已退出")


# 如果需要即時顯示滑鼠座標，取消下面這行的註解
show_mouse_position()


# %% 連續點擊


# 獲取Chrome瀏覽器窗口
def get_chrome_window():
    try:
        app = Application(backend="uia").connect(title_re=".*Panaloko.*Vegas Site.*")
        chrome_window = app.top_window()
        return chrome_window
    except Exception as e:
        print(f"無法找到Chrome窗口：{e}")
        return None


# 獲取Chrome窗口
chrome_window = get_chrome_window()
if chrome_window is None:
    print("請確保Chrome瀏覽器已開啟")
    exit()

print(f"已連接到Chrome窗口: {chrome_window.window_text()}")


# 設定相對座標點 (相對於Chrome窗口)
# 使用上面的 find_click_positions() 工具來找到正確的座標
# x1, y1 = 488, 307  # buy free dpin
x1, y1 = 481, 421  # buy super free dpin
x2, y2 = 1214, 760  # 確認
x3, y3 = 637, 347  # 空白

times = 5000  # 100沒有Retrigger大約兩場
paused = False
stop_program = False


def toggle_pause():
    global paused
    paused = not paused
    print(f"程式已{'暫停' if paused else '繼續'}")


def stop_program_func():
    global stop_program
    stop_program = True
    print("程式已停止")


def click_at_relative_position(window, x, y):
    """在窗口的相對位置點擊，不移動滑鼠游標"""
    try:
        # 直接發送點擊消息到窗口
        hwnd = window.handle
        lParam = win32api.MAKELONG(x, y)

        # 發送滑鼠按下和釋放事件
        win32gui.PostMessage(hwnd, win32con.WM_LBUTTONDOWN, win32con.MK_LBUTTON, lParam)
        time.sleep(0.01)
        win32gui.PostMessage(hwnd, win32con.WM_LBUTTONUP, 0, lParam)

    except Exception as e:
        print(f"點擊失敗: {e}")


# 設置熱鍵
keyboard.add_hotkey("space", toggle_pause)  # 空白鍵暫停/繼續
keyboard.add_hotkey("esc", stop_program_func)  # ESC鍵停止程式

print("開始自動點擊，您現在可以自由使用滑鼠")
print("按空白鍵暫停/繼續，按ESC停止程式")

# 修改你的迴圈
for it in range(times):
    if stop_program:
        break

    while paused:
        time.sleep(0.1)  # 暫停時等待
        if stop_program:
            break

    if stop_program:
        break

    print(f"開始第 {it+1} 次執行...")

    # 使用pywinauto點擊，不移動游標
    click_at_relative_position(chrome_window, x1, y1)
    time.sleep(0.3)
    click_at_relative_position(chrome_window, x2, y2)
    time.sleep(0.3)
    click_at_relative_position(chrome_window, x3, y3)
    time.sleep(0.3)

data_dict = {}

for request in driver.requests:
    if request.response and "gameService" in request.url:
        response_body = request.response.body.decode("utf-8", errors="ignore")
        parsed_data = urllib.parse.parse_qs(response_body)
        parsed_data = {k: v[0] if v else "" for k, v in parsed_data.items()}

        index = parsed_data.get("index")
        if index is not None:
            data_dict[index] = parsed_data

data_list = list(data_dict.values())

print("程式結束")


# 檢查資料並轉換數字類型
def convert_to_numeric(df):
    for column in df.columns:
        # 嘗試將每一列轉換為數字，若失敗則捕捉例外並保留原值
        try:
            df[column] = pd.to_numeric(df[column])
        except ValueError:
            pass  # 如果無法轉換，保留為原始資料
    return df


# 新增儲存為 Excel 且可新增工作表的功能
file_excel = "game_responses.xlsx"
file_json = "game_responses.json"
new_sheet_base = "game_data"

if data_list:
    df = pd.DataFrame(data_list)
    df = convert_to_numeric(df)  # 確保數據類型正確

    if os.path.exists(file_excel):
        # 載入現有 Excel
        book = load_workbook(file_excel)
        existing_sheets = book.sheetnames

        # 避免工作表名稱重複，動態產生新的名稱
        new_sheet_name = new_sheet_base
        i = 1
        while new_sheet_name in existing_sheets:
            new_sheet_name = f"{new_sheet_base}_{i}"
            i += 1

        # 使用 ExcelWriter 新增工作表（不直接設置 book）
        with pd.ExcelWriter(file_excel, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df.to_excel(writer, sheet_name=new_sheet_name, index=False)

        print(f"資料已新增至 {file_excel} 的工作表: {new_sheet_name}")
    else:
        # 檔案不存在，建立新的 Excel 檔案
        with pd.ExcelWriter(file_excel, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=new_sheet_base, index=False)
        print(f"資料已儲存到新檔案 {file_excel} 的工作表: {new_sheet_base}")

    # JSON 部分維持覆寫方式
    with open(file_json, "w") as json_file:
        json.dump(data_list, json_file, indent=4)
    print(f"資料已儲存到 {file_json}")
else:
    print("未找到符合條件的回應資料。")

# 關閉瀏覽器
# driver.quit()

# %%
