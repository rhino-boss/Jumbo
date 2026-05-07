# %%

from seleniumwire import webdriver
from selenium.webdriver.chrome.service import Service
import time
import pyautogui
import urllib.parse
import pandas as pd
import json
import os
from openpyxl import load_workbook
import gzip
import zlib
import keyboard
from pywinauto import Application
import win32gui
import win32api
import win32con


# %% 設定

TARGET_URL = "https://www.ko-panalo.com/common/opengame?payload=0xvl%2Fw9pt3bz0kTOkycqo7W9HxcKeYQvr2fcwC18jQWnJNGqt7udi3H4JrUHhyKjpXgVcPtYhFI7ExLRfkFK%2FeUHPL6wmD7jA3Uany7HqzxZKWPjESKPA6FIWbvd3yX%2FuvzjrL6d3jueKqYrxa%2FfnXIsZUduJ8Duk1l8SWNlFMz3qqJ%2FqftMxkh4%2FTJBdaMYhbPKZOlaAar%2B%2BSWl3fHc2ueGV6zO8EIrcYBWvAcnGqfDyAK3iUaEc18qsERpE4VzeSHcFwNMI0%2B9EKM%2BbX7nWQTFYQ1v0%2BFw0XwLn%2FeJ0cS3XtUGIYEHxCoZxJroITQZ6eA03iz8yWrJss9Z0rPp3z0UTpHxy5ySzxrR7B%2B5yplYjVu8q5yzrST8joN%2BMZllhre6zZd8fnX0K2ZetgON19rJbtDdzyoKvJMG3DzLoRit09Z%2BqJcncWWRm8z%2BuDZuxpu46exNUJBwR6jjyYiHMo51h6zbCp1HF3t7FmxhniBldb5wU%2Bdou%2Fwi5fzXPGBmOu4JOqMWphtQ3Ykyqffbbm3%2BhbztHw5DjMtjZWb6oBqHpNglZpVNNEPYL9%2FqywVF%2Fdqp%2BB6br18r3pabQb8Sh3jS2mjYqeJGOWMECyRJjuq3eZrLfzuOVgx%2FS8Wr9wCaX048j2GDJuPw%2B%2FimHb4KO%2BasdnHK8tPe2ZSVwPfE1S9PUKpsIGKj4JH%2FOrrUQD3tqYXPbkiGrVmLPR4XrIl0SUOsohzVrS1ZSagX0EVWwrF5Yy0NqErj2nd0LkD35UiCQNWP0t1%2BRqTHwl7%2F5tBrVH6WU2WHd6gUe27P8eU5R06ginTkBM2vuf%2BHoaJCz6cHYjH4085Y80vZ7sSxKLBeRNQ26aoMfDpx0WUcA2LhV%2FU%2BiRSjnENv1G%2BnQZU7LVomBoYh8nTdHLlgKdFqJLHjaMXGaDaArfS9jmiTlFwvBpVDTywZsI4zn1dnVyxwdAc0xfcQonyTmbn9hqPKzJwhZw%3D%3D&skipCustomerService=true"

# 這個 title_re 需要視你的 Chrome/站台標題調整；跟 Gates 的寫法保持一致
CHROME_TITLE_RE = ".*Panaloko.*Vegas Site.*"


# %% 開啟遊戲頁面

chrome_driver_path = os.path.join(os.path.dirname(__file__), "chromedriver-win64", "chromedriver.exe")
service = Service(chrome_driver_path)
options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
options.add_argument("--force-device-scale-factor=1.0")

driver = webdriver.Chrome(service=service, options=options)
driver.get(TARGET_URL)


# %% 找座標 / 背景點擊


def get_chrome_window():
    try:
        app = Application(backend="uia").connect(title_re=CHROME_TITLE_RE)
        return app.top_window()
    except Exception as e:
        print(f"找不到 Chrome 視窗：{e}")
        return None


def move_mouse_to_relative_position(window, x, y, duration=0.1):
    """
    把滑鼠移到 Chrome 視窗內的相對座標 (x, y)。
    回傳 (abs_x, abs_y) 絕對座標。
    """
    rect = window.rectangle()
    abs_x = rect.left + x
    abs_y = rect.top + y
    pyautogui.moveTo(abs_x, abs_y, duration=duration)
    return abs_x, abs_y


def move_mouse_to_x1_x2(window, duration=0.1, gap_seconds=0.05):
    """
    依序把滑鼠移到 (x1, y1) 與 (x2, y2)。
    x1/y1/x2/y2 需是以 Chrome 視窗左上角為 (0,0) 的相對座標。
    """
    abs1 = move_mouse_to_relative_position(window, x1, y1, duration=duration)
    time.sleep(gap_seconds)
    abs2 = move_mouse_to_relative_position(window, x2, y2, duration=duration)
    return abs1, abs2


def show_mouse_position():
    print("=== 即時滑鼠座標顯示 ===")
    print("移動滑鼠看座標")
    print("- 按 f1：印出當下座標")
    print("- 按 f2：退出")
    print("- 按 f3：把滑鼠移到 (x1, y1)")

    try:
        prev_f1 = False
        prev_f2 = False
        prev_f3 = False
        while True:
            pos = pyautogui.position()
            chrome_window = get_chrome_window()
            print(f"\r滑鼠位置: ({pos.x}, {pos.y})", end="", flush=True)

            f1_down = keyboard.is_pressed("f1")
            f2_down = keyboard.is_pressed("f2")
            f3_down = keyboard.is_pressed("f3")

            # key-down edge detection (避免按住鍵時重複觸發)
            if f1_down and not prev_f1:
                print(f"\n[f1] 目前座標: ({pos.x}, {pos.y})")

            if f3_down and not prev_f3:
                if not chrome_window:
                    print("\n[f3] 找不到 Chrome 視窗，無法移動滑鼠")
                else:
                    abs_x, abs_y = move_mouse_to_relative_position(chrome_window, x1, y1)
                    print(f"\n[f3] 已移動滑鼠到 (x1, y1): ({abs_x}, {abs_y})")

            if f2_down and not prev_f2:
                break

            prev_f1 = f1_down
            prev_f2 = f2_down
            prev_f3 = f3_down
            time.sleep(0.1)

    except KeyboardInterrupt:
        pass

    print("\n* 座標顯示已退出")


show_mouse_position()


# %% 背景連點

paused = False
stop_program = False
sleep_time_click = 0.2


def get_chrome_window():
    try:
        app = Application(backend="uia").connect(title_re=CHROME_TITLE_RE)
        return app.top_window()
    except Exception as e:
        print(f"找不到 Chrome 視窗：{e}")
        return None


def toggle_pause():
    paused = not paused
    print(f"程式已{'暫停' if paused else '繼續'}")


def stop_program_func():
    global stop_program
    stop_program = True
    print("程式已停止")


def click_at_relative_position(window, x, y):
    try:
        hwnd = window.handle
        lparam = win32api.MAKELONG(x, y)
        win32gui.PostMessage(hwnd, win32con.WM_LBUTTONDOWN, win32con.MK_LBUTTON, lparam)
        time.sleep(sleep_time_click)
        win32gui.PostMessage(hwnd, win32con.WM_LBUTTONUP, 0, lparam)
    except Exception as e:
        raise RuntimeError(f"點擊失敗: {e}") from e


x1, y1 = 954, 929
x2, y2 = 1128, 751
x3, y3 = 1065, 664
x4, y4 = 998, 897
times = 10**4 * 3  # 點擊次數

chrome_window = get_chrome_window()
if chrome_window is None:
    print("請確認 Chrome 已開啟且標題可被 CHROME_TITLE_RE 匹配")
    raise SystemExit(1)

print(f"已連到 Chrome 視窗: {chrome_window.window_text()}")
print("開始背景點擊：(按 f1 暫停/繼續；按 f2 停止)")

keyboard.add_hotkey("f1", toggle_pause)
keyboard.add_hotkey("f2", stop_program_func)
for it in range(times):
    if stop_program:
        break

    while paused:
        time.sleep(sleep_time_click)
        if stop_program:
            break

    print(f"點擊第 {it + 1}/{times} 次...")
    click_at_relative_position(chrome_window, x1, y1)
    # click_at_relative_position(chrome_window, x2, y2)
    # click_at_relative_position(chrome_window, x3, y3)
    # click_at_relative_position(chrome_window, x4, y4)
    time.sleep(sleep_time_click)


# %% 解析 Spin 回應並輸出 Excel/JSON（保留原本 Wild 的解析邏輯，但使用方式比照 Gates）


def decode_response_body(raw_body: bytes):
    if not raw_body:
        return "", "empty"

    try:
        if len(raw_body) >= 2 and raw_body[:2] == b"\x1f\x8b":
            return gzip.decompress(raw_body).decode("utf-8", errors="replace"), "gzip"
        if len(raw_body) >= 2 and raw_body[0:1] == b"\x78":
            return zlib.decompress(raw_body).decode("utf-8", errors="replace"), "deflate"
        return raw_body.decode("utf-8", errors="replace"), "utf8"
    except Exception:
        return raw_body.decode("utf-8", errors="replace"), "utf8_with_errors"


def extract_dt_si_fields(json_data):
    extracted = {}

    if isinstance(json_data, dict) and "dt" in json_data and json_data["dt"] and "si" in json_data["dt"]:
        si_data = json_data["dt"]["si"]
        for key, value in si_data.items():
            if isinstance(value, (dict, list)):
                extracted[f"si_{key}"] = json.dumps(value, ensure_ascii=False)
            else:
                extracted[f"si_{key}"] = value

    if not extracted:
        if isinstance(json_data, dict):
            for key in ["request_number", "request_url", "status_code", "error", "decode_method"]:
                if key in json_data:
                    extracted[key] = json_data[key]
            if not extracted:
                extracted["data_type"] = "unknown_structure"
                extracted["raw_data"] = json.dumps(json_data, ensure_ascii=False)[:200]
        else:
            extracted["data_type"] = "non_dict_data"
            extracted["raw_data"] = str(json_data)[:200]

    return extracted


def convert_to_numeric(df):
    for column in df.columns:
        try:
            df[column] = pd.to_numeric(df[column])
        except ValueError:
            pass
    return df


spin_data_list = []
spin_request_count = 0

for request in driver.requests:
    if not request.response:
        continue
    if "Spin" not in request.url:
        continue

    spin_request_count += 1
    try:
        raw_body = request.response.body or b""
        decoded_text, decode_method = decode_response_body(raw_body)

        parsed = None
        try:
            parsed = json.loads(decoded_text)
        except json.JSONDecodeError:
            parsed_qs = urllib.parse.parse_qs(decoded_text)
            if parsed_qs:
                parsed = {k: v[0] if v else "" for k, v in parsed_qs.items()}

        if parsed is None:
            parsed = {
                "request_number": spin_request_count,
                "request_url": request.url,
                "status_code": request.response.status_code,
                "decode_method": decode_method,
                "raw_data_length": len(raw_body),
                "raw_text_sample": decoded_text[:500],
            }

        spin_data_list.append(parsed)

    except Exception as e:
        spin_data_list.append(
            {
                "request_number": spin_request_count,
                "request_url": request.url,
                "status_code": request.response.status_code if request.response else None,
                "error": str(e),
            }
        )

print(f"共找到 {spin_request_count} 筆 Spin 回應")


file_excel = "spin_responses.xlsx"
file_json = "spin_responses.json"
new_sheet_base = "spin_data"

if spin_data_list:
    extracted_list = [extract_dt_si_fields(item) for item in spin_data_list]
    df = pd.DataFrame(extracted_list)
    df = convert_to_numeric(df)
    df = df.iloc[33:, 6:]

    if os.path.exists(file_excel):
        book = load_workbook(file_excel)
        existing_sheets = book.sheetnames

        new_sheet_name = new_sheet_base
        i = 1
        while new_sheet_name in existing_sheets:
            new_sheet_name = f"{new_sheet_base}_{i}"
            i += 1

        with pd.ExcelWriter(file_excel, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            df.to_excel(writer, sheet_name=new_sheet_name, index=False)

        print(f"資料已新增至 {file_excel} 工作表: {new_sheet_name}")
    else:
        with pd.ExcelWriter(file_excel, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=new_sheet_base, index=False)
        print(f"資料已寫入新檔 {file_excel} 工作表: {new_sheet_base}")

    with open(file_json, "w", encoding="utf-8") as json_file:
        json.dump(spin_data_list, json_file, indent=4, ensure_ascii=False)
    print(f"資料已寫入 {file_json}")
else:
    print("沒有抓到 Spin 回應資料")


# 關閉瀏覽器（如需要可取消註解）
# driver.quit()

# %%
