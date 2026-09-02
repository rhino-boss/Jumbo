from pathlib import Path
from openpyxl import load_workbook

root = Path(r"Project/Slots/H028_雷神爆金 1000/Source")
for name in ("H028188B.xlsx", "H028190B.xlsx", "H028192A.xlsx", "H028194A.xlsx"):
    path = root / name
    print(f"\n=== {name} ===")
    wb_f = load_workbook(path, read_only=False, data_only=False)
    wb_v = load_workbook(path, read_only=True, data_only=True)
    print("sheets:", wb_f.sheetnames)
    for sn in ("Overview", "Detail", "Detail_Newbie", "Multiplier_Weight"):
        if sn not in wb_f.sheetnames:
            continue
        ws_f, ws_v = wb_f[sn], wb_v[sn]
        print(f"-- {sn} {ws_f.max_row}x{ws_f.max_column}")
        if sn == "Overview":
            for r in range(1, 10):
                vals = [ws_v.cell(r, c).value for c in range(1, 9)]
                print(r, vals)
        elif sn in ("Detail", "Detail_Newbie"):
            for r in range(1, 12):
                vals = [ws_v.cell(r, c).value for c in range(1, 14)]
                print(r, vals)
            for r in (15, 16, 20, 30, 40, 50, 60, 70, 78, 79, 86, 87, 90, 100, 110, 120, 130, 140, 148, 149):
                vf = [ws_f.cell(r, c).value for c in range(1, 17)]
                vv = [ws_v.cell(r, c).value for c in range(1, 17)]
                print("row", r, "form", vf)
                print("row", r, "data", vv)
        else:
            for r in range(1, min(ws_f.max_row, 72) + 1):
                vals = [ws_v.cell(r, c).value for c in range(1, ws_f.max_column + 1)]
                if any(v is not None for v in vals):
                    print(r, vals)
    wb_f.close()
    wb_v.close()
