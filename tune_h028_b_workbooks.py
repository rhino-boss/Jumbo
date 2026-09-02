from __future__ import annotations

import json
import math
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(r"Project/Slots/H028_雷神爆金 1000")
SOURCE = ROOT / "Source"
BACKUP = ROOT / "Versions" / "3.2.0.0" / "Source_Backup"
sys.path.insert(0, str(SOURCE.resolve()))
import model_sync  # noqa: E402

TOTAL = 1_000_000_000
TARGETS = {
    "H028188B.xlsx": 0.16,
    "H028190B.xlsx": 0.18,
}


def largest_remainder(probabilities: list[float], total: int) -> list[int]:
    raw = [value * total for value in probabilities]
    result = [math.floor(value) for value in raw]
    missing = total - sum(result)
    order = sorted(
        range(len(raw)),
        key=lambda index: (raw[index] - result[index], -index),
        reverse=True,
    )
    for index in order[:missing]:
        result[index] += 1
    return result


def improve_integer_mean(
    weights: list[int], averages: list[float], target_numerator: float
) -> tuple[list[int], float]:
    result = weights[:]
    residual = target_numerator - sum(w * a for w, a in zip(result, averages))
    for _ in range(20):
        best = (abs(residual), None)
        for donor in range(len(result)):
            if result[donor] <= 0:
                continue
            for receiver in range(len(result)):
                delta = averages[receiver] - averages[donor]
                if residual * delta <= 0 or delta == 0:
                    continue
                estimate = residual / delta
                for amount in {
                    1,
                    max(1, min(result[donor], math.floor(estimate))),
                    max(1, min(result[donor], math.ceil(estimate))),
                }:
                    new_residual = residual - amount * delta
                    candidate = (abs(new_residual), (donor, receiver, amount, new_residual))
                    if candidate[0] < best[0]:
                        best = candidate
        if best[1] is None:
            break
        donor, receiver, amount, residual = best[1]
        result[donor] -= amount
        result[receiver] += amount
    return result, residual


def exponential_tilt(
    base_probabilities: list[float], averages: list[float], target_average: float
) -> list[float]:
    active = [i for i, value in enumerate(base_probabilities) if value > 0]
    if not active:
        raise ValueError("No active BG range weights")

    def tilted_mean(lam: float) -> tuple[float, list[float]]:
        logs = [math.log(base_probabilities[i]) + lam * averages[i] for i in active]
        peak = max(logs)
        raw = [math.exp(value - peak) for value in logs]
        scale = sum(raw)
        probs = [0.0] * len(base_probabilities)
        for i, value in zip(active, raw):
            probs[i] = value / scale
        mean = sum(probs[i] * averages[i] for i in active)
        return mean, probs

    low, high = -1.0, 1.0
    low_mean, _ = tilted_mean(low)
    high_mean, _ = tilted_mean(high)
    while target_average < low_mean:
        high = low
        low *= 2
        low_mean, _ = tilted_mean(low)
    while target_average > high_mean:
        low = high
        high *= 2
        high_mean, _ = tilted_mean(high)
    probabilities = base_probabilities
    for _ in range(120):
        middle = (low + high) / 2
        mean, probabilities = tilted_mean(middle)
        if mean < target_average:
            low = middle
        else:
            high = middle
    return probabilities


def workbook_averages(path: Path) -> tuple[list[float], float, list[float], list[str], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    detail = workbook["Detail"]
    bg_averages = [float(detail.cell(row, 5).value or 0) for row in range(15, 79)]
    trigger_average = float(detail["E79"].value)
    fg_averages = [float(detail.cell(row, 5).value or 0) for row in range(86, 150)]
    bg_labels = [str(detail.cell(row, 1).value) for row in range(15, 79)]
    fg_labels = [str(detail.cell(row, 1).value) for row in range(86, 150)]
    workbook.close()
    return bg_averages, trigger_average, fg_averages, bg_labels, fg_labels


def tune(path: Path, target_fg_rtp: float) -> None:
    config = model_sync.build_config(path)
    before = json.loads(json.dumps(config))
    oldhand = config["card_system"]["oldhand"]["normal_bet"]
    bg_cards = oldhand["weight_bg"]
    fg_cards = oldhand["weight_fg"]
    bg_averages, trigger_average, fg_averages, bg_labels, fg_labels = workbook_averages(path)

    range_cards = [card for card in bg_cards if card["type"] == "range"]
    trigger_card = next(card for card in bg_cards if card["type"] == "free_game")
    fg_range_cards = [card for card in fg_cards if card["type"] == "range"]
    if len(range_cards) != len(bg_averages) or len(fg_range_cards) != len(fg_averages):
        raise ValueError(
            f"{path.name}: card rows do not match Detail rows "
            f"(BG {len(range_cards)}/{len(bg_averages)}, FG {len(fg_range_cards)}/{len(fg_averages)})"
        )
    if [card_range(card) for card in range_cards] != bg_labels:
        raise ValueError(f"{path.name}: BG interval labels do not align")
    if [card_range(card) for card in fg_range_cards] != fg_labels:
        raise ValueError(f"{path.name}: FG interval labels do not align")

    fg_average = sum(
        int(card["weight"]) * average for card, average in zip(fg_range_cards, fg_averages)
    ) / TOTAL
    trigger_weight = round(target_fg_rtp / fg_average * TOTAL)
    regular_total = TOTAL - trigger_weight
    current_regular = [int(card["weight"]) for card in range_cards]
    current_regular_total = sum(current_regular)
    base_probabilities = [weight / current_regular_total for weight in current_regular]
    target_regular_average = (
        0.72 * TOTAL - trigger_weight * trigger_average
    ) / regular_total
    tilted = exponential_tilt(base_probabilities, bg_averages, target_regular_average)
    regular_weights = largest_remainder(tilted, regular_total)
    regular_weights, bg_residual_num = improve_integer_mean(
        regular_weights,
        bg_averages,
        0.72 * TOTAL - trigger_weight * trigger_average,
    )

    for card, weight in zip(range_cards, regular_weights):
        card["weight"] = weight
    trigger_card["weight"] = trigger_weight

    before_json = BACKUP / f"{path.stem}_before_weights.json"
    after_json = BACKUP / f"{path.stem}_after_weights.json"
    tuned_js = BACKUP / f"{path.stem}_tuned_{int(72)}-{int(round(target_fg_rtp * 100))}_config.js"
    before_json.write_text(json.dumps(before, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    after_json.write_text(json.dumps(config, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    model_sync.write_js_config(tuned_js, config)

    model_sync.run_import([
        "--config", str(tuned_js),
        "--source", str(path),
        "--in-place",
    ])
    verified = model_sync.build_config(path)
    if verified != config:
        raise ValueError(f"{path.name}: post-write config mismatch")

    final_bg_rtp = (
        sum(weight * average for weight, average in zip(regular_weights, bg_averages))
        + trigger_weight * trigger_average
    ) / TOTAL
    final_fg_rtp = trigger_weight / TOTAL * fg_average
    cycle = TOTAL / trigger_weight
    l1 = sum(abs(new / regular_total - old / current_regular_total)
             for new, old in zip(regular_weights, current_regular))
    max_change = max(
        abs(new / regular_total - old / current_regular_total)
        for new, old in zip(regular_weights, current_regular)
    )
    print(json.dumps({
        "workbook": path.name,
        "target": f"72:{target_fg_rtp * 100:.0f}",
        "fg_average_multiplier": fg_average,
        "fg_trigger_weight": trigger_weight,
        "fg_cycle": cycle,
        "expected_bg_rtp": final_bg_rtp,
        "expected_fg_rtp": final_fg_rtp,
        "expected_total_rtp": final_bg_rtp + final_fg_rtp,
        "bg_numerator_residual": bg_residual_num,
        "bg_conditional_l1_change": l1,
        "bg_conditional_max_bucket_change": max_change,
        "weight_total": sum(regular_weights) + trigger_weight,
    }, ensure_ascii=False, indent=2))


def card_range(card: dict) -> str:
    low = int(card["min"]) if float(card["min"]).is_integer() else card["min"]
    high = int(card["max"]) if float(card["max"]).is_integer() else card["max"]
    return f"({low}, {high}]"


if __name__ == "__main__":
    BACKUP.mkdir(parents=True, exist_ok=True)
    for workbook_name, target in TARGETS.items():
        tune(SOURCE / workbook_name, target)
