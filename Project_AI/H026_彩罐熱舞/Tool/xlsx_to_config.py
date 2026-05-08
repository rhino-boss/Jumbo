from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = PROJECT_DIR / "Source" / "H026192.xlsx"
DEFAULT_OUTPUT = PROJECT_DIR / "config.js"
STATIC_DEFAULTS: dict[str, Any] = {
    "source_box": "Project.Slots.Source.H026_Box",
    "game_name": "彩罐熱舞",
    "display_name": "彩罐熱舞",
    "bet_options": [1, 2, 3, 5, 10],
    "mode_normalbet": 0,
    "mode_featurebuy": 2,
    "scene_bg": 0,
    "scene_fg": 1,
    "scene_bf": 2,
    "special_pool_weight_base": 10000,
    "fg_table_rule": {
        "type": "multiplier_threshold",
        "thresholds": [10, 20],
        "table_ids": [3, 4, 5],
        "table_names": ["F1", "F2", "F3"],
        "rules": [
            {"min_multiplier": 0, "max_exclusive": 10, "table_id": 3, "table_name": "F1"},
            {"min_multiplier": 10, "max_exclusive": 20, "table_id": 4, "table_name": "F2"},
            {"min_multiplier": 20, "max_exclusive": None, "table_id": 5, "table_name": "F3"},
        ],
    },
    "frame_bg": "Source/Image/pinata-wins_symbol_s_frame_bg.png",
    "frame_top": "Source/Image/pinata-wins_symbol_s_frame.png",
    "asset_map": {
        "WW": "Source/Image/pinata-wins_symbol_s_wild.png",
        "C1": "Source/Image/pinata-wins_symbol_s_scatter.png",
        "M1": "Source/Image/pinata-wins_symbol_m1_skull.png",
        "M2": "Source/Image/pinata-wins_symbol_m2_sombrero.png",
        "M3": "Source/Image/pinata-wins_symbol_m3_maracas.png",
        "M4": "Source/Image/pinata-wins_symbol_m4_taco.png",
        "M5": "Source/Image/pinata-wins_symbol_m5_chilli.png",
        "A": "Source/Image/pinata-wins_symbol_l4_a.png",
        "K": "Source/Image/pinata-wins_symbol_l3_k.png",
        "Q": "Source/Image/pinata-wins_symbol_l2_q.png",
        "J": "Source/Image/pinata-wins_symbol_l1_j.png",
    },
}
STRIP_SHEETS = [
    "BG_Symbol",
    "BG_Symbol (2)",
    "BG_Symbol (3)",
    "FG_Symbol",
    "FG_Symbol (2)",
    "FG_Symbol (3)",
    "BF_Symbol",
]
STRIP_NAME_MAP = ["B1", "B2", "B3", "F1", "F2", "F3", "BF"]
MULTIPLIER_SECTION_ROWS = {
    "weight_multiple_special": (5, 11),
    "weight_multiple_r3_before": (16, 22),
    "weight_multiple_r3_after": (27, 33),
    "weight_multiple_before": (38, 44),
    "weight_multiple_after": (49, 55),
}


def load_template(path: Path) -> dict[str, Any]:
    if not path.exists():
        return dict(STATIC_DEFAULTS)
    raw = path.read_text(encoding="utf-8").strip()
    if raw.startswith("window.") and "=" in raw:
        raw = raw.split("=", 1)[1].strip()
    if raw.endswith(";"):
        raw = raw[:-1].strip()
    data = json.loads(raw)
    merged = dict(STATIC_DEFAULTS)
    merged.update(data)
    return merged


def to_int(value: Any) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, str):
        value = value.strip()
    return int(value)


def cumulative(values: list[int]) -> list[int]:
    total = 0
    result: list[int] = []
    for value in values:
        total += int(value)
        result.append(total)
    return result


def cumulative_matrix(matrix: list[list[int]]) -> list[list[int]]:
    if not matrix:
        return []
    cols = len(matrix[0])
    running = [0] * cols
    result: list[list[int]] = []
    for row in matrix:
        out_row: list[int] = []
        for col in range(cols):
            running[col] += int(row[col])
            out_row.append(running[col])
        result.append(out_row)
    return result


def find_row(ws: Any, column: int, value: str) -> int:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, column).value == value:
            return row
    raise ValueError(f"Could not find {value!r} in sheet {ws.title}")


def parse_overview(ws: Any) -> dict[str, Any]:
    pay_header_row = find_row(ws, 1, "Pay Table：")
    pay_row = pay_header_row + 1
    pay_table: list[list[int]] = []
    symbol_id: list[int] = []
    symbol_str: dict[str, str] = {}
    row = pay_row + 1
    while ws.cell(row, 1).value:
        symbol = str(ws.cell(row, 1).value).strip()
        symbol_id_value = to_int(ws.cell(row, 6).value)
        symbol_id.append(symbol_id_value)
        symbol_str[str(symbol_id_value)] = symbol
        pay_table.append([to_int(ws.cell(row, 3).value), to_int(ws.cell(row, 4).value), to_int(ws.cell(row, 5).value)])
        row += 1

    return {
        "game_id": str(ws["B2"].value).strip(),
        "default_coin_in": to_int(ws["A7"].value),
        "normalbet": to_int(ws["B11"].value),
        "featurebuy": to_int(ws["B12"].value),
        "reel_num": 5,
        "window_size": to_int(ws["B26"].value),
        "pay_table": pay_table,
        "symbol_id": symbol_id,
        "symbol_str": symbol_str,
    }


def parse_paylines(ws: Any) -> list[list[int]]:
    paylines: list[list[int]] = []
    row = 5
    while ws.cell(row, 2).value is not None:
        paylines.append([to_int(ws.cell(row, col).value) for col in range(3, 8)])
        row += 1
    return paylines


def parse_table_weights(ws: Any, start_row: int) -> list[int]:
    weights: list[int] = []
    row = start_row
    while ws.cell(row, 9).value:
        weights.append(to_int(ws.cell(row, 10).value))
        row += 1
    return weights


def parse_multiplier_range(ws: Any) -> list[int]:
    values: list[int] = []
    row = find_row(ws, 9, "Multiplier") + 1
    while ws.cell(row, 9).value is not None:
        values.append(to_int(ws.cell(row, 9).value))
        row += 1
    return values


def parse_multiplier_section(ws: Any, start_row: int, end_row: int) -> list[list[int]]:
    rows: list[list[int]] = []
    for row in range(start_row, end_row + 1):
        rows.append([to_int(ws.cell(row, col).value) for col in range(15, 28)])
    if not rows:
        return []
    return [list(column) for column in zip(*rows)]


def parse_special_pool_weight(ws: Any) -> list[list[int]]:
    rows: list[list[int]] = []
    for row in range(5, 12):
        rows.append([to_int(ws.cell(row, col).value) for col in range(30, 38)])
    if not rows:
        return []
    return [list(column) for column in zip(*rows)]


def parse_eliminate_table_weights(ws: Any) -> tuple[list[int], list[int], list[int]]:
    label_row = find_row(ws, 9, "Eliminate Table Weight")
    bg = [to_int(ws.cell(label_row + 2, 10).value), to_int(ws.cell(label_row + 2, 11).value)]
    fg = [to_int(ws.cell(label_row + 3, 10).value), to_int(ws.cell(label_row + 3, 11).value)]
    bf = list(fg)
    return bg, fg, bf


def parse_strip_sheet(ws: Any) -> dict[str, Any]:
    arr_reels: list[list[int]] = []
    arr_reels_weight: list[list[int]] = []
    row = 4
    while ws.cell(row, 10).value is not None:
        arr_reels.append([to_int(ws.cell(row, col).value) for col in range(17, 22)])
        arr_reels_weight.append([to_int(ws.cell(row, col).value) for col in range(23, 28)])
        row += 1

    drop_weight_a = [[to_int(ws.cell(row, col).value) for col in range(30, 35)] for row in range(30, 50)]
    drop_weight_b = [[to_int(ws.cell(row, col).value) for col in range(37, 42)] for row in range(30, 50)]
    return {
        "arr_reels": arr_reels,
        "arr_reels_weight": arr_reels_weight,
        "drop_weight_a": drop_weight_a,
        "drop_weight_b": drop_weight_b,
    }


def build_symbol_flags(symbol_codes: list[str]) -> tuple[list[int], list[int], list[int], list[int]]:
    code_to_id = {code: idx for idx, code in enumerate(symbol_codes)}
    base_symbol_of: list[int] = []
    is_gold_symbol: list[int] = []
    is_score_symbol: list[int] = []
    symbols_score: list[int] = []

    for idx, code in enumerate(symbol_codes):
        if code.startswith("G"):
            base_code = code[1:]
            if base_code == "":
                base_code = code
            if base_code == "1":
                base_code = "M1"
            elif base_code == "2":
                base_code = "M2"
            elif base_code == "3":
                base_code = "M3"
            elif base_code == "4":
                base_code = "M4"
            elif base_code == "5":
                base_code = "M5"
            base_symbol_of.append(code_to_id[base_code])
            is_gold_symbol.append(1)
            is_score_symbol.append(0)
            continue

        base_symbol_of.append(idx)
        is_gold_symbol.append(0)
        score_flag = 1 if code not in {"WW", "C1"} else 0
        is_score_symbol.append(score_flag)
        if score_flag:
            symbols_score.append(idx)

    return base_symbol_of, is_gold_symbol, is_score_symbol, symbols_score


def generate_config(xlsx_path: Path, template: dict[str, Any]) -> dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    overview = parse_overview(wb["Overview"])
    parameter = wb["Parameter"]

    weight_table_bg = parse_table_weights(parameter, 5)
    weight_table_fg = parse_table_weights(parameter, 12)
    weight_table_bf = parse_table_weights(parameter, 19)
    value_multiplier_range = parse_multiplier_range(parameter)
    eliminate_bg, eliminate_fg, eliminate_bf = parse_eliminate_table_weights(parameter)

    strip_data = [parse_strip_sheet(wb[sheet_name]) for sheet_name in STRIP_SHEETS]
    arr_reels = [item["arr_reels"] for item in strip_data]
    arr_reels_weight = [item["arr_reels_weight"] for item in strip_data]
    drop_weight_a = [item["drop_weight_a"] for item in strip_data]
    drop_weight_b = [item["drop_weight_b"] for item in strip_data]
    reels_len = [[len(table) for _ in range(overview["reel_num"])] for table in arr_reels]

    symbol_codes = [overview["symbol_str"][str(symbol_id)] for symbol_id in overview["symbol_id"]]
    base_symbol_of, is_gold_symbol, is_score_symbol, symbols_score = build_symbol_flags(symbol_codes)

    config: dict[str, Any] = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_box": template["source_box"],
        "game_id": overview["game_id"],
        "game_name": template["game_name"],
        "display_name": template["display_name"],
        "bet_options": template["bet_options"],
        "mode_normalbet": template["mode_normalbet"],
        "mode_featurebuy": template["mode_featurebuy"],
        "scene_bg": template["scene_bg"],
        "scene_fg": template["scene_fg"],
        "scene_bf": template["scene_bf"],
        "reel_num": overview["reel_num"],
        "window_size": overview["window_size"],
        "default_coin_in": overview["default_coin_in"],
        "normalbet": overview["normalbet"],
        "featurebuy": overview["featurebuy"],
        "special_pool_weight_base": template["special_pool_weight_base"],
        "paylines": parse_paylines(parameter),
        "pay_table": overview["pay_table"],
        "symbol_id": overview["symbol_id"],
        "symbol_str": overview["symbol_str"],
        "base_symbol_of": base_symbol_of,
        "is_gold_symbol": is_gold_symbol,
        "is_score_symbol": is_score_symbol,
        "symbols_score": symbols_score,
        "value_multiplier_range": value_multiplier_range,
        "weight_table_bg": weight_table_bg,
        "weight_table_fg": weight_table_fg,
        "weight_table_bf": weight_table_bf,
        "weight_cum_table_bg": cumulative(weight_table_bg),
        "weight_cum_table_fg": cumulative(weight_table_fg),
        "weight_cum_table_bf": cumulative(weight_table_bf),
        "fg_table_rule": template["fg_table_rule"],
        "weight_special_pool": parse_special_pool_weight(parameter),
        "arr_reels": arr_reels,
        "arr_reels_weight": arr_reels_weight,
        "arr_reels_weight_cum": [cumulative_matrix(table) for table in arr_reels_weight],
        "drop_weight_a": drop_weight_a,
        "drop_weight_b": drop_weight_b,
        "drop_weight_a_cum": [cumulative_matrix(table) for table in drop_weight_a],
        "drop_weight_b_cum": [cumulative_matrix(table) for table in drop_weight_b],
        "reels_len": reels_len,
        "strip_name_map": STRIP_NAME_MAP,
        "frame_bg": template["frame_bg"],
        "frame_top": template["frame_top"],
        "asset_map": template["asset_map"],
        "eliminate_table_weight_bg": eliminate_bg,
        "eliminate_table_weight_fg": eliminate_fg,
        "eliminate_table_weight_bf": eliminate_bf,
        "eliminate_table_weight_cum_bg": cumulative(eliminate_bg),
        "eliminate_table_weight_cum_fg": cumulative(eliminate_fg),
        "eliminate_table_weight_cum_bf": cumulative(eliminate_bf),
    }

    for key, rows in MULTIPLIER_SECTION_ROWS.items():
        matrix = parse_multiplier_section(parameter, rows[0], rows[1])
        config[key] = matrix
        config[key.replace("weight_", "weight_cum_")] = cumulative_matrix(matrix)

    return config


def write_config(path: Path, data: dict[str, Any]) -> None:
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    path.write_text(f"window.H026_BOX_DATA = {payload};\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate H026 config.js from the new xlsx layout")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--template", default=str(DEFAULT_OUTPUT))
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_path = Path(args.xlsx).resolve()
    output_path = Path(args.output).resolve()
    template = load_template(Path(args.template).resolve())
    config = generate_config(xlsx_path, template)
    write_config(output_path, config)
    print(f"Generated {output_path.name} from {xlsx_path.name}")


if __name__ == "__main__":
    main()
