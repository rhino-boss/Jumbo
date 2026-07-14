import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
PROJECT_DIR = BASE_DIR.parent
DEFAULT_SOURCE = BASE_DIR / "H027192A.xlsx"

METADATA = {
    "game_id": "101027",
    "parsheet_id": "H0271",
    "display_name": "Olympus 2500",
    "game_name_zh": "奧林帕斯 2500",
    "mode_normalbet": 0,
    "mode_extrabet": 1,
    "mode_featurebuy": 2,
    "supported_bet_modes": [0, 1, 2],
    "extra_bet_multiplier": 2,
    "extra_fg_probability_multiplier": 5,
    "has_super_feature": False,
    "has_wild": False,
    "has_jackpot": False,
    "c2_multiplier_levels": [2, 3, 4, 5, 8, 10, 12, 15, 20, 25, 50, 100, 500, 1000, 2500],
    "max_free_spins": 50,
    "reference_presentation": "文件/260630_Olympus 2500.pptx",
    "rule_document": "game_rule_H027.md",
    "model_status": "rules_confirmed_math_draft",
    "pending_math_items": [
        "Extra Bet dedicated reel and card weights",
        "WW-free recalibrated strips",
        "Initial 2500x C2 weight",
        "Max Win cap",
    ],
}

STRIP_SHEETS = [
    "BG_Symbol",
    "BG_Symbol (2)",
    "BG_Symbol (3)",
    "BG_Symbol (4)",
    "FG_Symbol",
    "FG_Symbol (2)",
    "FG_Symbol (3)",
    "FG_Symbol (4)",
    "BF_Symbol",
]


def to_int(value, default=0):
    if value is None or value == "":
        return default
    return int(value)


def cumulative(values):
    total = 0
    result = []
    for value in values:
        total += int(value)
        result.append(total)
    return result


def find_row(ws, column, value):
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, column).value).strip() == value:
            return row
    raise ValueError(f"Could not find {value!r} in {ws.title}")


def find_bet_row(ws, label):
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row, 4).value).strip() == label:
            return row
    raise ValueError(f"Could not find bet type {label!r} in {ws.title}")


def parse_overview(ws):
    pay_header = find_row(ws, 1, "Pay Table：")
    symbol_codes = []
    symbol_ids = []
    pay_table = []
    row = pay_header + 2
    while ws.cell(row, 1).value is not None:
        symbol_codes.append(str(ws.cell(row, 1).value).strip())
        symbol_ids.append(to_int(ws.cell(row, 9).value))
        pay_table.append([to_int(ws.cell(row, col).value) for col in range(3, 9)])
        row += 1

    visible_row = find_row(ws, 1, "Visible Window Size")
    bet_rows = {
        "normal_newbie": find_bet_row(ws, "Normal Bet Newbie"),
        "normal": find_bet_row(ws, "Normal Bet Oldhand"),
        "featurebuy": find_bet_row(ws, "Buy Feature"),
    }
    return {
        "model": str(ws["B2"].value).strip(),
        "excel_version": str(ws["B3"].value).strip(),
        "default_coin_in": to_int(ws["A7"].value),
        "reel_num": 6,
        "window_size": to_int(ws.cell(visible_row, 2).value),
        "symbol_codes": symbol_codes,
        "symbol_ids": symbol_ids,
        "pay_table": pay_table,
        "pay_count_bounds": [8, 10, 12],
        "scatter_pay_counts": [4, 5, 6],
        "normalbet": to_int(ws.cell(bet_rows["normal"], 2).value),
        "extrabet": 2,
        "featurebuy": to_int(ws.cell(bet_rows["featurebuy"], 2).value),
        "rtp_targets": {
            "normal_newbie": float(ws.cell(bet_rows["normal_newbie"], 3).value),
            "normal_oldhand": float(ws.cell(bet_rows["normal"], 3).value),
            "normal": float(ws.cell(bet_rows["normal"], 3).value),
            "extrabet": None,
            "featurebuy": float(ws.cell(bet_rows["featurebuy"], 3).value),
        },
    }


def parse_named_weights(ws, label_col, weight_col, start_row):
    names = []
    weights = []
    row = start_row
    while ws.cell(row, label_col).value is not None:
        names.append(str(ws.cell(row, label_col).value).strip())
        weights.append(to_int(ws.cell(row, weight_col).value))
        row += 1
    return names, weights


def parse_free_table(ws, name_col, free_col, retrigger_col):
    names = []
    free = []
    retrigger = []
    row = 14
    while ws.cell(row, name_col).value is not None:
        names.append(str(ws.cell(row, name_col).value).strip())
        free.append(to_int(ws.cell(row, free_col).value))
        retrigger.append(to_int(ws.cell(row, retrigger_col).value))
        row += 1
    return {"names": names, "initial": free, "retrigger": retrigger}


def parse_c2_block(ws, multiplier_col, first_weight_col):
    multipliers = []
    weights = [[] for _ in range(7)]
    row = 28
    while ws.cell(row, multiplier_col).value is not None:
        multipliers.append(to_int(ws.cell(row, multiplier_col).value))
        for index in range(7):
            weights[index].append(to_int(ws.cell(row, first_weight_col + index).value))
        row += 1
    names = ["base_direct", "base_wild", "free_direct", "free_wild", "super", "ultimate", "bad"]
    return {
        "multipliers": multipliers,
        "weights": {name: values for name, values in zip(names, weights)},
        "weights_cum": {name: cumulative(values) for name, values in zip(names, weights)},
    }


def parse_parameter(ws):
    normal_names, normal_bg_weights = parse_named_weights(ws, 2, 3, 6)
    buy_names, buy_bg_weights = parse_named_weights(ws, 12, 13, 6)
    return {
        "normal": {
            "base_reel_names": normal_names,
            "base_reel_weights": normal_bg_weights,
            "base_reel_weights_cum": cumulative(normal_bg_weights),
            "free_table": parse_free_table(ws, 2, 3, 4),
            "c2_mode_weights": {
                "base": [to_int(ws.cell(22, col).value) for col in range(3, 6)],
                "free": [to_int(ws.cell(23, col).value) for col in range(3, 6)],
            },
            "c2": parse_c2_block(ws, 2, 3),
        },
        "featurebuy": {
            "base_reel_names": buy_names,
            "base_reel_weights": buy_bg_weights,
            "base_reel_weights_cum": cumulative(buy_bg_weights),
            "free_table": parse_free_table(ws, 12, 13, 14),
            "c2_mode_weights": {
                "base": [to_int(ws.cell(22, col).value) for col in range(13, 16)],
                "free": [to_int(ws.cell(23, col).value) for col in range(13, 16)],
            },
            "c2": parse_c2_block(ws, 12, 13),
        },
    }


def apply_h027_c2_levels(parameter):
    levels = list(METADATA["c2_multiplier_levels"])
    active_weight_names = ("base_direct", "free_direct", "super", "ultimate", "bad")
    for profile in parameter.values():
        c2 = profile["c2"]
        original_index = {int(value): index for index, value in enumerate(c2["multipliers"])}
        filtered_weights = {}
        for name in active_weight_names:
            values = c2["weights"][name]
            filtered_weights[name] = [int(values[original_index[level]]) if level in original_index else 0 for level in levels]
        c2["multipliers"] = levels
        c2["weights"] = filtered_weights
        c2["weights_cum"] = {name: cumulative(values) for name, values in filtered_weights.items()}
    return parameter


def parse_card_range(label):
    if label is None:
        return None
    match = re.fullmatch(r"\(\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*\]", str(label).strip())
    if not match:
        return None
    return float(match.group(1)), float(match.group(2))


def parse_card_system(ws):
    profile_map = {
        3: ("newbie", "normal_bet", "weight_bg"),
        4: ("newbie", "normal_bet", "weight_fg"),
        5: ("oldhand", "normal_bet", "weight_bg"),
        6: ("oldhand", "normal_bet", "weight_fg"),
        7: ("oldhand", "buy_feature", "weight_fg"),
    }
    profiles = {
        "newbie": {"normal_bet": {"weight_bg": [], "weight_fg": []}},
        "oldhand": {
            "normal_bet": {"weight_bg": [], "weight_fg": []},
            "buy_feature": {"weight_fg": []},
        },
    }
    row = 4
    while ws.cell(row, 2).value is not None:
        label = str(ws.cell(row, 2).value).strip()
        range_pair = parse_card_range(label)
        for col, (player, mode, segment) in profile_map.items():
            weight = to_int(ws.cell(row, col).value)
            card = None
            if label.lower() == "free game":
                card = {"type": "free_game", "weight": weight}
            elif range_pair is not None:
                card = {"type": "range", "min": range_pair[0], "max": range_pair[1], "weight": weight}
            if card is not None:
                profiles[player][mode][segment].append(card)
        row += 1
    return {"enabled": True, "retry_limit": 5000, **profiles}


def parse_strip_sheet(ws):
    rows = []
    weights = []
    row = 4
    while any(ws.cell(row, col).value is not None for col in range(19, 25)):
        rows.append([to_int(ws.cell(row, col).value, -1) for col in range(19, 25)])
        weights.append([to_int(ws.cell(row, col).value) for col in range(26, 32)])
        row += 1
    reel_lengths = []
    for reel in range(6):
        reel_lengths.append(sum(1 for item in rows if item[reel] != -1))
    return {"symbols": rows, "weights": weights, "reel_lengths": reel_lengths}


def build_config(source_path):
    workbook = load_workbook(source_path, read_only=False, data_only=True)
    overview = parse_overview(workbook["Overview"])
    parameter = apply_h027_c2_levels(parse_parameter(workbook["Parameter"]))
    card_system = parse_card_system(workbook["Multiplier_Weight"])
    strip_data = [parse_strip_sheet(workbook[name]) for name in STRIP_SHEETS]
    workbook.close()

    config = dict(METADATA)
    config.update(overview)
    # H027192A is currently an H019-based draft workbook.  Use the actual
    # source filename as the model name so reports cannot be mistaken for H019.
    config["source_model"] = overview["model"]
    config["model"] = source_path.stem
    config["parameter"] = parameter
    config["card_system"] = card_system
    config["strip_names"] = STRIP_SHEETS
    config["strips"] = strip_data
    return config


def derive_output_path(source_path):
    match = re.fullmatch(r"H0271(?P<rtp>\d{2})(?P<variant>[A-Za-z0-9_-]*)", source_path.stem, re.IGNORECASE)
    if not match:
        raise ValueError(f"Unsupported xlsx name: {source_path.name}; expected a name such as H027192A.xlsx")
    return PROJECT_DIR / f"config_{match.group('rtp')}{match.group('variant')}.js"


def load_js_config(path):
    text = path.read_text(encoding="utf-8").strip()
    start = text.find("{")
    end = text.rfind("}")
    if start < 0 or end <= start:
        raise ValueError(f"Unsupported config format: {path}")
    return json.loads(text[start : end + 1])


def write_js_config(path, data):
    path.write_text("const data = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")


def process_source(source_path, output_path, check=False):
    generated = build_config(source_path)
    if check:
        current = load_js_config(output_path)
        if generated != current:
            changed = [key for key in generated if generated[key] != current.get(key)]
            raise ValueError(f"Config mismatch for {output_path.name}: {changed}")
        print(f"Config is current: {output_path}")
    else:
        write_js_config(output_path, generated)
        if load_js_config(output_path) != generated:
            raise ValueError(f"Write verification failed: {output_path}")
        print(f"Config written: {output_path} <- {source_path.name}")


def main():
    parser = argparse.ArgumentParser(description="Build H027 config files from xlsx")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--all", action="store_true")
    parser.add_argument("--check", action="store_true")
    args = parser.parse_args()

    sources = sorted(path for path in BASE_DIR.glob("H0271*.xlsx") if not path.name.startswith("~$")) if args.all else [args.source.resolve()]
    if not sources:
        raise FileNotFoundError(f"No H0271*.xlsx files found in {BASE_DIR}")

    for source_path in sources:
        output_path = args.output.resolve() if args.output and len(sources) == 1 else derive_output_path(source_path)
        process_source(source_path, output_path, args.check)
    print(f"Processed {len(sources)} xlsx file(s).")


if __name__ == "__main__":
    main()
