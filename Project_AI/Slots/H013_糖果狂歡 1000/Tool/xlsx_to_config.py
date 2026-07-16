from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = PROJECT_DIR / "Source" / "H013197.xlsx"
DEFAULT_OUTPUT = PROJECT_DIR / "config.js"

PARSHEET_ID = "H0131"
GAME_ID = "101001"
GAME_NAME = "糖果狂歡 1000"
ENGLISH_NAME = "Sugar Bonanza 1000"

STRIP_SHEETS = [
    "BG_strip", "BG_strip (2)", "BG_strip (3)", "BG_strip (4)",
    "EB_strip", "EB_strip (2)", "EB_strip (3)",
    "FG_strip", "FG_strip (2)",
    "FB_strip", "FB_strip (2)",
    "SB_strip", "SB_strip (2)",
]
WEIGHT_SHEETS = [name.replace("_strip", "_strip_weight") for name in STRIP_SHEETS]


def to_int(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    return int(round(float(value)))


def cumulative(values: list[int]) -> list[int]:
    result: list[int] = []
    total = 0
    for value in values:
        total += int(value)
        result.append(total)
    return result


def read_numeric_grid(ws: Any) -> list[list[int]]:
    rows: list[list[int]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = list(row[:6])
        if all(value is None for value in values):
            continue
        rows.append([to_int(value, -1) for value in values])
    return rows


def pad_tables(tables: list[list[list[int]]], fill: int) -> tuple[list[list[list[int]]], list[list[int]]]:
    max_rows = max(len(table) for table in tables)
    padded: list[list[list[int]]] = []
    lengths: list[list[int]] = []
    for table in tables:
        reel_lengths = []
        for reel in range(6):
            reel_lengths.append(sum(1 for row in table if row[reel] != -1))
        lengths.append(reel_lengths)
        padded.append(table + [[fill] * 6 for _ in range(max_rows - len(table))])
    return padded, lengths


def parse_paytable(ws: Any) -> dict[str, Any]:
    headers = [to_int(ws.cell(1, col).value) for col in range(2, 8)]
    symbol_str: dict[str, str] = {}
    symbol_id: list[int] = []
    pay_table: list[list[int]] = []
    for row in range(2, ws.max_row + 1):
        code = ws.cell(row, 1).value
        sid = ws.cell(row, 8).value
        if code is None or sid is None:
            continue
        sid_int = to_int(sid)
        while len(pay_table) <= sid_int:
            pay_table.append([0] * len(headers))
        pay_table[sid_int] = [to_int(ws.cell(row, col).value) for col in range(2, 8)]
        symbol_str[str(sid_int)] = str(code).strip()
        symbol_id.append(sid_int)
    return {
        "pay_awards": headers,
        "pay_table": pay_table,
        "symbol_str": symbol_str,
        "symbol_id": sorted(symbol_id),
        "symbols_special": [0, 1, 2],
        "symbols_score": [sid for sid in sorted(symbol_id) if sid >= 3],
    }


def parse_column(ws: Any, name: str) -> list[int]:
    headers = [cell.value for cell in ws[1]]
    if name not in headers:
        raise KeyError(f"{ws.title}: missing column {name}")
    col = headers.index(name) + 1
    return [to_int(ws.cell(row, col).value) for row in range(2, ws.max_row + 1) if ws.cell(row, col).value is not None]


def generate_config(xlsx_path: Path) -> dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=False)
    overview = wb["overview"]
    source_game_id = str(overview["B1"].value or "").strip()
    source_version = str(overview["B2"].value or "").strip()

    strip_tables = [read_numeric_grid(wb[name]) for name in STRIP_SHEETS]
    weight_tables = [read_numeric_grid(wb[name]) for name in WEIGHT_SHEETS]
    arr_reels, reel_lengths = pad_tables(strip_tables, -1)
    arr_weights, _ = pad_tables(weight_tables, 0)

    # Stop weights are cumulative per reel. Padding is zero and therefore does
    # not add selectable stops beyond the weight workbook's actual rows.
    arr_weights_cum: list[list[list[int]]] = []
    for table in arr_weights:
        running = [0] * 6
        cum_table = []
        for row in table:
            running = [running[i] + max(0, row[i]) for i in range(6)]
            cum_table.append(list(running))
        arr_weights_cum.append(cum_table)

    paytable = parse_paytable(wb["pay_table"])
    weight_ws = wb["weight"]
    value_ws = wb["value"]
    config: dict[str, Any] = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_xlsx": "Source/H013197.xlsx",
        "source_game_id": source_game_id,
        "game_id": GAME_ID,
        "parsheet_id": PARSHEET_ID,
        "game_name": GAME_NAME,
        "display_name": GAME_NAME,
        "english_name": ENGLISH_NAME,
        "game_version": source_version or "0004",
        "window_size": 5,
        "reel_num": 6,
        "default_coin_in": 100,
        "normalbet": 1.0,
        "extrabet": 1.25,
        "featurebuy": 75.0,
        "superfeaturebuy": 500.0,
        "mode_normalbet": 0,
        "mode_extrabet": 1,
        "mode_featurebuy": 2,
        "mode_superfeaturebuy": 3,
        "max_spin_free_game": 50,
        "initial_free_spins_low": 8,
        "initial_free_spins_high": 2,
        "retrigger_free_spins_low": 4,
        "retrigger_free_spins_high": 1,
        "strip_name_map": STRIP_SHEETS,
        "arr_reels": arr_reels,
        "arr_reels_weight_cum": arr_weights_cum,
        "reels_len": reel_lengths,
        "value_multiplier": parse_column(value_ws, "multiplier_range"),
        "weight_table_normal_bet": parse_column(weight_ws, "table_normal_bet"),
        "weight_table_extra_bet": parse_column(weight_ws, "table_extra_bet"),
        "weight_multiplier_fg_low": parse_column(weight_ws, "multiplier_range_FG_low"),
        "weight_multiplier_fg_high": parse_column(weight_ws, "multiplier_range_FG_high"),
        "weight_multiplier_fb_low": parse_column(weight_ws, "multiplier_range_FB_low"),
        "weight_multiplier_fb_high": parse_column(weight_ws, "multiplier_range_FB_high"),
        "weight_multiplier_sb_low": parse_column(weight_ws, "multiplier_range_SB_low"),
        "weight_multiplier_sb_high": parse_column(weight_ws, "multiplier_range_SB_high"),
        "official_list_featurebuy": 100.0,
        "source_conflicts": ["Official game list BF=100x; legacy math/simulator BF=75x. Simulator follows legacy math."],
    }
    config.update(paytable)
    return config


def validate_config(config: dict[str, Any]) -> None:
    if len(config["arr_reels"]) != 13 or len(config["arr_reels_weight_cum"]) != 13:
        raise ValueError("expected 13 strip and weight tables")
    for table_idx, table in enumerate(config["arr_reels_weight_cum"]):
        for reel in range(6):
            if table[-1][reel] <= 0:
                raise ValueError(f"table {table_idx}, reel {reel + 1}: no positive stop weight")
    if config["pay_awards"] != [4, 5, 6, 8, 10, 12]:
        raise ValueError(f"unexpected pay awards: {config['pay_awards']}")


def write_config(path: Path, config: dict[str, Any]) -> None:
    payload = json.dumps(config, ensure_ascii=False, indent=2)
    path.write_text(f"window.H013_BOX_DATA = {payload};\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate H013 config.js from H013197.xlsx")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--check", action="store_true", help="validate without writing")
    args = parser.parse_args()
    config = generate_config(Path(args.xlsx).resolve())
    validate_config(config)
    if args.check:
        print("OK: H013 xlsx mapping and generated config are valid")
    else:
        output = Path(args.output).resolve()
        write_config(output, config)
        print(f"Generated {output.name} from {Path(args.xlsx).name}")


if __name__ == "__main__":
    main()
