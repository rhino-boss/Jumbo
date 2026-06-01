from __future__ import annotations

import argparse
import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = PROJECT_DIR / "Source" / "H015192.xlsx"
DEFAULT_OUTPUT = PROJECT_DIR / "config.js"
TABLE_SHEETS = ["BG_Symbol", "BG_Symbol (2)", "FG_Symbol", "FG_Symbol (2)", "BF_Symbol"]
SYMBOL_TO_BASE = {
    "G1": "M1",
    "G2": "M2",
    "G3": "M3",
    "G4": "M4",
    "GA": "A",
    "GK": "K",
    "GQ": "Q",
    "GJ": "J",
}


def to_int(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return default


def cumulative(values: list[int]) -> list[int]:
    total = 0
    result: list[int] = []
    for value in values:
        total += int(value)
        result.append(total)
    return result


def cumulative_by_row(rows: list[list[int]]) -> list[list[int]]:
    if not rows:
        return []
    width = len(rows[0])
    running = [0] * width
    result: list[list[int]] = []
    for row in rows:
        out: list[int] = []
        for idx in range(width):
            running[idx] += int(row[idx])
            out.append(running[idx])
        result.append(out)
    return result


def load_template(path: Path) -> dict[str, Any]:
    defaults: dict[str, Any] = {
        "game_name": "賞金列車",
        "display_name": "賞金列車",
        "english_name": "Wild Bounty Showdown",
        "game_version": "2.0.0.1",
        "bet_options": [1, 2, 3, 4, 5, 6, 8, 10, 12, 16, 20, 30, 40, 60, 100, 200, 300, 600, 1000, 1500],
        "denom": 0.002,
        "mode_normalbet": 0,
        "mode_featurebuy": 2,
        "scene_bg": 0,
        "scene_fg": 1,
        "scene_bf": 2,
        "output_bg": 0,
        "output_fg": 1,
        "output_oa": 2,
        "max_spin_free_game": 50,
    }
    if not path.exists():
        return defaults
    raw = path.read_text(encoding="utf-8").strip()
    if raw.startswith("window.") and "=" in raw:
        raw = raw.split("=", 1)[1].strip()
    if raw.endswith(";"):
        raw = raw[:-1].strip()
    data = json.loads(raw)
    merged = dict(defaults)
    merged.update(data)
    return merged


def build_score_area(layout_visible: list[int], window_size: int) -> list[list[int]]:
    score_area = [[0 for _ in layout_visible] for _ in range(window_size)]
    for reel_idx, visible_count in enumerate(layout_visible):
        for row_idx in range(int(visible_count)):
            score_area[row_idx][reel_idx] = 1
    return score_area


def build_special_area(score_area: list[list[int]]) -> list[list[int]]:
    special_area: list[list[int]] = []
    for row in score_area:
        special_area.append([0 if cell == 1 else 99 for cell in row])
    return special_area


def parse_overview(ws: Any) -> dict[str, Any]:
    symbol_rows: list[dict[str, Any]] = []
    row = 77
    while ws.cell(row, 1).value:
        symbol_rows.append(
            {
                "code": str(ws.cell(row, 1).value).strip(),
                "pay": [to_int(ws.cell(row, col).value) for col in range(3, 7)],
                "id": to_int(ws.cell(row, 7).value),
            }
        )
        row += 1

    scatter_awards: dict[int, int] = {}
    row = 69
    while ws.cell(row, 1).value is not None:
        scatter_count = to_int(ws.cell(row, 1).value)
        free_spins = to_int(ws.cell(row, 2).value)
        if scatter_count > 0 and free_spins > 0:
            scatter_awards[scatter_count] = free_spins
        row += 1

    layout_visible = [to_int(ws.cell(26, col).value) for col in range(2, 8)]
    window_size = max(layout_visible)
    score_area = build_score_area(layout_visible, window_size)

    return {
        "game_id": str(ws["B2"].value).strip(),
        "excel_version": str(ws["B3"].value).strip(),
        "default_coin_in": to_int(ws["A7"].value),
        "normalbet": to_int(ws["B11"].value),
        "featurebuy": to_int(ws["B12"].value),
        "layout_visible": layout_visible,
        "window_size": window_size,
        "reel_num": len(layout_visible),
        "layout_shape": [window_size, len(layout_visible)],
        "score_area": score_area,
        "special_area": build_special_area(score_area),
        "max_spin_free_game": 50,
        "free_spin_awards": scatter_awards,
        "pay_table": [item["pay"] for item in symbol_rows],
        "symbol_id": [item["id"] for item in symbol_rows],
        "symbol_str": {str(item["id"]): item["code"] for item in symbol_rows},
    }


def parse_parameter(ws: Any) -> dict[str, Any]:
    multiplier_range = [to_int(ws.cell(5, col).value) for col in range(3, 14) if ws.cell(5, col).value is not None]

    weight_table_bg = [to_int(ws.cell(row, 3).value) for row in (10, 11)]
    weight_table_fg = [to_int(ws.cell(row, 3).value) for row in (16, 17)]
    weight_table_bf = [to_int(ws.cell(row, 3).value) for row in (22, 23)]

    weight_drop_choose_bg = [to_int(ws.cell(28, col).value) for col in range(3, 6)]
    weight_drop_choose_fg = [to_int(ws.cell(29, col).value) for col in range(3, 6)]

    guaranteed_rows = [to_int(ws.cell(50, col).value) for col in range(3, 14)]
    weight_cum_must_appear_1_fg = cumulative(guaranteed_rows)

    bg_multi = parse_multi_appear(ws, 34)
    fg_multi = parse_multi_appear(ws, 40)

    return {
        "value_multiplier_range": multiplier_range,
        "weight_table_bg": weight_table_bg,
        "weight_table_fg": weight_table_fg,
        "weight_table_bf": weight_table_bf,
        "weight_cum_table_bg": cumulative(weight_table_bg),
        "weight_cum_table_fg": cumulative(weight_table_fg),
        "weight_cum_table_bf": cumulative(weight_table_bf),
        "weight_drop_choose_bg": weight_drop_choose_bg,
        "weight_drop_choose_fg": weight_drop_choose_fg,
        "weight_cum_drop_choose_bg": cumulative(weight_drop_choose_bg),
        "weight_cum_drop_choose_fg": cumulative(weight_drop_choose_fg),
        "weight_cum_must_appear_1_fg": weight_cum_must_appear_1_fg,
        "weight_cum_multi_appear_bg": bg_multi,
        "weight_cum_multi_appear_fg": fg_multi,
    }


def parse_multi_appear(ws: Any, row_start: int) -> list[list[list[int]]]:
    block_starts = [4, 14, 24, 34, 44]
    tables: list[list[list[int]]] = []
    for block_start in block_starts:
        rows: list[list[int]] = []
        for bomb_count_offset in range(7):
            profile_values: list[int] = []
            for profile_idx in range(6):
                row = row_start + profile_idx
                col = block_start + bomb_count_offset
                profile_values.append(to_int(ws.cell(row, col).value))
            rows.append(profile_values)
        tables.append(cumulative_by_row(rows))
    return tables


def parse_table_sheet(ws: Any, symbol_code_to_id: dict[str, int]) -> dict[str, Any]:
    arr_reels: list[list[int]] = []
    arr_reels_weight: list[list[int]] = []
    for row in range(4, 46):
        if ws.cell(row, 12).value == "R1":
            continue
        line_values = [ws.cell(row, col).value for col in range(12, 18)]
        weight_values = [ws.cell(row, col).value for col in range(19, 25)]
        if all(value is None for value in line_values):
            continue
        arr_reels.append([symbol_code_to_id[str(value).strip()] for value in line_values])
        arr_reels_weight.append([to_int(value, 1) for value in weight_values])

    drop_rows: list[list[int]] = []
    for row in range(27, 45):
        weights: list[int] = []
        for reel_idx, col in enumerate(range(3, 9)):
            base_weight = ws.cell(row, col).value
            if base_weight is None:
                weights.append(0)
                continue
            value = float(base_weight)
            scaled = int(round(value * 10000)) if value <= 1 else to_int(value)
            allow_flag = to_int(ws.cell(row, 26 + reel_idx).value, 1)
            weights.append(scaled if allow_flag > 0 else 0)
        drop_rows.append(weights)

    drop_cum = cumulative_by_row(drop_rows)
    combo_map = {
        "combo0": drop_cum,
        "combo1": drop_cum,
        "combo2": drop_cum,
        "combo3_plus": drop_cum,
    }
    return {
        "arr_reels": arr_reels,
        "arr_reels_weight_cum": cumulative_by_row(arr_reels_weight),
        "reels_len": [len(arr_reels) for _ in range(6)],
        "drop_combo": combo_map,
    }


def build_symbol_groups(symbol_id: list[int], symbol_str: dict[str, str]) -> dict[str, Any]:
    ids = [int(item) for item in symbol_id]
    score_ids: list[int] = []
    gold_ids: list[int] = []
    for item in ids:
        code = symbol_str[str(item)]
        if code.startswith("G"):
            gold_ids.append(item)
        elif code not in {"WW", "C1", "C2"}:
            score_ids.append(item)

    return {
        "symbols_special": [0, 1, 2],
        "symbols_score": score_ids,
        "symbols_gold": gold_ids,
        "symbols_all": ids,
        "symbols_count": len(ids),
    }


def parse_free_spin_awards(awards_map: dict[int, int]) -> list[int]:
    max_scatter = max(6, max(awards_map.keys(), default=0))
    result = [0] * (max_scatter + 1)
    for scatter_count, free_spins in awards_map.items():
        if scatter_count < len(result):
            result[scatter_count] = int(free_spins)
    return result


def generate_config(xlsx_path: Path, template: dict[str, Any]) -> dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    overview = parse_overview(wb["Overview"])
    parameter = parse_parameter(wb["Parameter"])
    symbol_code_to_id = {code: int(symbol_id) for symbol_id, code in overview["symbol_str"].items()}

    table_data = [parse_table_sheet(wb[sheet_name], symbol_code_to_id) for sheet_name in TABLE_SHEETS]
    symbol_groups = build_symbol_groups(overview["symbol_id"], overview["symbol_str"])

    config: dict[str, Any] = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source_xlsx": str(xlsx_path),
        "source_box": "Source/H015192.xlsx",
        "game_id": overview["game_id"],
        "game_name": template["game_name"],
        "display_name": template["display_name"],
        "english_name": template["english_name"],
        "game_version": overview["excel_version"] or template["game_version"],
        "bet_options": template["bet_options"],
        "denom": template["denom"],
        "mode_normalbet": template["mode_normalbet"],
        "mode_featurebuy": template["mode_featurebuy"],
        "scene_bg": template["scene_bg"],
        "scene_fg": template["scene_fg"],
        "scene_bf": template["scene_bf"],
        "output_bg": template["output_bg"],
        "output_fg": template["output_fg"],
        "output_oa": template["output_oa"],
        "window_size": overview["window_size"],
        "reel_num": overview["reel_num"],
        "layout_shape": overview["layout_shape"],
        "layout_visible": overview["layout_visible"],
        "score_area": overview["score_area"],
        "special_area": overview["special_area"],
        "default_coin_in": overview["default_coin_in"],
        "normalbet": overview["normalbet"],
        "featurebuy": overview["featurebuy"],
        "max_spin_free_game": overview["max_spin_free_game"],
        "free_spin_awards": parse_free_spin_awards(overview["free_spin_awards"]),
        "pay_table": overview["pay_table"],
        "symbol_str": overview["symbol_str"],
        "symbol_id": overview["symbol_id"],
        "strip_name_map": TABLE_SHEETS,
        "strip_bg": 0,
        "strip_bg2": 1,
        "strip_fg": 2,
        "strip_fg2": 3,
        "strip_bf": 4,
        "arr_reels": [item["arr_reels"] for item in table_data],
        "arr_reels_weight_cum": [item["arr_reels_weight_cum"] for item in table_data],
        "reels_len": [item["reels_len"] for item in table_data],
        "weight_cum_drop_symbol_a": [item["drop_combo"] for item in table_data],
        "weight_cum_drop_symbol_b": [item["drop_combo"] for item in table_data],
        "weight_cum_drop_symbol_c": [item["drop_combo"] for item in table_data],
        "paytable_lines": [3, 4, 5, 6],
    }
    config.update(parameter)
    config.update(symbol_groups)
    return config


def write_config(path: Path, data: dict[str, Any]) -> None:
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    path.write_text(f"window.H015_BOX_DATA = {payload};\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate H015 config.js from H015192.xlsx")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--template", default=str(DEFAULT_OUTPUT))
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_path = Path(args.xlsx).resolve()
    output_path = Path(args.output).resolve()
    template = load_template(Path(args.template).resolve())
    config = generate_config(xlsx_path, template)
    write_config(output_path, config)
    print(f"Generated {output_path.name} from {xlsx_path.name}")


if __name__ == "__main__":
    main()
