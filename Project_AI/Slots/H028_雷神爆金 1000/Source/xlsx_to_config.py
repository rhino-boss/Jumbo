import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_SOURCE = BASE_DIR / "H028192A.xlsx"
DEFAULT_OUTPUT = BASE_DIR.parent / "config_92.js"

METADATA = {
    "game_id": "101016",
    "parsheet_id": "H0281",
    "display_name": "Thunder Boost 1000",
    "game_name_zh": "雷神爆金1000",
    "default_coin_in": 100,
    "mode_normalbet": 0,
    "mode_extrabet": 1,
    "mode_featurebuy": 2,
    "normalbet": 1,
    "featurebuy": 75,
    "supported_bet_modes": [0, 2],
}


def extract_transposed(sheet, range_text):
    rows = [[cell.value for cell in row] for row in sheet[range_text]]
    columns = [list(column) for column in zip(*rows)]
    if len(columns) == 1:
        return columns[0]
    return columns


def extract_no_transpose(sheet, range_text):
    return [[cell.value for cell in row] for row in sheet[range_text]]


def get_sheet(workbook, *names):
    for name in names:
        if name in workbook.sheetnames:
            return workbook[name]
    raise KeyError(f"Worksheet not found; tried {names}. Available: {workbook.sheetnames}")


def extract_linkpoint(overview):
    for row_index in range(1, overview.max_row + 1):
        if overview.cell(row_index, 1).value == "M1":
            return extract_no_transpose(overview, f"C{row_index}:F{row_index + 10}")
    raise ValueError("Pay-table start symbol M1 was not found in Overview")


def build_config(source_path):
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    output = dict(METADATA)

    overview = workbook["Overview"]
    excel_version = overview["B3"].value
    if excel_version is None or not str(excel_version).strip():
        workbook.close()
        raise ValueError("Overview!B3 (Version) is empty")
    output["excel_version"] = str(excel_version).strip()
    output["linkpoint"] = extract_linkpoint(overview)

    base = get_sheet(workbook, "Base Game Symbol", "BG_Symbol")
    base_ranges = {
        "BaseGameSymbol1": "T4:Z124",
        "BaseGameSymbolWeight1": "AB4:AH124",
        "BaseGameMegaWay1": "B33:G47",
        "BaseGameMY1": "B50:B62",
        "BaseGame1PostC1": "A65:B72",
        "BaseGameSymbol2": "BM4:BS124",
        "BaseGameSymbolWeight2": "BU4:CA124",
        "BaseGameMegaWay2": "AU33:AZ47",
        "BaseGameMY2": "AU50:AU62",
        "BaseGame2PostC1": "AT65:AU72",
    }
    for drop_index, first_row in enumerate((5, 34, 63, 92, 121), start=1):
        base_ranges[f"BaseGame1Drop{drop_index}"] = f"AK{first_row}:AQ{first_row + 25}"
        base_ranges[f"BaseGame2Drop{drop_index}"] = f"CD{first_row}:CJ{first_row + 25}"
    for key, range_text in base_ranges.items():
        output[key] = extract_transposed(base, range_text)

    description = workbook["Description"]
    output["ReelWeight"] = extract_transposed(description, "D5:D6")
    output["FreeReelWeight"] = extract_transposed(description, "G5:G7")
    output["FreeTriggerReel"] = extract_transposed(description, "D18:D20")

    free = get_sheet(workbook, "Free Game Symbol", "FG_Symbol")
    free_ranges = {
        "FreeGameSymbol1": "T4:Z124",
        "FreeGameSymbolWeight1": "AB4:AH124",
        "FreeGameMegaWay1": "B33:G47",
        "FreeGameMY1": "B50:B62",
        "FreeGame1PostC1": "A65:B72",
        "FreeGameSymbol2": "BM4:BS124",
        "FreeGameSymbolWeight2": "BU4:CA124",
        "FreeGameMegaWay2": "AU33:AZ47",
        "FreeGameMY2": "AU50:AU62",
        "FreeGame2PostC1": "AU65:AU72",
        "FreeGameSymbol3": "DF4:DL124",
        "FreeGameSymbolWeight3": "DN4:DT124",
        "FreeGameMegaWay3": "CN33:CS47",
        "FreeGameMY3": "CN50:CN62",
        "FreeGame3PostC1": "CM65:CN72",
    }
    for drop_index, first_row in enumerate((5, 34, 63, 92, 121), start=1):
        free_ranges[f"FreeGame1Drop{drop_index}"] = f"AK{first_row}:AQ{first_row + 25}"
        free_ranges[f"FreeGame2Drop{drop_index}"] = f"CD{first_row}:CJ{first_row + 25}"
        free_ranges[f"FreeGame3Drop{drop_index}"] = f"DW{first_row}:EC{first_row + 25}"
    for key, range_text in free_ranges.items():
        output[key] = extract_transposed(free, range_text)

    workbook.close()
    return output


def load_js_config(path):
    text = path.read_text(encoding="utf-8").strip()
    if not text.startswith("const data = "):
        raise ValueError(f"Unsupported config header: {path}")
    return json.loads(text[len("const data = ") :].rstrip(";"))


def write_js_config(path, data):
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "const data = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )


def derive_output_path(source_path):
    match = re.fullmatch(r"H0281(?P<rtp>\d{2})(?P<variant>[A-Za-z0-9_-]*)", source_path.stem, re.IGNORECASE)
    if not match:
        raise ValueError(
            f"Cannot derive config name from {source_path.name}; expected H0281 + two-digit RTP + optional variant, "
            "for example H028192A.xlsx"
        )
    return DEFAULT_OUTPUT.parent / f"config_{match.group('rtp')}{match.group('variant')}.js"


def compare_config(generated, output_path):
    current = load_js_config(output_path)
    changed_keys = [key for key in generated if generated[key] != current.get(key)]
    extra_keys = [key for key in current if key not in generated]
    if changed_keys or extra_keys:
        raise ValueError(f"Config mismatch for {output_path}. Changed keys: {changed_keys}; extra keys: {extra_keys}")


def process_source(source_path, output_path, check=False):
    generated = build_config(source_path)
    if check:
        compare_config(generated, output_path)
        print(f"Config is current: {output_path}")
    else:
        write_js_config(output_path, generated)
        compare_config(generated, output_path)
        print(f"Config written: {output_path} <- {source_path.name}")
    return generated


def main():
    parser = argparse.ArgumentParser(description="Build H028 config files from xlsx sources")
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path)
    parser.add_argument("--auto-output", action="store_true", help="Derive config_RTPVARIANT.js from the xlsx name")
    parser.add_argument("--all", action="store_true", help="Convert every H0281*.xlsx file in the Source folder")
    parser.add_argument(
        "--sync-default",
        action="store_true",
        help=f"Also update {DEFAULT_OUTPUT.name} when converting {DEFAULT_SOURCE.name}",
    )
    parser.add_argument("--check", action="store_true", help="Compare generated data with the existing output")
    args = parser.parse_args()

    if args.all:
        sources = sorted(
            path for path in BASE_DIR.glob("H0281*.xlsx") if not path.name.startswith("~$")
        )
        if not sources:
            raise FileNotFoundError(f"No H0281*.xlsx files found in {BASE_DIR}")
        for source_path in sources:
            output_path = derive_output_path(source_path)
            generated = process_source(source_path, output_path, args.check)
            if args.sync_default and source_path.name.casefold() == DEFAULT_SOURCE.name.casefold():
                if args.check:
                    compare_config(generated, DEFAULT_OUTPUT)
                    print(f"Config is current: {DEFAULT_OUTPUT}")
                else:
                    write_js_config(DEFAULT_OUTPUT, generated)
                    compare_config(generated, DEFAULT_OUTPUT)
                    print(f"Default config written: {DEFAULT_OUTPUT} <- {source_path.name}")
        print(f"Processed {len(sources)} xlsx file(s).")
        return

    source_path = args.source.resolve()
    output_path = (
        args.output.resolve()
        if args.output is not None
        else derive_output_path(source_path)
        if args.auto_output
        else DEFAULT_OUTPUT
    )
    generated = process_source(source_path, output_path, args.check)
    if args.sync_default and source_path.name.casefold() == DEFAULT_SOURCE.name.casefold() and output_path != DEFAULT_OUTPUT:
        if args.check:
            compare_config(generated, DEFAULT_OUTPUT)
            print(f"Config is current: {DEFAULT_OUTPUT}")
        else:
            write_js_config(DEFAULT_OUTPUT, generated)
            compare_config(generated, DEFAULT_OUTPUT)
            print(f"Default config written: {DEFAULT_OUTPUT} <- {source_path.name}")


if __name__ == "__main__":
    main()
