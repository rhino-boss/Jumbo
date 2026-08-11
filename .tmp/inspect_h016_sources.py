from pathlib import Path

from openpyxl import load_workbook


specs = [
    (Path("Project/Slots/H016_幸運王牌/Source/H0161.xlsx"), ["BG_Symbol", "FG_Symbol"]),
    (
        Path("C:/Users/rhinshen/Mine/個人工作區/市場資訊/H5/遊戲資源/JILI/JILI - Super Ace/遊戲資料/StripTable_SuperAce_還原.xlsx"),
        ["BG_Strip", "FG_Strip", "FillWeight"],
    ),
]

for path, sheet_names in specs:
    print(f"\nFILE {path}")
    workbook = load_workbook(path, read_only=True, data_only=False)
    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        print(f"\nSHEET {sheet_name}")
        for index, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=min(15, worksheet.max_row), values_only=True),
            start=1,
        ):
            print(index, repr(row))
