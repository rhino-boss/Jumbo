from openpyxl import load_workbook


before = load_workbook(".tmp/H0161_before_superace.xlsx", data_only=False)
after = load_workbook("Project/Slots/H016_幸運王牌/Source/H0161.xlsx", data_only=False)
for sheet_name in before.sheetnames:
    if sheet_name in {"BG_Symbol", "FG_Symbol"}:
        continue
    left, right = before[sheet_name], after[sheet_name]
    differences = []
    max_row = max(left.max_row, right.max_row)
    max_column = max(left.max_column, right.max_column)
    for row in range(1, max_row + 1):
        for column in range(1, max_column + 1):
            a, b = left.cell(row, column), right.cell(row, column)
            if a.value != b.value or a.number_format != b.number_format:
                differences.append((a.coordinate, repr(a.value), repr(b.value), a.number_format, b.number_format))
    if differences or left.row_dimensions != right.row_dimensions or left.column_dimensions != right.column_dimensions:
        print(sheet_name, "cell differences", differences[:20])
        print("before dims", left.max_row, left.max_column, dict(left.row_dimensions), dict(left.column_dimensions))
        print("after dims", right.max_row, right.max_column, dict(right.row_dimensions), dict(right.column_dimensions))
